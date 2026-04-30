from __future__ import annotations

from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def apply_fixplan_workflow_formatting(worksheet) -> None:
    headers = [cell.value for cell in worksheet[1]]
    header_to_col = {str(h): i + 1 for i, h in enumerate(headers) if h is not None}
    priority_col = header_to_col.get("Priority Score")
    points_col = header_to_col.get("Est. Sprint Points")
    aging_col = header_to_col.get("Aging/Priority")
    critical_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    warning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    good_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    edge_fill = PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid")
    for row_idx in range(2, worksheet.max_row + 1):
        if priority_col:
            cell = worksheet.cell(row=row_idx, column=priority_col)
            try:
                score = int(cell.value or 0)
                if score >= 100:
                    cell.fill = critical_fill
                elif score >= 65:
                    cell.fill = warning_fill
                else:
                    cell.fill = edge_fill
            except Exception:
                pass
        if points_col:
            cell = worksheet.cell(row=row_idx, column=points_col)
            try:
                points = int(cell.value or 0)
                if points >= 8:
                    cell.fill = critical_fill
                elif points >= 5:
                    cell.fill = warning_fill
                else:
                    cell.fill = good_fill
            except Exception:
                pass
        if aging_col:
            cell = worksheet.cell(row=row_idx, column=aging_col)
            value = str(cell.value or "").lower()
            if "immediate" in value:
                cell.fill = critical_fill
            elif "next sprint" in value:
                cell.fill = warning_fill
            elif "backlog" in value:
                cell.fill = edge_fill


def _header_map(worksheet) -> dict[str, int]:
    return {str(cell.value): idx for idx, cell in enumerate(worksheet[1], start=1) if cell.value}


def ensure_auto_filter(worksheet) -> None:
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"


def ensure_freeze_header(worksheet) -> None:
    if worksheet.max_row > 1:
        # Freeze header row and first column so URLs stay visible while scrolling.
        worksheet.freeze_panes = "B2"


def apply_global_conditional_formatting(worksheet) -> None:
    if worksheet.max_row <= 1:
        return
    headers = _header_map(worksheet)
    last_row = worksheet.max_row
    status_col = headers.get("Status Code") or headers.get("Target Status (if crawled)")
    if status_col:
        col = get_column_letter(status_col)
        rng = f"{col}2:{col}{last_row}"
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(operator="equal", formula=["200"], fill=PatternFill("solid", fgColor="C6EFCE")),
        )
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(operator="between", formula=["300", "399"], fill=PatternFill("solid", fgColor="FFEB9C")),
        )
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(operator="greaterThanOrEqual", formula=["400"], fill=PatternFill("solid", fgColor="FFC7CE")),
        )

    for load_header in ("Load Time (s)", "Load Time", "TTFB (ms)"):
        load_col = headers.get(load_header)
        if load_col:
            col = get_column_letter(load_col)
            rng = f"{col}2:{col}{last_row}"
            worksheet.conditional_formatting.add(
                rng,
                ColorScaleRule(
                    start_type="min",
                    start_color="63BE7B",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFEB84",
                    end_type="max",
                    end_color="F8696B",
                ),
            )
            worksheet.conditional_formatting.add(
                rng,
                DataBarRule(
                    start_type="min",
                    end_type="max",
                    color="638EC6",
                    showValue=True,
                ),
            )
            break

    for wc_header in ("Word Count", "Word Count (Body)"):
        wc_col = headers.get(wc_header)
        if wc_col:
            col = get_column_letter(wc_col)
            rng = f"{col}2:{col}{last_row}"
            worksheet.conditional_formatting.add(
                rng,
                ColorScaleRule(
                    start_type="min",
                    start_color="F8696B",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFEB84",
                    end_type="max",
                    end_color="63BE7B",
                ),
            )
            worksheet.conditional_formatting.add(
                rng,
                DataBarRule(
                    start_type="min",
                    end_type="max",
                    color="63BE7B",
                    showValue=True,
                ),
            )
            break

    priority_col = headers.get("Priority Score")
    if priority_col:
        col = get_column_letter(priority_col)
        rng = f"{col}2:{col}{last_row}"
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=["85"],
                font=Font(color="9C0006", bold=True),
                fill=PatternFill("solid", fgColor="FFC7CE"),
            ),
        )
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="between",
                formula=["65", "84"],
                font=Font(color="9C5700", bold=True),
                fill=PatternFill("solid", fgColor="FFEB9C"),
            ),
        )

    seo_score_col = headers.get("SEO Health Score")
    if seo_score_col:
        col = get_column_letter(seo_score_col)
        worksheet.conditional_formatting.add(
            f"{col}2:{col}{last_row}",
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )

    aeo_score_col = headers.get("AEO Readiness Score")
    if aeo_score_col:
        col = get_column_letter(aeo_score_col)
        worksheet.conditional_formatting.add(
            f"{col}2:{col}{last_row}",
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )

    action_col = headers.get("Action Needed")
    if action_col:
        col = get_column_letter(action_col)
        rng = f"{col}2:{col}{last_row}"
        worksheet.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'LOWER({col}2)="yes"'], stopIfTrue=True, fill=PatternFill("solid", fgColor="FFC7CE")),
        )
        worksheet.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'LOWER({col}2)="no"'], stopIfTrue=True, fill=PatternFill("solid", fgColor="C6EFCE")),
        )

    severity_badge_col = headers.get("Severity Badge")
    if severity_badge_col:
        col = get_column_letter(severity_badge_col)
        rng = f"{col}2:{col}{last_row}"
        worksheet.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'LOWER({col}2)="critical"'], stopIfTrue=True, fill=PatternFill("solid", fgColor="FFC7CE")),
        )
        worksheet.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'LOWER({col}2)="warning"'], stopIfTrue=True, fill=PatternFill("solid", fgColor="FFCC99")),
        )
        worksheet.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'LOWER({col}2)="pass"'], stopIfTrue=True, fill=PatternFill("solid", fgColor="C6EFCE")),
        )

    status_text_col = headers.get("Status")
    if status_text_col:
        col = get_column_letter(status_text_col)
        rng = f"{col}2:{col}{last_row}"
        worksheet.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'LOWER({col}2)="done"'], stopIfTrue=True, fill=PatternFill("solid", fgColor="D9EAD3")),
        )
        worksheet.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'OR(LOWER({col}2)="to do",LOWER({col}2)="in progress",LOWER({col}2)="in review")'],
                stopIfTrue=True,
                fill=PatternFill("solid", fgColor="FFF2CC"),
            ),
        )
