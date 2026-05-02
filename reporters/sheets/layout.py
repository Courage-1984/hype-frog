from __future__ import annotations

from collections.abc import Callable
from typing import Any

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from reporters.sheets.utils import header_index, to_int

MAIN_COLUMN_GROUP_DEFINITIONS: dict[str, list[str]] = {
    "Metadata Group": [
        "Title",
        "Meta Description",
        "Title Length",
        "Meta Desc Length",
        "OG-Image",
    ],
    "Heading Structure Group": [
        "H1 Content",
        "H1 Length",
        "H2 Content",
        "H2 Length",
        "H3 Content",
        "H3 Length",
        "H4 Content",
        "H4 Length",
        "H5 Content",
        "H5 Length",
        "H6 Content",
        "H6 Length",
    ],
    "Performance & CWV Group": [
        "CWV LCP (s)",
        "CWV CLS",
        "Field vs Lab",
        "Regional Authority Score",
        "Desktop PSI Score",
        "Mobile PSI Score",
        "Mobile LCP (s)",
        "Mobile CLS",
        "Mobile TTFB (s)",
    ],
    "Google Search Console Group": [
        "GSC Clicks",
        "GSC Impressions",
        "GSC CTR",
        "GSC Avg Position",
    ],
    "Crawl & Discovery Group": [
        "Click Depth",
        "Orphan Pages",
        "Internal PageRank",
        "Found via Sitemap",
        "Found via Crawl",
        "Discovery Source",
    ],
    "Raw State Group": [
        "Extraction State",
        "Extraction Source",
        "Technical Health",
        "Copy Score",
        "SEO Score",
        "Technical View",
    ],
}


def sort_worksheet_rows(
    worksheet: Worksheet, key_fn: Callable[[list[Any]], tuple[Any, ...]]
) -> None:
    """Sort all data rows in place while preserving the header row.

    Args:
        worksheet: Target worksheet containing a header row at index 1.
        key_fn: Sort-key callable receiving each row as a list of values.
    """
    if worksheet.max_row <= 2:
        return
    rows = [
        list(row)
        for row in worksheet.iter_rows(
            min_row=2, max_row=worksheet.max_row, values_only=True
        )
    ]
    rows.sort(key=key_fn)
    for row_idx, row_values in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_values, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)


def apply_intelligent_sorting(worksheet: Worksheet, sheet_name: str) -> None:
    """Apply sheet-specific deterministic sorting for report readability.

    Args:
        worksheet: Worksheet to sort.
        sheet_name: Canonical worksheet name controlling sort strategy.
    """
    headers = header_index(worksheet)
    if sheet_name == "FixPlan":
        severity_rank: dict[str, int] = {
            "Critical": 0,
            "High": 1,
            "Warning": 2,
            "Medium": 3,
            "Low": 4,
            "Observation": 5,
        }
        pcol = headers.get("Priority Score")
        scol = headers.get("Severity")
        ucol = headers.get("URL")
        if pcol:
            sort_worksheet_rows(
                worksheet,
                key_fn=lambda r: (
                    -to_int(r[pcol - 1], 0),
                    severity_rank.get(str(r[scol - 1]) if scol else "", 99),
                    str(r[ucol - 1] or "") if ucol else "",
                ),
            )
    elif sheet_name in {"Main", "Technical"}:
        sccol = headers.get("Status Code")
        ucol = headers.get("URL")
        if sccol:
            sort_worksheet_rows(
                worksheet,
                key_fn=lambda r: (
                    -to_int(r[sccol - 1], 0),
                    str(r[ucol - 1] or "") if ucol else "",
                ),
            )
    elif sheet_name == "AIOSEO":
        severity_rank = {"Critical": 0, "Warning": 1, "Observation": 2}
        sev_col = headers.get("Severity")
        pri_col = headers.get("Priority Score")
        panel_col = headers.get("AIOSEO Panel")
        ucol = headers.get("URL")
        if sev_col:
            sort_worksheet_rows(
                worksheet,
                key_fn=lambda r: (
                    severity_rank.get(str(r[sev_col - 1]) if sev_col else "", 99),
                    -to_int(r[pri_col - 1], 0) if pri_col else 0,
                    str(r[panel_col - 1] or "") if panel_col else "",
                    str(r[ucol - 1] or "") if ucol else "",
                ),
            )


def hide_noisy_columns(worksheet: Worksheet, sheet_name: str) -> None:
    """Hide low-signal raw/debug columns for client-facing exports.

    Args:
        worksheet: Worksheet to mutate.
        sheet_name: Canonical worksheet name used for token selection.
    """
    noisy_tokens_by_sheet: dict[str, list[str]] = {
        "Main": ["json-ld", "schema", "raw", "headers", "paragraph", "text extract"],
        "Technical": [
            "json-ld",
            "schema",
            "raw",
            "headers",
            "html",
            "paragraph",
            "text extract",
        ],
        "Content": ["paragraph", "full text", "raw", "json", "headers"],
    }
    tokens = noisy_tokens_by_sheet.get(sheet_name)
    if not tokens:
        return
    for idx, cell in enumerate(worksheet[1], start=1):
        header = str(cell.value or "").lower()
        if any(tok in header for tok in tokens):
            worksheet.column_dimensions[worksheet.cell(row=1, column=idx).column_letter].hidden = True


def reorder_columns(worksheet: Worksheet, sheet_name: str) -> None:
    """Reorder sheet columns according to preferred operational layouts.

    Args:
        worksheet: Worksheet to mutate.
        sheet_name: Canonical sheet name used to pick a preferred ordering.
    """
    preferred_orders: dict[str, list[str]] = {
        "Main": [
            "Health Icon",
            "URL",
            "Status Code",
            "Indexability",
            "Load Time (s)",
            "Title",
            "Meta Description",
            "Word Count (Body)",
            "SEO Health Score",
            "Severity Badge",
            "Action Needed",
            "Owner",
            "Status",
            "Sprint",
        ],
        "Technical": [
            "URL",
            "Status Code",
            "Status Class",
            "SEO Health Score",
            "Severity Badge",
            "Action Needed",
            "Indexability Reason",
            "TTFB (ms)",
            "Total Request Time (ms)",
            "Final URL",
            "Canonical URL",
            "Canonical Type",
            "Owner",
            "Status",
            "Sprint",
        ],
        "FixPlan": [
            "Issue Type",
            "Severity",
            "Priority Score",
            "Affected Count",
            "Affected URLs",
            "Detail Reference Tab",
            "Resolution Type",
            "URL",
            "Recommended Fix",
            "Likely Root Cause",
            "Owner",
            "Agency Owner",
            "Effort",
            "Est. Hours",
            "Est. Sprint Points",
            "Aging/Priority",
            "Status",
            "Verified By",
            "Date Resolved",
            "Revenue Risk",
            "Action Needed",
            "Jump to Details",
            "View Details",
            "Sprint",
        ],
        "Summary": [
            "Section",
            "Severity",
            "Issue",
            "Affected URL Count",
            "Reference Tab",
            "Affected URLs (sample)",
        ],
        "Priority URLs": [
            "URL",
            "Business Risk Score",
            "Severity Badge",
            "SEO Health Score",
            "GSC Impressions",
            "GSC CTR",
            "Revenue Intent",
            "Critical Issues Count",
            "Warning Issues Count",
            "Action Needed",
            "Why Prioritized",
            "Owner",
            "Status",
            "Sprint",
        ],
        "Content": [
            "URL",
            "Word Count",
            "Word Count Band",
            "Readability (Rough Flesch)",
            "H1 Count",
            "Missing H1 Flag",
            "Multiple H1 Flag",
            "Title Missing",
            "Meta Description Missing",
            "Thin Content Flag",
        ],
        "Links": [
            "URL",
            "Internal Links Count",
            "Unique Internal Links Count",
            "Broken Internal Links Count",
            "Unresolved Internal Links Count",
            "Generic Anchor Text Count",
            "External Links Count",
            "Nofollow Internal Links Count",
            "Nofollow External Links Count",
            "Internal Link Statuses",
        ],
        "AIOSEO": [
            "URL",
            "WordPress Post ID",
            "Direct Edit Link",
            "AIOSEO Panel",
            "Severity",
            "Issue",
            "Priority Score",
            "Current Value",
            "Recommended Target",
            "How to Fix in AIOSEO",
            "Reference Tab",
            "Reference Field",
            "Action Needed",
            "Owner",
            "Status",
            "Est. Hours",
            "Stable Issue ID",
        ],
        "Content Optimization Hub": [
            "Action Required",
            "Status",
            "Assigned Owner",
            "Content Cluster ID",
            "URL",
            "Current SEO Score",
            "Projected SEO Score",
            "Elementor Builder Link",
            "Target Keywords",
            "Current Page Copy Snippet",
            "Current Title",
            "Proposed Title (50-60 Chars)",
            "Title Count",
            "Current Meta Desc",
            "Proposed Meta Desc (120-160 Chars)",
            "Desc Count",
            "Current H-Tag Structure",
            "Proposed H-Tag Fixes",
            "AEO Answer Block Draft",
            "FAQ/QA Draft",
            "Current OG-Image URL",
            "OG Image Preview",
            "Social Share Note",
        ],
        "AEO": [
            "URL",
            "AEO Readiness Score",
            "AEO Badge",
            "Why It Matters",
            "FAQ Section Count",
            "Question Heading Count",
            "Paragraphs 40-60 Words Count",
            "QAPage/FAQ Schema Present",
            "Speakable Schema Present",
            "HowTo Signal",
            "Definition Signal",
            "List/Table Answer Signal",
        ],
        "Indexability": [
            "URL",
            "Status Code",
            "Status Class",
            "Indexability Reason",
            "Canonical URL",
            "Canonical Type",
            "Canonical Matches Final URL",
            "Canonical in Sitemap Match",
            "Meta Robots Raw",
            "X-Robots-Tag",
            "Final URL",
        ],
        "IssueInventory": [
            "URL",
            "Issue",
            "Severity",
            "Reference Tab",
            "Stable Issue ID",
            "Owner",
            "Status",
            "Sprint",
            "Open in Main",
            "Open in Reference",
        ],
    }
    preferred = preferred_orders.get(sheet_name)
    if not preferred or worksheet.max_row < 1:
        return
    current_headers = [
        worksheet.cell(row=1, column=i).value for i in range(1, worksheet.max_column + 1)
    ]
    if not any(h in current_headers for h in preferred):
        return
    ordered_headers = [h for h in preferred if h in current_headers]
    ordered_headers.extend([h for h in current_headers if h not in ordered_headers])
    if ordered_headers == current_headers:
        return
    idx_map = [current_headers.index(h) + 1 for h in ordered_headers]
    rows: list[list[Any]] = []
    for row_idx in range(1, worksheet.max_row + 1):
        rows.append(
            [worksheet.cell(row=row_idx, column=src_col).value for src_col in idx_map]
        )
    for row_idx, row_vals in enumerate(rows, start=1):
        for col_idx, val in enumerate(row_vals, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=val)


def apply_column_grouping(
    worksheet: Worksheet, group_definitions: dict[str, list[str]]
) -> None:
    """Apply collapsible column outlines for related metric groups.

    Args:
        worksheet: Worksheet where column grouping should be applied.
        group_definitions: Mapping of group name to ordered header names.
    """
    if worksheet.max_column <= 1:
        return
    if not group_definitions:
        return

    headers = header_index(worksheet)
    for _group_name, grouped_headers in group_definitions.items():
        present_indices: list[int] = []
        for header_name in grouped_headers:
            col_idx = headers.get(header_name)
            if col_idx is not None:
                present_indices.append(col_idx)
        if not present_indices:
            continue

        for col_idx in sorted(set(present_indices)):
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].outlineLevel = 1
            worksheet.column_dimensions[col_letter].hidden = True


def apply_column_widths(worksheet: Worksheet) -> None:
    """Set column widths from cell content, then clamp known-wide headers.

    Mirrors the original ``adjust_sheet_format`` auto-fit block: iterate
    columns for max string length (+2, cap 60), apply URL-like header minimums,
    then cap URL columns on wide sheets (30+ columns) to 36 for readability.
    """
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)
    headers = header_index(worksheet)
    for header_name, width in {
        "URL": 45,
        "Final URL": 45,
        "Canonical URL": 45,
        "Affected URLs": 55,
        "How to Fix in AIOSEO": 55,
    }.items():
        col_idx = headers.get(header_name)
        if col_idx:
            letter = get_column_letter(col_idx)
            current = worksheet.column_dimensions[letter].width or width
            worksheet.column_dimensions[letter].width = min(current, width)
    is_wide_sheet = worksheet.max_column >= 30
    if is_wide_sheet:
        for header_name in ("URL", "Final URL", "Canonical URL", "Affected URLs"):
            col_idx = headers.get(header_name)
            if col_idx:
                worksheet.column_dimensions[get_column_letter(col_idx)].width = 36


__all__ = [
    "MAIN_COLUMN_GROUP_DEFINITIONS",
    "sort_worksheet_rows",
    "apply_intelligent_sorting",
    "hide_noisy_columns",
    "reorder_columns",
    "apply_column_grouping",
    "apply_column_widths",
]
