from __future__ import annotations

from openpyxl.styles import Font, PatternFill


def apply_workbook_toc_and_links(
    writer,
    *,
    debug_excel_isolation_mode: bool,
    disable_non_core_freeze_panes: bool,
    std_navy: str,
    std_white: str,
    std_blue: str,
) -> None:
    def _sheet_link_target(name: str) -> str:
        return str(name).replace("'", "''")

    def _clear_orphaned_selection(ws) -> None:
        try:
            ws.views.sheetView[0].selection = []
        except Exception:
            pass

    if debug_excel_isolation_mode:
        return
    wb = writer.book
    wb.calculation.calcMode = "auto"
    if "Table of Contents" not in wb.sheetnames:
        toc_ws = wb.create_sheet("Table of Contents")
        toc_ws["A1"] = "Table of Contents"
        toc_ws["A1"].font = Font(color=std_navy, bold=True, size=14)
        toc_ws["A2"] = "Section"
        toc_ws["B2"] = "Open"
        toc_ws["C2"] = "Description"
        toc_ws["A2"].fill = PatternFill("solid", fgColor=std_navy)
        toc_ws["B2"].fill = PatternFill("solid", fgColor=std_navy)
        toc_ws["C2"].fill = PatternFill("solid", fgColor=std_navy)
        toc_ws["A2"].font = Font(color=std_white, bold=True)
        toc_ws["B2"].font = Font(color=std_white, bold=True)
        toc_ws["C2"].font = Font(color=std_white, bold=True)
        toc_descriptions = {
            "Dashboard": "Executive overview and core site metrics.",
            "Content Optimization Hub": "Copywriter workspace. Edit Meta Data, H-Tags, and AEO snippets.",
            "Quick Reference Guide": "SEO and AEO standards and target lengths.",
            "FixPlan": "Prioritized list of technical actions based on ROI.",
            "Main": "Primary URL inventory with key crawl metrics.",
            "Technical": "Status codes, response times, and server metrics.",
            "Content": "Word counts, readability, and content depth analysis.",
            "AEO": "Answer Engine Optimization and snippet extraction readiness.",
            "Schema & Metadata": "JSON-LD, Microdata, and OpenGraph validation.",
            "Links": "Internal and external link counts per page.",
            "Indexability": "Robots.txt, Canonical tags, and NoIndex directives.",
            "Redirects": "301/302 Redirect chains and loops.",
            "Priority URLs": "Highest business-value pages requiring immediate attention.",
            "AIOSEO": "All in One SEO plugin data extraction.",
            "Security": "SSL, mixed content, and header security.",
            "Summary": "High-level aggregate crawl data.",
            "LinksDetail": "Row-by-row internal outlink breakdown.",
            "Media": "Image sizes, alt text, and broken media.",
            "Pattern and Template Issues": "Sitewide structural flaws detected by folder path.",
            "Duplicates": "Exact and near-duplicate content detection.",
            "PSI Performance": "Core Web Vitals and Google PageSpeed metrics.",
            "IssueInventory": "Raw log of all detected errors.",
            "RunMetadata": "Crawl timestamp and configuration details.",
            "DeltaFromPreviousRun": "Changes detected since the previous crawl run.",
            "ResolvedIssues": "Issues that are now fixed versus prior runs.",
            "Glossary & Legend": "Definitions for metrics, statuses, and scoring labels.",
            "CrawlGraph": "Link relationships and crawl path visibility.",
            "SitemapQA": "Sitemap coverage and metadata validation.",
        }
        row_ptr = 3
        for sheet_name in wb.sheetnames:
            if sheet_name == "Table of Contents":
                continue
            toc_ws[f"A{row_ptr}"] = sheet_name
            toc_ws[f"B{row_ptr}"] = f'=HYPERLINK("#\'{_sheet_link_target(sheet_name)}\'!A1","Open")'
            toc_ws[f"C{row_ptr}"] = toc_descriptions.get(
                sheet_name, "Detailed URL diagnostic data."
            )
            toc_ws[f"B{row_ptr}"].font = Font(
                color=std_blue, underline="single", bold=True
            )
            row_ptr += 1
        toc_ws.column_dimensions["A"].width = 35
        toc_ws.column_dimensions["B"].width = 18
        toc_ws.column_dimensions["C"].width = 70
        toc_ws.freeze_panes = "A3"
    link_map = {
        "Summary": "Reference Tab",
        "FixPlan": "Reference Tab",
        "Dashboard": "Target Tab",
        "AIOSEO": "Reference Tab",
        "IssueInventory": "Reference Tab",
    }
    for sheet_name, col_header in link_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        if col_header not in headers:
            continue
        col_idx = headers.index(col_header) + 1
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            target = str(cell.value or "").strip()
            if target and target in wb.sheetnames:
                cell.hyperlink = f"#{target}!A1"
                cell.style = "Hyperlink"
    preferred_first_tabs = [
        "Table of Contents",
        "Dashboard",
        "Content Optimization Hub",
        "Quick Reference Guide",
        "FixPlan",
        "Main",
        "Technical",
        "Content",
        "AEO",
        "Schema & Metadata",
        "Links",
        "Indexability",
        "Redirects",
        "Priority URLs",
        "AIOSEO",
        "Security",
        "Summary",
        "LinksDetail",
        "Media",
        "Duplicates",
        "Pattern and Template Issues",
        "PSI Performance",
        "IssueInventory",
        "CrawlGraph",
        "SitemapQA",
        "DeltaFromPreviousRun",
        "RunMetadata",
    ]
    legend_name = "Glossary & Legend"
    if legend_name in wb.sheetnames and legend_name not in preferred_first_tabs:
        preferred_first_tabs.append(legend_name)
    for idx, tab_name in enumerate(preferred_first_tabs):
        if tab_name in wb.sheetnames:
            wb.move_sheet(wb[tab_name], offset=-wb.index(wb[tab_name]) + idx)
    if legend_name in wb.sheetnames:
        wb.move_sheet(
            wb[legend_name], offset=len(wb.sheetnames) - 1 - wb.index(wb[legend_name])
        )
    low_signal_tabs = {"RunMetadata", "DeltaFromPreviousRun"}
    for tab_name in low_signal_tabs:
        if tab_name in wb.sheetnames:
            wb[tab_name].sheet_state = "hidden"
    for tab_name in wb.sheetnames:
        ws = wb[tab_name]
        if disable_non_core_freeze_panes and tab_name not in {"Main", "Dashboard"}:
            ws.freeze_panes = None
            _clear_orphaned_selection(ws)
            continue
        if tab_name not in {"Main", "Dashboard"} and (
            ws.max_row < 10 or ws.max_column < 5
        ):
            ws.freeze_panes = None
            ws.auto_filter.ref = None
            _clear_orphaned_selection(ws)
            continue
        if tab_name == "Content Optimization Hub":
            ws.freeze_panes = "F3" if ws.max_row >= 3 and ws.max_column >= 6 else None
            if ws.freeze_panes is None:
                _clear_orphaned_selection(ws)
        elif tab_name in {
            "Main",
            "Technical",
            "Content",
            "Links",
            "AEO",
            "Schema & Metadata",
            "Indexability",
            "Redirects",
            "Priority URLs",
            "AIOSEO",
            "Security",
            "Summary",
            "LinksDetail",
            "Media",
            "Pattern and Template Issues",
            "Duplicates",
            "PSI Performance",
            "IssueInventory",
            "RunMetadata",
            "DeltaFromPreviousRun",
            "ResolvedIssues",
            "CrawlGraph",
            "SitemapQA",
            "FixPlan",
        }:
            ws.freeze_panes = "B2" if ws.max_row >= 2 and ws.max_column >= 2 else None
            if ws.freeze_panes is None:
                _clear_orphaned_selection(ws)
