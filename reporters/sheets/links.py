from __future__ import annotations

from typing import Any

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from utils import normalize_url_key


def is_safe_hyperlink_target(
    target: str, *, disable_external_links_and_images: bool
) -> bool:
    if disable_external_links_and_images:
        return False
    return bool(target) and len(target) <= 255


def sanitize_excel_url(url_value: Any) -> str:
    raw = str(url_value or "").strip()
    if not raw:
        return ""
    raw = "".join(ch for ch in raw if ord(ch) >= 32).replace('"', "").replace("'", "")
    return raw


def normalize_url_for_match(url_value: Any) -> str:
    return normalize_url_key(sanitize_excel_url(url_value))


def add_url_navigation_links(
    writer,
    worksheet,
    sheet_name: str,
    *,
    debug_excel_isolation_mode: bool,
    disable_external_links_and_images: bool,
    header_index_fn,
) -> None:
    if debug_excel_isolation_mode:
        return
    headers = header_index_fn(worksheet)
    url_col = headers.get("URL")
    if not url_col or worksheet.max_row <= 1:
        return
    for r in range(2, worksheet.max_row + 1):
        url_cell = worksheet.cell(row=r, column=url_col)
        url_val = str(url_cell.value or "").strip()
        if url_val.startswith(("http://", "https://")) and is_safe_hyperlink_target(
            url_val, disable_external_links_and_images=disable_external_links_and_images
        ):
            url_cell.hyperlink = url_val
            url_cell.style = "Hyperlink"
            url_cell.alignment = Alignment(wrap_text=True, vertical="top")

    if sheet_name not in {"Main", "Dashboard"} and "Main" in writer.book.sheetnames:
        if "Open in Main" not in headers:
            new_col = worksheet.max_column + 1
            worksheet.cell(row=1, column=new_col, value="Open in Main")
            url_col_letter = get_column_letter(url_col)
            for r in range(2, worksheet.max_row + 1):
                worksheet.cell(
                    row=r,
                    column=new_col,
                    value=f'=IFERROR(HYPERLINK("#Main!A"&MATCH({url_col_letter}{r},Main!A:A,0),"Open"),HYPERLINK("#Main!A1","Open"))',
                )


def apply_cross_sheet_links(
    writer,
    worksheet,
    sheet_name: str,
    *,
    debug_excel_isolation_mode: bool,
    header_index_fn,
) -> None:
    if debug_excel_isolation_mode:
        return
    headers = header_index_fn(worksheet)
    if sheet_name == "Summary":
        issue_col = headers.get("Issue")
        fix_ws = writer.book["FixPlan"] if "FixPlan" in writer.book.sheetnames else None
        fix_headers = header_index_fn(fix_ws) if fix_ws else {}
        fix_issue_col = fix_headers.get("Issue Type")
        fix_issue_to_row: dict[str, int] = {}
        if fix_ws and fix_issue_col:
            for r in range(2, fix_ws.max_row + 1):
                issue_name = str(fix_ws.cell(row=r, column=fix_issue_col).value or "").strip()
                if issue_name and issue_name not in fix_issue_to_row:
                    fix_issue_to_row[issue_name] = r
        if issue_col:
            for r in range(2, worksheet.max_row + 1):
                issue = str(worksheet.cell(row=r, column=issue_col).value or "").strip()
                if issue and not issue.startswith("==="):
                    issue_cell = worksheet.cell(row=r, column=issue_col)
                    fix_row = fix_issue_to_row.get(issue)
                    if fix_row:
                        issue_cell.hyperlink = f"#FixPlan!A{fix_row}"
                        issue_cell.style = "Hyperlink"
    if sheet_name == "Main":
        url_col = headers.get("URL")
        if url_col:
            target_col = worksheet.max_column + 1
            worksheet.cell(row=1, column=target_col, value="Technical View")
            col_letter = get_column_letter(url_col)
            for r in range(2, worksheet.max_row + 1):
                worksheet.cell(
                    row=r,
                    column=target_col,
                    value=f'=IFERROR(HYPERLINK("#Technical!A"&MATCH({col_letter}{r},Technical!A:A,0),"Open Technical"),HYPERLINK("#Technical!A1","Open Technical"))',
                )
    if sheet_name == "Priority URLs":
        url_col = headers.get("URL")
        if url_col:
            new_col = headers.get("Open in Technical")
            if not new_col:
                new_col = worksheet.max_column + 1
                worksheet.cell(row=1, column=new_col, value="Open in Technical")
            url_letter = get_column_letter(url_col)
            for r in range(2, worksheet.max_row + 1):
                worksheet.cell(
                    row=r,
                    column=new_col,
                    value=f'=IFERROR(HYPERLINK("#Technical!A"&MATCH({url_letter}{r},Technical!A:A,0),"Open"),HYPERLINK("#Technical!A1","Open"))',
                )
    if sheet_name == "IssueInventory":
        url_col = headers.get("URL")
        issue_col = headers.get("Issue")
        reference_tab_col = headers.get("Reference Tab")
        if url_col:
            new_col = headers.get("Open in Main")
            if not new_col:
                new_col = worksheet.max_column + 1
                worksheet.cell(row=1, column=new_col, value="Open in Main")
            url_letter = get_column_letter(url_col)
            for r in range(2, worksheet.max_row + 1):
                worksheet.cell(
                    row=r,
                    column=new_col,
                    value=f'=IFERROR(HYPERLINK("#Main!A"&MATCH({url_letter}{r},Main!A:A,0),"Open"),HYPERLINK("#Main!A1","Open"))',
                )
        if reference_tab_col:
            open_ref_col = headers.get("Open in Reference")
            if not open_ref_col:
                open_ref_col = worksheet.max_column + 1
                worksheet.cell(row=1, column=open_ref_col, value="Open in Reference")
            ref_letter = get_column_letter(reference_tab_col)
            url_letter = get_column_letter(url_col) if url_col else "A"
            for r in range(2, worksheet.max_row + 1):
                worksheet.cell(
                    row=r,
                    column=open_ref_col,
                    value=f'=IFERROR(HYPERLINK("#"&{ref_letter}{r}&"!A"&MATCH({url_letter}{r},INDIRECT("\'"&{ref_letter}{r}&"\'!A:A"),0),"Open"),HYPERLINK("#"&{ref_letter}{r}&"!A1","Open"))',
                )
        if issue_col and "FixPlan" in writer.book.sheetnames:
            fix_ws = writer.book["FixPlan"]
            fix_headers = header_index_fn(fix_ws)
            fix_issue_col = fix_headers.get("Issue Type")
            fix_issue_rows: dict[str, int] = {}
            if fix_issue_col:
                for r in range(2, fix_ws.max_row + 1):
                    key = str(fix_ws.cell(row=r, column=fix_issue_col).value or "").strip()
                    if key and key not in fix_issue_rows:
                        fix_issue_rows[key] = r
                for r in range(2, worksheet.max_row + 1):
                    issue = str(worksheet.cell(row=r, column=issue_col).value or "").strip()
                    target_row = fix_issue_rows.get(issue)
                    if target_row:
                        cell = worksheet.cell(row=r, column=issue_col)
                        cell.hyperlink = f"#FixPlan!A{target_row}"
                        cell.style = "Hyperlink"
    if sheet_name == "AIOSEO":
        url_col = headers.get("URL")
        if url_col:
            technical_col = headers.get("Open in Technical")
            if not technical_col:
                technical_col = worksheet.max_column + 1
                worksheet.cell(row=1, column=technical_col, value="Open in Technical")
            url_letter = get_column_letter(url_col)
            for r in range(2, worksheet.max_row + 1):
                worksheet.cell(
                    row=r,
                    column=technical_col,
                    value=f'=IFERROR(HYPERLINK("#Technical!A"&MATCH({url_letter}{r},Technical!A:A,0),"Open"),HYPERLINK("#Technical!A1","Open"))',
                )
    if sheet_name == "FixPlan" and "Content Optimization Hub" in writer.book.sheetnames:
        headers = header_index_fn(worksheet)
        url_col = headers.get("URL")
        hub_status_col = headers.get("Hub Status (Content Hub)")
        if not hub_status_col:
            hub_status_col = worksheet.max_column + 1
            worksheet.cell(row=1, column=hub_status_col, value="Hub Status (Content Hub)")
        if url_col:
            u_letter = get_column_letter(url_col)
            for r in range(2, worksheet.max_row + 1):
                worksheet.cell(
                    row=r,
                    column=hub_status_col,
                    value=f"=IFERROR(INDEX('Content Optimization Hub'!A:A,MATCH({u_letter}{r},'Content Optimization Hub'!C:C,0)),\"Not in Hub\")",
                )
