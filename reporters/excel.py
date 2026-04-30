from __future__ import annotations

from collections import Counter
from collections import defaultdict
from typing import Any

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from reporters.formatting import (
    apply_fixplan_workflow_formatting,
    apply_global_conditional_formatting,
    ensure_auto_filter,
    ensure_freeze_header,
)

STD_NAVY = "1F4E78"
STD_WHITE = "FFFFFF"
STD_BLUE = "0563C1"
DATA_HEAVY_TABS = {"Main", "Technical", "AEO", "AIOSEO", "FixPlan", "SnippetCandidates", "Summary"}


def _header_index(worksheet) -> dict[str, int]:
    return {str(cell.value): idx for idx, cell in enumerate(worksheet[1], start=1) if cell.value}


def _add_back_to_dashboard_link(worksheet, sheet_name: str) -> None:
    if sheet_name == "Dashboard":
        return
    target_col = worksheet.max_column + 1
    target_ref = f"{get_column_letter(target_col)}1"
    worksheet[target_ref] = "BACK TO DASHBOARD"
    worksheet[target_ref].hyperlink = "#Dashboard!A1"
    worksheet[target_ref].style = "Hyperlink"
    worksheet[target_ref].font = Font(color=STD_BLUE, underline="single", bold=True)
    worksheet[target_ref].alignment = Alignment(horizontal="left")


def _to_int(value: Any, default: int = 0) -> int:
    try:
        return int(float(value))
    except Exception:
        return default


def _sort_worksheet_rows(worksheet, key_fn) -> None:
    if worksheet.max_row <= 2:
        return
    rows = [list(row) for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True)]
    rows.sort(key=key_fn)
    for row_idx, row_values in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_values, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)


def _apply_intelligent_sorting(worksheet, sheet_name: str) -> None:
    headers = _header_index(worksheet)
    if sheet_name == "FixPlan":
        severity_rank = {"Critical": 0, "High": 1, "Warning": 2, "Medium": 3, "Low": 4, "Info": 5}
        pcol = headers.get("Priority Score")
        scol = headers.get("Severity")
        ucol = headers.get("URL")
        if pcol:
            _sort_worksheet_rows(
                worksheet,
                key_fn=lambda r: (
                    -_to_int(r[pcol - 1], 0),
                    severity_rank.get(str(r[scol - 1]) if scol else "", 99),
                    str(r[ucol - 1] or "") if ucol else "",
                ),
            )
    elif sheet_name in {"Main", "Technical"}:
        sccol = headers.get("Status Code")
        ucol = headers.get("URL")
        if sccol:
            _sort_worksheet_rows(
                worksheet,
                key_fn=lambda r: (
                    -_to_int(r[sccol - 1], 0),
                    str(r[ucol - 1] or "") if ucol else "",
                ),
            )
    elif sheet_name == "AIOSEO":
        severity_rank = {"Critical": 0, "Warning": 1, "Info": 2}
        sev_col = headers.get("Severity")
        pri_col = headers.get("Priority Score")
        panel_col = headers.get("AIOSEO Panel")
        ucol = headers.get("URL")
        if sev_col:
            _sort_worksheet_rows(
                worksheet,
                key_fn=lambda r: (
                    severity_rank.get(str(r[sev_col - 1]) if sev_col else "", 99),
                    -_to_int(r[pri_col - 1], 0) if pri_col else 0,
                    str(r[panel_col - 1] or "") if panel_col else "",
                    str(r[ucol - 1] or "") if ucol else "",
                ),
            )


def _hide_noisy_columns(worksheet, sheet_name: str) -> None:
    noisy_tokens_by_sheet = {
        "Main": ["json-ld", "schema", "raw", "headers", "paragraph", "text extract"],
        "Technical": ["json-ld", "schema", "raw", "headers", "html", "paragraph", "text extract"],
        "Content": ["paragraph", "full text", "raw", "json", "headers"],
    }
    tokens = noisy_tokens_by_sheet.get(sheet_name)
    if not tokens:
        return
    for idx, cell in enumerate(worksheet[1], start=1):
        header = str(cell.value or "").lower()
        if any(tok in header for tok in tokens):
            worksheet.column_dimensions[get_column_letter(idx)].hidden = True


def _friendly_metric_label(header: str) -> str:
    return (
        header.replace("_", " ")
        .replace("URL", "URL")
        .replace("SEO", "SEO")
        .replace("AEO", "AEO")
        .strip()
    )


def _tooltip_for_header(header: str) -> str:
    h = (header or "").strip()
    lower = h.lower()
    if not h:
        return "Column descriptor for this worksheet section."
    if "url" in lower:
        return f"{_friendly_metric_label(h)}. Use this URL to inspect the page directly or jump to related tabs."
    if "status code" in lower:
        return "HTTP response code returned for this page/resource. Fix: Resolve 4xx/5xx errors and reduce unnecessary redirects."
    if "status class" in lower:
        return "Grouped HTTP class (2xx/3xx/4xx/5xx). Use this to quickly triage crawl health and error concentration."
    if "health score" in lower:
        return "Composite SEO quality score for this URL. Higher is better. Guide: 90+ strong, 70-89 needs tuning, below 70 high priority."
    if "pass rate" in lower:
        return "Share of URLs marked as pass across the crawl. Use together with Error Rate and critical issue counts for decisions."
    if "severity" in lower:
        return "Issue impact level. Fix: Resolve Critical first, then Warning, then improvement opportunities."
    if "priority score" in lower:
        return "Execution priority score combining impact and effort. Fix: Start from highest scores."
    if "affected count" in lower:
        return "How many URLs are impacted by this issue. Larger counts usually indicate template-level or systemic problems."
    if "ttfb" in lower:
        return "Time to First Byte: server responsiveness signal. Fix: optimise backend performance, caching, and CDN usage."
    if "load time" in lower or "request time" in lower:
        return "Page/request speed metric. Fix: optimise assets, server response time, and blocking resources."
    if "indexability" in lower or "canonical" in lower:
        return "Indexing/canonicalization signal. Fix: ensure indexable pages use correct canonicals and non-conflicting directives."
    if "meta robots" in lower or "x-robots" in lower:
        return "Robots directives that influence indexation. Fix: remove unintended noindex/nofollow values on important pages."
    if "word count" in lower or "readability" in lower:
        return "Content quality depth metric. Fix: expand thin content and improve clarity for search intent."
    if "link" in lower or "anchor" in lower:
        return "Link quality and crawl-path metric. Fix: repair broken links and improve internal linking relevance."
    if "open in main" in lower or "technical view" in lower or "view details" in lower:
        return "Navigation helper. Click to jump directly to the related record in another tab."
    if "schema" in lower or "json-ld" in lower:
        return "Structured data signal. Fix: add valid schema types and correct parsing/validation errors."
    if "owner" in lower or "sprint" in lower or "status" == lower:
        return "Workflow management field for planning and tracking remediation progress."
    if "section" in lower or "reference tab" in lower:
        return "Summary grouping/navigation field. Use with hyperlinks to jump into detailed tabs."
    return f"{_friendly_metric_label(h)}. Review this metric to identify risk, then use linked tabs for details and remediation."


def _add_all_header_tooltips(worksheet) -> None:
    headers = _header_index(worksheet)
    for header, col_idx in headers.items():
        ref = f"{get_column_letter(col_idx)}1"
        dv = DataValidation(type="custom", formula1="TRUE", showInputMessage=True, allow_blank=True)
        dv.promptTitle = _friendly_metric_label(header)[:32] or "Column"
        dv.prompt = _tooltip_for_header(header)
        worksheet.add_data_validation(dv)
        dv.add(ref)


def _add_url_navigation_links(writer, worksheet, sheet_name: str) -> None:
    headers = _header_index(worksheet)
    url_col = headers.get("URL")
    if not url_col or worksheet.max_row <= 1:
        return
    for r in range(2, worksheet.max_row + 1):
        url_cell = worksheet.cell(row=r, column=url_col)
        url_val = str(url_cell.value or "").strip()
        if url_val.startswith(("http://", "https://")):
            url_cell.hyperlink = url_val
            url_cell.style = "Hyperlink"
            # Hyperlink style can reset wrapping; force wrapped top alignment for long URLs.
            url_cell.alignment = Alignment(wrap_text=True, vertical="top")

    if sheet_name not in {"Main", "Dashboard"} and "Main" in writer.book.sheetnames:
        if "Open in Main" not in headers:
            new_col = worksheet.max_column + 1
            worksheet.cell(row=1, column=new_col, value="Open in Main")
            url_col_letter = get_column_letter(url_col)
            for r in range(2, worksheet.max_row + 1):
                worksheet.cell(
                    row=r,
                    column=new_col,
                    value=f'=IFERROR(HYPERLINK("#Main!A"&MATCH({url_col_letter}{r},Main!A:A,0),"Open"),HYPERLINK("#Main!A1","Open"))',
                )


def _apply_fixplan_interactivity(writer, worksheet) -> None:
    headers = _header_index(worksheet)
    if worksheet.max_row <= 1:
        return
    if "Status" not in headers:
        new_col = worksheet.max_column + 1
        worksheet.cell(row=1, column=new_col, value="Status")
        for r in range(2, worksheet.max_row + 1):
            worksheet.cell(row=r, column=new_col, value="To Do")
        headers = _header_index(worksheet)
    status_col = headers.get("Status")
    if status_col:
        _apply_status_dropdown(worksheet, status_col)

    owner_col = headers.get("Agency Owner")
    if owner_col:
        owners = sorted({str(worksheet.cell(r, owner_col).value).strip() for r in range(2, worksheet.max_row + 1) if worksheet.cell(r, owner_col).value})
        if not owners:
            owners = ["Copy Writer", "Dev", "Server/Host"]
        owner_list = ",".join(owners[:8])
        owner_dv = DataValidation(type="list", formula1=f'"{owner_list}"', allow_blank=True)
        worksheet.add_data_validation(owner_dv)
        owner_dv.add(f"{get_column_letter(owner_col)}2:{get_column_letter(owner_col)}{worksheet.max_row}")

    url_col = headers.get("URL")
    jump_col = headers.get("Jump to Details")
    details_col = headers.get("View Details")
    if not details_col:
        details_col = worksheet.max_column + 1
        worksheet.cell(row=1, column=details_col, value="View Details")
        headers = _header_index(worksheet)
    for row_idx in range(2, worksheet.max_row + 1):
        if url_col:
            url_cell = worksheet.cell(row=row_idx, column=url_col)
            url = str(url_cell.value or "").strip()
            if url.startswith(("http://", "https://")):
                url_cell.hyperlink = url
                url_cell.style = "Hyperlink"
                url_cell.alignment = Alignment(wrap_text=True, vertical="top")
        if jump_col and url_col:
            url = str(worksheet.cell(row=row_idx, column=url_col).value or "").strip()
            jump_cell = worksheet.cell(row=row_idx, column=jump_col)
            if url:
                jump_cell.value = f'=IFERROR(HYPERLINK("#Main!A"&MATCH("{url}",Main!A:A,0),"Jump to Main"),HYPERLINK("#Main!A1","Open Main"))'
            else:
                jump_cell.value = '=HYPERLINK("#Main!A1","Open Main")'
        if details_col and url_col:
            url_col_letter = get_column_letter(url_col)
            worksheet.cell(
                row=row_idx,
                column=details_col,
                value=f'=IFERROR(HYPERLINK("#Main!A"&MATCH({url_col_letter}{row_idx},Main!A:A,0),"View URL Details"),HYPERLINK("#Main!A1","View URL Details"))',
            )


def _apply_status_dropdown(worksheet, status_col: int) -> None:
    if status_col <= 0 or worksheet.max_row <= 1:
        return
    status_dv = DataValidation(type="list", formula1='"To Do,In Progress,Fixed"', allow_blank=True)
    worksheet.add_data_validation(status_dv)
    status_dv.add(f"{get_column_letter(status_col)}2:{get_column_letter(status_col)}{worksheet.max_row}")


def _style_dashboard(worksheet, writer) -> None:
    worksheet.sheet_view.showGridLines = False
    # Keep metric context visible while scrolling large dashboard blocks.
    worksheet.freeze_panes = "B2"
    worksheet._charts = []
    for col_letter, width in {"A": 3, "B": 35, "C": 20}.items():
        worksheet.column_dimensions[col_letter].width = width

    headers = _header_index(worksheet)
    # Clear legacy helper blocks outside the B:C dashboard grid.
    for row_idx in range(1, 80):
        for col_idx in range(4, max(worksheet.max_column + 1, 26)):
            worksheet.cell(row=row_idx, column=col_idx, value=None)
    metric_col = headers.get("Metric")
    value_col = headers.get("Value")
    total_urls = 0
    pass_rate_pct = 0.0
    critical_urls = 0
    warning_urls = 0
    if metric_col and value_col:
        metric_to_value = {}
        for row_idx in range(2, worksheet.max_row + 1):
            metric_name = str(worksheet.cell(row=row_idx, column=metric_col).value or "").strip()
            if metric_name:
                metric_to_value[metric_name] = worksheet.cell(row=row_idx, column=value_col).value
        total_urls = _to_int(metric_to_value.get("URLs Crawled"), 0)
        try:
            pass_rate_pct = float(metric_to_value.get("Pass Rate (%)", 0) or 0.0)
        except Exception:
            pass_rate_pct = 0.0
        critical_urls = _to_int(metric_to_value.get("Critical URL Count"), 0)
        warning_urls = _to_int(metric_to_value.get("Warning URL Count"), 0)

        title = worksheet["B2"]
        title.value = "Executive SEO Performance Report"
        worksheet["B1"] = "Technical Audit Dashboard"
        worksheet["B1"].font = Font(color=STD_NAVY, bold=True, size=12)
        title.font = Font(color=STD_NAVY, bold=True, size=16)
        title.alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[2].height = 30

        header_fill = PatternFill("solid", fgColor=STD_NAVY)
        header_font = Font(color=STD_WHITE, bold=True, size=12)
        value_fill = PatternFill("solid", fgColor="F5F7FA")

        worksheet["B4"] = "EXECUTIVE METRICS"
        worksheet["C4"] = "Value"
        for ref in ("B4", "C4"):
            worksheet[ref].fill = header_fill
            worksheet[ref].font = header_font
            worksheet[ref].alignment = Alignment(horizontal="center", vertical="center")

        worksheet["B5"] = '=HYPERLINK("#Main!A1","Total URLs")'
        worksheet["C5"] = total_urls
        worksheet["B6"] = '=HYPERLINK("#Summary!A1","Overall Health Score")'
        worksheet["C6"] = "=IFERROR(0,0)"
        worksheet["C6"].number_format = "0.00%"
        worksheet["B7"] = '=HYPERLINK("#Summary!A1","SEO Pass Rate %")'
        worksheet["C7"] = "=IFERROR(0,0)"
        worksheet["C7"].number_format = "0.00%"
        worksheet["B8"] = '=HYPERLINK("#Technical!A1","Pass URLs")'
        worksheet["C8"] = 0
        worksheet["B9"] = '=HYPERLINK("#FixPlan!A1","Critical URLs")'
        worksheet["C9"] = critical_urls
        worksheet["B10"] = '=HYPERLINK("#Technical!A1","Warning URLs")'
        worksheet["C10"] = warning_urls
        worksheet["B11"] = '=HYPERLINK("#Technical!A1","Error Rate % (4xx/5xx)")'
        worksheet["C11"] = "=IFERROR(0,0)"
        worksheet["C11"].number_format = "0.00%"
        worksheet["B12"] = '=HYPERLINK("#Technical!A1","Crawl Success Rate % (2xx)")'
        worksheet["C12"] = "=IFERROR(0,0)"
        worksheet["C12"].number_format = "0.00%"
        worksheet["B13"] = '=HYPERLINK("#Technical!A1","Critical URL Rate %")'
        worksheet["C13"] = "=IFERROR(0,0)"
        worksheet["C13"].number_format = "0.00%"
        worksheet["B14"] = '=HYPERLINK("#Technical!A1","Warning URL Rate %")'
        worksheet["C14"] = "=IFERROR(0,0)"
        worksheet["C14"].number_format = "0.00%"
        for ref in ("B5", "C5", "B6", "C6", "B7", "C7", "B8", "C8", "B9", "C9", "B10", "C10", "B11", "C11", "B12", "C12", "B13", "C13", "B14", "C14"):
            worksheet[ref].fill = value_fill
            worksheet[ref].font = Font(color="1F2937", bold=True, size=12 if ref.startswith("B") else 14)
            worksheet[ref].alignment = Alignment(horizontal="center", vertical="center")
        for ref in ("B5", "B6", "B7", "B8", "B9", "B10", "B11", "B12", "B13", "B14"):
            worksheet[ref].font = Font(color=STD_BLUE, underline="single", bold=True, size=12)

    # In-cell dashboard tables: status and severity summaries.
    status_buckets = {"200 OK": 0, "3xx Redirects": 0, "4xx Errors": 0, "5xx Errors": 0, "Other": 0}
    pass_urls = 0
    avg_ttfb_ms = 0.0
    if "Technical" in writer.book.sheetnames:
        tech_ws = writer.book["Technical"]
        tech_headers = _header_index(tech_ws)
        sc_col = tech_headers.get("Status Code")
        sev_col = tech_headers.get("Severity Badge")
        ttfb_col = tech_headers.get("TTFB (ms)")
        ttfb_values: list[float] = []
        if sc_col:
            for r in range(2, tech_ws.max_row + 1):
                code = _to_int(tech_ws.cell(row=r, column=sc_col).value, 0)
                if 200 <= code < 300:
                    status_buckets["200 OK"] += 1
                elif 300 <= code < 400:
                    status_buckets["3xx Redirects"] += 1
                elif 400 <= code < 500:
                    status_buckets["4xx Errors"] += 1
                elif 500 <= code < 600:
                    status_buckets["5xx Errors"] += 1
                elif code:
                    status_buckets["Other"] += 1
                if sev_col:
                    sev_val = str(tech_ws.cell(row=r, column=sev_col).value or "").strip().lower()
                    if sev_val == "critical":
                        critical_urls += 1
                    elif sev_val == "warning":
                        warning_urls += 1
                if ttfb_col:
                    raw_ttfb = tech_ws.cell(row=r, column=ttfb_col).value
                    try:
                        if raw_ttfb is not None and str(raw_ttfb).strip() != "":
                            ttfb_values.append(float(raw_ttfb))
                    except Exception:
                        pass
        if ttfb_values:
            avg_ttfb_ms = round(sum(ttfb_values) / len(ttfb_values), 2)
        crit_col = tech_headers.get("Critical Issues Count")
        warn_col = tech_headers.get("Warning Issues Count")
        if crit_col and warn_col:
            for r in range(2, tech_ws.max_row + 1):
                crit_count = _to_int(tech_ws.cell(row=r, column=crit_col).value, 0)
                warn_count = _to_int(tech_ws.cell(row=r, column=warn_col).value, 0)
                # Treat Info-only URLs as pass; failing means critical/warning exists.
                if crit_count == 0 and warn_count == 0:
                    pass_urls += 1
        else:
            sev_badge_col = tech_headers.get("Severity Badge")
            if sev_badge_col:
                for r in range(2, tech_ws.max_row + 1):
                    sev = str(tech_ws.cell(row=r, column=sev_badge_col).value or "").strip().lower()
                    if sev in {"pass", "info"}:
                        pass_urls += 1

    severity_counts = Counter()
    if "FixPlan" in writer.book.sheetnames:
        fix_ws = writer.book["FixPlan"]
        fix_headers = _header_index(fix_ws)
        sev_col = fix_headers.get("Severity")
        if sev_col:
            for r in range(2, fix_ws.max_row + 1):
                sev = str(fix_ws.cell(r, sev_col).value or "").strip().lower()
                if sev == "critical":
                    severity_counts["Critical"] += 1
                elif sev in {"high", "warning"}:
                    severity_counts["High"] += 1
                elif sev == "medium":
                    severity_counts["Medium"] += 1
                else:
                    severity_counts["Low"] += 1

    # Real KPI formulas based on parsed status buckets and available score columns.
    status_total = sum(status_buckets.values())
    error_count = status_buckets["4xx Errors"] + status_buckets["5xx Errors"]
    success_count = status_buckets["200 OK"]
    crawl_denominator = max(1, status_total or total_urls)
    pass_rate_pct = round((pass_urls / crawl_denominator) * 100, 2)
    critical_rate_pct = round((critical_urls / crawl_denominator) * 100, 2)
    warning_rate_pct = round((warning_urls / crawl_denominator) * 100, 2)
    worksheet["C8"] = pass_urls
    worksheet["C9"] = critical_urls
    worksheet["C10"] = warning_urls
    worksheet["C11"] = f"=IFERROR({error_count}/{crawl_denominator},0)"
    worksheet["C11"].number_format = "0.00%"
    worksheet["C12"] = f"=IFERROR({success_count}/{crawl_denominator},0)"
    worksheet["C12"].number_format = "0.00%"
    worksheet["C13"] = f"=IFERROR({critical_urls}/{crawl_denominator},0)"
    worksheet["C13"].number_format = "0.00%"
    worksheet["C14"] = f"=IFERROR({warning_urls}/{crawl_denominator},0)"
    worksheet["C14"].number_format = "0.00%"
    worksheet["C7"] = f"=IFERROR({pass_urls}/{crawl_denominator},0)"
    worksheet["C7"].number_format = "0.00%"

    avg_health_score = None
    if "Technical" in writer.book.sheetnames:
        technical_ws = writer.book["Technical"]
        technical_headers = _header_index(technical_ws)
        score_col = technical_headers.get("SEO Health Score")
        if score_col:
            score_values: list[float] = []
            for r in range(2, technical_ws.max_row + 1):
                raw = technical_ws.cell(row=r, column=score_col).value
                try:
                    if raw is not None and str(raw).strip() != "":
                        score_values.append(float(raw))
                except Exception:
                    pass
            if score_values:
                avg_health_score = sum(score_values) / len(score_values)
    if avg_health_score is not None:
        worksheet["C6"] = f"=IFERROR({round(avg_health_score, 2)}/100,0)"
    else:
        worksheet["C6"] = f"=IFERROR({pass_rate_pct}/100,0)"
    worksheet["C6"].number_format = "0.00%"

    table_header_fill = PatternFill("solid", fgColor=STD_NAVY)
    table_header_font = Font(color=STD_WHITE, bold=True, size=11)
    for ref, val in {"B16": "Status", "C16": "Count", "B23": "Severity", "C23": "Count"}.items():
        cell = worksheet[ref]
        cell.value = val
        cell.fill = table_header_fill
        cell.font = table_header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    status_rows = [
        ("200 OK", status_buckets["200 OK"], "C6EFCE"),
        ("3xx Redirects", status_buckets["3xx Redirects"], "FFCC99"),
        ("4xx Errors", status_buckets["4xx Errors"], "FFC1C1"),
        ("5xx Errors", status_buckets["5xx Errors"], "FFC1C1"),
        ("Other", status_buckets["Other"], "FFCC99"),
    ]
    for idx, (label, count, color) in enumerate(status_rows, start=17):
        worksheet[f"B{idx}"] = label
        worksheet[f"C{idx}"] = count
        worksheet[f"B{idx}"].fill = PatternFill("solid", fgColor=color)
        worksheet[f"C{idx}"].fill = PatternFill("solid", fgColor=color)

    sev_rows = [
        ("Critical", severity_counts.get("Critical", 0), "FFC1C1"),
        ("Warning", severity_counts.get("High", 0), "FFCC99"),
        ("Medium", severity_counts.get("Medium", 0), "FFCC99"),
        ("Low", severity_counts.get("Low", 0), "C6EFCE"),
    ]
    for idx, (label, count, color) in enumerate(sev_rows, start=24):
        worksheet[f"B{idx}"] = label
        worksheet[f"C{idx}"] = count
        worksheet[f"B{idx}"].fill = PatternFill("solid", fgColor=color)
        worksheet[f"C{idx}"].fill = PatternFill("solid", fgColor=color)

    for row in range(17, 28):
        for col in ("B", "C"):
            worksheet[f"{col}{row}"].alignment = Alignment(horizontal="center", vertical="center")

    # High-priority operational snapshot.
    worksheet["B29"] = "PRIORITY SNAPSHOT"
    worksheet["C29"] = "Value"
    for ref in ("B29", "C29"):
        worksheet[ref].fill = table_header_fill
        worksheet[ref].font = table_header_font
        worksheet[ref].alignment = Alignment(horizontal="center", vertical="center")
    top_issue_name = "N/A"
    top_issue_affected = 0
    owner_rollup: dict[str, dict[str, int]] = defaultdict(lambda: {"issue_rows": 0, "affected_urls": 0, "critical": 0, "warning": 0, "info": 0})
    if "FixPlan" in writer.book.sheetnames:
        fix_ws = writer.book["FixPlan"]
        fix_headers = _header_index(fix_ws)
        issue_col = fix_headers.get("Issue Type")
        affected_col = fix_headers.get("Affected Count")
        owner_col = fix_headers.get("Owner")
        sev_col = fix_headers.get("Severity")
        if issue_col and affected_col:
            for r in range(2, fix_ws.max_row + 1):
                affected = _to_int(fix_ws.cell(row=r, column=affected_col).value, 0)
                if affected > top_issue_affected:
                    top_issue_affected = affected
                    top_issue_name = str(fix_ws.cell(row=r, column=issue_col).value or "N/A")
                owner_name = str(fix_ws.cell(row=r, column=owner_col).value or "Unassigned").strip() if owner_col else "Unassigned"
                sev_val = str(fix_ws.cell(row=r, column=sev_col).value or "").strip().lower() if sev_col else ""
                owner_rollup[owner_name]["issue_rows"] += 1
                owner_rollup[owner_name]["affected_urls"] += affected
                if sev_val == "critical":
                    owner_rollup[owner_name]["critical"] += 1
                elif sev_val in {"warning", "high", "medium"}:
                    owner_rollup[owner_name]["warning"] += 1
                else:
                    owner_rollup[owner_name]["info"] += 1
    worksheet["B30"] = "Top Blocking Issue"
    worksheet["C30"] = top_issue_name
    worksheet["B31"] = "Top Issue Affected URLs"
    worksheet["C31"] = top_issue_affected
    worksheet["B32"] = "4xx/5xx URLs"
    worksheet["C32"] = status_buckets["4xx Errors"] + status_buckets["5xx Errors"]
    worksheet["B33"] = "Avg TTFB (ms)"
    worksheet["C33"] = avg_ttfb_ms
    for row in range(30, 34):
        worksheet[f"B{row}"].fill = PatternFill("solid", fgColor="F5F7FA")
        worksheet[f"C{row}"].fill = PatternFill("solid", fgColor="F5F7FA")

    # Pivot-style owner summary block.
    worksheet["E4"] = "OWNER ISSUE SUMMARY (PIVOT)"
    worksheet["E5"] = "Owner"
    worksheet["F5"] = "Issue Rows"
    worksheet["G5"] = "Affected URLs"
    worksheet["H5"] = "Critical"
    worksheet["I5"] = "Warning"
    worksheet["J5"] = "Info"
    for ref in ("E4", "E5", "F5", "G5", "H5", "I5", "J5"):
        worksheet[ref].fill = table_header_fill
        worksheet[ref].font = table_header_font
        worksheet[ref].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.merge_cells("E4:J4")

    owner_rows_sorted = sorted(
        owner_rollup.items(),
        key=lambda x: (-x[1]["affected_urls"], -x[1]["critical"], -x[1]["warning"], x[0]),
    )
    owner_start_row = 6
    for idx, (owner_name, metrics) in enumerate(owner_rows_sorted[:8], start=owner_start_row):
        worksheet[f"E{idx}"] = f'=HYPERLINK("#FixPlan!A1","{owner_name}")'
        worksheet[f"F{idx}"] = metrics["issue_rows"]
        worksheet[f"G{idx}"] = metrics["affected_urls"]
        worksheet[f"H{idx}"] = metrics["critical"]
        worksheet[f"I{idx}"] = metrics["warning"]
        worksheet[f"J{idx}"] = metrics["info"]
        for col in ("E", "F", "G", "H", "I", "J"):
            worksheet[f"{col}{idx}"].fill = PatternFill("solid", fgColor="F5F7FA")
            worksheet[f"{col}{idx}"].alignment = Alignment(horizontal="center", vertical="center")
    if not owner_rows_sorted:
        worksheet["E6"] = "No owner data"
        worksheet["E6"].fill = PatternFill("solid", fgColor="F5F7FA")
        worksheet["E6"].alignment = Alignment(horizontal="center", vertical="center")

    for col, width in {"E": 22, "F": 12, "G": 14, "H": 10, "I": 10, "J": 10}.items():
        worksheet.column_dimensions[col].width = width

    # KPI emphasis colors.
    try:
        overall_health = float((worksheet["C6"].value or "0").replace("=IFERROR(", "").split("/")[0])
    except Exception:
        overall_health = 0.0
    health_fill = "C6EFCE" if overall_health >= 80 else "FFEB9C" if overall_health >= 60 else "FFC7CE"
    worksheet["C6"].fill = PatternFill("solid", fgColor=health_fill)
    worksheet["C7"].fill = PatternFill("solid", fgColor=health_fill)
    worksheet["C7"].fill = PatternFill("solid", fgColor="C6EFCE" if pass_rate_pct >= 70 else "FFEB9C" if pass_rate_pct >= 40 else "FFC7CE")
    worksheet["C11"].fill = PatternFill("solid", fgColor="FFC7CE" if error_count > 0 else "C6EFCE")
    worksheet["C12"].fill = PatternFill("solid", fgColor="C6EFCE" if success_count >= max(1, crawl_denominator * 0.9) else "FFEB9C")
    worksheet["C13"].fill = PatternFill("solid", fgColor="FFC7CE" if critical_rate_pct >= 0.1 else "FFEB9C" if critical_rate_pct > 0 else "C6EFCE")
    worksheet["C14"].fill = PatternFill("solid", fgColor="FFC7CE" if warning_rate_pct >= 0.5 else "FFEB9C" if warning_rate_pct > 0.2 else "C6EFCE")

    # Top issue list with direct links.
    worksheet["B42"] = "TOP ISSUES TO FIX FIRST"
    worksheet["C42"] = "Affected URLs"
    for ref in ("B42", "C42"):
        worksheet[ref].fill = table_header_fill
        worksheet[ref].font = table_header_font
        worksheet[ref].alignment = Alignment(horizontal="center", vertical="center")
    if "FixPlan" in writer.book.sheetnames:
        fix_ws = writer.book["FixPlan"]
        fix_headers = _header_index(fix_ws)
        issue_col = fix_headers.get("Issue Type")
        affected_col = fix_headers.get("Affected Count")
        priority_col = fix_headers.get("Priority Score")
        top_rows: list[tuple[str, int, int, int]] = []
        if issue_col and affected_col:
            for r in range(2, fix_ws.max_row + 1):
                issue_name = str(fix_ws.cell(row=r, column=issue_col).value or "").strip()
                affected = _to_int(fix_ws.cell(row=r, column=affected_col).value, 0)
                priority = _to_int(fix_ws.cell(row=r, column=priority_col).value, 0) if priority_col else 0
                if issue_name:
                    top_rows.append((issue_name, affected, priority, r))
        top_rows.sort(key=lambda x: (-x[1], -x[2], x[0]))
        for idx, (issue_name, affected, _priority, source_row) in enumerate(top_rows[:5], start=43):
            worksheet[f"B{idx}"] = f'=HYPERLINK("#FixPlan!A{source_row}","{issue_name}")'
            worksheet[f"C{idx}"] = affected
            worksheet[f"B{idx}"].fill = PatternFill("solid", fgColor="F5F7FA")
            worksheet[f"C{idx}"].fill = PatternFill("solid", fgColor="F5F7FA")
            worksheet[f"B{idx}"].font = Font(color=STD_BLUE, underline="single", bold=True)

    # Quick navigation block.
    quick_nav_fill = PatternFill("solid", fgColor=STD_NAVY)
    worksheet["B35"] = "Quick Navigation"
    worksheet["C35"] = "Open"
    for ref in ("B35", "C35"):
        worksheet[ref].fill = quick_nav_fill
        worksheet[ref].font = Font(color=STD_WHITE, bold=True, size=11)
        worksheet[ref].alignment = Alignment(horizontal="center", vertical="center")
    quick_links = [
        ("Fix Plan", '#FixPlan!A1'),
        ("Main URL Data", '#Main!A1'),
        ("Technical Diagnostics", '#Technical!A1'),
        ("Indexability", '#Indexability!A1'),
        ("AEO Opportunities", '#AEO!A1'),
    ]
    quick_links.append(("AIOSEO Action Queue", '#AIOSEO!A1'))
    for idx, (label, target) in enumerate(quick_links, start=36):
        worksheet[f"B{idx}"] = label
        worksheet[f"C{idx}"] = f'=HYPERLINK("{target}","Open")'
        worksheet[f"C{idx}"].font = Font(color=STD_BLUE, underline="single", bold=True)
        worksheet[f"B{idx}"].fill = PatternFill("solid", fgColor="F5F7FA")
        worksheet[f"C{idx}"].fill = PatternFill("solid", fgColor="F5F7FA")

    # Dashboard KPI tooltips on result cells.
    dashboard_tooltips = {
        "C5": "Total URLs crawled in this run. Calculated as the number of audited URL rows.",
        "C6": "Overall Health Score. Calculated as average SEO Health Score across Technical URLs; fallback to SEO Pass Rate if score data is unavailable.",
        "C7": "SEO Pass Rate %. Calculated as Pass URLs divided by Total URLs.",
        "C8": "Pass URL count. URL is pass when it has no Critical and no Warning issues.",
        "C9": "Critical URL count from Technical severity badge.",
        "C10": "Warning URL count from Technical severity badge.",
        "C11": "HTTP Error Rate %. Calculated as (4xx URLs + 5xx URLs) / Total URLs.",
        "C12": "Crawl Success Rate %. Calculated as 2xx URLs / Total URLs.",
        "C13": "Critical URL Rate %. Calculated as Critical URLs / Total URLs.",
        "C14": "Warning URL Rate %. Calculated as Warning URLs / Total URLs.",
        "C30": "Most widespread issue from FixPlan by affected URL count.",
        "C31": "Number of URLs impacted by the top blocking issue.",
        "C32": "Total URLs returning client/server errors (4xx + 5xx).",
        "C33": "Average Time to First Byte across Technical URLs (ms).",
        "C43": "Affected URL count for the highest-priority issue (linked to FixPlan).",
        "C44": "Affected URL count for the next issue in the priority list.",
        "C45": "Affected URL count for the next issue in the priority list.",
        "C46": "Affected URL count for the next issue in the priority list.",
        "C47": "Affected URL count for the next issue in the priority list.",
        "E5": "Owner responsible for remediation. Click to open FixPlan.",
        "F5": "Number of issue rows assigned to this owner.",
        "G5": "Total affected URLs across this owner's assigned issues.",
        "H5": "Count of critical issue types assigned to this owner.",
        "I5": "Count of warning issue types assigned to this owner.",
        "J5": "Count of info issue types assigned to this owner.",
    }
    for ref, message in dashboard_tooltips.items():
        dv = DataValidation(type="custom", formula1="TRUE", showInputMessage=True, allow_blank=True)
        dv.promptTitle = f"KPI {ref}"
        dv.prompt = message
        worksheet.add_data_validation(dv)
        dv.add(ref)


def _reorder_columns(worksheet, sheet_name: str) -> None:
    preferred_orders = {
        "Main": [
            "Health Icon",
            "URL",
            "Status Code",
            "Indexability",
            "Load Time (s)",
            "Title",
            "Meta Description",
            "Word Count (Body)",
            "SEO Health Score",
            "Severity Badge",
            "Action Needed",
            "Owner",
            "Status",
            "Sprint",
        ],
        "Technical": [
            "URL",
            "Status Code",
            "Status Class",
            "SEO Health Score",
            "Severity Badge",
            "Action Needed",
            "Indexability Reason",
            "TTFB (ms)",
            "Total Request Time (ms)",
            "Final URL",
            "Canonical URL",
            "Canonical Type",
            "Owner",
            "Status",
            "Sprint",
        ],
        "FixPlan": [
            "Issue Type",
            "Severity",
            "Priority Score",
            "Affected Count",
            "Affected URLs",
            "URL",
            "Recommended Fix",
            "Likely Root Cause",
            "Owner",
            "Agency Owner",
            "Effort",
            "Est. Hours",
            "Est. Sprint Points",
            "Aging/Priority",
            "Status",
            "Action Needed",
            "Jump to Details",
            "View Details",
            "Sprint",
        ],
        "Summary": [
            "Section",
            "Severity",
            "Issue",
            "Affected URL Count",
            "Reference Tab",
            "Affected URLs (sample)",
        ],
        "Priority URLs": [
            "URL",
            "Business Risk Score",
            "Severity Badge",
            "SEO Health Score",
            "Critical Issues Count",
            "Warning Issues Count",
            "Action Needed",
            "Why Prioritized",
            "Owner",
            "Status",
            "Sprint",
        ],
        "Content": [
            "URL",
            "Word Count",
            "Word Count Band",
            "Readability (Rough Flesch)",
            "H1 Count",
            "Missing H1 Flag",
            "Multiple H1 Flag",
            "Title Missing",
            "Meta Description Missing",
            "Thin Content Flag",
        ],
        "Links": [
            "URL",
            "Internal Links Count",
            "Unique Internal Links Count",
            "Broken Internal Links Count",
            "Unresolved Internal Links Count",
            "Generic Anchor Text Count",
            "External Links Count",
            "Nofollow Internal Links Count",
            "Nofollow External Links Count",
            "Internal Link Statuses",
        ],
        "AIOSEO": [
            "URL",
            "AIOSEO Panel",
            "Severity",
            "Issue",
            "Priority Score",
            "Current Value",
            "Recommended Target",
            "How to Fix in AIOSEO",
            "Reference Tab",
            "Reference Field",
            "Action Needed",
            "Owner",
            "Status",
            "Est. Hours",
            "Stable Issue ID",
        ],
        "AEO": [
            "URL",
            "AEO Readiness Score",
            "AEO Badge",
            "Why It Matters",
            "FAQ Section Count",
            "Question Heading Count",
            "Paragraphs 40-60 Words Count",
            "QAPage/FAQ Schema Present",
            "Speakable Schema Present",
            "HowTo Signal",
            "Definition Signal",
            "List/Table Answer Signal",
        ],
        "Indexability": [
            "URL",
            "Status Code",
            "Status Class",
            "Indexability Reason",
            "Canonical URL",
            "Canonical Type",
            "Canonical Matches Final URL",
            "Canonical in Sitemap Match",
            "Meta Robots Raw",
            "X-Robots-Tag",
            "Final URL",
        ],
        "IssueInventory": [
            "URL",
            "Issue",
            "Severity",
            "Reference Tab",
            "Stable Issue ID",
            "Owner",
            "Status",
            "Sprint",
            "Open in Main",
            "Open in Reference",
        ],
    }
    preferred = preferred_orders.get(sheet_name)
    if not preferred or worksheet.max_row < 1:
        return
    current_headers = [worksheet.cell(row=1, column=i).value for i in range(1, worksheet.max_column + 1)]
    if not any(h in current_headers for h in preferred):
        return
    ordered_headers = [h for h in preferred if h in current_headers]
    ordered_headers.extend([h for h in current_headers if h not in ordered_headers])
    if ordered_headers == current_headers:
        return
    idx_map = [current_headers.index(h) + 1 for h in ordered_headers]
    rows = []
    for row_idx in range(1, worksheet.max_row + 1):
        rows.append([worksheet.cell(row=row_idx, column=src_col).value for src_col in idx_map])
    for row_idx, row_vals in enumerate(rows, start=1):
        for col_idx, val in enumerate(row_vals, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=val)


def _collapse_technical_deep_dive_columns(worksheet, sheet_name: str) -> None:
    if sheet_name != "Technical" or worksheet.max_column <= 1:
        return
    headers = _header_index(worksheet)
    deep_dive_headers = [
        "Hreflang Present",
        "Hreflang Count",
        "Hreflang Self Reference",
        "Hreflang Reciprocal Check",
        "Hreflang Canonical Consistency",
        "x-default Present",
        "Pagination rel=next",
        "Pagination rel=prev",
    ]
    deep_dive_cols = sorted([headers[h] for h in deep_dive_headers if h in headers])
    if not deep_dive_cols:
        return

    range_start = deep_dive_cols[0]
    prev_col = deep_dive_cols[0]
    for col_idx in deep_dive_cols[1:]:
        if col_idx == prev_col + 1:
            prev_col = col_idx
            continue
        worksheet.column_dimensions.group(
            get_column_letter(range_start),
            get_column_letter(prev_col),
            hidden=True,
            outline_level=1,
        )
        range_start = col_idx
        prev_col = col_idx
    worksheet.column_dimensions.group(
        get_column_letter(range_start),
        get_column_letter(prev_col),
        hidden=True,
        outline_level=1,
    )


def _apply_cross_sheet_links(writer, worksheet, sheet_name: str) -> None:
    headers = _header_index(worksheet)
    if sheet_name == "Summary":
        issue_col = headers.get("Issue")
        fix_ws = writer.book["FixPlan"] if "FixPlan" in writer.book.sheetnames else None
        fix_headers = _header_index(fix_ws) if fix_ws else {}
        fix_issue_col = fix_headers.get("Issue Type")
        fix_issue_to_row: dict[str, int] = {}
        if fix_ws and fix_issue_col:
            for r in range(2, fix_ws.max_row + 1):
                issue_name = str(fix_ws.cell(row=r, column=fix_issue_col).value or "").strip()
                if issue_name and issue_name not in fix_issue_to_row:
                    fix_issue_to_row[issue_name] = r
        if issue_col:
            for r in range(2, worksheet.max_row + 1):
                issue = str(worksheet.cell(row=r, column=issue_col).value or "").strip()
                if issue and not issue.startswith("==="):
                    issue_cell = worksheet.cell(row=r, column=issue_col)
                    fix_row = fix_issue_to_row.get(issue)
                    if fix_row:
                        issue_cell.hyperlink = f"#FixPlan!A{fix_row}"
                        issue_cell.style = "Hyperlink"
    if sheet_name == "Main":
        url_col = headers.get("URL")
        if url_col:
            target_col = worksheet.max_column + 1
            worksheet.cell(row=1, column=target_col, value="Technical View")
            col_letter = get_column_letter(url_col)
            for r in range(2, worksheet.max_row + 1):
                worksheet.cell(
                    row=r,
                    column=target_col,
                    value=f'=IFERROR(HYPERLINK("#Technical!A"&MATCH({col_letter}{r},Technical!A:A,0),"Open Technical"),HYPERLINK("#Technical!A1","Open Technical"))',
                )
    if sheet_name == "Priority URLs":
        url_col = headers.get("URL")
        if url_col:
            new_col = headers.get("Open in Technical")
            if not new_col:
                new_col = worksheet.max_column + 1
                worksheet.cell(row=1, column=new_col, value="Open in Technical")
            url_letter = get_column_letter(url_col)
            for r in range(2, worksheet.max_row + 1):
                worksheet.cell(
                    row=r,
                    column=new_col,
                    value=f'=IFERROR(HYPERLINK("#Technical!A"&MATCH({url_letter}{r},Technical!A:A,0),"Open"),HYPERLINK("#Technical!A1","Open"))',
                )
    if sheet_name == "IssueInventory":
        url_col = headers.get("URL")
        issue_col = headers.get("Issue")
        reference_tab_col = headers.get("Reference Tab")
        if url_col:
            new_col = headers.get("Open in Main")
            if not new_col:
                new_col = worksheet.max_column + 1
                worksheet.cell(row=1, column=new_col, value="Open in Main")
            url_letter = get_column_letter(url_col)
            for r in range(2, worksheet.max_row + 1):
                worksheet.cell(
                    row=r,
                    column=new_col,
                    value=f'=IFERROR(HYPERLINK("#Main!A"&MATCH({url_letter}{r},Main!A:A,0),"Open"),HYPERLINK("#Main!A1","Open"))',
                )
        if reference_tab_col:
            open_ref_col = headers.get("Open in Reference")
            if not open_ref_col:
                open_ref_col = worksheet.max_column + 1
                worksheet.cell(row=1, column=open_ref_col, value="Open in Reference")
            ref_letter = get_column_letter(reference_tab_col)
            url_letter = get_column_letter(url_col) if url_col else "A"
            for r in range(2, worksheet.max_row + 1):
                worksheet.cell(
                    row=r,
                    column=open_ref_col,
                    value=f'=IFERROR(HYPERLINK("#"&{ref_letter}{r}&"!A"&MATCH({url_letter}{r},INDIRECT("\'"&{ref_letter}{r}&"\'!A:A"),0),"Open"),HYPERLINK("#"&{ref_letter}{r}&"!A1","Open"))',
                )
        if issue_col and "FixPlan" in writer.book.sheetnames:
            fix_ws = writer.book["FixPlan"]
            fix_headers = _header_index(fix_ws)
            fix_issue_col = fix_headers.get("Issue Type")
            fix_issue_rows: dict[str, int] = {}
            if fix_issue_col:
                for r in range(2, fix_ws.max_row + 1):
                    key = str(fix_ws.cell(row=r, column=fix_issue_col).value or "").strip()
                    if key and key not in fix_issue_rows:
                        fix_issue_rows[key] = r
                for r in range(2, worksheet.max_row + 1):
                    issue = str(worksheet.cell(row=r, column=issue_col).value or "").strip()
                    target_row = fix_issue_rows.get(issue)
                    if target_row:
                        cell = worksheet.cell(row=r, column=issue_col)
                        cell.hyperlink = f"#FixPlan!A{target_row}"
                        cell.style = "Hyperlink"
    if sheet_name == "AIOSEO":
        url_col = headers.get("URL")
        if url_col:
            technical_col = headers.get("Open in Technical")
            if not technical_col:
                technical_col = worksheet.max_column + 1
                worksheet.cell(row=1, column=technical_col, value="Open in Technical")
            url_letter = get_column_letter(url_col)
            for r in range(2, worksheet.max_row + 1):
                worksheet.cell(
                    row=r,
                    column=technical_col,
                    value=f'=IFERROR(HYPERLINK("#Technical!A"&MATCH({url_letter}{r},Technical!A:A,0),"Open"),HYPERLINK("#Technical!A1","Open"))',
                )


def _add_header_tooltips(worksheet) -> None:
    tooltip_messages = {
        "TTFB (ms)": "Time to First Byte. Measures server responsiveness. Fix: Optimize server-side code or use a CDN.",
        "AEO Readiness Score": "Composite Answer Engine Optimization quality score. Fix: Add concise answer sections, FAQ schema, and clear question headings.",
        "Indexability Reason": "Primary reason this URL may not be indexed. Fix: Resolve noindex directives, non-200 responses, and canonical mismatches.",
        "Status Code": "HTTP status returned for the URL. Fix: Resolve 4xx/5xx errors and remove unnecessary redirect chains.",
        "SEO Health Score": "Weighted technical SEO quality score for this URL. Fix: Prioritize critical issues and improve warnings in FixPlan.",
        "Priority Score": "Issue priority score for execution order. Fix: Start with the highest values to reduce risk fastest.",
        "Severity": "Impact level of the issue. Fix: Resolve Critical first, then Warning, then Info opportunities.",
        "Word Count": "Approximate body word count depth. Fix: Expand thin pages with original, search-intent-aligned content.",
        "Canonical Type": "Canonical relationship classification. Fix: Use self-canonical on indexable pages; avoid unintended cross-canonicals.",
        "Redirect Chain Length": "Number of redirect hops before final destination. Fix: Reduce to a single hop where possible.",
    }
    headers = _header_index(worksheet)
    for header, message in tooltip_messages.items():
        col_idx = headers.get(header)
        if not col_idx:
            continue
        ref = f"{get_column_letter(col_idx)}1"
        dv = DataValidation(type="custom", formula1="TRUE", showInputMessage=True, allow_blank=True)
        dv.promptTitle = header
        dv.prompt = message
        worksheet.add_data_validation(dv)
        dv.add(ref)


def _apply_wrapped_row_heights(worksheet) -> None:
    headers = _header_index(worksheet)
    wrapped_cols = []
    for header_name in ("URL", "Final URL", "Canonical URL", "Affected URLs", "Internal Link Statuses", "How to Fix in AIOSEO"):
        col_idx = headers.get(header_name)
        if col_idx:
            wrapped_cols.append(col_idx)
    if not wrapped_cols:
        return

    for row_idx in range(2, worksheet.max_row + 1):
        max_lines = 1
        for col_idx in wrapped_cols:
            value = worksheet.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            text = str(value)
            explicit_lines = text.count("\n") + 1
            estimated_wrap_lines = max(1, int(len(text) / 50) + 1)
            max_lines = max(max_lines, explicit_lines, estimated_wrap_lines)
        if max_lines > 1:
            worksheet.row_dimensions[row_idx].height = min(120, 15 * max_lines)


def _apply_south_african_formats(worksheet) -> None:
    headers = _header_index(worksheet)
    percent_headers = {
        "Pass Rate (%)",
        "SEO Pass Rate %",
        "Error Rate % (4xx/5xx)",
        "Crawl Success Rate % (2xx)",
        "Critical URL Rate %",
        "Warning URL Rate %",
        "Pass Rate",
    }
    integer_headers = {
        "URLs Crawled",
        "Value",
        "Critical URL Count",
        "Warning URL Count",
        "Pass URLs",
        "Critical URLs",
        "Warning URLs",
        "Top Issue Affected URLs",
        "Affected Count",
        "Priority Score",
        "Est. Sprint Points",
    }
    decimal_headers = {
        "SEO Health Score",
        "AEO Readiness Score",
        "TTFB (ms)",
        "Total Request Time (ms)",
        "Avg TTFB (ms)",
    }
    date_like_tokens = ("date", "timestamp", "lastmod", "updated")

    for col_idx, cell in enumerate(worksheet[1], start=1):
        header = str(cell.value or "").strip()
        col_letter = get_column_letter(col_idx)
        rng_start = 2
        rng_end = worksheet.max_row
        if rng_end < rng_start:
            continue
        if header in percent_headers or "%" in header:
            fmt = '[$-en-ZA]0.00%'
        elif header in decimal_headers:
            fmt = '[$-en-ZA]#,##0.00'
        elif header in integer_headers:
            fmt = '[$-en-ZA]#,##0'
        elif any(token in header.lower() for token in date_like_tokens):
            fmt = '[$-en-ZA]dd/mm/yyyy hh:mm:ss'
        else:
            continue
        for row_idx in range(rng_start, rng_end + 1):
            worksheet.cell(row=row_idx, column=col_idx).number_format = fmt


def adjust_sheet_format(writer, sheet_name):
    worksheet = writer.sheets[sheet_name]
    _reorder_columns(worksheet, sheet_name)
    header_fill = PatternFill(start_color=STD_NAVY, end_color=STD_NAVY, fill_type="solid")
    header_font = Font(color=STD_WHITE, bold=True)
    bad_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    warn_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    good_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    traffic_warn_fill = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
    edge_fill = PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid")
    zebra_fill = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
    headers = [cell.value for cell in worksheet[1]]
    if sheet_name in {"FixPlan", "Main", "Technical", "AIOSEO"}:
        _apply_intelligent_sorting(worksheet, sheet_name)
    ensure_auto_filter(worksheet)
    if sheet_name != "Dashboard":
        ensure_freeze_header(worksheet)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    is_wide_sheet = worksheet.max_column >= 30
    def parse_bool(value):
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.strip().lower() in {"true", "1", "yes", "y"}
        return bool(value)
    def is_bad_header(h):
        h = (h or "").lower()
        return any(t in h for t in ["error", "broken", "missing", "noindex", "disallow", "thin", "mixed content", "cross-canonical", "issue", "non-200", "loop", "out of"])
    def is_edge_header(h):
        h = (h or "").lower()
        return any(t in h for t in ["redirect chain", "param url", "edge", "unresolved"])
    def is_good_header(h):
        h = (h or "").lower()
        return any(t in h for t in ["accessible", "match", "enabled", "complete", "present", "indexable", "coverage (%)"])
    for row_idx in range(2, worksheet.max_row + 1):
        row_has_issue = False
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            val = cell.value
            h = str(header) if header is not None else ""
            is_url_like = "url" in h.lower() or h.lower().endswith("urls") or h.lower() in {"final url", "canonical url"}
            if is_wide_sheet and is_url_like:
                cell.alignment = Alignment(wrap_text=False, shrink_to_fit=True, vertical="top")
            elif "url" in h.lower() or "hops" in h.lower() or "images" == h.lower():
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")
            if "%" in h:
                try:
                    pct = float(val)
                    if pct == 100.0:
                        cell.fill = good_fill
                    elif pct < 80:
                        cell.fill = bad_fill
                        row_has_issue = True
                    else:
                        cell.fill = warn_fill
                except Exception:
                    pass
            if h in {"Status Code", "Target Status (if crawled)"} and isinstance(val, int):
                if val >= 400:
                    cell.fill = bad_fill
                    row_has_issue = True
                elif val >= 300:
                    cell.fill = warn_fill
                elif 200 <= val < 300:
                    cell.fill = good_fill
            if isinstance(val, bool) or (isinstance(val, str) and val.strip().lower() in {"true", "false"}):
                flag = parse_bool(val)
                if is_bad_header(h):
                    cell.fill = bad_fill if flag else good_fill
                    row_has_issue = row_has_issue or flag
                elif is_good_header(h):
                    cell.fill = good_fill if flag else warn_fill
                    row_has_issue = row_has_issue or (not flag)
                elif is_edge_header(h) and flag:
                    cell.fill = edge_fill
            if h in {"Broken Internal Links Count", "Image Filename Quality Issues", "Generic Anchor Text Count"}:
                try:
                    if int(val or 0) > 0:
                        cell.fill = bad_fill
                        row_has_issue = True
                    else:
                        cell.fill = good_fill
                except Exception:
                    pass
            if h in {"Word Count Band"} and isinstance(val, str):
                band = val.lower()
                if band == "thin":
                    cell.fill = bad_fill
                    row_has_issue = True
                elif band == "ok":
                    cell.fill = warn_fill
                elif band == "strong":
                    cell.fill = good_fill
            if h in {"Indexability Reason"} and isinstance(val, str):
                if "indexable" in val.lower() and "noindex" not in val.lower():
                    cell.fill = good_fill
                else:
                    cell.fill = bad_fill
                    row_has_issue = True
            if h in {"Severity", "Severity Badge"} and isinstance(val, str):
                sev = val.strip().lower()
                if sev == "critical":
                    cell.fill = bad_fill
                    row_has_issue = True
                elif sev == "warning":
                    cell.fill = traffic_warn_fill
                    row_has_issue = True
                elif sev == "info":
                    cell.fill = edge_fill
                elif sev == "pass":
                    cell.fill = good_fill
            if h == "SEO Health Score":
                try:
                    score = float(val)
                    if score < 70:
                        cell.fill = bad_fill
                        row_has_issue = True
                    elif score < 90:
                        cell.fill = traffic_warn_fill
                        row_has_issue = True
                    else:
                        cell.fill = good_fill
                except Exception:
                    pass
            if h == "Status" and isinstance(val, str):
                st = val.strip().lower()
                if st in {"to do", "todo", "open"}:
                    cell.fill = bad_fill
                    row_has_issue = True
                elif st == "in progress":
                    cell.fill = traffic_warn_fill
                    row_has_issue = True
                elif st in {"fixed", "done", "closed"}:
                    cell.fill = good_fill
        if not row_has_issue:
            worksheet.cell(row=row_idx, column=1).fill = good_fill
        if sheet_name != "Dashboard" and row_idx % 2 == 0:
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.fill.fill_type is None:
                    cell.fill = zebra_fill
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)
    # Keep URL-like columns readable and ensure wrapped text is visible.
    headers = _header_index(worksheet)
    for header_name, width in {"URL": 45, "Final URL": 45, "Canonical URL": 45, "Affected URLs": 55, "How to Fix in AIOSEO": 55}.items():
        col_idx = headers.get(header_name)
        if col_idx:
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(
                worksheet.column_dimensions[get_column_letter(col_idx)].width or width,
                width,
            )
    if is_wide_sheet:
        for header_name in ("URL", "Final URL", "Canonical URL", "Affected URLs"):
            col_idx = headers.get(header_name)
            if col_idx:
                worksheet.column_dimensions[get_column_letter(col_idx)].width = 36
    _apply_wrapped_row_heights(worksheet)
    if sheet_name == "FixPlan":
        apply_fixplan_workflow_formatting(worksheet)
        _apply_fixplan_interactivity(writer, worksheet)
    _hide_noisy_columns(worksheet, sheet_name)
    _apply_south_african_formats(worksheet)
    apply_global_conditional_formatting(worksheet)
    _collapse_technical_deep_dive_columns(worksheet, sheet_name)
    _add_url_navigation_links(writer, worksheet, sheet_name)
    _apply_cross_sheet_links(writer, worksheet, sheet_name)
    if sheet_name == "AIOSEO":
        status_col = _header_index(worksheet).get("Status")
        if status_col:
            _apply_status_dropdown(worksheet, status_col)
    _add_back_to_dashboard_link(worksheet, sheet_name)
    _add_all_header_tooltips(worksheet)
    if sheet_name in DATA_HEAVY_TABS:
        _add_header_tooltips(worksheet)
    if sheet_name == "Dashboard":
        _style_dashboard(worksheet, writer)


def apply_tab_hyperlinks(writer):
    wb = writer.book
    link_map = {
        "Summary": "Reference Tab",
        "FixPlan": "Reference Tab",
        "Dashboard": "Target Tab",
        "AIOSEO": "Reference Tab",
        "IssueInventory": "Reference Tab",
    }
    for sheet_name, col_header in link_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        if col_header not in headers:
            continue
        col_idx = headers.index(col_header) + 1
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            target = str(cell.value or "").strip()
            if target and target in wb.sheetnames:
                cell.hyperlink = f"#{target}!A1"
                cell.style = "Hyperlink"
    preferred_first_tabs = [
        "Dashboard",
        "Legend",
        "Summary",
        "Main",
        "Technical",
        "FixPlan",
        "Priority URLs",
        "Indexability",
        "Content",
        "Redirects",
        "Links",
        "LinksDetail",
        "AEO",
        "AIOSEO",
        "SnippetCandidates",
        "Media",
        "StructuredData",
        "Social",
        "Security",
        "Duplicates",
        "TemplateClusters",
        "IssueInventory",
        "CrawlGraph",
        "SitemapQA",
        "DeltaFromPreviousRun",
        "RunMetadata",
    ]
    for idx, tab_name in enumerate(preferred_first_tabs):
        if tab_name in wb.sheetnames:
            wb.move_sheet(wb[tab_name], offset=-wb.index(wb[tab_name]) + idx)
    low_signal_tabs = {"RunMetadata", "DeltaFromPreviousRun"}
    for tab_name in low_signal_tabs:
        if tab_name in wb.sheetnames:
            wb[tab_name].sheet_state = "hidden"


__all__ = ["adjust_sheet_format", "apply_tab_hyperlinks"]
