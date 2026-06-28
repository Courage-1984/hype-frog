"""Unit tests for crawl matrix audit helpers (no live network)."""

from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[2]
SCRIPT_PATH = ROOT / "scripts" / "crawl_matrix_audit.py"


def _load_crawl_matrix_audit():
    name = "crawl_matrix_audit_test_module"
    if name in sys.modules:
        return sys.modules[name]
    spec = importlib.util.spec_from_file_location(name, SCRIPT_PATH)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


@pytest.fixture(scope="module")
def matrix_audit():
    return _load_crawl_matrix_audit()


def test_header_map_reads_first_row(matrix_audit) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main"
    ws["A1"] = "URL"
    ws["B1"] = "Title"
    assert matrix_audit._header_map(ws) == {"URL": 1, "Title": 2}
    wb.close()


def test_scan_formula_refs_flags_missing_sheet(matrix_audit) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws["A1"] = "='Missing Sheet'!A1"
    issues = matrix_audit._scan_formula_refs(wb)
    assert any("Missing Sheet" in issue for issue in issues)
    wb.close()


def test_scan_formula_refs_ignores_indirect(matrix_audit) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws["A1"] = '=INDIRECT("'"&F2&"'!A:A")'
    assert matrix_audit._scan_formula_refs(wb) == []
    wb.close()


def test_audit_main_formulas_validates_vlookup(matrix_audit) -> None:
    wb = Workbook()
    main = wb.active
    main.title = "Main"
    td = wb.create_sheet("Technical Diagnostics")
    main["A1"] = "URL"
    main["B1"] = "Technical Health"
    main["A2"] = "https://example.com/"
    main["B2"] = (
        '=IFERROR(VLOOKUP(A2,\'Technical Diagnostics\'!$A:$E,5,FALSE),"")'
    )
    td["A1"] = "URL"
    td["E1"] = "Technical Health"

    assert matrix_audit._audit_main_formulas(wb) == []
    main["B2"] = "=1"
    assert matrix_audit._audit_main_formulas(wb)
    wb.close()


def test_deep_audit_reports_missing_full_suite_sheets(matrix_audit, tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main"
    ws["A1"] = "URL"
    ws["A2"] = "https://example.com/"
    path = tmp_path / "partial.xlsx"
    wb.save(path)
    wb.close()

    scenario = matrix_audit.Scenario(
        name="fast_minimal",
        config=matrix_audit._base_config(),
    )
    report = matrix_audit.deep_audit(path, scenario)

    assert report["scenario"] == "fast_minimal"
    assert report["ok"] is False
    assert any("Missing full-suite sheet" in issue for issue in report["issues"])
