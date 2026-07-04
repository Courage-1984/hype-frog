"""Cell-content validation of the real, end-to-end assembled full-suite workbook.

`export_workbook.py::write_full_suite_workbook` has no dedicated test file; the
existing coverage in `test_export_flow.py` only asserts sheet *names* on the real
export, and cell-level assertions elsewhere are made against hand-built in-memory
workbooks, not the actual pipeline output. This file closes that gap by running
the real crawl->enrichment->export pipeline against the deterministic
`full_smoke_fixtures` generator and asserting concrete cell values.
"""

from __future__ import annotations

from contextlib import contextmanager
from pathlib import Path
from typing import Iterator
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from hype_frog.app_orchestrator import (
    _build_aeo_rows,
    _build_aioseo_rows,
    _extract_subfolder,
    _value_or_default,
)
from hype_frog.crawler.gsc_engine import GSCEnrichmentContext
from hype_frog.diagnostics.full_smoke_fixtures import (
    FullSmokeFixture,
    build_full_smoke_fixture,
    build_smoke_crawl_payload,
    full_smoke_network_patches,
)
from hype_frog.orchestration.crawl_runner import CrawlExecutionResult
from hype_frog.orchestration.enrichment_flow import run_enrichment
from hype_frog.orchestration.export_flow import execute_export
from hype_frog.orchestration.run_setup import RunSetup


def _empty_gsc_context(_target: str) -> GSCEnrichmentContext:
    return GSCEnrichmentContext({}, False, None, None)


@contextmanager
def _offline_enrichment_patches(fixture: FullSmokeFixture) -> Iterator[None]:
    with (
        full_smoke_network_patches(fixture),
        patch(
            "hype_frog.orchestration.enrichment_flow.load_gsc_enrichment_context",
            new=_empty_gsc_context,
        ),
    ):
        yield


def _build_run_setup(*, target_input: str) -> RunSetup:
    return RunSetup(
        target_input=target_input,
        max_urls=3,
        max_psi_urls=2,
        high_value_slugs=[],
        crawl_mode="fast",
        render_wait_ms=1000,
        selector_wait_ms=500,
        workers_preset=2,
        request_delay_preset=0.0,
        full_suite_preset=True,
        previous_audit_path_preset="",
        checkpoint_every_preset=0,
        resume_checkpoint_mode="no",
        check_external_link_status=True,
        check_og_images=False,
        check_content_images=False,
    )


@pytest.fixture
async def full_suite_workbook_path(tmp_path: Path) -> Path:
    """Run the real crawl->enrichment->export pipeline and return the .xlsx path.

    Uses deterministic fixture indices 0-2: index 0 (homepage) is seeded with
    status "Timeout" (Non-Indexable) by `full_smoke_fixtures._status_for_index`;
    indices 1-2 are status 200 (Indexable). Word counts are `320 + index`.
    """
    fixture = build_full_smoke_fixture(url_count=5)
    urls = list(fixture.urls[:3])
    crawl_rows = [build_smoke_crawl_payload(fixture, url) for url in urls]
    output_path = tmp_path / "full_suite_content.xlsx"
    crawl_result = CrawlExecutionResult(
        output_filename=str(output_path),
        crawl_rows=crawl_rows,
        target_input=fixture.sitemap_url,
        max_psi_urls=2,
        crawl_urls=urls,
        sitemap_meta=fixture.sitemap_meta,
        sitemap_files_meta=fixture.sitemap_files_meta,
        source_label="example.com",
        workers=2,
        request_delay=0.0,
        full_suite=True,
        previous_audit_path="",
        checkpoint_every=0,
        crawl_completed=True,
        check_external_link_status=True,
        check_og_images=False,
        check_content_images=False,
    )
    setup = _build_run_setup(target_input=fixture.sitemap_url)

    with _offline_enrichment_patches(fixture):
        enrichment = await run_enrichment(crawl_result)

    execute_export(
        setup,
        crawl_result,
        enrichment,
        value_or_default_fn=_value_or_default,
        extract_subfolder_fn=_extract_subfolder,
        build_aeo_rows_fn=_build_aeo_rows,
        build_aioseo_rows_fn=_build_aioseo_rows,
    )
    return output_path


@pytest.mark.asyncio
async def test_main_sheet_header_row_matches_column_contract(
    full_suite_workbook_path: Path,
) -> None:
    workbook = load_workbook(full_suite_workbook_path, read_only=True)
    try:
        ws = workbook["Main"]
        header = [ws.cell(row=2, column=c).value for c in range(1, 12)]
    finally:
        workbook.close()
    assert header == [
        "Health Icon",
        "URL",
        "Status Code",
        "Indexability",
        "Load Time (s)",
        "Title",
        "Meta Description",
        "Word Count (Body)",
        "SEO Health Score",
        "Severity Badge",
        "Action Needed",
    ]


@pytest.mark.asyncio
async def test_main_sheet_rows_carry_real_per_url_content(
    full_suite_workbook_path: Path,
) -> None:
    workbook = load_workbook(full_suite_workbook_path, read_only=True)
    try:
        ws = workbook["Main"]
        rows = [
            {
                "URL": ws.cell(row=r, column=2).value,
                "Indexability": ws.cell(row=r, column=4).value,
                "Title": ws.cell(row=r, column=6).value,
                "Meta Description": ws.cell(row=r, column=7).value,
                "Word Count (Body)": ws.cell(row=r, column=8).value,
                "SEO Health Score": ws.cell(row=r, column=9).value,
                "Severity Badge": ws.cell(row=r, column=10).value,
            }
            for r in (3, 4, 5)
        ]
    finally:
        workbook.close()

    by_index = {int(row["Title"].rsplit(" ", 1)[-1]): row for row in rows}

    # Index 0 (homepage): fixture seeds status "Timeout" -> Non-Indexable, and the
    # lowest SEO Health Score / Critical severity of the three rows.
    homepage = by_index[0]
    assert homepage["Indexability"] == "Non-Indexable"
    assert homepage["Meta Description"] == "Smoke meta description for page 0."
    assert homepage["Word Count (Body)"] == 320
    assert homepage["Severity Badge"] == "Critical"

    # Indices 1-2: fixture seeds status 200 -> Indexable.
    for index in (1, 2):
        row = by_index[index]
        assert row["Indexability"] == "Indexable"
        assert row["Word Count (Body)"] == 320 + index
        assert row["Title"] == f"Smoke title {index}"

    # The homepage's health score must be strictly worse (lower) than both
    # healthy pages, proving Severity Badge/SEO Health Score are actually
    # derived per-row rather than a static placeholder.
    assert homepage["SEO Health Score"] < by_index[1]["SEO Health Score"]
    assert homepage["SEO Health Score"] < by_index[2]["SEO Health Score"]


@pytest.mark.asyncio
async def test_audit_run_details_reflects_real_run_metadata(
    full_suite_workbook_path: Path,
) -> None:
    workbook = load_workbook(full_suite_workbook_path, read_only=True)
    try:
        ws = workbook["Audit Run Details"]
        details = {
            ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
            for r in range(2, ws.max_row + 1)
        }
    finally:
        workbook.close()

    assert details["Total URLs"] == 3
    assert details["Mode"] == "Full Suite"
    assert details["Crawl Mode"] == "fast"
    assert details["Extraction Source Rendered Count"] == 3
    assert details["Workers"] == 2
