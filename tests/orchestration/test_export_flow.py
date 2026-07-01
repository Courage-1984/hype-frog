"""Export orchestration helper and integration tests."""

from __future__ import annotations

from contextlib import contextmanager
from dataclasses import replace
from pathlib import Path
from typing import Iterator
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

from hype_frog.app_orchestrator import (
    _build_aeo_rows,
    _build_aioseo_rows,
    _extract_subfolder,
    _value_or_default,
)
from hype_frog.core.url_normalization import normalize_url
from hype_frog.crawler.gsc_engine import GSCEnrichmentContext
from hype_frog.diagnostics.full_smoke_fixtures import (
    FullSmokeFixture,
    build_full_smoke_fixture,
    build_smoke_crawl_payload,
    full_smoke_network_patches,
)
from hype_frog.orchestration.crawl_runner import CrawlExecutionResult
from hype_frog.orchestration.enrichment_flow import EnrichmentResult, run_enrichment
from hype_frog.orchestration.export_flow import ExportSummary, execute_export, normalize_url_key
from hype_frog.orchestration.run_setup import RunSetup
from hype_frog.reporter.workbook_audit import REQUIRED_FULL_SUITE_SHEETS, audit_workbook


def _empty_gsc_context(_target: str) -> GSCEnrichmentContext:
    return GSCEnrichmentContext({}, False, None, None)


def _build_crawl_result(
    tmp_path: Path,
    fixture: FullSmokeFixture,
    *,
    output_name: str,
    url_count: int = 3,
    full_suite: bool = True,
) -> CrawlExecutionResult:
    urls = list(fixture.urls[:url_count])
    crawl_rows = [build_smoke_crawl_payload(fixture, url) for url in urls]
    return CrawlExecutionResult(
        output_filename=str(tmp_path / output_name),
        crawl_rows=crawl_rows,
        target_input=fixture.sitemap_url,
        max_psi_urls=2,
        crawl_urls=urls,
        sitemap_meta=fixture.sitemap_meta,
        sitemap_files_meta=fixture.sitemap_files_meta,
        source_label="example.com",
        workers=2,
        request_delay=0.0,
        full_suite=full_suite,
        previous_audit_path="",
        checkpoint_every=0,
        crawl_completed=True,
        check_external_link_status=True,
        check_og_images=False,
        check_content_images=False,
    )


@contextmanager
def _offline_enrichment_patches(fixture: FullSmokeFixture) -> Iterator[None]:
    with (
        full_smoke_network_patches(fixture),
        patch(
            "hype_frog.orchestration.enrichment_flow.load_gsc_enrichment_context",
            new=_empty_gsc_context,
        ),
    ):
        yield


def _build_run_setup(*, full_suite: bool, target_input: str) -> RunSetup:
    return RunSetup(
        target_input=target_input,
        max_urls=3,
        max_psi_urls=2,
        high_value_slugs=[],
        crawl_mode="fast",
        render_wait_ms=1000,
        selector_wait_ms=500,
        workers_preset=2,
        request_delay_preset=0.0,
        full_suite_preset=full_suite,
        previous_audit_path_preset="",
        checkpoint_every_preset=0,
        resume_checkpoint_mode="no",
        check_external_link_status=True,
        check_og_images=False,
        check_content_images=False,
    )


def test_normalize_url_key_normalises_scheme_and_host() -> None:
    assert normalize_url_key("https://Example.com/Page/?q=1") == "https://example.com/Page?q=1"


def test_export_summary_dataclass_fields() -> None:
    summary = ExportSummary(
        output_filename="audit.xlsx",
        main_rows_written=10,
        extra_rows_written=10,
        full_suite=True,
    )
    assert summary.main_rows_written == 10
    assert summary.full_suite is True


def test_execute_export_main_only_writes_workbook(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.delenv("HF_EXPORT_PDF", raising=False)
    monkeypatch.delenv("HF_EXPORT_HTML", raising=False)
    fixture = build_full_smoke_fixture(url_count=2)
    url = fixture.urls[0]
    payload = build_smoke_crawl_payload(fixture, url)
    output_path = tmp_path / "main_only.xlsx"
    crawl_result = CrawlExecutionResult(
        output_filename=str(output_path),
        crawl_rows=[payload],
        target_input=fixture.sitemap_url,
        max_psi_urls=0,
        crawl_urls=[url],
        sitemap_meta=fixture.sitemap_meta,
        sitemap_files_meta=fixture.sitemap_files_meta,
        source_label="example.com",
        workers=2,
        request_delay=0.0,
        full_suite=False,
        previous_audit_path="",
        checkpoint_every=0,
        crawl_completed=True,
        check_external_link_status=False,
    )
    enrichment = EnrichmentResult(
        typed_main_rows=[payload.main],
        typed_extra_rows=[payload.extra],
        status_by_url={normalize_url(url): payload.extra.values.get("Status Code")},
        sitemap_url_keys={normalize_url(url)},
    )
    setup = _build_run_setup(full_suite=False, target_input=fixture.sitemap_url)

    summary = execute_export(
        setup,
        crawl_result,
        enrichment,
        value_or_default_fn=_value_or_default,
        extract_subfolder_fn=_extract_subfolder,
        build_aeo_rows_fn=_build_aeo_rows,
        build_aioseo_rows_fn=_build_aioseo_rows,
    )

    assert summary.full_suite is False
    assert summary.main_rows_written == 1
    assert output_path.is_file()
    workbook = load_workbook(output_path, read_only=True)
    try:
        assert "Main" in workbook.sheetnames
        assert len(workbook.sheetnames) <= 2
    finally:
        workbook.close()


@pytest.mark.asyncio
async def test_execute_export_full_suite_writes_core_sheets(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.delenv("HF_EXPORT_PDF", raising=False)
    monkeypatch.delenv("HF_EXPORT_HTML", raising=False)
    fixture = build_full_smoke_fixture(url_count=5)
    crawl_result = _build_crawl_result(
        tmp_path,
        fixture,
        output_name="full_suite.xlsx",
        url_count=3,
        full_suite=True,
    )
    setup = _build_run_setup(full_suite=True, target_input=fixture.sitemap_url)

    with _offline_enrichment_patches(fixture):
        enrichment = await run_enrichment(crawl_result)

    summary = execute_export(
        setup,
        crawl_result,
        enrichment,
        value_or_default_fn=_value_or_default,
        extract_subfolder_fn=_extract_subfolder,
        build_aeo_rows_fn=_build_aeo_rows,
        build_aioseo_rows_fn=_build_aioseo_rows,
    )

    assert summary.full_suite is True
    assert summary.main_rows_written == 3
    assert Path(crawl_result.output_filename).is_file()
    workbook = load_workbook(crawl_result.output_filename, read_only=True)
    try:
        assert "Main" in workbook.sheetnames
        assert "Dashboard" in workbook.sheetnames
        assert "Table of Contents" in workbook.sheetnames
        assert REQUIRED_FULL_SUITE_SHEETS.issubset(set(workbook.sheetnames))
    finally:
        workbook.close()

    audit_errors = audit_workbook(crawl_result.output_filename, require_full_suite_sheets=True)
    assert not any("Missing required sheets" in err for err in audit_errors)


@pytest.mark.asyncio
async def test_execute_export_streaming_write_only_full_suite(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.delenv("HF_EXPORT_PDF", raising=False)
    monkeypatch.delenv("HF_EXPORT_HTML", raising=False)
    fixture = build_full_smoke_fixture(url_count=5)
    crawl_result = _build_crawl_result(
        tmp_path,
        fixture,
        output_name="streaming_suite.xlsx",
        url_count=3,
        full_suite=True,
    )
    crawl_result = replace(crawl_result, streaming=True)
    setup = replace(
        _build_run_setup(full_suite=True, target_input=fixture.sitemap_url),
        streaming=True,
    )

    with _offline_enrichment_patches(fixture):
        enrichment = await run_enrichment(crawl_result)

    summary = execute_export(
        setup,
        crawl_result,
        enrichment,
        value_or_default_fn=_value_or_default,
        extract_subfolder_fn=_extract_subfolder,
        build_aeo_rows_fn=_build_aeo_rows,
        build_aioseo_rows_fn=_build_aioseo_rows,
    )

    assert summary.full_suite is True
    assert Path(crawl_result.output_filename).is_file()
    audit_errors = audit_workbook(crawl_result.output_filename, require_full_suite_sheets=True)
    assert not any("Missing required sheets" in err for err in audit_errors)
