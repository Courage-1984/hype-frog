"""Integration tests for --regen-report orchestration branch."""

from __future__ import annotations

from pathlib import Path
from unittest.mock import AsyncMock, patch

import pytest
from openpyxl import load_workbook

from hype_frog.core import file_utils
from hype_frog.core.cli import UserConfig
from hype_frog.app_orchestrator import main as orchestrator_main
from hype_frog.core.run_config import CliRunOverrides
from hype_frog.snapshots.models import CrawlReplaySnapshot


def _snapshot(*, full_suite: bool = False) -> CrawlReplaySnapshot:
    return CrawlReplaySnapshot(
        snapshot_id="eeeeeeee-eeee-4eee-8eee-eeeeeeeeeeee",
        domain="example.com",
        run_timestamp="2026-06-28 12:00:00",
        source_output_path="reports/latest/SEO_AEO_Audit_example.com_20260628_120000.xlsx",
        main_rows=[
            {
                "URL": "https://example.com/",
                "Title": "Regenerated Homepage",
                "Extraction State": "complete",
            }
        ],
        extra_rows=[{"URL": "https://example.com/"}],
        crawl_context={
            "target_input": "https://example.com/",
            "crawl_urls": ["https://example.com/"],
            "full_suite": full_suite,
            "workers": 3,
            "request_delay": 0.0,
            "previous_audit_path": "",
            "checkpoint_every": 0,
            "check_external_link_status": True,
            "crawl_completed": True,
            "source_label": "example.com",
            "sitemap_meta": {},
            "sitemap_files_meta": {},
        },
        enrichment_context={"status_by_url": {}, "sitemap_url_keys": []},
        setup_context={},
    )


@pytest.mark.asyncio
async def test_regen_report_skips_crawl_and_writes_real_workbook_from_snapshot(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Exercises the real `execute_export` (not mocked) so this test verifies the
    regenerated workbook actually contains the snapshot's row data, not just that
    a mock was called. `execute_crawl`/`run_enrichment`/`save_crawl_snapshot` are
    still boundary-mocked since regen-report must never trigger a live crawl —
    that remains a legitimate assertion, just no longer the *only* one."""
    monkeypatch.setattr(file_utils, "REPORTS_LATEST_DIR", tmp_path)
    monkeypatch.delenv("HF_EXPORT_PDF", raising=False)
    monkeypatch.delenv("HF_EXPORT_HTML", raising=False)
    snapshot = _snapshot(full_suite=False)
    monkeypatch.setattr(
        "hype_frog.orchestration.run_setup.get_user_config",
        lambda: UserConfig(
            target_input="https://example.com/",
            max_urls=None,
            max_psi_urls=None,
            high_value_slugs=[],
            crawl_mode="fast",
            render_wait_ms=0,
            selector_wait_ms=0,
            check_external_link_status=True,
            check_og_images=False,
            check_content_images=False,
        ),
    )
    monkeypatch.setattr(
        "hype_frog.app_orchestrator.load_latest_crawl_snapshot_for_domain",
        lambda domain: snapshot if domain == "example.com" else None,
    )

    crawl_mock = patch(
        "hype_frog.app_orchestrator.execute_crawl",
        new_callable=AsyncMock,
    )
    enrich_mock = patch(
        "hype_frog.app_orchestrator.run_enrichment",
        new_callable=AsyncMock,
    )
    save_mock = patch("hype_frog.app_orchestrator.save_crawl_snapshot")

    with crawl_mock as crawl, enrich_mock as enrich, save_mock as save:
        await orchestrator_main(
            cli_overrides=CliRunOverrides(regen_report=True),
        )
        crawl.assert_not_called()
        enrich.assert_not_called()
        save.assert_not_called()

    written_files = list(tmp_path.glob("*_regen_*.xlsx"))
    assert len(written_files) == 1
    workbook = load_workbook(written_files[0], read_only=True)
    try:
        ws = workbook["Main"]
        rows = list(ws.iter_rows(values_only=True))
        # Row 1 is a "<- Return to Executive Briefing" nav strip; row 2 is the
        # real header; data starts at row 3 (see test_export_workbook.py).
        header = rows[1]
        data_row = dict(zip(header, rows[2]))
        assert data_row["URL"] == "https://example.com/"
        assert data_row["Title"] == "Regenerated Homepage"
    finally:
        workbook.close()


@pytest.mark.asyncio
async def test_regen_report_missing_snapshot_exits(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        "hype_frog.orchestration.run_setup.get_user_config",
        lambda: UserConfig(
            target_input="https://example.com/",
            max_urls=None,
            max_psi_urls=None,
            high_value_slugs=[],
            crawl_mode="fast",
            render_wait_ms=0,
            selector_wait_ms=0,
            check_external_link_status=True,
            check_og_images=False,
            check_content_images=False,
        ),
    )
    monkeypatch.setattr(
        "hype_frog.app_orchestrator.load_latest_crawl_snapshot_for_domain",
        lambda domain: None,
    )
    with pytest.raises(SystemExit) as exc:
        await orchestrator_main(cli_overrides=CliRunOverrides(regen_report=True))
    assert exc.value.code == 1
