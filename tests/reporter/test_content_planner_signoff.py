"""Tests for Content Planner sign-off dropdown and RAG CF rules."""

from __future__ import annotations

import openpyxl
import pytest

from hype_frog.reporter.sheets import conditional as cond_module
from hype_frog.reporter.sheets.conditional import apply_content_planner_signoff_rules
from hype_frog.reporter.sheets.config import (
    RAG_AMBER,
    RAG_AMBER_FONT,
    RAG_GREEN,
    RAG_GREEN_FONT,
    RAG_RED,
    RAG_RED_FONT,
)


def _rule_count(ws: openpyxl.worksheet.worksheet.Worksheet) -> int:
    return sum(len(rules) for rules in ws.conditional_formatting._cf_rules.values())


def _worksheet_with_data(rows: int = 5) -> openpyxl.worksheet.worksheet.Worksheet:
    """Return a worksheet matching the real Planner layout: a row-1 return-strip
    banner (Content Planner is not return-strip-exempt), real headers on row 2,
    data from row 3 — not the header-on-row-1 shape a synthetic sheet defaults to."""
    wb = openpyxl.Workbook()
    ws = wb.active
    from hype_frog.orchestration.content_planner import CONTENT_PLANNER_COLUMNS

    ws.append(["<- Return to Executive Briefing"] + [None] * (len(CONTENT_PLANNER_COLUMNS) - 1))
    ws.append(list(CONTENT_PLANNER_COLUMNS))
    for i in range(rows):
        ws.append(
            ["Home" if i == 0 else None, None, None, f"https://example.com/page-{i}/"]
            + [None] * (len(CONTENT_PLANNER_COLUMNS) - 4)
        )
    return ws


# ---------------------------------------------------------------------------
# Data validation
# ---------------------------------------------------------------------------

def test_dropdown_added_to_signoff_columns() -> None:
    ws = _worksheet_with_data()
    apply_content_planner_signoff_rules(ws)
    dvs = list(ws.data_validations.dataValidation)
    assert len(dvs) >= 1
    all_formula1 = [dv.formula1 for dv in dvs]
    assert any("Signed off" in (f or "") for f in all_formula1)
    assert any("In progress" in (f or "") for f in all_formula1)
    assert any("Not signed off" in (f or "") for f in all_formula1)


def test_dropdown_disabled_flag_suppresses_validation(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(cond_module, "DISABLE_DATA_VALIDATION", True)
    ws = _worksheet_with_data()
    apply_content_planner_signoff_rules(ws)
    dvs = list(ws.data_validations.dataValidation)
    assert len(dvs) == 0


def test_no_dropdown_on_empty_sheet() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    apply_content_planner_signoff_rules(ws)
    assert list(ws.data_validations.dataValidation) == []


# ---------------------------------------------------------------------------
# Conditional formatting — rule presence
# ---------------------------------------------------------------------------

def test_cf_rules_added_to_sheet(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(cond_module, "DISABLE_CONDITIONAL_FORMATTING", False)
    ws = _worksheet_with_data()
    apply_content_planner_signoff_rules(ws)
    assert _rule_count(ws) >= 3


def test_cf_disabled_flag_adds_no_rules(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(cond_module, "DISABLE_CONDITIONAL_FORMATTING", True)
    ws = _worksheet_with_data()
    apply_content_planner_signoff_rules(ws)
    assert _rule_count(ws) == 0


def test_zebra_banding_scoped_to_identity_columns_only(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Regression: columns A-E (Primary/Secondary/Tertiary/Page link/Copy Doc)
    had zero row-differentiation, unlike every other large data sheet in the
    workbook — only the sign-off columns get RAG conditional colouring. The
    zebra rule must stay confined to A:E so it doesn't visually fight the
    existing RAG fills on the sign-off block."""
    monkeypatch.setattr(cond_module, "DISABLE_CONDITIONAL_FORMATTING", False)
    ws = _worksheet_with_data()
    apply_content_planner_signoff_rules(ws)

    zebra_ranges = [
        str(rng)
        for rng, rules in ws.conditional_formatting._cf_rules.items()
        for rule in rules
        if rule.formula and "MOD(ROW(),2)=0" in rule.formula[0]
    ]
    assert len(zebra_ranges) == 1
    assert "A3:E" in zebra_ranges[0]
    assert not any(f"{col}3:" in zebra_ranges[0] for col in "FGHIJKLMNOPQRSTU")


def test_cf_formulas_reference_first_signoff_column(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(cond_module, "DISABLE_CONDITIONAL_FORMATTING", False)
    ws = _worksheet_with_data()
    apply_content_planner_signoff_rules(ws)
    all_rules = [
        rule
        for rules in ws.conditional_formatting._cf_rules.values()
        for rule in rules
    ]
    formulas = [str(getattr(r, "formula", None) or "") for r in all_rules]
    assert any("F3" in f for f in formulas)


def test_cf_covers_three_status_values(monkeypatch: pytest.MonkeyPatch) -> None:
    """Exactly three formula-rule branches must be present (signed off / in progress / not signed off)."""
    monkeypatch.setattr(cond_module, "DISABLE_CONDITIONAL_FORMATTING", False)
    ws = _worksheet_with_data()
    apply_content_planner_signoff_rules(ws)
    formulas_text = " ".join(
        str(getattr(r, "formula", ""))
        for rules in ws.conditional_formatting._cf_rules.values()
        for r in rules
    )
    assert "signed off" in formulas_text.lower()
    assert "in progress" in formulas_text.lower()
    assert "not signed off" in formulas_text.lower()


# ---------------------------------------------------------------------------
# RAG palette consistency
# ---------------------------------------------------------------------------

def test_cf_uses_rag_green_for_signed_off(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(cond_module, "DISABLE_CONDITIONAL_FORMATTING", False)
    ws = _worksheet_with_data()
    apply_content_planner_signoff_rules(ws)
    all_rules = [
        rule
        for rules in ws.conditional_formatting._cf_rules.values()
        for rule in rules
    ]
    # In openpyxl CF rules, fill lives at rule.dxf.fill (DifferentialStyle), not rule.fill
    rgb_values = set()
    for r in all_rules:
        dxf = getattr(r, "dxf", None)
        if dxf is None:
            continue
        fill = getattr(dxf, "fill", None)
        if fill is None:
            continue
        rgb = getattr(getattr(fill, "fgColor", None), "rgb", "")
        if rgb:
            rgb_values.add(rgb)
    assert rgb_values, "CF rules must have fill colours set"
    assert any(RAG_GREEN in rgb for rgb in rgb_values)
    assert any(RAG_AMBER in rgb for rgb in rgb_values)
    assert any(RAG_RED in rgb for rgb in rgb_values)


# ---------------------------------------------------------------------------
# Column range — freeze-safe check
# ---------------------------------------------------------------------------

def test_signoff_range_starts_at_column_f() -> None:
    """DataValidation sqref must include column F (index 6) for sign-off columns."""
    ws = _worksheet_with_data()
    apply_content_planner_signoff_rules(ws)
    dvs = list(ws.data_validations.dataValidation)
    sqrefs = [str(dv.sqref) for dv in dvs]
    assert any("F3" in s for s in sqrefs)


def test_freeze_panes_lock_columns_a_through_d() -> None:
    ws = _worksheet_with_data()
    apply_content_planner_signoff_rules(ws)
    assert ws.freeze_panes == "E3"


def test_empty_signoff_cells_backfilled_to_not_signed_off() -> None:
    ws = _worksheet_with_data(rows=3)
    apply_content_planner_signoff_rules(ws)
    for row_idx in range(3, ws.max_row + 1):
        for col_idx in range(6, 22):
            assert ws.cell(row=row_idx, column=col_idx).value == "Not signed off"


def test_column_widths_applied_for_planner_headers() -> None:
    ws = _worksheet_with_data()
    apply_content_planner_signoff_rules(ws)
    assert ws.column_dimensions["A"].width == 24.0
    assert ws.column_dimensions["D"].width == 52.0
    assert ws.column_dimensions["F"].width == 18.0  # Priority for MVP
    assert ws.column_dimensions["G"].width == 20.0  # Copywriter Sign off
    assert ws.column_dimensions["U"].width == 18.0  # Plugin Audit


def test_autofilter_covers_full_planner_grid() -> None:
    ws = _worksheet_with_data(rows=4)
    apply_content_planner_signoff_rules(ws)
    assert ws.auto_filter.ref == f"A2:U{ws.max_row}"


def test_header_row_2_used_when_banner_present() -> None:
    """Regression: this function used to hardcode header_row=1 unconditionally,
    silently reading the row-1 return-strip banner as if it were headers (a
    no-op on every real column). It must resolve headers from row 2 by default."""
    ws = _worksheet_with_data(rows=2)
    apply_content_planner_signoff_rules(ws)
    # Header row (2) styling applied — proves headers were found, not skipped.
    assert ws.row_dimensions[2].height == 42
    assert ws.cell(row=2, column=1).value == "Primary"
    assert ws.cell(row=2, column=6).value == "Priority for MVP"
