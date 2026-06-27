"""Workbook audit module smoke tests."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from hype_frog.reporter.sheets.config import CONTENT_OPTIMISATION_HUB_SHEET
from hype_frog.reporter.workbook_audit import REQUIRED_FULL_SUITE_SHEETS, audit_workbook, count_main_rows


def test_audit_flags_missing_toc(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main"
    ws.append(["URL", "Extraction State"])
    ws.append(["https://example.com/", "complete"])
    path = tmp_path / "bad.xlsx"
    wb.save(path)
    errors = audit_workbook(path, require_full_suite_sheets=False)
    assert any("TOC not at index 0" in err for err in errors)


def test_count_main_rows_excludes_header(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main"
    ws.append(["URL", "Extraction State"])
    ws.append(["https://example.com/a", "complete"])
    ws.append(["https://example.com/b", "partial"])
    path = tmp_path / "main.xlsx"
    wb.save(path)
    assert count_main_rows(path) == 2


def test_audit_require_full_suite_sheets_flags_missing_core_tabs(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main"
    ws.append(["URL", "Extraction State"])
    ws.append(["https://example.com/", "complete"])
    path = tmp_path / "partial.xlsx"
    wb.save(path)

    errors = audit_workbook(path, require_full_suite_sheets=True)
    assert any("Missing required sheets" in err for err in errors)


def test_audit_require_full_suite_sheets_passes_when_core_tabs_present(tmp_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name in REQUIRED_FULL_SUITE_SHEETS:
        wb.create_sheet(sheet_name)
    main = wb["Main"]
    main.append(["URL", "Extraction State"])
    main.append(["https://example.com/", "complete"])
    hub = wb[CONTENT_OPTIMISATION_HUB_SHEET]
    hub.append(["URL"])
    path = tmp_path / "core.xlsx"
    wb.save(path)

    errors = audit_workbook(path, require_full_suite_sheets=True)
    assert not any("Missing required sheets" in err for err in errors)
