"""Excel Report UX Overhaul regression guards.

Covers the column-width contract, single-line URL policy, empty-state
messaging, the phantom ``Column_4`` fix, and the non-overlapping Executive
Briefing chart grid.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from hype_frog.reporter.sheets import executive_dashboard as ed
from hype_frog.reporter.sheets.config import RETURN_TO_BRIEFING_LABEL
from hype_frog.reporter.sheets.config import CONTENT_OPTIMISATION_HUB_SHEET
from hype_frog.reporter.sheets.layout import (
    PROSE_HEADERS,
    URL_LIKE_HEADERS,
    _MAX_COL_WIDTH,
    _MIN_COL_WIDTH,
    _PROSE_COL_WIDTH,
    _URL_COL_WIDTH,
    apply_column_widths,
    apply_content_hub_heading_group,
)
from hype_frog.reporter.sheets.conditional import (
    apply_sheet_text_wrap_columns,
    apply_wrapped_row_heights,
)
from hype_frog.reporter.sheets.navigation import add_return_to_briefing_strip
from hype_frog.reporter.sheets.sheet_rows import sheet_data_header_row
from hype_frog.reporter.sheets.tables import normalize_table_headers
from hype_frog.reporter.sheets.tables_impl import (
    _EMPTY_STATE_BY_SHEET,
    _write_empty_state_message,
)


def _banner_sheet(title: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    return wb


def test_apply_column_widths_assigns_widths_and_url_single_line() -> None:
    wb = _banner_sheet("Summary")
    ws = wb.active
    hr = sheet_data_header_row("Summary")
    assert hr == 2
    ws.cell(row=1, column=1, value=RETURN_TO_BRIEFING_LABEL)
    ws.cell(row=hr, column=1, value="URL")
    ws.cell(row=hr, column=2, value="Issue")
    ws.cell(row=hr, column=3, value="Why It Matters")
    ws.cell(
        row=hr + 1,
        column=1,
        value="https://example.org/a/very/long/path/that/would/wrap/badly",
    )
    ws.cell(row=hr + 1, column=2, value="Broken link")
    ws.cell(row=hr + 1, column=3, value="x" * 200)

    apply_column_widths(ws)

    # URL columns are a fixed single-line width, never content-driven.
    assert ws.column_dimensions["A"].width == pytest.approx(_URL_COL_WIDTH)
    # Long prose columns get the generous wrapped width, capped.
    assert ws.column_dimensions["C"].width == pytest.approx(_PROSE_COL_WIDTH)
    # Ordinary columns get a concrete, clamped width (the old routine left None).
    width_b = ws.column_dimensions["B"].width
    assert width_b is not None
    assert _MIN_COL_WIDTH <= width_b <= _MAX_COL_WIDTH


def test_apply_column_widths_wraps_prose_column_body() -> None:
    """Prose columns must get ``wrap_text`` (not just the wide width), so long
    guidance reflows onto multiple lines instead of clipping at the column edge.
    URL columns stay single-line (no wrap)."""
    wb = _banner_sheet("Summary")
    ws = wb.active
    hr = sheet_data_header_row("Summary")
    ws.cell(row=1, column=1, value=RETURN_TO_BRIEFING_LABEL)
    ws.cell(row=hr, column=1, value="URL")
    ws.cell(row=hr, column=2, value="Why It Matters")
    ws.cell(row=hr + 1, column=1, value="https://example.org/")
    ws.cell(row=hr + 1, column=2, value="x" * 200)

    apply_column_widths(ws)

    assert ws.cell(row=hr + 1, column=2).alignment.wrap_text is True
    # URL column stays single-line by contract.
    assert ws.cell(row=hr + 1, column=1).alignment.wrap_text in (False, None)


def test_apply_column_widths_reads_resolved_header_row_not_row_one() -> None:
    wb = _banner_sheet("Priority URLs")
    ws = wb.active
    ws.cell(row=1, column=1, value=RETURN_TO_BRIEFING_LABEL)
    ws.cell(row=2, column=1, value="URL")
    ws.cell(row=3, column=1, value="https://example.org/")
    apply_column_widths(ws)
    # If the pass had assumed row 1 it would never recognise the URL header.
    assert ws.column_dimensions["A"].width == pytest.approx(_URL_COL_WIDTH)


def test_url_and_prose_header_sets_are_disjoint() -> None:
    assert not (URL_LIKE_HEADERS & PROSE_HEADERS)


def _hub_sheet_with_data() -> tuple:
    """Content Optimisation Hub headers on row 1, one data row on row 2.

    Mirrors the physical layout at the point in ``adjust_sheet_format`` where
    ``apply_column_widths`` runs BEFORE ``apply_content_hub_conditional_rules``
    inserts the row-1 banner — the exact state that produced the "Current
    Title"/"Priority Reason" default-width regression.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = CONTENT_OPTIMISATION_HUB_SHEET
    headers = [
        "Action Required",
        "On-Page Optimisation Score",
        "Current Title",
        "Priority Reason",
        "H2",
        "H2 Health",
        "H3",
        "H3 Health",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
        ws.cell(row=2, column=col_idx, value="x")
    return wb, ws, headers


def test_apply_column_widths_hub_pre_insert_still_resolves_real_headers() -> None:
    """Regression: Current Title / Priority Reason must not fall back to
    Excel's default width when called before the Hub's banner-row insert."""
    _wb, ws, headers = _hub_sheet_with_data()
    apply_column_widths(ws)
    ct_col = headers.index("Current Title") + 1
    pr_col = headers.index("Priority Reason") + 1
    ct_width = ws.column_dimensions[get_column_letter(ct_col)].width
    pr_width = ws.column_dimensions[get_column_letter(pr_col)].width
    assert ct_width is not None and ct_width >= 30.0
    assert pr_width == pytest.approx(_PROSE_COL_WIDTH)


def test_apply_column_widths_hub_post_insert_still_correct() -> None:
    """Same contract holds after the banner-row insert shifts headers to row 2."""
    _wb, ws, headers = _hub_sheet_with_data()
    ws.insert_rows(1)
    apply_column_widths(ws)
    ct_col = headers.index("Current Title") + 1
    pr_col = headers.index("Priority Reason") + 1
    ct_width = ws.column_dimensions[get_column_letter(ct_col)].width
    pr_width = ws.column_dimensions[get_column_letter(pr_col)].width
    assert ct_width is not None and ct_width >= 30.0
    assert pr_width == pytest.approx(_PROSE_COL_WIDTH)


def test_apply_wrapped_row_heights_resolves_hub_header_row_before_banner_insert() -> None:
    """Regression: ``apply_wrapped_row_heights`` runs (via ``adjust_sheet_format``)
    BEFORE ``apply_content_hub_conditional_rules`` inserts the Hub's row-1
    banner, so headers are still physically on row 1 at call time. A
    hardcoded row-2 lookup silently read the first data row as headers,
    matched nothing in PROSE_HEADERS, and left every Hub row's height unset —
    long "Recommended Action"/"Priority Reason" text would then rely entirely
    on Excel's own (unreliable) auto-fit instead of an explicit height."""
    wb = Workbook()
    ws = wb.active
    ws.title = CONTENT_OPTIMISATION_HUB_SHEET
    ws.cell(row=1, column=1, value="Action Required")
    ws.cell(row=1, column=2, value="Recommended Action")
    ws.cell(row=2, column=2, value="x" * 300)

    apply_wrapped_row_heights(ws)

    assert ws.row_dimensions[2].height is not None
    assert ws.row_dimensions[2].height > 15


def test_apply_wrapped_row_heights_hub_post_insert_still_correct() -> None:
    """Same contract holds after the banner-row insert shifts headers to row 2."""
    wb = Workbook()
    ws = wb.active
    ws.title = CONTENT_OPTIMISATION_HUB_SHEET
    ws.cell(row=1, column=1, value="banner")
    ws.cell(row=2, column=1, value="Action Required")
    ws.cell(row=2, column=2, value="Recommended Action")
    ws.cell(row=3, column=2, value="x" * 300)

    apply_wrapped_row_heights(ws)

    assert ws.row_dimensions[3].height is not None
    assert ws.row_dimensions[3].height > 15


def test_apply_content_hub_heading_group_collapses_h2_through_h6() -> None:
    _wb, ws, headers = _hub_sheet_with_data()
    ws.insert_rows(1)
    apply_column_widths(ws)
    apply_content_hub_heading_group(ws)
    h2_col = headers.index("H2") + 1
    dim = ws.column_dimensions[get_column_letter(h2_col)]
    assert dim.outline_level == 1
    assert dim.hidden is True


def test_apply_content_hub_heading_group_is_a_no_op_off_hub() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main"
    ws.cell(row=1, column=1, value="H2")
    apply_content_hub_heading_group(ws)
    assert ws.column_dimensions["A"].outline_level == 0


def test_empty_state_message_written_under_headers() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "FixPlan"
    ws.cell(row=1, column=1, value=RETURN_TO_BRIEFING_LABEL)
    ws.cell(row=2, column=1, value="Issue Type")
    ws.cell(row=2, column=2, value="Severity")
    ws.cell(row=2, column=3, value="Affected Count")

    _write_empty_state_message(ws, header_row=2)

    assert ws.cell(row=3, column=1).value == _EMPTY_STATE_BY_SHEET["FixPlan"]
    assert ws.cell(row=3, column=1).font.italic is True
    assert ws.cell(row=3, column=1).fill.start_color.rgb.endswith("E6F4EA")


@pytest.mark.parametrize(
    "sheet_name",
    [
        "Issue Register",
        "Priority URLs",
        "Broken Link Impact",
        "Snippet Opportunities",
        "Competitor Benchmarks",
    ],
)
def test_empty_state_message_covers_expanded_sheet_list(sheet_name: str) -> None:
    """Regression: empty-state guidance previously only existed for FixPlan
    and Quick Wins — other sheets that can legitimately be empty in a given
    run fell back to a bare header grid with no explanation."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.cell(row=1, column=1, value=RETURN_TO_BRIEFING_LABEL)
    ws.cell(row=2, column=1, value="URL")
    ws.cell(row=2, column=2, value="Issue")

    _write_empty_state_message(ws, header_row=2)

    assert ws.cell(row=3, column=1).value == _EMPTY_STATE_BY_SHEET[sheet_name]
    assert ws.cell(row=3, column=1).font.italic is True


@pytest.mark.parametrize(
    ("sheet_name", "narrative_header"),
    [
        ("Issue Register", "Affected URLs Sample"),
        ("Broken Link Impact", "Recommended Action"),
    ],
)
def test_narrative_columns_get_wrap_on_expanded_sheets(
    sheet_name: str, narrative_header: str
) -> None:
    """Regression: Issue Register/Broken Link Impact's narrative columns
    previously got no wrap/row-height treatment, unlike Technical/AEO/Content
    Hub Metrics, so long text spilled or got clipped instead of wrapping."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.cell(row=1, column=1, value=RETURN_TO_BRIEFING_LABEL)
    ws.cell(row=2, column=1, value=narrative_header)
    ws.cell(row=3, column=1, value="A long narrative value that should wrap onto multiple lines.")

    apply_sheet_text_wrap_columns(ws, sheet_name)

    assert ws.cell(row=3, column=1).alignment.wrap_text is True


def test_empty_state_message_skipped_when_data_present() -> None:
    wb = Workbook()
    ws = wb.active
    ws.cell(row=2, column=1, value="Issue Type")
    ws.cell(row=3, column=1, value="Broken internal link")

    _write_empty_state_message(ws, header_row=2)

    assert ws.cell(row=3, column=1).value == "Broken internal link"


def test_narrow_banner_sheet_has_no_phantom_column_4() -> None:
    wb = _banner_sheet("FixPlan")
    ws = wb.active
    ws.cell(row=1, column=1, value="Issue Type")
    ws.cell(row=1, column=2, value="Severity")
    ws.cell(row=1, column=3, value="Affected Count")

    add_return_to_briefing_strip(ws, "FixPlan")
    normalize_table_headers(ws, header_row=2)

    headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    assert "Column_4" not in headers
    assert ws.max_column == 3


def test_executive_briefing_chart_bands_do_not_overlap() -> None:
    bands = [
        ed._ROW_CH_HEALTH,
        ed._ROW_CH_ISSUES,
        ed._ROW_CH_ACTIONS,
        ed._ROW_CH_TOP_ISSUES,
    ]
    # Each chart is ~8.4 cm tall (~16-17 rows); require a clear gap between bands.
    for upper, lower in zip(bands, bands[1:]):
        assert lower - upper >= 17
    # The triage matrix sits below the final chart band, and the chart source
    # tables sit below the triage matrix.
    assert ed._BRIEFING_TRIAGE_START_ROW >= bands[-1] + 17
    assert ed.CHART_SOURCE_FIRST_ROW > ed._BRIEFING_TRIAGE_START_ROW
