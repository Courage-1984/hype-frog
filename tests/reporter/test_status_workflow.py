"""Phase 5 — unified workflow Status options and conditional formatting."""

from __future__ import annotations

from openpyxl import Workbook

from hype_frog.reporter.engine_formatting import apply_workflow_status_conditional_formatting
from hype_frog.reporter.sheets.config import (
    CONTENT_HUB_DATA_START_ROW,
    STATUS_OPTIONS,
    status_validation_list_formula,
)
from hype_frog.reporter.sheets.validation import apply_workflow_status_dropdown


def test_status_options_canonical_tuple() -> None:
    assert STATUS_OPTIONS == ("To Do", "In Progress", "In Review", "Done")


def test_status_validation_list_formula_matches_options() -> None:
    assert status_validation_list_formula() == '"To Do,In Progress,In Review,Done"'


def test_workflow_status_dropdown_uses_unified_list() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "FixPlan"
    ws.cell(row=2, column=1, value="URL")
    ws.cell(row=2, column=2, value="Status")
    ws.cell(row=3, column=1, value="https://example.com/a")

    apply_workflow_status_dropdown(ws, status_col=2, header_row=2)

    assert ws.data_validations.dataValidation
    dv = ws.data_validations.dataValidation[0]
    assert dv.formula1 == status_validation_list_formula()
    assert dv.sqref == "B3:B3"


def test_workflow_status_conditional_formatting_registers_three_rules() -> None:
    wb = Workbook()
    ws = wb.active
    ws.cell(row=CONTENT_HUB_DATA_START_ROW, column=6, value="To Do")

    apply_workflow_status_conditional_formatting(
        ws, status_col=6, first_row=CONTENT_HUB_DATA_START_ROW, last_row=5
    )

    rules = next(iter(ws.conditional_formatting._cf_rules.values()))
    assert len(rules) == 3
