"""Main sheet Performance / CWV column grouping and reorder contract."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from hype_frog.reporter.sheets.config import MAIN_TRIAGE_COLUMN_COUNT
from hype_frog.reporter.sheets.layout import (
    MAIN_COLUMN_GROUP_DEFINITIONS,
    PERFORMANCE_CWV_GROUP_COLUMNS,
    apply_main_triage_column_layout,
    reorder_columns,
)


def test_performance_cwv_group_lists_all_part3_columns_in_order() -> None:
    group = MAIN_COLUMN_GROUP_DEFINITIONS["Performance & CWV Group"]
    assert group == list(PERFORMANCE_CWV_GROUP_COLUMNS)
    assert "CrUX Level" in group
    assert "Lighthouse Performance (Mobile)" in group
    assert "Page Size (KB)" in group
    assert "Uses Modern Image Formats" in group
    assert "Reachable from Homepage" in group
    assert group.index("Uses Modern Image Formats") < group.index("Reachable from Homepage")
    assert group.index("CrUX Level") < group.index("Origin CrUX LCP (s)")
    assert group.index("Mobile PSI Score") < group.index("Lighthouse Performance (Mobile)")
    assert group.index("Lighthouse Performance (Mobile)") < group.index("Lab LCP (Mobile) (s)")


def test_reorder_columns_places_performance_block_after_sprint() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main"
    headers = [
        "Health Icon",
        "URL",
        "Sprint",
        "GSC Clicks",
        "Lab TBT (Mobile) (ms)",
        "CrUX Level",
        "CWV LCP (s)",
        "Title",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
        ws.cell(row=2, column=col_idx, value=f"val-{header}")

    reorder_columns(ws, "Main")

    reordered = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert reordered[0] == "Health Icon"
    assert reordered[1] == "URL"
    assert reordered.index("Sprint") < reordered.index("CWV LCP (s)")
    assert reordered.index("CWV LCP (s)") < reordered.index("CrUX Level")
    assert reordered.index("CrUX Level") < reordered.index("Lab TBT (Mobile) (ms)")
    assert reordered.index("Lab TBT (Mobile) (ms)") < reordered.index("GSC Clicks")


def test_main_triage_layout_collapses_columns_after_action_needed() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main"
    headers = [
        "Health Icon",
        "URL",
        "Status Code",
        "Indexability",
        "Load Time (s)",
        "Title",
        "Meta Description",
        "Word Count (Body)",
        "SEO Health Score",
        "Severity Badge",
        "Action Needed",
        "Owner",
        "Status",
        "GSC Clicks",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col_idx, value=header)

    apply_main_triage_column_layout(ws, header_row=2)

    for col in range(1, MAIN_TRIAGE_COLUMN_COUNT + 1):
        letter = get_column_letter(col)
        assert ws.column_dimensions[letter].hidden is not True

    first_advanced = get_column_letter(MAIN_TRIAGE_COLUMN_COUNT + 1)
    assert ws.column_dimensions[first_advanced].hidden is True
