"""Tests for `reporter/engine_io.py` — workbook-safe streaming I/O and sanitization.

Before this file, none of this module's 16 public functions had a direct test;
they were only exercised transitively whenever the full-suite export path ran.

Note on `_sanitize_excel_value`: it strips non-printable/control characters
but, as currently written, does NOT strip leading formula-injection prefixes
(`=`, `+`, `-`, `@`) from arbitrary string values — that stripping only exists
in `narrative_engine._sanitize_cell_text` for AI-generated narrative text.
`reporter/CLAUDE.md` describes formula-prefix stripping as mandatory "on every
cell write," which reads broader than the current implementation. The tests
below document current behavior faithfully rather than assert a stronger
guarantee that doesn't exist yet.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from hype_frog.checkpoint.cache import AuditCache
from hype_frog.core.models import ExtraRowPayload, MainRowPayload
from hype_frog.reporter.engine_io import (
    apply_link_intelligence_summary_broken_formulas,
    build_core_dataframes,
    load_cached_rows,
    write_cached_sheet_chunked,
    write_dataframe_sheet,
    write_dict_rows_sheet,
    write_link_inventory_sheet_streamed,
)
from hype_frog.reporter.stream_workbook import StreamingExcelWriter


# ---------------------------------------------------------------------------
# load_cached_rows / build_core_dataframes
# ---------------------------------------------------------------------------

@pytest.fixture
def populated_cache(tmp_path: Path) -> AuditCache:
    cache = AuditCache(str(tmp_path / "cache.sqlite"))
    cache.upsert_results(
        [
            {
                "main": {"URL": "https://example.com/a", "Title": "Page A"},
                "extra": {"URL": "https://example.com/a", "Status Code": 200},
            },
            {
                "main": {"URL": "https://example.com/b", "Title": "Page B"},
                "extra": {"URL": "https://example.com/b", "Status Code": 404},
            },
        ]
    )
    return cache


def test_load_cached_rows_splits_main_and_extra(populated_cache: AuditCache) -> None:
    main_rows, extra_rows = load_cached_rows(populated_cache)
    assert len(main_rows) == 2
    assert len(extra_rows) == 2
    titles = {row["Title"] for row in main_rows}
    assert titles == {"Page A", "Page B"}
    statuses = {row["Status Code"] for row in extra_rows}
    assert statuses == {200, 404}


def test_build_core_dataframes_returns_matching_dataframes_and_lists(
    populated_cache: AuditCache,
) -> None:
    main_df, extra_df, main_rows, extra_rows = build_core_dataframes(populated_cache)
    assert isinstance(main_df, pd.DataFrame)
    assert isinstance(extra_df, pd.DataFrame)
    assert len(main_df) == 2
    assert len(extra_df) == 2
    assert set(main_df["Title"]) == {row["Title"] for row in main_rows}
    assert list(extra_df["Status Code"]) == [row["Status Code"] for row in extra_rows]


# ---------------------------------------------------------------------------
# write_dict_rows_sheet (write-only streaming path)
# ---------------------------------------------------------------------------

def test_write_dict_rows_sheet_writes_header_and_rows(tmp_path: Path) -> None:
    writer = StreamingExcelWriter(tmp_path / "out.xlsx")
    rows = [
        {"URL": "https://example.com/a", "Title": "Page A"},
        {"URL": "https://example.com/b", "Title": "Page B"},
    ]
    write_dict_rows_sheet(writer, "Main", ["URL", "Title"], rows)
    writer.close()

    wb = load_workbook(tmp_path / "out.xlsx", read_only=True)
    try:
        ws = wb["Main"]
        assert [ws.cell(1, c).value for c in (1, 2)] == ["URL", "Title"]
        assert ws.cell(2, 1).value == "https://example.com/a"
        assert ws.cell(3, 2).value == "Page B"
    finally:
        wb.close()
    assert writer.sheet_row_count("Main") == 3


def test_write_dict_rows_sheet_accepts_typed_payloads(tmp_path: Path) -> None:
    writer = StreamingExcelWriter(tmp_path / "out.xlsx")
    payload = MainRowPayload.model_validate({"values": {"URL": "https://example.com/", "Title": "Home"}})
    write_dict_rows_sheet(writer, "Main", ["URL", "Title"], [payload])
    writer.close()

    wb = load_workbook(tmp_path / "out.xlsx", read_only=True)
    try:
        ws = wb["Main"]
        assert ws.cell(2, 1).value == "https://example.com/"
        assert ws.cell(2, 2).value == "Home"
    finally:
        wb.close()


def test_write_dict_rows_sheet_writes_header_only_when_no_rows(tmp_path: Path) -> None:
    writer = StreamingExcelWriter(tmp_path / "out.xlsx")
    write_dict_rows_sheet(writer, "Main", ["URL", "Title"], [])
    writer.close()

    wb = load_workbook(tmp_path / "out.xlsx", read_only=True)
    try:
        ws = wb["Main"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows == [("URL", "Title")]
    finally:
        wb.close()


def test_write_dict_rows_sheet_infers_columns_when_not_given(tmp_path: Path) -> None:
    writer = StreamingExcelWriter(tmp_path / "out.xlsx")
    rows = [{"URL": "https://example.com/", "Extra Col": "value"}]
    write_dict_rows_sheet(writer, "Main", [], rows)
    writer.close()

    wb = load_workbook(tmp_path / "out.xlsx", read_only=True)
    try:
        ws = wb["Main"]
        assert [ws.cell(1, c).value for c in (1, 2)] == ["URL", "Extra Col"]
    finally:
        wb.close()


def test_write_dict_rows_sheet_strips_illegal_control_characters(tmp_path: Path) -> None:
    writer = StreamingExcelWriter(tmp_path / "out.xlsx")
    rows = [{"URL": "https://example.com/\x00\x01bad"}]
    write_dict_rows_sheet(writer, "Main", ["URL"], rows)
    writer.close()

    wb = load_workbook(tmp_path / "out.xlsx", read_only=True)
    try:
        ws = wb["Main"]
        assert ws.cell(2, 1).value == "https://example.com/bad"
    finally:
        wb.close()


def test_write_dict_rows_sheet_converts_none_to_empty_string(tmp_path: Path) -> None:
    """`_sanitize_excel_value(None)` returns "" before the write; openpyxl then
    round-trips an empty-string cell back as None on reload, so the observable
    contract from disk is "blank", not the literal string "".
    """
    writer = StreamingExcelWriter(tmp_path / "out.xlsx")
    rows = [{"URL": "https://example.com/", "Title": None}]
    write_dict_rows_sheet(writer, "Main", ["URL", "Title"], rows)
    writer.close()

    wb = load_workbook(tmp_path / "out.xlsx", read_only=True)
    try:
        ws = wb["Main"]
        assert ws.cell(2, 2).value in (None, "")
    finally:
        wb.close()


def test_write_dict_rows_sheet_does_not_strip_formula_injection_prefix(
    tmp_path: Path,
) -> None:
    """Documents current behavior: only control chars are stripped here, not
    leading =/+/-/@ characters (that guard lives only in narrative_engine)."""
    writer = StreamingExcelWriter(tmp_path / "out.xlsx")
    rows = [{"Title": "=cmd|'/c calc'!A1"}]
    write_dict_rows_sheet(writer, "Main", ["Title"], rows)
    writer.close()

    wb = load_workbook(tmp_path / "out.xlsx", read_only=True)
    try:
        ws = wb["Main"]
        assert ws.cell(2, 1).value == "=cmd|'/c calc'!A1"
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# write_dataframe_sheet (write-only path)
# ---------------------------------------------------------------------------

def test_write_dataframe_sheet_write_only_path(tmp_path: Path) -> None:
    writer = StreamingExcelWriter(tmp_path / "out.xlsx")
    df = pd.DataFrame({"URL": ["https://example.com/"], "Score": [95]})
    write_dataframe_sheet(writer, df, "Scores")
    writer.close()

    wb = load_workbook(tmp_path / "out.xlsx", read_only=True)
    try:
        ws = wb["Scores"]
        assert [ws.cell(1, c).value for c in (1, 2)] == ["URL", "Score"]
        assert ws.cell(2, 2).value == 95
    finally:
        wb.close()


def test_write_dataframe_sheet_empty_dataframe_without_header_writes_nothing(
    tmp_path: Path,
) -> None:
    writer = StreamingExcelWriter(tmp_path / "out.xlsx")
    ws = writer.book.create_sheet(title="Empty")
    writer.sheets["Empty"] = ws
    write_dataframe_sheet(writer, pd.DataFrame(), "Empty", include_header=False)
    writer.close()

    wb = load_workbook(tmp_path / "out.xlsx", read_only=True)
    try:
        rows = list(wb["Empty"].iter_rows(values_only=True))
        assert rows == []
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# apply_link_intelligence_summary_broken_formulas
# ---------------------------------------------------------------------------

def test_apply_link_intelligence_summary_broken_formulas_writes_formula() -> None:
    wb = Workbook()
    li = wb.active
    li.title = "Link Intelligence"
    li.append(["Record Type", "Broken Internal Links Count", "Actionable Fixes"])
    li.append(["Summary", None, None])
    wb.create_sheet("Link Inventory")

    apply_link_intelligence_summary_broken_formulas(wb)

    formula = li.cell(row=2, column=2).value
    assert formula.startswith("=")
    # Self-referencing via ROW()/INDIRECT rather than a baked-in "$A2" literal,
    # so the reference stays correct even after a later insert_rows() shift
    # (see the regression test below — this is the H3 audit fix).
    assert 'INDIRECT("$A"&ROW())' in formula
    action_formula = li.cell(row=2, column=3).value
    assert action_formula.startswith("=IF(B2>0")


def test_apply_link_intelligence_summary_broken_formulas_survives_later_row_insert() -> None:
    """Regression: formulas must still reference their OWN row's URL cell even
    after a later worksheet.insert_rows(1) (e.g. from add_return_to_briefing_strip),
    which shifts cells without rewriting formula text."""
    wb = Workbook()
    li = wb.active
    li.title = "Link Intelligence"
    li.append(["Record Type", "Broken Internal Links Count", "Actionable Fixes"])
    li.append(["Summary", None, None])
    wb.create_sheet("Link Inventory")

    apply_link_intelligence_summary_broken_formulas(wb)
    formula_before = li.cell(row=2, column=2).value

    li.insert_rows(1)  # simulates add_return_to_briefing_strip's banner-row insert
    formula_after = li.cell(row=3, column=2).value  # same logical row, now shifted

    # The formula text itself is untouched by insert_rows (openpyxl doesn't
    # rewrite formulas), but because it resolves its own row via ROW() at
    # calc-time rather than a baked-in row number, it still points at the
    # correct (now-shifted) row's own URL cell.
    assert formula_before == formula_after
    assert 'INDIRECT("$A"&ROW())' in formula_after


def test_apply_link_intelligence_summary_broken_formulas_skips_non_summary_rows() -> None:
    wb = Workbook()
    li = wb.active
    li.title = "Link Intelligence"
    li.append(["Record Type", "Broken Internal Links Count"])
    li.append(["Detail", None])
    wb.create_sheet("Link Inventory")

    apply_link_intelligence_summary_broken_formulas(wb)

    assert li.cell(row=2, column=2).value is None


def test_apply_link_intelligence_summary_broken_formulas_noop_without_both_sheets() -> None:
    wb = Workbook()
    wb.active.title = "Link Intelligence"
    # No "Link Inventory" sheet present -> function must return without error.
    apply_link_intelligence_summary_broken_formulas(wb)


# ---------------------------------------------------------------------------
# write_cached_sheet_chunked
# ---------------------------------------------------------------------------

def test_write_cached_sheet_chunked_streams_all_rows(
    tmp_path: Path, populated_cache: AuditCache
) -> None:
    writer = StreamingExcelWriter(tmp_path / "out.xlsx")
    write_cached_sheet_chunked(
        writer, populated_cache, "Main", ["URL", "Title"], "main", chunk_size=1
    )
    writer.close()

    wb = load_workbook(tmp_path / "out.xlsx", read_only=True)
    try:
        ws = wb["Main"]
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 3  # header + 2 rows
        urls = {row[0] for row in rows[1:]}
        assert urls == {"https://example.com/a", "https://example.com/b"}
    finally:
        wb.close()
    assert writer.sheet_row_count("Main") == 3


# ---------------------------------------------------------------------------
# write_link_inventory_sheet_streamed
# ---------------------------------------------------------------------------

class _FakeLinkInventoryCache:
    def __init__(self, rows: list[dict[str, object]]) -> None:
        self._rows = rows

    def iter_rows(self, *, chunk_size: int) -> list[list[dict[str, object]]]:
        return [self._rows[i : i + chunk_size] for i in range(0, len(self._rows), chunk_size)]


def test_write_link_inventory_sheet_streamed_writes_all_rows_and_returns_count(
    tmp_path: Path,
) -> None:
    writer = StreamingExcelWriter(tmp_path / "out.xlsx")
    cache = _FakeLinkInventoryCache(
        [
            {"Source URL": "https://example.com/a", "Target URL": "https://example.com/b"},
            {"Source URL": "https://example.com/b", "Target URL": "https://example.com/c"},
        ]
    )
    rows_written = write_link_inventory_sheet_streamed(
        writer,
        cache,
        sheet_name="Link Inventory",
        columns=["Source URL", "Target URL"],
        chunk_size=1,
    )
    writer.close()

    assert rows_written == 3  # header + 2 rows
    wb = load_workbook(tmp_path / "out.xlsx", read_only=True)
    try:
        ws = wb["Link Inventory"]
        assert ws.cell(2, 1).value == "https://example.com/a"
        assert ws.cell(3, 2).value == "https://example.com/c"
    finally:
        wb.close()
