"""Tests for `reporter/stream_workbook.py`, the write_only workbook adapter.

Before this file, no test imported `StreamingExcelWriter`, `FormattingWorkbookWriter`,
`is_write_only_writer`, or `reopen_workbook_for_formatting` directly — they were
only exercised transitively whenever `export_workbook.py`'s streaming path ran.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from hype_frog.reporter.stream_workbook import (
    FormattingWorkbookWriter,
    StreamingExcelWriter,
    is_write_only_writer,
    reopen_workbook_for_formatting,
)


def test_streaming_excel_writer_is_write_only(tmp_path: Path) -> None:
    writer = StreamingExcelWriter(tmp_path / "out.xlsx")
    assert writer.write_only is True
    assert writer.book.write_only is True


def test_streaming_excel_writer_tracks_row_counts_per_sheet(tmp_path: Path) -> None:
    writer = StreamingExcelWriter(tmp_path / "out.xlsx")
    assert writer.sheet_row_count("Main") == 0
    writer.record_rows_appended("Main", 5)
    assert writer.sheet_row_count("Main") == 5
    writer.record_rows_appended("Main", 3)
    assert writer.sheet_row_count("Main") == 8
    assert writer.sheet_row_count("Other") == 0


def test_streaming_excel_writer_close_saves_real_file(tmp_path: Path) -> None:
    output_path = tmp_path / "streamed.xlsx"
    writer = StreamingExcelWriter(output_path)
    ws = writer.book.create_sheet(title="Main")
    ws.append(["URL", "Title"])
    ws.append(["https://example.com/", "Example"])
    writer.close()

    assert output_path.is_file()
    wb = load_workbook(output_path, read_only=True)
    try:
        ws_reloaded = wb["Main"]
        assert ws_reloaded.cell(row=1, column=1).value == "URL"
        assert ws_reloaded.cell(row=2, column=2).value == "Example"
    finally:
        wb.close()


def test_streaming_excel_writer_close_is_idempotent(tmp_path: Path) -> None:
    output_path = tmp_path / "streamed_twice.xlsx"
    writer = StreamingExcelWriter(output_path)
    writer.book.create_sheet(title="Main")
    writer.close()
    mtime_after_first_close = output_path.stat().st_mtime_ns
    writer.close()  # must not raise or re-save
    assert output_path.stat().st_mtime_ns == mtime_after_first_close


def test_streaming_excel_writer_save_is_alias_for_close(tmp_path: Path) -> None:
    output_path = tmp_path / "saved.xlsx"
    writer = StreamingExcelWriter(output_path)
    writer.book.create_sheet(title="Main")
    writer.save()
    assert output_path.is_file()


def test_formatting_workbook_writer_loads_sheets_by_name(tmp_path: Path) -> None:
    output_path = tmp_path / "existing.xlsx"
    wb = Workbook()
    wb.active.title = "Main"
    wb.create_sheet("Summary")
    wb.save(output_path)

    writer = FormattingWorkbookWriter(output_path)
    assert set(writer.sheets.keys()) == {"Main", "Summary"}


def test_formatting_workbook_writer_close_persists_edits(tmp_path: Path) -> None:
    output_path = tmp_path / "editable.xlsx"
    wb = Workbook()
    wb.active.title = "Main"
    wb.save(output_path)

    writer = FormattingWorkbookWriter(output_path)
    writer.sheets["Main"]["A1"] = "Edited"
    writer.close()

    reopened = load_workbook(output_path, read_only=True)
    try:
        assert reopened["Main"]["A1"].value == "Edited"
    finally:
        reopened.close()


def test_reopen_workbook_for_formatting_returns_formatting_writer(tmp_path: Path) -> None:
    output_path = tmp_path / "reopen.xlsx"
    Workbook().save(output_path)

    writer = reopen_workbook_for_formatting(output_path)
    assert isinstance(writer, FormattingWorkbookWriter)
    assert writer.path == str(output_path)


def test_is_write_only_writer_true_for_streaming_writer(tmp_path: Path) -> None:
    writer = StreamingExcelWriter(tmp_path / "out.xlsx")
    assert is_write_only_writer(writer) is True


def test_is_write_only_writer_false_for_formatting_writer(tmp_path: Path) -> None:
    output_path = tmp_path / "rw.xlsx"
    Workbook().save(output_path)
    writer = FormattingWorkbookWriter(output_path)
    assert is_write_only_writer(writer) is False


def test_is_write_only_writer_false_for_object_without_book() -> None:
    class _NotAWriter:
        pass

    assert is_write_only_writer(_NotAWriter()) is False
