"""
Regression harness for workbook post-processing and excel_engine guardrails.

Standalone: discovers an existing audit ``.xlsx`` under common locations, or builds
an in-memory mock workbook when none exist—never raises ``FileNotFoundError``
for a missing input file.

Usage::

  PYTHONPATH=src python tests/test_excel_engine.py
  PYTHONPATH=src python tests/test_excel_engine.py path/to/workbook.xlsx

  pytest tests/test_excel_engine.py -v
"""

from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "src"
if SRC.is_dir() and str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from hype_frog.reporter.excel_engine import (  # noqa: E402
    apply_header_tooltips,
    apply_workbook_export_guardrails,
    friendly_toc_description,
    refresh_toc_descriptions_dynamic,
)
from hype_frog.reporter.sheets.tables_impl import (  # noqa: E402
    adjust_sheet_format,
    apply_tab_hyperlinks,
)
from hype_frog.reporter.sheets.toc import PREFERRED_WORKBOOK_TAB_ORDER  # noqa: E402

_ADJUST_SHEET_ORDER: tuple[str, ...] = tuple(
    n for n in PREFERRED_WORKBOOK_TAB_ORDER if n != "Table of Contents"
)


def _dedupe_paths(paths: list[Path]) -> list[Path]:
    seen: set[str] = set()
    out: list[Path] = []
    for p in paths:
        try:
            key = str(p.resolve())
        except OSError:
            key = str(p)
        if key not in seen:
            seen.add(key)
            out.append(p)
    return out


def _collect_xlsx_under(root: Path) -> list[Path]:
    """Return all ``.xlsx`` files under ``root`` (recursive)."""
    if not root.is_dir():
        return []
    return sorted(root.rglob("*.xlsx"))


def _pick_latest_audit_xlsx(search_roots: list[Path]) -> Path | None:
    """
    Pick the newest ``.xlsx`` under the given roots (each tree searched
    recursively for ``*.xlsx``).

    ``main()`` passes ``reports/latest``, ``reports/archive``, ``output``,
    repo root, and ``archive_legacy``. Returns ``None`` when nothing matches.
    """
    candidates: list[Path] = []
    for root in _dedupe_paths(search_roots):
        if not root.is_dir():
            continue
        candidates.extend(root.glob("SEO_AEO_Audit*.xlsx"))
        candidates.extend(root.glob("*.xlsx"))
        candidates.extend(_collect_xlsx_under(root))

    candidates = _dedupe_paths([p for p in candidates if p.is_file()])
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def _fill_hex_upper(cell) -> str:
    """Best-effort solid fill RGB hex (``FF0000``) from a cell."""
    fill = cell.fill
    if fill is None or fill.fill_type is None:
        return ""
    sc = fill.start_color
    rgb = getattr(sc, "rgb", None)
    if rgb is None:
        return ""
    s = str(rgb).upper().replace("#", "")
    if len(s) == 8 and s.startswith("FF"):
        return s[2:]
    return s[-6:] if len(s) >= 6 else s


def build_mock_audit_workbook() -> Workbook:
    """
    Minimal workbook: ``Table of Contents`` + ``Content Optimization`` with
    ``Action Required`` column and mixed row values.
    """
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Content Optimization"

    headers = ["URL", "Status", "Action Required"]
    ws_data.append(headers)
    ws_data.append(["https://a.example/page1", "200", ""])  # empty → unchanged
    ws_data.append(
        ["https://b.example/page2", "404", "needs rewrite here"]
    )  # → Needs Copy
    ws_data.append(["https://c.example/page3", "200", "   "])  # whitespace only → skip
    ws_data.append(
        ["https://d.example/page4", "200", "random blocker text"]
    )  # → Needs Copy

    toc = wb.create_sheet("Table of Contents", 0)
    toc["A1"] = "Table of Contents"
    toc["A2"] = "Section"
    toc["B2"] = "Open"
    toc["C2"] = "Description"
    toc["A3"] = "Content Optimization"
    toc["B3"] = '=HYPERLINK("#\'Content Optimization\'!A1","Open")'
    toc["C3"] = "Detailed URL diagnostic data."  # must be rewritten by guardrails

    return wb


def assert_action_required_guardrails(ws: Worksheet) -> None:
    """Assert ``Action Required`` uses strict literals, red fill only for ``Needs Copy``."""
    headers = {str(c.value).strip(): c.column for c in ws[1] if c.value is not None}
    col = headers.get("Action Required")
    assert col is not None, "Action Required column missing"

    c2 = ws.cell(row=2, column=col)
    assert c2.value == "Complete"
    assert _fill_hex_upper(c2) == "C6EFCE"

    c3 = ws.cell(row=3, column=col)
    assert c3.value == "Needs Copy", f"row 3 expected Needs Copy, got {c3.value!r}"
    assert _fill_hex_upper(c3) == "FF0000", (
        f"row 3 fill expected FF0000, got {_fill_hex_upper(c3)!r}"
    )

    c4 = ws.cell(row=4, column=col)
    assert c4.value == "Complete"

    c5 = ws.cell(row=5, column=col)
    assert c5.value == "Needs Copy"
    assert _fill_hex_upper(c5) == "FF0000"


def test_friendly_toc_description_known_tabs() -> None:
    assert "Executive overview" in friendly_toc_description("Dashboard")
    assert "Diagnostic command center" in friendly_toc_description(
        "Content Optimisation Hub"
    )
    assert "Deep-dive diagnostic" in friendly_toc_description("Technical")


def test_friendly_toc_description_unknown_tab() -> None:
    assert friendly_toc_description("CustomClientTab") == (
        "Diagnostic metrics for CustomClientTab."
    )


def test_apply_header_tooltips_sets_comments() -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["URL", "TTFB (ms)", "Mobile LCP (s)", "Internal PageRank", "Click Depth"])
    apply_header_tooltips(ws, header_row=1)
    assert ws["B1"].comment is not None
    assert "Time to First Byte" in ws["B1"].comment.text
    assert ws["C1"].comment is not None
    assert "Largest Contentful Paint" in ws["C1"].comment.text


def test_refresh_toc_descriptions_dynamic_uses_friendly_map() -> None:
    wb = Workbook()
    technical = wb.active
    technical.title = "Technical"
    toc = wb.create_sheet("Table of Contents", 0)
    toc["A1"] = "Table of Contents"
    toc["A2"] = "Section"
    toc["B2"] = "Open"
    toc["C2"] = "Description"
    toc["A3"] = "Technical"
    toc["C3"] = "Primary columns: old fallback"

    refresh_toc_descriptions_dynamic(wb)

    desc = str(toc["C3"].value or "")
    assert "Deep-dive diagnostic" in desc
    assert "Primary columns" not in desc


def assert_toc_no_generic_fallback(wb: Workbook) -> None:
    toc = wb["Table of Contents"]
    c3 = toc["C3"].value or ""
    assert "Detailed URL diagnostic data" not in str(c3).lower()
    assert "Primary columns" not in str(c3)
    assert len(str(c3)) > 0


def test_mock_workbook_guardrails() -> None:
    wb = build_mock_audit_workbook()
    apply_workbook_export_guardrails(wb)
    assert_action_required_guardrails(wb["Content Optimization"])
    assert_toc_no_generic_fallback(wb)
    # C2 freeze on data sheet (TOC excluded inside helper)
    assert wb["Content Optimization"].freeze_panes == "C2"


def test_pick_latest_returns_none_when_no_xlsx(tmp_path: Path) -> None:
    empty = tmp_path / "empty_tree"
    empty.mkdir()
    picked = _pick_latest_audit_xlsx([empty])
    assert picked is None


def test_pick_latest_prefers_newest(tmp_path: Path) -> None:
    a = tmp_path / "old.xlsx"
    b = tmp_path / "new.xlsx"
    a.write_bytes(
        b"PK\x03\x04" + b"\x00" * 18
    )  # minimal zip-like garbage for mtime only
    b.write_bytes(b"PK\x03\x04" + b"\x00" * 18)
    import os
    import time

    os.utime(a, (time.time() - 100, time.time() - 100))
    os.utime(b, (time.time(), time.time()))
    picked = _pick_latest_audit_xlsx([tmp_path])
    assert picked == b.resolve()


def main() -> None:
    repo_root = ROOT
    dest = repo_root / "test_dashboard_fix.xlsx"

    if len(sys.argv) > 1:
        source = Path(sys.argv[1]).expanduser().resolve()
        if not source.is_file():
            raise FileNotFoundError(source)
        print(f"Source workbook (CLI): {source}")
        frames = pd.read_excel(source, sheet_name=None, engine="openpyxl")
        with pd.ExcelWriter(dest, engine="openpyxl") as writer:
            for sheet_name, df in frames.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            apply_tab_hyperlinks(writer)
            for sname in _ADJUST_SHEET_ORDER:
                if sname in writer.sheets:
                    print(f"  adjust_sheet_format -> {sname}")
                    adjust_sheet_format(writer, sname)
            apply_workbook_export_guardrails(writer.book)
        print(f"Wrote: {dest}")
        return

    search_roots = [
        repo_root / "reports" / "latest",
        repo_root / "reports" / "archive",
        repo_root / "output",
        repo_root,
        repo_root / "archive_legacy",
    ]
    source = _pick_latest_audit_xlsx(search_roots)

    if source is None:
        print("No .xlsx found under search paths; using in-memory mock workbook.")
        wb = build_mock_audit_workbook()
        apply_workbook_export_guardrails(wb)
        assert_action_required_guardrails(wb["Content Optimization"])
        assert_toc_no_generic_fallback(wb)
        wb.save(dest)
        print(f"Wrote mock-based workbook: {dest}")
        return

    print(f"Source workbook: {source}")
    frames = pd.read_excel(source, sheet_name=None, engine="openpyxl")
    main_df = frames.get("Main")
    if main_df is None:
        main_df = frames.get("Main_Data")
    inv_df = frames.get("IssueInventory")
    print(
        "Loaded sheets:",
        len(frames),
        "| Main:",
        "Main" in frames or "Main_Data" in frames,
        "| IssueInventory:",
        "IssueInventory" in frames,
    )
    if main_df is not None:
        print(f"  Main/Main_Data rows: {len(main_df)}")
    if inv_df is not None:
        print(f"  IssueInventory rows: {len(inv_df)}")

    with pd.ExcelWriter(dest, engine="openpyxl") as writer:
        for sheet_name, df in frames.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        apply_tab_hyperlinks(writer)
        for sname in _ADJUST_SHEET_ORDER:
            if sname in writer.sheets:
                print(f"  adjust_sheet_format -> {sname}")
                adjust_sheet_format(writer, sname)
        apply_workbook_export_guardrails(writer.book)

    print(f"Wrote: {dest}")


if __name__ == "__main__":
    main()
