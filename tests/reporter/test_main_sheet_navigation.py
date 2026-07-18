"""Main sheet navigation: row-1 return strip (no trailing BACK TO DASHBOARD column)."""

from __future__ import annotations

from openpyxl import Workbook

from hype_frog.reporter.sheets.config import RETURN_TO_BRIEFING_LABEL
from hype_frog.reporter.sheets.navigation import add_return_to_briefing_strip
from hype_frog.reporter.sheets.sheet_rows import sheet_data_header_row
from hype_frog.reporter.sheets.style_helpers import header_index
from hype_frog.reporter.sheets.tables import normalize_table_headers


class _WriterStub:
    def __init__(self, book: Workbook) -> None:
        self.book = book


def _header_labels(worksheet, header_row: int) -> list[str]:
    return [
        str(worksheet.cell(row=header_row, column=c).value or "")
        for c in range(1, worksheet.max_column + 1)
    ]


def _build_main_workbook() -> tuple[_WriterStub, Workbook]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main"
    ws.cell(row=1, column=1, value="URL")
    ws.cell(row=1, column=2, value="SEO Health Score")
    ws.cell(row=2, column=1, value="https://example.com/page-a")
    ws.cell(row=2, column=2, value=72.0)
    ws.cell(row=3, column=1, value="https://example.com/page-b")
    ws.cell(row=3, column=2, value=65.0)
    wb.create_sheet("Technical Diagnostics")
    wb.create_sheet("Executive Briefing")
    return _WriterStub(wb), wb


def test_main_return_strip_survives_double_format_pass() -> None:
    """Simulate export_flow calling adjust_sheet_format twice on Main."""
    from hype_frog.reporter.sheets.navigation import apply_cross_sheet_links

    writer, wb = _build_main_workbook()
    ws = wb["Main"]

    for _ in range(2):
        add_return_to_briefing_strip(ws, "Main")
        header_row = sheet_data_header_row("Main")
        apply_cross_sheet_links(writer, ws, "Main", header_row=header_row)
        normalize_table_headers(ws, header_row=header_row)

    assert str(ws["A1"].value) == RETURN_TO_BRIEFING_LABEL
    assert ws["A1"].hyperlink is not None

    header_row = sheet_data_header_row("Main")
    headers = _header_labels(ws, header_row)
    assert "BACK TO DASHBOARD" not in headers
    assert headers.count("Technical View") == 1
    assert not any(h.endswith("_1") for h in headers)

    tech_col = header_index(ws, header_row)["Technical View"]
    data_start = header_row + 1
    assert str(ws.cell(row=data_start, column=tech_col).value or "").startswith(
        "=IFERROR(HYPERLINK"
    )


def test_technical_view_jump_link_gets_visible_styling() -> None:
    """Regression: ``Technical View`` (and other HYPERLINK-formula jump-link
    columns) rendered as plain black text with no underline, giving the reader
    no visual cue the cell was clickable, even though the formula worked."""
    from hype_frog.reporter.sheets.links import style_unstyled_formula_hyperlinks
    from hype_frog.reporter.sheets.navigation import apply_cross_sheet_links

    writer, wb = _build_main_workbook()
    ws = wb["Main"]
    add_return_to_briefing_strip(ws, "Main")
    header_row = sheet_data_header_row("Main")
    apply_cross_sheet_links(writer, ws, "Main", header_row=header_row)
    style_unstyled_formula_hyperlinks(ws, header_row=header_row)

    tech_col = header_index(ws, header_row)["Technical View"]
    data_start = header_row + 1
    cell = ws.cell(row=data_start, column=tech_col)
    assert str(cell.value).startswith("=IFERROR(HYPERLINK")
    assert cell.font.underline == "single"


def test_return_strip_skips_when_already_present() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value=RETURN_TO_BRIEFING_LABEL)
    ws.cell(row=2, column=1, value="Section")
    ws.cell(row=2, column=2, value="Severity")

    add_return_to_briefing_strip(ws, "Summary")

    assert ws.max_row == 2
    assert str(ws["A1"].value) == RETURN_TO_BRIEFING_LABEL
