"""Editable-workflow-input header marking (Priority URLs Status/Sprint, Hub
Status/Assigned Owner).

Regression: these columns are seeded defaults ("Open", "", "To Do", "Dev")
that the tool never overwrites on re-export, but previously carried no visual
signal distinguishing them from computed columns.
"""

from __future__ import annotations

from openpyxl import Workbook

from hype_frog.reporter.sheets.config import EDITABLE_INPUT_HEADER_FILL
from hype_frog.reporter.sheets.tables_impl import _mark_editable_input_headers


def _sheet_with_headers(title: str, headers: list[str]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    return wb


def test_priority_urls_status_header_gets_editable_fill() -> None:
    # "Sprint" was removed from Priority URLs entirely — only "Status" remains
    # a manual editable-workflow-input field on this sheet.
    wb = _sheet_with_headers(
        "Priority URLs",
        ["URL", "Business Risk Score", "Status", "Owner"],
    )
    ws = wb.active
    _mark_editable_input_headers(ws, header_row=1)

    assert ws.cell(row=1, column=3).fill.fgColor.rgb.endswith(
        EDITABLE_INPUT_HEADER_FILL
    )
    # Non-editable computed columns must be untouched.
    assert ws.cell(row=1, column=1).fill.patternType is None
    assert ws.cell(row=1, column=2).fill.patternType is None
    assert ws.cell(row=1, column=4).fill.patternType is None


def test_hub_status_and_assigned_owner_headers_get_editable_fill() -> None:
    wb = _sheet_with_headers(
        "Content Optimisation Hub",
        ["Action Required", "Status", "Assigned Owner", "URL"],
    )
    ws = wb.active
    _mark_editable_input_headers(ws, header_row=1)

    assert ws.cell(row=1, column=2).fill.fgColor.rgb.endswith(
        EDITABLE_INPUT_HEADER_FILL
    )
    assert ws.cell(row=1, column=3).fill.fgColor.rgb.endswith(
        EDITABLE_INPUT_HEADER_FILL
    )
    assert ws.cell(row=1, column=1).fill.patternType is None


def test_unrelated_sheet_is_a_no_op() -> None:
    wb = _sheet_with_headers("Main", ["URL", "Status Code"])
    ws = wb.active
    _mark_editable_input_headers(ws, header_row=1)
    assert ws.cell(row=1, column=1).fill.patternType is None
    assert ws.cell(row=1, column=2).fill.patternType is None
