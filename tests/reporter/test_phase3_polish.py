"""Phase 3 polish regression guards: formulas, formats, visibility, display headers."""

from __future__ import annotations

from openpyxl import Workbook

from hype_frog.pipeline.broken_links import link_inventory_broken_internal_total_formula
from hype_frog.reporter.sheets.layout import (
    DISPLAY_HEADER_ALIASES,
    apply_display_header_aliases,
    link_intelligence_column_letter,
    sheet_data_column_range,
)
from hype_frog.reporter.sheets.links import _fixplan_hub_status_formula
from hype_frog.reporter.sheets.number_formats import apply_south_african_formats
from hype_frog.reporter.sheets.workbook_layout import (
    DASHBOARD_ADVANCED_SHEET_LINKS,
    SHEETS_EXCLUDED_FROM_TOC,
)


def test_fixplan_hub_status_uses_status_not_seo_score_column() -> None:
    # Column letters are resolved dynamically from the Hub's actual export order
    # (content_hub_column_letter), not hardcoded, so they track any approved
    # reorder (2.5 UX overhaul: Action Required, Status, Assigned Owner, URL, ...).
    formula = _fixplan_hub_status_formula("H", 4)
    assert "!B3:B10000" in formula
    assert "!D3:D10000" in formula
    assert "!C:C" not in formula


def test_link_inventory_formula_uses_header_resolved_columns() -> None:
    """Broken-link SUMPRODUCT now targets Link Intelligence's own Detail rows
    (folded in from the former standalone "Link Inventory" sheet), not a
    separate tab — and excludes Summary rows via an explicit Record Type term."""
    formula = link_inventory_broken_internal_total_formula()
    rt = link_intelligence_column_letter("Record Type")
    lt = link_intelligence_column_letter("Link Type")
    sc = link_intelligence_column_letter("Status Code")
    assert f"'Link Intelligence'!${rt}$2:${rt}$100000" in formula
    assert f"'Link Intelligence'!${lt}$2:${lt}$100000" in formula
    assert f"'Link Intelligence'!${sc}$2:${sc}$100000" in formula
    assert '="Detail"' in formula
    assert lt == "F" and sc == "G"


def test_link_intelligence_generic_anchor_column_is_dynamic() -> None:
    assert link_intelligence_column_letter("Generic Anchor Text Count") == "S"
    assert link_intelligence_column_letter("Record Type") == "B"


def test_sheet_data_column_range_is_header_driven() -> None:
    rng = sheet_data_column_range("Technical Diagnostics", "SEO Health Score")
    assert "MATCH(\"SEO Health Score\"" in rng
    assert "Technical Diagnostics" in rng


def test_issue_inventory_excluded_from_toc_and_dashboard_links() -> None:
    assert "IssueInventory" in SHEETS_EXCLUDED_FROM_TOC
    assert all(name != "IssueInventory" for name, _ in DASHBOARD_ADVANCED_SHEET_LINKS)


def test_issue_inventory_skipped_in_toc_advanced_section() -> None:
    from hype_frog.reporter.sheets.toc import ADVANCED_WORKBOOK_TAB_ORDER

    assert "IssueInventory" in ADVANCED_WORKBOOK_TAB_ORDER
    assert "IssueInventory" in SHEETS_EXCLUDED_FROM_TOC


def test_display_header_alias_rewrites_optimisation_label() -> None:
    ws = Workbook().active
    ws["A1"] = "On-Page Optimization Score"
    ws["B1"] = "URL"
    apply_display_header_aliases(ws)
    assert ws["A1"].value == DISPLAY_HEADER_ALIASES["On-Page Optimization Score"]
    assert ws["B1"].value == "URL"


def test_number_formats_apply_to_psi_and_citation_count() -> None:
    ws = Workbook().active
    ws["A1"] = "Mobile PSI Score"
    ws["A2"] = 0.82
    ws["B1"] = "Citation Candidate Count"
    ws["B2"] = 3
    apply_south_african_formats(ws)
    assert ws["A2"].number_format == "[$-en-ZA]#,##0"
    assert ws["B2"].number_format == "[$-en-ZA]#,##0"
