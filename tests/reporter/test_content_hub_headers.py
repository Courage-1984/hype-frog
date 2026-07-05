"""Phase 4A/4C — Content Hub 2-row header and large-sheet CF zebra."""

from __future__ import annotations

import time

import openpyxl
from openpyxl import Workbook

from hype_frog.reporter.engine_guardrails import apply_header_tooltips
from hype_frog.reporter.engine_rows import CONTENT_HUB_EXPORT_COLUMNS
from hype_frog.reporter.sheets.conditional import apply_content_hub_conditional_rules
from hype_frog.reporter.sheets.config import (
    CONTENT_HUB_DATA_START_ROW,
    CONTENT_OPTIMISATION_HUB_SHEET,
    LARGE_SHEET_ROW_THRESHOLD,
    RETURN_TO_BRIEFING_LABEL,
    ZEBRA_FAINT,
)
from hype_frog.reporter.sheets.large_sheet_presentation import (
    apply_large_sheet_presentation,
    should_apply_large_sheet_presentation,
)


def test_content_hub_two_row_header_without_scope_row() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = CONTENT_OPTIMISATION_HUB_SHEET
    columns = list(CONTENT_HUB_EXPORT_COLUMNS)
    ws.append(columns)
    data_row: list[object | None] = [None] * len(columns)
    data_row[columns.index("Action Required")] = "Needs Copy"
    data_row[columns.index("URL")] = "https://example.com/about/"
    ws.append(data_row)

    class _Writer:
        book = wb
        sheets = {CONTENT_OPTIMISATION_HUB_SHEET: ws}

    apply_content_hub_conditional_rules(ws, _Writer())
    # Mirrors tables_impl.py's real Content Hub pipeline order: header tooltips are
    # applied after the conditional-rules pass and win for any header present in both
    # dicts (see reporter/CLAUDE.md "Tooltip ownership").
    apply_header_tooltips(ws, header_row=2)

    header_values = [
        ws.cell(row=2, column=col_idx).value for col_idx in range(1, len(columns) + 1)
    ]
    assert header_values == columns
    assert ws.cell(row=3, column=1).value != "Scope note:"
    url_col = columns.index("URL") + 1
    assert ws.cell(row=CONTENT_HUB_DATA_START_ROW, column=url_col).value == (
        "https://example.com/about/"
    )
    action_col = columns.index("Action Required") + 1
    action_header = ws.cell(row=2, column=action_col)
    assert action_header.comment is not None
    assert "Needs Optimisation" in (action_header.comment.text or "")
    assert str(ws["A1"].value).startswith(RETURN_TO_BRIEFING_LABEL)


def test_large_sheet_presentation_uses_cf_zebra_not_static_fills() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Link Inventory"
    ws.append(["Source URL", "Target URL", "Anchor Text", "Status Code", "Link Type"])
    for i in range(LARGE_SHEET_ROW_THRESHOLD + 10):
        ws.append(
            [
                f"https://example.com/{i}",
                f"https://example.com/target-{i}",
                f"anchor {i}",
                200,
                "Internal",
            ]
        )

    assert should_apply_large_sheet_presentation("Link Inventory", ws)

    t0 = time.perf_counter()
    apply_large_sheet_presentation(ws, "Link Inventory")
    elapsed = time.perf_counter() - t0

    assert ws.sheet_view.showGridLines is False
    assert len(ws.conditional_formatting._cf_rules) == 1
    rules = next(iter(ws.conditional_formatting._cf_rules.values()))
    assert any("MOD(ROW(),2)=0" in (r.formula[0] if r.formula else "") for r in rules)
    zebra_rule = next(
        r for r in rules if r.formula and "MOD(ROW(),2)=0" in r.formula[0]
    )
    assert zebra_rule.dxf is not None
    assert zebra_rule.dxf.fill.fgColor.rgb.endswith(ZEBRA_FAINT)
    # Guardrail: pass must stay sub-second (no O(rows×cols) cell loops).
    assert elapsed < 1.0

    filled_data_cells = sum(
        1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column)
        for cell in row
        if cell.fill and cell.fill.fill_type == "solid"
    )
    assert filled_data_cells == 0
