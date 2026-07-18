"""Sheet workflow/formatting guardrails: autofilter coverage, workflow status dropdown, zebra banding, return strip."""

from __future__ import annotations

from openpyxl import Workbook

from hype_frog.reporter.engine_formatting import ensure_auto_filter
from hype_frog.reporter.sheets.conditional import (
    apply_generic_sheet_coloring,
)
from hype_frog.reporter.sheets.config import (
    AUTO_FILTER_SHEETS,
    RETURN_TO_BRIEFING_LABEL,
    ZEBRA_BAND,
)
from hype_frog.reporter.sheets.navigation import (
    add_return_to_briefing_strip,
    apply_return_strip_descriptions,
)
from hype_frog.reporter.sheets.sheet_descriptions import (
    SHEET_END_USER_DESCRIPTIONS,
    sheet_end_user_description,
)
from hype_frog.reporter.sheets.validation import apply_workflow_status_dropdown
from hype_frog.reporter.sheets.workbook_layout import PREFERRED_WORKBOOK_TAB_ORDER


def test_auto_filter_sheets_include_actionable_tabs() -> None:
    for name in (
        "Summary",
        "Priority URLs",
        "FixPlan",
        "Quick Wins",
        "Technical Diagnostics",
        "Broken Link Impact",
    ):
        assert name in AUTO_FILTER_SHEETS


def test_ensure_auto_filter_applies_to_sparse_fixplan() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "FixPlan"
    ws.cell(row=1, column=1, value=RETURN_TO_BRIEFING_LABEL)
    ws.cell(row=2, column=1, value="Issue Type")
    ws.cell(row=2, column=2, value="Severity")
    ws.cell(row=3, column=1, value="Missing Title")

    ensure_auto_filter(ws)

    assert ws.auto_filter.ref == "A2:B3"


def test_workflow_status_dropdown_shows_input_message() -> None:
    wb = Workbook()
    ws = wb.active
    ws.cell(row=2, column=1, value="Status")
    ws.cell(row=3, column=1, value="To Do")

    apply_workflow_status_dropdown(ws, status_col=1, header_row=2)

    assert ws.data_validations.dataValidation
    dv = ws.data_validations.dataValidation[0]
    assert dv.showInputMessage is True
    assert dv.promptTitle == "Workflow status"


def test_cf_zebra_banding_on_small_data_sheet() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value=RETURN_TO_BRIEFING_LABEL)
    ws.cell(row=2, column=1, value="Issue Type")
    ws.cell(row=2, column=2, value="Count")
    ws.cell(row=3, column=1, value="Missing Title")
    ws.cell(row=3, column=2, value=5)
    ws.cell(row=4, column=1, value="Thin Content")
    ws.cell(row=4, column=2, value=2)

    apply_generic_sheet_coloring(ws, "Summary", header_row=2)

    rules = next(iter(ws.conditional_formatting._cf_rules.values()))
    zebra_rule = next(
        r for r in rules if r.formula and "MOD(ROW(),2)=0" in r.formula[0]
    )
    assert zebra_rule.dxf is not None
    assert zebra_rule.dxf.fill.fgColor.rgb.endswith(ZEBRA_BAND)


def test_sheet_end_user_description_lookup() -> None:
    assert sheet_end_user_description("Main").startswith("Complete URL inventory")
    # Unknown / self-describing sheets return empty (no banner written).
    assert sheet_end_user_description("Table of Contents") == ""
    assert sheet_end_user_description("Not A Real Sheet") == ""


def test_return_strip_description_keys_are_canonical_sheet_names() -> None:
    """Guards the 3-way sheet-name lock: every banner key is a real tab."""
    canonical = set(PREFERRED_WORKBOOK_TAB_ORDER)
    unknown = set(SHEET_END_USER_DESCRIPTIONS) - canonical
    assert not unknown, f"description keys not in tab order: {sorted(unknown)}"


def test_return_strip_shows_description_banner() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "FixPlan"
    ws.cell(row=1, column=1, value="Issue Type")
    ws.cell(row=1, column=2, value="Severity")
    ws.cell(row=1, column=3, value="Count")
    ws.cell(row=2, column=1, value="Missing Title")

    add_return_to_briefing_strip(ws, "FixPlan")

    assert ws["A1"].value == RETURN_TO_BRIEFING_LABEL
    assert ws["A1"].hyperlink is not None
    assert ws["C1"].alignment.horizontal == "left"
    assert ws["C1"].alignment.wrap_text is True

    apply_return_strip_descriptions(wb)
    assert ws["C1"].value == sheet_end_user_description("FixPlan")
    assert not str(ws["C1"].value).startswith("=")
