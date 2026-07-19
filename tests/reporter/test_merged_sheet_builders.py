"""Merged workbook tabs must populate from paired Main + Extra crawl rows."""

from __future__ import annotations

from hype_frog.reporter.sheets.merged_builders import (
    TECHNICAL_DIAGNOSTICS_COLUMNS,
    TECHNICAL_DIAGNOSTICS_LIGHTHOUSE_COLUMNS,
    build_content_ai_readiness_rows,
    build_issue_register_rows,
    build_link_intelligence_rows,
    build_link_inventory_rows,
    build_technical_diagnostics_rows,
)


def _sample_main_extra() -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    url = "https://example.com/about"
    main_rows = [
        {
            "URL": url,
            "Title": "About Example Co",
            "Meta Description": "Learn about our team and mission.",
            "Word Count (Body)": 420,
            "H1 Content": "About Us",
            "Has Valid JSON-LD": True,
            "Mobile PSI Score": 88,
            "Desktop PSI Score": 92,
        }
    ]
    extra_rows = [
        {
            "URL": url,
            "Status Code": 200,
            "Severity Badge": "Warning",
            "SEO Health Score": 72.0,
            "Critical Issues Count": 0,
            "Warning Issues Count": 2,
            "Matched Issues": "Missing FAQ/QA Schema | Low AEO Readiness Score",
            "Internal Links Count": 12,
            "Broken Internal Links Count": 1,
            "Internal Inlinks": 3,
            "Generic Anchor Text Count": 2,
            "AEO Readiness Score": 55,
            "Question Heading Count": 1,
            "Word Count": 0,
            "PSI Data Status": "Lab only",
            "Mobile PSI Score": None,
            "Link Details": [
                {
                    "Source URL": url,
                    "Target URL": "https://example.com/missing",
                    "Anchor Text": "old team page",
                    "Status Code": 404,
                }
            ],
        }
    ]
    return main_rows, extra_rows


def test_technical_diagnostics_uses_main_fallbacks() -> None:
    main_rows, extra_rows = _sample_main_extra()
    rows = build_technical_diagnostics_rows(extra_rows, main_rows=main_rows)
    assert len(rows) == 1
    row = rows[0]
    assert row["URL"] == "https://example.com/about"
    assert row["Desktop PSI Score"] == 92
    assert row["Mobile PSI Score"] == 88
    assert row["SEO Health Score"] == 72
    assert "Performance" in str(row["Diagnostic Category"])


def test_technical_diagnostics_includes_lighthouse_columns_from_main() -> None:
    url = "https://example.com/slow"
    main_rows = [
        {
            "URL": url,
            "CrUX Level": "Origin",
            "Lab LCP (Mobile) (s)": 7.351,
            "Lab TBT (Mobile) (ms)": 3359,
            "Lab FCP (Mobile) (s)": 2.1,
            "Lab CLS (Mobile)": 0.12,
            "Lab TTFB (Mobile) (ms)": 450,
            "Lighthouse Accessibility (Mobile)": 85,
            "Lighthouse Best Practices (Mobile)": 78,
            "Lighthouse SEO Score (Mobile)": 92,
            "Lab LCP (Desktop) (s)": 3.2,
            "Lab TBT (Desktop) (ms)": 180,
            "Lighthouse Performance (Desktop)": 55,
            "Page Size (KB)": 6618.3,
            "DOM Size (nodes)": 842,
            "JS Execution (ms)": 1250,
            "Network Request Count": 42,
            "Origin CrUX LCP (s)": 11.852,
            "Origin CrUX INP (ms)": 210,
        }
    ]
    extra_rows = [
        {
            "URL": url,
            "Status Code": 200,
            "Severity Badge": "Critical",
            "SEO Health Score": 40.0,
            "PSI Data Status": "PSI + CrUX Field (Origin)",
            "Mobile PSI Score": 28,
        }
    ]
    rows = build_technical_diagnostics_rows(extra_rows, main_rows=main_rows)
    assert len(rows) == 1
    row = rows[0]
    for key in TECHNICAL_DIAGNOSTICS_LIGHTHOUSE_COLUMNS:
        assert key in row
    assert row["CrUX Level"] == "Origin"
    assert row["Lab LCP (Mobile) (s)"] == 7.351
    assert row["Origin CrUX LCP (s)"] == 11.852
    assert row["Page Size (KB)"] == 6618.3
    assert "Performance" in str(row["Diagnostic Category"])


def test_technical_diagnostics_includes_reachable_from_homepage() -> None:
    url = "https://example.com/cart"
    main_rows = [
        {
            "URL": url,
            "Reachable from Homepage": False,
            "Click Depth": -1,
        }
    ]
    extra_rows = [
        {
            "URL": url,
            "Status Code": 200,
            "Severity Badge": "Pass",
            "SEO Health Score": 80.0,
            "Orphan Pages": False,
        }
    ]
    rows = build_technical_diagnostics_rows(extra_rows, main_rows=main_rows)
    assert "Reachable from Homepage" in TECHNICAL_DIAGNOSTICS_COLUMNS
    assert rows[0]["Reachable from Homepage"] is False


def test_content_ai_readiness_backfills_word_count_and_title() -> None:
    main_rows, extra_rows = _sample_main_extra()
    rows = build_content_ai_readiness_rows(extra_rows, main_rows=main_rows)
    assert len(rows) == 1
    row = rows[0]
    assert row["Word Count"] == 420
    assert row["Title Missing"] is False
    assert row["H1 Count"] == 1
    assert "AEO" in str(row["Content Category"])


def test_content_ai_readiness_merges_hub_metrics_and_anchor_audit() -> None:
    """Content Hub Metrics + Anchor Text Audit columns fold in via left-join dicts."""
    main_rows, extra_rows = _sample_main_extra()
    url = "https://example.com/about"
    hub_metrics_by_url = {
        url: {
            "URL": url,
            "Search Intent": "Informational",
            "Search Intent Source": "Heuristic",
            "Instant Priority": "CRITICAL",
            "Potential Traffic Lift": 42,
            "AEO Visibility Gain": 7.5,
            "JS Dependent": False,
            "Raw Words": 400,
            "Rendered Words": 420,
            "Field LCP (ms)": 2500.0,
            "Field CLS": 0.05,
            "Anchor Text Diversity": "High",
        }
    }
    anchor_audit_by_url = {
        url: {
            "Destination URL": url,
            "Inbound Link Count": 6,
            "Generic Anchor Count": 4,
            "Generic Anchor %": 66.7,
            "Top Anchor Texts": "click here | read more",
            "Generic Anchor Dominance": True,
            "Recommended Action": (
                "Rewrite generic anchors ('click here', 'read more') with descriptive text."
            ),
        }
    }
    rows = build_content_ai_readiness_rows(
        extra_rows,
        main_rows=main_rows,
        hub_metrics_by_url=hub_metrics_by_url,
        anchor_audit_by_url=anchor_audit_by_url,
    )
    assert len(rows) == 1
    row = rows[0]
    assert row["Search Intent"] == "Informational"
    assert row["Instant Priority"] == "CRITICAL"
    assert row["Potential Traffic Lift"] == 42
    assert row["Inbound Link Count"] == 6
    assert row["Generic Anchor Dominance"] is True
    assert row["Recommended Action"].startswith("Rewrite generic anchors")


def test_content_ai_readiness_defaults_when_no_hub_metrics_or_anchor_audit() -> None:
    """Left-join defaults kick in for URLs with no inbound anchors/hub metrics — no error."""
    main_rows, extra_rows = _sample_main_extra()
    rows = build_content_ai_readiness_rows(extra_rows, main_rows=main_rows)
    assert len(rows) == 1
    row = rows[0]
    assert row["Search Intent"] == ""
    assert row["Inbound Link Count"] == 0
    assert row["Generic Anchor Dominance"] is False


def test_link_intelligence_summary_rows() -> None:
    """Summary rows only — Detail rows are streamed separately (see
    test_link_intelligence_detail_rows_stream_from_cache) since the merge with the
    former standalone "Link Inventory" sheet."""
    main_rows, extra_rows = _sample_main_extra()
    graph_rows = [
        {
            "URL": "https://example.com/about",
            "Inlinks Count": 3,
            "Inlinks URLs": "https://example.com/",
            "Orphan Candidate": False,
            "Click Depth": 1,
            "Internal PageRank": 0.12,
        }
    ]
    rows = build_link_intelligence_rows(
        extra_rows=extra_rows,
        crawlgraph_rows=graph_rows,
        main_rows=main_rows,
    )
    assert len(rows) == 1
    row = rows[0]
    assert row["Record Type"] == "Summary"
    assert row["Broken Internal Links Count"] == 1
    assert row["Inlinks Count"] == 3
    # No Link Equity Map data supplied -> defaults, not an error.
    assert row["Inbound Link Count"] == 0
    assert row["Equity Tier"] == ""


def test_link_intelligence_summary_rows_merge_link_equity_map_columns() -> None:
    """Link Equity Map's columns fold into Summary rows via link_equity_by_url."""
    main_rows, extra_rows = _sample_main_extra()
    url = "https://example.com/about"
    link_equity_by_url = {
        url: {
            "URL": url,
            "Inbound Link Count": 12,
            "Unique Source Pages": 5,
            "Anchor Texts (top 5)": "about us | our team",
            "PageRank Percentile": 82.5,
            "Equity Tier": "High",
            "Recommended Action": "Maintain equity; ensure key CTAs are not buried below the fold.",
        }
    }
    rows = build_link_intelligence_rows(
        extra_rows=extra_rows,
        crawlgraph_rows=[],
        main_rows=main_rows,
        link_equity_by_url=link_equity_by_url,
    )
    assert len(rows) == 1
    row = rows[0]
    assert row["Inbound Link Count"] == 12
    assert row["Unique Source Pages"] == 5
    assert row["Equity Tier"] == "High"
    assert row["Recommended Action"].startswith("Maintain equity")


def test_link_intelligence_detail_rows_stream_from_cache() -> None:
    """Detail rows (deduplicated, decorated) stream from the Link Inventory SQLite
    cache and append after Summary rows already written to the same sheet — folded
    in from the former standalone "Link Inventory" sheet."""
    from openpyxl import Workbook

    from hype_frog.checkpoint.link_inventory_cache import LinkInventoryCache
    from hype_frog.pipeline.link_inventory_stream import populate_link_inventory_cache
    from hype_frog.reporter.engine_io import (
        append_link_detail_rows_streamed,
        write_dict_rows_sheet,
    )
    from hype_frog.reporter.sheets.merged_builders import LINK_INTELLIGENCE_COLUMNS

    main_rows, extra_rows = _sample_main_extra()
    # Duplicate anchor instance to prove dedup on (URL, Target URL, Anchor Text).
    extra_rows[0]["Link Details"].append(dict(extra_rows[0]["Link Details"][0]))
    summary_rows = build_link_intelligence_rows(
        extra_rows=extra_rows,
        crawlgraph_rows=[],
        main_rows=main_rows,
    )

    class _Writer:
        def __init__(self, wb: Workbook) -> None:
            self.book = wb
            self.sheets: dict[str, object] = {}

    wb = Workbook()
    wb.remove(wb.active)
    writer = _Writer(wb)
    columns = list(LINK_INTELLIGENCE_COLUMNS)
    write_dict_rows_sheet(writer, "Link Intelligence", columns, summary_rows)

    cache = LinkInventoryCache(":memory:")
    try:
        populate_link_inventory_cache(cache, extra_rows)
        assert cache.row_count() == 1  # dedup collapsed the duplicate anchor
        append_link_detail_rows_streamed(
            writer,
            cache,
            sheet_name="Link Intelligence",
            columns=columns,
            status_by_url={},
        )
    finally:
        cache.close()

    ws = writer.book["Link Intelligence"]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    body_rows = [
        {header[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        for r in range(2, ws.max_row + 1)
    ]
    record_types = [r["Record Type"] for r in body_rows]
    assert record_types == ["Summary", "Detail"]
    detail_row = body_rows[1]
    assert detail_row["URL"] == "https://example.com/about"
    assert detail_row["Target URL"] == "https://example.com/missing"


def test_issue_register_includes_summary_and_inventory() -> None:
    summary_rows = [
        {
            "Section": "Issue Counts",
            "Severity": "Warning",
            "Issue": "Missing Meta Description",
            "Affected URL Count": 2,
            "Reference Tab": "Content & AI Readiness",
            "Affected URLs (sample)": "https://example.com/a",
        }
    ]
    issue_inventory_rows = [
        {
            "URL": "https://example.com/a",
            "Issue": "Missing Meta Description",
            "Severity": "Warning",
            "Reference Tab": "Content & AI Readiness",
            "Stable Issue ID": "https://example.com/a::missing_meta_description",
            "Owner": "Copy Writer",
            "Sprint": "",
            "Status": "Open",
        }
    ]
    rows = build_issue_register_rows(
        summary_rows=summary_rows,
        issue_inventory_rows=issue_inventory_rows,
    )
    assert any(r["Section"] == "Issue Counts" for r in rows)
    assert any(r["Section"] == "Issue Inventory" for r in rows)


def test_link_inventory_deduplicates_source_target_anchor() -> None:
    url = "https://example.com/page"
    target = "https://example.com/about"
    anchor = "About us"
    extra_rows = [
        {
            "URL": url,
            "Link Details": [
                {
                    "Target URL": target,
                    "Anchor Text": anchor,
                    "Status Code": 200,
                },
                {
                    "Target URL": target,
                    "Anchor Text": anchor,
                    "Status Code": 200,
                },
            ],
        }
    ]
    rows = build_link_inventory_rows(extra_rows)
    assert len(rows) == 1
    assert rows[0]["Source URL"] == url
    assert rows[0]["Target URL"] == target
    assert rows[0]["Anchor Text"] == anchor
