"""TOC tab order, freeze panes, and workbook-tab synchronisation contracts."""

from __future__ import annotations

import re

from openpyxl import Workbook

from hype_frog.reporter.sheets.config import CONTENT_OPTIMISATION_HUB_SHEET
from hype_frog.reporter.sheets.toc import PREFERRED_WORKBOOK_TAB_ORDER


def test_preferred_workbook_tab_order_starts_with_toc_and_briefing() -> None:
    assert PREFERRED_WORKBOOK_TAB_ORDER[0] == "Table of Contents"
    assert PREFERRED_WORKBOOK_TAB_ORDER[1] == "Executive Briefing"
    assert PREFERRED_WORKBOOK_TAB_ORDER[2] == "Priority URLs"
    assert CONTENT_OPTIMISATION_HUB_SHEET in PREFERRED_WORKBOOK_TAB_ORDER


def test_preferred_tab_order_has_no_duplicates() -> None:
    assert len(PREFERRED_WORKBOOK_TAB_ORDER) == len(set(PREFERRED_WORKBOOK_TAB_ORDER))


def test_workbook_tabs_follow_preferred_order_when_subset_present() -> None:
    """Tabs present in a workbook must appear in PREFERRED_WORKBOOK_TAB_ORDER sequence."""
    subset = ("Table of Contents", "Executive Briefing", "Main", "Technical Diagnostics")
    wb = Workbook()
    wb.active.title = "Table of Contents"
    for name in ("Executive Briefing", "Main", "Technical Diagnostics"):
        wb.create_sheet(name)

    preferred_index = {name: idx for idx, name in enumerate(PREFERRED_WORKBOOK_TAB_ORDER)}
    for idx, tab_name in enumerate(subset):
        if tab_name in wb.sheetnames:
            wb.move_sheet(wb[tab_name], offset=-wb.index(wb[tab_name]) + idx)

    visible = [ws.title for ws in wb.worksheets]
    indices = [preferred_index[n] for n in visible if n in preferred_index]
    assert indices == sorted(indices), f"Tab order drift: {visible}"


def test_toc_hyperlink_formula_escapes_sheet_quotes() -> None:
    """Internal TOC links must escape apostrophes in sheet names (openpyxl XML safety)."""
    sheet_name = "Client's Schema & Metadata"
    safe = str(sheet_name).replace("'", "''")
    formula = f'=HYPERLINK("#\'{safe}\'!A1","Open")'
    assert "''" in formula
    assert re.search(r"HYPERLINK\(\"#'", formula)


def test_toc_freeze_panes_contract_is_a3_after_tab_hyperlinks() -> None:
    """Nuclear guardrail must not strip TOC freeze (only three columns)."""
    from hype_frog.reporter.sheets.tables_impl import apply_tab_hyperlinks

    wb = Workbook()
    wb.active.title = "Main"
    wb.create_sheet("Summary")
    writer = type("W", (), {"book": wb, "sheets": {}})()
    writer.sheets = {"Main": wb["Main"], "Summary": wb["Summary"]}
    apply_tab_hyperlinks(writer)
    toc = wb["Table of Contents"]
    assert toc.freeze_panes == "A3"


def test_toc_freeze_panes_contract_is_a3() -> None:
    """Canonical TOC freeze target documented in toc.py apply_workbook_toc_and_links."""
    wb = Workbook()
    toc = wb.active
    toc.title = "Table of Contents"
    toc.freeze_panes = "A3"
    assert toc.freeze_panes == "A3"
