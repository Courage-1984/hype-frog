"""FixPlan → Content Hub Status join must resolve Status, not SEO Score (F1)."""

from __future__ import annotations

from openpyxl import Workbook

from hype_frog.reporter.engine_rows import content_hub_column_letter
from hype_frog.reporter.sheets.config import (
    CONTENT_HUB_DATA_START_ROW,
    CONTENT_OPTIMISATION_HUB_SHEET,
)
from hype_frog.reporter.sheets.links import (
    _fixplan_hub_status_formula,
    apply_cross_sheet_links,
)
from hype_frog.reporter.sheets.style_helpers import header_index


class _WriterStub:
    def __init__(self, book: Workbook) -> None:
        self.book = book


def test_fixplan_hub_status_formula_indexes_status_matches_url() -> None:
    status_l = content_hub_column_letter("Status")
    url_l = content_hub_column_letter("URL")
    formula = _fixplan_hub_status_formula("J", 5)

    assert status_l == "F"
    assert url_l == "I"
    assert f"'{CONTENT_OPTIMISATION_HUB_SHEET}'!{status_l}{CONTENT_HUB_DATA_START_ROW}:{status_l}10000" in formula
    assert f"'{CONTENT_OPTIMISATION_HUB_SHEET}'!{url_l}{CONTENT_HUB_DATA_START_ROW}:{url_l}10000" in formula
    assert "J5" in formula
    assert "!C:C" not in formula
    assert "Not in Hub" in formula


def test_apply_cross_sheet_links_writes_hub_status_on_fixplan() -> None:
    wb = Workbook()
    fix = wb.active
    fix.title = "FixPlan"
    fix.cell(row=1, column=1, value="URL")
    fix.cell(row=2, column=1, value="https://example.com/a")
    hub = wb.create_sheet(CONTENT_OPTIMISATION_HUB_SHEET)
    hub.cell(row=2, column=6, value="Status")
    hub.cell(row=2, column=9, value="URL")
    hub.cell(row=CONTENT_HUB_DATA_START_ROW, column=6, value="Done")
    hub.cell(row=CONTENT_HUB_DATA_START_ROW, column=9, value="https://example.com/a")
    writer = _WriterStub(wb)

    apply_cross_sheet_links(
        writer,
        fix,
        "FixPlan",
        debug_excel_isolation_mode=False,
        header_index_fn=header_index,
    )

    headers = header_index(fix)
    hub_status_col = headers["Hub Status (Content Hub)"]
    formula = str(fix.cell(row=2, column=hub_status_col).value or "")
    assert formula.startswith("=IFERROR(INDEX")
    assert content_hub_column_letter("Status") in formula
    assert content_hub_column_letter("URL") in formula
