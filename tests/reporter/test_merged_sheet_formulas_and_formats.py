"""Merged-sheet formula/format regression guards: header-resolved columns, display aliases, number formats."""

from __future__ import annotations

from openpyxl import Workbook

from hype_frog.pipeline.broken_links import link_inventory_broken_internal_total_formula
from hype_frog.reporter.sheets.layout import (
    DISPLAY_HEADER_ALIASES,
    apply_display_header_aliases,
    link_intelligence_column_letter,
    link_inventory_column_letter,
    sheet_data_column_range,
)
from hype_frog.reporter.sheets.links import _fixplan_hub_status_formula
from hype_frog.reporter.sheets.number_formats import apply_south_african_formats
from hype_frog.reporter.sheets.workbook_layout import (
    DASHBOARD_ADVANCED_SHEET_LINKS,
    SHEETS_EXCLUDED_FROM_TOC,
)


def test_fixplan_hub_status_uses_status_not_seo_score_column() -> None:
    from hype_frog.reporter.engine_rows import content_hub_column_letter

    formula = _fixplan_hub_status_formula("H", 4)
    status_l = content_hub_column_letter("Status")
    url_l = content_hub_column_letter("URL")
    seo_score_l = content_hub_column_letter("SEO Score")
    assert f"!{status_l}3:{status_l}10000" in formula
    assert f"!{url_l}3:{url_l}10000" in formula
    assert f"!{seo_score_l}:{seo_score_l}" not in formula


def test_link_inventory_formula_uses_header_resolved_columns() -> None:
    formula = link_inventory_broken_internal_total_formula()
    lt = link_inventory_column_letter("Link Type")
    sc = link_inventory_column_letter("Status Code")
    assert f"'Link Inventory'!${lt}$2:${lt}$100000" in formula
    assert f"'Link Inventory'!${sc}$2:${sc}$100000" in formula
    assert lt == "E" and sc == "F"


def test_link_intelligence_generic_anchor_column_is_dynamic() -> None:
    assert link_intelligence_column_letter("Generic Anchor Text Count") == "O"
    assert link_intelligence_column_letter("Record Type") == "B"


def test_sheet_data_column_range_is_header_driven() -> None:
    rng = sheet_data_column_range("Technical Diagnostics", "SEO Health Score")
    assert "MATCH(\"SEO Health Score\"" in rng
    assert "Technical Diagnostics" in rng
    assert "$2:$2" in rng


def test_sheet_data_column_range_main_uses_row2_header_and_row3_data() -> None:
    rng = sheet_data_column_range("Main", "SEO Score")
    assert "MATCH(\"SEO Score\",'Main'!$2:$2,0)" in rng
    assert "OFFSET('Main'!$A$1,2," in rng


def test_issue_register_is_canonical_backlog_tab() -> None:
    from hype_frog.reporter.sheets.toc import ADVANCED_WORKBOOK_TAB_ORDER

    assert "Issue Register" in ADVANCED_WORKBOOK_TAB_ORDER
    assert "IssueInventory" not in ADVANCED_WORKBOOK_TAB_ORDER
    assert "IssueInventory" not in SHEETS_EXCLUDED_FROM_TOC
    assert any(name == "Issue Register" for name, _ in DASHBOARD_ADVANCED_SHEET_LINKS)


def test_display_header_alias_rewrites_optimisation_label() -> None:
    ws = Workbook().active
    ws["A1"] = "On-Page Optimization Score"
    ws["B1"] = "URL"
    apply_display_header_aliases(ws)
    assert ws["A1"].value == DISPLAY_HEADER_ALIASES["On-Page Optimization Score"]
    assert ws["B1"].value == "URL"


def test_number_formats_apply_to_psi_and_citation_count() -> None:
    ws = Workbook().active
    ws["A1"] = "Mobile PSI Score"
    ws["A2"] = 0.82
    ws["B1"] = "Citation Candidate Count"
    ws["B2"] = 3
    apply_south_african_formats(ws)
    assert ws["A2"].number_format == "[$-en-ZA]#,##0"
    assert ws["B2"].number_format == "[$-en-ZA]#,##0"


def test_number_formats_read_resolved_header_row_not_row_one() -> None:
    """Regression: real data sheets carry a row-1 "Return to Executive
    Briefing" banner, pushing headers to row 2. The function used to read a
    hardcoded row 1, silently matching nothing on every such sheet — every
    percent/integer/decimal/date column fell back to Excel's default
    "General" format across the whole workbook."""
    ws = Workbook().active
    ws["A1"] = "← Return to Executive Briefing"
    ws["A2"] = "Mobile PSI Score"
    ws["A3"] = 49
    apply_south_african_formats(ws, header_row=2)
    assert ws["A3"].number_format == "[$-en-ZA]#,##0"


def test_date_headers_parse_inconsistent_raw_strings_to_matching_format() -> None:
    """Regression: date-like fields arrive as raw strings in whatever format
    the source gave (ISO 8601 with offset from schema/meta tags, RFC 2822 from
    the HTTP Last-Modified header) and render inconsistently side by side.
    Setting ``number_format`` alone is a no-op on a string cell — Excel only
    honours it on numeric/date-typed cells — so the value must be parsed into
    a real ``datetime`` for the format to take visual effect."""
    ws = Workbook().active
    ws["A1"] = "Schema Published Date"
    ws["A2"] = "2021-08-24T07:46:41+02:00"
    ws["B1"] = "Last Modified Date"
    ws["B2"] = "Wed, 08 Jul 2026 11:49:40 GMT"
    apply_south_african_formats(ws)

    import datetime as dt

    assert isinstance(ws["A2"].value, dt.datetime)
    assert isinstance(ws["B2"].value, dt.datetime)
    assert ws["A2"].number_format == "[$-en-ZA]dd/mm/yyyy hh:mm:ss"
    assert ws["B2"].number_format == "[$-en-ZA]dd/mm/yyyy hh:mm:ss"


def test_date_header_leaves_unparseable_value_untouched() -> None:
    ws = Workbook().active
    ws["A1"] = "Schema Published Date"
    ws["A2"] = "not a date"
    apply_south_african_formats(ws)
    assert ws["A2"].value == "not a date"
    assert ws["A2"].number_format == "General"


def test_number_formats_self_corrects_hub_header_row() -> None:
    """The Content Optimisation Hub's header row physically moves from row 1
    to row 2 partway through ``adjust_sheet_format`` (banner insert happens
    later than the caller's ``header_row`` snapshot), so trusting the passed
    value blindly would silently miss headers depending on call order."""
    from hype_frog.reporter.sheets.config import CONTENT_OPTIMISATION_HUB_SHEET

    ws = Workbook().active
    ws.title = CONTENT_OPTIMISATION_HUB_SHEET
    ws["A1"] = "Action Required"
    ws["B1"] = "Citation Candidate Count"
    ws["B2"] = 7
    # Caller passes header_row=2 (the sheet's eventual contract) even though
    # headers are still physically on row 1 at this point in the pipeline.
    apply_south_african_formats(ws, header_row=2)
    assert ws["B2"].number_format == "[$-en-ZA]#,##0"
