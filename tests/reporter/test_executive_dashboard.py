"""Executive Dashboard charts and visible source rows."""

from __future__ import annotations

import re
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from hype_frog.core.models import ExtraRowPayload, MainRowPayload, SummaryMetricsPayload
from hype_frog.reporter.chart_compat import patch_xlsx_app_xml_for_excel_compatibility
from hype_frog.reporter.sheets.config import EXECUTIVE_BRIEFING_SHEET
from hype_frog.reporter.excel_engine import (
    adjust_sheet_format,
    apply_tab_hyperlinks,
    apply_workbook_export_guardrails,
)
from hype_frog.reporter.dashboard_logic import FixPlanRowPayload, compute_dashboard_metrics
from hype_frog.reporter.sheets.executive_dashboard import (
    CHART_LABEL_COL,
    CHART_SOURCE_FIRST_ROW,
    LAB_LCP_MOBILE_TARGET_S,
    LIGHTHOUSE_ACCESSIBILITY_TARGET,
    _avg_lighthouse_performance_mobile,
    _avg_numeric_column,
    _is_low_value_priority_url,
    _meaningful_priority_rows,
    _owner_metrics,
    _project_component_score,
    _severity_metrics,
    _top_issues_by_impact,
    write_executive_dashboard,
)


def _chart_title_text(chart: object) -> str:
    title = getattr(chart, "title", None)
    tx = getattr(title, "tx", None)
    rich = getattr(tx, "rich", None)
    if rich is None:
        return ""
    parts: list[str] = []
    for paragraph in rich.p:
        for run in paragraph.r:
            text = getattr(run, "t", None)
            if text:
                parts.append(str(text))
    return "".join(parts)


@pytest.fixture
def sample_export_context() -> dict[str, object]:
    url = "https://example.com/pricing"
    extra = {
        "URL": url,
        "Status Code": 200,
        "Severity Badge": "Warning",
        "SEO Health Score": 62.0,
        "Critical Issues Count": 0,
        "Warning Issues Count": 2,
        "Observation Issues Count": 1,
        "Mobile PSI Score": 78,
        "Desktop PSI Score": 82,
        "Lighthouse Performance (Mobile)": 71,
        "Lab LCP (Mobile) (s)": 3.2,
        "Lighthouse Accessibility (Mobile)": 88,
        "AEO Readiness Score": 55,
        "Missing H1 Flag": False,
        "H1 Count": 1,
        "Meta Description Missing": False,
        "Paragraphs 40-60 Words Count": 2,
        "Schema Types Count": 1,
        "Image Alt Coverage (%)": 85,
    }
    return {
        "summary_metrics": SummaryMetricsPayload(
            urls_crawled=1,
            seo_pass_rate_pct=0.0,
            health_score_pct=62.0,
            critical_url_count=0,
            warning_url_count=1,
            projected_health_score_pct=78.0,
            projected_pass_rate_pct=40.0,
        ),
        "typed_main_rows": [MainRowPayload.model_validate({"URL": url, "SEO Health Score": 62.0})],
        "typed_extra_rows": [ExtraRowPayload.model_validate(extra)],
        "priority_rows": [
            {
                "URL": url,
                "Business Risk Score": 42,
                "Severity Badge": "Warning",
            }
        ],
        "fixplan_rows": [
            {
                "Issue Type": "Missing Meta Description",
                "Severity": "Warning",
                "Affected Count": 1,
                "Owner": "Copy Writer",
            }
        ],
        "hub_metrics_rows": [{"URL": url, "Potential Traffic Lift": 120}],
    }


def test_avg_lighthouse_performance_mobile_uses_lighthouse_not_desktop_blend() -> None:
    rows = [
        {
            "Lighthouse Performance (Mobile)": 44,
            "Mobile PSI Score": 39,
            "Desktop PSI Score": 82,
        },
        {
            "Lighthouse Performance (Mobile)": 28,
            "Mobile PSI Score": 28,
            "Desktop PSI Score": 90,
        },
    ]
    assert _avg_lighthouse_performance_mobile(rows) == 36.0


def test_avg_lighthouse_performance_mobile_falls_back_to_mobile_psi_score() -> None:
    rows = [{"Mobile PSI Score": 55, "Desktop PSI Score": 99}]
    assert _avg_lighthouse_performance_mobile(rows) == 55.0


def test_avg_numeric_column_skips_non_positive_values() -> None:
    rows = [
        {"Lab LCP (Mobile) (s)": 0.0},
        {"Lab LCP (Mobile) (s)": 4.1},
        {"Lab LCP (Mobile) (s)": 5.9},
    ]
    assert _avg_numeric_column(rows, "Lab LCP (Mobile) (s)", require_positive=True) == 5.0


def test_severity_metrics_separates_unique_urls_from_issue_instances() -> None:
    extra_rows = [
        {
            "URL": "https://example.com/a",
            "Severity Badge": "Critical",
            "Owner": "Dev",
        },
        {
            "URL": "https://example.com/b",
            "Severity Badge": "Warning",
            "Owner": "Copy Writer",
        },
    ]
    fixplan_rows = [
        {"Severity": "Critical", "Affected Count": 3, "Owner": "Dev"},
        {"Severity": "Warning", "Affected Count": 2, "Owner": "Copy Writer"},
        {"Severity": "Warning", "Affected Count": 1, "Owner": "Copy Writer"},
    ]
    severity = _severity_metrics(extra_rows, fixplan_rows)
    by_label = {label: (urls, instances) for label, urls, instances in severity}
    assert by_label["Critical"] == (1, 3)
    assert by_label["Warning"] == (1, 3)

    owners = _owner_metrics(extra_rows, fixplan_rows)
    owner_map = {name: (urls, instances) for name, urls, instances in owners}
    assert owner_map["Dev"] == (1, 3)
    assert owner_map["Copy Writer"] == (1, 3)


def test_top_issues_by_impact_tie_break_matches_key_insights_largest_theme() -> None:
    """Regression (L1): when several issues tie on Affected Count, the
    Executive Briefing's "Largest theme" sentence (dashboard_logic's
    top_issue_rows) and the "Top issues by URL impact" chart table
    (_top_issues_by_impact) must agree on which one is "the" biggest —
    previously they used different tie-break rules and could name two
    different issues for the exact same tied count."""
    fixplan_rows = [
        {"Issue Type": "No ETag Header", "Severity": "Warning", "Affected Count": 255},
        {"Issue Type": "No Terms Link", "Severity": "Warning", "Affected Count": 255},
        {"Issue Type": "No Consent Manager Detected", "Severity": "Warning", "Affected Count": 255},
        {"Issue Type": "No 40-60 Word Answer Paragraphs", "Severity": "Observation", "Affected Count": 255},
        {"Issue Type": "AI Crawlers Not Explicitly Allowed", "Severity": "Warning", "Affected Count": 255},
    ]
    summary = SummaryMetricsPayload(
        urls_crawled=255,
        seo_pass_rate_pct=0.0,
        health_score_pct=0.0,
        critical_url_count=0,
        warning_url_count=255,
        projected_health_score_pct=0.0,
        projected_pass_rate_pct=0.0,
    )
    fixplan_payloads = [
        FixPlanRowPayload.model_validate({**row, "source_row": idx})
        for idx, row in enumerate(fixplan_rows, start=2)
    ]
    dashboard_metrics = compute_dashboard_metrics(
        summary_metrics=summary,
        technical_main_rows=[],
        technical_extra_rows=[],
        fixplan_rows=fixplan_payloads,
        aeo_rows=[],
    )
    largest_theme = dashboard_metrics.top_issue_rows[0].issue_name

    chart_top_issue = _top_issues_by_impact(fixplan_rows, limit=8)[0][0]

    assert chart_top_issue == largest_theme


def test_meaningful_priority_rows_skip_low_value_paths() -> None:
    rows = [
        {"URL": "https://example.com/cart", "Business Risk Score": 99},
        {"URL": "https://example.com/pricing", "Business Risk Score": 42},
        {"URL": "https://example.com/thank-you", "Business Risk Score": 80},
    ]
    assert _is_low_value_priority_url("https://example.com/checkout")
    filtered = _meaningful_priority_rows(rows, limit=2)
    assert [row["URL"] for row in filtered] == ["https://example.com/pricing"]


def test_project_component_score_matches_export_flow_uplift() -> None:
    summary = SummaryMetricsPayload(
        urls_crawled=10,
        seo_pass_rate_pct=40.0,
        health_score_pct=50.0,
        critical_url_count=2,
        warning_url_count=3,
        projected_health_score_pct=72.5,
        projected_pass_rate_pct=60.0,
    )
    projected = _project_component_score(50.0, summary)
    assert projected == 72.5


def test_executive_dashboard_writes_charts_with_visible_source_rows(
    tmp_path: Path,
    sample_export_context: dict[str, object],
) -> None:
    out = tmp_path / "exec_dash.xlsx"
    writer = __import__("pandas").ExcelWriter(out, engine="openpyxl")
    write_executive_dashboard(
        writer,
        summary_metrics=sample_export_context["summary_metrics"],  # type: ignore[arg-type]
        typed_main_rows=sample_export_context["typed_main_rows"],  # type: ignore[arg-type]
        typed_extra_rows=sample_export_context["typed_extra_rows"],  # type: ignore[arg-type]
        priority_rows=sample_export_context["priority_rows"],  # type: ignore[arg-type]
        fixplan_rows=sample_export_context["fixplan_rows"],  # type: ignore[arg-type]
        hub_metrics_rows=sample_export_context["hub_metrics_rows"],  # type: ignore[arg-type]
    )
    writer.close()
    patch_xlsx_app_xml_for_excel_compatibility(out)

    wb = load_workbook(out)
    exec_ws = wb[EXECUTIVE_BRIEFING_SHEET]
    assert len(exec_ws._charts) >= 4
    titles = {_chart_title_text(chart) for chart in exec_ws._charts}
    assert "Health components — current vs illustrative projected" in titles
    assert any("Issue severity" in title for title in titles)
    assert any("Content & AEO readiness" in title for title in titles)
    assert any("Top issues by URL impact" in title for title in titles)
    assert "High-intent pages by business risk" in titles
    assert "Key insights:" in str(exec_ws.cell(row=11, column=1).value or "")
    assert exec_ws.cell(row=CHART_SOURCE_FIRST_ROW + 2, column=2).value is not None
    health_data_start = CHART_SOURCE_FIRST_ROW + 3
    health_labels = [
        str(exec_ws.cell(row=row, column=CHART_LABEL_COL).value or "")
        for row in range(health_data_start, health_data_start + 6)
    ]
    assert health_labels == [
        "SEO Health",
        "Technical Health",
        "Performance (PSI)",
        "LCP (Lab Mobile avg)",
        "Accessibility (avg)",
        "AEO Readiness",
    ]
    assert exec_ws.cell(row=health_data_start + 2, column=2).value == 71
    assert exec_ws.cell(row=health_data_start + 3, column=2).value == 3.2
    assert exec_ws.cell(row=health_data_start + 3, column=3).value == LAB_LCP_MOBILE_TARGET_S
    assert exec_ws.cell(row=health_data_start + 4, column=2).value == 88
    assert exec_ws.cell(row=health_data_start + 4, column=3).value == LIGHTHOUSE_ACCESSIBILITY_TARGET
    wb.close()


def test_executive_dashboard_chart_xml_uses_valid_rgb_and_visible_data(
    tmp_path: Path,
    sample_export_context: dict[str, object],
) -> None:
    out = tmp_path / "exec_dash_rgb.xlsx"
    writer = __import__("pandas").ExcelWriter(out, engine="openpyxl")
    write_executive_dashboard(
        writer,
        summary_metrics=sample_export_context["summary_metrics"],  # type: ignore[arg-type]
        typed_main_rows=sample_export_context["typed_main_rows"],  # type: ignore[arg-type]
        typed_extra_rows=sample_export_context["typed_extra_rows"],  # type: ignore[arg-type]
        priority_rows=sample_export_context["priority_rows"],  # type: ignore[arg-type]
        fixplan_rows=sample_export_context["fixplan_rows"],  # type: ignore[arg-type]
        hub_metrics_rows=sample_export_context["hub_metrics_rows"],  # type: ignore[arg-type]
    )
    writer.close()
    patch_xlsx_app_xml_for_excel_compatibility(out)

    with zipfile.ZipFile(out) as zf:
        app_xml = zf.read("docProps/app.xml").decode()
        assert "Compatible / Openpyxl" not in app_xml
        for name in zf.namelist():
            if not name.startswith("xl/charts/chart"):
                continue
            xml = zf.read(name).decode()
            for rgb in re.findall(r'srgbClr val="([^"]+)"', xml):
                assert len(rgb) == 6, f"invalid srgbClr {rgb!r} in {name}"
            assert 'plotVisOnly val="0"' in xml or "plotVisOnly" not in xml


def test_executive_dashboard_charts_have_per_point_or_per_series_colour(
    tmp_path: Path,
    sample_export_context: dict[str, object],
) -> None:
    """Regression: single-series bar charts must not render as one flat
    colour, and multi-series charts must distinguish their series by fill.

    Before this fix, 6 of 8 charts had no ``dPt``/series ``solidFill`` at
    all, so every bar in a chart rendered identically regardless of what it
    represented.
    """
    out = tmp_path / "exec_dash_colours.xlsx"
    writer = __import__("pandas").ExcelWriter(out, engine="openpyxl")
    write_executive_dashboard(
        writer,
        summary_metrics=sample_export_context["summary_metrics"],  # type: ignore[arg-type]
        typed_main_rows=sample_export_context["typed_main_rows"],  # type: ignore[arg-type]
        typed_extra_rows=sample_export_context["typed_extra_rows"],  # type: ignore[arg-type]
        priority_rows=sample_export_context["priority_rows"],  # type: ignore[arg-type]
        fixplan_rows=sample_export_context["fixplan_rows"],  # type: ignore[arg-type]
        hub_metrics_rows=sample_export_context["hub_metrics_rows"],  # type: ignore[arg-type]
    )
    writer.close()
    patch_xlsx_app_xml_for_excel_compatibility(out)

    with zipfile.ZipFile(out) as zf:
        chart_xmls = {
            name: zf.read(name).decode()
            for name in zf.namelist()
            if name.startswith("xl/charts/chart")
        }

    def _titled(fragment: str) -> str:
        for xml in chart_xmls.values():
            if fragment in xml:
                return xml
        raise AssertionError(f"no chart XML contains {fragment!r}")

    # Two-series charts: each series must carry its own solidFill, and the
    # two fills must differ (else the comparison reads as one colour).
    health_xml = _titled("Health components")
    series_fills = re.findall(r"<ser>.*?</ser>", health_xml, flags=re.S)
    assert len(series_fills) == 2
    fills = [re.findall(r'srgbClr val="([^"]+)"', s) for s in series_fills]
    assert all(fills), "expected an explicit solidFill on every series"
    assert fills[0][0] != fills[1][0]

    # Single-series charts: at least one <dPt> per-point override, proving
    # per-bar colouring (not just the theme's default single accent colour).
    status_xml = _titled("Status code")
    assert "<dPt>" in status_xml
    assert "70AD47" in status_xml  # 200 OK bucket -> green


def test_adjust_sheet_format_tolerates_executive_dashboard_merges(
    tmp_path: Path,
    sample_export_context: dict[str, object],
) -> None:
    out = tmp_path / "exec_dash_format.xlsx"
    writer = __import__("pandas").ExcelWriter(out, engine="openpyxl")
    write_executive_dashboard(
        writer,
        summary_metrics=sample_export_context["summary_metrics"],  # type: ignore[arg-type]
        typed_main_rows=sample_export_context["typed_main_rows"],  # type: ignore[arg-type]
        typed_extra_rows=sample_export_context["typed_extra_rows"],  # type: ignore[arg-type]
        priority_rows=sample_export_context["priority_rows"],  # type: ignore[arg-type]
        fixplan_rows=sample_export_context["fixplan_rows"],  # type: ignore[arg-type]
        hub_metrics_rows=sample_export_context["hub_metrics_rows"],  # type: ignore[arg-type]
    )
    adjust_sheet_format(writer, EXECUTIVE_BRIEFING_SHEET)
    writer.close()
    wb = load_workbook(out)
    assert wb[EXECUTIVE_BRIEFING_SHEET].freeze_panes == "A12"
    wb.close()


def test_executive_dashboard_survives_export_finalization(
    tmp_path: Path,
    sample_export_context: dict[str, object],
) -> None:
    out = tmp_path / "exec_dash_finalize.xlsx"
    writer = __import__("pandas").ExcelWriter(out, engine="openpyxl")
    write_executive_dashboard(
        writer,
        summary_metrics=sample_export_context["summary_metrics"],  # type: ignore[arg-type]
        typed_main_rows=sample_export_context["typed_main_rows"],  # type: ignore[arg-type]
        typed_extra_rows=sample_export_context["typed_extra_rows"],  # type: ignore[arg-type]
        priority_rows=sample_export_context["priority_rows"],  # type: ignore[arg-type]
        fixplan_rows=sample_export_context["fixplan_rows"],  # type: ignore[arg-type]
        hub_metrics_rows=sample_export_context["hub_metrics_rows"],  # type: ignore[arg-type]
    )
    apply_tab_hyperlinks(writer)
    adjust_sheet_format(writer, EXECUTIVE_BRIEFING_SHEET)
    apply_workbook_export_guardrails(writer.book)
    writer.close()
    patch_xlsx_app_xml_for_excel_compatibility(out)
    wb = load_workbook(out)
    assert wb[EXECUTIVE_BRIEFING_SHEET].freeze_panes == "A12"
    assert len(wb[EXECUTIVE_BRIEFING_SHEET]._charts) >= 4
    wb.close()
