"""TOC/tab-navigation regression guards: blurbs, landing tab, jump formulas, tooltips, freeze precedence."""

from __future__ import annotations

import openpyxl
import pytest
from openpyxl import Workbook

from hype_frog.reporter import engine_guardrails
from hype_frog.reporter.engine_guardrails import friendly_toc_description
from hype_frog.reporter.sheets import validation
from hype_frog.reporter.sheets.links import _reference_tab_jump_formula
from hype_frog.reporter.sheets.validation import (
    add_all_header_tooltips,
    resolve_curated_help_body,
)
from hype_frog.reporter.sheets.workbook_layout import (
    WORKBOOK_LANDING_SHEET,
    apply_workbook_active_tab,
)

_INVENTORY_SHEETS = [
    "Link Equity Map",
    "Anchor Text Audit",
    "Snippet Opportunities",
    "Script Inventory",
    "Image Inventory",
    "Broken Link Impact",
]


@pytest.mark.parametrize("sheet", _INVENTORY_SHEETS)
def test_inventory_sheets_have_curated_toc_blurbs(sheet: str) -> None:
    desc = friendly_toc_description(sheet)
    assert not desc.startswith("Diagnostic metrics for"), (
        f"{sheet} fell back to the generic TOC blurb"
    )
    assert len(desc) > 20


def test_apply_workbook_active_tab_lands_on_executive_briefing() -> None:
    wb = Workbook()
    wb.active.title = "Table of Contents"  # index 0 stays the TOC
    wb.create_sheet("Executive Briefing")
    wb.create_sheet("Main")

    apply_workbook_active_tab(wb)

    assert wb.sheetnames[0] == "Table of Contents"
    assert wb.active.title == WORKBOOK_LANDING_SHEET == "Executive Briefing"
    assert wb["Executive Briefing"].views.sheetView[0].tabSelected is True
    assert wb["Table of Contents"].views.sheetView[0].tabSelected is False


def test_apply_workbook_active_tab_falls_back_to_toc() -> None:
    wb = Workbook()
    wb.active.title = "Table of Contents"
    wb.create_sheet("Main")

    apply_workbook_active_tab(wb)  # no Dashboard present

    assert wb.active.title == "Table of Contents"


def test_reference_tab_jump_formula_quotes_and_escapes() -> None:
    formula = _reference_tab_jump_formula("E", "A", 2)
    # Sheet name (a runtime cell value) is escaped for embedding inside single quotes.
    assert "SUBSTITUTE(E2,\"'\",\"''\")" in formula
    # Both the HYPERLINK target and the INDIRECT lookup wrap the name in single quotes.
    assert formula.count("'\"&") >= 1
    assert "INDIRECT" in formula
    assert formula.startswith("=IFERROR(HYPERLINK(")


def test_curated_help_registered_for_phase2_sheets() -> None:
    assert resolve_curated_help_body("FixPlan", "Priority Score") is not None
    assert resolve_curated_help_body("Quick Wins", "Effort (hrs)") is not None
    assert resolve_curated_help_body("Broken Link Impact", "Inbound Link Count") is not None
    assert resolve_curated_help_body("Link Inventory", "Generic Anchor") is not None
    assert resolve_curated_help_body("SitemapQA", "Found via Crawl") is not None
    # Unknown pairs still defer to the generic tooltip path.
    assert resolve_curated_help_body("FixPlan", "Nonexistent Column") is None


def test_tooltip_flag_suppresses_header_comments(monkeypatch: pytest.MonkeyPatch) -> None:
    ws = openpyxl.Workbook().active
    ws["A1"] = "SEO Health Score"
    ws["B1"] = "URL"

    monkeypatch.setattr(validation, "_DISABLE_TOOLTIP_COMMENTS", True)
    add_all_header_tooltips(ws)
    assert ws["A1"].comment is None

    monkeypatch.setattr(validation, "_DISABLE_TOOLTIP_COMMENTS", False)
    add_all_header_tooltips(ws)
    assert ws["A1"].comment is not None


def test_add_all_header_tooltips_respects_header_row_with_banner() -> None:
    """Regression: most data sheets carry a row-1 "Return to Executive
    Briefing" banner, so real headers live on row 2 — the previous
    hardcoded row-1 lookup silently attached zero tooltips on every such
    sheet, including any curated entries for "Main"/etc."""
    ws = openpyxl.Workbook().active
    ws.title = "Main"
    ws["A1"] = "← Return to Executive Briefing"
    ws["A2"] = "Inbound Internal Link Count"
    ws["B2"] = "URL"

    add_all_header_tooltips(ws, header_row=2)

    assert ws["A2"].comment is not None
    assert "Inbound Internal Link Count" in ws["A2"].comment.text
    # The banner row itself must not receive a spurious header comment.
    assert ws["A1"].comment is None


def test_friendly_toc_description_unknown_still_names_sheet() -> None:
    # Guards the T1 fallback path used for quoted/odd sheet names.
    desc = friendly_toc_description("Totally New Sheet")
    assert "Totally New Sheet" in desc


def test_freeze_precedence_exempts_bespoke_sheets() -> None:
    from hype_frog.reporter.engine_guardrails import (
        apply_freeze_c2_data_sheets,
    )
    from hype_frog.reporter.sheets.config import (
        CONTENT_OPTIMISATION_HUB_SHEET,
        EXECUTIVE_BRIEFING_SHEET,
    )

    wb = Workbook()
    toc = wb.active
    toc.title = "Table of Contents"
    toc.freeze_panes = "A3"
    briefing = wb.create_sheet(EXECUTIVE_BRIEFING_SHEET)
    briefing.freeze_panes = "A22"
    hub = wb.create_sheet(CONTENT_OPTIMISATION_HUB_SHEET)
    hub.freeze_panes = "A4"
    data = wb.create_sheet("Technical Diagnostics")
    data["A1"] = "URL"
    data["A2"] = "https://example.org/"

    apply_freeze_c2_data_sheets(wb)

    # Bespoke freezes survive; ordinary data sheets normalise to C3 (return strip + header).
    assert toc.freeze_panes == "A3"
    assert briefing.freeze_panes == "A22"
    assert hub.freeze_panes == "A4"
    assert data.freeze_panes == "C3"


def test_banned_fallback_constant_present() -> None:
    # The refresh no longer references it directly, but it documents the legacy
    # fallback string the TOC must never show.
    assert engine_guardrails._BANNED_TOC_FALLBACK
