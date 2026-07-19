"""Workbook tab order, colours, and default visibility."""

from __future__ import annotations

from openpyxl import Workbook

from hype_frog.reporter.sheets.config import (
    AIOSEO_RECOMMENDATIONS_SHEET,
    AUDIT_RUN_DETAILS_SHEET,
    COMPETITOR_BENCHMARKS_SHEET,
    CONTENT_OPTIMISATION_HUB_SHEET,
    CONTENT_PLANNER_SHEET,
    IMAGE_INVENTORY_SHEET,
    ROBOTS_ANALYSIS_SHEET,
    SCRIPT_INVENTORY_SHEET,
    SNIPPET_OPPORTUNITIES_SHEET,
)
from hype_frog.reporter.sheets.workbook_layout import (
    ADVANCED_WORKBOOK_TAB_ORDER,
    HIDDEN_SHEETS_BY_DEFAULT,
    PREFERRED_WORKBOOK_TAB_ORDER,
    TAB_COLOR_MANAGEMENT,
    TAB_COLOR_TECHNICAL,
    VISIBLE_WORKBOOK_TAB_ORDER,
    _SHEET_TAB_COLORS,
    apply_workbook_tab_layout,
)


def test_visible_tab_order_matches_workflow_spec() -> None:
    visible = [n for n in VISIBLE_WORKBOOK_TAB_ORDER if n != "Table of Contents"]
    assert visible[:6] == [
        "Executive Briefing",
        "Playbook",
        "FixPlan",
        "Quick Wins",
        "Priority URLs",
        CONTENT_OPTIMISATION_HUB_SHEET,
    ]
    assert visible[6] == "Content & AI Readiness"
    assert visible[8] == "Broken Link Impact"
    assert visible[11] == "Main"
    assert AIOSEO_RECOMMENDATIONS_SHEET in visible
    assert CONTENT_PLANNER_SHEET in visible
    assert visible[-1] == IMAGE_INVENTORY_SHEET


def test_content_planner_positioned_after_content_optimisation_hub() -> None:
    visible = list(VISIBLE_WORKBOOK_TAB_ORDER)
    hub_idx = visible.index(CONTENT_OPTIMISATION_HUB_SHEET)
    content_ai_idx = visible.index("Content & AI Readiness")
    planner_idx = visible.index(CONTENT_PLANNER_SHEET)
    assert content_ai_idx == hub_idx + 1
    assert planner_idx > content_ai_idx


def test_content_planner_has_tab_colour() -> None:
    assert CONTENT_PLANNER_SHEET in _SHEET_TAB_COLORS


def test_advanced_tabs_hidden_by_default() -> None:
    assert "Technical Diagnostics" not in HIDDEN_SHEETS_BY_DEFAULT
    assert AUDIT_RUN_DETAILS_SHEET in HIDDEN_SHEETS_BY_DEFAULT
    assert "Dashboard" not in HIDDEN_SHEETS_BY_DEFAULT
    assert "Summary" not in HIDDEN_SHEETS_BY_DEFAULT


def test_aioseo_tab_colour_is_technical_persona() -> None:
    assert _SHEET_TAB_COLORS[AIOSEO_RECOMMENDATIONS_SHEET] == TAB_COLOR_TECHNICAL


def test_apply_workbook_tab_layout_orders_colors_and_hides() -> None:
    wb = Workbook()
    wb.active.title = "Main"
    for name in (
        "Executive Briefing",
        "Technical Diagnostics",
        AUDIT_RUN_DETAILS_SHEET,
        "Playbook",
    ):
        wb.create_sheet(name)
    apply_workbook_tab_layout(wb)
    titles = [ws.title for ws in wb.worksheets]
    assert titles.index("Executive Briefing") < titles.index("Main")
    assert titles.index("Playbook") < titles.index("Technical Diagnostics")
    briefing_color = str(wb["Executive Briefing"].sheet_properties.tabColor.rgb or "")
    assert briefing_color.upper().endswith(TAB_COLOR_MANAGEMENT.upper())
    assert wb["Technical Diagnostics"].sheet_state == "visible"
    assert wb[AUDIT_RUN_DETAILS_SHEET].sheet_state == "hidden"
    assert "Dashboard" not in wb.sheetnames
    assert len(PREFERRED_WORKBOOK_TAB_ORDER) == len(set(PREFERRED_WORKBOOK_TAB_ORDER))


def test_apply_workbook_tab_layout_shows_all_tabs_when_disabled() -> None:
    """New user-facing choice: hide_advanced_tabs=False must reveal every
    normally-hidden advanced/historical tab instead of hiding it."""
    wb = Workbook()
    wb.active.title = "Main"
    for name in (
        "Executive Briefing",
        "Issue Register",
        AUDIT_RUN_DETAILS_SHEET,
        "Playbook",
    ):
        wb.create_sheet(name)
    apply_workbook_tab_layout(wb, hide_advanced_tabs=False)
    assert wb["Issue Register"].sheet_state == "visible"
    assert wb[AUDIT_RUN_DETAILS_SHEET].sheet_state == "visible"


def test_all_ordered_tabs_have_a_tab_colour() -> None:
    """Every preferred tab except the TOC must carry a group tab colour (P1.6)."""
    for name in PREFERRED_WORKBOOK_TAB_ORDER:
        if name == "Table of Contents":
            continue
        assert name in _SHEET_TAB_COLORS, f"{name!r} has no tab colour"


def test_advanced_inventory_sheets_have_tab_colours() -> None:
    """Remaining advanced sheet (Snippet Opportunities/Script/Image Inventory were
    promoted to the visible primary workflow; Anchor Text Audit/Link Equity Map/
    Link Inventory were folded into Content & AI Readiness / Link Intelligence)."""
    for name in (COMPETITOR_BENCHMARKS_SHEET,):
        assert name in ADVANCED_WORKBOOK_TAB_ORDER
        assert name in _SHEET_TAB_COLORS


def test_promoted_advanced_sheets_are_now_visible() -> None:
    """Sheets promoted out of the hidden Advanced group by design (see
    workbook_layout.py) must appear in VISIBLE_WORKBOOK_TAB_ORDER, not
    ADVANCED_WORKBOOK_TAB_ORDER."""
    for name in (
        "Technical Diagnostics",
        "Content & AI Readiness",
        "Link Intelligence",
        ROBOTS_ANALYSIS_SHEET,
        SNIPPET_OPPORTUNITIES_SHEET,
        SCRIPT_INVENTORY_SHEET,
        IMAGE_INVENTORY_SHEET,
    ):
        assert name in VISIBLE_WORKBOOK_TAB_ORDER
        assert name not in ADVANCED_WORKBOOK_TAB_ORDER
