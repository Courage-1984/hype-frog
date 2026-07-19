"""Full-suite workbook sheet assembly for export orchestration."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any, Callable
from urllib.parse import urlparse

import pandas as pd
from openpyxl.utils import get_column_letter

from hype_frog.checkpoint.link_inventory_cache import LinkInventoryCache
from hype_frog.analysis.delta_engine import (
    RunSnapshot,
    build_delta_workbook_output,
    snapshot_from_current_run,
)
from hype_frog.analysis.link_equity import (
    build_anchor_text_audit_rows,
    build_link_equity_rows,
)
from hype_frog.analysis.snippet_opportunities import (
    SNIPPET_OPPORTUNITY_COLUMNS,
    build_snippet_opportunity_rows,
)
from hype_frog.analysis.third_party_scripts import (
    SCRIPT_INVENTORY_COLUMNS,
    build_script_inventory_rows,
)
from hype_frog.config import (
    DEFAULT_EFFORT_BY_SEVERITY,
    DEFAULT_OWNER_BY_SEVERITY,
    MAX_RETRIES,
    TIMEOUT_SECONDS,
)
from hype_frog.core import get_logger
from hype_frog.core.crawl_log import CRAWL_LOG_COLUMNS, crawl_log_sheet_rows
from hype_frog.core.models import ExtraRowPayload, MainRowPayload, SummaryMetricsPayload
from hype_frog.crawler.robots_mapping import (
    ROBOTS_ANALYSIS_COLUMNS,
    build_robots_analysis_rows,
)
from hype_frog.extractors.semantic_setup import probe_semantic_engine
from hype_frog.orchestration.crawl_runner import CrawlExecutionResult
from hype_frog.orchestration.enrichment_flow import EnrichmentResult
from hype_frog.orchestration.export_registry import (
    CMS_ACTION_URLS_COLUMNS,
    CMS_ACTION_URLS_SHEET,
    build_cms_action_url_rows,
    build_crawlgraph_rows,
    build_duplicates_rows,
    build_pattern_rows,
    build_priority_rows,
    build_sitemapqa_rows,
    get_merged_sheet_columns,
    get_standard_sheet_columns,
)
from hype_frog.orchestration.export_workbook_constants import (
    PLAYBOOK_LEGEND_ROWS,
    PLAYBOOK_QUICK_REFERENCE_ROWS,
)
from hype_frog.orchestration.run_setup import RunSetup
from hype_frog.pipeline.image_inventory import (
    IMAGE_INVENTORY_COLUMNS,
    build_image_inventory_rows,
)
from hype_frog.pipeline.link_inventory_stream import populate_link_inventory_cache
from hype_frog.pipeline.link_inventory import unique_external_health_counts
from hype_frog.reporter.engine_io import (
    append_link_detail_rows_streamed,
    apply_link_intelligence_summary_broken_formulas,
    write_dataframe_sheet,
)
from hype_frog.reporter.engine_rows import (
    CONTENT_HUB_EXPORT_COLUMNS,
    build_content_hub_metrics_for_all_urls,
)
from hype_frog.reporter.excel_engine import (
    build_content_optimisation_hub_rows,
    build_fixplan_rows,
    write_dict_rows_sheet,
)
from hype_frog.orchestration.content_planner import (
    CONTENT_PLANNER_COLUMNS,
    build_content_planner_rows,
)
from hype_frog.reporter.sheets.config import (
    AIOSEO_RECOMMENDATIONS_SHEET,
    AUDIT_RUN_DETAILS_SHEET,
    COMPETITOR_BENCHMARKS_SHEET,
    CONTENT_OPTIMISATION_HUB_SHEET,
    CONTENT_PLANNER_SHEET,
    CRAWL_LOG_SHEET,
    IMAGE_INVENTORY_SHEET,
    ROBOTS_ANALYSIS_SHEET,
    SCRIPT_INVENTORY_SHEET,
    SNIPPET_OPPORTUNITIES_SHEET,
)
from hype_frog.reporter.sheets.executive_dashboard import write_executive_briefing
from hype_frog.reporter.sheets.merged_builders import (
    LINK_INTELLIGENCE_COLUMNS,
    build_broken_link_impact_rows,
    build_content_ai_readiness_rows,
    build_issue_register_rows,
    build_link_intelligence_rows,
    build_quick_wins_rows,
    build_redirects_sheet_rows,
    build_technical_diagnostics_rows,
    build_template_duplication_risks_rows,
)
from hype_frog.reporter.stream_workbook import is_write_only_writer
from hype_frog.reporter.summary_builder import (
    build_issue_inventory_rows,
    build_summary_rows,
)
from hype_frog.rules import owner_for_issue, root_cause_and_fix
from hype_frog.rules.playbook_entries import (
    build_issue_playbook_rows,
    build_playbook_entry_index,
)

logger = get_logger(__name__)


def apply_deferred_readwrite_export_steps(
    writer: Any,
    *,
    summary_metrics: SummaryMetricsPayload | None,
    typed_main_rows: list[MainRowPayload],
    typed_extra_rows: list[ExtraRowPayload],
    priority_rows: list[dict[str, Any]],
    fixplan_rows: list[dict[str, Any]],
    hub_metrics_rows: list[dict[str, Any]] | None,
) -> None:
    """Run formula injection and briefing layout after a write_only streaming pass."""
    apply_link_intelligence_summary_broken_formulas(writer.book)
    if "Link Intelligence" in writer.book.sheetnames:
        _ws_li = writer.book["Link Intelligence"]
        _comp_idx = list(LINK_INTELLIGENCE_COLUMNS).index("Broken Links (computed)") + 1
        _ws_li.column_dimensions[get_column_letter(_comp_idx)].hidden = True
    if summary_metrics is not None:
        write_executive_briefing(
            writer,
            summary_metrics=summary_metrics,
            typed_main_rows=typed_main_rows,
            typed_extra_rows=typed_extra_rows,
            priority_rows=priority_rows,
            fixplan_rows=fixplan_rows,
            hub_metrics_rows=hub_metrics_rows,
        )

_AEO_ISSUE_NAMES = frozenset({
    "Low AEO Readiness Score",
    "Missing FAQ/QA Schema",
    "No Question Headings",
    "No Answer-Friendly Structure",
    "No 40-60 Word Answer Paragraphs",
})

_FIXPLAN_TOP_BLOCKER_COLS = ("Issue Type", "Severity", "Affected Count")


@dataclass(frozen=True)
class WorkbookExportContext:
    setup: RunSetup
    crawl_result: CrawlExecutionResult
    enrichment: EnrichmentResult
    output_filename: str
    main_rows: list[dict[str, Any]]
    extra_rows: list[dict[str, Any]]
    typed_main_rows: list[MainRowPayload]
    typed_extra_rows: list[ExtraRowPayload]
    status_by_url: dict[str, Any]
    main_by_url: dict[str, dict[str, Any]]
    summary_rules: list[Any]
    previous_snapshot: RunSnapshot | None
    high_value_slugs: list[str]
    value_or_default_fn: Callable[[object, float], float]
    extract_subfolder_fn: Callable[[str], str]
    build_aioseo_rows_fn: Callable[
        [list[dict[str, object]], dict[str, dict[str, object]], dict[str, str]],
        list[dict[str, object]],
    ]


@dataclass(frozen=True)
class FullSuiteExportResult:
    summary_rows: list[dict[str, Any]]
    fixplan_rows: list[dict[str, Any]]
    quick_wins_rows: list[dict[str, Any]]
    priority_rows: list[dict[str, Any]]
    broken_link_impact_rows: list[dict[str, Any]]
    run_timestamp: str
    summary_metrics: SummaryMetricsPayload | None
    current_snapshot: RunSnapshot | None
    hub_metrics_rows: list[dict[str, Any]] | None = None


def write_full_suite_workbook(
    writer: pd.ExcelWriter,
    ctx: WorkbookExportContext,
) -> FullSuiteExportResult:
    """Write all full-suite workbook tabs and return artefacts for executive reports."""
    setup = ctx.setup
    crawl_result = ctx.crawl_result
    enrichment = ctx.enrichment
    output_filename = ctx.output_filename
    main_rows = ctx.main_rows
    extra_rows = ctx.extra_rows
    typed_main_rows = ctx.typed_main_rows
    typed_extra_rows = ctx.typed_extra_rows
    status_by_url = ctx.status_by_url
    main_by_url = ctx.main_by_url
    summary_rules = ctx.summary_rules
    previous_snapshot = ctx.previous_snapshot
    high_value_slugs = ctx.high_value_slugs
    value_or_default_fn = ctx.value_or_default_fn
    extract_subfolder_fn = ctx.extract_subfolder_fn
    build_aioseo_rows_fn = ctx.build_aioseo_rows_fn
    urls = crawl_result.crawl_urls
    sitemap_meta = crawl_result.sitemap_meta
    sitemap_files_meta = crawl_result.sitemap_files_meta
    workers = crawl_result.workers
    request_delay = crawl_result.request_delay
    checkpoint_every = crawl_result.checkpoint_every
    previous_audit_path = crawl_result.previous_audit_path
    current_snapshot: RunSnapshot | None = None
    summary_rows: list[dict[str, Any]] = []
    fixplan_rows: list[dict[str, Any]] = []
    quick_wins_rows: list[dict[str, Any]] = []
    priority_rows: list[dict[str, Any]] = []
    broken_link_impact_rows: list[dict[str, Any]] = []
    run_timestamp = ""
    summary_metrics: SummaryMetricsPayload | None = None

    logger.info("Building full audit: %d URLs...", len(typed_extra_rows))
    sheet_columns = get_standard_sheet_columns()
    merged_columns = get_merged_sheet_columns()
    aioseo_rows = build_aioseo_rows_fn(extra_rows, main_by_url, DEFAULT_OWNER_BY_SEVERITY)
    write_dict_rows_sheet(
        writer,
        AIOSEO_RECOMMENDATIONS_SHEET,
        sheet_columns[AIOSEO_RECOMMENDATIONS_SHEET],
        aioseo_rows,
    )
    redirects_rows = build_redirects_sheet_rows(extra_rows)
    duplicate_rows = build_duplicates_rows(main_rows, extra_rows)
    pattern_rows, template_issue_counts = build_pattern_rows(
        extra_rows,
        extract_subfolder_fn=extract_subfolder_fn,
    )
    aeo_issue_names = _AEO_ISSUE_NAMES
    summary_rows = build_summary_rows(
        summary_rules,
        typed_extra_rows,
        template_issue_counts,
        value_or_default_fn,
        main_rows=typed_main_rows,
    )
    issue_inventory_rows = build_issue_inventory_rows(
        summary_rules,
        typed_extra_rows,
        main_rows=typed_main_rows,
    )
    issue_inventory_df = pd.DataFrame(issue_inventory_rows)
    run_timestamp = datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    register_snapshot = snapshot_from_current_run(
        issue_inventory_df=issue_inventory_df,
        main_rows=main_rows,
        extra_rows=extra_rows,
        source_path=output_filename,
        run_date=run_timestamp,
        previous_snapshot=previous_snapshot,
    )
    technical_diagnostics_rows = build_technical_diagnostics_rows(
        extra_rows, main_rows=main_rows
    )
    # Folded into Content & AI Readiness (formerly standalone "Content Hub Metrics"
    # and "Anchor Text Audit" sheets) — computed for every URL, then left-joined below.
    hub_metrics_by_url = {
        r["URL"]: r
        for r in build_content_hub_metrics_for_all_urls(typed_main_rows, typed_extra_rows)
        if r.get("URL")
    }
    anchor_audit_by_url = {
        r["Destination URL"]: r for r in build_anchor_text_audit_rows(extra_rows)
    }
    content_ai_rows = build_content_ai_readiness_rows(
        extra_rows,
        main_rows=main_rows,
        hub_metrics_by_url=hub_metrics_by_url,
        anchor_audit_by_url=anchor_audit_by_url,
    )
    issue_register_rows = build_issue_register_rows(
        summary_rows=summary_rows,
        issue_inventory_rows=issue_inventory_rows,
        issue_records=register_snapshot.issues,
        run_date=run_timestamp,
    )
    graph_rows = build_crawlgraph_rows(
        main_urls=[str(row.get("URL") or "") for row in main_rows if row.get("URL")],
        extra_rows=extra_rows,
    )
    # Folded into Link Intelligence's Summary rows (formerly standalone "Link
    # Equity Map" sheet).
    graph_metrics = enrichment.graph_metrics or {}
    link_equity_by_url = {
        r["URL"]: r for r in build_link_equity_rows(extra_rows, graph_metrics)
    }
    link_intelligence_rows = build_link_intelligence_rows(
        extra_rows=extra_rows,
        crawlgraph_rows=graph_rows,
        main_rows=main_rows,
        link_equity_by_url=link_equity_by_url,
    )
    logger.info("Writing summary and issue sheets...")
    # No standalone "Summary" tab: build_issue_register_rows() (above) already folds
    # every summary_rows entry (Issue Counts, AEO Opportunities, Top 10 Critical URLs,
    # Top Issues by Template) into Issue Register, so Summary was a strict subset.
    template_duplication_rows = build_template_duplication_risks_rows(
        duplicate_rows=duplicate_rows,
        pattern_rows=pattern_rows,
    )
    write_dict_rows_sheet(
        writer,
        "Issue Register",
        merged_columns["Issue Register"],
        issue_register_rows,
    )
    write_dict_rows_sheet(
        writer,
        "Technical Diagnostics",
        merged_columns["Technical Diagnostics"],
        technical_diagnostics_rows,
    )
    write_dict_rows_sheet(
        writer,
        "Content & AI Readiness",
        merged_columns["Content & AI Readiness"],
        content_ai_rows,
    )
    write_dict_rows_sheet(
        writer,
        "Link Intelligence",
        merged_columns["Link Intelligence"],
        link_intelligence_rows,
    )
    cms_action_rows = build_cms_action_url_rows(
        crawl_result.excluded_cms_action_urls,
        extra_rows,
    )
    write_dict_rows_sheet(
        writer,
        CMS_ACTION_URLS_SHEET,
        CMS_ACTION_URLS_COLUMNS,
        cms_action_rows,
    )
    logger.info("Writing link and duplication analysis sheets...")
    link_inventory_db = crawl_result.output_filename.replace(".xlsx", "_link_inventory.db")
    link_inventory_cache = LinkInventoryCache(link_inventory_db)
    try:
        populate_link_inventory_cache(link_inventory_cache, extra_rows)
        # Detail rows (deduplicated, streamed anchor-level rows) appended after the
        # Summary rows already written above — folded in from the former standalone
        # "Link Inventory" sheet.
        append_link_detail_rows_streamed(
            writer,
            link_inventory_cache,
            sheet_name="Link Intelligence",
            columns=list(merged_columns["Link Intelligence"]),
            status_by_url=status_by_url,
        )
        broken_link_impact_rows = build_broken_link_impact_rows(
            link_inventory_cache.iter_rows_flat(),
            extra_rows,
        )
        ok_unique_ext, total_unique_ext = unique_external_health_counts(
            link_inventory_cache.iter_rows_flat()
        )
    finally:
        link_inventory_cache.close(cleanup_file=True)
    write_dict_rows_sheet(
        writer,
        "Broken Link Impact",
        merged_columns["Broken Link Impact"],
        broken_link_impact_rows,
    )
    if not is_write_only_writer(writer):
        apply_link_intelligence_summary_broken_formulas(writer.book)
        _ws_li = writer.book["Link Intelligence"]
        _comp_idx = list(LINK_INTELLIGENCE_COLUMNS).index("Broken Links (computed)") + 1
        _ws_li.column_dimensions[get_column_letter(_comp_idx)].hidden = True

    write_dict_rows_sheet(
        writer,
        "Template & Duplication Risks",
        merged_columns["Template & Duplication Risks"],
        template_duplication_rows,
    )
    playbook_index = build_playbook_entry_index(summary_rules)
    fixplan_rows = build_fixplan_rows(
        summary_rules, typed_extra_rows, aeo_issue_names, root_cause_and_fix,
        DEFAULT_EFFORT_BY_SEVERITY, DEFAULT_OWNER_BY_SEVERITY,
        playbook_index,
    )
    _FIXPLAN_TOP_BLOCKER_COLS = ("Issue Type", "Severity", "Affected Count")
    if fixplan_rows:
        fixplan_df = pd.DataFrame(
            sorted(
                fixplan_rows,
                key=lambda item: (-item["Affected Count"], item["Severity"]),
            )
        )
    else:
        fixplan_df = pd.DataFrame(columns=list(_FIXPLAN_TOP_BLOCKER_COLS))
    write_dataframe_sheet(writer, fixplan_df, "FixPlan", startrow=1)
    # Computed here (rather than at its previous call site below) so Quick Wins
    # can look up the same Business Risk Score used by Priority URLs, instead
    # of reading a field that's never set on the raw extra_rows dicts.
    priority_rows = build_priority_rows(
        extra_rows,
        high_value_slugs=high_value_slugs,
        value_or_default_fn=value_or_default_fn,
        owner_for_issue_fn=owner_for_issue,
    )
    risk_score_by_url = {
        str(row.get("URL") or ""): row.get("Business Risk Score", 0)
        for row in priority_rows
    }
    quick_wins_rows = build_quick_wins_rows(
        extra_rows, fixplan_rows, summary_rules, playbook_index, risk_score_by_url
    )
    write_dict_rows_sheet(
        writer,
        "Quick Wins",
        merged_columns["Quick Wins"],
        quick_wins_rows,
    )
    hub_base_rows, hub_metrics_rows = build_content_optimisation_hub_rows(
        typed_main_rows, typed_extra_rows, fixplan_rows
    )
    content_hub_cols = list(CONTENT_HUB_EXPORT_COLUMNS)
    write_dict_rows_sheet(
        writer, CONTENT_OPTIMISATION_HUB_SHEET, content_hub_cols, hub_base_rows
    )
    # No standalone "Content Hub Metrics" tab: its 11 columns are now computed for
    # every URL (see hub_metrics_by_url above) and folded into Content & AI Readiness.
    # hub_metrics_rows (the Hub's curated-subset variant) still feeds Executive
    # Briefing's in-memory ROI KPI below.
    _parsed = urlparse(setup.target_input)
    _root_url = f"{_parsed.scheme}://{_parsed.netloc}/"
    content_planner_rows = build_content_planner_rows(typed_extra_rows, root_url=_root_url)
    write_dict_rows_sheet(
        writer,
        CONTENT_PLANNER_SHEET,
        list(CONTENT_PLANNER_COLUMNS),
        content_planner_rows,
    )
    # Remaining dashboard/delta/report tabs preserved from prior flow
    # (priority_rows already computed above, before Quick Wins)
    priority_df = pd.DataFrame(priority_rows)
    write_dataframe_sheet(writer, priority_df, "Priority URLs", startrow=1)
    total_urls = len(typed_extra_rows)
    pass_count = sum(
        1
        for row in typed_extra_rows
        if str(row.values.get("Severity Badge") or "") == "Pass"
    )
    critical_count = sum(
        1
        for row in typed_extra_rows
        if str(row.values.get("Severity Badge") or "") == "Critical"
    )
    warning_count = sum(
        1
        for row in typed_extra_rows
        if str(row.values.get("Severity Badge") or "") == "Warning"
    )
    seo_health_values: list[float] = []
    for row in typed_extra_rows:
        raw_hs = row.values.get("SEO Health Score")
        if raw_hs is None or str(raw_hs).strip() == "":
            continue
        try:
            seo_health_values.append(float(raw_hs))
        except (TypeError, ValueError):
            continue
    avg_seo_health_pct = (
        round(sum(seo_health_values) / len(seo_health_values), 2)
        if seo_health_values
        else 0.0
    )
    seo_pass_pct = round((pass_count / max(1, total_urls)) * 100.0, 2)
    to_do_share = (critical_count + warning_count) / max(1, total_urls)
    projected_health_pct = min(
        100.0,
        round(
            avg_seo_health_pct
            + max(0.0, 100.0 - avg_seo_health_pct) * to_do_share * 0.9,
            2,
        ),
    )
    projected_pass_pct = min(
        100.0,
        round(
            seo_pass_pct
            + max(0.0, 100.0 - seo_pass_pct) * to_do_share * 0.85,
            2,
        ),
    )
    summary_metrics = SummaryMetricsPayload(
        urls_crawled=total_urls,
        seo_pass_rate_pct=seo_pass_pct,
        health_score_pct=avg_seo_health_pct,
        critical_url_count=critical_count,
        warning_url_count=warning_count,
        projected_health_score_pct=projected_health_pct,
        projected_pass_rate_pct=projected_pass_pct,
    )
    if not is_write_only_writer(writer):
        write_executive_briefing(
            writer,
            summary_metrics=summary_metrics,
            typed_main_rows=typed_main_rows,
            typed_extra_rows=typed_extra_rows,
            priority_rows=priority_rows,
            fixplan_rows=fixplan_rows,
            hub_metrics_rows=hub_metrics_rows,
        )
    playbook_rows = list(PLAYBOOK_QUICK_REFERENCE_ROWS)

    playbook_rows.extend(
        [
            {"Section": "", "Item": "", "Guideline": "", "Why It Matters": ""},
            {
                "Section": "[Issue Playbook]",
                "Item": "",
                "Guideline": "",
                "Why It Matters": "",
            },
        ]
    )
    for issue_row in build_issue_playbook_rows(summary_rules):
        playbook_rows.append(
            {
                "Section": issue_row["Section"],
                "Item": issue_row["Issue"],
                "Guideline": (
                    f"What: {issue_row['What It Is']}\n"
                    f"Fix: {issue_row['How To Fix']}\n"
                    f"Verify: {issue_row['How To Verify']}"
                ),
                "Why It Matters": (
                    f"{issue_row['Why It Matters']} "
                    f"(Severity: {issue_row['Severity']}; "
                    f"Owner: {issue_row['Owner']}; "
                    f"Time: {issue_row['Time To Fix']})"
                ),
            }
        )
    semantic_probe = probe_semantic_engine()
    crawl_semantic_modes = {
        str(row.values.get("Semantic Analysis Mode") or "").strip()
        for row in typed_extra_rows
        if str(row.values.get("Semantic Analysis Mode") or "").strip()
    }
    gsc_freshness = next(
        (
            str(row.values.get("GSC Data Freshness") or "").strip()
            for row in typed_extra_rows
            if str(row.values.get("GSC Data Freshness") or "").strip()
        ),
        "",
    )
    gsc_matched_urls = sum(
        1
        for row in typed_extra_rows
        if str(row.values.get("GSC Coverage Note") or "").startswith("Matched in GSC")
    )
    gsc_unmatched_urls = sum(
        1
        for row in typed_extra_rows
        if "No Search Analytics row matched" in str(row.values.get("GSC Coverage Note") or "")
    )
    gsc_low_volume_urls = sum(
        1
        for row in typed_extra_rows
        if "low impressions" in str(row.values.get("GSC Coverage Note") or "").lower()
    )
    crawl_duration_s = round(crawl_result.crawl_duration_seconds, 1)
    rendered_source_count = sum(
        1 for row in typed_extra_rows if row.values.get("Extraction Source") == "rendered_browser"
    )
    raw_source_count = sum(
        1 for row in typed_extra_rows if row.values.get("Extraction Source") == "raw_http"
    )
    render_fallback_count = sum(
        1 for row in typed_extra_rows if row.values.get("Extraction Source Fallback")
    )
    logger.debug("Crawl duration for Dashboard RunMetadata: %.1fs", crawl_duration_s)
    target_site = urlparse(urls[0]).netloc if urls else "Unknown"
    run_meta_rows = [
        {"Key": "Target Site", "Value": target_site},
        {"Key": "Run Timestamp", "Value": run_timestamp},
        {"Key": "Total URLs", "Value": len(urls)},
        {"Key": "Duration (s)", "Value": crawl_duration_s},
        {"Key": "Crawl Mode", "Value": setup.crawl_mode},
        {
            "Key": "Extraction Source Rendered Count",
            "Value": rendered_source_count,
        },
        {"Key": "Extraction Source Raw HTTP Count", "Value": raw_source_count},
        {"Key": "Extraction Source Fallback Count", "Value": render_fallback_count},
        {"Key": "GSC Data Freshness", "Value": gsc_freshness or "Not available"},
        {
            "Key": "GSC Coverage Note",
            "Value": (
                f"{gsc_matched_urls} URL(s) matched Search Analytics; "
                f"{gsc_unmatched_urls} unmatched; "
                f"{gsc_low_volume_urls} low-impression (CTR directional only)"
                if gsc_freshness
                else "GSC Search Analytics not loaded for this run"
            ),
        },
        {"Key": "Mode", "Value": "Full Suite"},
        {
            "Key": "Semantic Engine (install probe)",
            "Value": semantic_probe.message,
        },
        {
            "Key": "Semantic Analysis Modes (crawl)",
            "Value": ", ".join(sorted(crawl_semantic_modes)) or "N/A",
        },
        {"Key": "Workers", "Value": workers},
        {"Key": "Delay Seconds", "Value": request_delay},
        {"Key": "Retries", "Value": MAX_RETRIES},
        {"Key": "Timeout Seconds", "Value": TIMEOUT_SECONDS},
        {"Key": "Checkpoint Every", "Value": checkpoint_every},
        {"Key": "Previous Audit Path", "Value": previous_audit_path or "Not supplied"},
        {
            "Key": "External Link Unique Denominator",
            "Value": int(total_unique_ext),
        },
        {"Key": "External Link Unique 200 OK", "Value": int(ok_unique_ext)},
        {
            "Key": "External Sniff Performed",
            "Value": int(1 if setup.check_external_link_status else 0),
        },
        {
            "Key": "OG Image Validation Performed",
            "Value": int(1 if setup.check_og_images else 0),
        },
    ]
    write_dataframe_sheet(
        writer, pd.DataFrame(run_meta_rows), AUDIT_RUN_DETAILS_SHEET, startrow=1
    )
    # No standalone "ResolvedIssues" tab: build_delta_sheet_rows() already folds
    # every resolved issue (with Stable Issue ID) into DeltaFromPreviousRun's own
    # "Resolved Issues" section, so the separate tab was a strict subset.
    delta_rows, _resolved_issues_df, current_snapshot = build_delta_workbook_output(
        issue_inventory_df=issue_inventory_df,
        main_rows=main_rows,
        extra_rows=extra_rows,
        summary_rules=summary_rules,
        previous_snapshot=previous_snapshot,
        baseline_report=previous_snapshot is None,
        output_path=output_filename,
        run_date=run_timestamp,
    )
    write_dataframe_sheet(writer, pd.DataFrame(delta_rows), "DeltaFromPreviousRun", startrow=1)
    playbook_rows.extend(
        [
            {"Section": "", "Item": "", "Guideline": "", "Why It Matters": ""},
            {
                "Section": "[Glossary & Legend]",
                "Item": "",
                "Guideline": "",
                "Why It Matters": "",
            },
        ]
    )
    for row in PLAYBOOK_LEGEND_ROWS:
        playbook_rows.append(
            {
                "Section": row.get("Section", ""),
                "Item": row.get("Term", ""),
                "Guideline": row.get("Values/Threshold", ""),
                "Why It Matters": row.get("Meaning", ""),
            }
        )
    write_dataframe_sheet(writer, pd.DataFrame(playbook_rows), "Playbook", startrow=1)
    sitemap_rows = build_sitemapqa_rows(
        sitemap_meta=sitemap_meta,
        sitemap_files_meta=sitemap_files_meta,
        extra_rows=extra_rows,
    )
    write_dataframe_sheet(writer, pd.DataFrame(sitemap_rows), "SitemapQA", startrow=1)
    write_dict_rows_sheet(
        writer,
        "Redirects",
        sheet_columns["Redirects"],
        redirects_rows,
    )
    robots_analysis_rows = build_robots_analysis_rows(
        robots_by_domain=crawl_result.robots_by_domain or {},
        extra_rows=extra_rows,
        sitemap_url_keys=enrichment.sitemap_url_keys,
    )
    write_dict_rows_sheet(
        writer,
        ROBOTS_ANALYSIS_SHEET,
        list(ROBOTS_ANALYSIS_COLUMNS),
        robots_analysis_rows,
    )
    write_dict_rows_sheet(
        writer,
        CRAWL_LOG_SHEET,
        list(CRAWL_LOG_COLUMNS),
        crawl_log_sheet_rows(enrichment.crawl_log_entries),
    )
    # No standalone "Link Equity Map" tab: its rows (computed once, above, as
    # link_equity_by_url) are now folded into Link Intelligence's Summary rows.
    # No standalone "Anchor Text Audit" tab: its rows (computed once, above, as
    # anchor_audit_by_url) are now folded into Content & AI Readiness.
    write_dict_rows_sheet(
        writer,
        SNIPPET_OPPORTUNITIES_SHEET,
        list(SNIPPET_OPPORTUNITY_COLUMNS),
        build_snippet_opportunity_rows(extra_rows),
    )
    write_dict_rows_sheet(
        writer,
        SCRIPT_INVENTORY_SHEET,
        list(SCRIPT_INVENTORY_COLUMNS),
        build_script_inventory_rows(extra_rows),
    )
    write_dict_rows_sheet(
        writer,
        IMAGE_INVENTORY_SHEET,
        list(IMAGE_INVENTORY_COLUMNS),
        build_image_inventory_rows(
            extra_rows,
            enrichment.image_probe_by_url or {},
        ),
    )
    if enrichment.competitor_benchmark_rows is not None:
        write_dict_rows_sheet(
            writer,
            COMPETITOR_BENCHMARKS_SHEET,
            list(enrichment.competitor_benchmark_columns or ("Metric", "Client Site")),
            enrichment.competitor_benchmark_rows,
        )
    return FullSuiteExportResult(
        summary_rows=summary_rows,
        fixplan_rows=fixplan_rows,
        quick_wins_rows=quick_wins_rows,
        priority_rows=priority_rows,
        broken_link_impact_rows=broken_link_impact_rows,
        run_timestamp=run_timestamp,
        summary_metrics=summary_metrics,
        current_snapshot=current_snapshot,
        hub_metrics_rows=hub_metrics_rows,
    )
