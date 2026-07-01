from __future__ import annotations

from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

from hype_frog.core import get_logger
from hype_frog.reporter.sheets.config import (
    CONTENT_OPTIMISATION_HUB_SHEET,
    CONTENT_PLANNER_SHEET,
    DATA_BAR_BLUE,
    DATA_SHEET_FREEZE_PANES,
    DISABLE_CONDITIONAL_FORMATTING,
    RAG_AMBER,
    RAG_AMBER_FONT,
    RAG_AMBER_SOFT,
    RAG_GREEN,
    RAG_GREEN_FONT,
    RAG_NEUTRAL,
    RAG_RED,
    RAG_RED_FONT,
    STATUS_TODO_FILL,
    STATUS_TODO_FONT,
)

logger = get_logger(__name__)


def apply_fixplan_workflow_formatting(worksheet: Worksheet) -> None:
    headers = [cell.value for cell in worksheet[1]]
    header_to_col = {str(h): i + 1 for i, h in enumerate(headers) if h is not None}
    priority_col = header_to_col.get("Priority Score")
    points_col = header_to_col.get("Est. Sprint Points")
    aging_col = header_to_col.get("Aging/Priority")
    critical_fill = PatternFill(start_color=RAG_RED, end_color=RAG_RED, fill_type="solid")
    warning_fill = PatternFill(start_color=RAG_AMBER, end_color=RAG_AMBER, fill_type="solid")
    good_fill = PatternFill(start_color=RAG_GREEN, end_color=RAG_GREEN, fill_type="solid")
    edge_fill = PatternFill(start_color=RAG_NEUTRAL, end_color=RAG_NEUTRAL, fill_type="solid")
    for row_idx in range(2, worksheet.max_row + 1):
        if priority_col:
            cell = worksheet.cell(row=row_idx, column=priority_col)
            try:
                score = int(cell.value or 0)
                if score >= 100:
                    cell.fill = critical_fill
                elif score >= 65:
                    cell.fill = warning_fill
                else:
                    cell.fill = edge_fill
            except Exception as exc:
                logger.debug(
                    "FixPlan priority score formatting skipped at row %s: %s",
                    row_idx,
                    exc,
                )
            cell = worksheet.cell(row=row_idx, column=points_col)
            try:
                points = int(cell.value or 0)
                if points >= 8:
                    cell.fill = critical_fill
                elif points >= 5:
                    cell.fill = warning_fill
                else:
                    cell.fill = good_fill
            except Exception as exc:
                logger.debug(
                    "FixPlan sprint points formatting skipped at row %s: %s",
                    row_idx,
                    exc,
                )
        if aging_col:
            cell = worksheet.cell(row=row_idx, column=aging_col)
            value = str(cell.value or "").lower()
            if "immediate" in value:
                cell.fill = critical_fill
            elif "next sprint" in value:
                cell.fill = warning_fill
            elif "backlog" in value:
                cell.fill = edge_fill


def apply_workflow_status_conditional_formatting(
    worksheet: Worksheet,
    status_col: int,
    *,
    first_row: int,
    last_row: int,
) -> None:
    """Uniform RAG styling for workflow ``Status`` cells (Phase 5)."""
    if DISABLE_CONDITIONAL_FORMATTING or last_row < first_row:
        return
    col = get_column_letter(status_col)
    rng = f"{col}{first_row}:{col}{last_row}"
    top = f"{col}{first_row}"
    done_fill = PatternFill("solid", fgColor=RAG_GREEN)
    done_font = Font(color=RAG_GREEN_FONT)
    progress_fill = PatternFill("solid", fgColor=RAG_AMBER)
    progress_font = Font(color=RAG_AMBER_FONT)
    todo_fill = PatternFill("solid", fgColor=STATUS_TODO_FILL)
    todo_font = Font(color=STATUS_TODO_FONT)
    legacy_done = (
        f'OR(LOWER({top})="done",LOWER({top})="completed",'
        f'LOWER({top})="fixed",LOWER({top})="closed")'
    )
    legacy_progress = (
        f'OR(LOWER({top})="in progress",LOWER({top})="in review",'
        f'LOWER({top})="review")'
    )
    legacy_todo = f'OR(LOWER({top})="to do",LOWER({top})="open")'
    for formula, fill, font in (
        (legacy_done, done_fill, done_font),
        (legacy_progress, progress_fill, progress_font),
        (legacy_todo, todo_fill, todo_font),
    ):
        worksheet.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[formula],
                stopIfTrue=True,
                fill=fill,
                font=font,
            ),
        )


def _legacy_sheet_header_index(
    worksheet: Worksheet, header_row: int = 1
) -> dict[str, int]:
    return {
        str(cell.value): idx
        for idx, cell in enumerate(worksheet[header_row], start=1)
        if cell.value
    }


def ensure_auto_filter(worksheet: Worksheet) -> None:
    from hype_frog.reporter.sheets.sheet_rows import sheet_data_header_row

    if worksheet.title == CONTENT_PLANNER_SHEET:
        if worksheet.max_row >= 2 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = (
                f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
            )
        else:
            worksheet.auto_filter.ref = None
        return

    if worksheet.title not in {"Main", "Dashboard", "Link Inventory"} and (
        worksheet.max_row < 10 or worksheet.max_column < 5
    ):
        worksheet.auto_filter.ref = None
        return

    header_row = sheet_data_header_row(worksheet.title)
    if worksheet.title == CONTENT_OPTIMISATION_HUB_SHEET:
        header_row = 2
    if worksheet.max_row >= header_row + 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
        )
    else:
        worksheet.auto_filter.ref = None


def _clear_orphaned_selection(worksheet: Worksheet) -> None:
    try:
        worksheet.views.sheetView[0].selection = []
    except Exception as exc:
        logger.debug("Could not clear worksheet selection for %s: %s", worksheet.title, exc)


def ensure_freeze_header(worksheet: Worksheet) -> None:
    if worksheet.title in {CONTENT_OPTIMISATION_HUB_SHEET, CONTENT_PLANNER_SHEET}:
        return
    if worksheet.title not in {"Main", "Dashboard", "Link Inventory"} and (
        worksheet.max_row < 10 or worksheet.max_column < 5
    ):
        worksheet.freeze_panes = None
        _clear_orphaned_selection(worksheet)
        return
    if worksheet.max_row > 1 and worksheet.max_column >= 1:
        worksheet.freeze_panes = DATA_SHEET_FREEZE_PANES
    else:
        worksheet.freeze_panes = None
        _clear_orphaned_selection(worksheet)


def ensure_print_setup(worksheet: Worksheet) -> None:
    """Fit-to-width landscape print layout so wide sheets paginate sanely.

    Without this, printing any sheet uses Excel's default pagination — the
    217-column ``Main`` sheet alone would spool dozens of near-blank pages.
    """
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    worksheet.print_area = (
        f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
    )


def apply_global_conditional_formatting(
    worksheet: Worksheet,
    *,
    merged_audit_tabs: bool = False,
    skip_headers: frozenset[str] = frozenset(),
    header_row: int = 1,
) -> None:
    # Honour the kill switch defensively, regardless of caller.
    if DISABLE_CONDITIONAL_FORMATTING or worksheet.max_row <= header_row:
        return
    headers = _legacy_sheet_header_index(worksheet, header_row)
    data_start = header_row + 1
    # Drop columns owned by a more specific pass (e.g. Main heatmaps) so we do not
    # stack two conflicting rules on the same range.
    for skipped in skip_headers:
        headers.pop(skipped, None)
    last_row = worksheet.max_row
    status_col = headers.get("Status Code") or headers.get("Target Status (if crawled)")
    if status_col:
        col = get_column_letter(status_col)
        rng = f"{col}{data_start}:{col}{last_row}"
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="equal",
                formula=["200"],
                fill=PatternFill("solid", fgColor=RAG_GREEN),
            ),
        )
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="between",
                formula=["300", "399"],
                fill=PatternFill("solid", fgColor=RAG_AMBER),
            ),
        )
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=["400"],
                fill=PatternFill("solid", fgColor=RAG_RED),
            ),
        )

    for load_header in ("Load Time (s)", "Load Time", "TTFB (ms)"):
        load_col = headers.get(load_header)
        if load_col:
            col = get_column_letter(load_col)
            rng = f"{col}{data_start}:{col}{last_row}"
            worksheet.conditional_formatting.add(
                rng,
                ColorScaleRule(
                    start_type="min",
                    start_color="63BE7B",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFEB84",
                    end_type="max",
                    end_color="F8696B",
                ),
            )
            worksheet.conditional_formatting.add(
                rng,
                DataBarRule(
                    start_type="min",
                    end_type="max",
                    color=DATA_BAR_BLUE,
                    showValue=True,
                ),
            )
            break

    for wc_header in ("Word Count", "Word Count (Body)"):
        wc_col = headers.get(wc_header)
        if wc_col:
            col = get_column_letter(wc_col)
            rng = f"{col}{data_start}:{col}{last_row}"
            worksheet.conditional_formatting.add(
                rng,
                DataBarRule(
                    start_type="min",
                    end_type="max",
                    color=DATA_BAR_BLUE,
                    showValue=True,
                ),
            )
            break

    priority_col = headers.get("Priority Score")
    if priority_col:
        col = get_column_letter(priority_col)
        rng = f"{col}{data_start}:{col}{last_row}"
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=["85"],
                font=Font(color=RAG_RED_FONT, bold=True),
                fill=PatternFill("solid", fgColor=RAG_RED),
            ),
        )
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="between",
                formula=["65", "84"],
                font=Font(color=RAG_AMBER_FONT, bold=True),
                fill=PatternFill("solid", fgColor=RAG_AMBER),
            ),
        )

    seo_score_col = headers.get("SEO Health Score")
    if seo_score_col and not merged_audit_tabs:
        col = get_column_letter(seo_score_col)
        worksheet.conditional_formatting.add(
            f"{col}{data_start}:{col}{last_row}",
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )

    aeo_score_col = headers.get("AEO Readiness Score")
    if aeo_score_col:
        col = get_column_letter(aeo_score_col)
        worksheet.conditional_formatting.add(
            f"{col}{data_start}:{col}{last_row}",
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )

    for score_header in ("Desktop Score", "Mobile Score"):
        score_col = headers.get(score_header)
        if score_col:
            col = get_column_letter(score_col)
            rng = f"{col}{data_start}:{col}{last_row}"
            worksheet.conditional_formatting.add(
                rng,
                CellIsRule(
                    operator="between",
                    formula=["90", "100"],
                    fill=PatternFill("solid", fgColor="C6EFCE"),
                ),
            )
            worksheet.conditional_formatting.add(
                rng,
                CellIsRule(
                    operator="between",
                    formula=["50", "89"],
                    fill=PatternFill("solid", fgColor="FFEB9C"),
                ),
            )
            worksheet.conditional_formatting.add(
                rng,
                CellIsRule(
                    operator="between",
                    formula=["0", "49"],
                    fill=PatternFill("solid", fgColor="FFC7CE"),
                ),
            )

    lcp_col = headers.get("Mobile LCP")
    if lcp_col:
        col = get_column_letter(lcp_col)
        rng = f"{col}{data_start}:{col}{last_row}"
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="lessThan",
                formula=["2.5"],
                fill=PatternFill("solid", fgColor="C6EFCE"),
            ),
        )
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="greaterThan",
                formula=["4.0"],
                fill=PatternFill("solid", fgColor="FFC7CE"),
            ),
        )

    answer_para_col = headers.get("Paragraphs 40-60 Words Count")
    if answer_para_col:
        col = get_column_letter(answer_para_col)
        rng = f"{col}{data_start}:{col}{last_row}"
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="equal",
                formula=["0"],
                fill=PatternFill("solid", fgColor="FFC7CE"),
            ),
        )
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="between",
                formula=["1", "2"],
                fill=PatternFill("solid", fgColor="FFEB9C"),
            ),
        )
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=["3"],
                fill=PatternFill("solid", fgColor="C6EFCE"),
            ),
        )

    action_col = headers.get("Action Needed")
    if action_col:
        col = get_column_letter(action_col)
        rng = f"{col}{data_start}:{col}{last_row}"
        worksheet.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'LOWER({col}2)="yes"'],
                stopIfTrue=True,
                fill=PatternFill("solid", fgColor=RAG_RED),
            ),
        )
        worksheet.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'LOWER({col}2)="no"'],
                stopIfTrue=True,
                fill=PatternFill("solid", fgColor=RAG_GREEN),
            ),
        )

    # Match the rich "Severity Badge" column and the plain "Severity" column used on
    # FixPlan / issue sheets, so severity reads consistently wherever it appears.
    severity_badge_col = headers.get("Severity Badge") or headers.get("Severity")
    if severity_badge_col:
        col = get_column_letter(severity_badge_col)
        rng = f"{col}{data_start}:{col}{last_row}"
        worksheet.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'LOWER({col}2)="critical"'],
                stopIfTrue=True,
                fill=PatternFill("solid", fgColor=RAG_RED),
            ),
        )
        worksheet.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'LOWER({col}2)="warning"'],
                stopIfTrue=True,
                fill=PatternFill("solid", fgColor=RAG_AMBER_SOFT),
            ),
        )
        worksheet.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'OR(LOWER({col}2)="pass",LOWER({col}2)="observation")'],
                stopIfTrue=True,
                fill=PatternFill("solid", fgColor=RAG_GREEN),
            ),
        )

    status_text_col = headers.get("Status")
    if status_text_col:
        apply_workflow_status_conditional_formatting(
            worksheet,
            status_text_col,
            first_row=data_start,
            last_row=last_row,
        )


# ---------------------------------------------------------------------------
# Sprint 6 — executive priority heatmap.
#
# Lives alongside ``apply_global_conditional_formatting`` because both
# share the openpyxl idiom (``worksheet.conditional_formatting.add(rng,
# Rule(...))``). Kept as a separate function rather than folded into
# the global helper because the Content Optimisation Hub uses
# ``header_row=2`` (banner row 1, headers row 2, data row 3) while
# every other audit sheet uses ``header_row=1`` — a single helper
# would have to branch on sheet identity, which we deliberately avoid
# here.
# ---------------------------------------------------------------------------


_CRITICAL_PRIORITY_FILL = PatternFill(
    start_color="C00000", end_color="C00000", fill_type="solid"
)
_CRITICAL_PRIORITY_FONT = Font(color="FFFFFF", bold=True)


def _header_index_at_row(worksheet: Worksheet, header_row: int) -> dict[str, int]:
    """Return ``{header_text: 1-indexed_column}`` for ``header_row``."""
    return {
        str(cell.value).strip(): cell.column
        for cell in worksheet[header_row]
        if cell.value is not None and str(cell.value).strip()
    }


def apply_executive_priority_formatting(
    worksheet: Worksheet,
    *,
    header_row: int = 2,
) -> None:
    """Apply Sprint 6 ROI-related conditional formatting.

    When ``Semantic AEO Score`` is present: **ColorScaleRule** (red →
    yellow → green) on that column (used on the Content Optimisation Hub).

    When ``Instant Priority`` is present: **FormulaRule** for whole-row
    red highlight when the value is ``CRITICAL`` (Hub no longer carries
    this column; primary use is ``Content Hub Metrics``).

    Missing headers skip the corresponding rule.
    """
    if DISABLE_CONDITIONAL_FORMATTING or worksheet.max_row <= header_row:
        return
    headers = _header_index_at_row(worksheet, header_row)
    last_row = worksheet.max_row
    first_data_row = header_row + 1

    aeo_col = headers.get("Semantic AEO Score")
    if aeo_col:
        aeo_letter = get_column_letter(aeo_col)
        worksheet.conditional_formatting.add(
            f"{aeo_letter}{first_data_row}:{aeo_letter}{last_row}",
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="F8696B",
                mid_type="num",
                mid_value=50,
                mid_color="FFEB84",
                end_type="num",
                end_value=100,
                end_color="63BE7B",
            ),
        )

    priority_col = headers.get("Instant Priority")
    if priority_col:
        priority_letter = get_column_letter(priority_col)
        last_col_letter = get_column_letter(worksheet.max_column)
        row_range = (
            f"A{first_data_row}:{last_col_letter}{last_row}"
        )
        # Whole-row formula must use a column-absolute / row-relative
        # reference so each row evaluates against ITS OWN priority cell.
        worksheet.conditional_formatting.add(
            row_range,
            FormulaRule(
                formula=[f'UPPER(${priority_letter}{first_data_row})="CRITICAL"'],
                stopIfTrue=False,
                fill=_CRITICAL_PRIORITY_FILL,
                font=_CRITICAL_PRIORITY_FONT,
            ),
        )
        # Belt-and-braces: also add a CellIsRule directly on the
        # priority column itself so the flag stays legible even if a
        # downstream stylesheet removes the row-wide rule.
        worksheet.conditional_formatting.add(
            f"{priority_letter}{first_data_row}:{priority_letter}{last_row}",
            CellIsRule(
                operator="equal",
                formula=['"CRITICAL"'],
                stopIfTrue=False,
                fill=_CRITICAL_PRIORITY_FILL,
                font=_CRITICAL_PRIORITY_FONT,
            ),
        )
