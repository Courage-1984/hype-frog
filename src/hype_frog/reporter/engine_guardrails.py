from __future__ import annotations

import re

from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from hype_frog.reporter.sheets.config import (
    CONTENT_HUB_METRICS_SHEET,
    CONTENT_OPTIMISATION_HUB_SHEET,
    CONTENT_PLANNER_SHEET,
    CRAWL_LOG_SHEET,
    EXECUTIVE_DASHBOARD_SHEET,
    RAG_AMBER,
    RAG_GREEN,
    RAG_RED,
    RAG_RED_FONT,
    REDIRECT_MAP_SHEET,
    ROBOTS_ANALYSIS_SHEET,
)
from hype_frog.reporter.sheets.view_state import set_freeze_panes_safe
from hype_frog.reporter.sheets.workbook_layout import apply_workbook_active_tab
from hype_frog.reporter.sheets.layout import (
    CONTENT_HUB_HEADER_COMMENT_OG_IMAGE_PREVIEW,
    CONTENT_HUB_HEADER_COMMENT_SEO_SCORE,
    CONTENT_HUB_HEADER_COMMENT_TECHNICAL_HEALTH,
)

_BANNED_TOC_FALLBACK = "Detailed URL diagnostic data"

# Manual TOC blurbs (authoritative for export guardrails and initial TOC seed).
_TOC_FRIENDLY_DESCRIPTIONS: dict[str, str] = {
    "Dashboard": (
        "Executive overview of site-wide SEO performance and critical alerts."
    ),
    "Executive Dashboard": (
        "Visual KPI cards and charts (health, issues, priority URLs, content readiness)."
    ),
    CONTENT_OPTIMISATION_HUB_SHEET: (
        "Diagnostic command center: live title, meta, and H1–H6 health plus on-page score."
    ),
    CONTENT_PLANNER_SHEET: (
        "Page-level workflow tracker: nav/footer site structure with "
        "copywriting, design, and client sign-off status columns."
    ),
    CONTENT_HUB_METRICS_SHEET: (
        "Per-URL crawl metrics (JS/rendered words, field CWV), anchor diversity, "
        "and executive ROI estimates aligned with the Content Hub URL set."
    ),
    "FixPlan": (
        "Prioritized list of technical and content fixes with estimated effort."
    ),
    "Quick Wins": (
        "Top 15 high-impact, low-effort fixes ranked by traffic and business risk."
    ),
    "Technical": (
        "Deep-dive diagnostic data for every crawled URL (Status, Load Time, etc.)."
    ),
    "AEO": (
        "Answer Engine Optimisation readiness scores and answer-block candidates."
    ),
    "Priority URLs": (
        "High-value pages requiring immediate attention based on business risk."
    ),
    "Main": "Primary URL inventory with titles, meta descriptions, and crawl signals.",
    "Summary": (
        "Aggregated issue counts, AEO opportunities, and top critical URLs from the run."
    ),
    "Content": "Content depth, readability, headings, and thin-content flags per URL.",
    "Links": "Internal and external link counts and anchor-text quality per URL.",
    "LinksDetail": "Row-level outbound internal links with crawl resolution status.",
    "Link Inventory": (
        "Every anchor-level outbound link with type, rel attributes, status codes, and generic-anchor flags."
    ),
    "Broken Link Impact": (
        "Broken destinations ranked by inbound link count and source-page GSC clicks."
    ),
    "Media": "Image inventory, alt coverage, filename quality, and mixed content flags.",
    "Schema & Metadata": "JSON-LD, microdata, Open Graph, and Twitter card signals.",
    "AIOSEO Recommendations": (
        "Plugin-aligned recommendations and edit links for AIOSEO users."
    ),
    "Security": "Transport and hardening headers (HSTS, CSP, XFO, referrer policy, etc.).",
    "PSI Performance": "Lab PageSpeed scores and mobile CWV-related proxies.",
    "Indexability": "Robots directives, canonicals, and indexability classification.",
    "Redirects": "Redirect chains, hop lists, HTTPS upgrades, and loop detection.",
    REDIRECT_MAP_SHEET: (
        "Per-source redirect chains with hop status codes, 302 flags, and SEO risk notes."
    ),
    ROBOTS_ANALYSIS_SHEET: (
        "Parsed robots.txt content, user-agent rules, blocked URLs, and sitemap conflicts."
    ),
    CRAWL_LOG_SHEET: (
        "Errors and warnings from fetch, render, PSI, and GSC phases for this audit run."
    ),
    "Link Equity Map": (
        "Internal PageRank distribution: inlink/outlink counts and equity flow per URL."
    ),
    "Anchor Text Audit": (
        "Anchor-text profile per destination: distribution, generic-anchor share, and diversity."
    ),
    "Snippet Opportunities": (
        "Pages with answer-block potential for featured snippets and AI answer extraction."
    ),
    "Script Inventory": (
        "Third-party and first-party scripts per URL with blocking and privacy implications."
    ),
    "Image Inventory": (
        "Every image with alt-text coverage, filename quality, dimensions, and mixed-content flags."
    ),
    "Duplicates": "Near-duplicate titles and meta descriptions across the crawl set.",
    "Pattern and Template Issues": (
        "Folder-level clusters and template-wide systemic issue patterns."
    ),
    "IssueInventory": (
        "Legacy flattened issue log (superseded by Issue Register for client workflows). "
        "Hidden from the Table of Contents; use Issue Register for backlog tracking."
    ),
    "SitemapQA": "Sitemap coverage, URL membership checks, and sitemap metadata QA.",
    "Quick Reference Guide": "In-workbook SEO and AEO copy standards and guardrails.",
    "Glossary & Legend": "Definitions for metrics, badges, and workbook conventions.",
    "Audit Run Details": "Run configuration, timestamps, and environment notes.",
    "Playbook": (
        "Client education reference: editorial standards plus per-issue playbooks "
        "with what/why/fix/verify guidance."
    ),
    "Competitor Benchmarks": (
        "Optional client-vs-competitor comparison on sampled pages when "
        "--competitors is configured."
    ),
    "DeltaFromPreviousRun": (
        "New, resolved, and persistent issues compared to a prior audit workbook or delta JSON. "
        "Sections: summary counts, new issues, resolved issues, metric changes, and SEO health trend. "
        "First runs show a baseline note when no previous export was supplied."
    ),
    "ResolvedIssues": (
        "Issues marked resolved when compared to the previous export. "
        "First runs show a baseline note when no previous export was supplied."
    ),
    "CrawlGraph": "Derived link graph metrics (click depth, inlinks, PageRank proxy).",
    "Issue Register": (
        "Tracked issue backlog with first-seen dates, days open, assignment fields, "
        "and client notes. Unified rollups from Summary plus per-URL rows from Issue Inventory."
    ),
    "Technical Diagnostics": (
        "Per-URL technical, indexability, redirect, security, PSI, and Search Console signals."
    ),
    "Content & AI Readiness": (
        "Per-URL content depth, schema, media, and answer-engine readiness metrics."
    ),
    "Link Intelligence": (
        "Per-URL internal link summary rows plus anchor-level detail for broken-link triage."
    ),
    "CMS Action URLs": (
        "WooCommerce and other CMS action-parameter URLs discovered during the crawl "
        "(for example add-to-cart) that were withheld from the crawl queue."
    ),
    "Template & Duplication Risks": (
        "Duplicate titles/meta, draft/copy pages, and folder-level template pattern risks."
    ),
}


def friendly_toc_description(sheet_name: str) -> str:
    """Return the fixed marketing-style blurb for TOC column C."""
    key = str(sheet_name or "").strip()
    if key in _TOC_FRIENDLY_DESCRIPTIONS:
        return _TOC_FRIENDLY_DESCRIPTIONS[key]
    return f"Diagnostic metrics for {key}."


_HEADER_TOOLTIP_MESSAGES: dict[str, str] = {
    "TTFB (ms)": "Time to First Byte: Measures server responsiveness.",
    "LCP (s)": "Largest Contentful Paint: Measures perceived load speed.",
    "Internal PageRank": ("Authority score based on internal linking structure."),
    "Click Depth": ("Number of clicks required to reach this URL from the homepage. "
                    "-1 means unreachable from the detected homepage within the crawl graph."),
    "AEO Readiness Score": (
        "Weighted 0–100 score: answer paragraphs under question H2/H3, FAQ/HowTo/Speakable "
        "schema, Flesch–Kincaid grade band 7–10, list/table structure, and robots.txt "
        "coverage for GPTBot/PerplexityBot/CCBot. 70+ targets high extraction probability."
    ),
    "Orphan Pages": (
        "Pages with zero internal incoming links; difficult for search engines to "
        "discover and value."
    ),
    "Reachable from Homepage": (
        "True when the URL has a link path from the detected homepage within the crawl "
        "graph. False when Click Depth is -1, even if the page has inbound links from "
        "other unreachable pages."
    ),
    "Canonical Type": (
        "Self: URL points to itself. Cross: URL points to another page. "
        "Missing: No canonical tag found."
    ),
    "Content Cluster ID": (
        "Groups pages by topical relevance (e.g. /about-us/) for bulk template editing."
    ),
    "Action Required": (
        "Formula-driven: shows Complete when On-Page Optimization Score reaches 85+."
    ),
    "Assigned Owner": (
        "Select from dropdown: Copy Writer (Content), Developer (Technical), or "
        "Server/Host (Infrastructure)."
    ),
    "Current SEO Score": ("Baseline score from the initial crawl. Does not change."),
    "Projected SEO Score": (
        "Live Score: Updates as you type. Reaches 100% when your Title, Description, "
        "and AEO drafts meet guidelines."
    ),
    "Elementor Builder Link": (
        "One-click access to the page editor. Requires active WP login."
    ),
    "Proposed Title (50-60 Chars)": (
        "Draft your optimised SEO title here. Keyword should be front-loaded."
    ),
    "Proposed Meta Desc (120-160 Chars)": (
        "Draft your optimised meta description here. Include a clear call-to-action."
    ),
    "AEO Answer Block Draft": (
        "A concise 40-60 word factual answer to the primary page question."
    ),
    "Status": (
        "Workflow state for this URL row. Use the list to move through To Do, "
        "In Progress, Review, and Completed."
    ),
    "URL": ("Canonical audited URL. Use links to open the live page or related tabs."),
    "Target Keywords": (
        "Enter one primary phrase plus optional supporting terms (comma- or pipe-separated). "
        "Use the same wording you want reflected in Title, Meta, and H1 so reviewers can "
        "spot intent drift. Keep phrases short; avoid stuffing. This column is editorial "
        "only and does not change crawl scores until you publish and re-audit."
    ),
    "Current Page Copy Snippet": (
        "Extracted body preview from the crawl. Reference when drafting changes."
    ),
    "Current Title": (
        "Title captured during the crawl. Compare with your proposed title."
    ),
    "Title Count": (
        "Character count of the proposed title (LEN). Target band 50-60 characters."
    ),
    "Current Meta Desc": (
        "Meta description from the crawl. Compare with your proposed description."
    ),
    "Desc Count": (
        "Character count of the proposed meta description (LEN). Target band 120-160."
    ),
    "Current H-Tag Structure": (
        "Heading outline from the crawl. Use when planning H-tag fixes."
    ),
    "Proposed H-Tag Fixes": (
        "Draft heading hierarchy or fixes before publishing in the CMS."
    ),
    "FAQ/QA Draft": ("Draft FAQ or Q&A pairs for structured answers and AEO coverage."),
    "Current OG-Image URL": (
        "Open Graph image URL detected on the page. Used for preview and sharing QA."
    ),
    "OG Image Health": (
        "Flags OG images that differ from the site-wide default or look like legacy social assets."
    ),
    "OG Image Preview": CONTENT_HUB_HEADER_COMMENT_OG_IMAGE_PREVIEW,
    "Social Share Note": ("Optional note for social snippets or share messaging."),
    "SEO Score": CONTENT_HUB_HEADER_COMMENT_SEO_SCORE,
    "Technical Health": CONTENT_HUB_HEADER_COMMENT_TECHNICAL_HEALTH,
    "Copy Score": (
        "Live copy readiness from proposed title and meta description length checks."
    ),
    "Open in Main": (
        "Jumps to the full diagnostic record for this URL on the Main tab."
    ),
}


def apply_header_tooltips(worksheet: Worksheet, *, header_row: int = 1) -> None:
    """Attach Excel cell comments to selected metric headers (UX guidance)."""
    lcp_body = _HEADER_TOOLTIP_MESSAGES["LCP (s)"]
    for col_idx in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=header_row, column=col_idx)
        header = str(cell.value or "").strip()
        if not header:
            continue
        tip = _HEADER_TOOLTIP_MESSAGES.get(header)
        if tip is None and "lcp" in header.lower() and "(s)" in header:
            tip = lcp_body
        if tip:
            cell.comment = Comment(tip, "hype-frog")


_ACTION_REQUIRED_FILL = PatternFill(
    start_color=RAG_RED,
    end_color=RAG_RED,
    fill_type="solid",
)
_ACTION_REQUIRED_FONT = Font(bold=True, color=RAG_RED_FONT)


def _action_header_map(ws: Worksheet, header_row: int = 1) -> dict[str, int]:
    out: dict[str, int] = {}
    for cell in ws[header_row]:
        val = cell.value
        if val is None:
            continue
        key = str(val).strip()
        if key:
            out[key] = cell.column
    return out


_ACTION_REQUIRED_ALLOWED: frozenset[str] = frozenset(
    {"Needs Copy", "Needs Optimisation", "Complete"}
)
_OPTIMISATION_FILL = PatternFill(
    start_color=RAG_AMBER, end_color=RAG_AMBER, fill_type="solid"
)
_COMPLETE_FILL = PatternFill(
    start_color=RAG_GREEN, end_color=RAG_GREEN, fill_type="solid"
)


def _normalize_action_required_cell_value(raw: object) -> str:
    """Map workbook cell content to strict Action Required literals (never None/empty)."""
    if raw is None:
        return "Complete"
    s = str(raw).strip()
    if not s:
        return "Complete"
    if s in _ACTION_REQUIRED_ALLOWED:
        return s
    lowered = s.lower()
    if lowered in {"ready to publish", "completed", "complete."}:
        return "Complete"
    if lowered in {"needs optimization", "needs optimisation"}:
        return "Needs Optimisation"
    # Legacy free-text or unknown labels still imply open work; default to copy track.
    return "Needs Copy"


def apply_action_required_guardrails(ws: Worksheet, *, header_row: int = 1) -> None:
    """Normalize **Action Required** to strict literals and apply fills (red only for ``Needs Copy``)."""
    if ws.title == CONTENT_OPTIMISATION_HUB_SHEET:
        return
    headers = _action_header_map(ws, header_row)
    col = headers.get("Action Required")
    if not col:
        return
    for r in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=r, column=col)
        literal = _normalize_action_required_cell_value(cell.value)
        cell.value = literal
        if literal == "Needs Copy":
            cell.font = _ACTION_REQUIRED_FONT
            cell.fill = _ACTION_REQUIRED_FILL
        elif literal == "Needs Optimisation":
            cell.font = Font(bold=True, color="000000")
            cell.fill = _OPTIMISATION_FILL
        else:
            cell.font = Font(bold=True, color="000000")
            cell.fill = _COMPLETE_FILL


# Production TOC column A stores ``=HYPERLINK("#'Sheet'!A1","Sheet")``; legacy/test rows
# may store the bare sheet name. Resolve the real sheet name from either form.
_TOC_HYPERLINK_TARGET_RE = re.compile(
    r"""HYPERLINK\(\s*"#'(?P<target>.+)'!A1"\s*,""", re.IGNORECASE
)
_TOC_HYPERLINK_DISPLAY_RE = re.compile(
    r'HYPERLINK\([^,]*,\s*"(?P<display>.*)"\s*\)\s*$', re.IGNORECASE
)


def _toc_sheet_name_from_cell_value(value: object) -> str:
    """Resolve the sheet name a TOC column-A cell points at.

    Returns ``""`` for empty rows and merged section-label rows (which are not
    HYPERLINK formulas and do not name a sheet).
    """
    raw = str(value or "").strip()
    if not raw:
        return ""
    if raw.startswith("="):
        target = _TOC_HYPERLINK_TARGET_RE.search(raw)
        if target:
            # Excel escapes a single quote in a sheet name as ``''`` inside the target.
            return target.group("target").replace("''", "'").strip()
        display = _TOC_HYPERLINK_DISPLAY_RE.search(raw)
        if display:
            return display.group("display").strip()
        return ""
    return raw


def refresh_toc_descriptions_dynamic(wb: Workbook) -> None:
    """Rewrite TOC column C using the canonical friendly description map."""
    if "Table of Contents" not in wb.sheetnames:
        return
    toc = wb["Table of Contents"]
    row = 3
    while row <= toc.max_row:
        name_cell = toc.cell(row=row, column=1)
        desc_cell = toc.cell(row=row, column=3)
        name = _toc_sheet_name_from_cell_value(name_cell.value)
        if not name or name not in wb.sheetnames:
            row += 1
            continue
        desc_cell.value = friendly_toc_description(name)
        row += 1


#: Sheets whose bespoke freeze contracts must survive the final C2 normalisation:
#: the Table of Contents (``A3``), the Content Optimisation Hub (banner-aware freeze),
#: the Content Planner (``E2`` — columns A–D locked for hierarchy + page link),
#: and the Executive Dashboard (``A8`` above the chart band).
FREEZE_C2_EXEMPT_SHEETS: frozenset[str] = frozenset(
    {
        "Table of Contents",
        CONTENT_OPTIMISATION_HUB_SHEET,
        CONTENT_PLANNER_SHEET,
        EXECUTIVE_DASHBOARD_SHEET,
    }
)


def apply_freeze_c2_data_sheets(
    wb: Workbook, *, skip_names: frozenset[str] | None = None
) -> None:
    """Normalise data-grid freezes to ``C2``; this is the **final** freeze authority.

    Runs last in :func:`apply_workbook_export_guardrails`, so it deliberately wins over
    earlier per-sheet freezes for ordinary data sheets (a single, predictable contract:
    keep the URL/key column and header visible). Sheets with bespoke layouts are exempt
    via :data:`FREEZE_C2_EXEMPT_SHEETS` and keep the freeze set during sheet assembly.
    """
    skip = skip_names or FREEZE_C2_EXEMPT_SHEETS
    for name in wb.sheetnames:
        if name in skip:
            continue
        set_freeze_panes_safe(wb[name], "C2")


def apply_workbook_export_guardrails(wb: Workbook) -> None:
    """Apply Action Required styling, dynamic TOC blurbs, C2 freezes, then landing tab."""
    for name in wb.sheetnames:
        if name == "Table of Contents":
            continue
        apply_action_required_guardrails(wb[name])
    refresh_toc_descriptions_dynamic(wb)
    apply_freeze_c2_data_sheets(wb)
    # Last write wins: open the workbook on the Dashboard (TOC stays left-most).
    apply_workbook_active_tab(wb)
