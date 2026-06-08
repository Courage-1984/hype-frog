from __future__ import annotations

import math
import re
from typing import Any
from urllib.parse import urlparse

from openpyxl.utils import get_column_letter

from hype_frog.checkpoint.cache import AuditCache
from hype_frog.core.models import ExtraRowPayload, MainRowPayload
from hype_frog.core.scoring import calculate_executive_roi
from hype_frog.pipeline.content_hub_metrics import resolve_content_hub_metrics
from hype_frog.pipeline.og_image_consistency import (
    build_og_image_site_profile,
    classify_og_image_consistency,
    resolve_og_image_url,
)
from hype_frog.reporter.engine_io import (
    _normalize_url_for_match,
    _safe_sheet_name,
    _sanitize_excel_url,
)
from hype_frog.rules import owner_for_issue, workflow_metrics_for_issue
from hype_frog.reporter.sheets.layout import (
    content_optimisation_hub_ordered_headers,
    main_sheet_url_column_letter,
)
from hype_frog.reporter.summary_builder import reference_tab_for_merged_workbook


def _hub_score_value(raw: object) -> float:
    """Normalize numeric SEO / Technical / Copy scores for Content Hub export."""
    if raw is None or raw == "":
        return 0.0
    try:
        x = float(raw)
    except (TypeError, ValueError):
        return 0.0
    if math.isnan(x) or math.isinf(x):
        return 0.0
    return round(x, 2)


def _round2(raw: object, default: float = 0.0) -> float:
    try:
        val = float(raw)
    except (TypeError, ValueError):
        return default
    if math.isnan(val) or math.isinf(val):
        return default
    return round(val, 2)


def _round4(raw: object, default: float = 0.0) -> float:
    try:
        val = float(raw)
    except (TypeError, ValueError):
        return default
    if math.isnan(val) or math.isinf(val):
        return default
    return round(val, 4)


_HUB_INVISIBLE_CHARS_RE = re.compile(r"[\u200b-\u200d\ufeff]")


def _hub_display_text(raw: object) -> str:
    """Strip zero-width / BOM characters from hub-facing copy (CMS-safe display)."""
    s = str(raw or "")
    s = _HUB_INVISIBLE_CHARS_RE.sub("", s)
    return s.strip()


def _heading_levels_from_h_tag_structure(structure: object) -> dict[int, list[str]]:
    """Parse ``Current H-Tag Structure`` lines like ``H2: text`` into level → texts."""
    out: dict[int, list[str]] = {1: [], 2: [], 3: [], 4: [], 5: [], 6: []}
    text = str(structure or "")
    for line in text.splitlines():
        line = line.strip()
        if ":" not in line:
            continue
        prefix, rest = line.split(":", 1)
        tag = prefix.strip().upper()
        if len(tag) == 2 and tag.startswith("H") and tag[1].isdigit():
            level = int(tag[1])
            heading_text = _hub_display_text(rest)
            if 1 <= level <= 6 and heading_text:
                out[level].append(heading_text)
    return out


def _first_non_empty(mapping: dict[str, Any], *keys: str) -> str:
    """Return the first non-empty string value among candidate mapping keys."""
    for key in keys:
        val = mapping.get(key)
        text = _hub_display_text(val)
        if text:
            return text
    return ""


def _join_pipe(values: list[str]) -> str:
    """Join heading values for display while preserving order and removing duplicates."""
    out: list[str] = []
    seen: set[str] = set()
    for raw in values:
        text = _hub_display_text(raw)
        if not text:
            continue
        key = text.casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(text)
    return " | ".join(out)


def _split_heading_pipe_segments(raw: object) -> list[str]:
    """Split Main ``H{n} Content`` pipe-joined values without treating ``|`` in copy as multiples."""
    text = _hub_display_text(raw)
    if not text:
        return []
    if "|" not in text:
        return [text]
    return [_hub_display_text(part) for part in text.split("|") if _hub_display_text(part)]


def _resolve_primary_h1_for_hub(
    main: dict[str, Any],
    extra: dict[str, Any],
    h_by_level: dict[int, list[str]],
) -> tuple[str, int]:
    """Return (primary H1 text, authoritative H1 count) for Content Hub export."""
    h1_count = _to_int(extra.get("H1 Count"), 0)
    primary = _hub_display_text(extra.get("Primary H1 Content"))
    if not primary:
        structure_h1 = h_by_level.get(1, [])
        if structure_h1:
            primary = structure_h1[0]
    if not primary:
        main_segments = _split_heading_pipe_segments(
            _first_non_empty(main, "H1 Content", "h1_content", "h1 content", "h1")
        )
        if main_segments:
            primary = main_segments[0]
            if h1_count <= 0:
                h1_count = len(main_segments)
    if primary and h1_count <= 0:
        h1_count = 1
    return primary, h1_count


def _to_int(value: object, default: int = 0) -> int:
    if value is None or str(value).strip() == "":
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _h1_health_label(h1_count: int, primary_h1: str) -> str:
    if not primary_h1:
        return "MISSING"
    if h1_count > 1:
        return "FIX: MULTIPLE"
    return "OK"


def _hyperlink_url_formula(display_url: str) -> str:
    """Excel HYPERLINK so the URL column is clickable (sanitized for formula quotes)."""
    s = str(display_url or "").strip()
    if not s:
        return ""
    safe = s.replace('"', '""')
    return f'=HYPERLINK("{safe}","{safe}")'


def _hyperlink_with_label_formula(target_url: str, label: str) -> str:
    """Excel HYPERLINK using supplied label as display text."""
    target = str(target_url or "").strip()
    display = str(label or "").strip()
    if not target:
        return display
    safe_target = target.replace('"', '""')
    safe_display = (display or target).replace('"', '""')
    return f'=HYPERLINK("{safe_target}","{safe_display}")'


def _og_image_hyperlink_formula(raw_url: object) -> str:
    """Excel HYPERLINK for OG image URLs with compact display label."""
    target = str(_sanitize_excel_url(raw_url) or "").strip()
    if not target:
        return ""
    safe_target = target.replace('"', '""')
    return f'=HYPERLINK("{safe_target}","View Image")'


def _proposed_slug_value(seed_text: str) -> str:
    """Generate a clean slug suggestion from normalized keyword guidance."""
    text = str(seed_text or "").strip().lower()
    if not text:
        return ""
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def _proposed_url_slug_for_hub(url: str, seed_text: str) -> str:
    """Slug suggestion from keywords, else from the last URL path segment."""
    slug = _proposed_slug_value(seed_text)
    if slug:
        return slug
    parts = [p for p in urlparse(str(url or "").strip()).path.strip("/").split("/") if p]
    if parts:
        return _proposed_slug_value(parts[-1].replace("-", " ").replace("_", " "))
    return ""


def _slug_normalization_link_label(url: str, keyword_phrase: str) -> str:
    """Human-readable HYPERLINK label for URL Slug Normalization (never duplicate full URL on /)."""
    cleaned = _hub_display_text(keyword_phrase)
    if cleaned:
        return cleaned
    parsed = urlparse(str(url or "").strip())
    segments = [s for s in parsed.path.strip("/").split("/") if s]
    if segments:
        return segments[-1].replace("-", " ").replace("_", " ").title()
    return "/"


def _fallback_keyword(url: str, h1_text: str) -> str:
    slug_parts = [p for p in urlparse(url).path.strip("/").split("/") if p]
    if slug_parts:
        slug = slug_parts[-1].replace("-", " ").replace("_", " ").strip()
        slug = re.sub(r"\s+", " ", slug)
        if slug:
            return slug.title()
    return ""


def build_fixplan_rows(
    summary_rules: list[tuple[str, str, Any]],
    extra_rows: list[ExtraRowPayload],
    aeo_issue_names: set[str],
    root_cause_resolver: Any,
    default_effort_by_severity: dict[str, str],
    default_owner_by_severity: dict[str, str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    status_by_severity = {
        "Critical": "To Do",
        "Warning": "To Do",
        "Observation": "In Review",
    }
    for severity, issue_name, _ in summary_rules:
        affected = [
            r
            for r in extra_rows
            if issue_name in str(r.values.get("Matched Issues") or "").split(" | ")
        ]
        if issue_name == "Broken Internal Links":
            affected_count = sum(
                int(r.values.get("Broken Internal Links Count") or 0) for r in affected
            )
        else:
            affected_count = len(affected)
        if affected_count <= 0:
            continue
        root_cause, recommended_fix = root_cause_resolver(issue_name)
        effort = default_effort_by_severity.get(severity, "S")
        workflow = workflow_metrics_for_issue(severity, effort)
        legacy_reference = (
            "Indexability"
            if ("Canonical" in issue_name or "Noindex" in issue_name)
            else (
                "Links"
                if ("Links" in issue_name or "Anchor" in issue_name)
                else (
                    "AEO"
                    if (
                        "AEO" in issue_name
                        or "Question" in issue_name
                        or "FAQ" in issue_name
                        or "Answer" in issue_name
                    )
                    else "Technical"
                )
            )
        )
        reference_tab = reference_tab_for_merged_workbook(legacy_reference)
        affected_urls = [
            str(r.values.get("URL") or "") for r in affected if r.values.get("URL")
        ]
        systemic_issue_tokens = (
            "CWV",
            "Schema",
            "Robots",
            "Compression",
            "Cache-Control",
            "ETag",
            "Canonical",
            "Redirect",
        )
        resolution_type = (
            "Global Template"
            if any(
                token.lower() in issue_name.lower() for token in systemic_issue_tokens
            )
            or len(affected_urls) > 10
            else "Manual Content"
        )
        rows.append(
            {
                "Category": "AEO" if issue_name in aeo_issue_names else "SEO",
                "Issue Type": issue_name,
                "Severity": severity,
                "Affected Count": affected_count,
                "Likely Root Cause": root_cause,
                "Recommended Fix": recommended_fix,
                "Owner": owner_for_issue(issue_name, severity),
                "URL": affected[0].values.get("URL") if affected else "",
                "Affected URLs": (
                    f"SEE DETAILS IN {reference_tab}"
                    if len(affected_urls) > 10
                    else "\n".join(affected_urls[:50])
                ),
                "Detail Reference Tab": reference_tab,
                "Resolution Type": resolution_type,
                "Effort": effort,
                "Action Needed": "Yes" if severity in {"Critical", "Warning"} else "No",
                "Sprint": "",
                "Status": status_by_severity.get(severity, "To Do"),
                "Verified By": "",
                "Date Resolved": "",
                "Revenue Risk": (
                    "High Risk"
                    if severity == "Critical" and workflow["Priority Score"] >= 100
                    else "Medium Risk" if severity == "Warning" else "Monitor"
                ),
                "Agency Owner": owner_for_issue(issue_name, severity),
                "Jump to Details": "Open in Main Tab",
                "Est. Sprint Points": workflow["Est. Sprint Points"],
                "Est. Hours": workflow["Est. Hours"],
                "Priority Score": workflow["Priority Score"],
                "Aging/Priority": workflow["Aging/Priority"],
            }
        )
    return rows


def write_snippet_candidates_chunked(
    writer: Any,
    cache: AuditCache,
    sheet_name: str = "SnippetCandidates",
    chunk_size: int = 500,
) -> None:
    columns = [
        "URL",
        "Heading (Question)",
        "Snippet (Answer)",
        "Word Count",
        "Snippet Preview Mockup",
    ]
    ws = writer.book.create_sheet(title=_safe_sheet_name(sheet_name))
    writer.sheets[sheet_name] = ws
    ws.append(columns)
    has_rows = False
    for chunk in cache.iter_results_chunked(chunk_size):
        for result in chunk:
            extra = result.get("extra", {})
            url = extra.get("URL")
            for snippet in extra.get("aeo_snippets", []) or []:
                ws.append(
                    [
                        url,
                        snippet.get("heading"),
                        snippet.get("snippet"),
                        snippet.get("word_count"),
                        f"{snippet.get('heading') or ''}\n{snippet.get('snippet') or ''}".strip(),
                    ]
                )
                has_rows = True
    if not has_rows:
        ws.append(["", "", "", 0, ""])


# Initial write order before ``reorder_columns``; kept so
# ``content_optimisation_hub_ordered_headers`` can derive the physical grid.
_CONTENT_HUB_FIELDS_PRE_REORDER: tuple[str, ...] = (
    "Action Required",
    "On-Page Optimization Score",
    "SEO Score",
    "Technical Health",
    "Copy Score",
    "Entity Density (%)",
    "Top Entities",
    "Citation Candidate Count",
    "Semantic AEO Score",
    # Sprint 5 diagnostics (JS/rendered words/field CWV/anchor diversity)
    # and Sprint 6 ROI columns live on ``Content Hub Metrics`` (same URL
    # set as the Hub) so the Hub stays an editorial command center.
    "Status",
    "Assigned Owner",
    "URL Slug Normalization",
    "URL",
    "Proposed URL Slug",
    "Current Title",
    "Title Health",
    "Current Meta Desc",
    "Meta Health",
    "H1",
    "H1 Health",
    "H2",
    "H2 Health",
    "H3",
    "H3 Health",
    "H4",
    "H4 Health",
    "H5",
    "H5 Health",
    "H6",
    "H6 Health",
    "Elementor Builder Link",
    "Current OG-Image URL",
    "OG Image Health",
    "OG Image Preview",
    "Open in Main",
)

CONTENT_HUB_EXPORT_COLUMNS: tuple[str, ...] = content_optimisation_hub_ordered_headers(
    _CONTENT_HUB_FIELDS_PRE_REORDER
)

# Per-URL metrics companion to the Hub (row 1 headers; no banner row).
CONTENT_HUB_METRICS_EXPORT_COLUMNS: tuple[str, ...] = (
    "URL",
    "JS Dependent",
    "Raw Words",
    "Rendered Words",
    "Field LCP (ms)",
    "Field CLS",
    "Anchor Text Diversity",
    "Potential Traffic Lift",
    "AEO Visibility Gain",
    "Instant Priority",
    "Search Intent",
)


def content_hub_column_letter(header: str) -> str:
    """Excel column letter for a Content Optimisation Hub header (post-reorder row 2)."""
    pos = CONTENT_HUB_EXPORT_COLUMNS.index(header) + 1
    return get_column_letter(pos)


def _content_hub_on_page_score_from_health_formula(row: int) -> str:
    """Weighted 0–100 score derived from live Title/Meta/H1–H6 *Health* formula columns."""
    th = content_hub_column_letter("Title Health")
    mh = content_hub_column_letter("Meta Health")
    h1 = content_hub_column_letter("H1 Health")
    h2 = content_hub_column_letter("H2 Health")
    h3 = content_hub_column_letter("H3 Health")
    h4 = content_hub_column_letter("H4 Health")
    h5 = content_hub_column_letter("H5 Health")
    h6 = content_hub_column_letter("H6 Health")
    r = row
    t_pts = (
        f'IF({th}{r}="",0,IF(LEFT({th}{r},7)="MISSING",0,'
        f'IF(ISNUMBER(SEARCH("OK",{th}{r})),100,IF(ISNUMBER(SEARCH("SHORT",{th}{r})),65,'
        f'IF(ISNUMBER(SEARCH("LONG",{th}{r})),50,70)))))'
    )
    m_pts = (
        f'IF({mh}{r}="",0,IF(LEFT({mh}{r},7)="MISSING",0,'
        f'IF(ISNUMBER(SEARCH("OK",{mh}{r})),100,IF(ISNUMBER(SEARCH("SHORT",{mh}{r})),65,'
        f'IF(ISNUMBER(SEARCH("LONG",{mh}{r})),50,70)))))'
    )
    h1_pts = (
        f'IF({h1}{r}="",0,IF(LEFT({h1}{r},2)="OK",100,0))'
    )

    def _h_pts(col: str) -> str:
        return f'IF({col}{r}="",0,IF(LEFT({col}{r},2)="OK",100,0))'

    h2p, h3p, h4p, h5p, h6p = (
        _h_pts(h2),
        _h_pts(h3),
        _h_pts(h4),
        _h_pts(h5),
        _h_pts(h6),
    )
    return (
        f"=MIN(100,ROUND(0.30*({t_pts})+0.25*({m_pts})+0.20*({h1_pts})+0.08*({h2p})+0.06*({h3p})"
        f"+0.05*({h4p})+0.03*({h5p})+0.03*({h6p}),0))"
    )


def _content_hub_open_in_main_formula(url_col_letter: str, row: int) -> str:
    """Cross-sheet jump to Main (URL column) with Technical Diagnostics fallback."""
    ml = main_sheet_url_column_letter()
    td = "Technical Diagnostics"
    u = url_col_letter
    r = row
    return (
        f'=IFERROR(HYPERLINK("#\'Main\'!{ml}"&MATCH(TRIM({u}{r}),'
        f'\'Main\'!{ml}:{ml},0),"Open in Main"),'
        f'IFERROR(HYPERLINK("#\'{td}\'!A"&MATCH(TRIM({u}{r}),'
        f'\'{td}\'!A:A,0),"Open in Technical"),"Not Found"))'
    )


def build_content_optimisation_hub_rows(
    main_rows: list[MainRowPayload],
    extra_rows: list[ExtraRowPayload],
    fixplan_rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    main_by_url = {
        _normalize_url_for_match(r.values.get("URL")): r
        for r in main_rows
        if r.values.get("URL")
    }
    extra_by_url = {
        _normalize_url_for_match(r.values.get("URL")): r
        for r in extra_rows
        if r.values.get("URL")
    }
    manual_content_urls = {
        _normalize_url_for_match(r.get("URL"))
        for r in fixplan_rows
        if str(r.get("Resolution Type") or "").strip().lower() == "manual content"
        and _normalize_url_for_match(r.get("URL"))
    }
    content_issue_tokens = (
        "title",
        "meta",
        "h1",
        "content",
        "question",
        "answer",
        "faq",
        "alt",
        "regional authority",
    )
    for e in extra_rows:
        row_values = e.values
        url = _normalize_url_for_match(row_values.get("URL"))
        issues = str(row_values.get("Matched Issues") or "").lower()
        if url and any(tok in issues for tok in content_issue_tokens):
            manual_content_urls.add(url)

    if not manual_content_urls:
        scored_urls: list[tuple[float, str]] = []
        for e in extra_rows:
            row_values = e.values
            raw_url = _normalize_url_for_match(row_values.get("URL"))
            if not raw_url:
                continue
            try:
                score = float(row_values.get("SEO Health Score") or 0.0)
            except Exception:
                score = 0.0
            scored_urls.append((round(score, 2), raw_url))
        scored_urls.sort(key=lambda item: item[0])
        for _score, url in scored_urls[:15]:
            manual_content_urls.add(url)

    g_l = content_hub_column_letter("Current Title")
    i_l = content_hub_column_letter("Current Meta Desc")
    k_l = content_hub_column_letter("H1")
    m_l = content_hub_column_letter("H2")
    o_l = content_hub_column_letter("H3")
    q_l = content_hub_column_letter("H4")
    s_l = content_hub_column_letter("H5")
    u_l = content_hub_column_letter("H6")
    w_l = content_hub_column_letter("On-Page Optimization Score")
    f_l = content_hub_column_letter("URL")

    og_profile = build_og_image_site_profile(
        [r.values for r in main_rows],
        [r.values for r in extra_rows],
    )
    rows: list[dict[str, Any]] = []
    metrics_rows: list[dict[str, Any]] = []
    for excel_row, url in enumerate(sorted(manual_content_urls), start=3):
        main_payload = main_by_url.get(url)
        extra_payload = extra_by_url.get(url)
        m = main_payload.values if main_payload else {}
        e = extra_payload.values if extra_payload else {}
        hub_metrics = resolve_content_hub_metrics(m, e)
        field_lcp_raw = hub_metrics.field_lcp_ms
        clicks_raw = e.get("GSC Clicks") or m.get("GSC Clicks")
        roi = calculate_executive_roi(
            clicks=clicks_raw,
            aeo_score=e.get("Semantic AEO Score"),
            lcp_ms=field_lcp_raw,
        )
        raw_url = ""
        if main_payload:
            raw_url = str(main_payload.values.get("URL") or "").strip()
        if not raw_url and extra_payload:
            raw_url = str(extra_payload.values.get("URL") or "").strip()
        if not raw_url:
            raw_url = str(url)
        post_id = e.get("WordPress Post ID")
        try:
            post_id = int(post_id) if post_id is not None else None
        except Exception:
            post_id = None
        elementor_link = None
        if post_id:
            parsed = urlparse(raw_url)
            if parsed.scheme and parsed.netloc:
                elementor_link = (
                    f"{parsed.scheme}://{parsed.netloc}/wp-admin/post.php?post={post_id}&action=elementor"
                )
        elementor_cell = ""
        if elementor_link:
            safe_url = str(elementor_link).replace('"', '""')
            elementor_cell = f'=HYPERLINK("{safe_url}","Open in Elementor")'
        target_keywords = _hub_display_text(
            str(e.get("Meta Keywords") or m.get("Meta Keywords") or "")
        )
        if not target_keywords:
            target_keywords = _hub_display_text(
                _fallback_keyword(
                    url,
                    str(e.get("Current H-Tag Structure") or m.get("H1 Content") or ""),
                )
            )

        h_by_level = _heading_levels_from_h_tag_structure(e.get("Current H-Tag Structure"))

        def _h_values(n: int) -> list[str]:
            items: list[str] = []
            from_main = _first_non_empty(
                m,
                f"H{n} Content",
                f"h{n}_content",
                f"h{n} content",
                f"h{n}",
            )
            if from_main:
                items.append(from_main)
            from_extra = _first_non_empty(
                e,
                f"H{n} Content",
                f"h{n}_content",
                f"h{n} content",
                f"h{n}",
            )
            if from_extra:
                items.append(from_extra)
            items.extend(h_by_level.get(n, []))
            return items

        def _h_joined(n: int) -> str:
            return _join_pipe(_h_values(n))

        primary_h1, h1_count = _resolve_primary_h1_for_hub(m, e, h_by_level)
        h1_health_label = _h1_health_label(h1_count, primary_h1)
        h2_values = _h_values(2)
        h3_values = _h_values(3)
        h4_values = _h_values(4)
        h5_values = _h_values(5)
        h6_values = _h_values(6)

        title_health = (
            f'=IF({g_l}{excel_row}="","MISSING",'
            f'IF(AND(LEN({g_l}{excel_row})>=50,LEN({g_l}{excel_row})<=60),"OK 50-60 chars",'
            f'IF(LEN({g_l}{excel_row})<50,"SHORT: "&LEN({g_l}{excel_row}),"LONG: "&LEN({g_l}{excel_row}))))'
        )
        meta_health = (
            f'=IF({i_l}{excel_row}="","MISSING",'
            f'IF(AND(LEN({i_l}{excel_row})>=120,LEN({i_l}{excel_row})<=160),"OK 120-160 chars",'
            f'IF(LEN({i_l}{excel_row})<120,"SHORT: "&LEN({i_l}{excel_row}),"LONG: "&LEN({i_l}{excel_row}))))'
        )
        # H1 health uses crawl-time ``H1 Count`` (static) — avoids false positives when
        # a single H1 legitimately contains a pipe character or Main uses ``|`` joining.
        h1_health = h1_health_label
        h2_health = f'=IF(ISBLANK({m_l}{excel_row}),"MISSING","OK")'
        h3_health = f'=IF(ISBLANK({o_l}{excel_row}),"MISSING","OK")'
        h4_health = f'=IF(ISBLANK({q_l}{excel_row}),"MISSING","OK")'
        h5_health = f'=IF(ISBLANK({s_l}{excel_row}),"MISSING","OK")'
        h6_health = f'=IF(ISBLANK({u_l}{excel_row}),"MISSING","OK")'
        on_page_score = _content_hub_on_page_score_from_health_formula(excel_row)
        action_formula = f'=IF({w_l}{excel_row}>=85,"Complete","Needs Copy")'
        open_main_formula = _content_hub_open_in_main_formula(f_l, excel_row)

        metrics_rows.append(
            {
                "URL": _hyperlink_url_formula(raw_url),
                "JS Dependent": hub_metrics.js_dependent,
                "Raw Words": hub_metrics.raw_words,
                "Rendered Words": hub_metrics.rendered_words,
                "Field LCP (ms)": _round2(
                    hub_metrics.field_lcp_ms,
                    default=0.0,
                ),
                "Field CLS": _round4(hub_metrics.field_cls, default=0.0),
                "Anchor Text Diversity": _hub_display_text(
                    str(e.get("Anchor Text Diversity") or "")
                ),
                "Potential Traffic Lift": int(round(roi["potential_traffic_lift"])),
                "AEO Visibility Gain": _round2(roi["aeo_visibility_gain"]),
                "Instant Priority": str(roi["instant_priority"]),
                "Search Intent": _hub_display_text(str(e.get("Search Intent") or "Unknown"))
                or "Unknown",
            }
        )
        rows.append(
            {
                "SEO Score": _hub_score_value(e.get("SEO Score")),
                "Technical Health": _hub_score_value(e.get("Technical Health")),
                "Copy Score": _hub_score_value(e.get("Copy Score")),
                "Entity Density (%)": _round2(e.get("Entity Density (%)")),
                "Top Entities": _hub_display_text(str(e.get("Top Entities") or "")),
                # Display-safe string: the shared number formatter treats any
                # header containing "date" as date-like, and "Candidate"
                # includes that substring.
                "Citation Candidate Count": str(
                    int(float(e.get("Citation Candidate Count") or 0))
                ),
                "Semantic AEO Score": _round2(e.get("Semantic AEO Score")),
                "Action Required": action_formula,
                "Status": "To Do",
                "Assigned Owner": "Copy Writer",
                "URL": _hyperlink_url_formula(raw_url),
                "URL Slug Normalization": _hyperlink_with_label_formula(
                    raw_url,
                    _slug_normalization_link_label(raw_url, target_keywords),
                ),
                "Proposed URL Slug": _proposed_url_slug_for_hub(raw_url, target_keywords),
                "Current Title": _hub_display_text(str(m.get("Title") or ""))
                or "MISSING TITLE",
                "Title Health": title_health,
                "Current Meta Desc": _hub_display_text(str(m.get("Meta Description") or ""))
                or "MISSING DESCRIPTION",
                "Meta Health": meta_health,
                "H1": primary_h1,
                "H1 Health": h1_health,
                "H2": _h_joined(2),
                "H2 Health": h2_health,
                "H3": _h_joined(3),
                "H3 Health": h3_health,
                "H4": _h_joined(4),
                "H4 Health": h4_health,
                "H5": _h_joined(5),
                "H5 Health": h5_health,
                "H6": _h_joined(6),
                "H6 Health": h6_health,
                "On-Page Optimization Score": on_page_score,
                "Elementor Builder Link": elementor_cell,
                # Extra uses ``OG Image``; Main uses ``OG-Image`` — prefer extra, fall back to Main.
                "Current OG-Image URL": _og_image_hyperlink_formula(
                    resolve_og_image_url(m, e)
                ),
                "OG Image Health": classify_og_image_consistency(
                    resolve_og_image_url(m, e),
                    og_profile,
                ),
                "OG Image Preview": "",
                "Open in Main": open_main_formula,
            }
        )
    return rows, metrics_rows


def build_content_optimization_hub_rows(
    main_rows: list[MainRowPayload],
    extra_rows: list[ExtraRowPayload],
    fixplan_rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Backward-compatible alias (US spelling) for :func:`build_content_optimisation_hub_rows`."""
    return build_content_optimisation_hub_rows(main_rows, extra_rows, fixplan_rows)
