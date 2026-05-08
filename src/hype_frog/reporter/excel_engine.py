"""
Monolithic Excel / openpyxl engine: workbook helpers, tab builders, conditional
formatting helpers, and strict export guardrails (Action Required, TOC, freeze).
"""

from __future__ import annotations

import re
from typing import Any
from urllib.parse import quote, unquote, urlparse, urlsplit, urlunsplit

import pandas as pd
from openpyxl.comments import Comment
from openpyxl.formatting.rule import (
    CellIsRule,
    ColorScaleRule,
    DataBarRule,
    FormulaRule,
)
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from hype_frog.checkpoint.cache import AuditCache
from hype_frog.reporter.sheets.config import CONTENT_OPTIMISATION_HUB_SHEET
from hype_frog.rules import owner_for_issue, workflow_metrics_for_issue
from hype_frog.utils import normalize_url_key

# ---------------------------------------------------------------------------
# Legacy formatting (formerly formatting.py)
# ---------------------------------------------------------------------------


def apply_fixplan_workflow_formatting(worksheet: Worksheet) -> None:
    headers = [cell.value for cell in worksheet[1]]
    header_to_col = {str(h): i + 1 for i, h in enumerate(headers) if h is not None}
    priority_col = header_to_col.get("Priority Score")
    points_col = header_to_col.get("Est. Sprint Points")
    aging_col = header_to_col.get("Aging/Priority")
    critical_fill = PatternFill(
        start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"
    )
    warning_fill = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
    )
    good_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    edge_fill = PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid")
    for row_idx in range(2, worksheet.max_row + 1):
        if priority_col:
            cell = worksheet.cell(row=row_idx, column=priority_col)
            try:
                score = int(cell.value or 0)
                if score >= 100:
                    cell.fill = critical_fill
                elif score >= 65:
                    cell.fill = warning_fill
                else:
                    cell.fill = edge_fill
            except Exception:
                pass
        if points_col:
            cell = worksheet.cell(row=row_idx, column=points_col)
            try:
                points = int(cell.value or 0)
                if points >= 8:
                    cell.fill = critical_fill
                elif points >= 5:
                    cell.fill = warning_fill
                else:
                    cell.fill = good_fill
            except Exception:
                pass
        if aging_col:
            cell = worksheet.cell(row=row_idx, column=aging_col)
            value = str(cell.value or "").lower()
            if "immediate" in value:
                cell.fill = critical_fill
            elif "next sprint" in value:
                cell.fill = warning_fill
            elif "backlog" in value:
                cell.fill = edge_fill


def _legacy_sheet_header_index(worksheet: Worksheet) -> dict[str, int]:
    return {
        str(cell.value): idx
        for idx, cell in enumerate(worksheet[1], start=1)
        if cell.value
    }


def ensure_auto_filter(worksheet: Worksheet) -> None:
    if worksheet.title not in {"Main", "Dashboard"} and (
        worksheet.max_row < 10 or worksheet.max_column < 5
    ):
        worksheet.auto_filter.ref = None
        return

    header_row = 2 if worksheet.title == CONTENT_OPTIMISATION_HUB_SHEET else 1
    if worksheet.max_row >= header_row + 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = f"A{header_row}:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
    else:
        worksheet.auto_filter.ref = None


def _clear_orphaned_selection(worksheet: Worksheet) -> None:
    try:
        worksheet.views.sheetView[0].selection = []
    except Exception:
        pass


def ensure_freeze_header(worksheet: Worksheet) -> None:
    if worksheet.title == CONTENT_OPTIMISATION_HUB_SHEET:
        return
    if worksheet.title not in {"Main", "Dashboard"} and (
        worksheet.max_row < 10 or worksheet.max_column < 5
    ):
        worksheet.freeze_panes = None
        _clear_orphaned_selection(worksheet)
        return
    if worksheet.max_row > 1 and worksheet.max_column >= 1:
        worksheet.freeze_panes = "A2"
    else:
        worksheet.freeze_panes = None
        _clear_orphaned_selection(worksheet)


def apply_global_conditional_formatting(worksheet: Worksheet) -> None:
    if worksheet.max_row <= 1:
        return
    headers = _legacy_sheet_header_index(worksheet)
    last_row = worksheet.max_row
    status_col = headers.get("Status Code") or headers.get("Target Status (if crawled)")
    if status_col:
        col = get_column_letter(status_col)
        rng = f"{col}2:{col}{last_row}"
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="equal",
                formula=["200"],
                fill=PatternFill("solid", fgColor="C6EFCE"),
            ),
        )
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="between",
                formula=["300", "399"],
                fill=PatternFill("solid", fgColor="FFEB9C"),
            ),
        )
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=["400"],
                fill=PatternFill("solid", fgColor="FFC7CE"),
            ),
        )

    for load_header in ("Load Time (s)", "Load Time", "TTFB (ms)"):
        load_col = headers.get(load_header)
        if load_col:
            col = get_column_letter(load_col)
            rng = f"{col}2:{col}{last_row}"
            worksheet.conditional_formatting.add(
                rng,
                ColorScaleRule(
                    start_type="min",
                    start_color="63BE7B",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFEB84",
                    end_type="max",
                    end_color="F8696B",
                ),
            )
            worksheet.conditional_formatting.add(
                rng,
                DataBarRule(
                    start_type="min",
                    end_type="max",
                    color="638EC6",
                    showValue=True,
                ),
            )
            break

    for wc_header in ("Word Count", "Word Count (Body)"):
        wc_col = headers.get(wc_header)
        if wc_col:
            col = get_column_letter(wc_col)
            rng = f"{col}2:{col}{last_row}"
            worksheet.conditional_formatting.add(
                rng,
                ColorScaleRule(
                    start_type="min",
                    start_color="F8696B",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFEB84",
                    end_type="max",
                    end_color="63BE7B",
                ),
            )
            worksheet.conditional_formatting.add(
                rng,
                DataBarRule(
                    start_type="min",
                    end_type="max",
                    color="63BE7B",
                    showValue=True,
                ),
            )
            break

    priority_col = headers.get("Priority Score")
    if priority_col:
        col = get_column_letter(priority_col)
        rng = f"{col}2:{col}{last_row}"
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=["85"],
                font=Font(color="9C0006", bold=True),
                fill=PatternFill("solid", fgColor="FFC7CE"),
            ),
        )
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="between",
                formula=["65", "84"],
                font=Font(color="9C5700", bold=True),
                fill=PatternFill("solid", fgColor="FFEB9C"),
            ),
        )

    seo_score_col = headers.get("SEO Health Score")
    if seo_score_col:
        col = get_column_letter(seo_score_col)
        worksheet.conditional_formatting.add(
            f"{col}2:{col}{last_row}",
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )

    aeo_score_col = headers.get("AEO Readiness Score")
    if aeo_score_col:
        col = get_column_letter(aeo_score_col)
        worksheet.conditional_formatting.add(
            f"{col}2:{col}{last_row}",
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )

    for score_header in ("Desktop Score", "Mobile Score"):
        score_col = headers.get(score_header)
        if score_col:
            col = get_column_letter(score_col)
            rng = f"{col}2:{col}{last_row}"
            worksheet.conditional_formatting.add(
                rng,
                CellIsRule(
                    operator="between",
                    formula=["90", "100"],
                    fill=PatternFill("solid", fgColor="C6EFCE"),
                ),
            )
            worksheet.conditional_formatting.add(
                rng,
                CellIsRule(
                    operator="between",
                    formula=["50", "89"],
                    fill=PatternFill("solid", fgColor="FFEB9C"),
                ),
            )
            worksheet.conditional_formatting.add(
                rng,
                CellIsRule(
                    operator="between",
                    formula=["0", "49"],
                    fill=PatternFill("solid", fgColor="FFC7CE"),
                ),
            )

    lcp_col = headers.get("Mobile LCP")
    if lcp_col:
        col = get_column_letter(lcp_col)
        rng = f"{col}2:{col}{last_row}"
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="lessThan",
                formula=["2.5"],
                fill=PatternFill("solid", fgColor="C6EFCE"),
            ),
        )
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="greaterThan",
                formula=["4.0"],
                fill=PatternFill("solid", fgColor="FFC7CE"),
            ),
        )

    answer_para_col = headers.get("Paragraphs 40-60 Words Count")
    if answer_para_col:
        col = get_column_letter(answer_para_col)
        rng = f"{col}2:{col}{last_row}"
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="equal",
                formula=["0"],
                fill=PatternFill("solid", fgColor="FFC7CE"),
            ),
        )
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="between",
                formula=["1", "2"],
                fill=PatternFill("solid", fgColor="FFEB9C"),
            ),
        )
        worksheet.conditional_formatting.add(
            rng,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=["3"],
                fill=PatternFill("solid", fgColor="C6EFCE"),
            ),
        )

    action_col = headers.get("Action Needed")
    if action_col:
        col = get_column_letter(action_col)
        rng = f"{col}2:{col}{last_row}"
        worksheet.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'LOWER({col}2)="yes"'],
                stopIfTrue=True,
                fill=PatternFill("solid", fgColor="FFC7CE"),
            ),
        )
        worksheet.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'LOWER({col}2)="no"'],
                stopIfTrue=True,
                fill=PatternFill("solid", fgColor="C6EFCE"),
            ),
        )

    severity_badge_col = headers.get("Severity Badge")
    if severity_badge_col:
        col = get_column_letter(severity_badge_col)
        rng = f"{col}2:{col}{last_row}"
        worksheet.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'LOWER({col}2)="critical"'],
                stopIfTrue=True,
                fill=PatternFill("solid", fgColor="FFC7CE"),
            ),
        )
        worksheet.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'LOWER({col}2)="warning"'],
                stopIfTrue=True,
                fill=PatternFill("solid", fgColor="FFCC99"),
            ),
        )
        worksheet.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'OR(LOWER({col}2)="pass",LOWER({col}2)="observation")'],
                stopIfTrue=True,
                fill=PatternFill("solid", fgColor="C6EFCE"),
            ),
        )

    status_text_col = headers.get("Status")
    if status_text_col:
        col = get_column_letter(status_text_col)
        rng = f"{col}2:{col}{last_row}"
        worksheet.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f'LOWER({col}2)="done"'],
                stopIfTrue=True,
                fill=PatternFill("solid", fgColor="D9EAD3"),
            ),
        )
        worksheet.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[
                    f'OR(LOWER({col}2)="to do",LOWER({col}2)="in progress",LOWER({col}2)="in review")'
                ],
                stopIfTrue=True,
                fill=PatternFill("solid", fgColor="FFF2CC"),
            ),
        )


# ---------------------------------------------------------------------------
# Tab / hub row builders (formerly tabs.py)
# ---------------------------------------------------------------------------

_ILLEGAL_XLSX_CHARS_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
_INVALID_SHEET_CHARS_RE = re.compile(r"[:\\/*?\[\]]")


def _safe_sheet_name(name: str) -> str:
    cleaned = _INVALID_SHEET_CHARS_RE.sub("_", str(name or "Sheet"))
    cleaned = cleaned[:31]
    return cleaned or "Sheet"


def _sanitize_excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return _ILLEGAL_XLSX_CHARS_RE.sub("", value)
    return value


def load_cached_rows(
    cache: AuditCache,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    main_rows: list[dict[str, Any]] = []
    extra_rows: list[dict[str, Any]] = []
    for result in cache.iter_results():
        main_rows.append(result["main"])
        extra_rows.append(result["extra"])
    return main_rows, extra_rows


def build_core_dataframes(
    cache: AuditCache,
) -> tuple[pd.DataFrame, pd.DataFrame, list[dict[str, Any]], list[dict[str, Any]]]:
    main_rows, extra_rows = load_cached_rows(cache)
    return pd.DataFrame(main_rows), pd.DataFrame(extra_rows), main_rows, extra_rows


def write_dict_rows_sheet(
    writer: Any, sheet_name: str, columns: list[str], rows: list[dict[str, Any]]
) -> None:
    ws = writer.book.create_sheet(title=_safe_sheet_name(sheet_name))
    writer.sheets[sheet_name] = ws
    ws.append(columns)
    for row in rows:
        ws.append([_sanitize_excel_value(row.get(col)) for col in columns])


def _sanitize_excel_url(url_value: Any) -> str:
    raw = str(url_value or "").strip()
    if not raw:
        return ""
    raw = "".join(ch for ch in raw if ord(ch) >= 32).replace('"', "").replace("'", "")
    if not raw.startswith(("http://", "https://")):
        return raw
    try:
        parts = urlsplit(raw)
        cleaned_path = quote(unquote(parts.path), safe="/:@-._~!$&()*+,;=")
        cleaned_query = quote(unquote(parts.query), safe="=&:@-._~!$()*+,;/?")
        cleaned_fragment = quote(unquote(parts.fragment), safe=":@-._~!$&()*+,;=/?")
        return urlunsplit(
            (parts.scheme, parts.netloc, cleaned_path, cleaned_query, cleaned_fragment)
        )
    except Exception:
        return raw


def _normalize_url_for_match(url_value: Any) -> str:
    return normalize_url_key(_sanitize_excel_url(url_value))


def _fallback_keyword(url: str, h1_text: str) -> str:
    slug_parts = [p for p in urlparse(url).path.strip("/").split("/") if p]
    if slug_parts:
        slug = slug_parts[-1].replace("-", " ").replace("_", " ").strip()
        slug = re.sub(r"\s+", " ", slug)
        if slug:
            return slug.title()
    return ""


def write_cached_sheet_chunked(
    writer: Any,
    cache: AuditCache,
    sheet_name: str,
    columns: list[str],
    payload_key: str,
    chunk_size: int = 500,
) -> None:
    ws = writer.book.create_sheet(title=_safe_sheet_name(sheet_name))
    writer.sheets[sheet_name] = ws
    ws.append(columns)
    for chunk in cache.iter_results_chunked(chunk_size):
        for result in chunk:
            payload = result.get(payload_key, {})
            ws.append([_sanitize_excel_value(payload.get(col)) for col in columns])


def build_fixplan_rows(
    summary_rules: list[tuple[str, str, Any]],
    extra_rows: list[dict[str, Any]],
    aeo_issue_names: set[str],
    root_cause_resolver: Any,
    default_effort_by_severity: dict[str, str],
    default_owner_by_severity: dict[str, str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    status_by_severity = {
        "Critical": "To Do",
        "Warning": "To Do",
        "Observation": "In Review",
    }
    for severity, issue_name, _ in summary_rules:
        affected = [
            r
            for r in extra_rows
            if issue_name in str(r.get("Matched Issues") or "").split(" | ")
        ]
        root_cause, recommended_fix = root_cause_resolver(issue_name)
        effort = default_effort_by_severity.get(severity, "S")
        workflow = workflow_metrics_for_issue(severity, effort)
        reference_tab = (
            "Indexability"
            if ("Canonical" in issue_name or "Noindex" in issue_name)
            else (
                "Links"
                if ("Links" in issue_name or "Anchor" in issue_name)
                else (
                    "AEO"
                    if (
                        "AEO" in issue_name
                        or "Question" in issue_name
                        or "FAQ" in issue_name
                        or "Answer" in issue_name
                    )
                    else "Technical"
                )
            )
        )
        affected_urls = [str(r.get("URL") or "") for r in affected if r.get("URL")]
        systemic_issue_tokens = (
            "CWV",
            "Schema",
            "Robots",
            "Compression",
            "Cache-Control",
            "ETag",
            "Canonical",
            "Redirect",
        )
        resolution_type = (
            "Global Template"
            if any(
                token.lower() in issue_name.lower() for token in systemic_issue_tokens
            )
            or len(affected_urls) > 10
            else "Manual Content"
        )
        rows.append(
            {
                "Category": "AEO" if issue_name in aeo_issue_names else "SEO",
                "Issue Type": issue_name,
                "Severity": severity,
                "Affected Count": len(affected),
                "Likely Root Cause": root_cause,
                "Recommended Fix": recommended_fix,
                "Owner": owner_for_issue(issue_name, severity),
                "URL": affected[0].get("URL") if affected else "",
                "Affected URLs": (
                    f"SEE DETAILS IN {reference_tab}"
                    if len(affected_urls) > 10
                    else "\n".join(affected_urls[:50])
                ),
                "Detail Reference Tab": reference_tab,
                "Resolution Type": resolution_type,
                "Effort": effort,
                "Action Needed": "Yes" if severity in {"Critical", "Warning"} else "No",
                "Sprint": "",
                "Status": status_by_severity.get(severity, "To Do"),
                "Verified By": "",
                "Date Resolved": "",
                "Revenue Risk": (
                    "High Risk"
                    if severity == "Critical" and workflow["Priority Score"] >= 100
                    else "Medium Risk" if severity == "Warning" else "Monitor"
                ),
                "Agency Owner": owner_for_issue(issue_name, severity),
                "Jump to Details": "Open in Main Tab",
                "Est. Sprint Points": workflow["Est. Sprint Points"],
                "Est. Hours": workflow["Est. Hours"],
                "Priority Score": workflow["Priority Score"],
                "Aging/Priority": workflow["Aging/Priority"],
            }
        )
    return rows


def write_snippet_candidates_chunked(
    writer: Any,
    cache: AuditCache,
    sheet_name: str = "SnippetCandidates",
    chunk_size: int = 500,
) -> None:
    columns = [
        "URL",
        "Heading (Question)",
        "Snippet (Answer)",
        "Word Count",
        "Snippet Preview Mockup",
    ]
    ws = writer.book.create_sheet(title=_safe_sheet_name(sheet_name))
    writer.sheets[sheet_name] = ws
    ws.append(columns)
    has_rows = False
    for chunk in cache.iter_results_chunked(chunk_size):
        for result in chunk:
            extra = result.get("extra", {})
            url = extra.get("URL")
            for snippet in extra.get("aeo_snippets", []) or []:
                ws.append(
                    [
                        url,
                        snippet.get("heading"),
                        snippet.get("snippet"),
                        snippet.get("word_count"),
                        f"{snippet.get('heading') or ''}\n{snippet.get('snippet') or ''}".strip(),
                    ]
                )
                has_rows = True
    if not has_rows:
        ws.append(["", "", "", 0, ""])


def build_content_optimisation_hub_rows(
    main_rows: list[dict[str, Any]],
    extra_rows: list[dict[str, Any]],
    fixplan_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    main_by_url = {
        _normalize_url_for_match(r.get("URL")): r for r in main_rows if r.get("URL")
    }
    extra_by_url = {
        _normalize_url_for_match(r.get("URL")): r for r in extra_rows if r.get("URL")
    }
    manual_content_urls = {
        _normalize_url_for_match(r.get("URL"))
        for r in fixplan_rows
        if str(r.get("Resolution Type") or "").strip().lower() == "manual content"
        and _normalize_url_for_match(r.get("URL"))
    }
    content_issue_tokens = (
        "title",
        "meta",
        "h1",
        "content",
        "question",
        "answer",
        "faq",
        "alt",
        "regional authority",
    )
    for e in extra_rows:
        url = _normalize_url_for_match(e.get("URL"))
        issues = str(e.get("Matched Issues") or "").lower()
        if url and any(tok in issues for tok in content_issue_tokens):
            manual_content_urls.add(url)

    if not manual_content_urls:
        scored_urls: list[tuple[float, str]] = []
        for e in extra_rows:
            raw_url = _normalize_url_for_match(e.get("URL"))
            if not raw_url:
                continue
            try:
                score = float(e.get("SEO Health Score") or 0.0)
            except Exception:
                score = 0.0
            scored_urls.append((score, raw_url))
        scored_urls.sort(key=lambda item: item[0])
        for _score, url in scored_urls[:15]:
            manual_content_urls.add(url)

    pt_l = _content_hub_column_letter("Proposed Title (50-60 Chars)")
    pmd_l = _content_hub_column_letter("Proposed Meta Desc (120-160 Chars)")
    th_l = _content_hub_column_letter("Technical Health")
    copy_l = _content_hub_column_letter("Copy Score")

    rows: list[dict[str, Any]] = []
    # Row 1 = instruction banner, row 2 = headers (after apply_content_hub_conditional_rules); data begins row 3.
    for excel_row, url in enumerate(sorted(manual_content_urls), start=3):
        m = main_by_url.get(url, {})
        e = extra_by_url.get(url, {})
        score = float(e.get("SEO Health Score") or 0)
        post_id = e.get("WordPress Post ID")
        try:
            post_id = int(post_id) if post_id is not None else None
        except Exception:
            post_id = None
        elementor_link = None
        if post_id:
            parsed = urlparse(url)
            if parsed.scheme and parsed.netloc:
                elementor_link = f"{parsed.scheme}://{parsed.netloc}/wp-admin/post.php?post={post_id}&action=elementor"
        elementor_cell = ""
        if elementor_link:
            safe_url = str(elementor_link).replace('"', '""')
            elementor_cell = f'=HYPERLINK("{safe_url}","Open in Elementor")'
        target_keywords = str(
            e.get("Meta Keywords") or m.get("Meta Keywords") or ""
        ).strip()
        if not target_keywords:
            target_keywords = _fallback_keyword(
                url, str(e.get("Current H-Tag Structure") or m.get("H1 Content") or "")
            )
        copy_formula = (
            f"=IF(AND(LEN({pt_l}{excel_row})>=50,LEN({pt_l}{excel_row})<=60),50,20)"
            f"+IF(AND(LEN({pmd_l}{excel_row})>=120,LEN({pmd_l}{excel_row})<=160),50,20)"
        )
        projected_formula = f"=(0.7*{th_l}{excel_row})+(0.3*{copy_l}{excel_row})"
        action_formula = (
            f'=IF(AND(F{excel_row}>=90,Y{excel_row}>=90),"Complete","Needs Copy")'
        )
        open_main_formula = (
            f'=IFERROR(HYPERLINK("#\'Main\'!A"&MATCH(D{excel_row},\'Main\'!A:A,0),'
            f'"Open"),"Not Found")'
        )
        rows.append(
            {
                "Action Required": action_formula,
                "Status": "To Do",
                "Assigned Owner": "Copy Writer",
                "URL": url,
                "Current SEO Score": score,
                "Projected SEO Score": projected_formula,
                "Elementor Builder Link": elementor_cell,
                "Target Keywords": target_keywords,
                "Current Page Copy Snippet": str(
                    e.get("Current Page Copy Snippet") or ""
                ).strip(),
                "Current Title": str(m.get("Title") or "").strip() or "MISSING TITLE",
                "Proposed Title (50-60 Chars)": "",
                "Title Count": f"=LEN(K{excel_row})",
                "Current Meta Desc": str(m.get("Meta Description") or "").strip()
                or "MISSING DESCRIPTION",
                "Proposed Meta Desc (120-160 Chars)": "",
                "Desc Count": f"=LEN(N{excel_row})",
                "Current H-Tag Structure": str(
                    e.get("Current H-Tag Structure") or m.get("H1 Content") or ""
                ).strip(),
                "Proposed H-Tag Fixes": "",
                "AEO Answer Block Draft": "",
                "FAQ/QA Draft": "",
                "Current OG-Image URL": _sanitize_excel_url(e.get("OG Image")),
                "OG Image Preview": "",
                "Social Share Note": "",
                "Copy Score": copy_formula,
                "Open in Main": open_main_formula,
            }
        )
    return rows


def build_content_optimization_hub_rows(
    main_rows: list[dict[str, Any]],
    extra_rows: list[dict[str, Any]],
    fixplan_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Backward-compatible alias (US spelling) for :func:`build_content_optimisation_hub_rows`."""
    return build_content_optimisation_hub_rows(main_rows, extra_rows, fixplan_rows)


# ---------------------------------------------------------------------------
# Strict export guardrails (Action Required, TOC, freeze)
# ---------------------------------------------------------------------------

_BANNED_TOC_FALLBACK = "Detailed URL diagnostic data"

# Manual TOC blurbs (authoritative for export guardrails and initial TOC seed).
_TOC_FRIENDLY_DESCRIPTIONS: dict[str, str] = {
    "Dashboard": (
        "Executive overview of site-wide SEO performance and critical alerts."
    ),
    CONTENT_OPTIMISATION_HUB_SHEET: (
        "Interactive workspace for drafting SEO titles, descriptions, and AEO content."
    ),
    "FixPlan": (
        "Prioritized list of technical and content fixes with estimated effort."
    ),
    "Technical": (
        "Deep-dive diagnostic data for every crawled URL (Status, Load Time, etc.)."
    ),
    "AEO": (
        "Answer Engine Optimisation readiness scores and answer-block candidates."
    ),
    "Priority URLs": (
        "High-value pages requiring immediate attention based on business risk."
    ),
    "Main": "Primary URL inventory with titles, meta descriptions, and crawl signals.",
    "Summary": (
        "Aggregated issue counts, AEO opportunities, and top critical URLs from the run."
    ),
    "Content": "Content depth, readability, headings, and thin-content flags per URL.",
    "Links": "Internal and external link counts and anchor-text quality per URL.",
    "LinksDetail": "Row-level outbound internal links with crawl resolution status.",
    "Media": "Image inventory, alt coverage, filename quality, and mixed content flags.",
    "Schema & Metadata": "JSON-LD, microdata, Open Graph, and Twitter card signals.",
    "AIOSEO": "Plugin-aligned recommendations and edit links for AIOSEO users.",
    "Security": "Transport and hardening headers (HSTS, CSP, XFO, referrer policy, etc.).",
    "PSI Performance": "Lab PageSpeed scores and mobile CWV-related proxies.",
    "Indexability": "Robots directives, canonicals, and indexability classification.",
    "Redirects": "Redirect chains, hop lists, HTTPS upgrades, and loop detection.",
    "Duplicates": "Near-duplicate titles and meta descriptions across the crawl set.",
    "Pattern and Template Issues": (
        "Folder-level clusters and template-wide systemic issue patterns."
    ),
    "IssueInventory": "Flattened issue log with severity, owner seed, and stable IDs.",
    "SitemapQA": "Sitemap coverage, URL membership checks, and sitemap metadata QA.",
    "Quick Reference Guide": "In-workbook SEO and AEO copy standards and guardrails.",
    "Glossary & Legend": "Definitions for metrics, badges, and workbook conventions.",
    "RunMetadata": "Run configuration, timestamps, and environment notes.",
    "DeltaFromPreviousRun": (
        "New, fixed, and persistent issues compared to a prior audit workbook."
    ),
    "ResolvedIssues": "Issues marked resolved when compared to the previous export.",
    "CrawlGraph": "Derived link graph metrics (click depth, inlinks, PageRank proxy).",
}


def friendly_toc_description(sheet_name: str) -> str:
    """Return the fixed marketing-style blurb for TOC column C."""
    key = str(sheet_name or "").strip()
    if key in _TOC_FRIENDLY_DESCRIPTIONS:
        return _TOC_FRIENDLY_DESCRIPTIONS[key]
    return f"Diagnostic metrics for {key}."


_HEADER_TOOLTIP_MESSAGES: dict[str, str] = {
    "TTFB (ms)": "Time to First Byte: Measures server responsiveness.",
    "LCP (s)": "Largest Contentful Paint: Measures perceived load speed.",
    "Internal PageRank": ("Authority score based on internal linking structure."),
    "Click Depth": ("Number of clicks required to reach this URL from the homepage."),
    "AEO Readiness Score": (
        "Proprietary score (0-100) measuring how likely an AI engine is to extract "
        "a direct answer from this URL."
    ),
    "Orphan Pages": (
        "Pages with zero internal incoming links; difficult for search engines to "
        "discover and value."
    ),
    "Canonical Type": (
        "Self: URL points to itself. Cross: URL points to another page. "
        "Missing: No canonical tag found."
    ),
    "Content Cluster ID": (
        "Groups pages by topical relevance (e.g. /about-us/) for bulk template editing."
    ),
    "Action Required": (
        "Automation: Flips to 'Complete' once projected SEO and copy scores reach "
        "90 or above."
    ),
    "Assigned Owner": (
        "Select from dropdown: Copy Writer (Content), Developer (Technical), or "
        "Server/Host (Infrastructure)."
    ),
    "Current SEO Score": ("Baseline score from the initial crawl. Does not change."),
    "Projected SEO Score": (
        "Live Score: Updates as you type. Reaches 100% when your Title, Description, "
        "and AEO drafts meet guidelines."
    ),
    "Elementor Builder Link": (
        "One-click access to the page editor. Requires active WP login."
    ),
    "Proposed Title (50-60 Chars)": (
        "Draft your optimised SEO title here. Keyword should be front-loaded."
    ),
    "Proposed Meta Desc (120-160 Chars)": (
        "Draft your optimised meta description here. Include a clear call-to-action."
    ),
    "AEO Answer Block Draft": (
        "A concise 40-60 word factual answer to the primary page question."
    ),
    "Status": (
        "Workflow state for this URL row. Use the list to move through To Do, "
        "In Progress, Review, and Completed."
    ),
    "URL": ("Canonical audited URL. Use links to open the live page or related tabs."),
    "Target Keywords": (
        "Primary and supporting phrases to align title, meta, and body copy with intent."
    ),
    "Current Page Copy Snippet": (
        "Extracted body preview from the crawl. Reference when drafting changes."
    ),
    "Current Title": (
        "Title captured during the crawl. Compare with your proposed title."
    ),
    "Title Count": (
        "Character count of the proposed title (LEN). Target band 50-60 characters."
    ),
    "Current Meta Desc": (
        "Meta description from the crawl. Compare with your proposed description."
    ),
    "Desc Count": (
        "Character count of the proposed meta description (LEN). Target band 120-160."
    ),
    "Current H-Tag Structure": (
        "Heading outline from the crawl. Use when planning H-tag fixes."
    ),
    "Proposed H-Tag Fixes": (
        "Draft heading hierarchy or fixes before publishing in the CMS."
    ),
    "FAQ/QA Draft": ("Draft FAQ or Q&A pairs for structured answers and AEO coverage."),
    "Current OG-Image URL": (
        "Open Graph image URL detected on the page. Used for preview and sharing QA."
    ),
    "OG Image Preview": (
        "In-workbook image preview when external content is enabled in Excel."
    ),
    "Social Share Note": ("Optional note for social snippets or share messaging."),
    "SEO Score": (
        "Composite SEO score from the crawl-era model. Shown for reference next to "
        "live projected SEO."
    ),
    "Technical Health": (
        "Technical-health subscore from the crawl. Used as a static input to "
        "projected SEO."
    ),
    "Copy Score": (
        "Live copy readiness from proposed title and meta description length checks."
    ),
    "Open in Main": (
        "Jumps to the full diagnostic record for this URL on the Main tab."
    ),
}


_CONTENT_HUB_COLUMN_ORDER: tuple[str, ...] = (
    "Action Required",
    "Status",
    "Assigned Owner",
    "URL",
    "Current SEO Score",
    "Projected SEO Score",
    "Elementor Builder Link",
    "Target Keywords",
    "Current Page Copy Snippet",
    "Current Title",
    "Proposed Title (50-60 Chars)",
    "Title Count",
    "Current Meta Desc",
    "Proposed Meta Desc (120-160 Chars)",
    "Desc Count",
    "Current H-Tag Structure",
    "Proposed H-Tag Fixes",
    "AEO Answer Block Draft",
    "FAQ/QA Draft",
    "Current OG-Image URL",
    "OG Image Preview",
    "Social Share Note",
    "SEO Score",
    "Technical Health",
    "Copy Score",
    "Open in Main",
)


def _content_hub_column_letter(header: str) -> str:
    """Stable A1 letter for Hub columns matching ``_CONTENT_HUB_COLUMN_ORDER``."""
    pos = _CONTENT_HUB_COLUMN_ORDER.index(header) + 1
    return get_column_letter(pos)


def apply_header_tooltips(worksheet: Worksheet, *, header_row: int = 1) -> None:
    """Attach Excel cell comments to selected metric headers (UX guidance)."""
    lcp_body = _HEADER_TOOLTIP_MESSAGES["LCP (s)"]
    for col_idx in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=header_row, column=col_idx)
        header = str(cell.value or "").strip()
        if not header:
            continue
        tip = _HEADER_TOOLTIP_MESSAGES.get(header)
        if tip is None and "lcp" in header.lower() and "(s)" in header:
            tip = lcp_body
        if tip:
            cell.comment = Comment(tip, "hype-frog")


_ACTION_REQUIRED_FILL = PatternFill(
    start_color="FF0000",
    end_color="FF0000",
    fill_type="solid",
)
_ACTION_REQUIRED_FONT = Font(bold=True, color="FFFFFF")


def _action_header_map(ws: Worksheet, header_row: int = 1) -> dict[str, int]:
    out: dict[str, int] = {}
    for cell in ws[header_row]:
        val = cell.value
        if val is None:
            continue
        key = str(val).strip()
        if key:
            out[key] = cell.column
    return out


_ACTION_REQUIRED_ALLOWED: frozenset[str] = frozenset(
    {"Needs Copy", "Needs Optimisation", "Complete"}
)
_OPTIMISATION_FILL = PatternFill(
    start_color="FFC000", end_color="FFC000", fill_type="solid"
)
_COMPLETE_FILL = PatternFill(
    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
)


def _describe_sheet_from_headers(
    ws: Worksheet, *, max_labels: int = 10, max_chars: int = 220
) -> str:
    """TOC description is driven by sheet name, not inferred column headers."""
    del max_labels, max_chars  # retained for backward-compatible call sites/tests
    return friendly_toc_description(ws.title)


def _normalize_action_required_cell_value(raw: object) -> str:
    """Map workbook cell content to strict Action Required literals (never None/empty)."""
    if raw is None:
        return "Complete"
    s = str(raw).strip()
    if not s:
        return "Complete"
    if s in _ACTION_REQUIRED_ALLOWED:
        return s
    lowered = s.lower()
    if lowered in {"ready to publish", "completed", "complete."}:
        return "Complete"
    if lowered in {"needs optimization", "needs optimisation"}:
        return "Needs Optimisation"
    # Legacy free-text or unknown labels still imply open work; default to copy track.
    return "Needs Copy"


def apply_action_required_guardrails(ws: Worksheet, *, header_row: int = 1) -> None:
    """Normalize **Action Required** to strict literals and apply fills (red only for ``Needs Copy``)."""
    if ws.title == CONTENT_OPTIMISATION_HUB_SHEET:
        return
    headers = _action_header_map(ws, header_row)
    col = headers.get("Action Required")
    if not col:
        return
    for r in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=r, column=col)
        literal = _normalize_action_required_cell_value(cell.value)
        cell.value = literal
        if literal == "Needs Copy":
            cell.font = _ACTION_REQUIRED_FONT
            cell.fill = _ACTION_REQUIRED_FILL
        elif literal == "Needs Optimisation":
            cell.font = Font(bold=True, color="000000")
            cell.fill = _OPTIMISATION_FILL
        else:
            cell.font = Font(bold=True, color="000000")
            cell.fill = _COMPLETE_FILL


def refresh_toc_descriptions_dynamic(wb: Workbook) -> None:
    """Rewrite TOC column C using the canonical friendly description map."""
    if "Table of Contents" not in wb.sheetnames:
        return
    toc = wb["Table of Contents"]
    row = 3
    while row <= toc.max_row:
        name_cell = toc.cell(row=row, column=1)
        desc_cell = toc.cell(row=row, column=3)
        sheet_name = name_cell.value
        if not sheet_name:
            row += 1
            continue
        name = str(sheet_name).strip()
        if name not in wb.sheetnames:
            row += 1
            continue
        desc_cell.value = friendly_toc_description(name)
        cur = str(desc_cell.value or "")
        if _BANNED_TOC_FALLBACK.lower() in cur.lower():
            desc_cell.value = friendly_toc_description(name)
        row += 1


def apply_freeze_c2_data_sheets(
    wb: Workbook, *, skip_names: frozenset[str] | None = None
) -> None:
    """``freeze_panes = 'C2'`` on every sheet except skips (default: TOC and Content Hub)."""
    skip = skip_names or frozenset(
        {"Table of Contents", CONTENT_OPTIMISATION_HUB_SHEET}
    )
    for name in wb.sheetnames:
        if name in skip:
            continue
        wb[name].freeze_panes = "C2"


def apply_workbook_export_guardrails(wb: Workbook) -> None:
    """Apply Action Required styling, dynamic TOC blurbs, then C2 freezes."""
    for name in wb.sheetnames:
        if name == "Table of Contents":
            continue
        apply_action_required_guardrails(wb[name])
    refresh_toc_descriptions_dynamic(wb)
    apply_freeze_c2_data_sheets(wb)


# ---------------------------------------------------------------------------
# Facades into sheets implementation (avoid circular import at module load)
# ---------------------------------------------------------------------------


def adjust_sheet_format(writer: Any, sheet_name: str) -> Any:
    from hype_frog.reporter.sheets.tables_impl import adjust_sheet_format as _impl

    return _impl(writer, sheet_name)


def apply_tab_hyperlinks(writer: Any) -> Any:
    from hype_frog.reporter.sheets.tables_impl import apply_tab_hyperlinks as _impl

    return _impl(writer)


__all__ = [
    "adjust_sheet_format",
    "apply_tab_hyperlinks",
    "apply_fixplan_workflow_formatting",
    "ensure_auto_filter",
    "ensure_freeze_header",
    "apply_global_conditional_formatting",
    "load_cached_rows",
    "build_core_dataframes",
    "write_dict_rows_sheet",
    "write_cached_sheet_chunked",
    "build_fixplan_rows",
    "write_snippet_candidates_chunked",
    "build_content_optimisation_hub_rows",
    "build_content_optimization_hub_rows",
    "apply_action_required_guardrails",
    "apply_freeze_c2_data_sheets",
    "apply_workbook_export_guardrails",
    "refresh_toc_descriptions_dynamic",
    "friendly_toc_description",
    "apply_header_tooltips",
]
