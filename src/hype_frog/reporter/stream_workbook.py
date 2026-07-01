"""openpyxl write_only workbook adapter for memory-flat export."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


class StreamingExcelWriter:
    """Thin wrapper around ``Workbook(write_only=True)`` for export orchestration."""

    def __init__(self, path: str | Path) -> None:
        self.path = str(path)
        self.book: Workbook = Workbook(write_only=True)
        self.sheets: dict[str, Any] = {}
        self._sheet_row_counts: dict[str, int] = {}
        self._closed = False

    @property
    def write_only(self) -> bool:
        return True

    def sheet_row_count(self, sheet_name: str) -> int:
        return self._sheet_row_counts.get(sheet_name, 0)

    def record_rows_appended(self, sheet_name: str, count: int) -> None:
        self._sheet_row_counts[sheet_name] = self.sheet_row_count(sheet_name) + count

    def close(self) -> None:
        if self._closed:
            return
        self.book.save(self.path)
        self._closed = True

    def save(self) -> None:
        self.close()


class FormattingWorkbookWriter:
    """Read-write workbook reopened after the streaming write pass."""

    def __init__(self, path: str | Path) -> None:
        self.path = str(path)
        self.book = load_workbook(self.path)
        self.sheets = {name: self.book[name] for name in self.book.sheetnames}

    def close(self) -> None:
        self.book.save(self.path)

    def save(self) -> None:
        self.close()


def is_write_only_writer(writer: Any) -> bool:
    book = getattr(writer, "book", None)
    return bool(getattr(book, "write_only", False))


def reopen_workbook_for_formatting(path: str | Path) -> FormattingWorkbookWriter:
    """Load a saved workbook for post-write formatting and guardrails."""
    return FormattingWorkbookWriter(path)


__all__ = [
    "FormattingWorkbookWriter",
    "StreamingExcelWriter",
    "is_write_only_writer",
    "reopen_workbook_for_formatting",
]
