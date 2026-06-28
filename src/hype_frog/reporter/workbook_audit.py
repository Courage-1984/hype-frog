"""Post-export workbook integrity checks (TOC, tab order, formula literals)."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from hype_frog.reporter.sheets.config import CONTENT_OPTIMISATION_HUB_SHEET
from hype_frog.reporter.sheets.toc import PREFERRED_WORKBOOK_TAB_ORDER
from hype_frog.reporter.sheets.workbook_layout import SHEETS_EXCLUDED_FROM_TOC

_ACTION_LITERALS = frozenset(
    {
        "Needs Copy",
        "Needs Optimisation",
        "Complete",
        "Ready to Publish",
    }
)
_FORMULA_ERROR_RE = re.compile(
    r"#(REF!|DIV/0!|NAME\?|NULL!|VALUE!|NUM!|N/A)", re.I
)
_PREFERRED_TAB_SET = frozenset(PREFERRED_WORKBOOK_TAB_ORDER)

REQUIRED_FULL_SUITE_SHEETS: frozenset[str] = frozenset(
    {
        "Table of Contents",
        "Dashboard",
        CONTENT_OPTIMISATION_HUB_SHEET,
        "Content Hub Metrics",
        "Main",
        "FixPlan",
        "SitemapQA",
    }
)


def _toc_listed_sheets(toc_ws) -> list[str]:
    names: list[str] = []
    row = 3
    while row <= toc_ws.max_row:
        cell = toc_ws.cell(row=row, column=1)
        val = cell.value
        if val is None or str(val).strip() == "":
            break
        text = str(val)
        if text.startswith("=HYPERLINK"):
            match = re.search(
                r'HYPERLINK\("#\'([^\']+(?:\'\'[^\']*)*)\'!A1","([^"]+)"\)',
                text,
            )
            if match:
                names.append(match.group(2).replace("''", "'"))
        elif (
            not text.strip().startswith("—")
            and "Primary workflow" not in text
            and "Technical & Historical" not in text
        ):
            # Skip TOC section headers; only hyperlink rows are sheet entries.
            pass
        row += 1
    return names


def audit_workbook(
    path: Path | str,
    *,
    require_full_suite_sheets: bool = True,
) -> list[str]:
    """Return a list of human-readable audit failures (empty if OK)."""
    workbook_path = Path(path)
    errors: list[str] = []
    wb = load_workbook(workbook_path, data_only=False)

    if require_full_suite_sheets:
        missing_core = REQUIRED_FULL_SUITE_SHEETS - set(wb.sheetnames)
        if missing_core:
            errors.append(f"Missing required sheets: {sorted(missing_core)}")

    if wb.sheetnames[0] != "Table of Contents":
        errors.append(f"TOC not at index 0; first tab is {wb.sheetnames[0]!r}")

    if "Table of Contents" not in wb.sheetnames:
        errors.append("Table of Contents sheet missing")
        wb.close()
        return errors

    toc = wb["Table of Contents"]
    if toc.freeze_panes != "A3":
        errors.append(f"TOC freeze_panes expected A3, got {toc.freeze_panes!r}")

    preferred_index = {name: idx for idx, name in enumerate(PREFERRED_WORKBOOK_TAB_ORDER)}
    workbook_tabs = [ws.title for ws in wb.worksheets]
    wb_indices = [preferred_index[t] for t in workbook_tabs if t in preferred_index]
    if wb_indices != sorted(wb_indices):
        errors.append(
            f"Workbook tab order violates PREFERRED_WORKBOOK_TAB_ORDER: {workbook_tabs}"
        )

    toc_tabs = _toc_listed_sheets(toc)
    expected_toc = [
        t
        for t in PREFERRED_WORKBOOK_TAB_ORDER
        if t != "Table of Contents"
        and t in wb.sheetnames
        and t not in SHEETS_EXCLUDED_FROM_TOC
    ]
    for tab in workbook_tabs:
        if (
            tab != "Table of Contents"
            and tab not in _PREFERRED_TAB_SET
            and tab not in SHEETS_EXCLUDED_FROM_TOC
        ):
            expected_toc.append(tab)
    if toc_tabs != expected_toc:
        missing_in_toc = (
            set(workbook_tabs)
            - {"Table of Contents"}
            - set(toc_tabs)
            - SHEETS_EXCLUDED_FROM_TOC
        )
        if missing_in_toc:
            errors.append(f"TOC missing sheets: {sorted(missing_in_toc)}")
        if set(toc_tabs) - set(workbook_tabs):
            errors.append(f"TOC lists absent sheets: {set(toc_tabs) - set(workbook_tabs)}")
        if not missing_in_toc and toc_tabs != expected_toc:
            errors.append(f"TOC order mismatch: expected {expected_toc}, got {toc_tabs}")

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and _FORMULA_ERROR_RE.search(cell.value):
                    errors.append(
                        f"{ws.title}!{cell.coordinate} literal error token: {cell.value}"
                    )

    hub_name = CONTENT_OPTIMISATION_HUB_SHEET
    if hub_name in wb.sheetnames:
        hub = wb[hub_name]
        headers = {str(c.value).strip(): c.column for c in hub[2] if c.value}
        action_col = headers.get("Action Required")
        if action_col is None:
            errors.append(f"{hub_name} missing Action Required column")
        else:
            # Row 3 is the scope-note row (merged); data starts at row 4.
            for r in range(4, hub.max_row + 1):
                val = hub.cell(row=r, column=action_col).value
                if val is None or str(val).strip() == "":
                    errors.append(
                        f"{hub_name}!{hub.cell(row=r, column=action_col).coordinate} "
                        "empty Action Required"
                    )
                elif isinstance(val, str) and not val.startswith("="):
                    if val.strip() not in _ACTION_LITERALS:
                        errors.append(
                            f"{hub_name}!{hub.cell(row=r, column=action_col).coordinate} "
                            f"unexpected Action Required literal: {val!r}"
                        )

    if "Main" in wb.sheetnames:
        main_ws = wb["Main"]
        main_headers = {str(c.value).strip(): c.column for c in main_ws[1] if c.value}
        state_col = main_headers.get("Extraction State")
        url_col = main_headers.get("URL")
        if state_col is None:
            errors.append("Main sheet missing Extraction State column")
        elif main_ws.max_row < 2:
            errors.append("Main sheet has no data rows")
        else:
            allowed_states = frozenset({"complete", "partial", "skipped"})
            for r in range(2, main_ws.max_row + 1):
                state = str(main_ws.cell(row=r, column=state_col).value or "").strip().lower()
                if state and state not in allowed_states:
                    errors.append(
                        f"Main!{main_ws.cell(row=r, column=state_col).coordinate} "
                        f"invalid Extraction State: {state!r}"
                    )
        if url_col is None:
            errors.append("Main sheet missing URL column")

    wb.close()
    return errors


def count_main_rows(path: Path | str) -> int:
    """Return data row count on the Main sheet (excluding header)."""
    wb = load_workbook(Path(path), read_only=True)
    try:
        if "Main" not in wb.sheetnames:
            return 0
        return max(0, wb["Main"].max_row - 1)
    finally:
        wb.close()


__all__ = [
    "REQUIRED_FULL_SUITE_SHEETS",
    "audit_workbook",
    "count_main_rows",
]
