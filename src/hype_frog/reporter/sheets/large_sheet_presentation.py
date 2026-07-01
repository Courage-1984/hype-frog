"""Large-sheet presentation pass (Phase 4C): gridlines off, CF zebra, header grid."""

from __future__ import annotations

from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from hype_frog.reporter.sheets.conditional import apply_cf_zebra_banding
from hype_frog.reporter.sheets.config import (
    CHART_DATA_SHEET,
    CONTENT_HUB_METRICS_SHEET,
    CONTENT_OPTIMISATION_HUB_SHEET,
    CONTENT_PLANNER_SHEET,
    DISABLE_CONDITIONAL_FORMATTING,
    EXECUTIVE_BRIEFING_SHEET,
    GRID_BORDER,
    LARGE_SHEET_ROW_THRESHOLD,
)
from hype_frog.reporter.sheets.sheet_rows import sheet_data_header_row

_EXEMPT_LARGE_SHEET_PRESENTATION: frozenset[str] = frozenset(
    {
        EXECUTIVE_BRIEFING_SHEET,
        "Table of Contents",
        CONTENT_OPTIMISATION_HUB_SHEET,
        CONTENT_PLANNER_SHEET,
        CONTENT_HUB_METRICS_SHEET,
        CHART_DATA_SHEET,
    }
)


def should_apply_large_sheet_presentation(sheet_name: str, worksheet: Worksheet) -> bool:
    """Return True when the sheet exceeds the large-inventory row threshold."""
    if sheet_name in _EXEMPT_LARGE_SHEET_PRESENTATION:
        return False
    if worksheet.max_row <= LARGE_SHEET_ROW_THRESHOLD:
        return False
    if worksheet.max_column < 2:
        return False
    return True


def apply_large_sheet_presentation(worksheet: Worksheet, sheet_name: str) -> None:
    """Apply CF zebra banding and a light header grid (no per-cell data fills)."""
    if not should_apply_large_sheet_presentation(sheet_name, worksheet):
        return

    worksheet.sheet_view.showGridLines = False

    header_row = sheet_data_header_row(sheet_name)
    if sheet_name == CONTENT_OPTIMISATION_HUB_SHEET:
        header_row = 2
    data_start = header_row + 1
    if worksheet.max_row < data_start:
        return

    max_col = worksheet.max_column
    last_col = get_column_letter(max_col)
    thin = Side(style="thin", color=GRID_BORDER)
    for col_idx in range(1, max_col + 1):
        cell = worksheet.cell(row=header_row, column=col_idx)
        cell.border = Border(
            bottom=thin,
            right=thin if col_idx < max_col else thin,
        )

    if DISABLE_CONDITIONAL_FORMATTING:
        return

    apply_cf_zebra_banding(worksheet, sheet_name, header_row=header_row)


__all__ = [
    "apply_large_sheet_presentation",
    "should_apply_large_sheet_presentation",
]
