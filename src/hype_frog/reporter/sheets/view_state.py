from __future__ import annotations

from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.views import Selection
from openpyxl.worksheet.worksheet import Worksheet

from hype_frog.reporter.sheets.config import CONTENT_OPTIMISATION_HUB_SHEET


def ranges_overlap(range_a: CellRange, range_b: CellRange) -> bool:
    return not (
        range_a.max_row < range_b.min_row
        or range_b.max_row < range_a.min_row
        or range_a.max_col < range_b.min_col
        or range_b.max_col < range_a.min_col
    )


def audit_non_overlapping_merges(worksheet: Worksheet) -> None:
    merge_ranges: list[CellRange] = list(worksheet.merged_cells.ranges)
    if len(merge_ranges) < 2:
        return

    kept: list[CellRange] = []
    for merge_range in merge_ranges:
        if any(ranges_overlap(merge_range, existing) for existing in kept):
            worksheet.unmerge_cells(str(merge_range))
            continue
        kept.append(merge_range)


def audit_freeze_merge_conflicts(worksheet: Worksheet) -> None:
    freeze = worksheet.freeze_panes
    if not freeze:
        return

    freeze_ref: str = freeze if isinstance(freeze, str) else freeze.coordinate
    freeze_row, freeze_col = coordinate_to_tuple(freeze_ref)
    for merge_range in list(worksheet.merged_cells.ranges):
        if (
            merge_range.min_row <= freeze_row <= merge_range.max_row
            and merge_range.min_col <= freeze_col <= merge_range.max_col
        ):
            worksheet.unmerge_cells(str(merge_range))


def set_freeze_panes_safe(worksheet: Worksheet, value: str | None) -> None:
    view = worksheet.views.sheetView[0]
    if not view.selection:
        view.selection = [Selection(activeCell="A1", sqref="A1")]
    worksheet.freeze_panes = value
    sanitize_sheet_view_selection(worksheet)


def sanitize_sheet_view_selection(worksheet: Worksheet) -> None:
    view = worksheet.views.sheetView[0]
    pane = worksheet.sheet_view.pane
    if not view.selection:
        return

    x_split: float = float(getattr(pane, "xSplit", 0) or 0) if pane is not None else 0.0
    y_split: float = float(getattr(pane, "ySplit", 0) or 0) if pane is not None else 0.0

    valid_panes: set[str | None] = {None, ""}
    if x_split > 0 and y_split > 0:
        valid_panes.update({"topRight", "bottomLeft", "bottomRight"})
    elif x_split > 0 and y_split == 0:
        valid_panes.add("topRight")
    elif x_split == 0 and y_split > 0:
        valid_panes.add("bottomLeft")

    sanitized = [
        sel for sel in view.selection if getattr(sel, "pane", None) in valid_panes
    ]
    if not sanitized:
        sanitized = [Selection(activeCell="A1", sqref="A1")]
    view.selection = sanitized


def apply_optimal_view_state(worksheet: Worksheet, sheet_name: str) -> None:
    """Apply governed freeze-pane defaults with ghost-pane sanitization.

    Args:
        worksheet: Worksheet to configure.
        sheet_name: Canonical sheet name for rule selection.
    """
    two_dimensional_freeze_sheets: set[str] = {"Main", "Content", "Technical"}
    should_clear_freeze = sheet_name not in {"Main", "Dashboard"} and (
        worksheet.max_row < 10 or worksheet.max_column < 5
    )

    if should_clear_freeze:
        set_freeze_panes_safe(worksheet, None)
        sanitize_sheet_view_selection(worksheet)
        return

    if sheet_name == CONTENT_OPTIMISATION_HUB_SHEET:
        set_freeze_panes_safe(worksheet, "H3")
        sanitize_sheet_view_selection(worksheet)
        return

    freeze_target = "B2" if sheet_name in two_dimensional_freeze_sheets else "A2"
    set_freeze_panes_safe(worksheet, freeze_target)
    sanitize_sheet_view_selection(worksheet)


__all__ = [
    "ranges_overlap",
    "audit_non_overlapping_merges",
    "audit_freeze_merge_conflicts",
    "set_freeze_panes_safe",
    "sanitize_sheet_view_selection",
    "apply_optimal_view_state",
]
