from __future__ import annotations

from collections.abc import Callable
from typing import Any

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from hype_frog.core import get_logger
from hype_frog.reporter.sheets.config import (
    CONTENT_OPTIMISATION_HUB_SHEET,
    MAIN_TRIAGE_COLUMN_COUNT,
    MAIN_TRIAGE_VISIBLE_HEADERS,
)
from hype_frog.reporter.sheets.style_helpers import header_index, header_row_index, to_int

logger = get_logger(__name__)

# Canonical Performance / CWV column order (LI-HF-PSI-P0 Part 3).
PERFORMANCE_CWV_GROUP_COLUMNS: tuple[str, ...] = (
    "CWV LCP (s)",
    "CWV CLS",
    "Field vs Lab",
    "CWV Data Source",
    "PSI Data Status",
    "CrUX Level",
    "CWV INP (ms)",
    "CWV FCP (ms)",
    "CWV TTFB (ms)",
    "CrUX LCP Category",
    "CrUX CLS Category",
    "CrUX INP Category",
    "Origin CrUX LCP (s)",
    "Origin CrUX CLS",
    "Origin CrUX INP (ms)",
    "Desktop PSI Score",
    "Mobile PSI Score",
    "Lighthouse Performance (Mobile)",
    "Lighthouse Accessibility (Mobile)",
    "Lighthouse Best Practices (Mobile)",
    "Lighthouse SEO Score (Mobile)",
    "Mobile LCP (s)",
    "Mobile CLS",
    "Mobile TTFB (s)",
    "Lab LCP (Mobile) (s)",
    "Lab CLS (Mobile)",
    "Lab TBT (Mobile) (ms)",
    "Lab INP (Mobile) (ms)",
    "Lab FCP (Mobile) (s)",
    "Lab Speed Index (Mobile) (s)",
    "Lab TTI (Mobile) (s)",
    "Lab TTFB (Mobile) (ms)",
    "Lighthouse Performance (Desktop)",
    "Lighthouse Accessibility (Desktop)",
    "Lighthouse Best Practices (Desktop)",
    "Lighthouse SEO Score (Desktop)",
    "Lab LCP (Desktop) (s)",
    "Lab CLS (Desktop)",
    "Lab TBT (Desktop) (ms)",
    "Lab INP (Desktop) (ms)",
    "Lab FCP (Desktop) (s)",
    "Lab Speed Index (Desktop) (s)",
    "Lab TTI (Desktop) (s)",
    "Lab TTFB (Desktop) (ms)",
    "Page Size (KB)",
    "DOM Size (nodes)",
    "JS Execution (ms)",
    "Network Request Count",
    "Has Text Compression",
    "Has Long Cache TTL Issues",
    "Has Render Blocking Resources",
    "Uses Modern Image Formats",
    "Reachable from Homepage",
    "Regional Authority Score",
)

# Canonical post-``reorder_columns`` layouts (single source for column helpers).
# Preferred column orders for active export sheets only.
# NOTE: "Issue Type" must stay at index 0 for "FixPlan" — Quick Wins' and
# FixPlan's own "Jump to Playbook"/"Jump to FixPlan" HYPERLINK formulas assume
# "Issue Type" lands in worksheet column A after reordering.
_PREFERRED_COLUMN_ORDERS: dict[str, list[str]] = {
    "Main": [
        "Health Icon",
        "URL",
        "Status Code",
        "Indexability",
        "Load Time (s)",
        "Title",
        "Meta Description",
        "Word Count (Body)",
        "SEO Health Score",
        "Severity Badge",
        "Action Needed",
        "Owner",
        "Status",
        "Sprint",
        *PERFORMANCE_CWV_GROUP_COLUMNS,
    ],
    "FixPlan": [
        "Issue Type",
        "Severity",
        "Priority Score",
        "Affected Count",
        "Affected Link Instances",
        "Affected URLs",
        "Detail Reference Tab",
        "Resolution Type",
        "URL",
        "Recommended Fix",
        "What It Is",
        "Likely Root Cause",
        "Owner",
        "Agency Owner",
        "Effort",
        "Est. Hours",
        "Est. Sprint Points",
        "Aging/Priority",
        "Status",
        "Verified By",
        "Date Resolved",
        "Revenue Risk",
        "Action Needed",
        "Jump to Details",
        "Jump to Playbook",
        "View Details",
        "Sprint",
    ],
    "Summary": [
        "Section",
        "Severity",
        "Issue",
        "Affected URL Count",
        "Reference Tab",
        "Affected URLs (sample)",
    ],
    "Priority URLs": [
        "URL",
        "Business Risk Score",
        "Severity Badge",
        "SEO Health Score",
        "GSC Impressions",
        "GSC CTR",
        "GSC Data Freshness",
        "GSC Coverage Note",
        "Revenue Intent",
        "Critical Issues Count",
        "Warning Issues Count",
        "Action Needed",
        "Why Prioritized",
        "Owner",
        "Status",
        "Sprint",
    ],
    "AIOSEO Recommendations": [
        "URL",
        "WordPress Post ID",
        "Direct Edit Link",
        "AIOSEO Panel",
        "Severity",
        "Issue",
        "Priority Score",
        "Current Value",
        "Recommended Target",
        "How to Fix in AIOSEO",
        "Reference Tab",
        "Reference Field",
        "Action Needed",
        "Owner",
        "Status",
        "Est. Hours",
        "Stable Issue ID",
    ],
    CONTENT_OPTIMISATION_HUB_SHEET: [
        "Action Required",
        "On-Page Optimization Score",
        "SEO Score",
        "Technical Health",
        "Copy Score",
        "Status",
        "Assigned Owner",
        "URL Slug Normalization",
        "URL",
        "Proposed URL Slug",
        "Current Title",
        "Title Health",
        "Current Meta Desc",
        "Meta Health",
        "H1",
        "H1 Health",
        "H2",
        "H2 Health",
        "H3",
        "H3 Health",
        "H4",
        "H4 Health",
        "H5",
        "H5 Health",
        "H6",
        "H6 Health",
        "Elementor Builder Link",
        "Current OG-Image URL",
        "OG Image Health",
        "OG Image Preview",
        "Open in Main",
    ],
}

# Deep header copy for Content Optimisation Hub (cell comments; see ``apply_header_tooltips``).
CONTENT_HUB_HEADER_COMMENT_SEO_SCORE: str = (
    "SEO Score (0–100) is the blended on-page and technical-quality signal from the crawl. "
    "It summarises how well this URL satisfied the audit model at export time: headings, "
    "metadata coverage, thin-content risk, and issue severity baked into the pipeline score. "
    "Use it as a triage anchor, not a live CMS measurement. When you rewrite titles or body "
    "copy here, watch Title Health and On-Page Optimization Score for live feedback; expect "
    "SEO Score to stay flat until you re-crawl. Interpret bands roughly as: 80+ strong "
    "baseline, 60–79 tune content and internal links, below 60 treat as a rewrite or "
    "template-level fix candidate."
)

CONTENT_HUB_HEADER_COMMENT_TECHNICAL_HEALTH: str = (
    "Technical Health (0–100) isolates crawl and response hygiene: status class, indexability, "
    "redirect behaviour, canonical consistency, and similar hard signals. High values mean "
    "fewer blocking technical defects; low values mean the page may be losing equity to "
    "errors, chains, or conflicting directives even when copy looks fine. Pair this column "
    "with Technical Diagnostics for evidence. It is intentionally decoupled from editorial "
    "drafts in this hub so copywriters can see whether engineering work is still a gate "
    "before declaring a page production-ready."
)

CONTENT_HUB_HEADER_COMMENT_OG_IMAGE_PREVIEW: str = (
    "Specs: 1200x630px, <300KB. Content: Use a high-contrast brand image with a centered "
    "focal point. Avoid placing text near the edges as it will be cropped on mobile social feeds."
)

# Row-2 header cell comments for Content Optimisation Hub (openpyxl Comment; not Data Validation).
#
# Tooltip ownership: any header also present in engine_guardrails._HEADER_TOOLTIP_MESSAGES is
# NOT listed here, because apply_header_tooltips() runs after this dict is applied (see
# tables_impl.py, apply_content_hub_conditional_rules then apply_header_tooltips) and would
# silently overwrite it. Keep the two dicts disjoint rather than relying on call order — see
# reporter/CLAUDE.md "Tooltip ownership" for the full rule.
CONTENT_HUB_ROW2_HEADER_COMMENTS: dict[str, str] = {
    "On-Page Optimization Score": (
        "Live calculation (0-100) of on-page health. Factors: Title/Meta length and H-tag hierarchy. Aim for 90+."
    ),
    "Title Health": (
        "Live formula: length vs 50–60 character target. Green when OK band; red when missing; "
        "orange for short/long."
    ),
    "Meta Health": "Live formula for meta length vs target band (same semantics as Title Health).",
    "H1": "Primary heading text from crawl (H1 line in Current H-Tag Structure when main H1 is absent).",
    "H1 Health": "Live formula: missing H1, long block warning, or OK when present.",
    "H2": "First H2 text from heading outline when not present on Main row.",
    "H2 Health": "Live formula: optional/tip/present states for section headings.",
    "H3": "First H3 from heading outline.",
    "H3 Health": "Optional heading guidance; Present when text exists.",
    "H4": "Parsed when present in heading outline.",
    "H4 Health": "Optional H4 guidance.",
    "H5": "Parsed when present in heading outline.",
    "H5 Health": "Optional H5 guidance.",
    "H6": "Parsed when present in heading outline.",
    "H6 Health": "Optional H6 guidance.",
    "URL Slug Normalization": (
        "Editorial slug guidance field. Capture normalized slug wording you want reflected in URL, "
        "title, and heading language before publishing."
    ),
    "Proposed URL Slug": (
        "Editorial URL-slug suggestion — not applied automatically; use alongside Title/H1 "
        "guidance before publishing."
    ),
}


def main_sheet_url_column_letter() -> str:
    """Excel column letter for ``URL`` on Main after :func:`reorder_columns`.

    Returns:
        Column letter (for example ``B`` when ``Health Icon`` occupies column A).
    """
    order = _PREFERRED_COLUMN_ORDERS["Main"]
    return get_column_letter(order.index("URL") + 1)


def sheet_data_column_range(
    sheet_name: str,
    header: str,
    *,
    start_row: int | None = None,
    end_row: int = 100_000,
) -> str:
    """Excel OFFSET range for ``header`` on ``sheet_name`` (dynamic column lookup)."""
    from hype_frog.reporter.sheets.sheet_rows import (
        sheet_data_header_row,
        sheet_data_start_row,
    )

    header_row = sheet_data_header_row(sheet_name)
    data_start = start_row if start_row is not None else sheet_data_start_row(sheet_name)
    span = max(1, end_row - data_start + 1)
    row_offset = data_start - 1
    header_row_ref = f"'{sheet_name}'!${header_row}:${header_row}"
    return (
        f"OFFSET('{sheet_name}'!$A$1,{row_offset},"
        f'MATCH("{header}",{header_row_ref},0)-1,{span},1)'
    )


def link_inventory_column_letter(header: str) -> str:
    """Column letter on Link Inventory from the strict seven-column export contract."""
    from hype_frog.reporter.sheets.merged_builders import LINK_INVENTORY_COLUMNS

    return get_column_letter(LINK_INVENTORY_COLUMNS.index(header) + 1)


def link_intelligence_column_letter(header: str) -> str:
    """Column letter on Link Intelligence from the merged export contract."""
    from hype_frog.reporter.sheets.merged_builders import LINK_INTELLIGENCE_COLUMNS

    return get_column_letter(LINK_INTELLIGENCE_COLUMNS.index(header) + 1)


# British English display labels for pipeline keys (append-only row keys unchanged).
DISPLAY_HEADER_ALIASES: dict[str, str] = {
    "On-Page Optimization Score": "On-Page Optimisation Score",
}


def apply_display_header_aliases(worksheet: Worksheet, *, header_row: int = 1) -> None:
    """Rewrite row headers to British English display labels where mapped."""
    for col_idx in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=header_row, column=col_idx)
        key = str(cell.value or "").strip()
        alias = DISPLAY_HEADER_ALIASES.get(key)
        if alias:
            cell.value = alias


MAIN_COLUMN_GROUP_DEFINITIONS: dict[str, list[str]] = {
    "Metadata Group": [
        "Title",
        "Meta Description",
        "Title Length",
        "Meta Desc Length",
    ],
    "Heading Structure Group": [
        "H1 Content",
        "H1 Length",
        "H2 Content",
        "H2 Length",
        "H3 Content",
        "H3 Length",
        "H4 Content",
        "H4 Length",
        "H5 Content",
        "H5 Length",
        "H6 Content",
        "H6 Length",
    ],
    "Performance & CWV Group": list(PERFORMANCE_CWV_GROUP_COLUMNS),
    "Google Search Console Group": [
        "GSC Index Status",
        "GSC Last Crawl Date",
        "GSC Mobile Usability",
        "GSC Rich Result Status",
        "GSC Coverage Reason",
        "Days Since Last Crawl",
        "GSC Clicks",
        "GSC Impressions",
        "GSC CTR",
        "GSC Avg Position",
        "GSC Data Freshness",
        "GSC Coverage Note",
    ],
    "Crawl & Discovery Group": [
        "Click Depth",
        "Orphan Pages",
        "Internal PageRank",
        "Found via Sitemap",
        "Found via Crawl",
        "Discovery Source",
        "Discovered On URL",
    ],
    "Raw State Group": [
        "Extraction State",
        "Extraction Source",
        "Technical Health",
        "Copy Score",
        "SEO Score",
        "Technical View",
    ],
    "Schema & Structured Data": [
        "Schema Present",
        "Schema Valid",
        "Has Valid JSON-LD",
        "Schema Types Found",
        "Schema Types Valid",
        "Schema Types With Errors",
        "Schema Error Count",
        "Schema Warning Count",
        "Schema Parse Error Detail",
        "Schema Validation Summary",
        "Schema Issues Detail",
    ],
    "E-E-A-T & Trust Signals": [
        "E-E-A-T Signal Score",
        "Schema Author Name",
        "Meta Author",
        "Has Byline Element",
        "Byline Text",
        "Schema Published Date",
        "Schema Modified Date",
        "OG Published Time",
        "OG Modified Time",
        "Has Time Element",
        "Has Privacy Policy Link",
        "Has Terms Link",
        "Has Social Links",
        "Social Profile Link Count",
        "Has Phone Number",
        "Has Email Address",
        "Links to About Page",
        "Has Authority External Links",
    ],
    "Content Quality": [
        "Is Thin Content",
        "Is Near Duplicate",
        "Near Duplicate Of",
        "Content Similarity Score",
        "Is Draft or Test Page",
        "Draft Signal",
        "Thin Content Flag",
        "Probable Duplicate Flag",
    ],
    "Social Cards": [
        "OG Title",
        "OG Description",
        "OG Type",
        "OG URL",
        "OG Image URL",
        "OG-Image",
        "OG Image Width",
        "OG Image Height",
        "OG Image OK",
        "OG Image Dimensions OK",
        "OG URL Mismatch",
        "OG Completeness Score",
        "Open Graph Complete",
        "Twitter Card Type",
        "Twitter Title",
        "Twitter Description",
        "Twitter Image",
    ],
    "Redirects": [
        "Final URL",
        "Redirect Chain",
        "Redirect Chain Length",
        "Redirect Chain Hops",
        "Has 302 in Chain",
        "Has Mixed Redirect Types",
        "Redirect Loop Flag",
    ],
    "Canonical Chain": [
        "Canonical URL",
        "Canonical Type",
        "Canonical Chain Depth",
        "Canonical Chain Final",
        "Canonical Chain",
        "Canonical Loop Detected",
        "Canonical Points to Redirect",
        "Canonical Points to Non-200",
    ],
    "Robots.txt": [
        "Robots.txt: Googlebot",
        "Robots.txt: Bingbot",
        "Robots.txt: GPTBot",
        "Robots.txt: ClaudeBot",
        "Robots.txt: PerplexityBot",
        "Crawl-Delay Applies",
        "Robots.txt Accessible",
    ],
}


def sort_worksheet_rows(
    worksheet: Worksheet, key_fn: Callable[[list[Any]], tuple[Any, ...]]
) -> None:
    """Sort all data rows in place while preserving the header row.

    Args:
        worksheet: Target worksheet containing a header row at index 1.
        key_fn: Sort-key callable receiving each row as a list of values.
    """
    if worksheet.max_row <= 2:
        return
    rows = [
        list(row)
        for row in worksheet.iter_rows(
            min_row=2, max_row=worksheet.max_row, values_only=True
        )
    ]
    rows.sort(key=key_fn)
    for row_idx, row_values in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_values, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)


def apply_intelligent_sorting(worksheet: Worksheet, sheet_name: str) -> None:
    """Apply sheet-specific deterministic sorting for report readability.

    Args:
        worksheet: Worksheet to sort.
        sheet_name: Canonical worksheet name controlling sort strategy.
    """
    headers = header_index(worksheet)
    if sheet_name == "FixPlan":
        severity_rank: dict[str, int] = {
            "Critical": 0,
            "High": 1,
            "Warning": 2,
            "Medium": 3,
            "Low": 4,
            "Observation": 5,
        }
        pcol = headers.get("Priority Score")
        scol = headers.get("Severity")
        ucol = headers.get("URL")
        rcol = headers.get("Discovery Rank")
        if pcol:
            sort_worksheet_rows(
                worksheet,
                key_fn=lambda r: (
                    -to_int(r[pcol - 1], 0),
                    severity_rank.get(str(r[scol - 1]) if scol else "", 99),
                    to_int(r[rcol - 1], 10**9) if rcol else 10**9,
                    str(r[ucol - 1] or "") if ucol else "",
                ),
            )
    elif sheet_name in {"Main", "Technical", "Technical Diagnostics"}:
        rank_col = headers.get("Discovery Rank")
        ucol = headers.get("URL")
        if rank_col:
            sort_worksheet_rows(
                worksheet,
                key_fn=lambda r: (
                    to_int(r[rank_col - 1], 10**9),
                    str(r[ucol - 1] or "") if ucol else "",
                ),
            )
    elif sheet_name == "AIOSEO Recommendations":
        severity_rank = {"Critical": 0, "Warning": 1, "Observation": 2}
        sev_col = headers.get("Severity")
        pri_col = headers.get("Priority Score")
        panel_col = headers.get("AIOSEO Panel")
        ucol = headers.get("URL")
        if sev_col:
            sort_worksheet_rows(
                worksheet,
                key_fn=lambda r: (
                    severity_rank.get(str(r[sev_col - 1]) if sev_col else "", 99),
                    -to_int(r[pri_col - 1], 0) if pri_col else 0,
                    str(r[panel_col - 1] or "") if panel_col else "",
                    str(r[ucol - 1] or "") if ucol else "",
                ),
            )


def hide_noisy_columns(worksheet: Worksheet, sheet_name: str) -> None:
    """Hide low-signal raw/debug columns for client-facing exports.

    Args:
        worksheet: Worksheet to mutate.
        sheet_name: Canonical worksheet name used for token selection.
    """
    noisy_tokens_by_sheet: dict[str, list[str]] = {
        "Main": ["json-ld", "schema", "raw", "headers", "paragraph", "text extract"],
        "Technical": [
            "json-ld",
            "schema",
            "raw",
            "headers",
            "html",
            "paragraph",
            "text extract",
        ],
        "Content": ["paragraph", "full text", "raw", "json", "headers"],
    }
    tokens = noisy_tokens_by_sheet.get(sheet_name)
    if tokens:
        for idx, cell in enumerate(worksheet[1], start=1):
            header = str(cell.value or "").lower()
            if any(tok in header for tok in tokens):
                if isinstance(cell, MergedCell):
                    continue
                worksheet.column_dimensions[get_column_letter(idx)].hidden = True

    # "Source Legacy Tab" is internal migration provenance (pre-merge tab names like
    # "LinksDetail"/"Duplicates"/"IssueInventory" that no longer exist as sheets) —
    # useful for debugging the merged-sheet builders, not for the end reader. Hide
    # rather than remove so column position/order stays stable for existing tests
    # and cross-sheet references.
    from hype_frog.reporter.sheets.sheet_rows import sheet_data_header_row

    header_row = sheet_data_header_row(sheet_name)
    legacy_col = header_row_index(worksheet, header_row).get("Source Legacy Tab")
    if legacy_col:
        worksheet.column_dimensions[get_column_letter(legacy_col)].hidden = True


def content_optimisation_hub_ordered_headers(
    all_headers: tuple[str, ...],
) -> tuple[str, ...]:
    """Return Content Optimisation Hub column order after :func:`reorder_columns`.

    The export pipeline writes row 1 headers then calls ``reorder_columns``,
    which places ``_PREFERRED_COLUMN_ORDERS[Content Optimisation Hub]`` first
    and appends any remaining headers in their original scan order. Formula
    builders must use this same ordering so column letters match the physical
    grid (not the pre-reorder DataFrame column list).

    Args:
        all_headers: Full header tuple in the initial write order (typically
            ``_CONTENT_HUB_FIELDS_PRE_REORDER`` from ``engine_rows``).

    Returns:
        Header names in final worksheet column order.
    """
    preferred = _PREFERRED_COLUMN_ORDERS[CONTENT_OPTIMISATION_HUB_SHEET]
    current = list(all_headers)
    ordered = [h for h in preferred if h in current]
    ordered.extend([h for h in current if h not in ordered])
    return tuple(ordered)


def reorder_columns(worksheet: Worksheet, sheet_name: str) -> None:
    """Reorder sheet columns according to preferred operational layouts.

    Args:
        worksheet: Worksheet to mutate.
        sheet_name: Canonical sheet name used to pick a preferred ordering.
    """
    preferred = _PREFERRED_COLUMN_ORDERS.get(sheet_name)
    if not preferred or worksheet.max_row < 1:
        return
    current_headers = [
        worksheet.cell(row=1, column=i).value
        for i in range(1, worksheet.max_column + 1)
    ]
    if not any(h in current_headers for h in preferred):
        return
    ordered_headers = [h for h in preferred if h in current_headers]
    ordered_headers.extend([h for h in current_headers if h not in ordered_headers])
    if ordered_headers == current_headers:
        return
    idx_map = [current_headers.index(h) + 1 for h in ordered_headers]
    rows: list[list[Any]] = []
    for row_idx in range(1, worksheet.max_row + 1):
        rows.append(
            [worksheet.cell(row=row_idx, column=src_col).value for src_col in idx_map]
        )
    for row_idx, row_vals in enumerate(rows, start=1):
        for col_idx, val in enumerate(row_vals, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=val)


def apply_column_grouping(
    worksheet: Worksheet,
    group_definitions: dict[str, list[str]],
    *,
    header_row: int = 1,
    min_col: int = 1,
) -> None:
    """Apply collapsible nested column outlines for related metric groups.

    Args:
        worksheet: Worksheet where column grouping should be applied.
        group_definitions: Mapping of group name to ordered header names.
        header_row: Row containing column headers.
        min_col: Only group columns at or after this index (1-based).
    """
    if worksheet.max_column <= 1:
        return
    if not group_definitions:
        return

    headers = header_row_index(worksheet, header_row)
    for _group_name, grouped_headers in group_definitions.items():
        present_indices: list[int] = [
            headers[header_name]
            for header_name in grouped_headers
            if header_name in headers and headers[header_name] >= min_col
        ]
        if not present_indices:
            continue

        present_indices = sorted(set(present_indices))
        range_start = present_indices[0]
        prev_col = present_indices[0]
        for col_idx in present_indices[1:]:
            if col_idx == prev_col + 1:
                prev_col = col_idx
                continue
            worksheet.column_dimensions.group(
                get_column_letter(range_start),
                get_column_letter(prev_col),
                hidden=True,
                outline_level=2,
            )
            range_start = col_idx
            prev_col = col_idx
        worksheet.column_dimensions.group(
            get_column_letter(range_start),
            get_column_letter(prev_col),
            hidden=True,
            outline_level=2,
        )


def apply_main_triage_column_layout(
    worksheet: Worksheet, *, header_row: int = 1
) -> None:
    """Collapse columns after triage block (L+) leaving A–K visible by default."""
    if worksheet.max_column <= MAIN_TRIAGE_COLUMN_COUNT:
        return

    headers = header_row_index(worksheet, header_row)
    triage_end = headers.get("Action Needed", MAIN_TRIAGE_COLUMN_COUNT)
    for header_name in MAIN_TRIAGE_VISIBLE_HEADERS:
        col_idx = headers.get(header_name)
        if col_idx is not None:
            triage_end = max(triage_end, col_idx)

    back_col = headers.get("BACK TO DASHBOARD")
    max_col = worksheet.max_column
    if back_col is not None and back_col == max_col:
        worksheet.delete_cols(back_col)
        max_col -= 1
        headers = header_row_index(worksheet, header_row)

    for col in range(1, triage_end + 1):
        letter = get_column_letter(col)
        dim = worksheet.column_dimensions[letter]
        dim.hidden = False
        dim.outlineLevel = 0

    if max_col > triage_end:
        worksheet.column_dimensions.group(
            get_column_letter(triage_end + 1),
            get_column_letter(max_col),
            hidden=True,
            outline_level=1,
        )
        apply_column_grouping(
            worksheet,
            MAIN_COLUMN_GROUP_DEFINITIONS,
            header_row=header_row,
            min_col=triage_end + 1,
        )

    worksheet.sheet_properties.outlinePr.summaryBelow = True
    worksheet.sheet_properties.outlinePr.summaryRight = True


# Cycling tint palette for hidden column-group headers (Main sheet). Applied to
# header cells only, so groups are visually distinguishable once a user expands
# the outline — the columns themselves stay collapsed by default.
_COLUMN_GROUP_HEADER_TINTS: tuple[str, ...] = (
    "EAF2FB",  # soft blue
    "EAF7EF",  # soft green
    "FDF3E4",  # soft amber
    "F3EAFB",  # soft purple
    "F0F0F0",  # soft grey
)
_COLUMN_GROUP_HEADER_TEXT: str = "222A35"


def apply_main_column_group_header_tints(
    worksheet: Worksheet, *, header_row: int = 1
) -> None:
    """Tint hidden column-group headers so an expanded outline shows group boundaries.

    Must run after any generic header styling pass (e.g. ``apply_mock_table_
    styling``) — that pass paints the whole header row a single colour and would
    otherwise overwrite these per-group tints.
    """
    if worksheet.max_column <= MAIN_TRIAGE_COLUMN_COUNT:
        return

    headers = header_row_index(worksheet, header_row)
    for tint_idx, grouped_headers in enumerate(MAIN_COLUMN_GROUP_DEFINITIONS.values()):
        present = sorted({headers[name] for name in grouped_headers if name in headers})
        if not present:
            continue
        tint = _COLUMN_GROUP_HEADER_TINTS[tint_idx % len(_COLUMN_GROUP_HEADER_TINTS)]
        fill = PatternFill(start_color=tint, end_color=tint, fill_type="solid")
        font = Font(bold=True, color=_COLUMN_GROUP_HEADER_TEXT)
        for col_idx in present:
            cell = worksheet.cell(row=header_row, column=col_idx)
            cell.fill = fill
            cell.font = font


# Column-width contract (character units approximating Excel widths).
# _MAX_COL_WIDTH is capped below openpyxl's un-clamped auto-fit so ordinary
# (non-URL, non-prose) columns don't force horizontal scrolling on a laptop
# display at the zoom levels in sheets/config.py::SHEET_ZOOM_OVERRIDES.
_MIN_COL_WIDTH = 10.0
_MAX_COL_WIDTH = 42.0
_URL_COL_WIDTH = 45.0
_PROSE_COL_WIDTH = 55.0

# Headers rendered on a single line at a fixed, readable width (no wrap).
URL_LIKE_HEADERS: frozenset[str] = frozenset(
    {
        "URL",
        "Final URL",
        "Canonical URL",
        "Page link",
        "Direct Edit Link",
        "Discovered On URL",
        "Current OG-Image URL",
        "OG Image URL",
        "Redirect Target",
        "Sitemap First Image",
    }
)

# Long narrative / policy columns that keep wrap at a generous width.
PROSE_HEADERS: frozenset[str] = frozenset(
    {
        "Affected URLs",
        "Affected URLs (sample)",
        "How to Fix in AIOSEO",
        "Why It Matters",
        "Why Prioritized",
        "What It Is",
        "Likely Root Cause",
        "Recommended Fix",
        "Recommended Target",
        "Recommended Action",
        "Priority Reason",
        "Current Value",
        "Internal Link Statuses",
        "GSC Coverage Note",
        "How To Verify",
        "Guideline",
    }
)


def _autofit_column_width(
    worksheet: Worksheet, col_idx: int, *, start_row: int
) -> float:
    """Longest single-line content length from ``start_row`` down (formulas skipped)."""
    max_length = 0
    for row_idx in range(start_row, min(worksheet.max_row, start_row + 400) + 1):
        cell = worksheet.cell(row=row_idx, column=col_idx)
        if isinstance(cell, MergedCell):
            continue
        value = cell.value
        if value is None:
            continue
        text = str(value)
        # Formula strings (hyperlinks, cross-sheet lookups) do not reflect the
        # rendered width, so they must not drive auto-fit.
        if text.startswith("="):
            continue
        for line in text.split("\n"):
            if len(line) > max_length:
                max_length = len(line)
    return float(max_length)


CONTENT_HUB_DENSITY_OVERRIDES: dict[str, float] = {
    "Action Required": 17.43,
    "On-Page Optimization Score": 12.0,
    "Assigned Owner": 15.0,
    "Elementor Builder Link": 18.14,
    "Current OG-Image URL": 15.0,
    "OG Image Health": 42.0,
    "Open in Main": 22.57,
    "URL Slug Normalization": 22.0,
}


def apply_column_widths(worksheet: Worksheet) -> None:
    """Size every column for readability using a header-aware width contract.

    Resolves the true header row (data sheets carry a row-1 banner and row-2
    headers), auto-fits from cell content (ignoring formulas), then applies a
    per-header contract: URL-like columns get a fixed single-line width, long
    prose columns get a generous wrapped width, and everything else is clamped
    to a readable range so headers are never starved of space.
    """
    from hype_frog.reporter.sheets.sheet_rows import sheet_data_header_row

    header_row = sheet_data_header_row(worksheet.title)
    if worksheet.title == CONTENT_OPTIMISATION_HUB_SHEET:
        header_row = 2
    data_start = header_row + 1
    headers = header_row_index(worksheet, header_row)
    header_by_col = {col_idx: name for name, col_idx in headers.items()}

    for col_idx in range(1, worksheet.max_column + 1):
        letter = get_column_letter(col_idx)
        header_name = header_by_col.get(col_idx, "")
        if header_name in URL_LIKE_HEADERS:
            worksheet.column_dimensions[letter].width = _URL_COL_WIDTH
            continue
        if header_name in PROSE_HEADERS:
            worksheet.column_dimensions[letter].width = _PROSE_COL_WIDTH
            continue
        content_len = _autofit_column_width(worksheet, col_idx, start_row=data_start)
        # A header floor keeps the label legible without forcing very wide
        # columns for long multi-word headers (those wrap onto two lines).
        header_floor = min(float(len(header_name)) + 2.0, 22.0) if header_name else 0.0
        width = max(_MIN_COL_WIDTH, content_len + 2.0, header_floor)
        worksheet.column_dimensions[letter].width = min(width, _MAX_COL_WIDTH)

    # Content Hub density overrides: preserve the compact, operational view.
    for header_name, width in CONTENT_HUB_DENSITY_OVERRIDES.items():
        col_idx = headers.get(header_name)
        if col_idx:
            worksheet.column_dimensions[get_column_letter(col_idx)].width = width


__all__ = [
    "DISPLAY_HEADER_ALIASES",
    "MAIN_COLUMN_GROUP_DEFINITIONS",
    "PERFORMANCE_CWV_GROUP_COLUMNS",
    "PROSE_HEADERS",
    "URL_LIKE_HEADERS",
    "apply_display_header_aliases",
    "apply_column_grouping",
    "apply_main_column_group_header_tints",
    "apply_main_triage_column_layout",
    "apply_column_widths",
    "apply_intelligent_sorting",
    "content_optimisation_hub_ordered_headers",
    "hide_noisy_columns",
    "link_intelligence_column_letter",
    "link_inventory_column_letter",
    "main_sheet_url_column_letter",
    "reorder_columns",
    "sheet_data_column_range",
    "sort_worksheet_rows",
]
