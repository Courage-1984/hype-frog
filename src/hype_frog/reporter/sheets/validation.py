from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from hype_frog.reporter.sheets.config import (
    CONTENT_OPTIMISATION_HUB_SHEET,
    DISABLE_DATA_VALIDATION,
    DISABLE_TOOLTIPS,
    STATUS_OPTIONS,
    status_validation_list_formula,
    triage_status_validation_list_formula,
)
from hype_frog.reporter.sheets.style_helpers import header_index

# Tooltips (cell comments) are suppressed by the dedicated tooltip flag, but the legacy
# data-validation flag is still honoured for backward compatibility. Dropdowns remain
# gated by ``DISABLE_DATA_VALIDATION`` only.
_DISABLE_TOOLTIP_COMMENTS: bool = DISABLE_TOOLTIPS or DISABLE_DATA_VALIDATION

HELP_DESCRIPTION_PREFIX: str = "Description:"
HELP_CALCULATION_PREFIX: str = "Calculation:"

_DEFAULT_COMMENT_WIDTH: float = 432.0
_DEFAULT_COMMENT_HEIGHT: float = 252.0


def format_help_layer(*, description: str, calculation: str) -> str:
    """Build a standardized two-part help string for Excel cell comments."""
    desc = (description or "").strip()
    calc = (calculation or "").strip()
    return (
        f"{HELP_DESCRIPTION_PREFIX} {desc}\n"
        f"{HELP_CALCULATION_PREFIX} {calc}"
    )


def apply_comment_dimensions(
    comment: Comment,
    *,
    width: float = _DEFAULT_COMMENT_WIDTH,
    height: float = _DEFAULT_COMMENT_HEIGHT,
) -> None:
    """Widen comment boxes so Description + Calculation text fits without manual resize."""
    comment.width = width
    comment.height = height


def _attach_header_comment(cell: object, text: str, author: str) -> None:
    comment = Comment(text, author)
    apply_comment_dimensions(comment)
    cell.comment = comment


# Curated bodies (Description + Calculation only; title line added in ``add_all_header_tooltips``).
_CONTENT_AI_READINESS_HELP: dict[str, str] = {
    "AEO Readiness Score": format_help_layer(
        description=(
            "0–100 score for how well the page supports answer-engine extraction "
            "(answer blocks, answer-focused schema, readability band, scannable structure, "
            "robots.txt AI-bot coverage)."
        ),
        calculation=(
            "Python ``compute_aeo_readiness_score`` sums capped sub-scores: "
            "min(30, 15 × Paragraphs 40-60 Words Count) + 25 if QAPage/FAQ OR HowTo OR "
            "Speakable schema + FK grade piecewise (up to 20) + 15 if List/Table Answer "
            "Signal + 10 × AEO Robots AI Bot Coverage ratio; result clamped to [0,100]. "
            "Non-scorable extraction states return a neutral 71 (Unmeasured badge)."
        ),
    ),
    "AEO Extractability Score": format_help_layer(
        description=(
            "How easily an LLM can map question-style headings to factual answer paragraphs "
            "on this URL (crawl-time label or numeric)."
        ),
        calculation=(
            "Export maps crawl labels to a 0–100 scale for reporting: High→85, Medium→55, "
            "Low→25; numeric inputs clamp to [0,100] (see ``_aeo_extractability_numeric`` in "
            "merged link-intelligence / Content-AI row builders)."
        ),
    ),
    "Answer Blocks": format_help_layer(
        description=(
            "Count of 40–60 word body paragraphs under question-oriented headings — "
            "the highest-leverage AEO pattern for answer-engine citation."
        ),
        calculation=(
            "Equals ``Paragraphs 40-60 Words Count`` from the crawl/HTML pipeline. "
            "When zero, prioritise rewriting H2–H4 as natural questions (Who/What/How) "
            "and place a concise 40–60 word factual answer directly underneath each."
        ),
    ),
    "Flesch-Kincaid Grade (Est.)": format_help_layer(
        description=(
            "Estimated U.S. school grade reading difficulty for visible body text; lower is "
            "easier. The AEO model treats grades 7–10 as the clarity sweet spot."
        ),
        calculation=(
            "Computed at crawl/enrichment from body text using the Flesch–Kincaid grade "
            "formula; surfaced here as a stored numeric estimate (not an Excel formula)."
        ),
    ),
    "Schema Types Count": format_help_layer(
        description=(
            "How many distinct structured-data types (for example JSON-LD @types) were "
            "detected for this URL."
        ),
        calculation=(
            "Integer count from the schema/metadata extraction stage; written to the row at "
            "export (no workbook formula)."
        ),
    ),
    # Folded in from the former standalone "Content Hub Metrics" sheet.
    "Search Intent Source": format_help_layer(
        description="How Search Intent was determined for this URL.",
        calculation=(
            "LLM when a hosted or local OpenAI-compatible model classified it; Heuristic "
            "when a URL/title/meta keyword rule matched instead (no LLM configured or the "
            "LLM call returned Unknown); Unknown when neither classified the page."
        ),
    ),
    "Instant Priority": format_help_layer(
        description="CRITICAL flags a high-traffic page with a specific AEO or performance risk.",
        calculation=(
            "CRITICAL when GSC clicks exceed the traffic threshold AND (AEO score is low OR "
            "field LCP is slow); Standard otherwise."
        ),
    ),
    # Folded in from the former standalone "Anchor Text Audit" sheet.
    "Inbound Link Count": format_help_layer(
        description="Raw count of internal anchor instances pointing at this URL.",
        calculation=(
            "Same raw, non-deduplicated counting method as Main!Inbound Internal Link Count — "
            "see that column's tooltip for why this differs from Link Inventory's per-"
            "(source, target, anchor) row count."
        ),
    ),
}

_TECHNICAL_DIAGNOSTICS_HELP: dict[str, str] = {
    "SEO Health Score": format_help_layer(
        description=(
            "Composite 0–100 technical SEO quality for this URL from matched audit rules."
        ),
        calculation=(
            "Python ``score_url_health``: start at 100, subtract capped diminishing "
            "penalties — Critical 20 for the first +10 each extra (cap 50), Warning 8 "
            "+5 (cap 30), Observation 3 each (cap 10); clamp to ≥0. 0 is reserved for "
            "non-200/404 pages. Unmeasured when Extraction State is not scorable."
        ),
    ),
    "Desktop PSI Score": format_help_layer(
        description="Lab PageSpeed Insights performance score (0–100) for the desktop run.",
        calculation=(
            "Passed through from PSI crawl telemetry for the resolved URL; persisted on the "
            "Technical Diagnostics row (not recalculated in Excel)."
        ),
    ),
    "Mobile PSI Score": format_help_layer(
        description="Lab PageSpeed Insights performance score (0–100) for the mobile run.",
        calculation=(
            "Passed through from PSI crawl telemetry for the resolved URL; persisted on the "
            "Technical Diagnostics row (not recalculated in Excel)."
        ),
    ),
    "GSC Index Status": format_help_layer(
        description=(
            "Search Console URL Inspection-style verdict for whether Google treats this URL as "
            "indexed or blocked (verbatim crawl field)."
        ),
        calculation=(
            "Copied from ``GSC Inspection Verdict`` on the underlying extra row when Technical "
            "Diagnostics rows are merged (string pass-through, not an Excel formula)."
        ),
    ),
    "Diagnostic Category": format_help_layer(
        description="Which diagnostic areas this row touches (Technical, Indexability, Redirect, Security, Performance, Search Console).",
        calculation="Joined tag list synthesised from presence checks across this row's fields at export time.",
    ),
    "Critical Issues Count": format_help_layer(
        description="Number of Critical-severity issues matched on this URL.",
        calculation="Count from score_url_health's matched-issue list.",
    ),
    "Warning Issues Count": format_help_layer(
        description="Number of Warning-severity issues matched on this URL.",
        calculation="Count from score_url_health's matched-issue list.",
    ),
    "Pass Flag": format_help_layer(
        description="Pass when Severity Badge is Pass, else Non-Pass — a quick binary triage filter.",
        calculation='"Pass" if Severity Badge == "Pass" else "Non-Pass".',
    ),
    "Extraction State": format_help_layer(
        description="How much of this row could be measured: complete, partial, or skipped.",
        calculation="Set by the crawl/extraction pipeline based on fetch and parse outcomes.",
    ),
    "Extraction Source": format_help_layer(
        description="Which fetch path produced this row: raw HTTP or a rendered browser session.",
        calculation="Set by the crawler based on which fetch mode ultimately succeeded for this URL.",
    ),
    "Indexability Reason": format_help_layer(
        description="Why this URL is or isn't indexable (canonical, robots, status signals).",
        calculation="Rule- and signal-driven text from the indexability pipeline.",
    ),
    "Redirect Chain Length": format_help_layer(
        description="Number of hops in this URL's redirect chain (0 if no redirect).",
        calculation="Count of hops traced by the redirect-following logic during the crawl.",
    ),
    "Redirect Loop Flag": format_help_layer(
        description="True when the redirect chain loops back on itself instead of resolving.",
        calculation="Detected when the redirect tracer revisits a URL already seen in the chain.",
    ),
    "Strict-Transport-Security": format_help_layer(
        description="Raw HSTS response header value; blank means the header was not sent.",
        calculation="Pass-through from the HTTP response headers captured during the crawl.",
    ),
    "Content-Security-Policy": format_help_layer(
        description="Raw CSP response header value; blank means the header was not sent.",
        calculation="Pass-through from the HTTP response headers captured during the crawl.",
    ),
    "X-Content-Type-Options": format_help_layer(
        description="Raw X-Content-Type-Options response header value; blank means it was not sent.",
        calculation="Pass-through from the HTTP response headers captured during the crawl.",
    ),
    "Mobile LCP (s)": format_help_layer(
        description="Field/CrUX Largest Contentful Paint for mobile, in seconds (lower is better).",
        calculation="Pass-through from CrUX/PSI field data for this URL.",
    ),
    "Mobile CLS": format_help_layer(
        description="Field/CrUX Cumulative Layout Shift for mobile (lower is better).",
        calculation="Pass-through from CrUX/PSI field data for this URL.",
    ),
    "Mobile TTFB (s)": format_help_layer(
        description="Field/CrUX Time to First Byte for mobile, in seconds (lower is better).",
        calculation="Pass-through from CrUX/PSI field data for this URL.",
    ),
    "GSC Last Crawl": format_help_layer(
        description="Date Google last crawled this URL, per Search Console URL Inspection.",
        calculation="Pass-through from GSC URL Inspection data when available.",
    ),
    "GSC Coverage Category": format_help_layer(
        description="Search Console's indexing coverage bucket for this URL (e.g. Indexed, Excluded, Error).",
        calculation="Pass-through from GSC coverage data when available.",
    ),
    "Discovered On URL": format_help_layer(
        description="The page this URL was first discovered on during the crawl.",
        calculation="Recorded at BFS discovery time from the referring page's link.",
    ),
    "Discovery Rank": format_help_layer(
        description="Crawl discovery order (lower = found earlier).",
        calculation="Sequential counter assigned as URLs are discovered during the BFS crawl.",
    ),
    "Reachable from Homepage": format_help_layer(
        description="Whether this URL is reachable from the homepage via internal links.",
        calculation="Derived from the internal link graph's reachability analysis.",
    ),
    "Crawl Depth": format_help_layer(
        description="Number of clicks from the homepage to reach this URL.",
        calculation="Shortest-path depth in the internal link graph from the homepage.",
    ),
    "Security: HSTS": format_help_layer(
        description="Whether this URL sends a Strict-Transport-Security header.",
        calculation="Boolean presence check on the HTTP response headers.",
    ),
    "Security: CSP": format_help_layer(
        description="Whether this URL sends a Content-Security-Policy header.",
        calculation="Boolean presence check on the HTTP response headers.",
    ),
    "Hreflang Signals": format_help_layer(
        description="Summary of hreflang annotations found on this URL, if any.",
        calculation="Parsed from <link rel=hreflang> tags/headers during extraction.",
    ),
    "Hreflang Declared Languages": format_help_layer(
        description="Language/region codes this URL declares hreflang alternates for.",
        calculation="Parsed from hreflang annotations during extraction.",
    ),
    "Hreflang Alternate URLs": format_help_layer(
        description="The alternate-language URLs this page's hreflang tags point to.",
        calculation="Parsed from hreflang annotations during extraction.",
    ),
    "Hreflang Reciprocal Status": format_help_layer(
        description="Whether this URL's hreflang alternates link back reciprocally.",
        calculation="Cross-checked against each alternate URL's own hreflang annotations.",
    ),
    "Hreflang Code Valid": format_help_layer(
        description="Whether this URL's hreflang language/region codes are valid ISO codes.",
        calculation="Validated against known ISO 639-1/3166-1 code lists.",
    ),
}

_LINK_INTELLIGENCE_HELP: dict[str, str] = {
    "Internal PageRank": format_help_layer(
        description=(
            "Directed-graph authority proxy: higher means more internal PageRank-style flow "
            "from other crawled pages."
        ),
        calculation=(
            "``networkx.pagerank`` on the internal-link DiGraph (α=0.85, max_iter=100); values "
            "rounded to 6 decimals in ``compute_internal_link_intelligence``."
        ),
    ),
    "Orphan Candidate": format_help_layer(
        description=(
            "Whether this URL behaves like an orphan in the internal graph—hard for crawlers "
            "to discover via internal paths."
        ),
        calculation=(
            "Summary rows: boolean ``Orphan Pages`` from the graph (in-degree 0 on crawled "
            "nodes). Graph rows: boolean ``Orphan Candidate`` from CrawlGraph export. "
            "Detail rows force FALSE because anchors are not graph nodes."
        ),
    ),
    # Folded in from the former standalone "Link Inventory" sheet — populated on
    # Detail rows only.
    "Generic Anchor": format_help_layer(
        description='Whether the anchor text is non-descriptive (e.g. "click here", "read more").',
        calculation="Boolean flag from anchor-text analysis on the outbound link.",
    ),
    "Rel Attribute": format_help_layer(
        description="The link's rel value (e.g. nofollow, sponsored, ugc) or blank.",
        calculation="Captured verbatim from the anchor element during extraction.",
    ),
    "Link Type": format_help_layer(
        description="Internal vs external classification for the link edge.",
        calculation="Derived from comparing the target host with the crawl domain.",
    ),
    # Folded in from the former standalone "Link Equity Map" sheet — populated on
    # Summary rows only.
    "Inbound Link Count": format_help_layer(
        description="Raw count of internal anchor instances pointing at this URL.",
        calculation=(
            "Same raw, non-deduplicated counting method as Main!Inbound Internal Link Count — "
            "see that column's tooltip for why this differs from this sheet's own Detail rows' "
            "per-(source, target, anchor) row count."
        ),
    ),
}

_FIXPLAN_HELP: dict[str, str] = {
    "Priority Score": format_help_layer(
        description="Remediation ranking (higher = fix sooner) blending severity and reach.",
        calculation=(
            "Computed in the pipeline from issue severity, affected-URL count, and business "
            "risk; persisted on the FixPlan row (not recalculated in Excel)."
        ),
    ),
    "Affected Link Instances": format_help_layer(
        description="Count of broken plus unresolved internal link instances for this issue/URL.",
        calculation="Sum of broken + unresolved internal links per URL from link analysis.",
    ),
    "Est. Sprint Points": format_help_layer(
        description="Relative delivery effort for this fix expressed in sprint points.",
        calculation="Mapped from the estimated hours band during FixPlan assembly.",
    ),
    "Aging/Priority": format_help_layer(
        description="Recommended sprint bucket for the fix.",
        calculation=(
            'One of "Immediate (Current Sprint)", "Next Sprint", or "Backlog", derived from '
            "Priority Score thresholds."
        ),
    ),
    "Action Needed": format_help_layer(
        description="Whether this row still requires work (Yes) or is clear (No).",
        calculation="Yes/No flag set during FixPlan assembly from the resolution state.",
    ),
    "Issue Type": format_help_layer(
        description="The rule/issue this row groups fixes for — one row per issue type, not per URL.",
        calculation="Rule name from the active IssueRule registry that matched one or more crawled URLs.",
    ),
    "Severity": format_help_layer(
        description="Critical, Warning, or Observation — how urgently this issue needs fixing.",
        calculation="Severity assigned to the matching rule in the IssueRule registry.",
    ),
    "Affected Count": format_help_layer(
        description="Number of crawled URLs this issue was found on.",
        calculation="Count of rows matching this rule at export time.",
    ),
    "URL": format_help_layer(
        description="A sample affected URL for this issue — see \"Affected URLs\" for the full list.",
        calculation="First matching row's URL, in discovery order.",
    ),
    "Affected URLs": format_help_layer(
        description=(
            "Every URL this issue affects, one per line. Column is deliberately not wrapped "
            "(text clips) so row height stays fixed — widen the column or open the cell to "
            "read the full list."
        ),
        calculation="All matching URLs newline-joined; truncated only at Excel's per-cell character limit.",
    ),
    "What It Is": format_help_layer(
        description="Plain-language explanation of the issue for non-technical stakeholders.",
        calculation="Playbook entry for this issue name, falling back to the root-cause text.",
    ),
    "Recommended Fix": format_help_layer(
        description="The concrete remediation step for this issue.",
        calculation="Resolved from the issue's root-cause/fix mapping in the rules layer.",
    ),
    "Resolution Type": format_help_layer(
        description=(
            "How the fix is typically delivered: Server Config, Site Config, Global Template "
            "(one change fixes every affected URL), or Manual Content (per-page edits)."
        ),
        calculation="Classified from the rule's scope and issue-name keywords, or URL fan-out (>10 URLs).",
    ),
    "Likely Root Cause": format_help_layer(
        description="The underlying technical/content cause behind this issue.",
        calculation="Resolved from the issue's root-cause mapping in the rules layer.",
    ),
    "Owner": format_help_layer(
        description="Suggested team to action this fix: Dev, Copy Writer, or Server/Host.",
        calculation="Mapped from issue name and severity in the rules layer.",
    ),
    "Effort": format_help_layer(
        description="Rough delivery effort band: S (small), M (medium), or L (large).",
        calculation="Classified by issue class (config fix, performance fix, content fix, etc.).",
    ),
    "Est. Hours": format_help_layer(
        description="Estimated hours to deliver this fix.",
        calculation="Mapped from the Effort band and affected-URL count.",
    ),
    "Status": format_help_layer(
        description="Workflow state: To Do, In Review, or Done. Editable — update as work progresses.",
        calculation="Seeded from severity at export (Critical/Warning → To Do, Observation → In Review).",
    ),
    "Hub Status (Content Hub)": format_help_layer(
        description="This issue's matching URL's status on the Content Optimisation Hub, if tracked there.",
        calculation="Live lookup (INDEX/MATCH) against the Content Optimisation Hub's Status column.",
    ),
    "Revenue Risk": format_help_layer(
        description="Business-impact framing: High Risk, Medium Risk, or Monitor.",
        calculation="Derived from severity and Priority Score at export time.",
    ),
    "Category": format_help_layer(
        description="Whether this issue is a traditional SEO fix or an AEO (Answer Engine) fix.",
        calculation="Set from a fixed list of AEO-specific issue names; everything else is SEO.",
    ),
    "Discovery Rank": format_help_layer(
        description="Crawl discovery order of the first affected URL (lower = found earlier).",
        calculation="Discovery Rank of the first matching row, in crawl BFS order.",
    ),
    "Detail Reference Tab": format_help_layer(
        description="The workbook tab with full per-URL detail for this issue — click to jump there.",
        calculation="Resolved from the issue category (Technical, Indexability, Links, or AEO).",
    ),
    "Jump to Details": format_help_layer(
        description="Shortcut link to this issue's detail tab (same target as Detail Reference Tab).",
        calculation="Hyperlinked to the resolved detail tab's first cell.",
    ),
    "Jump to Playbook": format_help_layer(
        description="Jumps to this issue's entry in the Playbook tab for full guidance.",
        calculation="HYPERLINK + MATCH formula against Playbook's Item column.",
    ),
    "Verified By": format_help_layer(
        description="Manual field — who confirmed the fix was applied and verified.",
        calculation="Blank at export; filled in by hand during workflow.",
    ),
    "Date Resolved": format_help_layer(
        description="Manual field — when this issue was marked resolved.",
        calculation="Blank at export; filled in by hand during workflow.",
    ),
    "Sprint": format_help_layer(
        description="Manual field — assign a sprint/iteration label for planning.",
        calculation="Blank at export; filled in by hand during workflow.",
    ),
    "Open in Main": format_help_layer(
        description="Jumps to this URL's row on the Main tab for full crawl detail.",
        calculation="HYPERLINK + MATCH formula against Main's URL column.",
    ),
}

_QUICK_WINS_HELP: dict[str, str] = {
    "Priority Score": format_help_layer(
        description=(
            "Quick-win ranking (higher = tackle sooner) blending severity, effort, traffic, "
            "and business risk."
        ),
        calculation=(
            "Severity weight (Critical=100, Warning=50, Observation=10) divided by effort "
            "hours, plus GSC clicks ÷ 10, plus Business Risk Score ÷ 100; computed during "
            "Quick Wins assembly (not recalculated in Excel)."
        ),
    ),
    "Effort (hrs)": format_help_layer(
        description="Estimated hands-on hours to complete this quick win.",
        calculation="Estimated during Quick Wins assembly from the issue's fix profile.",
    ),
    "Business Risk Score": format_help_layer(
        description="Relative business exposure of leaving this issue unresolved (higher = worse).",
        calculation="Blends severity with traffic/visibility signals in the pipeline.",
    ),
    "GSC Clicks (30d)": format_help_layer(
        description="Search Console clicks for this URL over the trailing 30 days.",
        calculation="Pass-through from GSC performance data; blank when GSC is unavailable.",
    ),
    "Revenue Risk": format_help_layer(
        description="Qualitative revenue exposure flag for the affected page.",
        calculation="Derived from page intent and traffic during Quick Wins assembly.",
    ),
    "Issue": format_help_layer(
        description="The rule that matched this URL — one row per URL+issue combination.",
        calculation="Rule name from the active IssueRule registry, coloured by this row's Severity.",
    ),
    "Severity": format_help_layer(
        description="Critical, Warning, or Observation — how urgently this fix is needed.",
        calculation="Severity assigned to the matching rule in the IssueRule registry.",
    ),
    "Owner": format_help_layer(
        description="Suggested team to action this fix: Dev, Copy Writer, or Server/Host.",
        calculation="Pulled from this issue's FixPlan row.",
    ),
    "What It Is": format_help_layer(
        description="Plain-language explanation of the issue for non-technical stakeholders.",
        calculation="Playbook entry for this issue name.",
    ),
    "Why It Matters": format_help_layer(
        description="Why this issue matters — the SEO/AEO or business rationale.",
        calculation="Playbook entry for this issue name.",
    ),
    "Recommended Fix": format_help_layer(
        description="The concrete remediation step for this issue.",
        calculation="Pulled from this issue's FixPlan row.",
    ),
    "How To Verify": format_help_layer(
        description="How to confirm the fix worked once applied.",
        calculation="Playbook entry for this issue name.",
    ),
    "Jump to FixPlan": format_help_layer(
        description="Jumps to this issue's row on the FixPlan tab.",
        calculation="HYPERLINK + MATCH formula against FixPlan's Issue Type column.",
    ),
    "Jump to Playbook": format_help_layer(
        description="Jumps to this issue's entry on the Playbook tab for full guidance.",
        calculation="HYPERLINK + MATCH formula against Playbook's Item column.",
    ),
    "Open in Main": format_help_layer(
        description="Jumps to this URL's row on the Main tab for full crawl detail.",
        calculation="HYPERLINK + MATCH formula against Main's URL column.",
    ),
}

_PRIORITY_URLS_HELP: dict[str, str] = {
    "Business Risk Score": format_help_layer(
        description="Relative urgency ranking that sorts this tab (higher = fix sooner).",
        calculation=(
            "Critical Issues Count x 30 + Warning Issues Count x 10 + "
            "(100 - SEO Health Score); computed at export time, not a workbook formula."
        ),
    ),
    "Severity Badge": format_help_layer(
        description="Worst issue severity found on this URL (Critical > Warning > Observation > Pass).",
        calculation=(
            "Presence-based: set to the highest severity tier with at least one matched "
            "issue rule from score_url_health. Unmeasured when Extraction State is not scorable."
        ),
    ),
    "Revenue Intent": format_help_layer(
        description="High when this page looks commercially important — worth prioritising fixes here.",
        calculation=(
            "High when the URL matches a configured high-value slug, OR Search Intent is "
            "Transactional/Commercial Investigation, OR GSC Impressions sit in this crawl's "
            "top quartile; otherwise Standard."
        ),
    ),
    "Why Prioritized": format_help_layer(
        description="Plain-language reasons this URL made the priority list.",
        calculation=(
            "Joined from: has critical issues, broken internal links, cross-canonical, "
            "noindex — whichever apply; \"Monitor\" when none do."
        ),
    ),
    "Owner": format_help_layer(
        description="Suggested team to action the top issue on this URL (Dev/Copy Writer/SEO Lead).",
        calculation="Mapped from the highest-severity matched issue's category.",
    ),
    "Status": format_help_layer(
        description="Editable — track triage here. The tool seeds \"Open\" and never overwrites your edits on re-export of this workbook copy.",
        calculation="Manual field; not recalculated.",
    ),
    "Action Needed": format_help_layer(
        description="Yes when this URL's Business Risk Score crosses the action threshold.",
        calculation="Yes when Business Risk Score >= 30, else No.",
    ),
    "SEO Health Score": format_help_layer(
        description="Overall 0–100 health score for this URL; blank when Extraction State is not scorable.",
        calculation="Pass-through from the crawl's scoring pipeline.",
    ),
    "Critical Issues Count": format_help_layer(
        description="Number of Critical-severity issues matched on this URL.",
        calculation="Count from score_url_health's matched-issue list.",
    ),
    "Warning Issues Count": format_help_layer(
        description="Number of Warning-severity issues matched on this URL.",
        calculation="Count from score_url_health's matched-issue list.",
    ),
    "GSC Impressions": format_help_layer(
        description="Search Console impressions for this URL over the reporting window.",
        calculation="Pass-through from GSC performance data; blank when GSC is unavailable.",
    ),
    "GSC CTR": format_help_layer(
        description="Search Console click-through rate for this URL.",
        calculation="Pass-through from GSC performance data, rounded to 4 decimal places.",
    ),
    "Indexability Reason": format_help_layer(
        description="Why this URL is or isn't indexable (canonical, robots, status signals).",
        calculation="Rule- and signal-driven text from the indexability pipeline.",
    ),
    "Broken Internal Links Count": format_help_layer(
        description="Number of broken internal links found pointing at or from this URL.",
        calculation="Pass-through from link analysis.",
    ),
}

_BROKEN_LINK_IMPACT_HELP: dict[str, str] = {
    "Priority Score": format_help_layer(
        description=(
            "Broken-destination ranking (higher = fix sooner) by inbound link volume and "
            "source-page traffic."
        ),
        calculation=(
            "Source Page Clicks Total plus Inbound Link Count × 10; computed during Broken "
            "Link Impact assembly (not recalculated in Excel)."
        ),
    ),
    "Inbound Link Count": format_help_layer(
        description="How many internal links point at this broken destination.",
        calculation="Count of internal link edges resolving to this broken target URL.",
    ),
    "Source Page Clicks Total": format_help_layer(
        description="Combined GSC clicks of the pages that link to this broken URL.",
        calculation="Sum of trailing-30-day GSC clicks across the linking source pages.",
    ),
    "Recommended Action": format_help_layer(
        description="Suggested remediation for the broken destination.",
        calculation="Heuristic from status code and link context (fix, redirect, or remove).",
    ),
    "Broken URL": format_help_layer(
        description="The broken destination URL itself — not the page(s) linking to it.",
        calculation="Aggregation key: one row per unique broken internal target URL.",
    ),
    "Status Code": format_help_layer(
        description="The HTTP status the broken destination itself returned (404, timeout, etc.).",
        calculation="Status code observed on the first internal link seen pointing at this target.",
    ),
    "Source Pages (first 5)": format_help_layer(
        description="Up to 5 pages that link to this broken destination.",
        calculation="First 5 unique source URLs from the link inventory, pipe-separated.",
    ),
    "Anchor Texts Used": format_help_layer(
        description="Up to 5 distinct anchor texts used to link to this broken destination.",
        calculation="First 5 unique anchor texts from the link inventory, pipe-separated.",
    ),
}

_MAIN_HELP: dict[str, str] = {
    "Inbound Internal Link Count": format_help_layer(
        description="Raw count of internal anchor instances pointing at this URL.",
        calculation=(
            "Every anchor-tag occurrence across the crawl is counted with no deduplication — "
            "identical desktop/mobile/footer nav markup to the same target all add up here. "
            "This is a different, larger number by design than Link Intelligence's Detail rows' "
            "per-(source, target, anchor) count, which collapses duplicate link instances into "
            "one row. Neither is 'wrong'; they answer different questions (raw link-instance "
            "volume vs. unique link relationships)."
        ),
    ),
    "Extraction State": format_help_layer(
        description="How much of this row could be measured: complete, partial, or skipped.",
        calculation="Set by the crawl/extraction pipeline based on fetch and parse outcomes.",
    ),
    "Title Length": format_help_layer(
        description="Character count of the page title. Target band: 50-60 characters.",
        calculation="len() of the extracted <title> text.",
    ),
    "Robots.txt: Googlebot": format_help_layer(
        description="Whether Googlebot may crawl this URL per robots.txt (Allow/Disallow/Not specified).",
        calculation="urllib.robotparser.can_fetch(\"Googlebot\", url) against this domain's robots.txt.",
    ),
    "Robots.txt: Bingbot": format_help_layer(
        description="Whether Bingbot may crawl this URL per robots.txt (Allow/Disallow/Not specified).",
        calculation="urllib.robotparser.can_fetch(\"Bingbot\", url) against this domain's robots.txt.",
    ),
    "Robots.txt: GPTBot": format_help_layer(
        description="Whether OpenAI's GPTBot may crawl this URL per robots.txt (Allow/Disallow/Not specified).",
        calculation="urllib.robotparser.can_fetch(\"GPTBot\", url) against this domain's robots.txt.",
    ),
    "Robots.txt: ClaudeBot": format_help_layer(
        description="Whether Anthropic's ClaudeBot may crawl this URL per robots.txt (Allow/Disallow/Not specified).",
        calculation="urllib.robotparser.can_fetch(\"ClaudeBot\", url) against this domain's robots.txt.",
    ),
    "Robots.txt: PerplexityBot": format_help_layer(
        description="Whether PerplexityBot may crawl this URL per robots.txt (Allow/Disallow/Not specified).",
        calculation="urllib.robotparser.can_fetch(\"PerplexityBot\", url) against this domain's robots.txt.",
    ),
    "Robots.txt: CCBot": format_help_layer(
        description="Whether Common Crawl's CCBot may crawl this URL per robots.txt (Allow/Disallow/Not specified).",
        calculation="urllib.robotparser.can_fetch(\"CCBot\", url) against this domain's robots.txt.",
    ),
    "Crawl-Delay Applies": format_help_layer(
        description="Whether robots.txt sets a Crawl-delay directive for Googlebot on this domain.",
        calculation="Parsed from this domain's robots.txt Crawl-delay directive.",
    ),
}

_SITEMAPQA_HELP: dict[str, str] = {
    "Found via Crawl": format_help_layer(
        description="Whether the URL was discovered by crawling internal links.",
        calculation="Boolean membership check against the crawl frontier.",
    ),
    "Found via Sitemap": format_help_layer(
        description="Whether the URL appears in an XML sitemap.",
        calculation="Boolean membership check against parsed sitemap entries.",
    ),
    "In Sitemap but Non-200": format_help_layer(
        description="Sitemap URL that did not return HTTP 200 (wastes crawl budget).",
        calculation="True when a sitemap URL's resolved status code is not 200.",
    ),
    "Crawled but Missing from Sitemap": format_help_layer(
        description="Indexable crawled URL absent from any sitemap (discoverability gap).",
        calculation="True when a crawled 200 URL is not present in sitemap entries.",
    ),
    "Lastmod vs HTTP Match": format_help_layer(
        description="Whether the sitemap <lastmod> agrees with the HTTP Last-Modified header.",
        calculation=(
            "Date comparison (day precision) between sitemap lastmod and HTTP header. "
            "A 'Mismatch' here is often expected, not a bug: many servers set a dynamic "
            "Last-Modified header to the response/crawl timestamp on every request, while "
            "the sitemap's <lastmod> reflects the actual content-edit date — these are two "
            "genuinely different signals, not two measurements of the same thing."
        ),
    ),
}

_CONTENT_HUB_SEMANTIC_HELP: dict[str, str] = {
    "Entity Density (%)": format_help_layer(
        description=(
            "The percentage of page content identified as named entities (people, orgs, places). "
            "Healthy pages usually sit in low single digits; the value is sanity-capped at 50%."
        ),
        calculation=(
            "(Unique named entities / total words) × 100 from semantic extraction "
            "(spaCy NER, or proper-noun/acronym fallback). Stored on a 0–100 scale "
            "and displayed with a literal % suffix."
        ),
    ),
    "Top Entities": format_help_layer(
        description="The most frequent named entities on the page (semantic relevance signal).",
        calculation="Ranked entity list from the semantic analyser for this URL.",
    ),
    "Citation Candidate Count": format_help_layer(
        description=(
            "Count of 25–90 word blocks suitable for answer-engine citation. "
            "0 means the page lacks citeable definition/answer blocks — add a "
            "25–90 word paragraph that defines or directly answers something."
        ),
        calculation=(
            "Blocks containing a definition trigger ('is', 'means', 'provides', "
            "'how to', …) or opening with a question followed by a ≥15-word "
            "answer, from the semantic pass."
        ),
    ),
    "Semantic AEO Score": format_help_layer(
        description="Weighted 0–100 score from entity density and citation readiness.",
        calculation="Composite semantic/AEO score from the crawl enrichment pass.",
    ),
}

_PLAYBOOK_HELP: dict[str, str] = {
    "Section": format_help_layer(
        description=(
            "Groups rows into a topic block — editorial standard, issue playbook entry, "
            "or glossary/legend item. Bracketed values (e.g. \"[Meta Data Standards]\") "
            "mark a block's header row; each block is tinted its own colour."
        ),
        calculation="Fixed reference copy plus per-issue rows from the active rule set.",
    ),
    "Item": format_help_layer(
        description="The specific standard, issue name, or glossary term this row documents.",
        calculation="Blank on section-header rows; populated on every content row beneath them.",
    ),
    "Guideline": format_help_layer(
        description="The concrete rule, threshold, or fix guidance for this item.",
        calculation=(
            "Static editorial copy for standards rows; What/Fix/Verify guidance for "
            "Issue Playbook rows; threshold value for glossary rows."
        ),
    ),
    "Why It Matters": format_help_layer(
        description="Why this standard or fix matters — the SEO/AEO or business rationale.",
        calculation=(
            "Static editorial copy for standards rows; issue rationale plus severity/owner/"
            "time-to-fix context for Issue Playbook rows; plain-language meaning for glossary rows."
        ),
    ),
}

_CONTENT_PLANNER_HELP: dict[str, str] = {
    "Primary": format_help_layer(
        description="Top-level nav/site-structure label for this URL (path depth <= 1).",
        calculation="Last URL path segment when the path has 0 or 1 segments.",
    ),
    "Secondary": format_help_layer(
        description="Second-level nav/site-structure label for this URL (path depth == 2).",
        calculation="Last URL path segment when the path has exactly 2 segments.",
    ),
    "Tertiary": format_help_layer(
        description="Third-level-or-deeper nav/site-structure label for this URL (path depth >= 3).",
        calculation="Last URL path segment when the path has 3 or more segments.",
    ),
    "Page link": format_help_layer(
        description="The crawled page itself — click to open it.",
        calculation="Raw crawled URL.",
    ),
    "Copy Doc": format_help_layer(
        description="Editable — paste a link to this page's copy document.",
        calculation="Manual field; seeded with a placeholder hint.",
    ),
    "Priority for MVP": format_help_layer(
        description="Editable — flag whether this page is required for MVP launch.",
        calculation="Manual field; not recalculated.",
    ),
    "Copywriter Sign off": format_help_layer(
        description="Editable workflow status: Not signed off / In progress / Signed off.",
        calculation="Manual field; coloured red/amber/green by status.",
    ),
    "Copy First Check": format_help_layer(
        description="Editable workflow status for the first editorial pass on this page's copy.",
        calculation="Manual field; coloured red/amber/green by status.",
    ),
    "2nd Revisions": format_help_layer(
        description="Editable workflow status for the second round of copy revisions.",
        calculation="Manual field; coloured red/amber/green by status.",
    ),
    "Client copy sign off": format_help_layer(
        description="Editable workflow status for client sign-off on the copy.",
        calculation="Manual field; coloured red/amber/green by status.",
    ),
    "Web design off": format_help_layer(
        description="Editable workflow status for web design sign-off on this page.",
        calculation="Manual field; coloured red/amber/green by status.",
    ),
    "UXI sign off": format_help_layer(
        description="Editable workflow status for UX/interaction design sign-off.",
        calculation="Manual field; coloured red/amber/green by status.",
    ),
    "Visual Design sign off": format_help_layer(
        description="Editable workflow status for visual design sign-off.",
        calculation="Manual field; coloured red/amber/green by status.",
    ),
    "Client final sign off": format_help_layer(
        description="Editable workflow status for the client's final sign-off on this page.",
        calculation="Manual field; coloured red/amber/green by status.",
    ),
    "Optimisations": format_help_layer(
        description="Editable workflow status for post-launch on-page optimisation work.",
        calculation="Manual field; coloured red/amber/green by status.",
    ),
    "Desktop": format_help_layer(
        description="Editable workflow status for desktop-layout QA on this page.",
        calculation="Manual field; coloured red/amber/green by status.",
    ),
    "Tablet": format_help_layer(
        description="Editable workflow status for tablet-layout QA on this page.",
        calculation="Manual field; coloured red/amber/green by status.",
    ),
    "Mobile": format_help_layer(
        description="Editable workflow status for mobile-layout QA on this page.",
        calculation="Manual field; coloured red/amber/green by status.",
    ),
    "SEO": format_help_layer(
        description="Editable workflow status for on-page SEO QA on this page.",
        calculation="Manual field; coloured red/amber/green by status.",
    ),
    "Performance": format_help_layer(
        description="Editable workflow status for performance QA on this page.",
        calculation="Manual field; coloured red/amber/green by status.",
    ),
    "Plugin Audit": format_help_layer(
        description="Editable — track whether this page's plugin/shortcode compatibility has been checked.",
        calculation="Manual field; coloured red/amber/green by status.",
    ),
}

_ROBOTS_ANALYSIS_HELP: dict[str, str] = {
    "Section": format_help_layer(
        description=(
            "Which part of the analysis this row belongs to: 1 (raw robots.txt file per "
            "domain), 2 (parsed user-agent/directive rules), 3 (crawled URLs blocked for a "
            "monitored bot), or 4 (sitemap/robots.txt mismatches)."
        ),
        calculation="Fixed section label assigned when the row is built.",
    ),
    "User Agent": format_help_layer(
        description=(
            "The bot this row concerns — a domain (Section 1), a directive's target agent "
            "(Section 2), or a monitored bot that was blocked (Section 3)."
        ),
        calculation="Parsed from the User-agent: line, or the monitored-bot list for Section 3.",
    ),
    "URL": format_help_layer(
        description="The robots.txt file location (Section 1) or the affected crawled URL (Section 3/4).",
        calculation="Domain + /robots.txt, or the crawled row's URL.",
    ),
    "Status": format_help_layer(
        description=(
            "Meaning depends on Section: Accessible/Unavailable/Fetched-body-unreadable "
            "(1), the directive name — user-agent/disallow/allow/crawl-delay/sitemap (2), "
            "Disallow/None (3), or the specific mismatch found/None (4)."
        ),
        calculation="Set per-row according to that row's Section.",
    ),
    "Detail": format_help_layer(
        description=(
            "The supporting text for this row: the raw robots.txt body (Section 1, up to "
            "32,000 characters — clipped/wrapped, not paginated), the directive's value "
            "(Section 2), or a short explanation (Section 3/4)."
        ),
        calculation="Pass-through from the fetched robots.txt or the row's own template text.",
    ),
    "Explanation": format_help_layer(
        description="Plain-language rationale for this row — which check matched and why.",
        calculation="Generated per-row from the specific rule/agent/directive that produced it.",
    ),
}

_SHEET_CURATED_HEADER_HELP: dict[str, dict[str, str]] = {
    "Content & AI Readiness": _CONTENT_AI_READINESS_HELP,
    "Playbook": _PLAYBOOK_HELP,
    "Content Planner": _CONTENT_PLANNER_HELP,
    "Robots.txt Analysis": _ROBOTS_ANALYSIS_HELP,
    "Technical Diagnostics": _TECHNICAL_DIAGNOSTICS_HELP,
    "Link Intelligence": _LINK_INTELLIGENCE_HELP,
    CONTENT_OPTIMISATION_HUB_SHEET: _CONTENT_HUB_SEMANTIC_HELP,
    "Priority URLs": _PRIORITY_URLS_HELP,
    "FixPlan": _FIXPLAN_HELP,
    "Quick Wins": _QUICK_WINS_HELP,
    "Broken Link Impact": _BROKEN_LINK_IMPACT_HELP,
    "SitemapQA": _SITEMAPQA_HELP,
    "Main": _MAIN_HELP,
}

SCHEMA_METADATA_HEADER_TOOLTIP_BODIES: dict[str, str] = {
    "TTFB (ms)": format_help_layer(
        description="Time to First Byte in milliseconds: server/network responsiveness.",
        calculation=(
            "Measured during the HTTP fetch for this URL and stored on the row (not derived "
            "in Excel)."
        ),
    ),
    "AEO Readiness Score": _CONTENT_AI_READINESS_HELP["AEO Readiness Score"],
    "Indexability Reason": format_help_layer(
        description="Primary explanation when a URL is treated as non-indexable or risky.",
        calculation=(
            "Rule- and signal-driven text from the indexability pipeline (canonical, robots, "
            "status); exported as a string field."
        ),
    ),
    "Status Code": format_help_layer(
        description="Final HTTP status observed for this URL during the crawl.",
        calculation=(
            "Integer from the HTTP client response stored on the crawl row; Technical sheets "
            "reference that stored value."
        ),
    ),
    "SEO Health Score": _TECHNICAL_DIAGNOSTICS_HELP["SEO Health Score"],
    "Priority Score": format_help_layer(
        description="Relative urgency for FixPlan-style remediation rows.",
        calculation=(
            "``workflow_metrics_for_issue``: severity base (Critical=100, Warning=65, "
            "Observation=35, default=25) plus sprint points from effort band "
            "(S=2, M=5, L=8, default=2)."
        ),
    ),
    "Severity": format_help_layer(
        description="Impact tier for the issue attached to this row.",
        calculation=(
            "Categorical label from rules (Critical / Warning / Observation / etc.); "
            "drives SEO Health Score penalties when aggregated per URL."
        ),
    ),
    "Word Count": format_help_layer(
        description="Approximate visible body word count used for thin-content signals.",
        calculation=(
            "Tokenised word count from extracted main content at crawl time; stored numeric "
            "on the row."
        ),
    ),
    "Canonical Type": format_help_layer(
        description="How the declared canonical relates to the crawled URL.",
        calculation=(
            "Classifier output (self / cross-canonical / missing) from canonical tag "
            "analysis; string field on the export row."
        ),
    ),
    "Redirect Chain Length": format_help_layer(
        description="Count of hops before the crawler reached the final URL.",
        calculation=(
            "Integer from redirect tracing during fetch; stored on the row (lower is better)."
        ),
    ),
}


def curated_help_keys_by_sheet() -> dict[str, frozenset[str]]:
    """Canonical header keys that have curated help (for contract tests vs layout tuples)."""
    return {name: frozenset(body.keys()) for name, body in _SHEET_CURATED_HEADER_HELP.items()}


def resolve_curated_help_body(sheet_title: str, header: str) -> str | None:
    """Return curated help text when this sheet/header pair is in the registry.

    Matching is case-sensitive on the stored header text (Excel headers are stable literals).
    Content & AI also maps any header starting with ``Flesch-Kincaid Grade`` to the FK help.
    """
    sheet = (sheet_title or "").strip()
    h = (header or "").strip()
    mapping = _SHEET_CURATED_HEADER_HELP.get(sheet)
    if not mapping:
        return None
    if h in mapping:
        return mapping[h]
    if sheet == "Content & AI Readiness" and h.startswith("Flesch-Kincaid Grade"):
        return mapping.get("Flesch-Kincaid Grade (Est.)")
    return None


def friendly_metric_label(header: str) -> str:
    """Return a user-facing metric label used in validation prompt titles.

    Args:
        header: Raw worksheet header text.

    Returns:
        Cleaned and presentation-friendly label text.
    """
    return (
        header.replace("_", " ")
        .replace("URL", "URL")
        .replace("SEO", "SEO")
        .replace("AEO", "AEO")
        .strip()
    )


def tooltip_for_header(header: str) -> str:
    """Generate contextual tooltip guidance for a header (Description + Calculation)."""
    h = (header or "").strip()
    lower = h.lower()
    label = friendly_metric_label(h)
    if not h:
        return format_help_layer(
            description="Placeholder for an unnamed column in this section.",
            calculation="No formula: bind a header text in row 1 so metrics stay traceable.",
        )
    if "url" in lower:
        return format_help_layer(
            description=f"{label} identifies the audited page for drill-down and hyperlinks.",
            calculation="String URL from the crawl/export row; workbook hyperlinks wrap this value where enabled.",
        )
    if "status code" in lower:
        return format_help_layer(
            description="HTTP response code returned for this page or resource.",
            calculation="Integer status from the crawler HTTP layer; compare to Technical Diagnostics for evidence.",
        )
    if "status class" in lower:
        return format_help_layer(
            description="Bucketed HTTP class (2xx/3xx/4xx/5xx) for fast volume triage.",
            calculation="Derived categorisation from the numeric Status Code field in the pipeline or pivot logic.",
        )
    if "health score" in lower:
        return format_help_layer(
            description="Composite SEO quality score for this URL (higher is better).",
            calculation=(
                "When labelled ``SEO Health Score``, matches ``score_url_health`` "
                "(100 − capped diminishing penalties: Critical 20+10/extra cap 50, "
                "Warning 8+5/extra cap 30, Observation 3 each cap 10; floored at 0). "
                "Other *Health* columns may use different pipeline blends—see sheet context."
            ),
        )
    if "pass rate" in lower:
        return format_help_layer(
            description="Share of URLs classified as pass for this slice of the crawl.",
            calculation="Pass URL count ÷ crawl denominator (see Dashboard / Summary formulas for the active definition).",
        )
    if "severity" in lower:
        return format_help_layer(
            description="Issue impact level used to prioritise remediation.",
            calculation="Categorical rule output; when aggregated per URL it feeds SEO Health Score deductions.",
        )
    if "priority score" in lower:
        return format_help_layer(
            description="Relative execution priority combining severity and effort.",
            calculation="``SEVERITY_PRIORITY_BASE[severity] + EFFORT_TO_SPRINT_POINTS[effort]`` (see ``workflow_metrics_for_issue``).",
        )
    if "affected count" in lower:
        return format_help_layer(
            description="How many URLs are impacted by this issue row.",
            calculation="Integer rollup from FixPlan / inventory builders counting matching URLs for the issue.",
        )
    if "ttfb" in lower:
        return format_help_layer(
            description="Time to First Byte: server responsiveness signal.",
            calculation="Milliseconds from the first response byte timestamp in the HTTP fetch metrics.",
        )
    if "load time" in lower or "request time" in lower:
        return format_help_layer(
            description="Elapsed time for the page or request path being measured.",
            calculation="Stopwatch-style timing from the crawler session, stored as a numeric field on the row.",
        )
    if "indexability" in lower or "canonical" in lower:
        return format_help_layer(
            description="Indexing and canonical consistency signal for this URL.",
            calculation="Classifier + directive parser outputs (canonical URL, robots, status) merged into the export row.",
        )
    if "meta robots" in lower or "x-robots" in lower:
        return format_help_layer(
            description="Robots directives that influence crawling and indexing.",
            calculation="Parsed header/tag text from the crawl stored verbatim or normalised on the row.",
        )
    if "word count" in lower or "readability" in lower:
        return format_help_layer(
            description="Content depth / readability signal for editorial quality.",
            calculation="Word / readability metrics computed from extracted visible text at crawl time.",
        )
    if "link" in lower or "anchor" in lower:
        return format_help_layer(
            description="Link graph or anchor-quality signal for crawl equity and clarity.",
            calculation="Aggregated from per-URL link extraction (counts, flags, targets) in the pipeline export.",
        )
    if "open in main" in lower or "technical view" in lower or "view details" in lower:
        return format_help_layer(
            description="Navigation helper to jump to the related record on another tab.",
            calculation="Workbook ``HYPERLINK`` or equivalent jump wiring resolved at export time.",
        )
    if "schema" in lower or "json-ld" in lower:
        return format_help_layer(
            description="Structured-data coverage or validation signal.",
            calculation="JSON-LD / microdata parsers populate typed fields (counts, errors) on the crawl row.",
        )
    if "owner" in lower or "sprint" in lower or lower == "status":
        return format_help_layer(
            description="Workflow field for ownership, cadence, or row state.",
            calculation="Editorial / process value (dropdowns may apply); does not change crawl scores until republished.",
        )
    if "section" in lower or "reference tab" in lower:
        return format_help_layer(
            description="Summary grouping or pointer toward a deeper tab.",
            calculation="String label from Summary / inventory builders for navigation context.",
        )
    return format_help_layer(
        description=f"{label}: auxiliary diagnostic field for this worksheet.",
        calculation="See Glossary & Legend and the crawl export schema for the precise field definition.",
    )


def apply_curated_header_tooltips(
    worksheet: Worksheet,
    sheet_title: str,
    *,
    header_row: int = 1,
    only_headers: frozenset[str] | None = None,
) -> None:
    """Attach curated help comments for ``sheet_title`` (single registry entry point)."""
    if _DISABLE_TOOLTIP_COMMENTS:
        return
    author = "hype-frog"
    for col_idx in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=header_row, column=col_idx)
        header = str(cell.value or "").strip()
        if not header or (only_headers is not None and header not in only_headers):
            continue
        body = resolve_curated_help_body(sheet_title, header)
        if body is None:
            continue
        title = friendly_metric_label(header)[:32] or "Column"
        _attach_header_comment(cell, f"{title}\n\n{body}", author)


def add_all_header_tooltips(worksheet: Worksheet, *, header_row: int = 1) -> None:
    """Attach generic header guidance as cell comments on the header row.

    Cell comments render above the grid and avoid freeze-pane clipping that affects
    Data Validation input messages.

    Args:
        worksheet: Worksheet where comments are added.
        header_row: 1-based row holding column headers — most data sheets carry
            a row-1 "Return to Executive Briefing" banner, so real headers (and
            therefore curated tooltips) live on row 2; passing the wrong row
            silently attaches nothing (see ``sheet_data_header_row``).
    """
    if _DISABLE_TOOLTIP_COMMENTS:
        return
    headers = header_index(worksheet, header_row)
    author = "hype-frog"
    sheet_title = worksheet.title or ""
    for header, col_idx in headers.items():
        cell = worksheet.cell(row=header_row, column=col_idx)
        title = friendly_metric_label(header)[:32] or "Column"
        curated = resolve_curated_help_body(sheet_title, header)
        body = curated if curated is not None else tooltip_for_header(header)
        text = f"{title}\n\n{body}" if title else body
        _attach_header_comment(cell, text, author)


def add_header_tooltips(worksheet: Worksheet) -> None:
    """Attach specialized schema-focused header tooltips.

    Args:
        worksheet: Worksheet where schema prompts are added.
    """
    # Lazy import avoids import cycle: ``schema`` consumes ``SCHEMA_METADATA_HEADER_TOOLTIP_BODIES``.
    from hype_frog.reporter.sheets.schema import add_schema_header_tooltips

    add_schema_header_tooltips(
        worksheet,
        disable_data_validation=_DISABLE_TOOLTIP_COMMENTS,
        header_index_fn=header_index,
    )


def apply_workflow_status_dropdown(
    worksheet: Worksheet,
    status_col: int,
    *,
    header_row: int = 1,
) -> None:
    """Apply unified workflow status list to a ``Status`` column.

    Permitted values: ``STATUS_OPTIONS`` (To Do, In Progress, In Review, Done).
    """
    if DISABLE_DATA_VALIDATION:
        return
    if status_col <= 0 or worksheet.max_row <= header_row:
        return
    data_start = header_row + 1
    status_dv = DataValidation(
        type="list",
        formula1=status_validation_list_formula(),
        allow_blank=True,
    )
    status_dv.showInputMessage = True
    status_dv.promptTitle = "Workflow status"
    status_dv.prompt = "Track workflow state: To Do → In Progress → In Review → Done"
    status_dv.showErrorMessage = True
    status_dv.errorTitle = "Invalid Status"
    status_dv.error = "Select a value from the dropdown list."
    worksheet.add_data_validation(status_dv)
    status_dv.add(
        f"{get_column_letter(status_col)}{data_start}:"
        f"{get_column_letter(status_col)}{worksheet.max_row}"
    )


def apply_status_dropdown(worksheet: Worksheet, status_col: int, *, header_row: int = 1) -> None:
    """Add controlled status dropdown validation for remediation fields."""
    apply_workflow_status_dropdown(worksheet, status_col, header_row=header_row)


def apply_triage_status_dropdown(
    worksheet: Worksheet,
    status_col: int,
    *,
    header_row: int = 1,
) -> None:
    """Apply the lightweight triage status list to a ``Status`` column.

    Permitted values: ``TRIAGE_STATUS_OPTIONS`` (Open, In Progress, Resolved, Won't Fix).
    Distinct from :func:`apply_workflow_status_dropdown` — Priority URLs seeds every
    row "Open" (a triage flag), not the FixPlan/Hub "To Do" workflow.
    """
    if DISABLE_DATA_VALIDATION:
        return
    if status_col <= 0 or worksheet.max_row <= header_row:
        return
    data_start = header_row + 1
    status_dv = DataValidation(
        type="list",
        formula1=triage_status_validation_list_formula(),
        allow_blank=True,
    )
    status_dv.showInputMessage = True
    status_dv.promptTitle = "Triage status"
    status_dv.prompt = "Track triage state: Open → In Progress → Resolved (or Won't Fix)"
    status_dv.showErrorMessage = True
    status_dv.errorTitle = "Invalid Status"
    status_dv.error = "Select a value from the dropdown list."
    worksheet.add_data_validation(status_dv)
    status_dv.add(
        f"{get_column_letter(status_col)}{data_start}:"
        f"{get_column_letter(status_col)}{worksheet.max_row}"
    )


def apply_status_dropdown_to_inventory(
    worksheet: Worksheet, *, header_row: int = 1
) -> None:
    """Apply strict status dropdown validation using the sheet's Status column."""
    if DISABLE_DATA_VALIDATION:
        return
    if worksheet.max_row <= header_row:
        return

    headers = header_index(worksheet, header_row)
    status_col = headers.get("Status")
    if status_col is None:
        return

    apply_workflow_status_dropdown(worksheet, status_col, header_row=header_row)


__all__ = [
    "HELP_CALCULATION_PREFIX",
    "HELP_DESCRIPTION_PREFIX",
    "SCHEMA_METADATA_HEADER_TOOLTIP_BODIES",
    "apply_comment_dimensions",
    "apply_workflow_status_dropdown",
    "apply_status_dropdown",
    "apply_triage_status_dropdown",
    "apply_status_dropdown_to_inventory",
    "STATUS_OPTIONS",
    "add_all_header_tooltips",
    "add_header_tooltips",
    "curated_help_keys_by_sheet",
    "format_help_layer",
    "friendly_metric_label",
    "resolve_curated_help_body",
    "tooltip_for_header",
]
