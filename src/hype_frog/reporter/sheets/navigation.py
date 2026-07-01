from __future__ import annotations

from typing import Any

from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from hype_frog.reporter.sheets.config import (
    CONTENT_OPTIMISATION_HUB_SHEET,
    DEBUG_EXCEL_ISOLATION_MODE,
    DISABLE_EXTERNAL_LINKS_AND_IMAGES,
    RETURN_TO_BRIEFING_LABEL,
    STD_BLUE,
    WORKBOOK_NAV_TARGET_SHEET,
)
from hype_frog.reporter.sheets.links import (
    add_url_navigation_links as add_url_navigation_links_impl,
    apply_cross_sheet_links as apply_cross_sheet_links_impl,
)
from hype_frog.reporter.sheets.sheet_rows import (
    sheet_data_header_row,
    sheet_data_start_row,
    sheet_uses_return_strip,
)
from hype_frog.reporter.sheets.style_helpers import header_row_index
from hype_frog.reporter.sheets.workbook_layout import excel_sheet_link_target

_LEGACY_BACK_NAV_HEADER = "BACK TO DASHBOARD"


def _has_return_strip(worksheet: Worksheet) -> bool:
    value = worksheet.cell(row=1, column=1).value
    return str(value or "").strip() == RETURN_TO_BRIEFING_LABEL


def _remove_legacy_back_to_dashboard_column(
    worksheet: Worksheet, *, header_row: int
) -> None:
    """Drop trailing legacy navigation column when present."""
    headers = header_row_index(worksheet, header_row)
    back_col = headers.get(_LEGACY_BACK_NAV_HEADER)
    if back_col is None:
        return
    if back_col != worksheet.max_column:
        return
    worksheet.delete_cols(back_col)


def add_return_to_briefing_strip(worksheet: Worksheet, sheet_name: str) -> None:
    """Insert row-1 return navigation (blue italic hyperlink, no trailing column)."""
    if DEBUG_EXCEL_ISOLATION_MODE:
        return
    if not sheet_uses_return_strip(sheet_name):
        return
    if _has_return_strip(worksheet):
        return

    header_row = 1
    _remove_legacy_back_to_dashboard_column(worksheet, header_row=header_row)
    worksheet.insert_rows(1)

    # Merge across the sheet's real columns only (capped at 8). Forcing a
    # minimum of 4 columns previously inflated ``max_column`` on narrow sheets
    # (e.g. the 3-column empty FixPlan), which then produced a stray
    # ``Column_4`` header via ``normalize_table_headers``.
    merge_end = get_column_letter(min(max(worksheet.max_column, 1), 8))
    worksheet.merge_cells(f"A1:{merge_end}1")
    cell = worksheet["A1"]
    cell.value = RETURN_TO_BRIEFING_LABEL
    safe_target = excel_sheet_link_target(WORKBOOK_NAV_TARGET_SHEET)
    cell.hyperlink = f"#'{safe_target}'!A1"
    cell.font = Font(color=STD_BLUE, italic=True, underline="single")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 20


def add_back_to_dashboard_link(worksheet: Worksheet, sheet_name: str) -> None:
    """Deprecated: use :func:`add_return_to_briefing_strip` (Phase 3 navigation)."""
    add_return_to_briefing_strip(worksheet, sheet_name)


def add_url_navigation_links(
    writer: Any,
    worksheet: Worksheet,
    sheet_name: str,
    *,
    header_row: int | None = None,
) -> None:
    """Delegate URL/link column generation to link helpers."""
    resolved_header = header_row if header_row is not None else sheet_data_header_row(
        sheet_name
    )
    data_start = resolved_header + 1

    def _header_index_fn(ws: Worksheet) -> dict[str, int]:
        return header_row_index(ws, resolved_header)

    add_url_navigation_links_impl(
        writer,
        worksheet,
        sheet_name,
        debug_excel_isolation_mode=DEBUG_EXCEL_ISOLATION_MODE,
        disable_external_links_and_images=DISABLE_EXTERNAL_LINKS_AND_IMAGES,
        header_index_fn=_header_index_fn,
        data_start_row=data_start,
        header_row=resolved_header,
    )


def apply_cross_sheet_links(
    writer: Any,
    worksheet: Worksheet,
    sheet_name: str,
    *,
    header_row: int | None = None,
) -> None:
    """Delegate cross-sheet link generation with project defaults."""
    resolved_header = header_row if header_row is not None else sheet_data_header_row(
        sheet_name
    )
    data_start = resolved_header + 1

    def _header_index_fn(ws: Worksheet) -> dict[str, int]:
        return header_row_index(ws, resolved_header)

    apply_cross_sheet_links_impl(
        writer,
        worksheet,
        sheet_name,
        debug_excel_isolation_mode=DEBUG_EXCEL_ISOLATION_MODE,
        header_index_fn=_header_index_fn,
        data_start_row=data_start,
        header_row=resolved_header,
    )


__all__ = [
    "add_back_to_dashboard_link",
    "add_return_to_briefing_strip",
    "add_url_navigation_links",
    "apply_cross_sheet_links",
    "sheet_data_header_row",
    "sheet_data_start_row",
    "sheet_uses_return_strip",
]
