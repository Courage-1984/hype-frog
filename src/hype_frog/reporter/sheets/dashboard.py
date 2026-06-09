from __future__ import annotations

from typing import Any

from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from hype_frog.core.models import ExtraRowPayload, MainRowPayload, SummaryMetricsPayload
from hype_frog.reporter.dashboard_logic import (
    FixPlanRowPayload,
    compute_dashboard_metrics,
)
from hype_frog.reporter.narrative_engine import NarrativeEngine, average_seo_score_pct
from hype_frog.reporter.sheets.dashboard_config import (
    ALERT_COLOR,
    DASHBOARD_COLUMN_WIDTHS,
    DASHBOARD_KPI_ROW_COMMENTS,
    DASHBOARD_TOOLTIPS,
    GOOD_COLOR,
    LIGHT_HEADER_COLOR,
    PANEL_BG_COLOR,
    QUICK_LINKS,
    SEVERITY_ROW_STYLE,
    SOFT_ALERT_COLOR,
    SOFT_WARN_COLOR,
    STATUS_ROW_STYLE,
    TABLE_HEADER_COLOR,
    VALUE_BLOCK_COLOR,
    WARN_COLOR,
)
from hype_frog.reporter.sheets.config import (
    AUDIT_RUN_DETAILS_SHEET,
    CONTENT_HUB_METRICS_SHEET,
    CONTENT_OPTIMISATION_HUB_SHEET,
    STD_BLUE,
    STD_NAVY,
)
from hype_frog.reporter.sheets.workbook_layout import (
    DASHBOARD_ADVANCED_SHEET_LINKS,
    DASHBOARD_ADVANCED_SHEETS_NOTE,
    excel_sheet_link_target,
)
from hype_frog.reporter.sheets.style_helpers import header_index, to_int
from hype_frog.reporter.sheets.validation import apply_comment_dimensions
from hype_frog.reporter.sheets.view_state import set_freeze_panes_safe
from hype_frog.pipeline.broken_links import (
    count_broken_internal_instances,
    link_inventory_broken_internal_total_formula,
)
from hype_frog.reporter.sheets.conditional import (
    apply_dashboard_metric_conditional_rules,
)

# Technical Diagnostics URL denominators: see ``_technical_url_row_denominator``.
_MAIN_DYNAMIC_COLUMN = (
    'INDEX(\'Main\'!$1:$1048576,0,MATCH("{header}",\'Main\'!$1:$1,0))'
)
_TD_DYNAMIC_COLUMN = (
    'INDEX(\'Technical Diagnostics\'!$1:$1048576,0,'
    'MATCH("{header}",\'Technical Diagnostics\'!$1:$1,0))'
)
_CONTENT_HUB_METRICS_DYNAMIC_COLUMN = (
    f'INDEX(\'{CONTENT_HUB_METRICS_SHEET}\'!$1:$1048576,0,'
    f'MATCH("{{header}}",\'{CONTENT_HUB_METRICS_SHEET}\'!$1:$1,0))'
)
# Hub data starts row 3 (row 1 banner, row 2 headers); F = Status per preferred column order.
_CONTENT_HUB_STATUS_RANGE = f"'{CONTENT_OPTIMISATION_HUB_SHEET}'!F3:F10000"

DASHBOARD_BRAND_A1 = "🐸 HYPE FROG: SEO & AEO Intelligence Report"

# KPI cells hold 0–1 fractions (or formulas yielding 0–1); Excel ``0%`` displays e.g. 0.6657 as 67%.
_DASHBOARD_PERCENT_KPI_CELLS: tuple[str, ...] = (
    "B5",  # Average SEO Score
    "B6",  # Technical Health
    "B12",  # Crawl Success Rate % (2xx)
    "B15",  # Projected health (Python-aligned)
    "B16",  # Projected pass rate (Python-aligned)
    "B19",  # AEO Opportunity Gap (executive metrics row 19)
)

# Narrative/metrics read merged technical rows from this sheet (legacy alias: "Technical").
_TECHNICAL_SOURCE_SHEET_NAMES: tuple[str, ...] = ("Technical Diagnostics", "Technical")

# Average numeric AEO Extractability column without hard-coded column letters (headers may shift).
_CONTENT_AI_AEO_EXTRACTABILITY_AVG = (
    "IFERROR(AVERAGE(INDEX('Content & AI Readiness'!$1:$1048576,0,"
    'MATCH("AEO Extractability Score",\'Content & AI Readiness\'!$1:$1,0))),0)'
)

# Dashboard KPI: optional external HEAD sniff ratio (values live on Audit Run Details).
_AUDIT_DETAILS_SHEET_ESC = excel_sheet_link_target(AUDIT_RUN_DETAILS_SHEET)
_EXCEL_EXTERNAL_LINK_HEALTH_PCT = (
    f'=IF(IFERROR(INDEX(\'{_AUDIT_DETAILS_SHEET_ESC}\'!$B:$B,MATCH("External Sniff Performed",'
    f"\'{_AUDIT_DETAILS_SHEET_ESC}\'!$A:$A,0)),0)=0,\"\","
    f'IF(IFERROR(INDEX(\'{_AUDIT_DETAILS_SHEET_ESC}\'!$B:$B,MATCH("External Link Unique Denominator",'
    f"\'{_AUDIT_DETAILS_SHEET_ESC}\'!$A:$A,0)),0)=0,\"\","
    f'IFERROR(IFERROR(INDEX(\'{_AUDIT_DETAILS_SHEET_ESC}\'!$B:$B,MATCH("External Link Unique 200 OK",'
    f"\'{_AUDIT_DETAILS_SHEET_ESC}\'!$A:$A,0)),0)/"
    f'IFERROR(INDEX(\'{_AUDIT_DETAILS_SHEET_ESC}\'!$B:$B,MATCH("External Link Unique Denominator",'
    f"\'{_AUDIT_DETAILS_SHEET_ESC}\'!$A:$A,0)),1),0)))"
)

_TD_SHEET = "Technical Diagnostics"


def _technical_diagnostics_data_column(header: str) -> str:
    """Excel OFFSET range for ``header`` on Technical Diagnostics (data rows 2–100000)."""
    return (
        f"OFFSET('{_TD_SHEET}'!$A$1,1,MATCH(\"{header}\",'{_TD_SHEET}'!$1:$1,0)-1,99999,1)"
    )


def _main_sheet_data_column(header: str) -> str:
    """Excel OFFSET range for ``header`` on Main (data rows 2–100000)."""
    return (
        f"OFFSET('Main'!$A$1,1,MATCH(\"{header}\",'Main'!$1:$1,0)-1,99999,1)"
    )


def _technical_url_row_denominator() -> str:
    """Stable URL count for Technical Diagnostics rate denominators (at least 1)."""
    return f"MAX(1,COUNTA({_technical_diagnostics_data_column('URL')}))"


def _safe_float(value: Any, default: float = 0.0) -> float:
    try:
        return float(value or default)
    except (TypeError, ValueError):
        return default


def _technical_sheet_rows(writer: Any) -> list[dict[str, Any]]:
    """Return data rows from Technical Diagnostics (preferred) or legacy Technical sheet."""
    try:
        names = list(writer.book.sheetnames)
    except Exception:
        return []
    for sheet_title in _TECHNICAL_SOURCE_SHEET_NAMES:
        if sheet_title in names:
            return _sheet_rows(writer.book[sheet_title])
    return []


def _sheet_rows(worksheet: Worksheet) -> list[dict[str, Any]]:
    headers = header_index(worksheet)
    ordered = sorted(headers.items(), key=lambda item: item[1])
    rows: list[dict[str, Any]] = []
    for row_idx in range(2, worksheet.max_row + 1):
        row_dict = {
            key: worksheet.cell(row=row_idx, column=column).value
            for key, column in ordered
        }
        rows.append(row_dict)
    return rows


def _format_priority_urls_ctr(writer: Any) -> None:
    """Format Priority URLs CTR values to four decimals for readability."""
    if "Priority URLs" not in writer.book.sheetnames:
        return
    priority_ws = writer.book["Priority URLs"]
    headers = header_index(priority_ws)
    ctr_col = headers.get("GSC CTR")
    if not ctr_col or priority_ws.max_row < 2:
        return
    for row_idx in range(2, priority_ws.max_row + 1):
        priority_ws.cell(row=row_idx, column=ctr_col).number_format = "0.0000"


def style_dashboard(worksheet: Worksheet, writer: Any) -> None:
    """Render the executive dashboard sheet with KPI and ownership blocks.

    Ported from legacy ``_style_dashboard`` in ``tables_impl.legacy.old.py``
    (aggregation order, D/E status tables, owner summary, tooltips, row heights).

    Args:
        worksheet: Dashboard worksheet instance.
        writer: Pandas ExcelWriter-like object exposing workbook via ``book``.
    """
    worksheet.sheet_view.showGridLines = False
    set_freeze_panes_safe(worksheet, "A2")
    worksheet._charts = []
    light_header_fill = PatternFill("solid", fgColor=LIGHT_HEADER_COLOR)
    headers = header_index(worksheet)
    health_from_feed: float | None = None
    projected_health_from_feed: float | None = None
    projected_pass_from_feed: float | None = None
    seo_pass_rate_from_run: float | None = None
    for row_idx in range(1, 80):
        for col_idx in range(4, max(worksheet.max_column + 1, 26)):
            worksheet.cell(row=row_idx, column=col_idx, value=None)
    metric_col = headers.get("Metric")
    value_col = headers.get("Value")
    total_urls = 0
    pass_rate_pct = 0.0
    if metric_col and value_col:
        metric_to_value: dict[str, Any] = {}
        for row_idx in range(2, worksheet.max_row + 1):
            metric_name = str(
                worksheet.cell(row=row_idx, column=metric_col).value or ""
            ).strip()
            if metric_name:
                metric_to_value[metric_name] = worksheet.cell(
                    row=row_idx, column=value_col
                ).value
        if worksheet.max_row > 0:
            worksheet.delete_rows(1, worksheet.max_row)
        for col_letter, width in DASHBOARD_COLUMN_WIDTHS.items():
            worksheet.column_dimensions[col_letter].width = width
        worksheet.column_dimensions["A"].width = 25
        worksheet.column_dimensions["B"].width = 25
        total_urls = to_int(
            metric_to_value.get("URLs Crawled") or metric_to_value.get("Total URLs"),
            0,
        )
        pass_raw = metric_to_value.get("SEO Pass Rate %") or metric_to_value.get(
            "Pass Rate (%)"
        )
        try:
            pass_rate_pct = float(pass_raw or 0.0)
        except Exception:
            pass_rate_pct = 0.0
        seo_pass_rate_from_run = pass_rate_pct
        critical_urls = to_int(metric_to_value.get("Critical URL Count"), 0)
        warning_urls = to_int(metric_to_value.get("Warning URL Count"), 0)
        try:
            if "Health Score %" in metric_to_value:
                health_from_feed = float(metric_to_value.get("Health Score %") or 0.0)
        except Exception:
            health_from_feed = None
        try:
            if "Projected Health Score %" in metric_to_value:
                projected_health_from_feed = float(
                    metric_to_value.get("Projected Health Score %") or 0.0
                )
        except Exception:
            projected_health_from_feed = None
        try:
            if "Projected Pass Rate %" in metric_to_value:
                projected_pass_from_feed = float(
                    metric_to_value.get("Projected Pass Rate %") or 0.0
                )
        except Exception:
            projected_pass_from_feed = None

        title = worksheet["A1"]
        title.value = DASHBOARD_BRAND_A1
        worksheet["A2"] = "Executive SEO & AEO Dashboard"
        worksheet["A2"].font = Font(color=STD_NAVY, bold=True, size=12)
        title.font = Font(color=STD_NAVY, bold=True, size=16)
        title.alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[1].height = 45

        header_fill = PatternFill("solid", fgColor=TABLE_HEADER_COLOR)
        header_font = Font(color="000000", bold=True, size=12)
        value_fill = PatternFill("solid", fgColor=VALUE_BLOCK_COLOR)

        worksheet["A4"] = "EXECUTIVE METRICS"
        worksheet["B4"] = "Value"
        for ref in ("A4", "B4"):
            worksheet[ref].fill = header_fill
            worksheet[ref].font = header_font
            worksheet[ref].alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        worksheet["A5"] = '=HYPERLINK("#\'Main\'!A1","Average SEO Score")'
        worksheet["B5"] = (
            f'=IFERROR(AVERAGE({_MAIN_DYNAMIC_COLUMN.format(header="SEO Score")}),'
            f'IFERROR(AVERAGE({_MAIN_DYNAMIC_COLUMN.format(header="SEO Health Score")}),0))/100'
        )
        worksheet["B5"].number_format = "0%"
        worksheet["A6"] = '=HYPERLINK("#\'Technical Diagnostics\'!A1","Technical Health")'
        worksheet["B6"] = (
            f'=IFERROR(AVERAGE({_TD_DYNAMIC_COLUMN.format(header="SEO Health Score")}),0)/100'
        )
        worksheet["B6"].number_format = "0%"
        worksheet["A7"] = '=HYPERLINK("#\'Technical Diagnostics\'!A1","Performance (PSI)")'
        worksheet["B7"] = (
            f'=IFERROR(('
            f'IFERROR(AVERAGE({_TD_DYNAMIC_COLUMN.format(header="Mobile PSI Score")}),0)+'
            f'IFERROR(AVERAGE({_TD_DYNAMIC_COLUMN.format(header="Desktop PSI Score")}),0)'
            f')/2,0)/100'
        )
        worksheet["B7"].number_format = "0%"
        worksheet["A8"] = (
            '=HYPERLINK("#\'Technical Diagnostics\'!A1","Mobile vs. Desktop Variance")'
        )
        worksheet["B8"] = (
            f"=ROUND(IFERROR("
            f"IFERROR(AVERAGE({_TD_DYNAMIC_COLUMN.format(header='Mobile PSI Score')}),0)-"
            f"IFERROR(AVERAGE({_TD_DYNAMIC_COLUMN.format(header='Desktop PSI Score')}),0)"
            f",0),1)"
        )
        worksheet["B8"].number_format = "0.0"
        _td_sev = _technical_diagnostics_data_column("Severity Badge")
        _td_sc = _technical_diagnostics_data_column("Status Code")
        worksheet["A9"] = '=HYPERLINK("#\'Priority URLs\'!A1","Critical URLs")'
        worksheet["B9"] = f'=COUNTIFS({_td_sev},"Critical")'
        worksheet["A10"] = '=HYPERLINK("#\'Technical Diagnostics\'!A1","Warning URLs")'
        worksheet["B10"] = (
            f'=COUNTIFS({_td_sev},"Warning")+COUNTIFS({_td_sev},"Needs Work")'
        )
        worksheet["A11"] = '=HYPERLINK("#\'Technical Diagnostics\'!A1","Error Rate % (4xx/5xx)")'
        worksheet["B11"] = (
            f"=IFERROR(("
            f"COUNTIFS({_td_sc},\">=400\",{_td_sc},\"<500\")"
            f"+COUNTIFS({_td_sc},\">=500\",{_td_sc},\"<600\"))"
            f"/MAX(1,COUNTIFS({_td_sc},\">=200\",{_td_sc},\"<600\")),0)"
        )
        worksheet["B11"].number_format = "0%"
        worksheet["A12"] = (
            '=HYPERLINK("#\'Technical Diagnostics\'!A1","Crawl Success Rate % (2xx)")'
        )
        worksheet["B12"] = (
            f"=IFERROR(COUNTIFS({_td_sc},\">=200\",{_td_sc},\"<300\")"
            f"/MAX(1,COUNTIFS({_td_sc},\">=200\",{_td_sc},\"<600\")),0)"
        )
        worksheet["B12"].number_format = "0%"
        worksheet["A13"] = '=HYPERLINK("#\'Technical Diagnostics\'!A1","Critical URL Rate %")'
        worksheet["B13"] = f"=IFERROR(B9/{_technical_url_row_denominator()},0)"
        worksheet["B13"].number_format = "0%"
        worksheet["A14"] = '=HYPERLINK("#\'Technical Diagnostics\'!A1","Warning URL Rate %")'
        worksheet["B14"] = f"=IFERROR(B10/{_technical_url_row_denominator()},0)"
        worksheet["B14"].number_format = "0%"
        worksheet["A15"] = "Projected Health Score % (if all To Do done)"
        worksheet["A16"] = "Projected Pass Rate % (if all To Do done)"
        worksheet["A17"] = (
            f'=HYPERLINK("#\'{CONTENT_OPTIMISATION_HUB_SHEET}\'!A1",'
            f'"Content Hub Readiness (%)")'
        )
        worksheet["B17"] = (
            f"=IF(COUNTA({_CONTENT_HUB_STATUS_RANGE})=0,0,"
            f'COUNTIF({_CONTENT_HUB_STATUS_RANGE},"Completed")/'
            f"COUNTA({_CONTENT_HUB_STATUS_RANGE}))"
        )
        worksheet["B17"].number_format = "0%"
        worksheet["A18"] = '=HYPERLINK("#\'Main\'!A1","URLs with Schema")'
        worksheet["B18"] = (
            f"=IFERROR(COUNTIF({_main_sheet_data_column('Has Valid JSON-LD')},TRUE),0)"
        )
        worksheet["B18"].number_format = "0"
        worksheet["A19"] = '=HYPERLINK("#\'Content & AI Readiness\'!A1","AEO Opportunity Gap")'
        worksheet["B19"] = (
            f"=MAX(0, IF({_CONTENT_AI_AEO_EXTRACTABILITY_AVG}>0, "
            f"(100-{_CONTENT_AI_AEO_EXTRACTABILITY_AVG})/100, "
            f"1-IFERROR(B5,0)))"
        )
        worksheet["B19"].number_format = "0%"
        worksheet["A20"] = (
            '=HYPERLINK("#\'Link Inventory\'!A1","Broken Internal Links (instances)")'
        )
        worksheet["B20"] = link_inventory_broken_internal_total_formula()
        worksheet["A21"] = (
            '=HYPERLINK("#\'Link Intelligence\'!A1","Generic Anchor Links (Total)")'
        )
        worksheet["B21"] = "=SUMIFS('Link Intelligence'!$O:$O,'Link Intelligence'!$B:$B,\"Summary\")"
        worksheet["B21"].number_format = "0"
        worksheet["A22"] = '=HYPERLINK("#\'Link Inventory\'!A1","External Link Health %")'
        worksheet["B22"] = _EXCEL_EXTERNAL_LINK_HEALTH_PCT
        worksheet["B22"].number_format = "0%"
        # Sprint 6 — Executive ROI roll-ups sourced from Content Hub Metrics
        # (same URL rows as the Hub; row-1 headers). INDEX/MATCH keeps
        # formulas stable if column order changes.
        worksheet["A23"] = (
            f'=HYPERLINK("#\'{CONTENT_HUB_METRICS_SHEET}\'!A1",'
            f'"Total Potential Traffic Lift (sum)")'
        )
        worksheet["B23"] = (
            f"=IFERROR(SUM("
            f'{_CONTENT_HUB_METRICS_DYNAMIC_COLUMN.format(header="Potential Traffic Lift")}'
            f"),0)"
        )
        worksheet["B23"].number_format = "#,##0"
        worksheet["A24"] = (
            f'=HYPERLINK("#\'{CONTENT_HUB_METRICS_SHEET}\'!A1",'
            f'"Critical Priority Pages")'
        )
        worksheet["B24"] = (
            f"=IFERROR(SUMPRODUCT(--("
            f'{_CONTENT_HUB_METRICS_DYNAMIC_COLUMN.format(header="Instant Priority")}'
            f'="CRITICAL")),0)'
        )
        worksheet["B24"].number_format = "0"
        for ref in (
            "A5",
            "B5",
            "A6",
            "B6",
            "A7",
            "B7",
            "A8",
            "B8",
            "A9",
            "B9",
            "A10",
            "B10",
            "A11",
            "B11",
            "A12",
            "B12",
            "A13",
            "B13",
            "A14",
            "B14",
            "A15",
            "B15",
            "A16",
            "B16",
            "A17",
            "B17",
            "A18",
            "B18",
            "A19",
            "B19",
            "A20",
            "B20",
            "A21",
            "B21",
            "A22",
            "B22",
            "A23",
            "B23",
            "A24",
            "B24",
        ):
            worksheet[ref].fill = value_fill
            worksheet[ref].font = Font(
                color="1F2937", bold=True, size=12 if ref.startswith("A") else 14
            )
            worksheet[ref].alignment = Alignment(horizontal="center", vertical="center")
        for ref in (
            "A5",
            "A6",
            "A7",
            "A8",
            "A9",
            "A10",
            "A11",
            "A12",
            "A13",
            "A14",
            "A17",
            "A18",
            "A19",
            "A20",
            "A21",
            "A22",
            "A23",
            "A24",
        ):
            worksheet[ref].font = Font(
                color=STD_BLUE, underline="single", bold=True, size=12
            )
        for row in range(5, 25):
            worksheet[f"A{row}"].alignment = Alignment(
                horizontal="center", vertical="center"
            )
            worksheet.row_dimensions[row].height = 24
        for addr, body in DASHBOARD_KPI_ROW_COMMENTS.items():
            kpi_comment = Comment(body, "hype-frog")
            apply_comment_dimensions(kpi_comment)
            worksheet[addr].comment = kpi_comment
        # Sprint 6 — inline tooltips for the new executive ROI KPIs.
        # Inlined here (not in ``DASHBOARD_KPI_ROW_COMMENTS``) so this
        # sprint stays inside its 5-file budget without touching
        # ``dashboard_config.py``.
        _exec_roi_tooltips: dict[str, str] = {
            "A23": (
                "Total Potential Traffic Lift (sum)\n\n"
                "Sum of 'Potential Traffic Lift' across every URL on the "
                "Content Hub Metrics sheet (aligned with the Hub URL set). "
                "Each row uses the same rule as the crawl export: "
                "GSC Clicks * ((100 - Semantic AEO Score) / 100) * 0.25 "
                "(25% AEO ceiling), in click units for the GSC window used in the run. "
                "Pages with no GSC traffic or no AEO score contribute 0."
            ),
            "A24": (
                "Critical Priority Pages\n\n"
                "Count of Content Hub Metrics rows whose 'Instant Priority' is "
                "'CRITICAL' — i.e. GSC Clicks > 500 AND (Semantic AEO "
                "Score < 50 OR Field LCP > 2500ms)."
            ),
        }
        for addr, body in _exec_roi_tooltips.items():
            roi_comment = Comment(body, "hype-frog")
            apply_comment_dimensions(roi_comment)
            worksheet[addr].comment = roi_comment

    summary_metrics = SummaryMetricsPayload(
        urls_crawled=max(0, total_urls),
        seo_pass_rate_pct=max(0.0, min(100.0, _safe_float(seo_pass_rate_from_run, 0.0))),
        health_score_pct=max(0.0, min(100.0, _safe_float(health_from_feed, 0.0))),
        critical_url_count=max(0, to_int(metric_to_value.get("Critical URL Count"), 0))
        if metric_col and value_col
        else 0,
        warning_url_count=max(0, to_int(metric_to_value.get("Warning URL Count"), 0))
        if metric_col and value_col
        else 0,
        projected_health_score_pct=max(
            0.0, min(100.0, _safe_float(projected_health_from_feed, 0.0))
        ),
        projected_pass_rate_pct=max(
            0.0, min(100.0, _safe_float(projected_pass_from_feed, 0.0))
        ),
    )

    technical_main_rows: list[MainRowPayload] = []
    technical_extra_rows: list[ExtraRowPayload] = []
    technical_rows = _technical_sheet_rows(writer)
    if technical_rows:
        technical_main_rows = [
            MainRowPayload.model_validate(row_dict) for row_dict in technical_rows
        ]
        technical_extra_rows = [
            ExtraRowPayload.model_validate(row_dict) for row_dict in technical_rows
        ]

    fixplan_rows: list[FixPlanRowPayload] = []
    if "FixPlan" in writer.book.sheetnames:
        raw_fixplan_rows = _sheet_rows(writer.book["FixPlan"])
        fixplan_rows = [
            FixPlanRowPayload.model_validate({**row_dict, "source_row": idx})
            for idx, row_dict in enumerate(raw_fixplan_rows, start=2)
        ]

    aeo_rows: list[ExtraRowPayload] = []
    if "AEO" in writer.book.sheetnames:
        aeo_rows = [
            ExtraRowPayload.model_validate(row_dict)
            for row_dict in _sheet_rows(writer.book["AEO"])
        ]

    dashboard_metrics = compute_dashboard_metrics(
        summary_metrics=summary_metrics,
        technical_main_rows=technical_main_rows,
        technical_extra_rows=technical_extra_rows,
        fixplan_rows=fixplan_rows,
        aeo_rows=aeo_rows,
    )

    content_ai_readiness_rows: list[dict[str, Any]] = []
    if "Content & AI Readiness" in writer.book.sheetnames:
        content_ai_readiness_rows = _sheet_rows(writer.book["Content & AI Readiness"])
    link_inventory_rows_narrative: list[dict[str, Any]] = []
    if "Link Inventory" in writer.book.sheetnames:
        link_inventory_rows_narrative = _sheet_rows(writer.book["Link Inventory"])

    avg_seo_pct_narrative = average_seo_score_pct(technical_main_rows)
    if (
        not technical_main_rows
        and summary_metrics.urls_crawled > 0
        and summary_metrics.health_score_pct > 0.0
    ):
        avg_seo_pct_narrative = summary_metrics.health_score_pct
    business_impact_narrative = NarrativeEngine.build_business_impact(
        total_urls=summary_metrics.urls_crawled,
        link_inventory_rows=link_inventory_rows_narrative,
        technical_extra_rows=technical_extra_rows,
        content_ai_rows=content_ai_readiness_rows,
        avg_seo_score_pct=avg_seo_pct_narrative,
    )
    strategic_narrative_text = NarrativeEngine.build_strategic_narrative(
        total_urls=summary_metrics.urls_crawled,
        avg_seo_score_pct=avg_seo_pct_narrative,
        critical_url_count=dashboard_metrics.critical_urls,
    )
    broken_internal_instances = count_broken_internal_instances(
        link_inventory_rows_narrative
    )
    if broken_internal_instances > 0:
        strategic_narrative_text = (
            f"{strategic_narrative_text}\n\n"
            f"Broken internal link instances: {broken_internal_instances} "
            f"(see KPI B20 and Link Inventory)."
        )

    status_buckets = dashboard_metrics.status_buckets
    error_count = dashboard_metrics.error_count
    success_count = dashboard_metrics.success_count
    crawl_denominator = dashboard_metrics.crawl_denominator
    pass_rate_pct = dashboard_metrics.pass_rate_pct
    critical_rate_pct = dashboard_metrics.critical_rate_pct
    warning_rate_pct = dashboard_metrics.warning_rate_pct
    _td_sev2 = _technical_diagnostics_data_column("Severity Badge")
    _td_sc2 = _technical_diagnostics_data_column("Status Code")
    worksheet["B8"] = (
        f"=ROUND(IFERROR("
        f"IFERROR(AVERAGE({_TD_DYNAMIC_COLUMN.format(header='Mobile PSI Score')}),0)-"
        f"IFERROR(AVERAGE({_TD_DYNAMIC_COLUMN.format(header='Desktop PSI Score')}),0)"
        f",0),1)"
    )
    worksheet["B8"].number_format = "0.0"
    worksheet["B9"] = f'=COUNTIFS({_td_sev2},"Critical")'
    worksheet["B10"] = (
        f'=COUNTIFS({_td_sev2},"Warning")+COUNTIFS({_td_sev2},"Needs Work")'
    )
    worksheet["B17"] = (
        f"=IF(COUNTA({_CONTENT_HUB_STATUS_RANGE})=0,0,"
        f'COUNTIF({_CONTENT_HUB_STATUS_RANGE},"Completed")/'
        f"COUNTA({_CONTENT_HUB_STATUS_RANGE}))"
    )
    worksheet["B17"].number_format = "0%"
    worksheet["B18"] = (
        f"=IFERROR(COUNTIF({_main_sheet_data_column('Has Valid JSON-LD')},TRUE),0)"
    )
    worksheet["B18"].number_format = "0"
    worksheet["B19"] = (
        f"=MAX(0, IF({_CONTENT_AI_AEO_EXTRACTABILITY_AVG}>0, "
        f"(100-{_CONTENT_AI_AEO_EXTRACTABILITY_AVG})/100, "
        f"1-IFERROR(B5,0)))"
    )
    worksheet["B19"].number_format = "0%"
    worksheet["B20"] = link_inventory_broken_internal_total_formula()
    worksheet["B21"] = "=SUMIFS('Link Intelligence'!$O:$O,'Link Intelligence'!$B:$B,\"Summary\")"
    worksheet["B21"].number_format = "0"
    worksheet["B22"] = _EXCEL_EXTERNAL_LINK_HEALTH_PCT
    worksheet["B22"].number_format = "0%"
    worksheet["B11"] = (
        f"=IFERROR(("
        f"COUNTIFS({_td_sc2},\">=400\",{_td_sc2},\"<500\")"
        f"+COUNTIFS({_td_sc2},\">=500\",{_td_sc2},\"<600\"))"
        f"/MAX(1,COUNTIFS({_td_sc2},\">=200\",{_td_sc2},\"<600\")),0)"
    )
    worksheet["B11"].number_format = "0%"
    worksheet["B12"] = (
        f"=IFERROR(COUNTIFS({_td_sc2},\">=200\",{_td_sc2},\"<300\")"
        f"/MAX(1,COUNTIFS({_td_sc2},\">=200\",{_td_sc2},\"<600\")),0)"
    )
    worksheet["B12"].number_format = "0%"
    worksheet["B13"] = f"=IFERROR(B9/{_technical_url_row_denominator()},0)"
    worksheet["B13"].number_format = "0%"
    worksheet["B14"] = f"=IFERROR(B10/{_technical_url_row_denominator()},0)"
    worksheet["B14"].number_format = "0%"
    worksheet["B7"] = (
        f'=IFERROR(('
        f'IFERROR(AVERAGE({_TD_DYNAMIC_COLUMN.format(header="Mobile PSI Score")}),0)+'
        f'IFERROR(AVERAGE({_TD_DYNAMIC_COLUMN.format(header="Desktop PSI Score")}),0)'
        f')/2,0)/100'
    )
    worksheet["B7"].number_format = "0%"
    for row in range(5, 15):
        worksheet[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[row].height = 24

    avg_health_score = dashboard_metrics.avg_health_score
    worksheet["B6"] = (
        f'=IFERROR(AVERAGE({_TD_DYNAMIC_COLUMN.format(header="SEO Health Score")}),0)/100'
    )
    worksheet["B6"].number_format = "0%"
    worksheet["B5"] = (
        f'=IFERROR(AVERAGE({_MAIN_DYNAMIC_COLUMN.format(header="SEO Score")}),'
        f'IFERROR(AVERAGE({_MAIN_DYNAMIC_COLUMN.format(header="SEO Health Score")}),0))/100'
    )
    worksheet["B5"].number_format = "0%"

    table_header_fill = PatternFill("solid", fgColor=TABLE_HEADER_COLOR)
    table_header_font = Font(color="000000", bold=True, size=11)
    for ref, val in {
        "D4": "Status",
        "E4": "Count",
        "D12": "Severity",
        "E12": "Count",
    }.items():
        cell = worksheet[ref]
        cell.value = val
        cell.fill = table_header_fill
        cell.font = table_header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    status_rows = [
        (label, status_buckets[label], color) for label, color in STATUS_ROW_STYLE
    ]
    for idx, (label, count, color) in enumerate(status_rows, start=5):
        worksheet[f"D{idx}"] = label
        worksheet[f"E{idx}"] = count
        worksheet[f"D{idx}"].fill = PatternFill("solid", fgColor=color)
        worksheet[f"E{idx}"].fill = PatternFill("solid", fgColor=color)

    sev_rows = [
        (
            label,
            dashboard_metrics.severity_counts.get(
                "High" if label == "Warning" else label,
                0,
            ),
            color,
        )
        for label, color in SEVERITY_ROW_STYLE
    ]
    for idx, (label, count, color) in enumerate(sev_rows, start=13):
        worksheet[f"D{idx}"] = label
        worksheet[f"E{idx}"] = count
        worksheet[f"D{idx}"].fill = PatternFill("solid", fgColor=color)
        worksheet[f"E{idx}"].fill = PatternFill("solid", fgColor=color)

    for row in range(5, 16):
        for col in ("D", "E"):
            worksheet[f"{col}{row}"].alignment = Alignment(
                horizontal="center", vertical="center"
            )

    worksheet["M4"] = "PRIORITY SNAPSHOT"
    worksheet["N4"] = "Value"
    for ref in ("M4", "N4"):
        worksheet[ref].fill = table_header_fill
        worksheet[ref].font = table_header_font
        worksheet[ref].alignment = Alignment(horizontal="center", vertical="center")
    top_issue_name = dashboard_metrics.top_issue_name
    top_issue_affected = dashboard_metrics.top_issue_affected
    worksheet["M5"] = "Primary blocking issue"
    worksheet["N5"] = top_issue_name
    worksheet["M6"] = "Top Issue Affected URLs"
    worksheet["N6"] = top_issue_affected
    worksheet["M7"] = "4xx/5xx URLs"
    worksheet["N7"] = dashboard_metrics.error_count
    worksheet["M8"] = "Avg TTFB (ms)"
    worksheet["N8"] = dashboard_metrics.avg_ttfb_ms
    for row in range(5, 9):
        worksheet[f"M{row}"].fill = PatternFill("solid", fgColor=PANEL_BG_COLOR)
        worksheet[f"N{row}"].fill = PatternFill("solid", fgColor=PANEL_BG_COLOR)

    worksheet["G22"] = "OWNER ISSUE SUMMARY"
    worksheet["G23"] = "Owner"
    worksheet["H23"] = "Issue Rows"
    worksheet["I23"] = "Affected URLs"
    worksheet["J23"] = "Critical"
    worksheet["K23"] = "Warning"
    for ref in ("G22",):
        worksheet[ref].fill = light_header_fill
        worksheet[ref].font = Font(color="000000", bold=True, size=12)
        worksheet[ref].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    owner_header_fill = PatternFill("solid", fgColor=TABLE_HEADER_COLOR)
    for ref in ("G23", "H23", "I23", "J23", "K23"):
        worksheet[ref].fill = owner_header_fill
        worksheet[ref].font = Font(color="000000", bold=True, size=11)
        worksheet[ref].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    worksheet.merge_cells("G22:K22")
    worksheet["G22"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    owner_rows_sorted = sorted(
        dashboard_metrics.owner_rollup.items(),
        key=lambda x: (
            -x[1].affected_urls,
            -x[1].critical,
            -x[1].warning,
            x[0],
        ),
    )
    owner_start_row = 24
    for idx, (owner_name, metrics) in enumerate(
        owner_rows_sorted[:8], start=owner_start_row
    ):
        worksheet[f"G{idx}"] = f'=HYPERLINK("#\'FixPlan\'!A1","{owner_name}")'
        worksheet[f"H{idx}"] = metrics.issue_rows
        worksheet[f"I{idx}"] = metrics.affected_urls
        worksheet[f"J{idx}"] = metrics.critical
        worksheet[f"K{idx}"] = metrics.warning
        for col in ("G", "H", "I", "J", "K"):
            worksheet[f"{col}{idx}"].fill = PatternFill("solid", fgColor=PANEL_BG_COLOR)
            worksheet[f"{col}{idx}"].alignment = Alignment(
                horizontal="center", vertical="center"
            )
    if not owner_rows_sorted:
        worksheet["G24"] = "No owner data"
        worksheet["G24"].fill = PatternFill("solid", fgColor=PANEL_BG_COLOR)
        worksheet["G24"].alignment = Alignment(horizontal="center", vertical="center")

    for col, width in {"G": 25, "H": 25, "I": 30, "J": 30, "K": 30}.items():
        worksheet.column_dimensions[col].width = width

    # Crawl metadata sits below executive KPIs (rows 5–22) so link metrics on A21:B22
    # are not overwritten (External Link Health % stays on B22 with RunMetadata lookup).
    worksheet["A25"] = "CRAWL METADATA"
    worksheet["B25"] = "Value"
    for ref in ("A25", "B25"):
        worksheet[ref].fill = table_header_fill
        worksheet[ref].font = table_header_font
        worksheet[ref].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["A26"] = "Run Date"
    worksheet["B26"] = (
        f'=IFERROR(INDEX(\'{_AUDIT_DETAILS_SHEET_ESC}\'!$B:$B,'
        f'MATCH("Run Timestamp",\'{_AUDIT_DETAILS_SHEET_ESC}\'!$A:$A,0)),"")'
    )
    worksheet["A27"] = "URL Count"
    worksheet["B27"] = (
        f'=IFERROR(INDEX(\'{_AUDIT_DETAILS_SHEET_ESC}\'!$B:$B,'
        f'MATCH("Total URLs",\'{_AUDIT_DETAILS_SHEET_ESC}\'!$A:$A,0)),'
        f"MAX(0,COUNTA({_main_sheet_data_column('URL')})))"
    )
    worksheet["A28"] = "Duration"
    worksheet["B28"] = (
        f'=IFERROR(INDEX(\'{_AUDIT_DETAILS_SHEET_ESC}\'!$B:$B,'
        f'MATCH("Duration (s)",\'{_AUDIT_DETAILS_SHEET_ESC}\'!$A:$A,0)),"")'
    )
    for row in range(26, 29):
        worksheet[f"A{row}"].fill = PatternFill("solid", fgColor=PANEL_BG_COLOR)
        worksheet[f"B{row}"].fill = PatternFill("solid", fgColor=PANEL_BG_COLOR)
        worksheet[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
        worksheet[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")

    overall_health = dashboard_metrics.overall_health
    health_fill = (
        GOOD_COLOR
        if overall_health >= 80
        else WARN_COLOR if overall_health >= 60 else ALERT_COLOR
    )
    worksheet["B6"].fill = PatternFill("solid", fgColor=health_fill)
    b7_pass_display = dashboard_metrics.b7_pass_display
    worksheet["B7"].fill = PatternFill(
        "solid",
        fgColor=(
            GOOD_COLOR
            if b7_pass_display >= 70
            else WARN_COLOR if b7_pass_display >= 40 else ALERT_COLOR
        ),
    )
    worksheet["B11"].fill = PatternFill(
        "solid", fgColor=ALERT_COLOR if error_count > 0 else GOOD_COLOR
    )
    worksheet["B12"].fill = PatternFill(
        "solid",
        fgColor=(
            GOOD_COLOR if success_count >= max(1, crawl_denominator * 0.9) else WARN_COLOR
        ),
    )
    worksheet["B13"].fill = PatternFill(
        "solid",
        fgColor=(
            ALERT_COLOR
            if dashboard_metrics.critical_rate_pct >= 10
            else WARN_COLOR if critical_rate_pct > 0 else GOOD_COLOR
        ),
    )
    worksheet["B14"].fill = PatternFill(
        "solid",
        fgColor=(
            ALERT_COLOR
            if dashboard_metrics.warning_rate_pct >= 50
            else WARN_COLOR if warning_rate_pct > 20 else GOOD_COLOR
        ),
    )
    projected_health_pct = summary_metrics.projected_health_score_pct
    projected_pass_rate_pct = summary_metrics.projected_pass_rate_pct
    worksheet["B15"] = round(
        max(0.0, min(1.0, projected_health_pct / 100.0)),
        6,
    )
    worksheet["B15"].number_format = "0%"
    worksheet["B16"] = round(
        max(0.0, min(1.0, projected_pass_rate_pct / 100.0)),
        6,
    )
    worksheet["B16"].number_format = "0%"
    worksheet["B15"].fill = PatternFill(
        "solid",
        fgColor=(
            GOOD_COLOR
            if projected_health_pct >= 80
            else WARN_COLOR if projected_health_pct >= 60 else ALERT_COLOR
        ),
    )
    worksheet["B16"].fill = PatternFill(
        "solid",
        fgColor=(
            GOOD_COLOR
            if projected_pass_rate_pct >= 70
            else WARN_COLOR if projected_pass_rate_pct >= 40 else ALERT_COLOR
        ),
    )
    top_issues_header_fill = PatternFill("solid", fgColor=SOFT_ALERT_COLOR)
    top_issues_header_font = Font(color="000000", bold=True, size=12)
    worksheet.merge_cells("G13:H13")
    worksheet["G13"] = "TOP ISSUES TO FIX FIRST"
    worksheet["G13"].fill = top_issues_header_fill
    worksheet["G13"].font = top_issues_header_font
    worksheet["G13"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[13].height = 28
    worksheet["G14"] = "Issue (Fix Plan link)"
    worksheet["H14"] = "Affected URLs"
    for ref in ("G14", "H14"):
        worksheet[ref].fill = table_header_fill
        worksheet[ref].font = table_header_font
        worksheet[ref].alignment = Alignment(horizontal="center", vertical="center")
    for idx, top_issue in enumerate(dashboard_metrics.top_issue_rows, start=15):
        worksheet[f"G{idx}"] = (
            f'=HYPERLINK("#FixPlan!A{top_issue.source_row}","{top_issue.issue_name}")'
        )
        worksheet[f"H{idx}"] = top_issue.affected_urls
        row_fill = PatternFill(
            "solid",
            fgColor=SOFT_WARN_COLOR if idx == 15 else PANEL_BG_COLOR,
        )
        worksheet[f"G{idx}"].fill = row_fill
        worksheet[f"H{idx}"].fill = row_fill
        worksheet[f"G{idx}"].font = Font(color=STD_BLUE, underline="single", bold=True)
        worksheet.row_dimensions[idx].height = 22

    quick_nav_fill = PatternFill("solid", fgColor=STD_NAVY)
    worksheet["I12"] = "Quick Navigation"
    worksheet["J12"] = "Open"
    for ref in ("I12", "J12"):
        worksheet[ref].fill = quick_nav_fill
        worksheet[ref].font = Font(color="000000", bold=True, size=11)
        worksheet[ref].alignment = Alignment(horizontal="center", vertical="center")
    target_remap = {
        "#Technical!A1": "#'Technical Diagnostics'!A1",
        "#Indexability!A1": "#'Technical Diagnostics'!A1",
        "#AEO!A1": "#'Content & AI Readiness'!A1",
    }
    for idx, (label, target) in enumerate(QUICK_LINKS, start=13):
        target = target_remap.get(target, target)
        worksheet[f"I{idx}"] = label
        worksheet[f"I{idx}"].alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
        worksheet[f"J{idx}"] = f'=HYPERLINK("{target}","Open")'
        worksheet[f"J{idx}"].font = Font(color=STD_BLUE, underline="single", bold=True)
        worksheet[f"I{idx}"].fill = PatternFill("solid", fgColor=PANEL_BG_COLOR)
        worksheet[f"J{idx}"].fill = PatternFill("solid", fgColor=PANEL_BG_COLOR)
        worksheet[f"K{idx}"].fill = PatternFill("solid", fgColor=PANEL_BG_COLOR)

    worksheet["I19"] = "Content Hub"
    worksheet["J19"] = f'=HYPERLINK("#\'{CONTENT_OPTIMISATION_HUB_SHEET}\'!A1","Open")'
    worksheet["I19"].fill = PatternFill("solid", fgColor=PANEL_BG_COLOR)
    worksheet["J19"].fill = PatternFill("solid", fgColor=PANEL_BG_COLOR)
    worksheet["I19"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["J19"].font = Font(color=STD_BLUE, underline="single", bold=True)
    worksheet["J19"].alignment = Alignment(horizontal="center", vertical="center")

    worksheet["I20"] = "Advanced Sheets"
    worksheet["I20"].font = Font(bold=True, color=STD_NAVY)
    worksheet.merge_cells("I20:K20")
    worksheet["I21"] = DASHBOARD_ADVANCED_SHEETS_NOTE
    worksheet.merge_cells("I21:K22")
    worksheet["I21"].alignment = Alignment(wrap_text=True, vertical="top")
    adv_row = 23
    for sheet_name, label in DASHBOARD_ADVANCED_SHEET_LINKS:
        safe = excel_sheet_link_target(sheet_name)
        worksheet[f"I{adv_row}"] = label
        worksheet[f"J{adv_row}"] = f'=HYPERLINK("#\'{safe}\'!A1","Open")'
        for col in ("I", "J"):
            worksheet[f"{col}{adv_row}"].fill = PatternFill("solid", fgColor=PANEL_BG_COLOR)
        worksheet[f"I{adv_row}"].alignment = Alignment(horizontal="left", vertical="center")
        worksheet[f"J{adv_row}"].font = Font(color=STD_BLUE, underline="single", bold=True)
        worksheet[f"J{adv_row}"].alignment = Alignment(horizontal="center", vertical="center")
        adv_row += 1

    worksheet["I4"] = "BUSINESS IMPACT SUMMARY"
    worksheet["I5"] = business_impact_narrative
    worksheet.merge_cells("I4:K4")
    worksheet.merge_cells("I5:K10")
    worksheet["I4"].fill = table_header_fill
    worksheet["I4"].font = table_header_font
    worksheet["I4"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["I5"].fill = PatternFill("solid", fgColor=PANEL_BG_COLOR)
    worksheet["I5"].font = Font(color="000000", bold=False, size=11)
    worksheet["I5"].alignment = Alignment(
        horizontal="left", vertical="top", wrap_text=True
    )

    worksheet["G4"] = "Traditional SEO vs. 2026 AEO Readiness"
    worksheet["G5"] = "Dimension"
    worksheet["H5"] = "Score"
    for ref in ("G4", "G5", "H5"):
        worksheet[ref].fill = table_header_fill
        worksheet[ref].font = table_header_font
        worksheet[ref].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.merge_cells("G4:H4")
    worksheet["G4"].font = Font(color="000000", bold=True, size=10)
    worksheet["G4"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    traditional_score = dashboard_metrics.traditional_score
    aeo_readiness = dashboard_metrics.aeo_readiness
    worksheet["G6"] = "Traditional SEO"
    worksheet["H6"] = traditional_score / 100.0
    worksheet["H6"].number_format = "0%"
    worksheet["G7"] = "2026 AEO Readiness"
    worksheet["H7"] = aeo_readiness / 100.0
    worksheet["H7"].number_format = "0%"
    for ref in ("G6", "H6", "G7", "H7"):
        worksheet[ref].fill = PatternFill("solid", fgColor=PANEL_BG_COLOR)
        worksheet[ref].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["G9"] = "Strategic Narrative"
    worksheet["G10"] = strategic_narrative_text
    worksheet.merge_cells("G10:H12")
    worksheet["G9"].fill = table_header_fill
    worksheet["G9"].font = table_header_font
    worksheet["G9"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    worksheet["G10"].fill = PatternFill("solid", fgColor=PANEL_BG_COLOR)
    worksheet["G10"].font = Font(color="000000", bold=False, size=11)
    worksheet["G10"].alignment = Alignment(
        horizontal="left", vertical="top", wrap_text=True
    )

    for ref, message in DASHBOARD_TOOLTIPS.items():
        cell = worksheet[ref]
        title = f"KPI {ref}"
        kpi_comment = Comment(f"{title}\n\n{message}", "hype-frog")
        apply_comment_dimensions(kpi_comment)
        cell.comment = kpi_comment
    worksheet.conditional_formatting.add(
        "B20",
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            stopIfTrue=True,
            fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            font=Font(color="FFFFFF", bold=True),
        ),
    )
    apply_dashboard_metric_conditional_rules(worksheet)
    _format_priority_urls_ctr(writer)
    # Lock percentage display for executive KPIs (0–1 → ``67%`` via ``0%`` format).
    for addr in _DASHBOARD_PERCENT_KPI_CELLS:
        worksheet[addr].number_format = "0%"
    banner = worksheet["A1"]
    banner.value = DASHBOARD_BRAND_A1
    banner.font = Font(color=STD_NAVY, bold=True, size=16)
    banner.alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = max(
        float(worksheet.row_dimensions[1].height or 0), 45.0
    )
    for row_idx in range(4, 60):
        if row_idx in (15, 16):
            continue
        worksheet.row_dimensions[row_idx].height = 20
    for row_idx in range(5, 11):
        worksheet.row_dimensions[row_idx].height = max(
            float(worksheet.row_dimensions[row_idx].height or 20), 26.0
        )
    for row_idx in (9, 10, 11, 12):
        worksheet.row_dimensions[row_idx].height = max(
            float(worksheet.row_dimensions[row_idx].height or 20), 24.0
        )
    for row_idx in range(13, 18):
        worksheet.row_dimensions[row_idx].height = max(
            float(worksheet.row_dimensions[row_idx].height or 20), 22.0
        )
    for row_idx in range(14, 20):
        worksheet.row_dimensions[row_idx].height = max(
            float(worksheet.row_dimensions[row_idx].height or 20), 22.0
        )
    worksheet.row_dimensions[15].height = 35
    worksheet.row_dimensions[16].height = 35
    worksheet.row_dimensions[5].height = max(
        float(worksheet.row_dimensions[5].height or 0), 60.0
    )
    for _rn in range(6, 10):
        worksheet.row_dimensions[_rn].height = max(
            float(worksheet.row_dimensions[_rn].height or 0), 24.0
        )
    for _rn in (10, 11, 12):
        worksheet.row_dimensions[_rn].height = max(
            float(worksheet.row_dimensions[_rn].height or 0), 36.0
        )
    worksheet.row_dimensions[21].height = max(
        float(worksheet.row_dimensions[21].height or 0), 60.0
    )
    for row_idx in range(1, 70):
        for col_idx in range(1, 12):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value and cell.font and cell.font.bold:
                cell.alignment = Alignment(
                    horizontal=(
                        cell.alignment.horizontal if cell.alignment else "center"
                    ),
                    vertical=cell.alignment.vertical if cell.alignment else "center",
                    wrap_text=True,
                )


__all__ = ["DASHBOARD_BRAND_A1", "style_dashboard"]
