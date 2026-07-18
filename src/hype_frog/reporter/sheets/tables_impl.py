from __future__ import annotations

from typing import Any

from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from hype_frog.core import get_logger
from hype_frog.reporter.help_layer import apply_semantic_aeo_tooltips
from hype_frog.reporter.engine_formatting import (
    apply_executive_priority_formatting,
    apply_fixplan_workflow_formatting,
    ensure_auto_filter,
    ensure_freeze_header,
    ensure_print_setup,
)
from hype_frog.reporter.engine_guardrails import apply_header_tooltips
from hype_frog.reporter.sheets.toc import apply_workbook_toc_and_links
from hype_frog.reporter.sheets.conditional import (
    MERGED_TAB_NAMES,
    apply_content_hub_conditional_rules,
    apply_content_planner_signoff_rules,
    apply_generic_sheet_coloring,
    apply_main_sheet_heatmaps,
    apply_merged_tabs_conditional_formatting,
    apply_psi_conditional_rules,
    apply_sheet_text_wrap_columns,
    apply_wrapped_row_heights,
    finalize_content_hub_after_normalized_headers,
)
from hype_frog.reporter.sheets.config import (
    AIOSEO_RECOMMENDATIONS_SHEET,
    CONTENT_HUB_DATA_START_ROW,
    EXECUTIVE_BRIEFING_FREEZE_PANES,
    EXECUTIVE_BRIEFING_SHEET,
    CONTENT_HUB_METRICS_SHEET,
    CONTENT_OPTIMISATION_HUB_SHEET,
    CONTENT_PLANNER_SHEET,
    DATA_HEAVY_TABS,
    DEBUG_EXCEL_ISOLATION_MODE,
    DISABLE_DATA_VALIDATION,
    DISABLE_EXTERNAL_LINKS_AND_IMAGES,
    RAG_GREEN,
    RAG_GREEN_FONT,
    DISABLE_NON_CORE_FREEZE_PANES,
    EDITABLE_INPUT_HEADER_FILL,
    EDITABLE_INPUT_HEADER_FONT,
    STD_BLUE,
    STD_NAVY,
    STD_WHITE,
)
from hype_frog.reporter.sheets.layout import (
    CONTENT_HUB_DENSITY_OVERRIDES,
    apply_column_widths,
    apply_content_hub_heading_group,
    apply_display_header_aliases,
    apply_intelligent_sorting,
    apply_main_column_group_header_tints,
    apply_main_triage_column_layout,
    hide_noisy_columns,
    reorder_columns,
    sheet_data_column_range,
)
from hype_frog.reporter.engine_rows import content_hub_column_letter
from hype_frog.reporter.sheets.links import (
    apply_editor_url_column_hyperlinks,
    style_unstyled_formula_hyperlinks,
)
from hype_frog.reporter.sheets.navigation import (
    add_return_to_briefing_strip,
    add_url_navigation_links,
    apply_cross_sheet_links,
)
from hype_frog.reporter.sheets.sheet_rows import (
    sheet_data_header_row,
    sheet_data_start_row,
)
from hype_frog.reporter.sheets.number_formats import apply_south_african_formats
from hype_frog.reporter.sheets.tables import (
    apply_mock_table_styling,
    compute_exact_table_ref,
    normalize_table_headers,
)
from hype_frog.reporter.sheets.technical import collapse_technical_deep_dive_columns
from hype_frog.reporter.sheets.validation import (
    add_all_header_tooltips,
    add_header_tooltips,
    apply_status_dropdown,
    apply_triage_status_dropdown,
)
from hype_frog.reporter.sheets.view_state import (
    apply_optimal_view_state,
    audit_freeze_merge_conflicts,
    audit_non_overlapping_merges,
    sanitize_sheet_view_selection,
    set_freeze_panes_safe,
)
from hype_frog.reporter.sheets.style_helpers import header_index

logger = get_logger(__name__)

# Structural / security / i18n diagnostic tooltips, covering the "ghost
# data" surfaced by the diagnostics-rendering pipeline. A single dict feeds
# both the Content Hub (header_row=2) and the Technical Diagnostics tab
# (header_row=1) — ``_apply_diagnostic_header_tooltips`` does header-name
# lookup so each call only attaches comments to columns that actually exist
# on the given sheet. Inlined here rather than in ``reporter.help_layer`` to
# keep this dict next to its two call sites.
_DIAGNOSTIC_HEADER_TOOLTIPS: dict[str, str] = {
    "Crawl Depth": (
        "Description: BFS hop distance from the seed URL during this audit "
        "(0 = seed). Higher values indicate pages buried deeper in the site "
        "structure, which can hurt discoverability and crawl budget."
    ),
    "Security: HSTS": (
        "Description: TRUE when a non-empty Strict-Transport-Security "
        "response header was returned, signalling enforced HTTPS to "
        "browsers (HSTS)."
    ),
    "Security: CSP": (
        "Description: TRUE when a non-empty Content-Security-Policy "
        "response header was returned, signalling browser-side defence "
        "against XSS / script-injection."
    ),
    "Anchor Text Diversity": (
        "Description: Summary of the internal-link anchor pool on this "
        "page, formatted as '<unique> unique / <total> total'. Low "
        "diversity often means heavy reuse of generic anchors "
        "(e.g. 'click here') across the internal link graph."
    ),
    "Hreflang Signals": (
        "Description: On-page hreflang cluster joined as "
        "'lang: url; lang: url'. Extraction is HTML-only — no extra "
        "network requests are issued for cluster validation."
    ),
    "JS Dependent": (
        "Description: Flags pages where JavaScript adds more than 100 "
        "words or > 50% additional content compared to the raw HTML. "
        "Strong indicator that bots without a JS renderer (some "
        "answer-engine crawlers, monitoring tools) will see a thin "
        "version of the page."
    ),
    "Raw Words": (
        "Description: Word count of the raw HTML body returned by the "
        "initial HTTP fetch, before JavaScript execution. Compare with "
        "Rendered Words to gauge JS dependency."
    ),
    "Rendered Words": (
        "Description: Word count of the body after Playwright "
        "rendering, ``networkidle`` wait, and hydration settle. The "
        "delta vs Raw Words drives the JS Dependent flag."
    ),
    "Field LCP (ms)": (
        "Description: Largest Contentful Paint in milliseconds, "
        "captured live in the browser via PerformanceObserver during "
        "the rendered fetch. Blank when the observer was blocked "
        "(CSP) or the page never reached a stable LCP."
    ),
    "Field CLS": (
        "Description: Cumulative Layout Shift captured live in the "
        "browser via PerformanceObserver during the rendered fetch. "
        "Blank when the observer was blocked or the page produced no "
        "shift events in the observation window."
    ),
    # Executive ROI tooltips. Numbers reference the constants in
    # ``hype_frog.core.scoring`` so any future re-tuning only needs to edit
    # one place; the tooltip text is duplicated here for the workbook reader.
    "Potential Traffic Lift": (
        "Description: Estimated monthly clicks recoverable by closing "
        "the AEO gap on this URL.\n"
        "Calculation: GSC Clicks * ((100 - Semantic AEO Score) / 100) "
        "* 0.25 (assumed 25% maximum AEO-driven lift).\n"
        "Returns 0 when GSC traffic or AEO score is missing."
    ),
    "AEO Visibility Gain": (
        "Description: Semantic readiness headroom on a 0-100 scale "
        "(100 - Semantic AEO Score). Higher = more upside if the AEO "
        "issues on this page are addressed."
    ),
    "Instant Priority": (
        "Description: 'CRITICAL' when GSC Clicks > 500 AND (Semantic "
        "AEO Score < 50 OR Field LCP > 2500ms). High-traffic pages "
        "with weak readiness or poor field web vitals are flagged for "
        "immediate executive attention; everything else is 'Standard'."
    ),
}


def _apply_diagnostic_header_tooltips(
    worksheet: Worksheet,
    *,
    header_row: int = 1,
) -> None:
    """Attach structural/security/i18n diagnostic header tooltips by name lookup."""
    for col_idx in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=header_row, column=col_idx)
        header = str(cell.value or "").strip()
        tooltip = _DIAGNOSTIC_HEADER_TOOLTIPS.get(header)
        if tooltip:
            cell.comment = Comment(tooltip, "hype-frog")


def _apply_link_inventory_client_polish(worksheet: Worksheet) -> None:
    """Navy header row and readable widths aligned with other data sheets."""
    header_fill = PatternFill(
        start_color=STD_NAVY,
        end_color=STD_NAVY,
        fill_type="solid",
    )
    header_font = Font(color=STD_WHITE, bold=True)
    for col_idx in range(1, 8):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    worksheet.column_dimensions["A"].width = 50.0
    worksheet.column_dimensions["B"].width = 50.0
    worksheet.column_dimensions["C"].width = 30.0


_COPY_HUB_WIDE_HEADERS: frozenset[str] = frozenset(
    {
        "Current Title",
        "Current Meta Desc",
        "H1",
        "H2",
        "H3",
        "H4",
        "H5",
        "H6",
        "Target Keywords",
    }
)


def _hub_headers_row2(worksheet) -> dict[str, int]:
    return {
        str(c.value).strip(): i for i, c in enumerate(worksheet[2], start=1) if c.value
    }


def _apply_content_hub_assigned_owner_validation(worksheet) -> None:
    if DISABLE_DATA_VALIDATION:
        return
    headers = _hub_headers_row2(worksheet)
    col = headers.get("Assigned Owner")
    if not col or worksheet.max_row < CONTENT_HUB_DATA_START_ROW:
        return
    letter = get_column_letter(col)
    dv = DataValidation(
        type="list",
        formula1='"Copy Writer,Developer,Server/Host"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid owner",
        error="Choose Copy Writer, Developer, or Server/Host.",
    )
    dv.errorStyle = "stop"
    worksheet.add_data_validation(dv)
    dv.add(f"{letter}{CONTENT_HUB_DATA_START_ROW}:{letter}{worksheet.max_row}")


def _apply_content_hub_copywriter_column_layout(worksheet) -> None:
    headers = _hub_headers_row2(worksheet)
    for name, col_idx in headers.items():
        if name in CONTENT_HUB_DENSITY_OVERRIDES:
            letter = get_column_letter(col_idx)
            worksheet.column_dimensions[letter].width = CONTENT_HUB_DENSITY_OVERRIDES[name]
            continue
        low = name.lower()
        if name in _COPY_HUB_WIDE_HEADERS or "proposed" in low:
            letter = get_column_letter(col_idx)
            cur = worksheet.column_dimensions[letter].width or 8.0
            floor = 56.0 if name == "Target Keywords" else 45.0
            worksheet.column_dimensions[letter].width = max(floor, float(cur))
    header_lower_by_col: dict[int, str] = {
        col_idx: str(name).lower() for name, col_idx in headers.items()
    }
    for r in range(CONTENT_HUB_DATA_START_ROW, worksheet.max_row + 1):
        for c in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=r, column=c)
            prev = cell.alignment
            h = prev.horizontal if prev else "left"
            hdr_low = header_lower_by_col.get(c, "")
            wrap = "proposed" in hdr_low
            cell.alignment = Alignment(horizontal=h, vertical="top", wrap_text=wrap)


def _link_main_technical_health_to_diagnostics(worksheet) -> None:
    """Keep Main Technical Health synced from Technical Diagnostics by URL lookup."""
    header_row = sheet_data_header_row("Main")
    headers = header_index(worksheet, header_row)
    url_col = headers.get("URL")
    technical_health_col = headers.get("Technical Health")
    if not url_col or not technical_health_col:
        return
    url_letter = get_column_letter(url_col)
    technical_health_letter = get_column_letter(technical_health_col)
    td_url = sheet_data_column_range("Technical Diagnostics", "URL")
    td_health = sheet_data_column_range("Technical Diagnostics", "SEO Health Score")
    data_start = sheet_data_start_row("Main")
    for row_idx in range(data_start, worksheet.max_row + 1):
        worksheet[f"{technical_health_letter}{row_idx}"] = (
            f"=IFERROR(INDEX({td_health},MATCH({url_letter}{row_idx},{td_url},0)),\"\")"
        )


def _link_hub_scores_from_main(worksheet: Worksheet) -> None:
    """Replace static Hub score copies with live INDEX/MATCH lookups on Main."""
    if worksheet.max_row < CONTENT_HUB_DATA_START_ROW:
        return
    hub_url_l = content_hub_column_letter("URL")
    score_pairs = (
        ("SEO Score", "SEO Score"),
        ("Technical Health", "Technical Health"),
        ("Copy Score", "Copy Score"),
    )
    main_url = sheet_data_column_range("Main", "URL")
    for hub_header, main_header in score_pairs:
        hub_l = content_hub_column_letter(hub_header)
        main_col = sheet_data_column_range("Main", main_header)
        for row_idx in range(CONTENT_HUB_DATA_START_ROW, worksheet.max_row + 1):
            worksheet[f"{hub_l}{row_idx}"] = (
                f"=IFERROR(INDEX({main_col},MATCH(TRIM({hub_url_l}{row_idx}),{main_url},0)),\"\")"
            )


_EMPTY_STATE_BY_SHEET: dict[str, str] = {
    "FixPlan": (
        "No issues qualified for FixPlan in this run. "
        "See Summary and Issue Register for the full issue list."
    ),
    "Quick Wins": (
        "No quick wins met the ranking threshold this run. "
        "See Summary for all flagged issues."
    ),
    "Issue Register": "No issues to report for this run. See Summary for site-wide issue counts.",
    "Priority URLs": (
        "No URLs met the priority-risk threshold this run. "
        "See Summary for site-wide issue counts."
    ),
    "Broken Link Impact": "No broken internal links detected this run — nothing to report.",
    "Snippet Opportunities": "No featured-snippet candidates identified this run.",
    "Competitor Benchmarks": "No competitor benchmark data available for this run.",
}
_EMPTY_STATE_SHEETS: frozenset[str] = frozenset(_EMPTY_STATE_BY_SHEET.keys())
_EMPTY_STATE_FILL = PatternFill(
    start_color=RAG_GREEN, end_color=RAG_GREEN, fill_type="solid"
)


def _write_empty_state_message(worksheet: Worksheet, header_row: int) -> None:
    """Replace a bare empty grid with a friendly, muted 'nothing to report' note."""
    max_col = worksheet.max_column
    if max_col < 1:
        return
    data_start = header_row + 1
    for row_idx in range(data_start, worksheet.max_row + 1):
        for col_idx in range(1, max_col + 1):
            value = worksheet.cell(row=row_idx, column=col_idx).value
            if value is not None and str(value).strip():
                return
    message = _EMPTY_STATE_BY_SHEET.get(
        worksheet.title, "No items to report for this run."
    )
    cell = worksheet.cell(row=data_start, column=1, value=message)
    cell.font = Font(italic=True, color=RAG_GREEN_FONT)
    cell.fill = _EMPTY_STATE_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    end_col = get_column_letter(max(1, min(max_col, 8)))
    if end_col != "A":
        try:
            worksheet.merge_cells(f"A{data_start}:{end_col}{data_start}")
        except Exception as exc:  # pragma: no cover - defensive merge guard
            logger.debug("Empty-state merge skipped on %s: %s", worksheet.title, exc)
    worksheet.row_dimensions[data_start].height = 20


# Headers that are editable workflow inputs — the tool seeds a default value
# but never overwrites operator edits on re-export. Marked with a distinct
# header fill so they read as "yours to fill in", not computed output.
_EDITABLE_INPUT_HEADERS_BY_SHEET: dict[str, tuple[str, ...]] = {
    "Priority URLs": ("Status", "Sprint"),
    CONTENT_OPTIMISATION_HUB_SHEET: ("Status", "Assigned Owner"),
}
_EDITABLE_INPUT_FILL = PatternFill(
    start_color=EDITABLE_INPUT_HEADER_FILL,
    end_color=EDITABLE_INPUT_HEADER_FILL,
    fill_type="solid",
)


def _mark_editable_input_headers(worksheet: Worksheet, header_row: int) -> None:
    """Tint editable-workflow-input column headers so they read as inputs."""
    headers = _EDITABLE_INPUT_HEADERS_BY_SHEET.get(worksheet.title)
    if not headers:
        return
    header_positions = header_index(worksheet, header_row)
    for header in headers:
        col_idx = header_positions.get(header)
        if not col_idx:
            continue
        cell = worksheet.cell(row=header_row, column=col_idx)
        cell.fill = _EDITABLE_INPUT_FILL
        cell.font = Font(
            bold=cell.font.bold if cell.font else True,
            color=EDITABLE_INPUT_HEADER_FONT,
        )


def adjust_sheet_format(writer: Any, sheet_name: str) -> None:
    worksheet = writer.sheets[sheet_name]
    if sheet_name == EXECUTIVE_BRIEFING_SHEET:
        set_freeze_panes_safe(worksheet, EXECUTIVE_BRIEFING_FREEZE_PANES)
        ensure_print_setup(worksheet)
        return
    reorder_columns(worksheet, sheet_name)
    add_return_to_briefing_strip(worksheet, sheet_name)
    header_row = sheet_data_header_row(sheet_name)
    if sheet_name == "Main":
        apply_main_triage_column_layout(worksheet, header_row=header_row)
    if sheet_name in {
        "FixPlan",
        "Main",
        "Technical",
        "Technical Diagnostics",
        AIOSEO_RECOMMENDATIONS_SHEET,
    }:
        apply_intelligent_sorting(worksheet, sheet_name)
    apply_generic_sheet_coloring(worksheet, sheet_name, header_row=header_row)
    if sheet_name == "Priority URLs":
        _mark_editable_input_headers(worksheet, header_row)
    apply_column_widths(worksheet)
    if sheet_name == "Main":
        _link_main_technical_health_to_diagnostics(worksheet)
        apply_main_sheet_heatmaps(worksheet, header_row=header_row)
    apply_wrapped_row_heights(worksheet)
    if sheet_name == "FixPlan":
        apply_fixplan_workflow_formatting(worksheet)
    hide_noisy_columns(worksheet, sheet_name)
    apply_south_african_formats(worksheet, header_row=header_row)
    collapse_technical_deep_dive_columns(
        worksheet, sheet_name, header_index_fn=header_index
    )
    add_url_navigation_links(writer, worksheet, sheet_name, header_row=header_row)
    apply_cross_sheet_links(writer, worksheet, sheet_name, header_row=header_row)
    if sheet_name in {AIOSEO_RECOMMENDATIONS_SHEET, "FixPlan", "Issue Register"}:
        status_col = header_index(worksheet, header_row).get("Status")
        if status_col:
            apply_status_dropdown(worksheet, status_col, header_row=header_row)
    if sheet_name == "Priority URLs":
        # Priority URLs seeds "Open" (a lightweight triage flag), not the FixPlan/
        # Hub "To Do" workflow, so it gets its own dropdown list.
        status_col = header_index(worksheet, header_row).get("Status")
        if status_col:
            apply_triage_status_dropdown(worksheet, status_col, header_row=header_row)
    if sheet_name == CONTENT_OPTIMISATION_HUB_SHEET:
        apply_content_hub_conditional_rules(worksheet, writer)
        _link_hub_scores_from_main(worksheet)
        # Semantic AEO heatmap on the Hub (Instant Priority moved to Content
        # Hub Metrics). Runs AFTER the Hub conditional pipeline so the banner
        # row insert in ``apply_content_hub_conditional_rules`` has already pushed
        # headers to row 2 / data to row 3.
        apply_executive_priority_formatting(worksheet, header_row=2)
        _mark_editable_input_headers(worksheet, header_row=2)
    elif sheet_name == CONTENT_HUB_METRICS_SHEET:
        # Content Hub Metrics carries a row-1 return banner, so the real
        # headers (incl. "Instant Priority") live on row 2.
        apply_executive_priority_formatting(worksheet, header_row=2)
    apply_sheet_text_wrap_columns(worksheet, sheet_name)
    if sheet_name in {
        CONTENT_OPTIMISATION_HUB_SHEET,
        CONTENT_HUB_METRICS_SHEET,
        AIOSEO_RECOMMENDATIONS_SHEET,
    }:
        apply_editor_url_column_hyperlinks(
            worksheet,
            sheet_name,
            disable_external_links_and_images=DISABLE_EXTERNAL_LINKS_AND_IMAGES,
        )
    if sheet_name == "PSI Performance":
        apply_psi_conditional_rules(worksheet)
    if sheet_name != CONTENT_OPTIMISATION_HUB_SHEET:
        add_all_header_tooltips(worksheet, header_row=header_row)
    if sheet_name in DATA_HEAVY_TABS and sheet_name != CONTENT_OPTIMISATION_HUB_SHEET:
        add_header_tooltips(worksheet)
    if sheet_name in {"Technical", "Main", "AEO"}:
        apply_header_tooltips(worksheet, header_row=1)
    if sheet_name == CONTENT_OPTIMISATION_HUB_SHEET:
        header_row = 2
    style_unstyled_formula_hyperlinks(worksheet, header_row=header_row)
    normalize_table_headers(worksheet, header_row=header_row)
    header_values = [
        worksheet.cell(row=header_row, column=c).value
        for c in range(1, worksheet.max_column + 1)
    ]
    valid_table_headers = all(
        isinstance(v, str) and v.strip() for v in header_values
    )
    if (
        worksheet.max_row > header_row
        and worksheet.max_column > 0
        and valid_table_headers
    ):
        ref_string = compute_exact_table_ref(worksheet, header_row)
        if ref_string:
            start_ref, end_ref = ref_string.split(":")
            min_row, min_col = coordinate_to_tuple(start_ref)
            max_row, max_col = coordinate_to_tuple(end_ref)
            apply_mock_table_styling(
                worksheet,
                min_col=min_col,
                max_col=max_col,
                min_row=min_row,
                max_row=max_row,
            )
    if sheet_name == "Main":
        apply_main_column_group_header_tints(worksheet, header_row=header_row)
    if sheet_name in MERGED_TAB_NAMES:
        apply_merged_tabs_conditional_formatting(
            worksheet, sheet_name, header_row=header_row
        )
    if sheet_name == "Technical Diagnostics":
        # Attach tooltips for the migrated diagnostic columns (Crawl Depth /
        # Security: HSTS / Security: CSP / Hreflang Signals). The helper is
        # name-keyed so it skips any header it doesn't recognise.
        _apply_diagnostic_header_tooltips(worksheet, header_row=header_row)
    if sheet_name == "Link Inventory":
        _apply_link_inventory_client_polish(worksheet)
    if sheet_name == CONTENT_PLANNER_SHEET:
        apply_content_planner_signoff_rules(worksheet)
    if sheet_name == CONTENT_OPTIMISATION_HUB_SHEET:
        finalize_content_hub_after_normalized_headers(worksheet)
        apply_display_header_aliases(worksheet, header_row=2)
        _apply_content_hub_assigned_owner_validation(worksheet)
        _apply_content_hub_copywriter_column_layout(worksheet)
        # Runs after apply_content_hub_conditional_rules (see finalize_content_hub_after_normalized_headers
        # above / CONTENT_HUB_ROW2_HEADER_COMMENTS in layout.py); any header present in both
        # engine_guardrails._HEADER_TOOLTIP_MESSAGES and CONTENT_HUB_ROW2_HEADER_COMMENTS gets this
        # dict's text. Keep the two dicts disjoint rather than relying on call order.
        apply_header_tooltips(worksheet, header_row=2)
        apply_semantic_aeo_tooltips(worksheet, header_row=2)
        _apply_diagnostic_header_tooltips(worksheet, header_row=2)
    if sheet_name in _EMPTY_STATE_SHEETS:
        _write_empty_state_message(worksheet, header_row)
    # Re-run: add_url_navigation_links/apply_cross_sheet_links (above) can append new
    # columns (e.g. Main's "Technical View", Priority URLs' "Open in Technical") after
    # the first apply_column_widths pass, leaving them at Excel's default width.
    apply_column_widths(worksheet)
    if sheet_name == CONTENT_OPTIMISATION_HUB_SHEET:
        apply_content_hub_heading_group(worksheet)
    ensure_auto_filter(worksheet)
    ensure_freeze_header(worksheet)
    ensure_print_setup(worksheet)
    apply_optimal_view_state(worksheet, sheet_name)
    sanitize_sheet_view_selection(worksheet)
    audit_non_overlapping_merges(worksheet)
    audit_freeze_merge_conflicts(worksheet)


def apply_tab_hyperlinks(writer: Any, *, hide_advanced_tabs: bool = True) -> None:
    apply_workbook_toc_and_links(
        writer,
        debug_excel_isolation_mode=DEBUG_EXCEL_ISOLATION_MODE,
        disable_non_core_freeze_panes=DISABLE_NON_CORE_FREEZE_PANES,
        std_navy=STD_NAVY,
        std_white=STD_WHITE,
        std_blue=STD_BLUE,
        hide_advanced_tabs=hide_advanced_tabs,
    )


__all__ = ["adjust_sheet_format", "apply_tab_hyperlinks", "set_freeze_panes_safe"]
