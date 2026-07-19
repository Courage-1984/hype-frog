"""Workbook tab order, colours, and default visibility (single source of truth)."""

from __future__ import annotations

from openpyxl.workbook.workbook import Workbook

from hype_frog.reporter.sheets.config import (
    AIOSEO_RECOMMENDATIONS_SHEET,
    AUDIT_RUN_DETAILS_SHEET,
    CONTENT_OPTIMISATION_HUB_SHEET,
    CONTENT_PLANNER_SHEET,
    CRAWL_LOG_SHEET,
    COMPETITOR_BENCHMARKS_SHEET,
    EXECUTIVE_BRIEFING_SHEET,
    IMAGE_INVENTORY_SHEET,
    ROBOTS_ANALYSIS_SHEET,
    SCRIPT_INVENTORY_SHEET,
    SNIPPET_OPPORTUNITIES_SHEET,
    STD_NAVY,
)

# Excel tab colours (RRGGBB — openpyxl ``tabColor``) — persona grouping (Phase 1).
# Six genuinely distinct hues (no grey-on-grey ambiguity in the tab bar); grey is
# reserved for the archival/historical group only.
TAB_COLOR_MANAGEMENT = "2C3E50"   # executive / summary / playbook (dark blue-grey)
TAB_COLOR_CONTENT = "27AE60"    # content teams (green)
TAB_COLOR_TECHNICAL = "2980B9"  # technical SEO / developers (blue)
TAB_COLOR_INVENTORY = "E67E22"  # raw crawl inventories (orange)
TAB_COLOR_ADVANCED = "8E44AD"   # hidden diagnostics (purple)
TAB_COLOR_HISTORICAL = "95A5A6" # crawl log / delta / resolved (grey — archival only)

# Backward-compatible aliases (tests and external refs during transition).
TAB_COLOR_EXECUTIVE = TAB_COLOR_MANAGEMENT
TAB_COLOR_ACTIONABLE = TAB_COLOR_TECHNICAL
TAB_COLOR_DIAGNOSTIC = TAB_COLOR_TECHNICAL
TAB_COLOR_REFERENCE = TAB_COLOR_MANAGEMENT
TAB_COLOR_PLUGIN = TAB_COLOR_TECHNICAL

# Visible left-to-right (after Table of Contents).
VISIBLE_WORKBOOK_TAB_ORDER: tuple[str, ...] = (
    "Table of Contents",
    EXECUTIVE_BRIEFING_SHEET,
    "Playbook",
    "FixPlan",
    "Quick Wins",
    "Priority URLs",
    CONTENT_OPTIMISATION_HUB_SHEET,
    "Content & AI Readiness",
    CONTENT_PLANNER_SHEET,
    "Broken Link Impact",
    ROBOTS_ANALYSIS_SHEET,
    "SitemapQA",
    "Main",
    AIOSEO_RECOMMENDATIONS_SHEET,
    "Technical Diagnostics",
    "Template & Duplication Risks",
    SNIPPET_OPPORTUNITIES_SHEET,
    "Link Intelligence",
    SCRIPT_INVENTORY_SHEET,
    IMAGE_INVENTORY_SHEET,
)

# Technical / historical tabs (hidden by default; linked from Executive Briefing + TOC).
ADVANCED_WORKBOOK_TAB_ORDER: tuple[str, ...] = (
    "Issue Register",
    "CMS Action URLs",
    "Redirects",
    CRAWL_LOG_SHEET,
    COMPETITOR_BENCHMARKS_SHEET,
    "DeltaFromPreviousRun",
    AUDIT_RUN_DETAILS_SHEET,
)

PREFERRED_WORKBOOK_TAB_ORDER: tuple[str, ...] = (
    VISIBLE_WORKBOOK_TAB_ORDER + ADVANCED_WORKBOOK_TAB_ORDER
)

_PREFERRED_TAB_SET: frozenset[str] = frozenset(PREFERRED_WORKBOOK_TAB_ORDER)

HIDDEN_SHEETS_BY_DEFAULT: frozenset[str] = frozenset(ADVANCED_WORKBOOK_TAB_ORDER)

SHEETS_EXCLUDED_FROM_TOC: frozenset[str] = frozenset()

TOC_PRIMARY_SECTION_LABEL = "Primary workflow"
TOC_ADVANCED_SECTION_LABEL = (
    "Technical & Historical (Advanced) — tabs hidden; use Open links or "
    "right-click tab bar → Unhide Sheet"
)
TOC_ADVANCED_SECTION_LABEL_SHOWN = (
    "Technical & Historical (Advanced) — tabs shown (--show-all-tabs); use Open links "
    "or right-click tab bar → Hide Sheet"
)

DASHBOARD_ADVANCED_SHEETS_NOTE = (
    "Technical & historical tabs are hidden to reduce clutter. Use the links below or "
    "Home → Format → Hide & Unhide → Unhide Sheet."
)

# (sheet_name, dashboard_label) — hyperlinks work even when the tab is hidden.
DASHBOARD_ADVANCED_SHEET_LINKS: tuple[tuple[str, str], ...] = (
    ("Issue Register", "Issue Register (canonical backlog)"),
    ("Technical Diagnostics", "Technical Diagnostics"),
    ("Content & AI Readiness", "Content & AI Readiness"),
    ("Link Intelligence", "Link Intelligence"),
    ("CMS Action URLs", "CMS Action URLs"),
    ("Redirects", "Redirects"),
    (ROBOTS_ANALYSIS_SHEET, "Robots.txt Analysis"),
    (CRAWL_LOG_SHEET, "Crawl Log"),
    (AUDIT_RUN_DETAILS_SHEET, "Audit Run Details"),
    ("DeltaFromPreviousRun", "Delta From Previous Run (includes resolved issues)"),
)

_SHEET_TAB_COLORS: dict[str, str] = {
    EXECUTIVE_BRIEFING_SHEET: TAB_COLOR_MANAGEMENT,
    "Playbook": TAB_COLOR_MANAGEMENT,
    "Priority URLs": TAB_COLOR_TECHNICAL,
    "FixPlan": TAB_COLOR_TECHNICAL,
    "Quick Wins": TAB_COLOR_TECHNICAL,
    "SitemapQA": TAB_COLOR_TECHNICAL,
    "Template & Duplication Risks": TAB_COLOR_TECHNICAL,
    AIOSEO_RECOMMENDATIONS_SHEET: TAB_COLOR_TECHNICAL,
    CONTENT_OPTIMISATION_HUB_SHEET: TAB_COLOR_CONTENT,
    CONTENT_PLANNER_SHEET: TAB_COLOR_CONTENT,
    "Main": TAB_COLOR_INVENTORY,
    "Broken Link Impact": TAB_COLOR_INVENTORY,
    "Issue Register": TAB_COLOR_ADVANCED,
    "Technical Diagnostics": TAB_COLOR_TECHNICAL,
    "Content & AI Readiness": TAB_COLOR_CONTENT,
    "Link Intelligence": TAB_COLOR_INVENTORY,
    "CMS Action URLs": TAB_COLOR_ADVANCED,
    "Redirects": TAB_COLOR_ADVANCED,
    ROBOTS_ANALYSIS_SHEET: TAB_COLOR_TECHNICAL,
    SNIPPET_OPPORTUNITIES_SHEET: TAB_COLOR_TECHNICAL,
    COMPETITOR_BENCHMARKS_SHEET: TAB_COLOR_ADVANCED,
    SCRIPT_INVENTORY_SHEET: TAB_COLOR_INVENTORY,
    IMAGE_INVENTORY_SHEET: TAB_COLOR_INVENTORY,
    CRAWL_LOG_SHEET: TAB_COLOR_HISTORICAL,
    "DeltaFromPreviousRun": TAB_COLOR_HISTORICAL,
    AUDIT_RUN_DETAILS_SHEET: TAB_COLOR_HISTORICAL,
}


def excel_sheet_link_target(name: str) -> str:
    """Escape single quotes for internal ``HYPERLINK("#'…'!A1")`` targets."""
    return str(name).replace("'", "''")


def reorder_workbook_tabs(wb: Workbook) -> None:
    """Move tabs into ``PREFERRED_WORKBOOK_TAB_ORDER``; unknown tabs append at the end."""
    desired_order = list(PREFERRED_WORKBOOK_TAB_ORDER)
    for tab_name in wb.sheetnames:
        if tab_name not in _PREFERRED_TAB_SET:
            desired_order.append(tab_name)
    for idx, tab_name in enumerate(desired_order):
        if tab_name in wb.sheetnames:
            wb.move_sheet(wb[tab_name], offset=-wb.index(wb[tab_name]) + idx)


def apply_workbook_tab_colors(wb: Workbook) -> None:
    """Apply group tab colours; Table of Contents stays default."""
    for name, rgb in _SHEET_TAB_COLORS.items():
        if name not in wb.sheetnames:
            continue
        wb[name].sheet_properties.tabColor = rgb


def apply_workbook_tab_visibility(wb: Workbook, *, hide_advanced_tabs: bool = True) -> None:
    """Hide advanced / historical tabs by default, or show every tab when disabled."""
    from hype_frog.reporter.sheets.config import EXECUTIVE_DASHBOARD_SHEET

    if EXECUTIVE_DASHBOARD_SHEET in wb.sheetnames:
        del wb[EXECUTIVE_DASHBOARD_SHEET]
    for name in wb.sheetnames:
        ws = wb[name]
        if name in HIDDEN_SHEETS_BY_DEFAULT:
            ws.sheet_state = "hidden" if hide_advanced_tabs else "visible"
        elif ws.sheet_state == "hidden" and name in VISIBLE_WORKBOOK_TAB_ORDER:
            ws.sheet_state = "visible"


# Tab the workbook should open on. The Table of Contents stays at index 0 (left-most),
# but the client lands on Executive Briefing rather than a wall of links.
WORKBOOK_LANDING_SHEET: str = EXECUTIVE_BRIEFING_SHEET


def apply_workbook_active_tab(
    wb: Workbook, *, landing_sheet: str = WORKBOOK_LANDING_SHEET
) -> None:
    """Open the workbook on ``landing_sheet`` (default Executive Briefing); TOC stays index 0.

    Falls back to the Table of Contents when the landing sheet is absent. Ensures
    exactly one tab is selected so Excel lands deterministically on the target.
    """
    target = landing_sheet if landing_sheet in wb.sheetnames else "Table of Contents"
    if target not in wb.sheetnames:
        return
    idx = wb.sheetnames.index(target)
    wb.active = idx
    for i, name in enumerate(wb.sheetnames):
        try:
            wb[name].views.sheetView[0].tabSelected = i == idx
        except (IndexError, AttributeError):
            continue


def apply_workbook_tab_layout(wb: Workbook, *, hide_advanced_tabs: bool = True) -> None:
    """Reorder tabs, apply colours, and set default visibility."""
    reorder_workbook_tabs(wb)
    apply_workbook_tab_colors(wb)
    apply_workbook_tab_visibility(wb, hide_advanced_tabs=hide_advanced_tabs)


__all__ = [
    "ADVANCED_WORKBOOK_TAB_ORDER",
    "DASHBOARD_ADVANCED_SHEET_LINKS",
    "DASHBOARD_ADVANCED_SHEETS_NOTE",
    "HIDDEN_SHEETS_BY_DEFAULT",
    "PREFERRED_WORKBOOK_TAB_ORDER",
    "TOC_ADVANCED_SECTION_LABEL",
    "TOC_ADVANCED_SECTION_LABEL_SHOWN",
    "TOC_PRIMARY_SECTION_LABEL",
    "SHEETS_EXCLUDED_FROM_TOC",
    "VISIBLE_WORKBOOK_TAB_ORDER",
    "WORKBOOK_LANDING_SHEET",
    "apply_workbook_active_tab",
    "apply_workbook_tab_colors",
    "apply_workbook_tab_layout",
    "apply_workbook_tab_visibility",
    "excel_sheet_link_target",
    "reorder_workbook_tabs",
]
