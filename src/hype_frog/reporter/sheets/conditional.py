from __future__ import annotations

import re
from typing import Any

from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.formatting.rule import (
    CellIsRule,
    ColorScaleRule,
    DataBarRule,
    FormulaRule,
)

from hype_frog.core import get_logger
from hype_frog.reporter.engine_formatting import (
    apply_global_conditional_formatting,
    apply_workflow_status_conditional_formatting,
)
from hype_frog.reporter.sheets.config import (
    CHART_DATA_SHEET,
    CONTENT_HUB_DATA_START_ROW,
    CONTENT_HUB_FREEZE_PANES,
    CONTENT_OPTIMISATION_HUB_SHEET,
    CONTENT_PLANNER_SHEET,
    DATA_BAR_BLUE,
    DEBUG_EXCEL_ISOLATION_MODE,
    DISABLE_CONDITIONAL_FORMATTING,
    DISABLE_DATA_VALIDATION,
    DISABLE_EXTERNAL_LINKS_AND_IMAGES,
    EXECUTIVE_BRIEFING_SHEET,
    HEATMAP_HIGH,
    HEATMAP_LOW,
    HEATMAP_MID,
    RAG_AMBER,
    RAG_AMBER_FONT,
    RAG_AMBER_SOFT,
    RAG_GREEN,
    RAG_GREEN_FONT,
    RAG_NEUTRAL,
    RAG_RED,
    RAG_RED_FONT,
    RAG_RED_SOFT,
    RETURN_TO_BRIEFING_LABEL,
    SEVERITY_OBSERVATION_FILL,
    SEVERITY_UNMEASURED_FILL,
    STD_BLUE,
    STD_NAVY,
    STD_WHITE,
    THEME_HEADER_BG,
    WORKBOOK_NAV_TARGET_SHEET,
    LARGE_SHEET_ROW_THRESHOLD,
    ZEBRA_BAND,
    ZEBRA_FAINT,
    HUB_BANNER_FILL,
    HUB_OWNER_COPYWRITER_FILL,
    HUB_OWNER_DEVELOPER_FILL,
    HUB_OWNER_SERVER_FILL,
    HUB_SCOPE_NOTE_FONT,
    STATUS_TODO_FILL,
    STATUS_TODO_FONT,
    status_validation_list_formula,
    HTTP_STATUS_ERROR_FONT,
    HTTP_STATUS_TIMEOUT_FONT,
)
from hype_frog.reporter.sheets.layout import (
    CONTENT_HUB_ROW2_HEADER_COMMENTS,
    PROSE_HEADERS,
    resolve_content_hub_header_row,
)
from hype_frog.reporter.sheets.links import (
    is_safe_hyperlink_target,
    sanitize_excel_url,
)
from hype_frog.reporter.sheets.style_helpers import header_index
from hype_frog.reporter.sheets.view_state import set_freeze_panes_safe
from hype_frog.reporter.sheets.workbook_layout import excel_sheet_link_target

logger = get_logger(__name__)


def apply_wrapped_row_heights(worksheet: Worksheet) -> None:
    """Increase row heights only for genuinely wrapped long-prose columns.

    URL-like columns are single-line by contract (see ``apply_column_widths``)
    and must not inflate row heights into tall strips.

    Args:
        worksheet: Worksheet to update.
    """
    from hype_frog.reporter.sheets.sheet_rows import sheet_data_header_row

    header_row = sheet_data_header_row(worksheet.title)
    if worksheet.title == CONTENT_OPTIMISATION_HUB_SHEET:
        # Self-correcting, not a hardcoded ``2``: this function runs (via
        # ``adjust_sheet_format``) before ``apply_content_hub_conditional_rules``
        # inserts the Hub's row-1 banner, so headers are still physically on row 1
        # at call time. A hardcoded row 2 silently read the first *data* row as
        # headers, matched nothing in PROSE_HEADERS, and skipped every Hub row —
        # the exact bug ``resolve_content_hub_header_row`` exists to prevent (see
        # ``apply_column_widths``, which already uses it for the same reason).
        header_row = resolve_content_hub_header_row(worksheet)
    data_start = header_row + 1
    if worksheet.max_row < data_start:
        return

    # Inflate row heights for the same prose columns ``apply_column_widths`` wraps
    # (PROSE_HEADERS), so long guidance text — Recommended Action, How To Verify,
    # Priority Reason, etc. — is not clipped when the previously hardcoded subset of
    # prose columns happens to be short on that row. URL-like columns stay single-line.
    headers = header_index(worksheet, header_row)
    wrapped_cols: list[int] = [
        col_idx
        for header_name in PROSE_HEADERS
        if (col_idx := headers.get(header_name))
    ]
    if not wrapped_cols:
        return

    for row_idx in range(data_start, worksheet.max_row + 1):
        max_lines = 1
        for col_idx in wrapped_cols:
            value = worksheet.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            text = str(value)
            if text.startswith("="):
                continue
            explicit_lines = text.count("\n") + 1
            estimated_wrap_lines = max(1, int(len(text) / 50) + 1)
            max_lines = max(max_lines, explicit_lines, estimated_wrap_lines)
        if max_lines > 1:
            worksheet.row_dimensions[row_idx].height = min(120, 15 * max_lines)


_SHEET_WRAP_TARGETS: dict[str, tuple[int, tuple[str, ...]]] = {
    "Technical": (1, ("Redirect Hops", "X-Robots-Tag", "Content-Security-Policy")),
    "AEO": (1, ("Why It Matters", "Snippet Preview Mockup")),
    # Folded in from the former standalone "Content Hub Metrics" sheet; header row 2
    # (not 1) since this is an ordinary data sheet with a row-1 return-to-briefing strip.
    "Content & AI Readiness": (2, ("Anchor Text Diversity", "Search Intent")),
    "Issue Register": (2, ("Affected URLs Sample",)),
    "Broken Link Impact": (
        2,
        ("Source Pages (first 5)", "Anchor Texts Used", "Recommended Action"),
    ),
    "Robots.txt Analysis": (2, ("Detail", "Explanation")),
}


def apply_sheet_text_wrap_columns(worksheet: Worksheet, sheet_name: str) -> None:
    """Wrap and cap row heights for high-volume narrative / policy columns."""
    if sheet_name == CONTENT_OPTIMISATION_HUB_SHEET:
        if worksheet.max_row <= 2:
            return
        headers: dict[str, int] = {}
        for c in range(1, worksheet.max_column + 1):
            v = worksheet.cell(row=2, column=c).value
            if v is not None and str(v).strip():
                headers[str(v).strip()] = c
        _hub_wrap = frozenset(
            {
                "Current Title",
                "Current Meta Desc",
                "H1",
                "H2",
                "H3",
                "H4",
                "H5",
                "H6",
                "URL Slug Normalization",
            }
        )
        for _name, col_idx in headers.items():
            if _name not in _hub_wrap and "proposed" not in str(_name).lower():
                continue
            for row_idx in range(CONTENT_HUB_DATA_START_ROW, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                prev = cell.alignment
                cell.alignment = Alignment(
                    horizontal=prev.horizontal if prev else "left",
                    vertical="top",
                    wrap_text=True,
                )
        return
    spec = _SHEET_WRAP_TARGETS.get(sheet_name)
    if not spec:
        return
    header_row, col_names = spec
    if worksheet.max_row <= header_row:
        return
    headers: dict[str, int] = {}
    for c in range(1, worksheet.max_column + 1):
        v = worksheet.cell(row=header_row, column=c).value
        if v is not None and str(v).strip():
            headers[str(v).strip()] = c
    wrap_cols = [headers[n] for n in col_names if n in headers]
    if not wrap_cols:
        return
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        max_lines = 1
        for col_idx in wrap_cols:
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(
                wrap_text=True, vertical="top", horizontal="left"
            )
            val = cell.value
            if val is None:
                continue
            text = str(val)
            explicit_lines = text.count("\n") + 1
            estimated_wrap_lines = max(1, int(len(text) / 45) + 1)
            max_lines = max(max_lines, explicit_lines, estimated_wrap_lines)
        if max_lines > 1:
            worksheet.row_dimensions[row_idx].height = min(110, 13 * max_lines)


# Headers whose semantic colouring is already owned by a conditional-formatting
# rule elsewhere (``apply_main_sheet_heatmaps`` / ``apply_merged_tabs_conditional_
# formatting``) for that specific sheet. ``apply_generic_sheet_coloring`` skips
# writing a *static* per-cell fill for these so sort/filter doesn't strand a fill
# that no longer matches the row's actual value — the CF rule follows the row,
# a static fill does not.
_CF_OWNED_COLUMNS_BY_SHEET: dict[str, frozenset[str]] = {
    "Main": frozenset({"Status Code", "Severity Badge"}),
    "Technical Diagnostics": frozenset({"Severity Badge", "Indexability Reason"}),
    "Content & AI Readiness": frozenset({"Thin Content Flag", "Title Missing"}),
    "Issue Register": frozenset({"Severity", "Status"}),
    "Template & Duplication Risks": frozenset({"Severity"}),
    "Quick Wins": frozenset({"Severity"}),
    "Broken Link Impact": frozenset({"Status Code"}),
    "FixPlan": frozenset({"Severity", "Status"}),
    "Priority URLs": frozenset({"Severity Badge", "Status", "Indexability Reason"}),
}


def _cf_owns_static_fill(sheet_name: str, header: str) -> bool:
    return header in _CF_OWNED_COLUMNS_BY_SHEET.get(sheet_name, frozenset())


_CF_ZEBRA_EXEMPT_SHEETS: frozenset[str] = frozenset(
    {
        EXECUTIVE_BRIEFING_SHEET,
        "Table of Contents",
        CONTENT_OPTIMISATION_HUB_SHEET,
        CONTENT_PLANNER_SHEET,
        CHART_DATA_SHEET,
        # Playbook gets its own per-section static colour instead (see
        # ``apply_playbook_section_colors``) — zebra striping would fight that.
        "Playbook",
    }
)

# Light, distinct fills for Playbook's "Section" groupings — rotated in first-seen
# order so each block (quick-reference topic, Issue Playbook, Glossary & Legend)
# reads as visually separate while scrolling a long reference sheet.
_PLAYBOOK_SECTION_PALETTE: tuple[str, ...] = (
    "D6E4F0",  # soft blue
    "E2F0D9",  # soft green
    "FCE4D6",  # soft orange
    "E4DFEC",  # soft purple
    "FFF2CC",  # soft yellow
    "D9E2F3",  # soft indigo
    "F2DCDB",  # soft red
    "DAEEF3",  # soft teal
    "EAD1DC",  # soft pink
    "D8E4BC",  # soft olive
)


def apply_playbook_section_colors(worksheet: Worksheet, *, header_row: int) -> None:
    """Tint each distinct Playbook ``Section`` value with its own light fill.

    Applies to whichever rows actually carry a ``Section`` value: the bracketed
    label row only for the quick-reference/Issue Playbook/Glossary blocks (their
    body rows leave ``Section`` blank), or every row for the Glossary & Legend
    block (which repeats ``Section`` on each line) — matching how the row data
    is already authored rather than inferring block boundaries.
    """
    if DISABLE_CONDITIONAL_FORMATTING:
        return
    headers = _merged_headers(worksheet, header_row)
    section_col = headers.get("Section")
    if not section_col:
        return
    last_col = worksheet.max_column
    color_by_section: dict[str, str] = {}
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        section_val = str(worksheet.cell(row=row_idx, column=section_col).value or "").strip()
        if not section_val:
            continue
        color = color_by_section.get(section_val)
        if color is None:
            color = _PLAYBOOK_SECTION_PALETTE[
                len(color_by_section) % len(_PLAYBOOK_SECTION_PALETTE)
            ]
            color_by_section[section_val] = color
        fill = PatternFill("solid", fgColor=color)
        for col_idx in range(1, last_col + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if isinstance(cell, MergedCell):
                continue
            cell.fill = fill


def _worksheet_has_cf_zebra(worksheet: Worksheet) -> bool:
    for rules in worksheet.conditional_formatting._cf_rules.values():
        for rule in rules:
            if not rule.formula:
                continue
            if any("MOD(ROW(),2)=0" in formula for formula in rule.formula):
                return True
    return False


def apply_cf_zebra_banding(
    worksheet: Worksheet,
    sheet_name: str,
    *,
    header_row: int,
) -> None:
    """Apply sort/filter-safe zebra banding via conditional formatting."""
    if DISABLE_CONDITIONAL_FORMATTING:
        return
    if sheet_name in _CF_ZEBRA_EXEMPT_SHEETS:
        return
    if _worksheet_has_cf_zebra(worksheet):
        return
    data_start = header_row + 1
    if worksheet.max_row < data_start or worksheet.max_column < 2:
        return
    data_rows = worksheet.max_row - header_row
    fill_color = ZEBRA_FAINT if data_rows > LARGE_SHEET_ROW_THRESHOLD else ZEBRA_BAND
    last_col = get_column_letter(worksheet.max_column)
    worksheet.conditional_formatting.add(
        f"A{data_start}:{last_col}{worksheet.max_row}",
        FormulaRule(
            formula=["MOD(ROW(),2)=0"],
            stopIfTrue=False,
            fill=PatternFill("solid", fgColor=fill_color),
        ),
    )


def apply_generic_sheet_coloring(
    worksheet: Worksheet, sheet_name: str, *, header_row: int = 1
) -> None:
    """Apply base per-cell semantic coloring and global conditional formatting.

    Args:
        worksheet: Worksheet to style.
        sheet_name: Current sheet name.
        header_row: 1-based header row (2 when a return strip occupies row 1).
    """
    if worksheet.max_row <= header_row:
        return

    bad_fill = PatternFill(start_color=RAG_RED, end_color=RAG_RED, fill_type="solid")
    warn_fill = PatternFill(start_color=RAG_AMBER, end_color=RAG_AMBER, fill_type="solid")
    good_fill = PatternFill(start_color=RAG_GREEN, end_color=RAG_GREEN, fill_type="solid")
    todo_fill = PatternFill(
        start_color=STATUS_TODO_FILL, end_color=STATUS_TODO_FILL, fill_type="solid"
    )
    review_fill = PatternFill(
        start_color=SEVERITY_OBSERVATION_FILL,
        end_color=SEVERITY_OBSERVATION_FILL,
        fill_type="solid",
    )
    traffic_warn_fill = PatternFill(
        start_color=RAG_AMBER_SOFT, end_color=RAG_AMBER_SOFT, fill_type="solid"
    )
    edge_fill = PatternFill(start_color=RAG_NEUTRAL, end_color=RAG_NEUTRAL, fill_type="solid")
    headers = [
        worksheet.cell(row=header_row, column=c).value
        for c in range(1, worksheet.max_column + 1)
    ]
    is_wide_sheet = worksheet.max_column >= 30

    def parse_bool(value: Any) -> bool:
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.strip().lower() in {"true", "1", "yes", "y"}
        return bool(value)

    def is_bad_header(header_name: str) -> bool:
        h = (header_name or "").lower()
        return any(
            t in h
            for t in [
                "error",
                "broken",
                "missing",
                "noindex",
                "disallow",
                "thin",
                "mixed content",
                "cross-canonical",
                "issue",
                "non-200",
                "loop",
                "out of",
            ]
        )

    def is_edge_header(header_name: str) -> bool:
        h = (header_name or "").lower()
        return any(
            t in h for t in ["redirect chain", "param url", "edge", "unresolved"]
        )

    def is_good_header(header_name: str) -> bool:
        h = (header_name or "").lower()
        return any(
            t in h
            for t in [
                "accessible",
                "match",
                "enabled",
                "complete",
                "present",
                "indexable",
                "coverage (%)",
            ]
        )

    data_start = header_row + 1
    for row_idx in range(data_start, worksheet.max_row + 1):
        # Baseline breathing room for top-aligned cells; wrapped-prose passes that
        # run later (``apply_wrapped_row_heights`` etc.) only raise this further.
        if worksheet.row_dimensions[row_idx].height is None:
            worksheet.row_dimensions[row_idx].height = 18
        row_has_issue = False
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            val = cell.value
            h = str(header) if header is not None else ""
            is_url_like = (
                "url" in h.lower()
                or h.lower().endswith("urls")
                or h.lower() in {"final url", "canonical url"}
            )
            if is_wide_sheet and is_url_like:
                cell.alignment = Alignment(
                    wrap_text=False, shrink_to_fit=True, vertical="top"
                )
            elif "url" in h.lower() or "hops" in h.lower() or "images" == h.lower():
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")

            if "%" in h:
                try:
                    pct = float(val)
                    if pct == 100.0:
                        cell.fill = good_fill
                    elif pct < 80:
                        cell.fill = bad_fill
                        row_has_issue = True
                    else:
                        cell.fill = warn_fill
                except Exception as exc:
                    logger.debug(
                        "Percent CF skipped at %s!%s (%s=%r): %s",
                        worksheet.title,
                        cell.coordinate,
                        h,
                        val,
                        exc,
                    )
            if h in {"Status Code", "Target Status (if crawled)"} and isinstance(
                val, int
            ):
                skip_fill = _cf_owns_static_fill(sheet_name, h)
                if val >= 400:
                    if not skip_fill:
                        cell.fill = bad_fill
                    row_has_issue = True
                elif val >= 300:
                    if not skip_fill:
                        cell.fill = warn_fill
                elif 200 <= val < 300:
                    if not skip_fill:
                        cell.fill = good_fill
            if isinstance(val, bool) or (
                isinstance(val, str) and val.strip().lower() in {"true", "false"}
            ):
                flag = parse_bool(val)
                skip_fill = _cf_owns_static_fill(sheet_name, h)
                if is_bad_header(h):
                    if not skip_fill:
                        cell.fill = bad_fill if flag else good_fill
                    row_has_issue = row_has_issue or flag
                elif is_good_header(h):
                    if not skip_fill:
                        cell.fill = good_fill if flag else warn_fill
                    row_has_issue = row_has_issue or (not flag)
                elif is_edge_header(h) and flag and not skip_fill:
                    cell.fill = edge_fill
            if h in {
                "Broken Internal Links Count",
                "Image Filename Quality Issues",
                "Generic Anchor Text Count",
            }:
                try:
                    if int(val or 0) > 0:
                        cell.fill = bad_fill
                        row_has_issue = True
                    else:
                        cell.fill = good_fill
                except Exception as exc:
                    logger.debug(
                        "Count CF skipped at %s!%s (%s=%r): %s",
                        worksheet.title,
                        cell.coordinate,
                        h,
                        val,
                        exc,
                    )
            if h in {"Word Count Band"} and isinstance(val, str):
                band = val.lower()
                if band == "thin":
                    cell.fill = bad_fill
                    row_has_issue = True
                elif band == "ok":
                    cell.fill = warn_fill
                elif band == "strong":
                    cell.fill = good_fill
            if h in {"Indexability Reason"} and isinstance(val, str):
                skip_fill = _cf_owns_static_fill(sheet_name, h)
                if "indexable" in val.lower() and "noindex" not in val.lower():
                    if not skip_fill:
                        cell.fill = good_fill
                else:
                    if not skip_fill:
                        cell.fill = bad_fill
                    row_has_issue = True
            if h in {"Severity", "Severity Badge"} and isinstance(val, str):
                sev = val.strip().lower()
                skip_fill = _cf_owns_static_fill(sheet_name, h)
                if sev == "critical":
                    if not skip_fill:
                        cell.fill = bad_fill
                    row_has_issue = True
                elif sev == "warning":
                    if not skip_fill:
                        cell.fill = traffic_warn_fill
                    row_has_issue = True
                elif sev in {"info", "observation"}:
                    if not skip_fill:
                        cell.fill = edge_fill
                elif sev == "pass":
                    if not skip_fill:
                        cell.fill = good_fill
            if h in {"SEO Score", "Technical Health", "Copy Score"}:
                try:
                    score = float(val)
                    if score < 70:
                        cell.fill = bad_fill
                        row_has_issue = True
                    elif score < 90:
                        cell.fill = traffic_warn_fill
                        row_has_issue = True
                    else:
                        cell.fill = good_fill
                except Exception as exc:
                    logger.debug(
                        "Score CF skipped at %s!%s (%s=%r): %s",
                        worksheet.title,
                        cell.coordinate,
                        h,
                        val,
                        exc,
                    )
            if h == "Status" and isinstance(val, str) and not _cf_owns_static_fill(
                sheet_name, h
            ):
                st = val.strip().lower()
                if st in {"done", "fixed", "closed", "completed"}:
                    cell.fill = good_fill
                elif st == "in progress":
                    cell.fill = warn_fill
                elif st in {"in review", "review"}:
                    cell.fill = review_fill
                elif st in {"to do", "todo", "open"}:
                    cell.fill = todo_fill
            if (
                h == "Direct Edit Link"
                and isinstance(val, str)
                and val.startswith(("http://", "https://"))
            ):
                if is_safe_hyperlink_target(
                    val,
                    disable_external_links_and_images=DISABLE_EXTERNAL_LINKS_AND_IMAGES,
                ):
                    # Render as a clean blue hyperlink (no dark button fill):
                    # apply_editor_url_column_hyperlinks re-applies a blue font
                    # afterwards, and a dark fill left readers with dark-on-dark.
                    cell.hyperlink = val
                    cell.style = "Hyperlink"
                    cell.font = Font(color=STD_BLUE, underline="single")
                    cell.alignment = Alignment(
                        horizontal="left", vertical="center"
                    )

        if not row_has_issue:
            worksheet.cell(row=row_idx, column=1).fill = good_fill

    if not DISABLE_CONDITIONAL_FORMATTING:
        # On Main, apply_main_sheet_heatmaps owns these columns; tell the global pass
        # to skip them so we never stack two conflicting rules on one range.
        if sheet_name == "Main":
            skip_headers = _MAIN_HEATMAP_OWNED_HEADERS
        elif sheet_name == "Broken Link Impact":
            # This sheet's "Priority Score" is clicks_total + inbound_count*10
            # (unbounded), not the 0-100-scale score this global rule assumes
            # (>=85/65-84 bands) — it gets its own data-bar-based CF instead,
            # in apply_merged_tabs_conditional_formatting.
            skip_headers = frozenset({"Priority Score"})
        else:
            skip_headers = frozenset()
        apply_global_conditional_formatting(
            worksheet,
            merged_audit_tabs=sheet_name in _MERGED_TAB_NAMES,
            skip_headers=skip_headers,
            header_row=header_row,
        )
    apply_cf_zebra_banding(worksheet, sheet_name, header_row=header_row)


def apply_content_hub_conditional_rules(worksheet: Worksheet, writer: Any) -> None:
    """Apply Content Hub-specific validations, formulas, and conditional rules.

    Args:
        worksheet: Content Optimisation Hub worksheet.
        writer: Pandas ExcelWriter-like object.
    """
    if worksheet.max_row < 2:
        return

    worksheet.insert_rows(1)
    max_col = worksheet.max_column
    instruction = (
        "CONTENT HUB (DIAGNOSTIC): Edit Title, Meta, and H1-H6 in-place; health columns "
        "and On-Page score update live. | Use Elementor link for CMS edits. | "
        "Set Status to Done when changes are live in the CMS."
    )
    security_note = (
        "NOTE: If images show '#BLOCKED!', enable external content in Excel security."
    )
    if max_col > 1:
        return_end = get_column_letter(4)
        worksheet.merge_cells(f"A1:{return_end}1")
        return_cell = worksheet["A1"]
        return_cell.value = RETURN_TO_BRIEFING_LABEL
        safe_target = excel_sheet_link_target(WORKBOOK_NAV_TARGET_SHEET)
        return_cell.hyperlink = f"#'{safe_target}'!A1"
        return_cell.font = Font(color=STD_BLUE, italic=True, underline="single")
        return_cell.alignment = Alignment(horizontal="left", vertical="center")

        # Give the security note its own highlighted cell (roughly the last third
        # of the remaining width) so it doesn't get lost inside a long pipe-joined
        # instruction string — this is the note users most often miss.
        instr_start = 5
        note_cols = max(2, (max_col - instr_start + 1) // 3) if max_col >= instr_start + 1 else 0
        note_start = max_col - note_cols + 1 if note_cols and max_col > instr_start else None
        instr_end = (note_start - 1) if note_start else max_col

        worksheet.merge_cells(
            f"{get_column_letter(instr_start)}1:{get_column_letter(instr_end)}1"
        )
        worksheet[f"{get_column_letter(instr_start)}1"] = instruction
        worksheet[f"{get_column_letter(instr_start)}1"].hyperlink = None
        worksheet[f"{get_column_letter(instr_start)}1"].font = Font(
            color=STD_NAVY, bold=True
        )
        worksheet[f"{get_column_letter(instr_start)}1"].alignment = Alignment(
            horizontal="left", vertical="center"
        )

        banner_fill = PatternFill(
            start_color=HUB_BANNER_FILL, end_color=HUB_BANNER_FILL, fill_type="solid"
        )
        note_fill = PatternFill(
            start_color=RAG_AMBER, end_color=RAG_AMBER, fill_type="solid"
        )
        for col in range(1, max_col + 1):
            worksheet.cell(row=1, column=col).fill = (
                note_fill if (note_start and col >= note_start) else banner_fill
            )

        if note_start:
            worksheet.merge_cells(
                f"{get_column_letter(note_start)}1:{get_column_letter(max_col)}1"
            )
            note_cell = worksheet[f"{get_column_letter(note_start)}1"]
            note_cell.value = security_note
            note_cell.hyperlink = None
            note_cell.font = Font(color=RAG_AMBER_FONT, bold=True)
            note_cell.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
    else:
        worksheet["A1"] = f"{instruction} | {security_note}"
        worksheet["A1"].hyperlink = None
        worksheet["A1"].fill = PatternFill(
            start_color=HUB_BANNER_FILL, end_color=HUB_BANNER_FILL, fill_type="solid"
        )
        worksheet["A1"].font = Font(color=STD_NAVY, bold=True)
        worksheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    worksheet.row_dimensions[1].height = 28
    set_freeze_panes_safe(worksheet, CONTENT_HUB_FREEZE_PANES)
    headers = {
        str(cell.value): idx
        for idx, cell in enumerate(worksheet[2], start=1)
        if cell.value
    }
    status_col = headers.get("Status")
    start_row = CONTENT_HUB_DATA_START_ROW
    end_row = worksheet.max_row

    if status_col and not DISABLE_DATA_VALIDATION and end_row >= start_row:
        dv = DataValidation(
            type="list",
            formula1=status_validation_list_formula(),
            allow_blank=True,
        )
        dv.showInputMessage = True
        dv.promptTitle = "Workflow status"
        dv.prompt = "Track workflow state: To Do → In Progress → In Review → Done"
        worksheet.add_data_validation(dv)
        dv.add(
            f"{get_column_letter(status_col)}{start_row}:{get_column_letter(status_col)}{end_row}"
        )
    if status_col and end_row >= start_row and not DISABLE_CONDITIONAL_FORMATTING:
        apply_workflow_status_conditional_formatting(
            worksheet,
            status_col,
            first_row=start_row,
            last_row=end_row,
        )
    if end_row >= start_row and not DISABLE_CONDITIONAL_FORMATTING:
        for score_header in (
            "On-Page Optimization Score",
            "SEO Score",
            "Technical Health",
            "Copy Score",
            "Semantic AEO Score",
        ):
            col_idx = headers.get(score_header)
            if not col_idx:
                continue
            letter = get_column_letter(col_idx)
            rng = f"{letter}{start_row}:{letter}{end_row}"
            worksheet.conditional_formatting.add(
                rng,
                ColorScaleRule(
                    start_type="num",
                    start_value=0,
                    start_color=HEATMAP_LOW,
                    mid_type="num",
                    mid_value=50,
                    mid_color=HEATMAP_MID,
                    end_type="num",
                    end_value=100,
                    end_color=HEATMAP_HIGH,
                ),
            )
        # Entity Density (%) is sanity-capped at 50 (see header tooltip) — a
        # 0-100 scale would compress everything into the low end, so use its
        # own realistic range instead. Citation Candidate Count is a raw count
        # (0 = no citeable blocks, worth flagging) so a data bar fits better
        # than a colour scale.
        entity_density_col = headers.get("Entity Density (%)")
        if entity_density_col:
            letter = get_column_letter(entity_density_col)
            rng = f"{letter}{start_row}:{letter}{end_row}"
            worksheet.conditional_formatting.add(
                rng,
                ColorScaleRule(
                    start_type="num",
                    start_value=0,
                    start_color=HEATMAP_LOW,
                    mid_type="num",
                    mid_value=5,
                    mid_color=HEATMAP_MID,
                    end_type="num",
                    end_value=15,
                    end_color=HEATMAP_HIGH,
                ),
            )
        citation_col = headers.get("Citation Candidate Count")
        if citation_col:
            letter = get_column_letter(citation_col)
            rng = f"{letter}{start_row}:{letter}{end_row}"
            worksheet.conditional_formatting.add(
                rng,
                CellIsRule(
                    operator="equal",
                    formula=["0"],
                    fill=PatternFill(start_color=RAG_RED, end_color=RAG_RED, fill_type="solid"),
                ),
            )
            worksheet.conditional_formatting.add(
                rng,
                DataBarRule(start_type="min", end_type="max", color=DATA_BAR_BLUE, showValue=True),
            )
    owner_col = headers.get("Assigned Owner")
    if owner_col and end_row >= start_row and not DISABLE_CONDITIONAL_FORMATTING:
        ol = get_column_letter(owner_col)
        o_rng = f"{ol}{start_row}:{ol}{end_row}"
        for needle, bg_hex, font_color in (
            ("copy writer", HUB_OWNER_COPYWRITER_FILL, STD_WHITE),
            ("developer", HUB_OWNER_DEVELOPER_FILL, STD_WHITE),
            ("server", HUB_OWNER_SERVER_FILL, STD_WHITE),
        ):
            worksheet.conditional_formatting.add(
                o_rng,
                FormulaRule(
                    formula=[
                        f'ISNUMBER(SEARCH("{needle}",LOWER(SUBSTITUTE({ol}{start_row},"/",""))))'
                    ],
                    stopIfTrue=True,
                    fill=PatternFill(
                        start_color=bg_hex, end_color=bg_hex, fill_type="solid"
                    ),
                    font=Font(color=font_color, bold=True),
                ),
            )
    action_required_col = headers.get("Action Required")
    if (
        action_required_col
        and end_row >= start_row
        and not DISABLE_CONDITIONAL_FORMATTING
    ):
        ar_letter = get_column_letter(action_required_col)
        ar_range = f"{ar_letter}{start_row}:{ar_letter}{end_row}"
        worksheet.conditional_formatting.add(
            ar_range,
            CellIsRule(
                operator="equal",
                formula=['"Needs Copy"'],
                fill=PatternFill(
                    start_color=RAG_RED, end_color=RAG_RED, fill_type="solid"
                ),
                font=Font(color=RAG_RED_FONT, bold=True),
            ),
        )
        for optimisation_literal in ('"Needs Optimisation"',):
            worksheet.conditional_formatting.add(
                ar_range,
                CellIsRule(
                    operator="equal",
                    formula=[optimisation_literal],
                    fill=PatternFill(
                        start_color=RAG_AMBER, end_color=RAG_AMBER, fill_type="solid"
                    ),
                    font=Font(color=RAG_AMBER_FONT, bold=True),
                ),
            )
        for complete_literal in ('"Complete"', '"Ready to Publish"'):
            worksheet.conditional_formatting.add(
                ar_range,
                CellIsRule(
                    operator="equal",
                    formula=[complete_literal],
                    fill=PatternFill(
                        start_color=RAG_GREEN, end_color=RAG_GREEN, fill_type="solid"
                    ),
                    font=Font(color=RAG_GREEN_FONT, bold=True),
                ),
            )

    health_headers = (
        "Title Health",
        "Meta Health",
        "H1 Health",
        "H2 Health",
        "H3 Health",
        "H4 Health",
        "H5 Health",
        "H6 Health",
    )
    if not DISABLE_CONDITIONAL_FORMATTING and end_row >= start_row:
        green_h = PatternFill(start_color=RAG_GREEN, end_color=RAG_GREEN, fill_type="solid")
        red_h = PatternFill(start_color=RAG_RED, end_color=RAG_RED, fill_type="solid")
        orange_h = PatternFill(
            start_color=RAG_AMBER, end_color=RAG_AMBER, fill_type="solid"
        )
        for hname in health_headers:
            cix = headers.get(hname)
            if not cix:
                continue
            hl = get_column_letter(cix)
            h_rng = f"{hl}{start_row}:{hl}{end_row}"
            top = f"{hl}{start_row}"
            worksheet.conditional_formatting.add(
                h_rng,
                FormulaRule(
                    formula=[
                        f'OR(ISNUMBER(SEARCH("OK",{top})),ISNUMBER(SEARCH("Perfect",{top})))'
                    ],
                    stopIfTrue=True,
                    fill=green_h,
                    font=Font(color=RAG_GREEN_FONT, bold=False),
                ),
            )
            worksheet.conditional_formatting.add(
                h_rng,
                FormulaRule(
                    formula=[
                        f'OR(ISNUMBER(SEARCH("FIX",{top})),ISNUMBER(SEARCH("MISSING",{top})))'
                    ],
                    stopIfTrue=True,
                    fill=red_h,
                    font=Font(color=RAG_RED_FONT, bold=True),
                ),
            )
            worksheet.conditional_formatting.add(
                h_rng,
                FormulaRule(
                    formula=[
                        f'OR(ISNUMBER(SEARCH("SHORT",{top})),ISNUMBER(SEARCH("LONG",{top})),'
                        f'ISNUMBER(SEARCH("REVIEW",{top})),LEFT({top},3)="Tip")'
                    ],
                    stopIfTrue=True,
                    fill=orange_h,
                    font=Font(color=RAG_AMBER_FONT, bold=True),
                ),
            )

    og_health_col = headers.get("OG Image Health")
    if og_health_col and end_row >= start_row and not DISABLE_CONDITIONAL_FORMATTING:
        health_letter = get_column_letter(og_health_col)
        health_rng = f"{health_letter}{start_row}:{health_letter}{end_row}"
        worksheet.conditional_formatting.add(
            health_rng,
            FormulaRule(
                formula=[
                    f'OR(ISNUMBER(SEARCH("Outlier",{health_letter}{start_row})),'
                    f'ISNUMBER(SEARCH("Legacy",{health_letter}{start_row})),'
                    f'ISNUMBER(SEARCH("legacy",{health_letter}{start_row})),'
                    f'ISNUMBER(SEARCH("Generic",{health_letter}{start_row})),'
                    f'ISNUMBER(SEARCH("Missing",{health_letter}{start_row})))'
                ],
                stopIfTrue=True,
                fill=PatternFill("solid", fgColor=RAG_AMBER),
                font=Font(color=RAG_AMBER_FONT, bold=True),
            ),
        )

    og_url_col = headers.get("Current OG-Image URL")
    og_preview_col = headers.get("OG Image Preview")
    if og_preview_col and og_url_col and end_row >= start_row:
        og_url_letter = get_column_letter(og_url_col)
        for r in range(start_row, end_row + 1):
            og_url_cell = worksheet.cell(row=r, column=og_url_col)
            og_url_formula = str(og_url_cell.value or "").strip()
            target_url = ""
            if og_url_formula.upper().startswith("=HYPERLINK("):
                m = re.match(r'^=HYPERLINK\("([^"]+)"\s*,', og_url_formula, re.IGNORECASE)
                if m:
                    target_url = sanitize_excel_url(m.group(1))
            else:
                target_url = sanitize_excel_url(og_url_cell.value)
            worksheet.cell(
                row=r,
                column=og_preview_col,
                value=(
                    str(target_url or "")
                    if (DEBUG_EXCEL_ISOLATION_MODE or DISABLE_EXTERNAL_LINKS_AND_IMAGES)
                    else (
                        ""
                        if not target_url
                        else f'=IF(LEN("{target_url}")>0,IFERROR(_xlfn.IMAGE("{target_url}"),"{target_url}"),"")'
                    )
                ),
            )

    url_col_idx = headers.get("URL")
    if url_col_idx and end_row >= start_row:
        link_font = Font(color=STD_BLUE, underline="single")
        for rr in range(start_row, end_row + 1):
            ucell = worksheet.cell(row=rr, column=url_col_idx)
            uv = ucell.value
            if isinstance(uv, str) and uv.strip().upper().startswith("=HYPERLINK("):
                ucell.font = link_font
    open_in_main_col_idx = headers.get("Open in Main")
    if open_in_main_col_idx and end_row >= start_row:
        link_font = Font(color=STD_BLUE, underline="single")
        for rr in range(start_row, end_row + 1):
            worksheet.cell(row=rr, column=open_in_main_col_idx).font = link_font
    og_image_url_col_idx = headers.get("Current OG-Image URL")
    if og_image_url_col_idx and end_row >= start_row:
        link_font = Font(color=STD_BLUE, underline="single")
        for rr in range(start_row, end_row + 1):
            worksheet.cell(row=rr, column=og_image_url_col_idx).font = link_font

    for hdr_name, cidx in headers.items():
        tip = CONTENT_HUB_ROW2_HEADER_COMMENTS.get(hdr_name)
        if tip:
            hcell = worksheet.cell(row=2, column=cidx)
            hcell.comment = Comment(tip, "hype-frog")

    worksheet["A1"].hyperlink = None


def finalize_content_hub_after_normalized_headers(worksheet: Worksheet) -> None:
    """Strip stray header hyperlinks after mock table (hub formulas live in row data)."""
    if worksheet.title != CONTENT_OPTIMISATION_HUB_SHEET or worksheet.max_row < 3:
        return
    for col_idx in range(1, worksheet.max_column + 1):
        hdr = worksheet.cell(row=2, column=col_idx)
        hdr.hyperlink = None
        try:
            if hdr.style and str(hdr.style) == "Hyperlink":
                hdr.style = "Normal"
        except (AttributeError, TypeError):
            pass
    # Reassert Content Hub freeze target after downstream header normalization passes.
    set_freeze_panes_safe(worksheet, CONTENT_HUB_FREEZE_PANES)


def apply_psi_conditional_rules(worksheet: Worksheet) -> None:
    """Apply PSI score/LCP-specific conditional formatting rules.

    Args:
        worksheet: PSI Performance worksheet.
    """
    if worksheet.max_row < 2:
        return

    start_row = 2
    end_row = worksheet.max_row
    if end_row < start_row:
        return

    headers = header_index(worksheet)
    mobile_lcp_col = headers.get("Mobile LCP")
    for header_name, tooltip in {
        "Mobile LCP": "Largest Contentful Paint. Target: < 2.5 seconds.",
        "Mobile CLS": "Cumulative Layout Shift. Target: < 0.1.",
    }.items():
        cidx = headers.get(header_name)
        if cidx:
            worksheet.cell(row=1, column=cidx).comment = Comment(
                tooltip, "hype-frog"
            )
    score_cols = [
        cidx
        for h, cidx in headers.items()
        if ("score" in str(h).lower())
        and ("desktop" in str(h).lower() or "mobile" in str(h).lower())
    ]
    for cidx in score_cols:
        col = get_column_letter(cidx)
        rng = f"{col}{start_row}:{col}{end_row}"
        if not DISABLE_CONDITIONAL_FORMATTING:
            worksheet.conditional_formatting.add(
                rng,
                FormulaRule(
                    formula=[f"{col}{start_row}>=90"],
                    stopIfTrue=True,
                    fill=PatternFill("solid", fgColor=RAG_GREEN),
                    font=Font(color=RAG_GREEN_FONT),
                ),
            )
            worksheet.conditional_formatting.add(
                rng,
                FormulaRule(
                    formula=[f"AND({col}{start_row}>=50,{col}{start_row}<90)"],
                    stopIfTrue=True,
                    fill=PatternFill("solid", fgColor=RAG_AMBER),
                    font=Font(color=RAG_AMBER_FONT),
                ),
            )
            worksheet.conditional_formatting.add(
                rng,
                FormulaRule(
                    formula=[f"{col}{start_row}<50"],
                    stopIfTrue=True,
                    fill=PatternFill("solid", fgColor=RAG_RED),
                    font=Font(color=RAG_RED_FONT),
                ),
            )
    if mobile_lcp_col and not DISABLE_CONDITIONAL_FORMATTING:
        col = get_column_letter(mobile_lcp_col)
        rng = f"{col}{start_row}:{col}{end_row}"
        worksheet.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f"{col}{start_row}<=2.5"],
                stopIfTrue=True,
                fill=PatternFill("solid", fgColor=RAG_GREEN),
            ),
        )
        worksheet.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[f"{col}{start_row}>4.0"],
                stopIfTrue=True,
                fill=PatternFill("solid", fgColor=RAG_RED),
            ),
        )


_MERGED_TAB_NAMES: frozenset[str] = frozenset(
    {
        "Technical Diagnostics",
        "Content & AI Readiness",
        "Link Intelligence",
        "Broken Link Impact",
        "Quick Wins",
        "Issue Register",
        "Template & Duplication Risks",
        "Priority URLs",
        "FixPlan",
        "Robots.txt Analysis",
    }
)

# Columns whose conditional formatting is owned by ``apply_main_sheet_heatmaps`` on the
# Main sheet. The global pass skips these to avoid stacking two rules on one range.
_MAIN_HEATMAP_OWNED_HEADERS: frozenset[str] = frozenset(
    {
        "Status Code",
        "SEO Health Score",
        "AEO Readiness Score",
        "Word Count",
        "Word Count (Body)",
        "Severity Badge",
    }
)

_HIGHER_BETTER_HEADERS: tuple[str, ...] = (
    "SEO Health Score",
    "Desktop PSI Score",
    "Readability (Rough Flesch)",
    "PageRank Score",
    "PageRank Percentile",
)

_LOWER_BETTER_HEADERS: tuple[str, ...] = (
    "Mobile LCP (s)",
    "Mobile TTFB (s)",
    "Mobile CLS",
    "Click Depth",
    "Generic Anchor %",
)

_DATA_BAR_HEADERS: tuple[str, ...] = (
    "Word Count",
    "Image Count",
    "Inlinks Count",
    "Internal PageRank",
    "Internal Links Count",
    "Redirect Chain Length",
    "Priority Score",
    "Inbound Link Count",
    "Business Risk Score",
)


def _merged_headers(worksheet: Worksheet, header_row: int) -> dict[str, int]:
    return {
        str(cell.value).strip(): idx
        for idx, cell in enumerate(worksheet[header_row], start=1)
        if cell.value is not None and str(cell.value).strip()
    }


def _column_range(headers: dict[str, int], header_name: str, start_row: int, end_row: int) -> str | None:
    col_idx = headers.get(header_name)
    if not col_idx or end_row < start_row:
        return None
    letter = get_column_letter(col_idx)
    return f"{letter}{start_row}:{letter}{end_row}"


def _add_color_scale_higher_better(
    worksheet: Worksheet, rng: str
) -> None:
    """Fixed 0–100 scale so SEO/PSI/readability gradients are smooth, not data-relative bands."""
    worksheet.conditional_formatting.add(
        rng,
        ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color=HEATMAP_LOW,
            mid_type="num",
            mid_value=50,
            mid_color=HEATMAP_MID,
            end_type="num",
            end_value=100,
            end_color=HEATMAP_HIGH,
        ),
    )


def _add_color_scale_lower_better(worksheet: Worksheet, rng: str) -> None:
    worksheet.conditional_formatting.add(
        rng,
        ColorScaleRule(
            start_type="min",
            start_color=HEATMAP_HIGH,
            mid_type="percentile",
            mid_value=50,
            mid_color=HEATMAP_MID,
            end_type="max",
            end_color=HEATMAP_LOW,
        ),
    )


def _add_data_bar_blue(worksheet: Worksheet, rng: str) -> None:
    worksheet.conditional_formatting.add(
        rng,
        DataBarRule(
            start_type="min",
            end_type="max",
            color=DATA_BAR_BLUE,
            showValue=True,
        ),
    )


def _add_text_semantic_highlights(
    worksheet: Worksheet,
    headers: dict[str, int],
    header_names: tuple[str, ...],
    start_row: int,
    end_row: int,
) -> None:
    """Red/pink for Critical / Non-Pass / Error; yellow/orange for Warning / Needs Work."""
    critical_fill = PatternFill(start_color=RAG_RED, end_color=RAG_RED, fill_type="solid")
    warn_fill = PatternFill(start_color=RAG_AMBER, end_color=RAG_AMBER, fill_type="solid")
    for name in header_names:
        col_idx = headers.get(name)
        if not col_idx:
            continue
        col = get_column_letter(col_idx)
        rng = f"{col}{start_row}:{col}{end_row}"
        top = f"{col}{start_row}"
        worksheet.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[
                    f'OR(NOT(ISERROR(SEARCH("Critical",{top}))),NOT(ISERROR(SEARCH("Non-Pass",{top}))),NOT(ISERROR(SEARCH("Error",{top}))))'
                ],
                stopIfTrue=True,
                fill=critical_fill,
            ),
        )
        worksheet.conditional_formatting.add(
            rng,
            FormulaRule(
                formula=[
                    f'OR(NOT(ISERROR(SEARCH("Warning",{top}))),NOT(ISERROR(SEARCH("Needs Work",{top}))))'
                ],
                stopIfTrue=True,
                fill=warn_fill,
            ),
        )


def _discrete_fill_by_value(
    worksheet: Worksheet,
    headers: dict[str, int],
    header: str,
    mapping: dict[str, str],
    start_row: int,
    end_row: int,
) -> None:
    """Exact-match CellIsRule per literal value — for columns with a small, fixed,
    real value set (workflow status, owner, risk band, etc.), as opposed to
    ``_add_text_semantic_highlights``'s substring search over free-form text."""
    col_idx = headers.get(header)
    if not col_idx:
        return
    letter = get_column_letter(col_idx)
    cell_range = f"{letter}{start_row}:{letter}{end_row}"
    for value, color in mapping.items():
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="equal",
                formula=[f'"{value}"'],
                fill=PatternFill(start_color=color, end_color=color, fill_type="solid"),
            ),
        )


def _substring_fill_by_keyword(
    worksheet: Worksheet,
    headers: dict[str, int],
    header: str,
    mapping: dict[str, str],
    start_row: int,
    end_row: int,
) -> None:
    """SEARCH()-based fill per keyword — for columns whose real values are joined
    strings (e.g. "Noindex | HTTP 404") where exact match won't hit."""
    col_idx = headers.get(header)
    if not col_idx:
        return
    letter = get_column_letter(col_idx)
    top = f"{letter}{start_row}"
    cell_range = f"{letter}{start_row}:{letter}{end_row}"
    for keyword, color in mapping.items():
        worksheet.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[f'NOT(ISERROR(SEARCH("{keyword}",{top})))'],
                stopIfTrue=True,
                fill=PatternFill(start_color=color, end_color=color, fill_type="solid"),
            ),
        )


def apply_merged_tabs_conditional_formatting(
    worksheet: Worksheet,
    sheet_name: str,
    *,
    header_row: int = 1,
) -> None:
    """Conditional formatting for Wave 3 merged workbook tabs (color scales, data bars, text highlights)."""
    if DISABLE_CONDITIONAL_FORMATTING:
        return
    if sheet_name not in _MERGED_TAB_NAMES:
        return
    if worksheet.max_row <= header_row:
        return

    headers = _merged_headers(worksheet, header_row)
    start_row = header_row + 1
    end_row = worksheet.max_row

    for hdr in _HIGHER_BETTER_HEADERS:
        rng = _column_range(headers, hdr, start_row, end_row)
        if rng:
            _add_color_scale_higher_better(worksheet, rng)

    for hdr in _LOWER_BETTER_HEADERS:
        rng = _column_range(headers, hdr, start_row, end_row)
        if rng:
            _add_color_scale_lower_better(worksheet, rng)

    for hdr in _DATA_BAR_HEADERS:
        rng = _column_range(headers, hdr, start_row, end_row)
        if rng:
            _add_data_bar_blue(worksheet, rng)

    if sheet_name == "Technical Diagnostics":
        _add_text_semantic_highlights(worksheet, headers, ("Pass Flag",), start_row, end_row)
        # Severity Badge — discrete exact-match (not substring) so Observation/
        # Unmeasured get their own colour too, matching Main's heatmap treatment
        # (previously only Critical/Warning were coloured here).
        _discrete_fill_by_value(
            worksheet, headers, "Severity Badge",
            {
                "Critical": RAG_RED,
                "Warning": RAG_AMBER,
                "Observation": SEVERITY_OBSERVATION_FILL,
                "Unmeasured": RAG_NEUTRAL,
            },
            start_row, end_row,
        )
        # Indexability Reason — confirmed bug: the old substring highlighter
        # searched for "Critical"/"Error", but this field's real values
        # (Indexable/Noindex/HTTP 4xx/5xx/Request Timeout) never contain those
        # words, so it never actually coloured anything. Match the real values.
        _substring_fill_by_keyword(
            worksheet, headers, "Indexability Reason",
            {"Noindex": RAG_RED, "HTTP 4": RAG_RED, "HTTP 5": RAG_RED, "Timeout": RAG_AMBER},
            start_row, end_row,
        )
        # SEO Health Score — unblocked here (apply_global_conditional_formatting
        # skips it globally for merged-audit-tab sheets); column position (E) is
        # preserved to honour the Main -> Technical Diagnostics VLOOKUP contract.
        seo_rng = _column_range(headers, "SEO Health Score", start_row, end_row)
        if seo_rng:
            _add_color_scale_higher_better(worksheet, seo_rng)

        for hdr in ("Critical Issues Count", "Warning Issues Count"):
            r = _column_range(headers, hdr, start_row, end_row)
            if r:
                _add_data_bar_blue(worksheet, r)
        for hdr in ("Redirect Chain Length", "Discovery Rank", "Crawl Depth"):
            r = _column_range(headers, hdr, start_row, end_row)
            if r:
                _add_color_scale_lower_better(worksheet, r)
        for hdr in (
            "Lab LCP (Mobile) (s)", "Lab TBT (Mobile) (ms)", "Lab FCP (Mobile) (s)",
            "Lab CLS (Mobile)", "Lab TTFB (Mobile) (ms)", "Lab LCP (Desktop) (s)",
            "Lab TBT (Desktop) (ms)", "Mobile LCP (s)", "Mobile CLS", "Mobile TTFB (s)",
            "Page Size (KB)", "DOM Size (nodes)", "JS Execution (ms)",
            "Network Request Count", "Origin CrUX LCP (s)", "Origin CrUX INP (ms)",
        ):
            r = _column_range(headers, hdr, start_row, end_row)
            if r:
                _add_color_scale_lower_better(worksheet, r)
        for hdr in (
            "Desktop PSI Score", "Mobile PSI Score",
            "Lighthouse Accessibility (Mobile)", "Lighthouse Best Practices (Mobile)",
            "Lighthouse SEO Score (Mobile)", "Lighthouse Performance (Desktop)",
        ):
            r = _column_range(headers, hdr, start_row, end_row)
            if r:
                _add_color_scale_higher_better(worksheet, r)
        for hdr in ("Redirect Loop Flag",):
            _bool_fill_if(
                worksheet, headers, hdr, when=True, color=RAG_RED,
                start_row=start_row, end_row=end_row,
            )
        _bool_fill_if(
            worksheet, headers, "Reachable from Homepage", when=False, color=RAG_RED,
            start_row=start_row, end_row=end_row,
        )
        for hdr in ("Security: HSTS", "Security: CSP", "Hreflang Code Valid"):
            _bool_fill_if(
                worksheet, headers, hdr, when=True, color=RAG_GREEN,
                start_row=start_row, end_row=end_row,
            )
        # Raw security header text (absent = missing hardening, not just "empty").
        for hdr in ("Strict-Transport-Security", "Content-Security-Policy", "X-Content-Type-Options"):
            col_idx = headers.get(hdr)
            if not col_idx:
                continue
            letter = get_column_letter(col_idx)
            worksheet.conditional_formatting.add(
                f"{letter}{start_row}:{letter}{end_row}",
                FormulaRule(
                    formula=[f"LEN({letter}{start_row})=0"],
                    fill=PatternFill(start_color=RAG_AMBER_SOFT, end_color=RAG_AMBER_SOFT, fill_type="solid"),
                ),
            )
        _discrete_fill_by_value(
            worksheet, headers, "Canonical Type",
            {"self": RAG_GREEN, "cross-canonical": RAG_RED, "missing": RAG_RED},
            start_row, end_row,
        )
        _discrete_fill_by_value(
            worksheet, headers, "Extraction State",
            {"complete": RAG_GREEN, "partial": RAG_AMBER, "skipped": RAG_RED},
            start_row, end_row,
        )
        _substring_fill_by_keyword(
            worksheet, headers, "GSC Coverage Category",
            {"Error": RAG_RED, "Excluded": RAG_RED, "Indexed": RAG_GREEN, "Valid": RAG_GREEN},
            start_row, end_row,
        )
        _substring_fill_by_keyword(
            worksheet, headers, "GSC Index Status",
            {"Error": RAG_RED, "Excluded": RAG_RED, "Indexed": RAG_GREEN},
            start_row, end_row,
        )
        for hdr in ("Meta Robots Raw", "X-Robots-Tag"):
            _substring_fill_by_keyword(
                worksheet, headers, hdr, {"noindex": RAG_RED}, start_row, end_row,
            )
        # "Missing Reciprocal" contains both keywords — check the bad case first
        # with stopIfTrue so it wins over the generic "Reciprocal" match below.
        recip_col = headers.get("Hreflang Reciprocal Status")
        if recip_col:
            letter = get_column_letter(recip_col)
            recip_top = f"{letter}{start_row}"
            recip_rng = f"{letter}{start_row}:{letter}{end_row}"
            worksheet.conditional_formatting.add(
                recip_rng,
                FormulaRule(
                    formula=[f'NOT(ISERROR(SEARCH("Missing",{recip_top})))'],
                    stopIfTrue=True,
                    fill=PatternFill(start_color=RAG_RED, end_color=RAG_RED, fill_type="solid"),
                ),
            )
            worksheet.conditional_formatting.add(
                recip_rng,
                FormulaRule(
                    formula=[f'NOT(ISERROR(SEARCH("Reciprocal",{recip_top})))'],
                    stopIfTrue=True,
                    fill=PatternFill(start_color=RAG_GREEN, end_color=RAG_GREEN, fill_type="solid"),
                ),
            )
    elif sheet_name == "Content & AI Readiness":
        _add_text_semantic_highlights(
            worksheet,
            headers,
            ("AEO Badge", "Thin Content Flag", "Title Missing"),
            start_row,
            end_row,
        )
        # Folded in from the former standalone "Anchor Text Audit" sheet.
        dominance_col = headers.get("Generic Anchor Dominance")
        if dominance_col:
            dominance_letter = get_column_letter(dominance_col)
            dominance_range = f"{dominance_letter}{start_row}:{dominance_letter}{end_row}"
            worksheet.conditional_formatting.add(
                dominance_range,
                CellIsRule(
                    operator="equal",
                    formula=["TRUE"],
                    fill=PatternFill(start_color=RAG_AMBER, end_color=RAG_AMBER, fill_type="solid"),
                ),
            )
    elif sheet_name == "Link Intelligence":
        _add_text_semantic_highlights(
            worksheet,
            headers,
            ("Record Type", "Crawlable"),
            start_row,
            end_row,
        )
        # Folded in from the former standalone "Link Inventory" sheet.
        _add_text_semantic_highlights(
            worksheet,
            headers,
            ("Link Type", "Generic Anchor"),
            start_row,
            end_row,
        )
        # Folded in from the former standalone "Link Equity Map" sheet.
        tier_col = headers.get("Equity Tier")
        if tier_col:
            tier_letter = get_column_letter(tier_col)
            tier_range = f"{tier_letter}{start_row}:{tier_letter}{end_row}"
            worksheet.conditional_formatting.add(
                tier_range,
                CellIsRule(
                    operator="equal",
                    formula=['"Orphan"'],
                    fill=PatternFill(start_color=RAG_RED, end_color=RAG_RED, fill_type="solid"),
                    font=Font(bold=True, color=RAG_RED_FONT),
                ),
            )
            worksheet.conditional_formatting.add(
                tier_range,
                CellIsRule(
                    operator="equal",
                    formula=['"Low"'],
                    fill=PatternFill(start_color=RAG_AMBER, end_color=RAG_AMBER, fill_type="solid"),
                ),
            )
            worksheet.conditional_formatting.add(
                tier_range,
                CellIsRule(
                    operator="equal",
                    formula=['"High"'],
                    fill=PatternFill(start_color=RAG_GREEN, end_color=RAG_GREEN, fill_type="solid"),
                ),
            )
    elif sheet_name == "Issue Register":
        days_col = headers.get("Days Open")
        if days_col:
            days_letter = get_column_letter(days_col)
            days_range = f"{days_letter}{start_row}:{days_letter}{end_row}"
            worksheet.conditional_formatting.add(
                days_range,
                CellIsRule(
                    operator="greaterThan",
                    formula=["60"],
                    fill=PatternFill(start_color=RAG_RED, end_color=RAG_RED, fill_type="solid"),
                    font=Font(bold=True, color=RAG_RED_FONT),
                ),
            )
            worksheet.conditional_formatting.add(
                days_range,
                CellIsRule(
                    operator="between",
                    formula=["31", "60"],
                    fill=PatternFill(start_color=RAG_AMBER, end_color=RAG_AMBER, fill_type="solid"),
                ),
            )
        status_col = headers.get("Status")
        if status_col and worksheet.max_row >= 2:
            last_col = get_column_letter(worksheet.max_column)
            worksheet.auto_filter.ref = f"A1:{last_col}{worksheet.max_row}"
        _add_text_semantic_highlights(
            worksheet,
            headers,
            ("Severity", "Issue", "Status", "Section"),
            start_row,
            end_row,
        )
    elif sheet_name == "Template & Duplication Risks":
        _add_text_semantic_highlights(
            worksheet,
            headers,
            ("Issue", "Severity", "Exact Action", "Risk Category"),
            start_row,
            end_row,
        )
    elif sheet_name == "Quick Wins":
        # "Business Risk Score" now gets its data bar from the generic
        # _DATA_BAR_HEADERS pass above (shared with Priority URLs) — only
        # "GSC Clicks (30d)" needs a sheet-specific rule here.
        rng = _column_range(headers, "GSC Clicks (30d)", start_row, end_row)
        if rng:
            _add_data_bar_blue(worksheet, rng)
        effort_col = headers.get("Effort (hrs)")
        if effort_col:
            effort_letter = get_column_letter(effort_col)
            effort_range = f"{effort_letter}{start_row}:{effort_letter}{end_row}"
            worksheet.conditional_formatting.add(
                effort_range,
                CellIsRule(
                    operator="lessThanOrEqual",
                    formula=["2"],
                    fill=PatternFill(start_color=RAG_GREEN, end_color=RAG_GREEN, fill_type="solid"),
                ),
            )
            worksheet.conditional_formatting.add(
                effort_range,
                CellIsRule(
                    operator="between",
                    formula=["2.01", "4"],
                    fill=PatternFill(start_color=RAG_AMBER, end_color=RAG_AMBER, fill_type="solid"),
                ),
            )
        _add_text_semantic_highlights(
            worksheet,
            headers,
            ("Severity",),
            start_row,
            end_row,
        )
        # "Issue" has no fixed value set of its own (arbitrary rule names), so
        # colour it by cross-referencing this row's own Severity cell instead —
        # gives the same at-a-glance urgency signal without duplicating the
        # substring-match approach (issue names don't contain "Critical"/"Warning").
        issue_col = headers.get("Issue")
        severity_col = headers.get("Severity")
        if issue_col and severity_col:
            issue_letter = get_column_letter(issue_col)
            severity_letter = get_column_letter(severity_col)
            issue_range = f"{issue_letter}{start_row}:{issue_letter}{end_row}"
            severity_top = f"{severity_letter}{start_row}"
            worksheet.conditional_formatting.add(
                issue_range,
                FormulaRule(
                    formula=[f'{severity_top}="Critical"'],
                    stopIfTrue=True,
                    fill=PatternFill(start_color=RAG_RED, end_color=RAG_RED, fill_type="solid"),
                ),
            )
            worksheet.conditional_formatting.add(
                issue_range,
                FormulaRule(
                    formula=[f'{severity_top}="Warning"'],
                    stopIfTrue=True,
                    fill=PatternFill(start_color=RAG_AMBER, end_color=RAG_AMBER, fill_type="solid"),
                ),
            )
        owner_col = headers.get("Owner")
        if owner_col:
            owner_letter = get_column_letter(owner_col)
            owner_range = f"{owner_letter}{start_row}:{owner_letter}{end_row}"
            for value, color in {
                "Dev": "D9E2F3",
                "Copy Writer": "E2F0D9",
                "Server/Host": "FCE4D6",
            }.items():
                worksheet.conditional_formatting.add(
                    owner_range,
                    CellIsRule(
                        operator="equal",
                        formula=[f'"{value}"'],
                        fill=PatternFill(start_color=color, end_color=color, fill_type="solid"),
                    ),
                )
        revenue_risk_col = headers.get("Revenue Risk")
        if revenue_risk_col:
            revenue_risk_letter = get_column_letter(revenue_risk_col)
            revenue_risk_range = (
                f"{revenue_risk_letter}{start_row}:{revenue_risk_letter}{end_row}"
            )
            for value, color in {
                "High Risk": RAG_RED,
                "Medium Risk": RAG_AMBER,
                "Monitor": RAG_GREEN,
            }.items():
                worksheet.conditional_formatting.add(
                    revenue_risk_range,
                    CellIsRule(
                        operator="equal",
                        formula=[f'"{value}"'],
                        fill=PatternFill(start_color=color, end_color=color, fill_type="solid"),
                    ),
                )
    elif sheet_name == "Broken Link Impact":
        status_col = headers.get("Status Code")
        if status_col:
            status_letter = get_column_letter(status_col)
            status_range = f"{status_letter}{start_row}:{status_letter}{end_row}"
            worksheet.conditional_formatting.add(
                status_range,
                CellIsRule(
                    operator="greaterThanOrEqual",
                    formula=["400"],
                    fill=PatternFill(start_color=RAG_RED, end_color=RAG_RED, fill_type="solid"),
                    font=Font(bold=True, color=RAG_RED_FONT),
                ),
            )
        # "Inbound Link Count" already gets a data bar from the generic
        # _DATA_BAR_HEADERS pass above; "Source Page Clicks Total" has no
        # generic entry (name isn't shared with any other sheet) so it needs
        # its own here.
        rng = _column_range(headers, "Source Page Clicks Total", start_row, end_row)
        if rng:
            _add_data_bar_blue(worksheet, rng)
    elif sheet_name == "FixPlan":
        # "Severity" values (Critical/Warning/Observation) match the shared
        # substring highlighter; every other FixPlan column below uses its own
        # exact, real values (To Do/Yes/High Risk/etc.), which don't contain
        # "Critical"/"Warning" text, so they need dedicated CellIsRules.
        _add_text_semantic_highlights(worksheet, headers, ("Severity",), start_row, end_row)

        for hdr in ("Affected Count", "Affected Link Instances", "Est. Sprint Points"):
            rng = _column_range(headers, hdr, start_row, end_row)
            if rng:
                _add_data_bar_blue(worksheet, rng)

        for hdr in ("Discovery Rank", "Est. Hours"):
            rng = _column_range(headers, hdr, start_row, end_row)
            if rng:
                _add_color_scale_lower_better(worksheet, rng)

        _discrete_fill_by_value(
            worksheet, headers, "Status",
            {"To Do": RAG_RED, "In Review": RAG_AMBER, "Done": RAG_GREEN},
            start_row, end_row,
        )
        _discrete_fill_by_value(
            worksheet, headers, "Action Needed",
            {"Yes": RAG_AMBER, "No": RAG_GREEN},
            start_row, end_row,
        )
        _discrete_fill_by_value(
            worksheet, headers, "Revenue Risk",
            {"High Risk": RAG_RED, "Medium Risk": RAG_AMBER, "Monitor": RAG_GREEN},
            start_row, end_row,
        )
        _discrete_fill_by_value(
            worksheet, headers, "Aging/Priority",
            {
                "Immediate (Current Sprint)": RAG_RED,
                "Next Sprint": RAG_AMBER,
                "Backlog": RAG_GREEN,
            },
            start_row, end_row,
        )
        _discrete_fill_by_value(
            worksheet, headers, "Effort",
            {"S": RAG_GREEN, "M": RAG_AMBER, "L": RAG_RED},
            start_row, end_row,
        )
        _discrete_fill_by_value(
            worksheet, headers, "Owner",
            {"Dev": "D9E2F3", "Copy Writer": "E2F0D9", "Server/Host": "FCE4D6"},
            start_row, end_row,
        )
        _discrete_fill_by_value(
            worksheet, headers, "Category",
            {"AEO": "D9E2F3", "SEO": "E2F0D9"},
            start_row, end_row,
        )
    elif sheet_name == "Robots.txt Analysis":
        # "Status" carries different vocabularies per section (Accessible/
        # Unavailable/Fetched-body-unreadable in Section 1; user-agent/
        # disallow/allow/crawl-delay/sitemap directives in Section 2; Disallow/
        # None in Section 3; In sitemap but Disallow/Sitemap not declared/None
        # in Section 4) — substring keyword matching covers all of them without
        # needing per-section CF branches.
        _substring_fill_by_keyword(
            worksheet, headers, "Status",
            {
                "Unavailable": RAG_RED,
                "unreadable": RAG_AMBER,
                "Disallow": RAG_RED,
                "not declared": RAG_AMBER,
            },
            start_row, end_row,
        )
    elif sheet_name == "Priority URLs":
        _discrete_fill_by_value(
            worksheet, headers, "Action Needed",
            {"Yes": RAG_AMBER, "No": RAG_GREEN},
            start_row, end_row,
        )
        _discrete_fill_by_value(
            worksheet, headers, "Severity Badge",
            {
                "Critical": RAG_RED,
                "Warning": RAG_AMBER,
                "Observation": SEVERITY_OBSERVATION_FILL,
                "Unmeasured": RAG_NEUTRAL,
            },
            start_row, end_row,
        )
        _discrete_fill_by_value(
            worksheet, headers, "Owner",
            {"Dev": "D9E2F3", "Copy Writer": "E2F0D9", "Server/Host": "FCE4D6"},
            start_row, end_row,
        )
        _discrete_fill_by_value(
            worksheet, headers, "Status",
            {
                "Open": RAG_AMBER,
                "In Progress": RAG_AMBER,
                "Resolved": RAG_GREEN,
                "Won't Fix": RAG_NEUTRAL,
            },
            start_row, end_row,
        )
        _discrete_fill_by_value(
            worksheet, headers, "Revenue Intent",
            {"High": RAG_RED, "Standard": RAG_NEUTRAL},
            start_row, end_row,
        )
        _substring_fill_by_keyword(
            worksheet, headers, "Indexability Reason",
            {"Noindex": RAG_RED, "HTTP 4": RAG_RED, "HTTP 5": RAG_RED, "Timeout": RAG_AMBER},
            start_row, end_row,
        )
        for hdr in ("Critical Issues Count", "Warning Issues Count", "Broken Internal Links Count"):
            rng = _column_range(headers, hdr, start_row, end_row)
            if rng:
                _add_color_scale_lower_better(worksheet, rng)
        for hdr in ("GSC Impressions", "GSC CTR"):
            rng = _column_range(headers, hdr, start_row, end_row)
            if rng:
                _add_data_bar_blue(worksheet, rng)


def apply_main_sheet_heatmaps(
    worksheet: Worksheet, *, header_row: int = 1
) -> None:
    """Apply traffic-light and data-bar formatting to Main sheet key columns."""
    if DISABLE_CONDITIONAL_FORMATTING:
        return
    if worksheet.max_row <= header_row:
        return

    start_row = header_row + 1
    end_row = worksheet.max_row
    if end_row < start_row:
        return

    headers = header_index(worksheet, header_row)

    def col_range(col_name: str) -> str | None:
        col_idx = headers.get(col_name)
        if not col_idx:
            return None
        letter = get_column_letter(col_idx)
        return f"{letter}{start_row}:{letter}{end_row}"

    red_green_scale = ColorScaleRule(
        start_type="num",
        start_value=0,
        start_color=HEATMAP_LOW,
        mid_type="num",
        mid_value=50,
        mid_color=HEATMAP_MID,
        end_type="num",
        end_value=100,
        end_color=HEATMAP_HIGH,
    )

    for col_name in (
        "SEO Health Score",
        "Mobile PSI Score",
        "Desktop PSI Score",
        "Lighthouse Performance (Mobile)",
        "Lighthouse Accessibility (Mobile)",
        "Lighthouse Best Practices (Mobile)",
        "Lighthouse SEO Score (Mobile)",
        "AEO Readiness Score",
    ):
        rng = col_range(col_name)
        if rng:
            worksheet.conditional_formatting.add(rng, red_green_scale)

    lcp_rng = col_range("Lab LCP (Mobile) (s)")
    if lcp_rng:
        worksheet.conditional_formatting.add(
            lcp_rng,
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color=HEATMAP_HIGH,
                mid_type="num",
                mid_value=2.5,
                mid_color=HEATMAP_MID,
                end_type="num",
                end_value=10.0,
                end_color=HEATMAP_LOW,
            ),
        )

    status_rng = col_range("Status Code")
    if status_rng:
        worksheet.conditional_formatting.add(
            status_rng,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=["400"],
                fill=PatternFill(start_color=RAG_RED_SOFT, end_color=RAG_RED_SOFT, fill_type="solid"),
                font=Font(bold=True, color=HTTP_STATUS_ERROR_FONT),
            ),
        )
        worksheet.conditional_formatting.add(
            status_rng,
            CellIsRule(
                operator="equal",
                formula=['"Timeout"'],
                fill=PatternFill(start_color=RAG_AMBER_SOFT, end_color=RAG_AMBER_SOFT, fill_type="solid"),
                font=Font(bold=True, color=HTTP_STATUS_TIMEOUT_FONT),
            ),
        )

    word_rng = col_range("Word Count (Body)")
    if word_rng:
        worksheet.conditional_formatting.add(
            word_rng,
            DataBarRule(
                start_type="num",
                start_value=0,
                end_type="num",
                end_value=2000,
                color=DATA_BAR_BLUE,
            ),
        )

    eeat_rng = col_range("E-E-A-T Signal Score")
    if eeat_rng:
        worksheet.conditional_formatting.add(
            eeat_rng,
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color=HEATMAP_LOW,
                mid_type="num",
                mid_value=5,
                mid_color=HEATMAP_MID,
                end_type="num",
                end_value=10,
                end_color=HEATMAP_HIGH,
            ),
        )

    schema_err_rng = col_range("Schema Error Count")
    if schema_err_rng:
        worksheet.conditional_formatting.add(
            schema_err_rng,
            CellIsRule(
                operator="greaterThan",
                formula=["0"],
                fill=PatternFill(start_color=RAG_RED, end_color=RAG_RED, fill_type="solid"),
            ),
        )
        worksheet.conditional_formatting.add(
            schema_err_rng,
            CellIsRule(
                operator="equal",
                formula=["0"],
                fill=PatternFill(start_color=RAG_GREEN, end_color=RAG_GREEN, fill_type="solid"),
            ),
        )

    depth_rng = col_range("Click Depth")
    if depth_rng:
        worksheet.conditional_formatting.add(
            depth_rng,
            CellIsRule(
                operator="equal",
                formula=["-1"],
                fill=PatternFill(start_color=RAG_AMBER_SOFT, end_color=RAG_AMBER_SOFT, fill_type="solid"),
                font=Font(italic=True),
            ),
        )

    page_size_rng = col_range("Page Size (KB)")
    if page_size_rng:
        worksheet.conditional_formatting.add(
            page_size_rng,
            CellIsRule(
                operator="greaterThan",
                formula=["1024"],
                fill=PatternFill(start_color=RAG_RED, end_color=RAG_RED, fill_type="solid"),
            ),
        )

    badge_rng = col_range("Severity Badge")
    if badge_rng:
        for val, colour in (
            ("Critical", RAG_RED_SOFT),
            ("Warning", RAG_AMBER_SOFT),
            ("Observation", SEVERITY_OBSERVATION_FILL),
            ("Unmeasured", SEVERITY_UNMEASURED_FILL),
        ):
            worksheet.conditional_formatting.add(
                badge_rng,
                CellIsRule(
                    operator="equal",
                    formula=[f'"{val}"'],
                    fill=PatternFill(start_color=colour, end_color=colour, fill_type="solid"),
                ),
            )

    for col_name in ("Is Thin Content", "Is Near Duplicate", "Is Draft or Test Page"):
        flag_rng = col_range(col_name)
        if flag_rng:
            worksheet.conditional_formatting.add(
                flag_rng,
                CellIsRule(
                    operator="equal",
                    formula=["TRUE"],
                    fill=PatternFill(start_color=RAG_RED, end_color=RAG_RED, fill_type="solid"),
                ),
            )

    age_rng = col_range("Content Age (days)")
    if age_rng:
        worksheet.conditional_formatting.add(
            age_rng,
            CellIsRule(
                operator="greaterThan",
                formula=["730"],
                fill=PatternFill(start_color=RAG_RED, end_color=RAG_RED, fill_type="solid"),
            ),
        )
        worksheet.conditional_formatting.add(
            age_rng,
            CellIsRule(
                operator="greaterThan",
                formula=["365"],
                fill=PatternFill(start_color=RAG_AMBER_SOFT, end_color=RAG_AMBER_SOFT, fill_type="solid"),
            ),
        )

    meta_desc_length_col = headers.get("Meta Desc Length")
    if meta_desc_length_col:
        mdl_col_letter = get_column_letter(meta_desc_length_col)
        mdl_range = f"{mdl_col_letter}{start_row}:{mdl_col_letter}{end_row}"
        worksheet.conditional_formatting.add(
            mdl_range,
            CellIsRule(
                operator="greaterThan",
                formula=["160"],
                fill=PatternFill(
                    start_color=RAG_RED, end_color=RAG_RED, fill_type="solid"
                ),
                font=Font(color=RAG_RED_FONT),
            ),
        )


# Headers apply_main_sheet_heatmaps() (or the generic static-fill pass in
# apply_generic_sheet_coloring, keyed by the exact same names) already owns —
# skipped here to avoid stacking a second conflicting rule on the same range.
_MAIN_GROUP_CF_ALREADY_OWNED: frozenset[str] = frozenset(
    {
        "Status Code", "SEO Health Score", "AEO Readiness Score", "Word Count",
        "Word Count (Body)", "Severity Badge", "Mobile PSI Score", "Desktop PSI Score",
        "Lighthouse Performance (Mobile)", "Lighthouse Accessibility (Mobile)",
        "Lighthouse Best Practices (Mobile)", "Lighthouse SEO Score (Mobile)",
        "Lab LCP (Mobile) (s)", "E-E-A-T Signal Score", "Schema Error Count",
        "Click Depth", "Page Size (KB)", "Is Thin Content", "Is Near Duplicate",
        "Is Draft or Test Page", "Content Age (days)", "Meta Desc Length",
        "Load Time (s)", "Load Time", "TTFB (ms)",
        "Broken Internal Links Count", "Image Filename Quality Issues",
        "Generic Anchor Text Count", "Word Count Band",
        "Technical Health", "Copy Score", "SEO Score",
    }
)


def _bool_fill_if(
    worksheet: Worksheet,
    headers: dict[str, int],
    header: str,
    *,
    when: bool,
    color: str,
    start_row: int,
    end_row: int,
) -> None:
    """Fill a boolean column's cells when its value equals ``when``."""
    col_idx = headers.get(header)
    if not col_idx:
        return
    letter = get_column_letter(col_idx)
    rng = f"{letter}{start_row}:{letter}{end_row}"
    worksheet.conditional_formatting.add(
        rng,
        CellIsRule(
            operator="equal",
            formula=["TRUE" if when else "FALSE"],
            fill=PatternFill(start_color=color, end_color=color, fill_type="solid"),
        ),
    )


def apply_main_group_conditional_formatting(
    worksheet: Worksheet, *, header_row: int = 1
) -> None:
    """Group-based CF rollout for Main's ~150 columns beyond the always-visible
    triage block (which apply_main_sheet_heatmaps already owns).

    One rule pattern per column-type, organised by MAIN_COLUMN_GROUP_DEFINITIONS:
    0-100 scores -> fixed colour scale; unbounded numeric metrics (seconds/ms/
    counts) -> data-relative colour scale or data bar; booleans -> red/green;
    known enums -> discrete FormulaRule sets matching real values (not
    substring keyword matching, which is how Technical Diagnostics' old
    Indexability Reason CF went wrong for its actual values).
    """
    if DISABLE_CONDITIONAL_FORMATTING:
        return
    if worksheet.max_row <= header_row:
        return
    start_row = header_row + 1
    end_row = worksheet.max_row
    if end_row < start_row:
        return

    headers = header_index(worksheet, header_row)

    def rng(header: str) -> str | None:
        if header in _MAIN_GROUP_CF_ALREADY_OWNED:
            return None
        return _column_range(headers, header, start_row, end_row)

    def higher_100(header: str) -> None:
        r = rng(header)
        if r:
            _add_color_scale_higher_better(worksheet, r)

    def lower_relative(header: str) -> None:
        r = rng(header)
        if r:
            _add_color_scale_lower_better(worksheet, r)

    def data_bar(header: str) -> None:
        r = rng(header)
        if r:
            _add_data_bar_blue(worksheet, r)

    def good_if_true(header: str) -> None:
        if header not in _MAIN_GROUP_CF_ALREADY_OWNED:
            _bool_fill_if(
                worksheet, headers, header, when=True, color=RAG_GREEN,
                start_row=start_row, end_row=end_row,
            )

    def bad_if_true(header: str) -> None:
        if header not in _MAIN_GROUP_CF_ALREADY_OWNED:
            _bool_fill_if(
                worksheet, headers, header, when=True, color=RAG_RED,
                start_row=start_row, end_row=end_row,
            )

    # --- Metadata Group ---
    title_len_col = headers.get("Title Length")
    if title_len_col:
        letter = get_column_letter(title_len_col)
        title_rng = f"{letter}{start_row}:{letter}{end_row}"
        title_cell = f"{letter}{start_row}"
        worksheet.conditional_formatting.add(
            title_rng,
            FormulaRule(
                formula=[f"OR({title_cell}<50,{title_cell}>60)"],
                stopIfTrue=True,
                fill=PatternFill(start_color=RAG_AMBER_SOFT, end_color=RAG_AMBER_SOFT, fill_type="solid"),
            ),
        )
        worksheet.conditional_formatting.add(
            title_rng,
            FormulaRule(
                formula=[f"AND({title_cell}>=50,{title_cell}<=60)"],
                stopIfTrue=True,
                fill=PatternFill(start_color=RAG_GREEN, end_color=RAG_GREEN, fill_type="solid"),
            ),
        )

    # --- Heading Structure Group ---
    h1_col = headers.get("H1 Content")
    if h1_col:
        letter = get_column_letter(h1_col)
        worksheet.conditional_formatting.add(
            f"{letter}{start_row}:{letter}{end_row}",
            FormulaRule(
                formula=[f"LEN({letter}{start_row})=0"],
                fill=PatternFill(start_color=RAG_RED, end_color=RAG_RED, fill_type="solid"),
            ),
        )
    for h in ("H2 Content", "H3 Content", "H4 Content", "H5 Content", "H6 Content"):
        col_idx = headers.get(h)
        if not col_idx:
            continue
        letter = get_column_letter(col_idx)
        worksheet.conditional_formatting.add(
            f"{letter}{start_row}:{letter}{end_row}",
            FormulaRule(
                formula=[f"LEN({letter}{start_row})=0"],
                fill=PatternFill(start_color=RAG_AMBER_SOFT, end_color=RAG_AMBER_SOFT, fill_type="solid"),
            ),
        )
    for h in ("H1 Length", "H2 Length", "H3 Length", "H4 Length", "H5 Length", "H6 Length"):
        data_bar(h)

    # --- Performance & CWV Group ---
    for h in (
        "Lighthouse Performance (Desktop)", "Lighthouse Accessibility (Desktop)",
        "Lighthouse Best Practices (Desktop)", "Lighthouse SEO Score (Desktop)",
        "Regional Authority Score",
    ):
        higher_100(h)
    for h in (
        "CWV LCP (s)", "CWV INP (ms)", "CWV FCP (ms)", "CWV TTFB (ms)",
        "Origin CrUX LCP (s)", "Origin CrUX CLS", "Origin CrUX INP (ms)",
        "Mobile LCP (s)", "Mobile CLS", "Mobile TTFB (s)",
        "Lab CLS (Mobile)", "Lab TBT (Mobile) (ms)", "Lab INP (Mobile) (ms)",
        "Lab FCP (Mobile) (s)", "Lab Speed Index (Mobile) (s)", "Lab TTI (Mobile) (s)",
        "Lab TTFB (Mobile) (ms)", "Lab LCP (Desktop) (s)", "Lab CLS (Desktop)",
        "Lab TBT (Desktop) (ms)", "Lab INP (Desktop) (ms)", "Lab FCP (Desktop) (s)",
        "Lab Speed Index (Desktop) (s)", "Lab TTI (Desktop) (s)", "Lab TTFB (Desktop) (ms)",
        "DOM Size (nodes)", "JS Execution (ms)", "Network Request Count",
    ):
        lower_relative(h)
    for h in ("Has Text Compression", "Uses Modern Image Formats", "Reachable from Homepage"):
        good_if_true(h)
    for h in ("Has Long Cache TTL Issues", "Has Render Blocking Resources"):
        bad_if_true(h)

    # --- Google Search Console Group ---
    _substring_fill_by_keyword(
        worksheet, headers, "GSC Index Status",
        {"Excluded": RAG_RED, "Error": RAG_RED, "Indexed": RAG_GREEN},
        start_row, end_row,
    )
    for h in ("GSC Clicks", "GSC Impressions", "GSC CTR"):
        data_bar(h)
    lower_relative("GSC Avg Position")
    lower_relative("Days Since Last Crawl")

    # --- Crawl & Discovery Group ---
    bad_if_true("Orphan Pages")
    data_bar("Internal PageRank")
    for h in ("Found via Sitemap", "Found via Crawl"):
        good_if_true(h)

    # --- Raw State Group ---
    _discrete_fill_by_value(
        worksheet, headers, "Extraction State",
        {"complete": RAG_GREEN, "partial": RAG_AMBER, "skipped": RAG_RED},
        start_row, end_row,
    )

    # --- Schema & Structured Data ---
    for h in ("Schema Present", "Schema Valid", "Has Valid JSON-LD"):
        good_if_true(h)
    data_bar("Schema Warning Count")
    _substring_fill_by_keyword(
        worksheet, headers, "Schema Types With Errors",
        {"error": RAG_RED},
        start_row, end_row,
    )

    # --- E-E-A-T & Trust Signals ---
    for h in (
        "Has Byline Element", "Has Time Element", "Has Privacy Policy Link",
        "Has Terms Link", "Has Social Links", "Has Phone Number",
        "Has Email Address", "Has Authority External Links", "Links to About Page",
    ):
        good_if_true(h)
    data_bar("Social Profile Link Count")

    # --- Content Quality ---
    lower_relative("Content Similarity Score")
    for h in ("Thin Content Flag", "Probable Duplicate Flag"):
        bad_if_true(h)

    # --- Social Cards ---
    for h in ("OG Image OK", "OG Image Dimensions OK", "Open Graph Complete"):
        good_if_true(h)
    bad_if_true("OG URL Mismatch")
    higher_100("OG Completeness Score")

    # --- Redirects ---
    for h in ("Redirect Chain Length", "Redirect Chain Hops"):
        lower_relative(h)
    for h in ("Has 302 in Chain", "Has Mixed Redirect Types", "Redirect Loop Flag"):
        bad_if_true(h)

    # --- Canonical Chain ---
    _discrete_fill_by_value(
        worksheet, headers, "Canonical Type",
        {"self": RAG_GREEN, "cross-canonical": RAG_RED, "missing": RAG_RED},
        start_row, end_row,
    )
    lower_relative("Canonical Chain Depth")
    for h in ("Canonical Loop Detected", "Canonical Points to Redirect", "Canonical Points to Non-200"):
        bad_if_true(h)

    # --- Robots.txt ---
    for _agent, column in (
        ("Googlebot", "Robots.txt: Googlebot"),
        ("Bingbot", "Robots.txt: Bingbot"),
        ("GPTBot", "Robots.txt: GPTBot"),
        ("ClaudeBot", "Robots.txt: ClaudeBot"),
        ("PerplexityBot", "Robots.txt: PerplexityBot"),
        ("CCBot", "Robots.txt: CCBot"),
    ):
        _discrete_fill_by_value(
            worksheet, headers, column,
            {"Allow": RAG_GREEN, "Disallow": RAG_RED},
            start_row, end_row,
        )
    good_if_true("Robots.txt Accessible")


def apply_dashboard_metric_conditional_rules(worksheet: Worksheet) -> None:
    """Apply at-a-glance color scales to primary dashboard metric values."""
    if DISABLE_CONDITIONAL_FORMATTING:
        return
    if worksheet.title != "Dashboard":
        return
    if worksheet.max_row < 8:
        return

    completion_scale = dict(
        start_type="num",
        start_value=0,
        start_color=HEATMAP_LOW,
        mid_type="num",
        mid_value=0.5,
        mid_color=HEATMAP_MID,
        end_type="num",
        end_value=1,
        end_color=HEATMAP_HIGH,
    )
    worksheet.conditional_formatting.add("B5:B7", ColorScaleRule(**completion_scale))
    worksheet.conditional_formatting.add(
        "B8:B8",
        CellIsRule(
            operator="lessThan",
            formula=["-10"],
            fill=PatternFill(start_color=RAG_RED, end_color=RAG_RED, fill_type="solid"),
        ),
    )
    worksheet.conditional_formatting.add(
        "B8:B8",
        CellIsRule(
            operator="lessThan",
            formula=["0"],
            font=Font(color=RAG_RED_FONT, bold=True),
        ),
    )
    worksheet.conditional_formatting.add("B17:B17", ColorScaleRule(**completion_scale))
    worksheet.conditional_formatting.add("B22:B22", ColorScaleRule(**completion_scale))
    worksheet.conditional_formatting.add(
        "B20:B20",
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            stopIfTrue=True,
            fill=PatternFill(start_color=RAG_RED, end_color=RAG_RED, fill_type="solid"),
            font=Font(color=RAG_RED_FONT, bold=True),
        ),
    )


_CONTENT_PLANNER_COL_WIDTHS: dict[str, float] = {
    "Primary": 24.0,
    "Secondary": 24.0,
    "Tertiary": 24.0,
    "Page link": 52.0,
    "Copy Doc": 30.0,
    "Priority for MVP": 18.0,
    "Copywriter Sign off": 20.0,
    "Copy First Check": 18.0,
    "2nd Revisions": 16.0,
    "Client copy sign off": 20.0,
    "Web design off": 18.0,
    "UXI sign off": 16.0,
    "Visual Design sign off": 20.0,
    "Client final sign off": 20.0,
    "Optimisations": 16.0,
    "Desktop": 14.0,
    "Tablet": 14.0,
    "Mobile": 14.0,
    "SEO": 14.0,
    "Performance": 16.0,
    "Plugin Audit": 18.0,
}

# Teal accent used for the hierarchy (Primary/Secondary/Tertiary) header cells.
_PLANNER_TEAL: str = "BFE9E4"
_PLANNER_TEAL_FONT: str = "1A4A47"
# Soft amber for the Copy Doc workflow column header.
_PLANNER_COPYDOC_HDR: str = "FFF3CD"
_PLANNER_COPYDOC_HDR_FONT: str = "7A5C00"
# Soft indigo band for sign-off / QA column headers.
_PLANNER_SIGNOFF_HDR: str = "E8EAF6"
_PLANNER_SIGNOFF_HDR_FONT: str = "1A237E"
_CONTENT_PLANNER_SIGNOFF_STATUS: str = "Not signed off"
# Column indices (1-based) for orchestration.content_planner.CONTENT_PLANNER_SIGNOFF_COLUMNS
# (CONTENT_PLANNER_COLUMNS[5:] — "Priority for MVP" through "Plugin Audit", 16
# columns: F through U). Kept as separate constants (not derived from the column
# tuple, since reporter/ must not import orchestration/) — update both in lockstep
# with any future column insertion/removal in that slice.
_CONTENT_PLANNER_SIGNOFF_FIRST_COL: int = 6
_CONTENT_PLANNER_SIGNOFF_LAST_COL: int = 21


def _apply_content_planner_header_accents(
    worksheet: Worksheet, headers: dict[str, int], *, header_row: int
) -> None:
    """Reapply section header colours after mock table styling."""
    for hier_name in ("Primary", "Secondary", "Tertiary"):
        col_idx = headers.get(hier_name)
        if not col_idx:
            continue
        hdr = worksheet.cell(row=header_row, column=col_idx)
        hdr.fill = PatternFill(
            start_color=_PLANNER_TEAL, end_color=_PLANNER_TEAL, fill_type="solid"
        )
        hdr.font = Font(color=_PLANNER_TEAL_FONT, bold=True)

    copy_doc_col = headers.get("Copy Doc")
    if copy_doc_col:
        hdr = worksheet.cell(row=header_row, column=copy_doc_col)
        hdr.fill = PatternFill(
            start_color=_PLANNER_COPYDOC_HDR,
            end_color=_PLANNER_COPYDOC_HDR,
            fill_type="solid",
        )
        hdr.font = Font(color=_PLANNER_COPYDOC_HDR_FONT, bold=True)

    for col_idx in range(_CONTENT_PLANNER_SIGNOFF_FIRST_COL, _CONTENT_PLANNER_SIGNOFF_LAST_COL + 1):
        hdr = worksheet.cell(row=header_row, column=col_idx)
        hdr.fill = PatternFill(
            start_color=_PLANNER_SIGNOFF_HDR,
            end_color=_PLANNER_SIGNOFF_HDR,
            fill_type="solid",
        )
        hdr.font = Font(color=_PLANNER_SIGNOFF_HDR_FONT, bold=True)


def apply_content_planner_signoff_rules(
    worksheet: Worksheet, *, header_row: int = 2
) -> None:
    """Column widths, row heights, hyperlinks, and RAG sign-off formatting.

    Column layout: A=Primary, B=Secondary, C=Tertiary, D=Page link, E=Copy Doc,
    F-T = 16 sign-off/QA columns (Priority for MVP, Copywriter Sign off …
    Performance, Plugin Audit). ``header_row`` defaults to 2 because Content
    Planner is not return-strip-exempt — export always inserts a row-1 banner
    before this function runs, pushing real headers to row 2 (previously this
    function read row 1 unconditionally, silently no-op'ing on the banner text
    instead of the real headers).
    Freeze panes at ``E{header_row+1}`` lock columns A–D while scrolling the
    workflow grid.
    """
    if worksheet.max_row <= header_row:
        return

    headers = header_index(worksheet, header_row)
    last_row = worksheet.max_row
    data_start = header_row + 1
    signoff_col_indices = list(
        range(_CONTENT_PLANNER_SIGNOFF_FIRST_COL, _CONTENT_PLANNER_SIGNOFF_LAST_COL + 1)
    )

    # ── Column widths ────────────────────────────────────────────────────────
    for col_name, width in _CONTENT_PLANNER_COL_WIDTHS.items():
        col_idx = headers.get(col_name)
        if col_idx:
            worksheet.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Header row: height + center alignment ────────────────────────────────
    worksheet.row_dimensions[header_row].height = 42
    for col_idx in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=header_row, column=col_idx)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    _apply_content_planner_header_accents(worksheet, headers, header_row=header_row)

    # ── Hierarchy columns: bold populated labels ─────────────────────────────
    for hier_name in ("Primary", "Secondary", "Tertiary"):
        col_idx = headers.get(hier_name)
        if not col_idx:
            continue
        for row_idx in range(data_start, last_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if cell.value:
                cell.font = Font(bold=True)

    # ── Page link column: hyperlinks + shrink-to-fit ────────────────────────
    page_link_col = headers.get("Page link")
    if page_link_col:
        link_font = Font(color=STD_BLUE, underline="single")
        for row_idx in range(data_start, last_row + 1):
            cell = worksheet.cell(row=row_idx, column=page_link_col)
            url_val = str(cell.value or "").strip()
            if url_val.startswith(("http://", "https://")):
                cell.hyperlink = url_val
                cell.font = link_font
            cell.alignment = Alignment(
                wrap_text=False, shrink_to_fit=True, vertical="center"
            )

    # ── Copy Doc column: left-aligned placeholder hint ─────────────────────
    copy_doc_col = headers.get("Copy Doc")
    if copy_doc_col:
        placeholder_font = Font(color="808080", italic=True)
        for row_idx in range(data_start, last_row + 1):
            cell = worksheet.cell(row=row_idx, column=copy_doc_col)
            if cell.value is None or str(cell.value).strip() == "":
                cell.value = "Paste doc link"
                cell.font = placeholder_font
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # ── Sign-off data cells: default status + centre alignment ───────────────
    for row_idx in range(data_start, last_row + 1):
        worksheet.row_dimensions[row_idx].height = 22
        for col_idx in signoff_col_indices:
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value is None or str(cell.value).strip() == "":
                cell.value = _CONTENT_PLANNER_SIGNOFF_STATUS
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── DataValidation for sign-off columns ───────────────────────────────
    if not DISABLE_DATA_VALIDATION:
        dv = DataValidation(
            type="list",
            formula1='"Signed off,In progress,Not signed off"',
            allow_blank=False,
        )
        dv.showErrorMessage = True
        dv.errorTitle = "Invalid status"
        dv.error = "Select: Signed off, In progress, or Not signed off."
        worksheet.add_data_validation(dv)
        for col_idx in signoff_col_indices:
            letter = get_column_letter(col_idx)
            dv.add(f"{letter}{data_start}:{letter}{last_row}")

    # ── Zebra banding: identity/copy columns only ────────────────────────────
    # The sign-off columns (below) already carry their own RAG fills, so a
    # second zebra layer there would visually fight that colouring. Columns
    # before the sign-off block (Primary/Secondary/Tertiary/Page link/Copy Doc)
    # have no row-differentiation otherwise, unlike every other large data
    # sheet in the workbook, which makes it easy to lose your place scrolling
    # a long workflow list.
    if not DISABLE_CONDITIONAL_FORMATTING:
        identity_last_col = get_column_letter(_CONTENT_PLANNER_SIGNOFF_FIRST_COL - 1)
        data_rows = last_row - header_row
        zebra_fill_color = ZEBRA_FAINT if data_rows > LARGE_SHEET_ROW_THRESHOLD else ZEBRA_BAND
        worksheet.conditional_formatting.add(
            f"A{data_start}:{identity_last_col}{last_row}",
            FormulaRule(
                formula=["MOD(ROW(),2)=0"],
                stopIfTrue=False,
                fill=PatternFill("solid", fgColor=zebra_fill_color),
            ),
        )

    # ── Conditional formatting: RAG for the sign-off column block ────────────
    if not DISABLE_CONDITIONAL_FORMATTING:
        cf_range = (
            f"{get_column_letter(_CONTENT_PLANNER_SIGNOFF_FIRST_COL)}{data_start}:"
            f"{get_column_letter(_CONTENT_PLANNER_SIGNOFF_LAST_COL)}{last_row}"
        )
        signed_off_fill = PatternFill(
            start_color=RAG_GREEN, end_color=RAG_GREEN, fill_type="solid"
        )
        in_progress_fill = PatternFill(
            start_color=RAG_AMBER, end_color=RAG_AMBER, fill_type="solid"
        )
        not_signed_off_fill = PatternFill(
            start_color=RAG_RED, end_color=RAG_RED, fill_type="solid"
        )
        first_signoff_letter = get_column_letter(_CONTENT_PLANNER_SIGNOFF_FIRST_COL)
        first_signoff_top = f"{first_signoff_letter}{data_start}"
        worksheet.conditional_formatting.add(
            cf_range,
            FormulaRule(
                formula=[f'LOWER({first_signoff_top})="signed off"'],
                fill=signed_off_fill,
                font=Font(color=RAG_GREEN_FONT),
                stopIfTrue=True,
            ),
        )
        worksheet.conditional_formatting.add(
            cf_range,
            FormulaRule(
                formula=[f'LOWER({first_signoff_top})="in progress"'],
                fill=in_progress_fill,
                font=Font(color=RAG_AMBER_FONT),
                stopIfTrue=True,
            ),
        )
        worksheet.conditional_formatting.add(
            cf_range,
            FormulaRule(
                formula=[f'LOWER({first_signoff_top})="not signed off"'],
                fill=not_signed_off_fill,
                font=Font(color=RAG_RED_FONT),
                stopIfTrue=True,
            ),
        )

    # ── Autofilter + freeze (columns A–D pinned) ────────────────────────────
    worksheet.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(worksheet.max_column)}{last_row}"
    )
    set_freeze_panes_safe(worksheet, f"E{data_start}")


__all__ = [
    "apply_wrapped_row_heights",
    "apply_sheet_text_wrap_columns",
    "apply_cf_zebra_banding",
    "apply_generic_sheet_coloring",
    "apply_content_hub_conditional_rules",
    "finalize_content_hub_after_normalized_headers",
    "apply_psi_conditional_rules",
    "apply_merged_tabs_conditional_formatting",
    "apply_main_sheet_heatmaps",
    "apply_main_group_conditional_formatting",
    "apply_dashboard_metric_conditional_rules",
    "apply_content_planner_signoff_rules",
    "apply_playbook_section_colors",
    "MERGED_TAB_NAMES",
]

# Public alias — tables_impl.py's dispatcher must call apply_merged_tabs_conditional_
# formatting for exactly this sheet set, so both share one source of truth.
MERGED_TAB_NAMES = _MERGED_TAB_NAMES
