"""Executive Dashboard — KPI cards and charts backed by visible source rows."""

from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from typing import Any
from urllib.parse import urlparse

from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from hype_frog.core.models import ExtraRowPayload, MainRowPayload, SummaryMetricsPayload
from hype_frog.reporter.chart_compat import configure_openpyxl_chart_for_excel
from hype_frog.reporter.dashboard_logic import (
    DashboardComputationResult,
    FixPlanRowPayload,
    compute_dashboard_metrics,
)
from hype_frog.reporter.engine_io import _safe_sheet_name, _sanitize_excel_value
from hype_frog.reporter.sheets.config import (
    AUDIT_RUN_DETAILS_SHEET,
    DATA_BAR_BLUE,
    DEBUG_EXCEL_ISOLATION_MODE,
    EXECUTIVE_BRIEFING_FREEZE_PANES,
    EXECUTIVE_BRIEFING_SHEET,
    EXECUTIVE_DASHBOARD_SHEET,
    STD_BLUE,
    STD_NAVY,
)
from hype_frog.reporter.sheets.dashboard import (
    DASHBOARD_BRAND_A1,
    apply_executive_briefing_triage,
)
from hype_frog.reporter.sheets.dashboard_config import (
    LIGHT_HEADER_COLOR,
    PANEL_BG_COLOR,
    VALUE_BLOCK_COLOR,
)
from hype_frog.reporter.sheets.view_state import set_freeze_panes_safe

# Reuse the shared Dashboard structural palette so both dashboards read consistently.
_KPI_FILL = PatternFill("solid", fgColor=VALUE_BLOCK_COLOR)
_SECTION_FILL = PatternFill("solid", fgColor=LIGHT_HEADER_COLOR)
_SOURCE_FILL = PatternFill("solid", fgColor="F3F4F6")
_INSIGHT_FILL = PatternFill("solid", fgColor=PANEL_BG_COLOR)

# Curated tooltips for the less self-evident KPI cards (T5).
_EXEC_CARD_TOOLTIPS: dict[str, str] = {
    "SEO Health (now)": "Average SEO health across all crawled URLs (0–100).",
    "SEO Health (illustrative projected)": (
        "Illustrative SEO health if the prioritised fixes are completed. "
        "A planning estimate, not a guarantee."
    ),
    "AEO Readiness": (
        "Average AEO readiness across measured URLs only (Unmeasured rows excluded). "
        "The KPI label states how many URLs were dropped from this average."
    ),
    "Traffic Lift (est.)": (
        "Rough estimated monthly organic visit uplift from remediating priority pages."
    ),
    "Performance (PSI)": "Average mobile PageSpeed Insights performance score (0–100).",
}

_KEY_INSIGHTS_ROW = 13

# 2.5 UX overhaul — column split, not a row-stacked layout: columns A-H hold
# stats (KPI cards, key insights, owner/nav triage) and are frozen; columns I
# onward hold the 6 chart sections side by side, reached by scrolling right
# (not down). Everything in A-H is compact enough to need no vertical scroll.
_KPI_CARDS_PER_ROW = 4
_STATS_ZONE_LAST_COL = 8  # "H" — freeze boundary
_ROW_KPI1 = 7
_ROW_KPI2 = 9
_ROW_KPI3 = 11
_BRIEFING_TRIAGE_START_ROW = 14

_SOURCE_DATA_HINT_ROW = 153

# Chart zone: one shared row band (all 6 sections start at the same row) with
# each section occupying its own column width, one after another starting at
# column I — see _chart_section_start_cols().
_CHART_ZONE_START_COL = 9  # "I"
_ROW_SEC_CHARTS = 15
_ROW_CH_CHARTS = 16
_CHART_BAND_ROW_HEIGHT_PT = 24.0
_CHART_BAND_ROWS = 2

# Column width (in sheet column-count units) reserved per chart section,
# left-to-right: Health, Issues (severity+owner doughnut), Actions (priority+
# content readiness), Top Issues, Status, Traditional-vs-AEO. Wide enough for
# each section's chart(s) at the configured column width (see
# _apply_executive_column_grid) without overlapping the next section.
_CHART_SECTION_COLUMN_SPANS: tuple[int, ...] = (10, 12, 12, 10, 6, 6)

# Chart size presets (cm): full-width and half-section.
_SIZE_HEALTH_FULL = (18.0, 8.4)
_SIZE_HALF_CHART = (9.2, 8.4)
_SIZE_DOUGHNUT = (9.0, 8.4)
_SIZE_TOP_ISSUES = (18.0, 8.4)


def _chart_section_start_cols() -> list[int]:
    """1-based column index where each of the 6 chart sections starts."""
    starts: list[int] = []
    col = _CHART_ZONE_START_COL
    for span in _CHART_SECTION_COLUMN_SPANS:
        starts.append(col)
        col += span
    return starts

_LOW_VALUE_URL_PATH_TOKENS: tuple[str, ...] = (
    "cart",
    "checkout",
    "thank-you",
    "thankyou",
    "thanks",
    "order-received",
    "sponsorship",
    "sponsor",
    "donate",
    "login",
    "sign-in",
    "signin",
    "my-account",
    "account",
    "wp-json",
    "/feed",
    "attachment",
    "privacy-policy",
    "terms",
)

# Chart tables in visible columns A–C well below the chart area and the triage
# matrix (Excel ignores hidden cols by default).
CHART_SOURCE_FIRST_ROW = 155
CHART_LABEL_COL = 1
CHART_VALUE_COL = 2
CHART_VALUE2_COL = 3

# Doughnut slice colours — 6-digit RRGGBB only (Excel ``srgbClr`` rejects 8-digit ARGB).
_SEVERITY_SLICE_COLORS: tuple[str, ...] = ("C00000", "ED7D31", "BFBFBF")
_OWNER_SLICE_COLORS: tuple[str, ...] = (
    "4472C4",
    "70AD47",
    "7030A0",
    "A6A6A6",
    "FFC000",
)

# Bar/column chart palettes. Deliberately more saturated than the pastel CF
# fills in sheets/config.py (STATUS_TODO_FILL etc.) — chart bars need to read
# at a glance, not sit quietly behind text. Same 6-digit RRGGBB constraint.
_CHART_SERIES_PRIMARY: str = DATA_BAR_BLUE
_CHART_SERIES_SECONDARY: str = "9DC3E6"
# For charts with no inherent severity/pass-fail meaning per bar (priority
# pages, content-readiness signals, top issues) — a categorical ramp so each
# bar is visually distinct instead of one flat colour.
_CHART_CATEGORICAL_COLORS: tuple[str, ...] = (
    "4472C4",
    "70AD47",
    "FFC000",
    "7030A0",
    "ED7D31",
    "A6A6A6",
    "5B9BD5",
    "C00000",
)
# Status-code buckets ("200 OK", "3xx Redirects", "4xx Errors", "5xx Errors",
# "Other" — see dashboard_logic.compute_dashboard_metrics) coloured by class.
_STATUS_CODE_COLOR_RULES: tuple[tuple[str, str], ...] = (
    ("2", "70AD47"),
    ("3", "FFC000"),
    ("4", "C00000"),
    ("5", "C00000"),
)
_STATUS_CODE_OTHER_COLOR: str = "A6A6A6"


def _status_bucket_color(label: str) -> str:
    """Map a status-code bucket label (e.g. '4xx Errors') to a chart colour."""
    stripped = label.strip()
    for prefix, color in _STATUS_CODE_COLOR_RULES:
        if stripped.startswith(prefix):
            return color
    return _STATUS_CODE_OTHER_COLOR


@dataclass(frozen=True)
class ChartDataLayout:
    """1-based row anchors for chart source tables (columns A–C)."""

    health_start: int = CHART_SOURCE_FIRST_ROW + 1
    health_rows: int = 6
    severity_start: int = CHART_SOURCE_FIRST_ROW + 9
    severity_rows: int = 3
    owner_start: int = CHART_SOURCE_FIRST_ROW + 16
    owner_rows: int = 0
    priority_start: int = CHART_SOURCE_FIRST_ROW + 25
    priority_rows: int = 0
    content_start: int = CHART_SOURCE_FIRST_ROW + 37
    content_rows: int = 5
    projected_start: int = CHART_SOURCE_FIRST_ROW + 45
    projected_rows: int = 2
    top_issues_start: int = CHART_SOURCE_FIRST_ROW + 50
    top_issues_rows: int = 0
    status_start: int = CHART_SOURCE_FIRST_ROW + 62
    status_rows: int = 0
    tradaeo_start: int = CHART_SOURCE_FIRST_ROW + 70
    tradaeo_rows: int = 2


def _label_col() -> int:
    return CHART_LABEL_COL


def _value_col() -> int:
    return CHART_VALUE_COL


def _value_col_2() -> int:
    return CHART_VALUE2_COL


def _normalize_chart_rgb(color: str) -> str:
    cleaned = color.strip().lstrip("#").upper()
    if len(cleaned) == 8:
        cleaned = cleaned[2:]
    return cleaned[:6]


def _write_label_cell(ws: Worksheet, row: int, value: object) -> None:
    cell = ws.cell(row=row, column=_label_col(), value=_sanitize_excel_value(str(value)))
    cell.data_type = "s"


def _to_float(value: object, default: float = 0.0) -> float:
    if value is None or str(value).strip() == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _optional_float(value: object) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        out = float(value)
    except (TypeError, ValueError):
        return None
    if out != out:
        return None
    return out


def _short_url_label(url: object, *, max_len: int = 36) -> str:
    raw = str(url or "").strip()
    if not raw:
        return "(no url)"
    path = urlparse(raw).path or "/"
    if path == "/":
        label = raw.replace("https://", "").replace("http://", "").rstrip("/") or "/"
    else:
        label = path.strip("/")
        if len(label) > max_len:
            label = "…" + label[-(max_len - 1) :]
    return _sanitize_excel_value(label) or "(url)"


LAB_LCP_MOBILE_TARGET_S = 2.5
LIGHTHOUSE_ACCESSIBILITY_TARGET = 90.0


def _avg_numeric_column(
    extra_rows: list[dict[str, Any]],
    key: str,
    *,
    require_positive: bool = False,
) -> float | None:
    values: list[float] = []
    for row in extra_rows:
        val = _optional_float(row.get(key))
        if val is None:
            continue
        if require_positive and val <= 0:
            continue
        values.append(val)
    if not values:
        return None
    return round(sum(values) / len(values), 1)


def _avg_lighthouse_performance_mobile(extra_rows: list[dict[str, Any]]) -> float | None:
    scores: list[float] = []
    for row in extra_rows:
        val = _optional_float(row.get("Lighthouse Performance (Mobile)"))
        if val is None:
            val = _optional_float(row.get("Mobile PSI Score"))
        if val is not None and val > 0:
            scores.append(val)
    if not scores:
        return None
    return round(sum(scores) / len(scores), 1)


def _content_readiness_percentages(
    extra_rows: list[dict[str, Any]],
) -> list[tuple[str, float]]:
    total = len(extra_rows)
    if total <= 0:
        return [
            ("Good H1", 0.0),
            ("Meta description", 0.0),
            ("Answer paragraphs", 0.0),
            ("Schema present", 0.0),
            ("Image alt ≥80%", 0.0),
            ("Question headings", 0.0),
        ]

    def pct(count: int) -> float:
        return round((count / total) * 100.0, 1)

    good_h1 = sum(
        1
        for row in extra_rows
        if not bool(row.get("Missing H1 Flag")) and int(row.get("H1 Count") or 0) > 0
    )
    good_meta = sum(1 for row in extra_rows if not bool(row.get("Meta Description Missing")))
    answer_blocks = sum(
        1 for row in extra_rows if int(row.get("Paragraphs 40-60 Words Count") or 0) > 0
    )
    schema = sum(1 for row in extra_rows if int(row.get("Schema Types Count") or 0) > 0)
    alt_ok = sum(
        1 for row in extra_rows if _to_float(row.get("Image Alt Coverage (%)")) >= 80.0
    )
    question_headings = sum(
        1 for row in extra_rows if int(row.get("Question Heading Count") or 0) > 0
    )
    return [
        ("Good H1", pct(good_h1)),
        ("Meta description", pct(good_meta)),
        ("Answer paragraphs", pct(answer_blocks)),
        ("Schema present", pct(schema)),
        ("Image alt ≥80%", pct(alt_ok)),
        ("Question headings", pct(question_headings)),
    ]


def _gsc_totals(extra_rows: list[dict[str, Any]]) -> tuple[int, int, float]:
    """Return (clicks_total, impressions_total, avg_position) across all crawled URLs."""
    clicks_total = int(sum(_to_float(row.get("GSC Clicks"), 0.0) for row in extra_rows))
    impressions_total = int(sum(_to_float(row.get("GSC Impressions"), 0.0) for row in extra_rows))
    positions = [
        _to_float(row.get("GSC Avg Position"), 0.0)
        for row in extra_rows
        if _to_float(row.get("GSC Avg Position"), 0.0) > 0
    ]
    avg_position = round(sum(positions) / len(positions), 1) if positions else 0.0
    return clicks_total, impressions_total, avg_position


def _traffic_lift_total(hub_metrics_rows: list[dict[str, Any]] | None) -> int:
    if not hub_metrics_rows:
        return 0
    return int(
        round(
            sum(_to_float(row.get("Potential Traffic Lift"), 0.0) for row in hub_metrics_rows)
        )
    )


def _normalize_fixplan_severity(raw: object) -> str:
    sev = str(raw or "").strip().lower()
    if sev == "critical":
        return "Critical"
    if sev in {"warning", "high"}:
        return "Warning"
    if sev in {"observation", "medium", "low", "info"}:
        return "Observation"
    return "Observation"


def _severity_metrics(
    extra_rows: list[dict[str, Any]],
    fixplan_rows: list[dict[str, Any]],
) -> list[tuple[str, int, int]]:
    """Return (label, unique_urls, issue_instances) per severity tier."""
    unique_urls: Counter[str] = Counter()
    for row in extra_rows:
        badge = str(row.get("Severity Badge") or "").strip()
        if badge == "Critical":
            unique_urls["Critical"] += 1
        elif badge == "Warning":
            unique_urls["Warning"] += 1
        elif badge in {"Observation", "Info"}:
            unique_urls["Observation"] += 1

    issue_instances: Counter[str] = Counter()
    for row in fixplan_rows:
        tier = _normalize_fixplan_severity(row.get("Severity"))
        issue_instances[tier] += int(row.get("Affected Count") or 0)

    rows: list[tuple[str, int, int]] = []
    for label in ("Critical", "Warning", "Observation"):
        url_count = unique_urls.get(label, 0)
        instance_count = issue_instances.get(label, 0)
        if url_count > 0 or instance_count > 0:
            rows.append((label, url_count, instance_count))
    if not rows:
        return [("No open issues", 0, 0)]
    return rows


def _owner_metrics(
    extra_rows: list[dict[str, Any]],
    fixplan_rows: list[dict[str, Any]],
) -> list[tuple[str, int, int]]:
    """Return (owner, unique_urls, issue_instances) sorted by unique URL count."""
    urls_by_owner: dict[str, set[str]] = defaultdict(set)
    for row in extra_rows:
        badge = str(row.get("Severity Badge") or "").strip()
        if badge not in {"Critical", "Warning", "Observation", "Info"}:
            continue
        owner = str(row.get("Owner") or "Unassigned").strip() or "Unassigned"
        url = str(row.get("URL") or "").strip()
        if url:
            urls_by_owner[owner].add(url)

    instances_by_owner: Counter[str] = Counter()
    for row in fixplan_rows:
        owner = str(row.get("Owner") or "Unassigned").strip() or "Unassigned"
        instances_by_owner[owner] += int(row.get("Affected Count") or 0)

    owners = set(urls_by_owner) | set(instances_by_owner)
    if not owners:
        return [("Unassigned", 0, 0)]

    ranked = sorted(
        owners,
        key=lambda name: (
            -len(urls_by_owner.get(name, set())),
            -instances_by_owner.get(name, 0),
            name,
        ),
    )
    return [
        (
            owner,
            len(urls_by_owner.get(owner, set())),
            instances_by_owner.get(owner, 0),
        )
        for owner in ranked[:8]
    ]


def _to_do_share(summary_metrics: SummaryMetricsPayload) -> float:
    total = max(1, summary_metrics.urls_crawled)
    return (summary_metrics.critical_url_count + summary_metrics.warning_url_count) / total


def _project_component_score(current: float, summary_metrics: SummaryMetricsPayload) -> float:
    """Mirror export_flow projected-health uplift (0.9 × open-issue share)."""
    share = _to_do_share(summary_metrics)
    return min(
        100.0,
        round(current + max(0.0, 100.0 - current) * share * 0.9, 1),
    )


def _avg_technical_health(extra_rows: list[dict[str, Any]], fallback: float) -> float:
    values: list[float] = []
    for row in extra_rows:
        for key in ("Technical Health", "SEO Health Score"):
            val = _optional_float(row.get(key))
            if val is not None and val >= 0:
                values.append(val)
                break
    if not values:
        return round(fallback, 1)
    return round(sum(values) / len(values), 1)


def _is_low_value_priority_url(url: object) -> bool:
    path = urlparse(str(url or "")).path.lower()
    return any(token in path for token in _LOW_VALUE_URL_PATH_TOKENS)


def _meaningful_priority_rows(
    priority_rows: list[dict[str, Any]],
    *,
    limit: int = 8,
) -> list[dict[str, Any]]:
    """Prefer high-intent URLs for the business-risk chart (exclude cart/checkout/etc.)."""
    ranked = sorted(
        priority_rows,
        key=lambda item: _to_float(item.get("Business Risk Score"), 0.0),
        reverse=True,
    )
    filtered = [row for row in ranked if not _is_low_value_priority_url(row.get("URL"))]
    if filtered:
        return filtered[:limit]
    return ranked[:limit]


def _format_aeo_readiness_kpi(dashboard_metrics: DashboardComputationResult) -> str:
    value = f"{dashboard_metrics.aeo_readiness:.0f}%"
    dropped = dashboard_metrics.aeo_unmeasured_count
    if dropped > 0:
        noun = "URL" if dropped == 1 else "URLs"
        return f"{value} (excludes {dropped} unmeasured {noun})"
    return value


def _build_key_insights(
    summary_metrics: SummaryMetricsPayload,
    dashboard_metrics: DashboardComputationResult,
) -> list[str]:
    insights: list[str] = []
    if summary_metrics.critical_url_count > 0:
        insights.append(
            f"{summary_metrics.critical_url_count} URL(s) carry a Critical badge and need urgent attention."
        )
    if summary_metrics.warning_url_count > 0:
        insights.append(
            f"{summary_metrics.warning_url_count} URL(s) are in Warning state — schedule fixes before they escalate."
        )
    if dashboard_metrics.top_issue_rows:
        lead = dashboard_metrics.top_issue_rows[0]
        insights.append(
            f"Largest theme: \"{lead.issue_name}\" touches {lead.affected_urls} URL(s) (see Fix Plan)."
        )
    if dashboard_metrics.aeo_readiness < 70:
        aeo_note = ""
        if dashboard_metrics.aeo_unmeasured_count > 0:
            aeo_note = (
                f" ({dashboard_metrics.aeo_unmeasured_count} unmeasured URL(s) excluded from average)"
            )
        insights.append(
            f"AEO readiness averages {dashboard_metrics.aeo_readiness:.0f}%{aeo_note} — "
            "highest-leverage fix: add question-style H2–H4 headings with concise 40–60 word "
            "answers directly underneath."
        )
    if dashboard_metrics.error_count > 0:
        insights.append(
            f"{dashboard_metrics.error_count} URL(s) returned 4xx/5xx during the crawl."
        )
    if not insights:
        insights.append(
            f"Site SEO health is {summary_metrics.health_score_pct:.0f}% with no Critical URLs in this crawl."
        )
    return insights[:4]


def _reserve_chart_band(
    exec_ws: Worksheet,
    start_row: int,
    row_count: int,
    *,
    row_height_pt: float = _CHART_BAND_ROW_HEIGHT_PT,
) -> None:
    """Reserve vertical space so openpyxl charts (sized in cm) do not overlap."""
    for offset in range(row_count):
        exec_ws.row_dimensions[start_row + offset].height = row_height_pt


def _apply_executive_column_grid(exec_ws: Worksheet) -> None:
    """Column widths for the 2.5 UX overhaul's two zones: A-H (frozen stats,
    even width so the KPI card grid/owner table line up) and I onward (chart
    zone, sized to roughly match _CHART_SECTION_COLUMN_SPANS' cm assumptions —
    see _apply_chart_size call sites for each section's actual chart width)."""
    for col_idx in range(1, _STATS_ZONE_LAST_COL + 1):
        exec_ws.column_dimensions[get_column_letter(col_idx)].width = 12.0
    last_chart_col = _CHART_ZONE_START_COL + sum(_CHART_SECTION_COLUMN_SPANS)
    for col_idx in range(_CHART_ZONE_START_COL, last_chart_col + 1):
        exec_ws.column_dimensions[get_column_letter(col_idx)].width = 10.0


def _write_source_data_hint(exec_ws: Worksheet) -> None:
    exec_ws.merge_cells(f"A{_SOURCE_DATA_HINT_ROW}:L{_SOURCE_DATA_HINT_ROW}")
    hint = exec_ws.cell(
        row=_SOURCE_DATA_HINT_ROW,
        column=1,
        value=f"▼ Chart source data begins row {CHART_SOURCE_FIRST_ROW} (columns A–C)",
    )
    hint.font = Font(italic=True, size=9, color="666666")
    hint.fill = _SOURCE_FILL
    hint.alignment = Alignment(horizontal="center", vertical="center")
    exec_ws.row_dimensions[_SOURCE_DATA_HINT_ROW].height = 18


def _apply_chart_size(chart: BarChart | DoughnutChart, width_cm: float, height_cm: float) -> None:
    chart.width = width_cm
    chart.height = height_cm


def _write_section_header(
    exec_ws: Worksheet,
    row: int,
    title: str,
    *,
    start_col: int = 1,
    merge_to_col: int = 8,
) -> None:
    merge = f"{get_column_letter(start_col)}{row}:{get_column_letter(merge_to_col)}{row}"
    exec_ws.merge_cells(merge)
    cell = exec_ws.cell(row=row, column=start_col, value=title)
    cell.fill = _SECTION_FILL
    cell.font = Font(bold=True, color=STD_NAVY, size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    exec_ws.row_dimensions[row].height = 24


def _write_key_insights(
    exec_ws: Worksheet,
    insights: list[str],
) -> None:
    exec_ws.merge_cells(f"A{_KEY_INSIGHTS_ROW}:H{_KEY_INSIGHTS_ROW}")
    body = exec_ws.cell(
        row=_KEY_INSIGHTS_ROW,
        column=1,
        value="Key insights: " + " | ".join(insights),
    )
    body.fill = _INSIGHT_FILL
    body.font = Font(size=9, color="1F2937")
    body.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
    exec_ws.row_dimensions[_KEY_INSIGHTS_ROW].height = 36


def _top_issues_by_impact(
    fixplan_rows: list[dict[str, Any]],
    *,
    limit: int = 8,
) -> list[tuple[str, int]]:
    # Tie-break must match dashboard_logic.compute_dashboard_metrics()'s
    # top_issue_rows sort — (-affected, -source_row, issue_name), where
    # source_row is enumerate(fixplan_rows, start=2) — the same list, same
    # order, same 1-based row offset — so a tied count never disagrees with
    # the "Largest theme" sentence in _build_key_insights (L1 audit fix).
    indexed_rows = list(enumerate(fixplan_rows, start=2))
    ranked = sorted(
        indexed_rows,
        key=lambda item: (
            -int(item[1].get("Affected Count") or 0),
            -item[0],
            str(item[1].get("Issue Type") or ""),
        ),
    )
    issues: list[tuple[str, int]] = []
    for _source_row, row in ranked:
        count = int(row.get("Affected Count") or 0)
        if count <= 0:
            continue
        name = _sanitize_excel_value(str(row.get("Issue Type") or "Unknown")) or "Unknown"
        if len(name) > 42:
            name = name[:39] + "…"
        issues.append((name, count))
        if len(issues) >= limit:
            break
    return issues


def _build_projection_narrative(summary_metrics: SummaryMetricsPayload) -> str:
    total = max(1, summary_metrics.urls_crawled)
    open_urls = summary_metrics.critical_url_count + summary_metrics.warning_url_count
    open_pct = round((open_urls / total) * 100.0)
    gap = max(
        0.0,
        summary_metrics.projected_health_score_pct - summary_metrics.health_score_pct,
    )
    return (
        f"SEO health is {summary_metrics.health_score_pct:.0f}% across {total} crawled URLs. "
        f"{open_urls} URL(s) ({open_pct}%) carry Critical or Warning badges. "
        f"If those were cleared, blended SEO health could reach ~"
        f"{summary_metrics.projected_health_score_pct:.0f}% (+{gap:.0f} pts) — "
        f"an illustrative ceiling from the export model (90% of remaining headroom × open-issue share). "
        f"Technical, PSI, and AEO components use the same conservative formula in the chart below; "
        f"they are not guaranteed to move in lockstep. Source tables: row {CHART_SOURCE_FIRST_ROW}+."
    )


def populate_chart_data_sheet(
    ws: Worksheet,
    *,
    summary_metrics: SummaryMetricsPayload,
    dashboard_metrics: DashboardComputationResult,
    extra_rows: list[dict[str, Any]],
    priority_rows: list[dict[str, Any]],
    fixplan_rows: list[dict[str, Any]],
    hub_metrics_rows: list[dict[str, Any]] | None,
) -> ChartDataLayout:
    """Write chart source tables in visible columns A–C (no merged cells in data)."""
    layout = ChartDataLayout()
    psi = _avg_lighthouse_performance_mobile(extra_rows)
    lab_lcp = _avg_numeric_column(extra_rows, "Lab LCP (Mobile) (s)", require_positive=True)
    accessibility = _avg_numeric_column(
        extra_rows,
        "Lighthouse Accessibility (Mobile)",
        require_positive=True,
    )
    aeo = round(dashboard_metrics.aeo_readiness, 1)
    current_health = round(summary_metrics.health_score_pct, 1)
    projected_health = round(summary_metrics.projected_health_score_pct, 1)
    technical_health = _avg_technical_health(extra_rows, dashboard_metrics.overall_health)
    value_col = _value_col()
    value_col_2 = _value_col_2()
    projected_technical = _project_component_score(technical_health, summary_metrics)
    projected_psi = _project_component_score(psi or 0.0, summary_metrics) if psi else 0.0
    projected_aeo = _project_component_score(aeo, summary_metrics)

    ws.merge_cells(
        start_row=CHART_SOURCE_FIRST_ROW,
        start_column=CHART_LABEL_COL,
        end_row=CHART_SOURCE_FIRST_ROW,
        end_column=CHART_VALUE2_COL,
    )
    banner = ws.cell(row=CHART_SOURCE_FIRST_ROW, column=CHART_LABEL_COL)
    banner.value = "Chart source data (scroll down — used by charts above)"
    banner.font = Font(bold=True, size=9, color="666666")
    banner.fill = _SOURCE_FILL

    _write_label_cell(ws, layout.health_start, "Health comparison")
    hr = layout.health_start + 1
    _write_label_cell(ws, hr, "Metric")
    ws.cell(row=hr, column=value_col, value="Current")
    ws.cell(row=hr, column=value_col_2, value="Illustrative projected")
    health_rows = [
        ("SEO Health", current_health, projected_health),
        ("Technical Health", technical_health, projected_technical),
        ("Performance (PSI)", psi or 0.0, projected_psi),
        ("LCP (Lab Mobile avg)", lab_lcp or 0.0, LAB_LCP_MOBILE_TARGET_S),
        ("Accessibility (avg)", accessibility or 0.0, LIGHTHOUSE_ACCESSIBILITY_TARGET),
        ("AEO Readiness", aeo, projected_aeo),
    ]
    data_row = hr + 1
    for metric, current, projected in health_rows:
        _write_label_cell(ws, data_row, metric)
        ws.cell(row=data_row, column=value_col, value=current)
        ws.cell(row=data_row, column=value_col_2, value=projected)
        data_row += 1
    layout = ChartDataLayout(
        health_start=layout.health_start,
        health_rows=len(health_rows),
        severity_start=layout.severity_start,
    )

    _write_label_cell(ws, layout.severity_start, "Severity — unique URLs vs issue instances")
    sr = layout.severity_start + 1
    _write_label_cell(ws, sr, "Severity")
    ws.cell(row=sr, column=value_col, value="Unique URLs")
    ws.cell(row=sr, column=value_col_2, value="Issue instances")
    severity = _severity_metrics(extra_rows, fixplan_rows)
    layout = ChartDataLayout(
        health_start=layout.health_start,
        health_rows=layout.health_rows,
        severity_start=layout.severity_start,
        severity_rows=len(severity),
    )
    for idx, (label, unique_urls, issue_instances) in enumerate(severity, start=sr + 1):
        _write_label_cell(ws, idx, label)
        ws.cell(row=idx, column=value_col, value=unique_urls)
        ws.cell(row=idx, column=value_col_2, value=issue_instances)

    _write_label_cell(ws, layout.owner_start, "Owner workload — unique URLs vs issue instances")
    orow = layout.owner_start + 1
    _write_label_cell(ws, orow, "Owner")
    ws.cell(row=orow, column=value_col, value="Unique URLs")
    ws.cell(row=orow, column=value_col_2, value="Issue instances")
    owners = _owner_metrics(extra_rows, fixplan_rows)
    layout = ChartDataLayout(
        health_start=layout.health_start,
        health_rows=layout.health_rows,
        severity_start=layout.severity_start,
        severity_rows=layout.severity_rows,
        owner_start=layout.owner_start,
        owner_rows=len(owners),
    )
    for idx, (owner, unique_urls, issue_instances) in enumerate(owners, start=orow + 1):
        _write_label_cell(ws, idx, owner)
        ws.cell(row=idx, column=value_col, value=unique_urls)
        ws.cell(row=idx, column=value_col_2, value=issue_instances)

    _write_label_cell(
        ws,
        layout.priority_start,
        "Top business risk URLs (chart excludes cart/checkout/thank-you pages)",
    )
    pr = layout.priority_start + 1
    _write_label_cell(ws, pr, "Page")
    ws.cell(row=pr, column=value_col, value="Business Risk Score")
    top_priority = _meaningful_priority_rows(priority_rows, limit=8)
    layout = ChartDataLayout(
        health_start=layout.health_start,
        health_rows=layout.health_rows,
        severity_start=layout.severity_start,
        severity_rows=layout.severity_rows,
        owner_start=layout.owner_start,
        owner_rows=layout.owner_rows,
        priority_start=layout.priority_start,
        priority_rows=len(top_priority),
    )
    for idx, row_dict in enumerate(top_priority, start=pr + 1):
        _write_label_cell(ws, idx, _short_url_label(row_dict.get("URL")))
        ws.cell(row=idx, column=value_col, value=_to_float(row_dict.get("Business Risk Score"), 0.0))

    _write_label_cell(ws, layout.content_start, "Content readiness")
    cr = layout.content_start + 1
    _write_label_cell(ws, cr, "Metric")
    ws.cell(row=cr, column=value_col, value="Percent")
    content = _content_readiness_percentages(extra_rows)
    for idx, (label, pct) in enumerate(content, start=cr + 1):
        _write_label_cell(ws, idx, label)
        ws.cell(row=idx, column=value_col, value=pct)

    _write_label_cell(ws, layout.projected_start, "SEO health — export projection model")
    pj = layout.projected_start + 1
    _write_label_cell(ws, pj, "Metric")
    ws.cell(row=pj, column=value_col, value="Percent")
    _write_label_cell(ws, pj + 1, "Current SEO Health")
    ws.cell(row=pj + 1, column=value_col, value=current_health)
    _write_label_cell(ws, pj + 2, "Illustrative projected SEO Health")
    ws.cell(row=pj + 2, column=value_col, value=projected_health)
    note_row = pj + 3
    _write_label_cell(
        ws,
        note_row,
        "Model: current + (100-current) × open-issue share × 0.9 (matches export_flow)",
    )

    top_issues_start = CHART_SOURCE_FIRST_ROW + 50
    _write_label_cell(ws, top_issues_start, "Top issues by URL impact")
    tir = top_issues_start + 1
    _write_label_cell(ws, tir, "Issue")
    ws.cell(row=tir, column=value_col, value="Affected URLs")
    top_issues = _top_issues_by_impact(fixplan_rows, limit=8)
    for idx, (issue_name, affected) in enumerate(top_issues, start=tir + 1):
        _write_label_cell(ws, idx, issue_name)
        ws.cell(row=idx, column=value_col, value=affected)

    status_start = CHART_SOURCE_FIRST_ROW + 62
    _write_label_cell(ws, status_start, "Status code breakdown")
    str_row = status_start + 1
    _write_label_cell(ws, str_row, "Status")
    ws.cell(row=str_row, column=value_col, value="URLs")
    status_buckets = [
        (label, count)
        for label, count in dashboard_metrics.status_buckets.items()
        if count > 0
    ]
    for idx, (label, count) in enumerate(status_buckets, start=str_row + 1):
        _write_label_cell(ws, idx, label)
        ws.cell(row=idx, column=value_col, value=count)

    tradaeo_start = CHART_SOURCE_FIRST_ROW + 70
    _write_label_cell(ws, tradaeo_start, "Traditional SEO vs AEO readiness")
    tar = tradaeo_start + 1
    _write_label_cell(ws, tar, "Model")
    ws.cell(row=tar, column=value_col, value="Score")
    _write_label_cell(ws, tar + 1, "Traditional SEO")
    ws.cell(row=tar + 1, column=value_col, value=round(dashboard_metrics.traditional_score, 1))
    aeo_label = (
        f"AEO Readiness (excl. {dashboard_metrics.aeo_unmeasured_count} unmeasured)"
        if dashboard_metrics.aeo_unmeasured_count > 0
        else "AEO Readiness"
    )
    _write_label_cell(ws, tar + 2, aeo_label)
    ws.cell(row=tar + 2, column=value_col, value=round(dashboard_metrics.aeo_readiness, 1))

    return ChartDataLayout(
        health_start=layout.health_start,
        health_rows=layout.health_rows,
        severity_start=layout.severity_start,
        severity_rows=layout.severity_rows,
        owner_start=layout.owner_start,
        owner_rows=layout.owner_rows,
        priority_start=layout.priority_start,
        priority_rows=layout.priority_rows,
        content_start=layout.content_start,
        content_rows=len(content),
        projected_start=layout.projected_start,
        projected_rows=2,
        top_issues_start=top_issues_start,
        top_issues_rows=len(top_issues),
        status_start=status_start,
        status_rows=len(status_buckets),
        tradaeo_start=tradaeo_start,
        tradaeo_rows=2,
    )


def _attach_chart(exec_ws: Worksheet, chart: BarChart | DoughnutChart, anchor: str) -> None:
    configure_openpyxl_chart_for_excel(chart)
    exec_ws.add_chart(chart, anchor)


def _apply_bar_point_colors(
    chart: BarChart,
    colors: tuple[str, ...],
    *,
    point_count: int,
) -> None:
    """Per-bar colours for a single-series BarChart (openpyxl defaults every
    bar to one flat colour otherwise)."""
    if point_count <= 0 or not chart.series:
        return
    chart.varyColors = True
    series = chart.series[0]
    for point_idx in range(point_count):
        color = _normalize_chart_rgb(colors[point_idx % len(colors)])
        pt = DataPoint(idx=point_idx)
        pt.graphicalProperties.solidFill = color
        series.data_points.append(pt)


def _apply_series_fills(chart: BarChart, colors: tuple[str, ...]) -> None:
    """Explicit per-series fill for a multi-series BarChart (current vs
    projected, unique URLs vs instances) so each series reads distinctly."""
    for series, color in zip(chart.series, colors):
        series.graphicalProperties.solidFill = _normalize_chart_rgb(color)


def _apply_doughnut_colors(
    chart: DoughnutChart,
    colors: tuple[str, ...],
    *,
    point_count: int,
) -> None:
    if point_count <= 0 or not chart.series:
        return
    series = chart.series[0]
    for point_idx in range(point_count):
        color = _normalize_chart_rgb(colors[point_idx % len(colors)])
        pt = DataPoint(idx=point_idx)
        pt.graphicalProperties.solidFill = color
        series.data_points.append(pt)


def _add_health_comparison_chart(
    exec_ws: Worksheet,
    layout: ChartDataLayout,
    anchor: str,
) -> None:
    header_row = layout.health_start + 1
    first_data = header_row + 1
    last_data = first_data + layout.health_rows - 1
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Health components — current vs illustrative projected"
    chart.y_axis.title = "Score"
    chart.x_axis.title = "Metric"
    chart.legend.position = "b"
    _apply_chart_size(chart, *_SIZE_HEALTH_FULL)
    data = Reference(
        exec_ws,
        min_col=_value_col(),
        min_row=header_row,
        max_col=_value_col_2(),
        max_row=last_data,
    )
    cats = Reference(
        exec_ws,
        min_col=_label_col(),
        min_row=first_data,
        max_row=last_data,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    _apply_series_fills(chart, (_CHART_SERIES_PRIMARY, _CHART_SERIES_SECONDARY))
    _attach_chart(exec_ws, chart, anchor)


def _add_doughnut_chart(
    exec_ws: Worksheet,
    *,
    title: str,
    header_row: int,
    data_rows: int,
    anchor: str,
    colors: tuple[str, ...],
    value_col: int | None = None,
) -> None:
    if data_rows <= 0:
        return
    first_data = header_row + 1
    last_data = header_row + data_rows
    series_col = value_col if value_col is not None else _value_col()
    total = sum(
        _to_float(exec_ws.cell(row=r, column=series_col).value, 0.0)
        for r in range(first_data, last_data + 1)
    )
    if total <= 0:
        return
    chart = DoughnutChart()
    chart.title = title
    chart.legend.position = "r"
    _apply_chart_size(chart, *_SIZE_DOUGHNUT)
    data = Reference(exec_ws, min_col=series_col, min_row=header_row, max_row=last_data)
    cats = Reference(exec_ws, min_col=_label_col(), min_row=first_data, max_row=last_data)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showCatName = True
    _apply_doughnut_colors(chart, colors, point_count=data_rows)
    _attach_chart(exec_ws, chart, anchor)


def _add_severity_comparison_chart(
    exec_ws: Worksheet,
    layout: ChartDataLayout,
    anchor: str,
) -> None:
    if layout.severity_rows <= 0:
        return
    header_row = layout.severity_start + 1
    first_data = header_row + 1
    last_data = header_row + layout.severity_rows
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Issue severity — unique URLs vs instances"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Severity"
    chart.legend.position = "b"
    _apply_chart_size(chart, *_SIZE_HALF_CHART)
    data = Reference(
        exec_ws,
        min_col=_value_col(),
        min_row=header_row,
        max_col=_value_col_2(),
        max_row=last_data,
    )
    cats = Reference(
        exec_ws,
        min_col=_label_col(),
        min_row=first_data,
        max_row=last_data,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    _apply_series_fills(chart, (_CHART_SERIES_PRIMARY, _CHART_SERIES_SECONDARY))
    _attach_chart(exec_ws, chart, anchor)


def _add_priority_bar_chart(
    exec_ws: Worksheet,
    layout: ChartDataLayout,
    anchor: str,
) -> None:
    if layout.priority_rows <= 0:
        return
    header_row = layout.priority_start + 1
    first_data = header_row + 1
    last_data = header_row + layout.priority_rows
    chart = BarChart()
    chart.type = "bar"
    chart.title = "High-intent pages by business risk"
    chart.x_axis.title = "Business risk score"
    chart.y_axis.title = "Page"
    chart.legend = None
    _apply_chart_size(chart, *_SIZE_HALF_CHART)
    data = Reference(exec_ws, min_col=_value_col(), min_row=header_row, max_row=last_data)
    cats = Reference(exec_ws, min_col=_label_col(), min_row=first_data, max_row=last_data)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    _apply_bar_point_colors(
        chart, _CHART_CATEGORICAL_COLORS, point_count=layout.priority_rows
    )
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    _attach_chart(exec_ws, chart, anchor)


def _add_content_readiness_bar_chart(
    exec_ws: Worksheet,
    layout: ChartDataLayout,
    anchor: str,
) -> None:
    if layout.content_rows <= 0:
        return
    header_row = layout.content_start + 1
    first_data = header_row + 1
    last_data = header_row + layout.content_rows
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Content & AEO readiness (% URLs)"
    chart.x_axis.title = "Percent of URLs"
    chart.y_axis.title = "Signal"
    chart.legend = None
    _apply_chart_size(chart, *_SIZE_HALF_CHART)
    data = Reference(exec_ws, min_col=_value_col(), min_row=header_row, max_row=last_data)
    cats = Reference(exec_ws, min_col=_label_col(), min_row=first_data, max_row=last_data)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    _apply_bar_point_colors(
        chart, _CHART_CATEGORICAL_COLORS, point_count=layout.content_rows
    )
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    _attach_chart(exec_ws, chart, anchor)


def _add_top_issues_chart(
    exec_ws: Worksheet,
    layout: ChartDataLayout,
    anchor: str,
) -> None:
    if layout.top_issues_rows <= 0:
        return
    header_row = layout.top_issues_start + 1
    first_data = header_row + 1
    last_data = header_row + layout.top_issues_rows
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Top issues by URL impact (Fix Plan)"
    chart.x_axis.title = "Affected URLs"
    chart.y_axis.title = "Issue"
    chart.legend = None
    _apply_chart_size(chart, *_SIZE_TOP_ISSUES)
    data = Reference(exec_ws, min_col=_value_col(), min_row=header_row, max_row=last_data)
    cats = Reference(exec_ws, min_col=_label_col(), min_row=first_data, max_row=last_data)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    _apply_bar_point_colors(
        chart, _CHART_CATEGORICAL_COLORS, point_count=layout.top_issues_rows
    )
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    _attach_chart(exec_ws, chart, anchor)


def _add_status_code_chart(
    exec_ws: Worksheet,
    layout: ChartDataLayout,
    anchor: str,
) -> None:
    if layout.status_rows <= 0:
        return
    header_row = layout.status_start + 1
    first_data = header_row + 1
    last_data = header_row + layout.status_rows
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Status code & technical health breakdown"
    chart.y_axis.title = "URLs"
    chart.x_axis.title = "Status"
    chart.legend = None
    _apply_chart_size(chart, *_SIZE_HALF_CHART)
    data = Reference(exec_ws, min_col=_value_col(), min_row=header_row, max_row=last_data)
    cats = Reference(exec_ws, min_col=_label_col(), min_row=first_data, max_row=last_data)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    status_colors = tuple(
        _status_bucket_color(str(exec_ws.cell(row=r, column=_label_col()).value or ""))
        for r in range(first_data, last_data + 1)
    )
    _apply_bar_point_colors(chart, status_colors, point_count=layout.status_rows)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    _attach_chart(exec_ws, chart, anchor)


def _add_traditional_vs_aeo_chart(
    exec_ws: Worksheet,
    layout: ChartDataLayout,
    anchor: str,
) -> None:
    if layout.tradaeo_rows <= 0:
        return
    header_row = layout.tradaeo_start + 1
    first_data = header_row + 1
    last_data = header_row + layout.tradaeo_rows
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Traditional SEO vs AEO Readiness"
    chart.y_axis.title = "Score"
    chart.x_axis.title = "Model"
    chart.legend = None
    _apply_chart_size(chart, *_SIZE_HALF_CHART)
    data = Reference(exec_ws, min_col=_value_col(), min_row=header_row, max_row=last_data)
    cats = Reference(exec_ws, min_col=_label_col(), min_row=first_data, max_row=last_data)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    _apply_bar_point_colors(
        chart,
        (_CHART_SERIES_PRIMARY, _CHART_SERIES_SECONDARY),
        point_count=layout.tradaeo_rows,
    )
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    _attach_chart(exec_ws, chart, anchor)


def _write_briefing_header(
    exec_ws: Worksheet,
    *,
    summary_metrics: SummaryMetricsPayload,
) -> None:
    """Rows 1–5: branded title block and audit run metadata."""
    audit_esc = AUDIT_RUN_DETAILS_SHEET.replace("'", "''")
    exec_ws.merge_cells("A1:H1")
    title = exec_ws["A1"]
    title.value = DASHBOARD_BRAND_A1
    title.font = Font(bold=True, size=18, color=STD_NAVY)
    title.alignment = Alignment(horizontal="left", vertical="center")
    exec_ws.row_dimensions[1].height = 45

    exec_ws.merge_cells("A2:H2")
    subtitle = exec_ws["A2"]
    subtitle.value = _build_projection_narrative(summary_metrics)
    subtitle.alignment = Alignment(wrap_text=True, vertical="top")
    subtitle.font = Font(size=13, color="374151")
    exec_ws.row_dimensions[2].height = 52

    meta_header_fill = PatternFill("solid", fgColor=LIGHT_HEADER_COLOR)
    meta_header_font = Font(color="000000", bold=True, size=11)
    exec_ws["A4"] = "Audit Run Details"
    exec_ws["B4"] = "Value"
    for ref in ("A4", "B4"):
        exec_ws[ref].fill = meta_header_fill
        exec_ws[ref].font = meta_header_font
        exec_ws[ref].alignment = Alignment(horizontal="center", vertical="center")
    exec_ws["A5"] = "Run Date"
    exec_ws["B5"] = (
        f'=IFERROR(INDEX(\'{audit_esc}\'!$B:$B,'
        f'MATCH("Run Timestamp",\'{audit_esc}\'!$A:$A,0)),"")'
    )
    exec_ws.merge_cells("A3:H3")
    exec_ws["A3"] = (
        f'=IFERROR(INDEX(\'{audit_esc}\'!$B:$B,'
        f'MATCH("Total URLs",\'{audit_esc}\'!$A:$A,0)),"") & " URLs crawled"'
    )
    exec_ws["A3"].font = Font(size=10, color=STD_NAVY, bold=True)
    exec_ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    for ref in ("A5", "B5"):
        exec_ws[ref].fill = PatternFill("solid", fgColor=PANEL_BG_COLOR)
        exec_ws[ref].alignment = Alignment(horizontal="center", vertical="center")


def _write_kpi_card_row(
    exec_ws: Worksheet,
    cards: list[tuple[str, str]],
    *,
    row: int,
    cards_per_row: int = _KPI_CARDS_PER_ROW,
) -> None:
    """Lay ``cards`` out 2 columns wide each, starting at column A.

    ``cards_per_row`` * 2 must stay within the frozen A–H stats band (2.5 UX
    overhaul) — 4 cards/row fits exactly (columns A–H); the previous 6/row
    contract spanned A–L, outside the new freeze boundary.
    """
    positions = [(1 + 2 * i, row) for i in range(cards_per_row)]
    exec_ws.row_dimensions[row].height = 22
    exec_ws.row_dimensions[row + 1].height = 30
    for (col, card_row), (label, value) in zip(positions, cards, strict=False):
        end_col = col + 1
        label_merge = (
            f"{get_column_letter(col)}{card_row}:{get_column_letter(end_col)}{card_row}"
        )
        value_merge = (
            f"{get_column_letter(col)}{card_row + 1}:{get_column_letter(end_col)}{card_row + 1}"
        )
        exec_ws.merge_cells(label_merge)
        exec_ws.merge_cells(value_merge)
        label_cell = exec_ws.cell(row=card_row, column=col, value=label)
        label_cell.font = Font(bold=True, size=11, color=STD_NAVY)
        label_cell.fill = _KPI_FILL
        label_cell.alignment = Alignment(horizontal="center", wrap_text=True)
        tip = _EXEC_CARD_TOOLTIPS.get(label)
        if tip:
            label_cell.comment = Comment(tip, "hype-frog")
        value_cell = exec_ws.cell(row=card_row + 1, column=col, value=value)
        value_cell.font = Font(bold=True, size=16)
        value_cell.fill = _KPI_FILL
        value_cell.alignment = Alignment(horizontal="center")


def _kpi_cards_row1(
    *,
    summary_metrics: SummaryMetricsPayload,
    dashboard_metrics: DashboardComputationResult,
    traffic_lift: int,
    psi: float | None,
) -> list[tuple[str, str]]:
    return [
        ("SEO Health (now)", f"{summary_metrics.health_score_pct:.0f}%"),
        (
            "SEO Health (illustrative projected)",
            f"{summary_metrics.projected_health_score_pct:.0f}%",
        ),
        ("AEO Readiness", _format_aeo_readiness_kpi(dashboard_metrics)),
        ("Critical URLs", str(summary_metrics.critical_url_count)),
        ("Traffic Lift (est.)", str(traffic_lift)),
        (
            "Performance (PSI)",
            f"{psi:.0f}" if psi is not None else "N/A",
        ),
    ]


def _kpi_cards_row2(
    *,
    dashboard_metrics: DashboardComputationResult,
    extra_rows: list[dict[str, Any]],
) -> list[tuple[str, str]]:
    clicks_total, impressions_total, avg_position = _gsc_totals(extra_rows)
    return [
        ("Avg TTFB (ms)", f"{dashboard_metrics.avg_ttfb_ms:.0f}"),
        ("Schema-Present URLs", str(dashboard_metrics.schema_urls)),
        ("Broken Link Instances", str(dashboard_metrics.broken_link_instances_total)),
        ("GSC Clicks (total)", str(clicks_total)),
        ("GSC Impressions (total)", str(impressions_total)),
        ("GSC Avg Position", f"{avg_position:.1f}" if avg_position > 0 else "N/A"),
    ]


def _write_kpi_card_grid(
    exec_ws: Worksheet,
    cards: list[tuple[str, str]],
    *,
    first_row: int,
    cards_per_row: int = _KPI_CARDS_PER_ROW,
) -> int:
    """Write ``cards`` in a grid confined to columns A–H, ``cards_per_row`` at a
    time, each occupying 2 rows (label + value). Returns the next free row."""
    row = first_row
    for start in range(0, len(cards), cards_per_row):
        chunk = cards[start : start + cards_per_row]
        _write_kpi_card_row(exec_ws, chunk, row=row, cards_per_row=cards_per_row)
        row += 2
    return row


def write_executive_briefing(
    writer: Any,
    *,
    summary_metrics: SummaryMetricsPayload,
    typed_main_rows: list[MainRowPayload],
    typed_extra_rows: list[ExtraRowPayload],
    priority_rows: list[dict[str, Any]],
    fixplan_rows: list[dict[str, Any]],
    hub_metrics_rows: list[dict[str, Any]] | None = None,
) -> None:
    """Create Executive Briefing — merged visual dashboard and triage navigation."""
    book = writer.book
    for legacy_name in (EXECUTIVE_DASHBOARD_SHEET, EXECUTIVE_BRIEFING_SHEET):
        if legacy_name in book.sheetnames:
            del book[legacy_name]

    exec_ws = book.create_sheet(title=_safe_sheet_name(EXECUTIVE_BRIEFING_SHEET))
    writer.sheets[EXECUTIVE_BRIEFING_SHEET] = exec_ws

    extra_rows = [dict(row.values) for row in typed_extra_rows]
    fixplan_payloads = [
        FixPlanRowPayload.model_validate({**row, "source_row": idx})
        for idx, row in enumerate(fixplan_rows, start=2)
    ]
    dashboard_metrics = compute_dashboard_metrics(
        summary_metrics=summary_metrics,
        technical_main_rows=typed_main_rows,
        technical_extra_rows=typed_extra_rows,
        fixplan_rows=fixplan_payloads,
        aeo_rows=typed_extra_rows,
    )
    layout = populate_chart_data_sheet(
        exec_ws,
        summary_metrics=summary_metrics,
        dashboard_metrics=dashboard_metrics,
        extra_rows=extra_rows,
        priority_rows=priority_rows,
        fixplan_rows=fixplan_rows,
        hub_metrics_rows=hub_metrics_rows,
    )

    psi = _avg_lighthouse_performance_mobile(extra_rows)
    traffic_lift = _traffic_lift_total(hub_metrics_rows)
    _write_briefing_header(exec_ws, summary_metrics=summary_metrics)
    all_kpi_cards = _kpi_cards_row1(
        summary_metrics=summary_metrics,
        dashboard_metrics=dashboard_metrics,
        traffic_lift=traffic_lift,
        psi=psi,
    ) + _kpi_cards_row2(
        dashboard_metrics=dashboard_metrics,
        extra_rows=extra_rows,
    )
    _write_kpi_card_grid(exec_ws, all_kpi_cards, first_row=_ROW_KPI1)
    _apply_executive_column_grid(exec_ws)
    _write_key_insights(
        exec_ws,
        _build_key_insights(summary_metrics, dashboard_metrics),
    )

    # 2.5 UX overhaul: all 6 chart sections start at the same row band
    # (_ROW_SEC_CHARTS / _ROW_CH_CHARTS) and are placed side by side going
    # right from column I — scrolling right (not down) reveals each one, so
    # they never fight the frozen A-H stats band above for vertical space.
    section_cols = _chart_section_start_cols()
    _reserve_chart_band(exec_ws, _ROW_CH_CHARTS, _CHART_BAND_ROWS)

    health_col, issues_col, actions_col, top_issues_col, status_col, tradaeo_col = section_cols

    _write_section_header(
        exec_ws,
        _ROW_SEC_CHARTS,
        "Health & illustrative projection",
        start_col=health_col,
        merge_to_col=health_col + _CHART_SECTION_COLUMN_SPANS[0] - 1,
    )
    _add_health_comparison_chart(
        exec_ws,
        layout,
        f"{get_column_letter(health_col)}{_ROW_CH_CHARTS}",
    )

    _write_section_header(
        exec_ws,
        _ROW_SEC_CHARTS,
        "Issue distribution — unique URLs vs issue instances (see source row 60+)",
        start_col=issues_col,
        merge_to_col=issues_col + _CHART_SECTION_COLUMN_SPANS[1] - 1,
    )
    owner_header = layout.owner_start + 1
    _add_severity_comparison_chart(
        exec_ws,
        layout,
        f"{get_column_letter(issues_col)}{_ROW_CH_CHARTS}",
    )
    _add_doughnut_chart(
        exec_ws,
        # L2 audit fix: explicitly scoped to unique-URL share, not total issue
        # volume — Copy Writer typically owns fewer *unique* URLs but a larger
        # share of total *issue instances* (see the "Issue instances" column
        # in the source data a few rows below), so an unqualified "workload"
        # title would misleadingly hide that behind a small doughnut slice.
        title="Owner workload — share of unique URLs (not total issue volume)",
        header_row=owner_header,
        data_rows=layout.owner_rows,
        anchor=f"{get_column_letter(issues_col + 6)}{_ROW_CH_CHARTS}",
        colors=_OWNER_SLICE_COLORS,
        value_col=_value_col(),
    )

    _write_section_header(
        exec_ws,
        _ROW_SEC_CHARTS,
        "Priority pages & content readiness",
        start_col=actions_col,
        merge_to_col=actions_col + _CHART_SECTION_COLUMN_SPANS[2] - 1,
    )
    _add_priority_bar_chart(
        exec_ws,
        layout,
        f"{get_column_letter(actions_col)}{_ROW_CH_CHARTS}",
    )
    _add_content_readiness_bar_chart(
        exec_ws,
        layout,
        f"{get_column_letter(actions_col + 6)}{_ROW_CH_CHARTS}",
    )

    _write_section_header(
        exec_ws,
        _ROW_SEC_CHARTS,
        "Top issues to fix first (by affected URLs)",
        start_col=top_issues_col,
        merge_to_col=top_issues_col + _CHART_SECTION_COLUMN_SPANS[3] - 1,
    )
    _add_top_issues_chart(
        exec_ws,
        layout,
        f"{get_column_letter(top_issues_col)}{_ROW_CH_CHARTS}",
    )

    _write_section_header(
        exec_ws,
        _ROW_SEC_CHARTS,
        "Status code & technical health",
        start_col=status_col,
        merge_to_col=status_col + _CHART_SECTION_COLUMN_SPANS[4] - 1,
    )
    _add_status_code_chart(
        exec_ws,
        layout,
        f"{get_column_letter(status_col)}{_ROW_CH_CHARTS}",
    )

    _write_section_header(
        exec_ws,
        _ROW_SEC_CHARTS,
        "Traditional SEO vs AEO Readiness",
        start_col=tradaeo_col,
        merge_to_col=tradaeo_col + _CHART_SECTION_COLUMN_SPANS[5] - 1,
    )
    _add_traditional_vs_aeo_chart(
        exec_ws,
        layout,
        f"{get_column_letter(tradaeo_col)}{_ROW_CH_CHARTS}",
    )

    apply_executive_briefing_triage(
        exec_ws,
        dashboard_metrics=dashboard_metrics,
        row_start=_BRIEFING_TRIAGE_START_ROW,
    )
    _write_source_data_hint(exec_ws)
    set_freeze_panes_safe(exec_ws, EXECUTIVE_BRIEFING_FREEZE_PANES)
    exec_ws.sheet_view.showGridLines = False


def write_executive_dashboard(
    writer: Any,
    *,
    summary_metrics: SummaryMetricsPayload,
    typed_main_rows: list[MainRowPayload],
    typed_extra_rows: list[ExtraRowPayload],
    priority_rows: list[dict[str, Any]],
    fixplan_rows: list[dict[str, Any]],
    hub_metrics_rows: list[dict[str, Any]] | None = None,
) -> None:
    """Deprecated alias — writes :data:`EXECUTIVE_BRIEFING_SHEET` (Phase 2 merge)."""
    write_executive_briefing(
        writer,
        summary_metrics=summary_metrics,
        typed_main_rows=typed_main_rows,
        typed_extra_rows=typed_extra_rows,
        priority_rows=priority_rows,
        fixplan_rows=fixplan_rows,
        hub_metrics_rows=hub_metrics_rows,
    )


__all__ = [
    "CHART_LABEL_COL",
    "CHART_SOURCE_FIRST_ROW",
    "CHART_VALUE2_COL",
    "CHART_VALUE_COL",
    "ChartDataLayout",
    "populate_chart_data_sheet",
    "write_executive_briefing",
    "write_executive_dashboard",
]
