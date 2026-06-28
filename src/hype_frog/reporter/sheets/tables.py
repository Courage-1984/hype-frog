from __future__ import annotations

from typing import Any

import math

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from hype_frog.reporter.sheets.config import (
    CONTENT_HUB_FREEZE_PANES,
    CONTENT_OPTIMISATION_HUB_SHEET,
    STD_NAVY,
    STD_WHITE,
)
from hype_frog.reporter.sheets.view_state import set_freeze_panes_safe


def normalize_table_headers(worksheet: Worksheet, header_row: int = 1) -> None:
    """Normalize table headers into unique, Excel-safe string values.

    Args:
        worksheet: Worksheet containing the target header row.
        header_row: 1-based row index containing the table headers.
    """
    seen: dict[str, int] = {}
    for col_idx in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=header_row, column=col_idx)
        raw = cell.value
        if (
            raw is None
            or (isinstance(raw, float) and math.isnan(raw))
            or not str(raw).strip()
        ):
            header = f"Column_{col_idx}"
        else:
            header = str(raw).replace("\n", " ").replace("\r", " ").strip()
            if len(header) > 255:
                header = header[:255].strip() or f"Column_{col_idx}"
        if header in seen:
            seen[header] += 1
            header = f"{header}_{seen[header]}"
        else:
            seen[header] = 0
        cell.value = header


def compute_exact_table_ref(worksheet: Worksheet, header_row: int) -> str | None:
    """Compute the exact A1 table reference from populated worksheet cells.

    Args:
        worksheet: Worksheet containing table-like data.
        header_row: 1-based header row.

    Returns:
        A1-style ref string (for example ``A1:H123``), or ``None`` when invalid.
    """
    if worksheet.max_row < header_row + 1:
        return None
    max_col = 1
    for row_idx in range(header_row, worksheet.max_row + 1):
        for col_idx in range(worksheet.max_column, 0, -1):
            value = worksheet.cell(row=row_idx, column=col_idx).value
            if value is not None and str(value) != "":
                if col_idx > max_col:
                    max_col = col_idx
                break
    if max_col < 1:
        return None
    return f"A{header_row}:{get_column_letter(max_col)}{worksheet.max_row}"


def apply_mock_table_styling(
    worksheet: Worksheet, min_col: int, max_col: int, min_row: int, max_row: int
) -> None:
    """Apply stable table-like styling without creating native Excel table objects.

    Args:
        worksheet: Worksheet to style.
        min_col: 1-based start column.
        max_col: 1-based end column.
        min_row: 1-based start row.
        max_row: 1-based end row.
    """
    if min_col > max_col or min_row > max_row:
        return
    if max_row >= min_row + 1:
        ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        worksheet.auto_filter.ref = ref

    header_fill = PatternFill(
        start_color=STD_NAVY, end_color=STD_NAVY, fill_type="solid"
    )
    header_font = Font(color=STD_WHITE, bold=True)
    for col_idx in range(min_col, max_col + 1):
        cell = worksheet.cell(row=min_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    # Per-cell zebra banding is O(rows × cols) — skip for large sheets to avoid
    # multi-minute formatting passes. Sheets above the threshold keep the navy
    # header and auto-filter; row striping is a cosmetic-only omission.
    if (max_row - min_row) <= 500:
        band_fill = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
        for row_idx in range(min_row + 1, max_row + 1):
            if row_idx % 2 == 0:
                for col_idx in range(min_col, max_col + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.fill.fill_type is None:
                        cell.fill = band_fill

    if max_row >= min_row:
        if worksheet.title == CONTENT_OPTIMISATION_HUB_SHEET:
            set_freeze_panes_safe(worksheet, CONTENT_HUB_FREEZE_PANES)
        else:
            set_freeze_panes_safe(worksheet, "A2")


__all__ = [
    "adjust_sheet_format",
    "apply_tab_hyperlinks",
    "normalize_table_headers",
    "compute_exact_table_ref",
    "apply_mock_table_styling",
]


def adjust_sheet_format(writer: Any, sheet_name: str) -> None:
    """Backward-compatible facade for legacy import paths.

    Args:
        writer: Pandas ExcelWriter-like object.
        sheet_name: Worksheet name to format.
    """
    from hype_frog.reporter.sheets.tables_impl import adjust_sheet_format as _impl

    return _impl(writer, sheet_name)


def apply_tab_hyperlinks(writer: Any) -> None:
    """Backward-compatible facade for legacy import paths.

    Args:
        writer: Pandas ExcelWriter-like object.
    """
    from hype_frog.reporter.sheets.tables_impl import apply_tab_hyperlinks as _impl

    return _impl(writer)
