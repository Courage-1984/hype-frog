from __future__ import annotations

from collections.abc import Callable
from typing import Any

from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from hype_frog.core.url_normalization import normalize_url_key
from hype_frog.reporter.engine_rows import content_hub_column_letter
from hype_frog.reporter.sheets.config import (
    CONTENT_HUB_DATA_START_ROW,
    CONTENT_OPTIMISATION_HUB_SHEET,
)
from hype_frog.reporter.sheets.layout import main_sheet_url_column_letter

TECHNICAL_DIAGNOSTICS_SHEET = "Technical Diagnostics"


def _technical_diagnostics_jump_formula(url_col_letter: str, row: int, *, link_label: str) -> str:
    """Build a HYPERLINK+MATCH formula targeting the merged Technical Diagnostics sheet."""
    td = TECHNICAL_DIAGNOSTICS_SHEET
    return (
        f'=IFERROR(HYPERLINK("#\'{td}\'!A"&MATCH({url_col_letter}{row},\'{td}\'!A:A,0),'
        f'"{link_label}"),HYPERLINK("#\'{td}\'!A1","{link_label}"))'
    )


def _fixplan_hub_status_formula(fixplan_url_letter: str, row: int) -> str:
    """Return Hub ``Status`` for a FixPlan URL via INDEX/MATCH on the Content Hub.

    Hub data starts at row ``CONTENT_HUB_DATA_START_ROW`` (banner + headers).
    Column letters are resolved from the canonical export order so reordering does not
    silently join on SEO Score.
    """
    hub = CONTENT_OPTIMISATION_HUB_SHEET
    status_l = content_hub_column_letter("Status")
    url_l = content_hub_column_letter("URL")
    data_start = CONTENT_HUB_DATA_START_ROW
    return (
        f"=IFERROR(INDEX('{hub}'!{status_l}{data_start}:{status_l}10000,"
        f"MATCH({fixplan_url_letter}{row},'{hub}'!{url_l}{data_start}:{url_l}10000,0)),"
        f'"Not in Hub")'
    )


def _reference_tab_jump_formula(ref_letter: str, url_letter: str, row: int) -> str:
    """Build a HYPERLINK+MATCH jump into a per-row reference tab named in ``ref_letter``.

    The reference tab name is a runtime cell value, so it is wrapped in single quotes
    for both the HYPERLINK target and the INDIRECT lookup, and any embedded apostrophes
    are doubled with ``SUBSTITUTE`` so names with spaces or apostrophes resolve correctly.
    """
    # ``q`` is the sheet name with single quotes escaped (``'`` -> ``''``) for safe
    # quoting inside an Excel sheet reference.
    q = f'SUBSTITUTE({ref_letter}{row},"\'","\'\'")'
    return (
        f'=IFERROR(HYPERLINK("#\'"&{q}&"\'!A"&'
        f'MATCH({url_letter}{row},INDIRECT("\'"&{q}&"\'!A:A"),0),"Open"),'
        f'HYPERLINK("#\'"&{q}&"\'!A1","Open"))'
    )


def is_safe_hyperlink_target(
    target: str, *, disable_external_links_and_images: bool
) -> bool:
    """Validate whether a hyperlink target is safe to emit into a workbook.

    Args:
        target: Hyperlink target candidate.
        disable_external_links_and_images: Feature flag disabling outbound links.

    Returns:
        True when the link should be written; otherwise False.
    """
    if disable_external_links_and_images:
        return False
    return bool(target) and len(target) <= 255


def sanitize_excel_url(url_value: Any) -> str:
    """Strip control characters and quote marks from URL-like cell text.

    Args:
        url_value: Raw URL-like value from dataframe/worksheet content.

    Returns:
        A cleaned string safe for workbook formulas/hyperlinks.
    """
    raw = str(url_value or "").strip()
    if not raw:
        return ""
    raw = "".join(ch for ch in raw if ord(ch) >= 32).replace('"', "").replace("'", "")
    return raw


def normalize_url_for_match(url_value: Any) -> str:
    """Normalize a URL key for deterministic cross-sheet matching.

    Args:
        url_value: Raw URL-like value.

    Returns:
        Canonicalized key used for lookups.
    """
    return normalize_url_key(sanitize_excel_url(url_value))


from hype_frog.reporter.sheets.sheet_rows import sheet_data_start_row


def add_url_navigation_links(
    writer: Any,
    worksheet: Worksheet,
    sheet_name: str,
    *,
    debug_excel_isolation_mode: bool,
    disable_external_links_and_images: bool,
    header_index_fn: Callable[..., dict[str, int]],
    data_start_row: int = 2,
    header_row: int = 1,
) -> None:
    """Add per-row URL hyperlinks and optional ``Open in Main`` helper links.

    Args:
        writer: Pandas ExcelWriter-like object containing ``book`` and sheets.
        worksheet: Worksheet currently being formatted.
        sheet_name: Name of current sheet.
        debug_excel_isolation_mode: Flag that disables cross-sheet link creation.
        disable_external_links_and_images: Flag that disables external hyperlinks.
        header_index_fn: Callable returning header-to-column index mapping.
    """
    if debug_excel_isolation_mode:
        return
    if sheet_name == CONTENT_OPTIMISATION_HUB_SHEET:
        headers_r2: dict[str, int] = {}
        for c in range(1, worksheet.max_column + 1):
            raw = worksheet.cell(row=2, column=c).value
            if raw is not None and str(raw).strip():
                headers_r2[str(raw).strip()] = c
        url_col = headers_r2.get("URL")
        if not url_col or worksheet.max_row < 3:
            return
        for r in range(3, worksheet.max_row + 1):
            url_cell = worksheet.cell(row=r, column=url_col)
            url_val = str(url_cell.value or "").strip()
            if url_val.startswith(("http://", "https://")) and is_safe_hyperlink_target(
                url_val,
                disable_external_links_and_images=disable_external_links_and_images,
            ):
                url_cell.hyperlink = url_val
                url_cell.style = "Hyperlink"
                url_cell.alignment = Alignment(wrap_text=False, vertical="center")
        return

    headers = header_index_fn(worksheet)
    url_col = headers.get("URL")
    if not url_col or worksheet.max_row < data_start_row:
        return
    for r in range(data_start_row, worksheet.max_row + 1):
        url_cell = worksheet.cell(row=r, column=url_col)
        url_val = str(url_cell.value or "").strip()
        if url_val.startswith(("http://", "https://")) and is_safe_hyperlink_target(
            url_val, disable_external_links_and_images=disable_external_links_and_images
        ):
            url_cell.hyperlink = url_val
            url_cell.style = "Hyperlink"
            url_cell.alignment = Alignment(wrap_text=False, vertical="center")

    if sheet_name not in {"Main", "Dashboard"} and "Main" in writer.book.sheetnames:
        if "Open in Main" not in headers:
            new_col = worksheet.max_column + 1
            worksheet.cell(row=header_row, column=new_col, value="Open in Main")
            url_col_letter = get_column_letter(url_col)
            main_u = main_sheet_url_column_letter()
            for r in range(data_start_row, worksheet.max_row + 1):
                worksheet.cell(
                    row=r,
                    column=new_col,
                    value=(
                        f'=IFERROR(HYPERLINK("#\'Main\'!{main_u}"&MATCH(TRIM({url_col_letter}{r}),'
                        f'\'Main\'!{main_u}:{main_u},0),"Open"),HYPERLINK("#\'Main\'!{main_u}1","Open"))'
                    ),
                )


def apply_cross_sheet_links(
    writer: Any,
    worksheet: Worksheet,
    sheet_name: str,
    *,
    debug_excel_isolation_mode: bool,
    header_index_fn: Callable[..., dict[str, int]],
    data_start_row: int = 2,
    header_row: int = 1,
) -> None:
    """Attach cross-sheet helper links used for analyst navigation.

    Args:
        writer: Pandas ExcelWriter-like object containing workbook metadata.
        worksheet: Worksheet currently being formatted.
        sheet_name: Name of current sheet.
        debug_excel_isolation_mode: Flag that disables cross-sheet links.
        header_index_fn: Callable returning header-to-column index mapping.
    """
    if debug_excel_isolation_mode:
        return
    headers = header_index_fn(worksheet)
    fix_data_start = sheet_data_start_row("FixPlan")
    if sheet_name in {"Summary", "Issue Register"}:
        issue_col = headers.get("Issue")
        fix_ws = writer.book["FixPlan"] if "FixPlan" in writer.book.sheetnames else None
        fix_headers = header_index_fn(fix_ws) if fix_ws else {}
        fix_issue_col = fix_headers.get("Issue Type")
        fix_issue_to_row: dict[str, int] = {}
        if fix_ws and fix_issue_col:
            for r in range(fix_data_start, fix_ws.max_row + 1):
                issue_name = str(
                    fix_ws.cell(row=r, column=fix_issue_col).value or ""
                ).strip()
                if issue_name and issue_name not in fix_issue_to_row:
                    fix_issue_to_row[issue_name] = r
        if issue_col:
            for r in range(data_start_row, worksheet.max_row + 1):
                issue = str(worksheet.cell(row=r, column=issue_col).value or "").strip()
                if issue and not issue.startswith("==="):
                    issue_cell = worksheet.cell(row=r, column=issue_col)
                    fix_row = fix_issue_to_row.get(issue)
                    if fix_row:
                        issue_cell.hyperlink = f"#FixPlan!A{fix_row}"
                        issue_cell.style = "Hyperlink"
    if sheet_name == "Main":
        url_col = headers.get("URL")
        if url_col:
            existing_col = headers.get("Technical View")
            if existing_col:
                target_col = existing_col
            else:
                target_col = worksheet.max_column + 1
                worksheet.cell(row=header_row, column=target_col, value="Technical View")
            col_letter = get_column_letter(url_col)
            for r in range(data_start_row, worksheet.max_row + 1):
                worksheet.cell(
                    row=r,
                    column=target_col,
                    value=_technical_diagnostics_jump_formula(
                        col_letter, r, link_label="Open Technical"
                    ),
                )
    if sheet_name == "Priority URLs":
        url_col = headers.get("URL")
        if url_col:
            new_col = headers.get("Open in Technical")
            if not new_col:
                new_col = worksheet.max_column + 1
                worksheet.cell(row=header_row, column=new_col, value="Open in Technical")
            url_letter = get_column_letter(url_col)
            for r in range(data_start_row, worksheet.max_row + 1):
                worksheet.cell(
                    row=r,
                    column=new_col,
                    value=_technical_diagnostics_jump_formula(
                        url_letter, r, link_label="Open"
                    ),
                )
    if sheet_name in {"IssueInventory", "Issue Register"}:
        url_col = headers.get("URL")
        issue_col = headers.get("Issue")
        reference_tab_col = headers.get("Reference Tab") or headers.get(
            "Reference Area"
        )
        if url_col:
            new_col = headers.get("Open in Main")
            if not new_col:
                new_col = worksheet.max_column + 1
                worksheet.cell(row=header_row, column=new_col, value="Open in Main")
            url_letter = get_column_letter(url_col)
            main_u = main_sheet_url_column_letter()
            for r in range(data_start_row, worksheet.max_row + 1):
                worksheet.cell(
                    row=r,
                    column=new_col,
                    value=(
                        f'=IFERROR(HYPERLINK("#\'Main\'!{main_u}"&MATCH(TRIM({url_letter}{r}),'
                        f'\'Main\'!{main_u}:{main_u},0),"Open"),HYPERLINK("#\'Main\'!{main_u}1","Open"))'
                    ),
                )
        if reference_tab_col:
            open_ref_col = headers.get("Open in Reference")
            if not open_ref_col:
                open_ref_col = worksheet.max_column + 1
                worksheet.cell(row=header_row, column=open_ref_col, value="Open in Reference")
            ref_letter = get_column_letter(reference_tab_col)
            url_letter = get_column_letter(url_col) if url_col else "A"
            for r in range(data_start_row, worksheet.max_row + 1):
                worksheet.cell(
                    row=r,
                    column=open_ref_col,
                    value=_reference_tab_jump_formula(ref_letter, url_letter, r),
                )
        if issue_col and "FixPlan" in writer.book.sheetnames:
            fix_ws = writer.book["FixPlan"]
            fix_headers = header_index_fn(fix_ws)
            fix_issue_col = fix_headers.get("Issue Type")
            fix_issue_rows: dict[str, int] = {}
            if fix_issue_col:
                for r in range(fix_data_start, fix_ws.max_row + 1):
                    key = str(
                        fix_ws.cell(row=r, column=fix_issue_col).value or ""
                    ).strip()
                    if key and key not in fix_issue_rows:
                        fix_issue_rows[key] = r
                for r in range(data_start_row, worksheet.max_row + 1):
                    issue = str(
                        worksheet.cell(row=r, column=issue_col).value or ""
                    ).strip()
                    target_row = fix_issue_rows.get(issue)
                    if target_row:
                        cell = worksheet.cell(row=r, column=issue_col)
                        cell.hyperlink = f"#FixPlan!A{target_row}"
                        cell.style = "Hyperlink"
    if sheet_name == "AIOSEO Recommendations":
        url_col = headers.get("URL")
        if url_col:
            technical_col = headers.get("Open in Technical")
            if not technical_col:
                technical_col = worksheet.max_column + 1
                worksheet.cell(row=header_row, column=technical_col, value="Open in Technical")
            url_letter = get_column_letter(url_col)
            for r in range(data_start_row, worksheet.max_row + 1):
                worksheet.cell(
                    row=r,
                    column=technical_col,
                    value=_technical_diagnostics_jump_formula(
                        url_letter, r, link_label="Open"
                    ),
                )
    if sheet_name == "FixPlan" and CONTENT_OPTIMISATION_HUB_SHEET in writer.book.sheetnames:
        headers = header_index_fn(worksheet)
        url_col = headers.get("URL")
        hub_status_col = headers.get("Hub Status (Content Hub)")
        if not hub_status_col:
            hub_status_col = worksheet.max_column + 1
            worksheet.cell(
                row=header_row, column=hub_status_col, value="Hub Status (Content Hub)"
            )
        if url_col:
            u_letter = get_column_letter(url_col)
            for r in range(data_start_row, worksheet.max_row + 1):
                worksheet.cell(
                    row=r,
                    column=hub_status_col,
                    value=_fixplan_hub_status_formula(u_letter, r),
                )


def apply_editor_url_column_hyperlinks(
    worksheet: Worksheet,
    sheet_name: str,
    *,
    disable_external_links_and_images: bool,
) -> None:
    """Turn Elementor / AIOSEO edit URLs into real outbound hyperlinks (when enabled)."""
    if sheet_name == CONTENT_OPTIMISATION_HUB_SHEET:
        header_row, first_data_row, column_name = 2, 3, "Elementor Builder Link"
    elif sheet_name == "AIOSEO Recommendations":
        # AIOSEO carries a row-1 return banner, so headers sit on row 2.
        header_row, first_data_row, column_name = 2, 3, "Direct Edit Link"
    else:
        return
    if worksheet.max_row < first_data_row:
        return
    headers: dict[str, int] = {}
    for col_idx in range(1, worksheet.max_column + 1):
        raw = worksheet.cell(row=header_row, column=col_idx).value
        if raw is not None:
            headers[str(raw).strip()] = col_idx
    col_idx = headers.get(column_name)
    if not col_idx:
        return
    link_font = Font(color="0563C1", underline="single")
    for row_idx in range(first_data_row, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_idx, column=col_idx)
        raw_val = cell.value
        if isinstance(raw_val, str) and raw_val.strip().upper().startswith(
            "=HYPERLINK("
        ):
            cell.font = link_font
            cell.alignment = Alignment(wrap_text=False, vertical="center")
            continue
        val = sanitize_excel_url(raw_val)
        if val.startswith(("http://", "https://")) and is_safe_hyperlink_target(
            val, disable_external_links_and_images=disable_external_links_and_images
        ):
            cell.hyperlink = val
            cell.style = "Hyperlink"
            cell.font = link_font
            cell.alignment = Alignment(wrap_text=False, vertical="center")
