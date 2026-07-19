from __future__ import annotations

from typing import Any

from openpyxl.cell.cell import MergedCell
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.views import Selection

from hype_frog.core import get_logger
from hype_frog.reporter.engine_guardrails import friendly_toc_description
from hype_frog.reporter.sheets.config import (
    AIOSEO_RECOMMENDATIONS_SHEET,
    CONTENT_HUB_FREEZE_PANES,
    CONTENT_OPTIMISATION_HUB_SHEET,
    CONTENT_PLANNER_SHEET,
    CRAWL_LOG_SHEET,
    EXECUTIVE_BRIEFING_FREEZE_PANES,
    EXECUTIVE_BRIEFING_SHEET,
    ROBOTS_ANALYSIS_SHEET,
    SHEET_ZOOM_OVERRIDES,
    ZEBRA_BAND,
)
from hype_frog.reporter.sheets.sheet_rows import sheet_data_header_row
from hype_frog.reporter.sheets.workbook_layout import (
    ADVANCED_WORKBOOK_TAB_ORDER,
    PREFERRED_WORKBOOK_TAB_ORDER,
    SHEETS_EXCLUDED_FROM_TOC,
    TOC_ADVANCED_SECTION_LABEL,
    TOC_ADVANCED_SECTION_LABEL_SHOWN,
    TOC_PRIMARY_SECTION_LABEL,
    VISIBLE_WORKBOOK_TAB_ORDER,
    apply_workbook_tab_layout,
    excel_sheet_link_target,
)

# Re-export for workbook_audit and export_flow.
__all__ = [
    "PREFERRED_WORKBOOK_TAB_ORDER",
    "apply_workbook_toc_and_links",
]

_PREFERRED_TAB_SET = frozenset(PREFERRED_WORKBOOK_TAB_ORDER)

logger = get_logger(__name__)


def apply_workbook_toc_and_links(
    writer: Any,
    *,
    debug_excel_isolation_mode: bool,
    disable_non_core_freeze_panes: bool,
    std_navy: str,
    std_white: str,
    std_blue: str,
    hide_advanced_tabs: bool = True,
) -> None:
    def _clear_orphaned_selection(ws) -> None:
        try:
            ws.views.sheetView[0].selection = []
        except Exception as exc:
            logger.debug("Could not clear orphaned selection on %s: %s", ws.title, exc)

    def _set_freeze_panes_safe(ws, value: str | None) -> None:
        view = ws.views.sheetView[0]
        if not view.selection:
            view.selection = [Selection(activeCell="A1", sqref="A1")]
        ws.freeze_panes = value

    def _append_toc_row(
        toc_ws,
        wb_ref,
        row_ptr: int,
        sheet_name: str,
        *,
        std_blue_color: str,
        section_label: str | None = None,
    ) -> int:
        if section_label:
            toc_ws.cell(row=row_ptr, column=1, value=section_label)
            toc_ws.cell(row=row_ptr, column=1).font = Font(bold=True, color=std_navy)
            toc_ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=3)
            return row_ptr + 1
        if sheet_name not in wb_ref.sheetnames:
            return row_ptr
        safe = excel_sheet_link_target(sheet_name)
        disp = str(sheet_name).replace('"', "'")
        a_cell = toc_ws.cell(row=row_ptr, column=1)
        a_cell.value = f'=HYPERLINK("#\'{safe}\'!A1","{disp}")'
        a_cell.font = Font(color=std_blue_color, underline="single", bold=True)
        b_cell = toc_ws.cell(row=row_ptr, column=2)
        b_cell.value = f'=HYPERLINK("#\'{safe}\'!A1","Open")'
        b_cell.font = Font(color=std_blue_color, underline="single", bold=True)
        toc_ws.cell(
            row=row_ptr, column=3, value=friendly_toc_description(sheet_name)
        )
        return row_ptr + 1

    def _rebuild_toc_body(toc_ws, wb_ref) -> None:
        """Primary workflow first, then advanced (hidden) diagnostics."""
        while toc_ws.max_row >= 3:
            toc_ws.delete_rows(3)
        row_ptr = 3
        row_ptr = _append_toc_row(
            toc_ws,
            wb_ref,
            row_ptr,
            "",
            std_blue_color=std_blue,
            section_label=TOC_PRIMARY_SECTION_LABEL,
        )
        for sheet_name in VISIBLE_WORKBOOK_TAB_ORDER:
            if sheet_name == "Table of Contents":
                continue
            row_ptr = _append_toc_row(
                toc_ws, wb_ref, row_ptr, sheet_name, std_blue_color=std_blue
            )
        row_ptr = _append_toc_row(
            toc_ws,
            wb_ref,
            row_ptr,
            "",
            std_blue_color=std_blue,
            section_label=(
                TOC_ADVANCED_SECTION_LABEL
                if hide_advanced_tabs
                else TOC_ADVANCED_SECTION_LABEL_SHOWN
            ),
        )
        for sheet_name in ADVANCED_WORKBOOK_TAB_ORDER:
            if sheet_name in SHEETS_EXCLUDED_FROM_TOC:
                continue
            row_ptr = _append_toc_row(
                toc_ws, wb_ref, row_ptr, sheet_name, std_blue_color=std_blue
            )
        for sheet_name in wb_ref.sheetnames:
            if (
                sheet_name == "Table of Contents"
                or sheet_name in _PREFERRED_TAB_SET
                or sheet_name in SHEETS_EXCLUDED_FROM_TOC
            ):
                continue
            row_ptr = _append_toc_row(
                toc_ws, wb_ref, row_ptr, sheet_name, std_blue_color=std_blue
            )

    if debug_excel_isolation_mode:
        return
    wb = writer.book
    wb.calculation.calcMode = "auto"

    if "Table of Contents" not in wb.sheetnames:
        wb.create_sheet("Table of Contents", 0)

    apply_workbook_tab_layout(wb, hide_advanced_tabs=hide_advanced_tabs)

    toc_ws = wb["Table of Contents"]
    toc_ws["A1"] = "Table of Contents"
    toc_ws["A1"].font = Font(color=std_navy, bold=True, size=14)
    toc_ws["A2"] = "Section"
    toc_ws["B2"] = "Open"
    toc_ws["C2"] = "Description"
    for ref in ("A2", "B2", "C2"):
        toc_ws[ref].fill = PatternFill("solid", fgColor=std_navy)
        toc_ws[ref].font = Font(color=std_white, bold=True)
    toc_ws.column_dimensions["A"].width = 45
    toc_ws.column_dimensions["B"].width = 12
    # 100: friendly_toc_description() entries are curated to fit within this width
    # (longest is ~94 chars) so every description renders on one line — no wrapping,
    # no inflated row heights.
    _toc_desc_width = 100
    toc_ws.column_dimensions["C"].width = _toc_desc_width
    toc_ws.freeze_panes = "A3"
    _rebuild_toc_body(toc_ws, wb)
    section_label_rows: set[int] = set()
    for row_idx in range(3, toc_ws.max_row + 1):
        cell = toc_ws.cell(row=row_idx, column=3)
        if isinstance(cell, MergedCell):
            section_label_rows.add(row_idx)
            continue
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        text = str(cell.value or "")
        wrap_lines = max(1, (len(text) // _toc_desc_width) + 1)
        if wrap_lines > 1:
            toc_ws.row_dimensions[row_idx].height = min(60, 15 * wrap_lines)

    # Zebra banding on sheet-link rows only — skip the bold navy section-label rows
    # (merged A:C) so their own styling isn't double-painted. Each contiguous run of
    # link rows between section labels gets its own CF range; MOD(ROW(),2) still
    # alternates naturally across runs since it keys off the absolute row number.
    # Guarded against re-adding on repeated builds (e.g. --regen-report re-invokes
    # this function): _rebuild_toc_body rewrites row *values* but does not clear
    # previously-attached conditional-formatting rules from an earlier build.
    _has_zebra_rule = any(
        any("MOD(ROW(),2)=0" in (rule.formula or []) for rule in rules)
        for rules in toc_ws.conditional_formatting._cf_rules.values()
    )
    if not _has_zebra_rule:
        run_start: int | None = None
        for row_idx in range(3, toc_ws.max_row + 2):
            is_label_or_end = row_idx > toc_ws.max_row or row_idx in section_label_rows
            if not is_label_or_end and run_start is None:
                run_start = row_idx
            elif is_label_or_end and run_start is not None:
                toc_ws.conditional_formatting.add(
                    f"A{run_start}:C{row_idx - 1}",
                    FormulaRule(
                        formula=["MOD(ROW(),2)=0"],
                        stopIfTrue=False,
                        fill=PatternFill("solid", fgColor=ZEBRA_BAND),
                    ),
                )
                run_start = None

    link_map: list[tuple[str, str]] = [
        ("Issue Register", "Reference Area"),
        ("FixPlan", "Detail Reference Tab"),
        ("FixPlan", "Jump to Details"),
        ("Executive Briefing", "Target Tab"),
        (AIOSEO_RECOMMENDATIONS_SHEET, "Reference Tab"),
    ]
    for sheet_name, col_header in link_map:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header_row = sheet_data_header_row(sheet_name)
        headers = [c.value for c in ws[header_row]]
        if col_header not in headers:
            continue
        col_idx = headers.index(col_header) + 1
        for row_idx in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            target = str(cell.value or "").strip()
            if target and target in wb.sheetnames:
                safe = excel_sheet_link_target(target)
                cell.hyperlink = f"#'{safe}'!A1"
                cell.style = "Hyperlink"
    for tab_name in wb.sheetnames:
        ws = wb[tab_name]
        # Uniform, final pass — overrides any earlier per-builder gridline/zoom
        # settings so every tab looks consistent regardless of build order.
        ws.sheet_view.showGridLines = False
        zoom = SHEET_ZOOM_OVERRIDES.get(tab_name)
        if zoom:
            ws.sheet_view.zoomScale = zoom
        if tab_name == "Table of Contents":
            _set_freeze_panes_safe(ws, "A3")
            continue
        if tab_name == EXECUTIVE_BRIEFING_SHEET:
            _set_freeze_panes_safe(ws, EXECUTIVE_BRIEFING_FREEZE_PANES)
            _clear_orphaned_selection(ws)
            continue
        if disable_non_core_freeze_panes and tab_name not in {
            "Main",
            EXECUTIVE_BRIEFING_SHEET,
        }:
            _set_freeze_panes_safe(ws, None)
            _clear_orphaned_selection(ws)
            continue
        if tab_name not in {"Main"} and (
            ws.max_row < 10 or ws.max_column < 5
        ):
            _set_freeze_panes_safe(ws, None)
            ws.auto_filter.ref = None
            _clear_orphaned_selection(ws)
            continue
        wide_sheets = {
            "Main",
            "Technical",
            "Technical Diagnostics",
            CONTENT_OPTIMISATION_HUB_SHEET,
            "FixPlan",
            "Content",
            "Links",
            "Schema & Metadata",
            "Indexability",
            "Priority URLs",
            "AEO",
            "Redirects",
            ROBOTS_ANALYSIS_SHEET,
            CRAWL_LOG_SHEET,
            AIOSEO_RECOMMENDATIONS_SHEET,
            "Security",
            "Summary",
            "Quick Reference Guide",
            "LinksDetail",
            "Media",
            "Pattern and Template Issues",
            "Duplicates",
            "PSI Performance",
            "Issue Register",
            "Content & AI Readiness",
            "Link Intelligence",
            "Template & Duplication Risks",
            "Playbook",
            "Audit Run Details",
            "DeltaFromPreviousRun",
            "CrawlGraph",
            "SitemapQA",
        }
        standard_data_sheets = {
            "Content",
            "Links",
            "Schema & Metadata",
            "Indexability",
            "Priority URLs",
        }
        if tab_name == CONTENT_OPTIMISATION_HUB_SHEET:
            target_freeze = (
                CONTENT_HUB_FREEZE_PANES
                if ws.max_row >= 3 and ws.max_column >= 9
                else None
            )
            _set_freeze_panes_safe(ws, target_freeze)
        elif tab_name == CONTENT_PLANNER_SHEET:
            _set_freeze_panes_safe(ws, "E2")
        elif tab_name in standard_data_sheets:
            target_freeze = "B2" if ws.max_row >= 2 and ws.max_column >= 2 else None
            _set_freeze_panes_safe(ws, target_freeze)
        elif tab_name in wide_sheets:
            target_freeze = "C2" if ws.max_row >= 2 and ws.max_column >= 3 else None
            _set_freeze_panes_safe(ws, target_freeze)
