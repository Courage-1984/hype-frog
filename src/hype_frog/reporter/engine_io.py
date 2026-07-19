from __future__ import annotations

import gc
import re
from collections.abc import Iterable, Iterator
from typing import Any
from urllib.parse import quote, unquote, urlsplit, urlunsplit

import pandas as pd
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter

from hype_frog.checkpoint.cache import AuditCache
from hype_frog.core import get_logger
from hype_frog.core.models import ExtraRowPayload, MainRowPayload
from hype_frog.core.url_normalization import normalize_url_key
from hype_frog.pipeline.broken_links import link_inventory_broken_per_source_formula
from hype_frog.reporter.sheets.config import THEME_HEADER_BG, THEME_HEADER_TEXT
from hype_frog.reporter.stream_workbook import is_write_only_writer

logger = get_logger(__name__)

_HEADER_STYLE_NAME = "hf_table_header"

_ILLEGAL_XLSX_CHARS_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F]")
_INVALID_SHEET_CHARS_RE = re.compile(r"[:\\/*?\[\]]")

# Chunks at or above this size trigger an explicit ``gc.collect()`` after
# their rows are flushed. Tuned to balance teardown cost against the cost of
# 10k+ row audits keeping deserialised JSON dicts alive for too long.
_GC_COLLECT_CHUNK_THRESHOLD: int = 1000


def _safe_sheet_name(name: str) -> str:
    cleaned = _INVALID_SHEET_CHARS_RE.sub("_", str(name or "Sheet"))
    cleaned = cleaned[:31]
    return cleaned or "Sheet"


def _sanitize_excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return _ILLEGAL_XLSX_CHARS_RE.sub("", value)
    return value


def load_cached_rows(
    cache: AuditCache,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    main_rows: list[dict[str, Any]] = []
    extra_rows: list[dict[str, Any]] = []
    for result in cache.iter_results():
        main_rows.append(result["main"])
        extra_rows.append(result["extra"])
    return main_rows, extra_rows


def build_core_dataframes(
    cache: AuditCache,
) -> tuple[pd.DataFrame, pd.DataFrame, list[dict[str, Any]], list[dict[str, Any]]]:
    main_rows, extra_rows = load_cached_rows(cache)
    return pd.DataFrame(main_rows), pd.DataFrame(extra_rows), main_rows, extra_rows


def _row_to_mapping(
    row: dict[str, Any] | MainRowPayload | ExtraRowPayload,
) -> dict[str, Any]:
    if isinstance(row, (MainRowPayload, ExtraRowPayload)):
        return row.values
    return row


def apply_link_intelligence_summary_broken_formulas(workbook: Any) -> None:
    """Set Broken Internal Links Count and Actionable Fixes on Summary rows.

    Reads Link Intelligence's own Detail rows (folded in from the former standalone
    "Link Inventory" sheet — same sheet now, not a cross-sheet reference).
    """
    try:
        names = list(workbook.sheetnames)
    except Exception as exc:
        logger.debug("Could not read workbook sheet names: %s", exc)
        return
    if "Link Intelligence" not in names:
        return
    ws = workbook["Link Intelligence"]
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        key = str(ws.cell(1, col).value or "").strip()
        if key:
            headers[key] = col
    rt_col = headers.get("Record Type")
    brk_col = headers.get("Broken Internal Links Count")
    act_col = headers.get("Actionable Fixes")
    if not rt_col or not brk_col:
        return
    brk_letter = get_column_letter(brk_col)
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row, rt_col).value or "").strip() != "Summary":
            continue
        # Self-referencing via ROW()/INDIRECT (not a baked-in "$A{row}" literal):
        # add_return_to_briefing_strip() later calls worksheet.insert_rows(1),
        # which shifts cells but does not rewrite formula text — a literal row
        # number here would point one row too high after that insert.
        formula = link_inventory_broken_per_source_formula('INDIRECT("$A"&ROW())')
        ws.cell(row=row, column=brk_col, value=formula)
        if act_col:
            ws.cell(
                row=row,
                column=act_col,
                value=(
                    f'=IF({brk_letter}{row}>0,"Fix "&{brk_letter}{row}&'
                    f'" broken links (see Detail rows below).","")'
                ),
            )


def _ensure_header_named_style(workbook: Any) -> str:
    if _HEADER_STYLE_NAME in workbook.named_styles:
        return _HEADER_STYLE_NAME
    header_style = NamedStyle(
        name=_HEADER_STYLE_NAME,
        font=Font(bold=True, color=THEME_HEADER_TEXT),
        fill=PatternFill("solid", fgColor=THEME_HEADER_BG),
    )
    workbook.add_named_style(header_style)
    return _HEADER_STYLE_NAME


def _append_header_row(ws: Any, workbook: Any, columns: list[str]) -> None:
    if getattr(workbook, "write_only", False):
        style_name = _ensure_header_named_style(workbook)
        header_cells = []
        for col in columns:
            cell = WriteOnlyCell(ws, value=col)
            cell.style = style_name
            header_cells.append(cell)
        ws.append(header_cells)
        return
    ws.append(columns)


def _get_or_create_sheet(writer: Any, sheet_name: str) -> Any:
    safe_name = _safe_sheet_name(sheet_name)
    if sheet_name in writer.sheets:
        return writer.sheets[sheet_name]
    ws = writer.book.create_sheet(title=safe_name)
    writer.sheets[sheet_name] = ws
    return ws


def _pad_write_only_sheet(writer: Any, sheet_name: str, target_row: int) -> Any:
    """Pad a write-only sheet with blank rows until ``target_row`` (1-based)."""
    ws = _get_or_create_sheet(writer, sheet_name)
    if not is_write_only_writer(writer):
        return ws
    current = writer.sheet_row_count(sheet_name)
    pad_needed = max(0, target_row - 1 - current)
    for _ in range(pad_needed):
        ws.append([])
    if pad_needed and hasattr(writer, "record_rows_appended"):
        writer.record_rows_appended(sheet_name, pad_needed)
    return ws


def write_dataframe_sheet(
    writer: Any,
    df: pd.DataFrame,
    sheet_name: str,
    *,
    startrow: int = 1,
    include_header: bool = True,
) -> None:
    """Write a DataFrame via append (write-only safe) or pandas fallback."""
    if df.empty and not include_header:
        return
    if not is_write_only_writer(writer):
        from hype_frog.pipeline.export import to_excel_safe

        kwargs: dict[str, Any] = {"index": False}
        if startrow > 1:
            kwargs["startrow"] = startrow - 1
        to_excel_safe(df, writer, sheet_name, **kwargs)
        return

    ws = _pad_write_only_sheet(writer, sheet_name, startrow)
    columns = [str(col) for col in df.columns]
    rows_appended = 0
    if include_header:
        _append_header_row(ws, writer.book, columns)
        rows_appended += 1
    for row in df.itertuples(index=False, name=None):
        ws.append([_sanitize_excel_value(value) for value in row])
        rows_appended += 1
    if hasattr(writer, "record_rows_appended"):
        writer.record_rows_appended(sheet_name, rows_appended)


def write_dict_rows_sheet(
    writer: Any,
    sheet_name: str,
    columns: list[str],
    rows: Iterable[dict[str, Any] | MainRowPayload | ExtraRowPayload],
) -> None:
    """Write row payloads to sheet; accepts iterators for streaming export."""
    ws = _get_or_create_sheet(writer, sheet_name)
    row_iter = iter(rows)
    try:
        first_row = next(row_iter)
    except StopIteration:
        if columns:
            _append_header_row(ws, writer.book, columns)
            if hasattr(writer, "record_rows_appended"):
                writer.record_rows_appended(sheet_name, 1)
        return

    if not columns:
        columns = list(_row_to_mapping(first_row).keys())
    _append_header_row(ws, writer.book, columns)
    rows_written = 1

    def _append_mapping(row_mapping: dict[str, Any]) -> None:
        nonlocal rows_written
        ws.append([_sanitize_excel_value(row_mapping.get(col)) for col in columns])
        rows_written += 1

    _append_mapping(_row_to_mapping(first_row))
    for row in row_iter:
        _append_mapping(_row_to_mapping(row))

    if hasattr(writer, "record_rows_appended"):
        writer.record_rows_appended(sheet_name, rows_written)


def _sanitize_excel_url(url_value: Any) -> str:
    raw = str(url_value or "").strip()
    if not raw:
        return ""
    raw = "".join(ch for ch in raw if ord(ch) >= 32).replace('"', "").replace("'", "")
    if not raw.startswith(("http://", "https://")):
        return raw
    try:
        parts = urlsplit(raw)
        cleaned_path = quote(unquote(parts.path), safe="/:@-._~!$&()*+,;=")
        cleaned_query = quote(unquote(parts.query), safe="=&:@-._~!$()*+,;/?")
        cleaned_fragment = quote(unquote(parts.fragment), safe=":@-._~!$&()*+,;=/?")
        return urlunsplit(
            (parts.scheme, parts.netloc, cleaned_path, cleaned_query, cleaned_fragment)
        )
    except Exception as exc:
        logger.debug("Could not sanitise Excel URL %r: %s", raw, exc)
        return raw


def _normalize_url_for_match(url_value: Any) -> str:
    return normalize_url_key(_sanitize_excel_url(url_value))


def write_cached_sheet_chunked(
    writer: Any,
    cache: AuditCache,
    sheet_name: str,
    columns: list[str],
    payload_key: str,
    chunk_size: int = 500,
) -> None:
    """Stream cached rows into ``sheet_name`` without holding the full set in RAM.

    Each chunk is materialised once, written to the worksheet, then
    explicitly cleared so deserialised JSON dictionaries become collectable
    immediately. When ``chunk_size`` exceeds ``_GC_COLLECT_CHUNK_THRESHOLD``
    a manual ``gc.collect()`` is triggered after every chunk to keep the
    high-water mark flat on 10k+ page audits.
    """
    ws = _get_or_create_sheet(writer, sheet_name)
    _append_header_row(ws, writer.book, columns)
    rows_written = 1
    aggressive_gc = chunk_size >= _GC_COLLECT_CHUNK_THRESHOLD
    for chunk in cache.iter_results_chunked(chunk_size):
        for result in chunk:
            payload = result.get(payload_key) or {}
            ws.append([_sanitize_excel_value(payload.get(col)) for col in columns])
            rows_written += 1
            if isinstance(payload, dict):
                payload.clear()
            if isinstance(result, dict):
                result.clear()
        chunk.clear()
        del chunk
        if aggressive_gc:
            collected = gc.collect()
            logger.debug(
                "Streamed chunk to %s; gc reclaimed %s objects (rows=%s).",
                sheet_name,
                collected,
                rows_written,
            )
    if hasattr(writer, "record_rows_appended"):
        writer.record_rows_appended(sheet_name, rows_written)


def write_link_inventory_sheet_streamed(
    writer: Any,
    cache: Any,
    *,
    sheet_name: str,
    columns: list[str],
    chunk_size: int = 500,
) -> int:
    """Stream deduped Link Inventory rows from SQLite into the workbook."""
    from hype_frog.core.memory_guard import memory_circuit_breaker

    ws = _get_or_create_sheet(writer, sheet_name)
    _append_header_row(ws, writer.book, columns)
    rows_written = 1
    aggressive_gc = chunk_size >= _GC_COLLECT_CHUNK_THRESHOLD
    for chunk in cache.iter_rows(chunk_size=chunk_size):
        for row in chunk:
            ws.append([_sanitize_excel_value(row.get(col)) for col in columns])
            rows_written += 1
        chunk.clear()
        memory_circuit_breaker()
        if aggressive_gc:
            gc.collect()
    return rows_written


def append_link_detail_rows_streamed(
    writer: Any,
    cache: Any,
    *,
    sheet_name: str,
    columns: list[str],
    status_by_url: dict[str, Any],
    chunk_size: int = 500,
) -> int:
    """Append deduped, streamed Detail rows to an already-written Link Intelligence sheet.

    Folded in from the former standalone "Link Inventory" sheet: same batching /
    ``memory_circuit_breaker()`` / ``gc.collect()`` memory-safety characteristics as
    :func:`write_link_inventory_sheet_streamed`, but targets a sheet whose Summary
    rows (and header) were already written by :func:`write_dict_rows_sheet` — so no
    header row is written here, only data rows. Columns not present on a decorated
    anchor row (the Summary-only columns, e.g. ``Click Depth``) are written blank via
    the same ``row.get(col)`` pattern used elsewhere.
    """
    from hype_frog.core.memory_guard import memory_circuit_breaker
    from hype_frog.pipeline.link_inventory_stream import iter_rows_decorated

    ws = _get_or_create_sheet(writer, sheet_name)
    record_type_idx = columns.index("Record Type")
    rows_written = 0
    aggressive_gc = chunk_size >= _GC_COLLECT_CHUNK_THRESHOLD
    for chunk in iter_rows_decorated(cache, status_by_url, chunk_size=chunk_size):
        for row in chunk:
            values = [_sanitize_excel_value(row.get(col)) for col in columns]
            values[record_type_idx] = "Detail"
            ws.append(values)
            rows_written += 1
        chunk.clear()
        memory_circuit_breaker()
        if aggressive_gc:
            gc.collect()
    return rows_written
