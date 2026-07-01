"""Run small crawl matrix with varied flags and deep-audit workbooks."""

from __future__ import annotations

import asyncio
import json
import re
import sys
import time
from dataclasses import dataclass, field, replace
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from hype_frog.core.path_bootstrap import bootstrap_src_path, repo_root

bootstrap_src_path(anchor=Path(__file__))
ROOT = repo_root()

from hype_frog.config import load_environment  # noqa: E402
from hype_frog.core import configure_logging, get_logger  # noqa: E402
from hype_frog.diagnostics.quick_test import _run_pipeline  # noqa: E402
from hype_frog.core.env_vars import get_psi_api_key  # noqa: E402
from hype_frog.core.run_config import RunConfig  # noqa: E402
from hype_frog.reporter.workbook_audit import audit_workbook, count_main_rows  # noqa: E402

logger = get_logger(__name__)

SITEMAP = "https://africanmarketingconfederation.org/page-sitemap.xml"
OUT_DIR = ROOT / "reports" / "matrix_test"
_FORMULA_ERR = re.compile(r"#(REF!|DIV/0!|NAME\?|NULL!|VALUE!|NUM!|N/A)", re.I)
_SHEET_REF = re.compile(r"'([^']+)'!")

FULL_SUITE_SHEETS = (
    "Table of Contents",
    "Dashboard",
    "Content Optimisation Hub",
    "Content Hub Metrics",
    "Main",
    "Technical Diagnostics",
    "Content & AI Readiness",
    "Issue Register",
    "FixPlan",
    "SitemapQA",
    "Script Inventory",
    "Image Inventory",
    "Snippet Opportunities",
    "Link Equity Map",
)

MAIN_ENRICHMENT_COLUMNS = (
    "Third Party Script Count",
    "Hreflang Declared Languages",
    "Hreflang Code Valid",
    "Broken Image Count",
    "PageRank Percentile",
    "Equity Tier",
    "Featured Snippet Readiness",
    "Top TF-IDF Terms",
    "Technical Health",
    "SEO Health Score",
    "Extraction State",
    "URL",
)


@dataclass(frozen=True)
class Scenario:
    name: str
    config: RunConfig
    expect_psi: bool = False
    expect_image_inventory: bool = False
    expect_script_inventory: bool = False
    expect_competitor_sheet: bool = False


def _base_config(**overrides: Any) -> RunConfig:
    base = dict(
        target_input=SITEMAP,
        max_urls=3,
        max_psi_urls=0,
        high_value_slugs=["about", "contact"],
        crawl_mode="fast",
        render_wait_ms=2000,
        selector_wait_ms=1500,
        workers=3,
        request_delay=0.5,
        full_suite=True,
        previous_audit_path="",
        checkpoint_every=0,
        resume_checkpoint="no",
        check_external_link_status=False,
        check_og_images=False,
        check_content_images=False,
        bfs_max_depth=1,
        gsc_url_inspection=None,
        max_memory_mb=None,
        streaming=False,
        competitor_domains=(),
    )
    base.update(overrides)
    return RunConfig(**base)


def _scenarios() -> list[Scenario]:
    psi_urls = 2 if get_psi_api_key() else 0
    return [
        Scenario(
            name="fast_minimal",
            config=_base_config(crawl_mode="fast", max_psi_urls=0),
        ),
        Scenario(
            name="accurate_psi_images",
            config=_base_config(
                crawl_mode="accurate",
                max_psi_urls=psi_urls,
                check_content_images=True,
                check_og_images=True,
                check_external_link_status=True,
            ),
            expect_psi=psi_urls > 0,
            expect_image_inventory=True,
            expect_script_inventory=psi_urls > 0,
        ),
        Scenario(
            name="accurate_competitors",
            config=_base_config(
                crawl_mode="accurate",
                max_psi_urls=min(1, psi_urls),
                competitor_domains=("ticonafrica.org",),
            ),
            expect_psi=psi_urls > 0,
            expect_competitor_sheet=True,
        ),
        Scenario(
            name="streaming_gsc_limited",
            config=_base_config(
                crawl_mode="accurate",
                max_psi_urls=0,
                streaming=True,
                gsc_url_inspection="limited",
            ),
        ),
    ]


def _header_map(ws, header_row: int = 1) -> dict[str, int]:
    return {
        str(c.value).strip(): c.column
        for c in ws[header_row]
        if c.value is not None and str(c.value).strip()
    }


def _data_rows(ws, header_row: int = 1) -> int:
    return max(0, ws.max_row - header_row)


def _column_nonempty_count(ws, col: int, start_row: int = 2) -> int:
    count = 0
    for r in range(start_row, ws.max_row + 1):
        val = ws.cell(row=r, column=col).value
        if val not in (None, "", False):
            count += 1
    return count


def _scan_formula_refs(wb) -> list[str]:
    issues: list[str] = []
    names = set(wb.sheetnames)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if not isinstance(val, str) or not val.startswith("="):
                    continue
                if _FORMULA_ERR.search(val):
                    issues.append(f"{ws.title}!{cell.coordinate} contains error token in formula")
                # Skip dynamic sheet refs (e.g. INDIRECT("'"&F2&"'!A:A")) — valid at runtime.
                if "INDIRECT(" in val.upper():
                    continue
                for ref in _SHEET_REF.findall(val):
                    sheet_name = ref.replace("''", "'")
                    if sheet_name not in names:
                        issues.append(
                            f"{ws.title}!{cell.coordinate} references missing sheet {sheet_name!r}"
                        )
    return issues


def _audit_main_formulas(wb) -> list[str]:
    issues: list[str] = []
    if "Main" not in wb.sheetnames or "Technical Diagnostics" not in wb.sheetnames:
        return issues
    main = wb["Main"]
    headers = _header_map(main)
    th_col = headers.get("Technical Health")
    url_col = headers.get("URL")
    if not th_col or not url_col:
        return issues
    url_letter = get_column_letter(url_col)
    th_letter = get_column_letter(th_col)
    expected_prefix = (
        f'=IFERROR(VLOOKUP({url_letter}2,'
        "'Technical Diagnostics'!$A:$E,5,FALSE),\"\")"
    ).replace("2", "")
    for r in range(2, min(main.max_row, 5) + 1):
        formula = main.cell(row=r, column=th_col).value
        if not isinstance(formula, str) or not formula.startswith("=IFERROR(VLOOKUP"):
            issues.append(f"Main!{th_letter}{r} missing Technical Health VLOOKUP formula")
            break
        if "'Technical Diagnostics'!$A:$E" not in formula:
            issues.append(f"Main!{th_letter}{r} VLOOKUP range incorrect")
            break
        if f"{url_letter}{r}" not in formula:
            issues.append(f"Main!{th_letter}{r} VLOOKUP URL reference incorrect")
            break
    return issues


def deep_audit(
    path: Path,
    scenario: Scenario,
) -> dict[str, Any]:
    issues: list[str] = []
    stats: dict[str, Any] = {}

    issues.extend(audit_workbook(path, require_full_suite_sheets=True))
    stats["main_rows"] = count_main_rows(path)

    wb = load_workbook(path, data_only=False)
    try:
        issues.extend(_scan_formula_refs(wb))
        issues.extend(_audit_main_formulas(wb))

        for sheet in FULL_SUITE_SHEETS:
            if sheet not in wb.sheetnames:
                issues.append(f"Missing full-suite sheet: {sheet}")
        stats["sheet_rows"] = {
            name: _data_rows(wb[name], header_row=1 if name != "Content Optimisation Hub" else 2)
            for name in wb.sheetnames
            if name in FULL_SUITE_SHEETS
        }

        if "Main" in wb.sheetnames:
            main = wb["Main"]
            headers = _header_map(main)
            missing_cols = [c for c in MAIN_ENRICHMENT_COLUMNS if c not in headers]
            if missing_cols:
                issues.append(f"Main missing columns: {missing_cols}")
            stats["main_enrichment_populated"] = {
                col: _column_nonempty_count(main, headers[col])
                for col in MAIN_ENRICHMENT_COLUMNS
                if col in headers
            }

        if scenario.expect_script_inventory and "Script Inventory" in wb.sheetnames:
            rows = _data_rows(wb["Script Inventory"])
            stats["script_inventory_rows"] = rows
            if rows == 0:
                issues.append("Script Inventory empty but PSI was enabled")

        if scenario.expect_image_inventory and "Image Inventory" in wb.sheetnames:
            rows = _data_rows(wb["Image Inventory"])
            stats["image_inventory_rows"] = rows
            if rows == 0:
                issues.append("Image Inventory empty but --check-images was enabled")

        if scenario.expect_competitor_sheet:
            if "Competitor Benchmarks" not in wb.sheetnames:
                issues.append("Competitor Benchmarks sheet missing")
            elif _data_rows(wb["Competitor Benchmarks"]) == 0:
                issues.append("Competitor Benchmarks has no data rows")

        if "Dashboard" in wb.sheetnames:
            dash = wb["Dashboard"]
            hyperlink_cells = [
                c.coordinate
                for row in dash.iter_rows(min_row=1, max_row=15)
                for c in row
                if isinstance(c.value, str) and c.value.startswith("=HYPERLINK")
            ]
            stats["dashboard_hyperlinks"] = len(hyperlink_cells)
            if not hyperlink_cells:
                issues.append("Dashboard has no HYPERLINK formulas in top rows")

        if "Technical Diagnostics" in wb.sheetnames:
            td = wb["Technical Diagnostics"]
            td_headers = _header_map(td)
            for col_name in (
                "Hreflang Declared Languages",
                "Hreflang Reciprocal Status",
                "SEO Health Score",
            ):
                if col_name not in td_headers:
                    issues.append(f"Technical Diagnostics missing column {col_name}")
    finally:
        wb.close()

    return {
        "scenario": scenario.name,
        "path": str(path),
        "issues": issues,
        "stats": stats,
        "ok": not issues,
    }


async def _run_scenario(scenario: Scenario) -> dict[str, Any]:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / f"matrix_{scenario.name}.xlsx"
    if out_path.exists():
        out_path.unlink()

    config = replace(scenario.config, output_filename=str(out_path))

    started = time.perf_counter()
    logger.info("=== Scenario %s ===", scenario.name)
    try:
        crawl_result = await _run_pipeline(config)
    except Exception as exc:
        return {
            "scenario": scenario.name,
            "path": str(out_path),
            "issues": [f"Pipeline failed: {type(exc).__name__}: {exc}"],
            "stats": {},
            "ok": False,
            "elapsed_s": round(time.perf_counter() - started, 1),
        }
    elapsed = time.perf_counter() - started
    report = deep_audit(Path(crawl_result.output_filename), scenario)
    report["elapsed_s"] = round(elapsed, 1)
    report["urls_crawled"] = len(crawl_result.crawl_rows)
    return report


async def main() -> int:
    configure_logging()
    load_environment()
    reports: list[dict[str, Any]] = []
    for scenario in _scenarios():
        reports.append(await _run_scenario(scenario))

    summary_path = OUT_DIR / "matrix_audit_summary.json"
    summary_path.write_text(json.dumps(reports, indent=2), encoding="utf-8")

    print("\n" + "=" * 72)
    print(" CRAWL MATRIX AUDIT SUMMARY")
    print("=" * 72)
    failed = 0
    for rep in reports:
        status = "PASS" if rep["ok"] else "FAIL"
        if not rep["ok"]:
            failed += 1
        print(f"\n[{status}] {rep['scenario']} ({rep.get('elapsed_s', '?')}s)")
        print(f"  Workbook: {rep['path']}")
        print(f"  URLs: {rep.get('urls_crawled', '?')} | Main rows: {rep.get('stats', {}).get('main_rows', '?')}")
        if rep.get("stats", {}).get("sheet_rows"):
            sparse = {
                k: v
                for k, v in rep["stats"]["sheet_rows"].items()
                if v == 0 and k not in {"Snippet Opportunities", "Competitor Benchmarks"}
            }
            if sparse:
                print(f"  Empty sheets: {sparse}")
        if rep.get("stats", {}).get("main_enrichment_populated"):
            pop = rep["stats"]["main_enrichment_populated"]
            print(f"  Main enrichment populated rows: {pop}")
        for issue in rep.get("issues", []):
            print(f"  - {issue}")
    print("\n" + "=" * 72)
    print(f" Overall: {'FAIL' if failed else 'PASS'} ({len(reports) - failed}/{len(reports)} scenarios)")
    print(f" JSON: {summary_path}")
    print("=" * 72 + "\n")
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(asyncio.run(main()))
