"""Phase 0 empirical workbook inspection (throwaway script)."""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from hype_frog.reporter.sheets.config import RETURN_TO_BRIEFING_LABEL
from hype_frog.reporter.sheets.sheet_rows import sheet_data_header_row

# Heuristic: row is a header if it has >=3 non-empty string cells and no row above
# looks more header-like (unless row 1 is return strip / title block).
_KNOWN_HEADER_MARKERS = frozenset(
    {
        "URL",
        "Issue Type",
        "Source URL",
        "Target URL",
        "Metric",
        "Key",
        "Section",
        "Tab",
        "Sheet",
        "Rule ID",
        "Playbook ID",
        "Status",
        "Health Icon",
        "Priority Score",
        "Field",
        "Value",
    }
)


def _cell_text(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _row_values(ws, row: int, max_col: int | None = None) -> list[str]:
    mc = max_col or ws.max_column
    return [_cell_text(ws.cell(row=row, column=c).value) for c in range(1, mc + 1)]


def _score_header_row(ws, row: int) -> float:
    vals = _row_values(ws, row)
    non_empty = [v for v in vals if v]
    if len(non_empty) < 2:
        return 0.0
    score = len(non_empty)
    for v in non_empty:
        if v in _KNOWN_HEADER_MARKERS:
            score += 5
        if len(v) <= 60 and not v.startswith("="):
            score += 0.5
        if v.startswith("←") or "Return to" in v:
            score -= 10
        if v.startswith("http"):
            score -= 3
    return score


def detect_header_row(ws, sheet_name: str) -> tuple[int, list[str]]:
    """Empirical header detection with code-contract cross-check."""
    code_row = sheet_data_header_row(sheet_name)
    # Scan first 15 rows
    best_row, best_score = 1, -1.0
    limit = min(15, ws.max_row or 1)
    for r in range(1, limit + 1):
        sc = _score_header_row(ws, r)
        if sc > best_score:
            best_score = sc
            best_row = r
    # Return strip: row 1 nav, headers often row 2
    if _cell_text(ws.cell(row=1, column=1).value) == RETURN_TO_BRIEFING_LABEL:
        if best_row == 1:
            best_row = 2
    headers = [h for h in _row_values(ws, best_row) if h]
    return best_row, headers


def inspect_workbook(path: Path) -> dict:
    wb = load_workbook(path, data_only=False)
    out: dict = {
        "path": str(path),
        "sheets": [],
        "named_ranges": [],
        "totals": {
            "merged_ranges": 0,
            "hyperlinks": 0,
            "charts": 0,
            "formulas": 0,
            "freeze_panes": 0,
        },
    }
    try:
        if wb.defined_names:
            for name, dn in wb.defined_names.items():
                out["named_ranges"].append(
                    {"name": name, "attr_text": getattr(dn, "attr_text", str(dn))}
                )
        for name in wb.sheetnames:
            ws = wb[name]
            state = getattr(ws, "sheet_state", "visible")
            used = f"{get_column_letter(ws.max_column or 1)}{ws.max_row or 1}"
            hdr_row, headers = detect_header_row(ws, name)
            merges = [str(r) for r in ws.merged_cells.ranges]
            links: list[dict] = []
            formulas: list[dict] = []
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.hyperlink and cell.hyperlink.target:
                        links.append(
                            {
                                "cell": cell.coordinate,
                                "target": cell.hyperlink.target,
                                "display": _cell_text(cell.value)[:80],
                            }
                        )
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formulas.append({"cell": cell.coordinate, "formula": cell.value[:120]})
            charts = []
            if hasattr(ws, "_charts") and ws._charts:
                for i, ch in enumerate(ws._charts):
                    anchor = getattr(ch, "anchor", None)
                    charts.append({"index": i, "type": type(ch).__name__, "anchor": str(anchor)})
            freeze = ws.freeze_panes
            sheet_info = {
                "name": name,
                "state": state,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "used_range": f"A1:{used}",
                "detected_header_row": hdr_row,
                "code_expected_header_row": sheet_data_header_row(name),
                "headers": headers[:40],
                "header_count": len(headers),
                "freeze_panes": freeze,
                "merged_count": len(merges),
                "merged_ranges_sample": merges[:15],
                "hyperlink_count": len(links),
                "hyperlinks_sample": links[:12],
                "chart_count": len(charts),
                "charts": charts,
                "formula_count": len(formulas),
                "formulas_sample": formulas[:8],
            }
            out["sheets"].append(sheet_info)
            out["totals"]["merged_ranges"] += len(merges)
            out["totals"]["hyperlinks"] += len(links)
            out["totals"]["charts"] += len(charts)
            out["totals"]["formulas"] += len(formulas)
            if freeze:
                out["totals"]["freeze_panes"] += 1
    finally:
        wb.close()
    return out


def main() -> None:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else None
    if path is None:
        latest = sorted(
            Path("reports/latest").glob("SEO_AEO_Audit_*.xlsx"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        if not latest:
            print("No workbook found", file=sys.stderr)
            sys.exit(1)
        path = latest[0]
    result = inspect_workbook(path)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
