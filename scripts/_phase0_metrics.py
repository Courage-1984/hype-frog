"""Extract AEO/content metric distributions from workbook Main + Content & AI Readiness."""
from __future__ import annotations

import json
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

from hype_frog.reporter.sheets.style_helpers import header_row_index


def col_values(ws, header: str, hdr_row: int) -> list:
    headers = header_row_index(ws, hdr_row)
    col = headers.get(header)
    if col is None:
        return []
    vals = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v is not None and str(v).strip() != "":
            vals.append(v)
    return vals


def summarize(vals: list) -> dict:
    if not vals:
        return {"count": 0, "unique": 0, "degenerate": "empty"}
    nums = []
    for v in vals:
        try:
            nums.append(float(v))
        except (TypeError, ValueError):
            pass
    c = Counter(vals)
    deg = None
    if len(c) == 1:
        deg = f"all identical: {next(iter(c))}"
    elif nums and len(set(nums)) == 1:
        deg = f"all numeric identical: {nums[0]}"
    elif nums and sum(1 for n in nums if n == 0) == len(nums):
        deg = "all zero"
    elif nums and sum(1 for n in nums if n == 100) == len(nums):
        deg = "all 100"
    return {
        "count": len(vals),
        "unique": len(c),
        "sample": vals[:5],
        "degenerate": deg,
        "distribution": dict(c.most_common(8)) if len(c) <= 8 else {"top": c.most_common(8)},
    }


def main() -> None:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else sorted(
        Path("reports/latest").glob("SEO_AEO_Audit_*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )[0]
    metrics = [
        "AEO Readiness Score",
        "Question Heading Count",
        "Image Alt Coverage (%)",
        "Schema Types Count",
        "SEO Health Score",
        "Technical Health",
        "Answer Paragraph Count",
        "Citation Block Count",
        "Semantic AEO Score",
        "Extraction Source",
        "Extraction State",
        "H1 Count",
        "Generic Anchor Text Count",
        "Robots.txt: GPTBot",
        "AI Crawlers Allowed (GPTBot/ClaudeBot/PerplexityBot)",
    ]
    wb = load_workbook(path, read_only=True, data_only=True)
    out = {"path": str(path)}
    try:
        for sheet, hdr_row in [("Main", 2), ("Content & AI Readiness", 2)]:
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            out[sheet] = {}
            for m in metrics:
                out[sheet][m] = summarize(col_values(ws, m, hdr_row))
    finally:
        wb.close()
    print(json.dumps(out, indent=2))

if __name__ == "__main__":
    main()
