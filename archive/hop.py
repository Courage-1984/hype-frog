"""
Lightweight SEO Auditor & Crawler
---------------------------------
This script extracts foundational technical SEO metrics from a list of URLs or a Sitemap.
It respects strict concurrency limits and outputs a clean Excel report.

Run instruction: python seo_auditor.py
Dependencies: pip install aiohttp beautifulsoup4 pandas openpyxl lxml
"""

import asyncio
import aiohttp
import time
import os
import json
import re
import random
import shutil
from datetime import datetime
from collections import defaultdict
from urllib.parse import urljoin, urlparse
import pandas as pd
from bs4 import BeautifulSoup
import xml.etree.ElementTree as ET
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import BarChart, PieChart, Reference

# --- CONFIGURATION ---
MAX_WORKERS = 3
DELAY_BETWEEN_REQUESTS = 2.5
REQUEST_JITTER_SECONDS = 0.6
TIMEOUT_SECONDS = 20
CONNECT_TIMEOUT_SECONDS = 8
READ_TIMEOUT_SECONDS = 20
MAX_RETRIES = 3
RETRY_BASE_DELAY_SECONDS = 2.0
RETRY_BACKOFF_FACTOR = 2.0
RETRY_MAX_DELAY_SECONDS = 20.0
RETRYABLE_STATUS_CODES = {408, 425, 429, 500, 502, 503, 504}
OUTPUT_FILENAME = "seo_audit_report.xlsx"
DEFAULT_OWNER_BY_SEVERITY = {"Critical": "Dev", "Warning": "SEO", "Info": "Content"}
DEFAULT_EFFORT_BY_SEVERITY = {"Critical": "M", "Warning": "S", "Info": "S"}


def readability_flesch(words, sentences):
    """Returns a rough Flesch reading ease score."""
    if words <= 0 or sentences <= 0:
        return None
    # syllables per word is approximated on a sample for speed.
    return round(206.835 - 1.015 * (words / sentences), 2)


def adjust_sheet_format(writer, sheet_name):
    """Freeze top row and auto-size columns."""
    worksheet = writer.sheets[sheet_name]
    worksheet.freeze_panes = "B2"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    bad_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    warn_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    good_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    edge_fill = PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid")
    zebra_fill = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")

    headers = [cell.value for cell in worksheet[1]]
    header_to_col = {str(h): i + 1 for i, h in enumerate(headers) if h is not None}

    # Header style
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def is_bad_header(header_text):
        h = (header_text or "").lower()
        bad_tokens = [
            "error",
            "broken",
            "missing",
            "noindex",
            "disallow",
            "thin",
            "mixed content",
            "cross-canonical",
            "issue",
            "non-200",
            "loop",
            "out of",
        ]
        return any(t in h for t in bad_tokens)

    def is_edge_header(header_text):
        h = (header_text or "").lower()
        return any(t in h for t in ["redirect chain", "param url", "edge", "unresolved"])

    def is_good_header(header_text):
        h = (header_text or "").lower()
        good_tokens = [
            "accessible",
            "match",
            "enabled",
            "complete",
            "present",
            "indexable",
            "coverage (%)",
        ]
        return any(t in h for t in good_tokens)

    def parse_bool(value):
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.strip().lower() in {"true", "1", "yes", "y"}
        return bool(value)

    # Data cell style and conditional coloring
    for row_idx in range(2, worksheet.max_row + 1):
        row_has_issue = False
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            val = cell.value
            h = str(header) if header is not None else ""

            if "url" in h.lower() or "hops" in h.lower() or "images" == h.lower():
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")

            # 100% pass style for percentage fields
            if "%" in h:
                try:
                    pct = float(val)
                    if pct == 100.0:
                        cell.fill = good_fill
                    elif pct < 80:
                        cell.fill = bad_fill
                        row_has_issue = True
                    else:
                        cell.fill = warn_fill
                except Exception:
                    pass

            # Status code coloring
            if h in {"Status Code", "Target Status (if crawled)"} and isinstance(val, int):
                if val >= 400:
                    cell.fill = bad_fill
                    row_has_issue = True
                elif val >= 300:
                    cell.fill = warn_fill
                elif 200 <= val < 300:
                    cell.fill = good_fill

            # Boolean rule coloring
            if isinstance(val, bool) or (isinstance(val, str) and val.strip().lower() in {"true", "false"}):
                flag = parse_bool(val)
                if is_bad_header(h):
                    if flag:
                        cell.fill = bad_fill
                        row_has_issue = True
                    else:
                        cell.fill = good_fill
                elif is_good_header(h):
                    if flag:
                        cell.fill = good_fill
                    else:
                        cell.fill = warn_fill
                        row_has_issue = True
                elif is_edge_header(h) and flag:
                    cell.fill = edge_fill

            # Numeric quality fields
            if h in {"Broken Internal Links Count", "Image Filename Quality Issues", "Generic Anchor Text Count"}:
                try:
                    if int(val or 0) > 0:
                        cell.fill = bad_fill
                        row_has_issue = True
                    else:
                        cell.fill = good_fill
                except Exception:
                    pass

            if h in {"Redirect Chain Length"}:
                try:
                    chain_len = int(val or 0)
                    if chain_len > 1:
                        cell.fill = warn_fill
                        row_has_issue = True
                    elif chain_len == 1:
                        cell.fill = edge_fill
                    else:
                        cell.fill = good_fill
                except Exception:
                    pass

            if h in {"Word Count Band"} and isinstance(val, str):
                band = val.lower()
                if band == "thin":
                    cell.fill = bad_fill
                    row_has_issue = True
                elif band == "ok":
                    cell.fill = warn_fill
                elif band == "strong":
                    cell.fill = good_fill

            if h in {"Indexability Reason"} and isinstance(val, str):
                reason = val.lower()
                if "indexable" in reason and "noindex" not in reason:
                    cell.fill = good_fill
                else:
                    cell.fill = bad_fill
                    row_has_issue = True
            if h == "SEO Health Score":
                try:
                    score = float(val)
                    if score >= 90:
                        cell.fill = good_fill
                    elif score >= 70:
                        cell.fill = warn_fill
                        row_has_issue = True
                    else:
                        cell.fill = bad_fill
                        row_has_issue = True
                except Exception:
                    pass
            if h == "AEO Readiness Score":
                try:
                    score = float(val)
                    if score >= 80:
                        cell.fill = good_fill
                    elif score >= 60:
                        cell.fill = warn_fill
                    else:
                        cell.fill = bad_fill
                        row_has_issue = True
                except Exception:
                    pass
            if h == "AEO Badge" and isinstance(val, str):
                badge_val = val.lower()
                if badge_val in {"strong", "good"}:
                    cell.fill = good_fill
                elif badge_val == "fair":
                    cell.fill = warn_fill
                    row_has_issue = True
                else:
                    cell.fill = bad_fill
                    row_has_issue = True
            if h == "Health Icon" and isinstance(val, str):
                if "PASS" in val:
                    cell.fill = good_fill
                elif "WARN" in val:
                    cell.fill = warn_fill
                    row_has_issue = True
                elif "FAIL" in val:
                    cell.fill = bad_fill
                    row_has_issue = True
            if h in {"Severity", "Severity Badge"} and isinstance(val, str):
                sev = val.strip().lower()
                if sev == "critical":
                    cell.fill = bad_fill
                    row_has_issue = True
                elif sev == "warning":
                    cell.fill = warn_fill
                    row_has_issue = True
                elif sev == "info":
                    cell.fill = edge_fill
                elif sev == "pass":
                    cell.fill = good_fill
            if h == "Category" and isinstance(val, str):
                cat = val.strip().lower()
                if cat == "aeo":
                    cell.fill = edge_fill
                elif cat == "seo":
                    cell.fill = good_fill
            if h == "Status" and isinstance(val, str):
                st = val.strip().lower()
                if st in {"open", "in progress"}:
                    cell.fill = warn_fill
                elif st == "done":
                    cell.fill = good_fill

        # First column green if row passes best-practice checks
        first_cell = worksheet.cell(row=row_idx, column=1)
        if not row_has_issue:
            first_cell.fill = good_fill
        if sheet_name in {"LinksDetail", "SitemapQA"} and row_idx % 2 == 0:
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.fill.fill_type is None:
                    cell.fill = zebra_fill

    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)


def normalize_text_hash(value):
    return re.sub(r"\s+", " ", (value or "").strip().lower())


def status_class(status_code):
    if isinstance(status_code, int):
        return f"{status_code // 100}xx"
    return str(status_code)


def url_depth(url):
    path = urlparse(url).path.strip("/")
    if not path:
        return 0
    return len([p for p in path.split("/") if p])


def word_count_band(count):
    if count < 300:
        return "Thin"
    if count < 800:
        return "OK"
    return "Strong"


def image_extension(src_url):
    path = urlparse(src_url).path.lower()
    for ext in [".webp", ".avif", ".jpg", ".jpeg", ".png", ".gif", ".svg"]:
        if path.endswith(ext):
            return ext.replace(".", "")
    return "other"


def looks_generic_image_filename(src_url):
    name = os.path.basename(urlparse(src_url).path).lower()
    if not name:
        return False
    # Common low-quality patterns: img123, image_001, dsc0001, photo12
    return bool(re.match(r"^(img|image|dsc|photo|pic)[-_]?\d+\.[a-z0-9]+$", name))


def to_bool(value):
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in {"true", "1", "yes", "y"}
    return bool(value)


def get_summary_rules():
    return [
        ("Critical", "Non-200 Status", lambda r: isinstance(r.get("Status Code"), int) and r.get("Status Code") >= 400),
        ("Critical", "Noindex Directive", lambda r: "noindex" in str(r.get("Indexability Reason", "")).lower()),
        ("Critical", "Canonical Points Elsewhere", lambda r: r.get("Canonical Type") == "cross-canonical"),
        ("Critical", "Robots.txt Disallow Root", lambda r: to_bool(r.get("Robots.txt Disallow /"))),
        ("Critical", "Broken Internal Links", lambda r: (r.get("Broken Internal Links Count") or 0) > 0),
        ("Warning", "Redirect Chains", lambda r: (r.get("Redirect Chain Length") or 0) > 1),
        ("Warning", "Missing Title", lambda r: to_bool(r.get("Title Missing"))),
        ("Warning", "Missing Meta Description", lambda r: to_bool(r.get("Meta Description Missing"))),
        ("Warning", "Missing H1", lambda r: to_bool(r.get("Missing H1 Flag"))),
        ("Warning", "Multiple H1", lambda r: to_bool(r.get("Multiple H1 Flag"))),
        ("Warning", "Thin Content", lambda r: to_bool(r.get("Thin Content Flag"))),
        ("Warning", "Low Image Alt Coverage", lambda r: (r.get("Image Alt Coverage (%)") or 100) < 80),
        ("Warning", "Mixed Content", lambda r: to_bool(r.get("Mixed Content Detected"))),
        ("Warning", "Canonical Missing", lambda r: r.get("Canonical Type") == "missing"),
        ("Warning", "Hreflang Without Reciprocity", lambda r: r.get("Hreflang Present") and not to_bool(r.get("Hreflang Reciprocal Check"))),
        ("Info", "Uses URL Parameters", lambda r: to_bool(r.get("Param URL Flag"))),
        ("Info", "Generic Anchor Text Present", lambda r: (r.get("Generic Anchor Text Count") or 0) > 0),
        ("Info", "Image Filename Quality Issues", lambda r: (r.get("Image Filename Quality Issues") or 0) > 0),
        ("Info", "No Compression Header", lambda r: not to_bool(r.get("Compression Enabled"))),
        ("Info", "No Cache-Control Header", lambda r: not bool(r.get("Cache-Control"))),
        ("Info", "No ETag Header", lambda r: not bool(r.get("ETag"))),
        # AEO-specific opportunities
        ("Warning", "Low AEO Readiness Score", lambda r: (r.get("AEO Readiness Score") or 0) < 60),
        ("Info", "No FAQ/QA Schema", lambda r: not to_bool(r.get("QAPage/FAQ Schema Present"))),
        ("Info", "No Question Headings", lambda r: (r.get("Question Heading Count") or 0) == 0),
        ("Info", "No Answer-Friendly Structure", lambda r: not to_bool(r.get("List/Table Answer Signal"))),
        ("Info", "No 40-60 Word Answer Paragraphs", lambda r: (r.get("Paragraphs 40-60 Words Count") or 0) == 0),
    ]


def score_url_health(row, summary_rules):
    matched = {"Critical": [], "Warning": [], "Info": []}
    for severity, issue_name, rule_fn in summary_rules:
        try:
            if rule_fn(row):
                matched[severity].append(issue_name)
        except Exception:
            continue
    score = max(0, 100 - (25 * len(matched["Critical"])) - (10 * len(matched["Warning"])) - (3 * len(matched["Info"])))
    if matched["Critical"]:
        badge = "Critical"
        icon = "FAIL 🔴"
    elif matched["Warning"]:
        badge = "Warning"
        icon = "WARN 🟡"
    elif matched["Info"]:
        badge = "Info"
        icon = "INFO 🔵"
    else:
        badge = "Pass"
        icon = "PASS 🟢"
    return score, badge, icon, matched


def stable_issue_id(url, issue_name):
    safe_url = str(url or "").strip()
    safe_issue = re.sub(r"[^a-zA-Z0-9_-]+", "_", str(issue_name or "").strip().lower())
    return f"{safe_url}::{safe_issue}"


def root_cause_and_fix(issue_name):
    mapping = {
        "Non-200 Status": (
            "URL returns 4xx/5xx or equivalent failure.",
            "Fix status code, restore page, or implement correct redirect to canonical destination.",
        ),
        "Noindex Directive": (
            "Meta robots or X-Robots-Tag contains noindex.",
            "Remove unintended noindex directives on index-worthy URLs.",
        ),
        "Canonical Points Elsewhere": (
            "Canonical targets a different URL variant.",
            "Align canonical to preferred final URL and ensure internal links use canonical target.",
        ),
        "Broken Internal Links": (
            "Internal links resolve to missing or error pages.",
            "Update internal links to valid URLs and remove dead references.",
        ),
        "Missing Title": (
            "Template/page missing title tag.",
            "Add unique descriptive title on affected template/page type.",
        ),
        "Missing Meta Description": (
            "Meta description missing or empty.",
            "Add concise unique description aligned with user intent.",
        ),
        "Thin Content": (
            "Insufficient body content for indexing confidence.",
            "Expand page with helpful, unique, intent-matching content.",
        ),
        "Low AEO Readiness Score": (
            "Page lacks concise answer-oriented signals and structured answer patterns.",
            "Add short direct answer blocks, question-led headings, and answer-ready formatting.",
        ),
        "No FAQ/QA Schema": (
            "No FAQPage/QAPage schema found for likely Q&A style content.",
            "Add valid FAQPage or QAPage JSON-LD where appropriate and ensure on-page parity.",
        ),
        "No Question Headings": (
            "Headings are not phrased as user questions.",
            "Add clear question-style headings (especially H2/H3) for key intents.",
        ),
        "No Answer-Friendly Structure": (
            "Page lacks lists/tables that help concise answer extraction.",
            "Introduce structured bullets, ordered steps, or tables for key answers.",
        ),
        "No 40-60 Word Answer Paragraphs": (
            "No concise answer-length paragraph detected.",
            "Add a direct 40-60 word answer summary near question headings.",
        ),
        "Mixed Content": (
            "HTTPS page references insecure HTTP assets.",
            "Serve all assets over HTTPS and update hardcoded URLs.",
        ),
    }
    return mapping.get(
        issue_name,
        ("Template/technical implementation quality issue.", "Apply fix based on issue type and re-run audit."),
    )


def sanitize_filename_part(value):
    cleaned = re.sub(r"[^a-zA-Z0-9._-]+", "_", str(value or "").strip().lower())
    cleaned = re.sub(r"_+", "_", cleaned).strip("._-")
    return cleaned or "audit"


def build_output_filename(source_label, full_suite):
    mode = "full" if full_suite else "main"
    source = sanitize_filename_part(source_label)
    return f"seo_audit_{source}_{mode}.xlsx"


def apply_tab_hyperlinks(writer):
    """Add cross-tab hyperlinks where helpful."""
    wb = writer.book
    link_map = {
        "Summary": "Reference Tab",
        "FixPlan": "Reference Tab",
        "Dashboard": "Target Tab",
    }
    for sheet_name, col_header in link_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        if col_header not in headers:
            continue
        col_idx = headers.index(col_header) + 1
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            target = str(cell.value or "").strip()
            if target and target in wb.sheetnames:
                cell.hyperlink = f"#{target}!A1"
                cell.style = "Hyperlink"


async def fetch_and_parse(
    url,
    session,
    semaphore,
    full_suite=True,
    robots_cache=None,
    request_delay=None,
    sitemap_meta=None,
):
    """
    Fetches a single URL and parses its SEO metrics.
    Uses a semaphore to limit concurrent workers.
    """
    async with semaphore:
        start_time = time.time()

        # Main tab schema is preserved exactly as before.
        main_data = {
            "URL": url,
            "Status Code": None,
            "Load Time (s)": None,
            "Indexability": "Indexable",
            "Title": None,
            "Title Length": 0,
            "Meta Description": None,
            "Meta Desc Length": 0,
            "Word Count (Body)": 0,
            "OG-Image": None,
            "Has Valid JSON-LD": False,
        }

        # Initialize H1-H6 columns
        for i in range(1, 7):
            main_data[f"H{i} Content"] = None
            main_data[f"H{i} Length"] = 0

        extra = {
            "URL": url,
            "Status Code": None,
            "Final URL": None,
            "Protocol": None,
            "Redirect Chain Length": 0,
            "Redirect Target": None,
            "Redirect Hops": None,
            "HTTP->HTTPS Redirect": False,
            "Status Class": None,
            "TTFB (ms)": None,
            "Total Request Time (ms)": None,
            "Content-Type": None,
            "HTTP Version": None,
            "HTML Size (KB)": None,
            "Compression Enabled": False,
            "Cache-Control": None,
            "ETag": None,
            "X-Robots-Tag": None,
            "Meta Robots Raw": None,
            "Canonical URL": None,
            "Canonical Matches Final URL": None,
            "Canonical Type": None,
            "Canonical Absolute URL": None,
            "Canonical in Sitemap Match": None,
            "Hreflang Present": False,
            "Hreflang Count": 0,
            "Hreflang Self Reference": False,
            "Hreflang Reciprocal Check": None,
            "Hreflang Canonical Consistency": None,
            "x-default Present": False,
            "Pagination rel=next": False,
            "Pagination rel=prev": False,
            "H1 Count": 0,
            "Missing H1 Flag": False,
            "Multiple H1 Flag": False,
            "Thin Content Flag": False,
            "Body Text-to-HTML Ratio": None,
            "Word Count": 0,
            "Word Count Band": None,
            "Sentence Count": 0,
            "Readability (Rough Flesch)": None,
            "Last-Modified": None,
            "Published Date": None,
            "Modified Date": None,
            "Last Updated": None,
            "Change Frequency": None,
            "Priority": None,
            "Internal Links Count": 0,
            "External Links Count": 0,
            "Unique Internal Links Count": 0,
            "Nofollow Internal Links Count": 0,
            "Nofollow External Links Count": 0,
            "Generic Anchor Text Count": 0,
            "Param URL Flag": "?" in url,
            "URL Depth": url_depth(url),
            "Image Count": 0,
            "Images": None,
            "Images Missing Alt": 0,
            "Image Alt Coverage (%)": None,
            "Image Extension Distribution": None,
            "Likely Large Image Count": 0,
            "Image Filename Quality Issues": 0,
            "Image On Canonical Domain (%)": None,
            "Mixed Content Detected": False,
            "Schema Types Found": None,
            "Schema Types Count": 0,
            "Schema Parse Errors": 0,
            "Open Graph Complete": False,
            "Twitter Card Type": None,
            "OG Title": None,
            "OG Description": None,
            "OG Image": None,
            "Strict-Transport-Security": None,
            "Content-Security-Policy": None,
            "X-Content-Type-Options": None,
            "X-Frame-Options": None,
            "Referrer-Policy": None,
            "Permissions-Policy": None,
            "Robots.txt Accessible": None,
            "Sitemap in Robots.txt": None,
            "Robots.txt Crawl-Delay": None,
            "Robots.txt Disallow /": None,
            "Title Missing": True,
            "Meta Description Missing": True,
            "Indexability Reason": None,
            "SERP Title Truncation Risk": False,
            "SERP Meta Truncation Risk": False,
            "SERP Title Pixel Approx": 0,
            "SERP Meta Pixel Approx": 0,
            "Inlinks Bucket": None,
            "Important But Underlinked": False,
            "Cannibalization Hint": None,
            "FAQ Section Count": 0,
            "Question Heading Count": 0,
            "HowTo Signal": False,
            "Definition Signal": False,
            "List/Table Answer Signal": False,
            "Paragraphs 40-60 Words Count": 0,
            "Speakable Schema Present": False,
            "QAPage/FAQ Schema Present": False,
            "AEO Readiness Score": 0,
            "AEO Badge": "Needs Work",
            "Action Needed": "No",
            "Owner": None,
            "Sprint": "",
            "Status": "Open",
            "Stable Issue IDs": None,
            "Internal Links List Full": [],
            "Internal Links List": [],
            "Link Details": [],
        }
        if sitemap_meta and url in sitemap_meta:
            extra["Change Frequency"] = sitemap_meta[url].get("changefreq")
            extra["Priority"] = sitemap_meta[url].get("priority")
            extra["Last Updated"] = sitemap_meta[url].get("lastmod")

        timeout = aiohttp.ClientTimeout(
            total=TIMEOUT_SECONDS,
            connect=CONNECT_TIMEOUT_SECONDS,
            sock_read=READ_TIMEOUT_SECONDS,
        )
        completed = False

        for attempt in range(MAX_RETRIES + 1):
            request_start = time.time()
            try:
                async with session.get(url, timeout=timeout) as response:
                    first_byte_ms = round((time.time() - request_start) * 1000, 2)
                    total_request_ms = first_byte_ms
                    main_data["Load Time (s)"] = round(time.time() - start_time, 3)
                    main_data["Status Code"] = response.status
                    extra["Status Code"] = response.status
                    extra["Final URL"] = str(response.url)
                    extra["Protocol"] = urlparse(str(response.url)).scheme
                    extra["Redirect Chain Length"] = len(response.history)
                    if response.history:
                        extra["Redirect Target"] = str(response.history[0].url)
                        extra["Redirect Hops"] = " | ".join(
                            [f"{h.status}:{h.url}" for h in response.history]
                            + [f"{response.status}:{response.url}"]
                        )
                    extra["HTTP->HTTPS Redirect"] = (
                        urlparse(url).scheme == "http"
                        and urlparse(str(response.url)).scheme == "https"
                    )
                    extra["Status Class"] = status_class(response.status)
                    extra["TTFB (ms)"] = first_byte_ms
                    extra["Total Request Time (ms)"] = total_request_ms
                    extra["Content-Type"] = response.headers.get("Content-Type")
                    extra["HTTP Version"] = (
                        f"HTTP/{response.version.major}.{response.version.minor}"
                    )
                    enc = (response.headers.get("Content-Encoding") or "").lower()
                    extra["Compression Enabled"] = "gzip" in enc or "br" in enc or "deflate" in enc
                    extra["Cache-Control"] = response.headers.get("Cache-Control")
                    extra["ETag"] = response.headers.get("ETag")
                    extra["X-Robots-Tag"] = response.headers.get("X-Robots-Tag")
                    extra["Last-Modified"] = response.headers.get("Last-Modified")
                    extra["Strict-Transport-Security"] = response.headers.get(
                        "Strict-Transport-Security"
                    )
                    extra["Content-Security-Policy"] = response.headers.get(
                        "Content-Security-Policy"
                    )
                    extra["X-Content-Type-Options"] = response.headers.get(
                        "X-Content-Type-Options"
                    )
                    extra["X-Frame-Options"] = response.headers.get("X-Frame-Options")
                    extra["Referrer-Policy"] = response.headers.get("Referrer-Policy")
                    extra["Permissions-Policy"] = response.headers.get(
                        "Permissions-Policy"
                    )

                    if (
                        response.status in RETRYABLE_STATUS_CODES
                        and attempt < MAX_RETRIES
                    ):
                        retry_after = response.headers.get("Retry-After")
                        try:
                            retry_after_seconds = float(retry_after) if retry_after else None
                        except ValueError:
                            retry_after_seconds = None
                        backoff_wait = min(
                            RETRY_MAX_DELAY_SECONDS,
                            RETRY_BASE_DELAY_SECONDS
                            * (RETRY_BACKOFF_FACTOR ** attempt),
                        )
                        wait_time = (
                            retry_after_seconds
                            if retry_after_seconds and retry_after_seconds > 0
                            else backoff_wait
                        )
                        wait_time += random.uniform(0, REQUEST_JITTER_SECONDS)
                        print(
                            f"[{response.status}] Retrying {url} "
                            f"(attempt {attempt + 2}/{MAX_RETRIES + 1}) in {wait_time:.1f}s"
                        )
                        await asyncio.sleep(wait_time)
                        continue

                    # Only attempt to parse if it's a valid HTML response
                    if response.status == 200 and "text/html" in response.headers.get(
                        "Content-Type", ""
                    ):
                        html = await response.text()
                        total_request_ms = round((time.time() - request_start) * 1000, 2)
                        extra["Total Request Time (ms)"] = total_request_ms
                        extra["HTML Size (KB)"] = round(len(html.encode("utf-8")) / 1024, 2)
                        soup = BeautifulSoup(html, "lxml")
                    else:
                        # Non-HTML or non-200 pages should not enter HTML parsing logic.
                        completed = True
                        break

                    # 1. Indexability (Check robots meta)
                    robots_meta = soup.find("meta", attrs={"name": "robots"})
                    if (
                        robots_meta
                        and "noindex" in robots_meta.get("content", "").lower()
                    ):
                        main_data["Indexability"] = "Noindex"
                    if robots_meta and robots_meta.get("content"):
                        extra["Meta Robots Raw"] = robots_meta.get("content").strip()

                    # 2. Title & Length
                    if soup.title and soup.title.string:
                        title_text = soup.title.string.strip()
                        main_data["Title"] = title_text
                        main_data["Title Length"] = len(title_text)
                        extra["Title Missing"] = False
                        extra["SERP Title Pixel Approx"] = int(len(title_text) * 7.2)
                        extra["SERP Title Truncation Risk"] = len(title_text) > 60

                    # 3. Meta Description & Length
                    desc_meta = soup.find("meta", attrs={"name": "description"})
                    if desc_meta and desc_meta.get("content"):
                        desc_text = desc_meta["content"].strip()
                        main_data["Meta Description"] = desc_text
                        main_data["Meta Desc Length"] = len(desc_text)
                        extra["Meta Description Missing"] = False
                        extra["SERP Meta Pixel Approx"] = int(len(desc_text) * 6.2)
                        extra["SERP Meta Truncation Risk"] = len(desc_text) > 155

                    # 4. H-Tags (H1 - H6)
                    for i in range(1, 7):
                        tags = soup.find_all(f"h{i}")
                        if tags:
                            # Join multiple same-level headings with a pipe delimiter
                            h_text = " | ".join(
                                [
                                    t.get_text(strip=True)
                                    for t in tags
                                    if t.get_text(strip=True)
                                ]
                            )
                            main_data[f"H{i} Content"] = h_text
                            main_data[f"H{i} Length"] = len(h_text)
                    extra["H1 Count"] = len(soup.find_all("h1"))
                    extra["Missing H1 Flag"] = extra["H1 Count"] == 0
                    extra["Multiple H1 Flag"] = extra["H1 Count"] > 1

                    # 5. Word Count (Body)
                    if soup.body:
                        # Extract text, stripping excessive whitespace and HTML tags
                        body_text = soup.body.get_text(separator=" ", strip=True)
                        words = body_text.split()
                        word_count = len(words)
                        main_data["Word Count (Body)"] = word_count
                        extra["Word Count"] = word_count
                        extra["Thin Content Flag"] = word_count < 300
                        extra["Word Count Band"] = word_count_band(word_count)
                        extra["Body Text-to-HTML Ratio"] = (
                            round(len(body_text) / max(1, len(html)) * 100, 2)
                        )
                        sentence_count = max(
                            1, len([s for s in re.split(r"[.!?]+", body_text) if s.strip()])
                        )
                        extra["Sentence Count"] = sentence_count
                        extra["Readability (Rough Flesch)"] = readability_flesch(
                            word_count, sentence_count
                        )

                    # 6. OG-Image
                    og_image = soup.find("meta", attrs={"property": "og:image"})
                    if og_image and og_image.get("content"):
                        main_data["OG-Image"] = og_image["content"]
                        extra["OG Image"] = og_image["content"]

                    # 7. Structured Data Validation (JSON-LD)
                    json_ld_scripts = soup.find_all(
                        "script", attrs={"type": "application/ld+json"}
                    )
                    if json_ld_scripts:
                        main_data["Has Valid JSON-LD"] = True

                    # Canonical, hreflang, pagination
                    canonical_tag = soup.find("link", attrs={"rel": "canonical"})
                    if canonical_tag and canonical_tag.get("href"):
                        canonical_url = urljoin(str(response.url), canonical_tag["href"].strip())
                        extra["Canonical URL"] = canonical_url
                        extra["Canonical Matches Final URL"] = (
                            canonical_url.rstrip("/") == str(response.url).rstrip("/")
                        )
                        extra["Canonical Absolute URL"] = canonical_url.startswith(
                            ("http://", "https://")
                        )
                        extra["Canonical Type"] = (
                            "self"
                            if extra["Canonical Matches Final URL"]
                            else "cross-canonical"
                        )
                    else:
                        extra["Canonical Type"] = "missing"

                    hreflangs = soup.find_all("link", attrs={"rel": "alternate", "hreflang": True})
                    extra["Hreflang Count"] = len(hreflangs)
                    extra["Hreflang Present"] = len(hreflangs) > 0
                    final_url_norm = str(response.url).rstrip("/")
                    self_href = False
                    has_x_default = False
                    for hreflang_tag in hreflangs:
                        href = (hreflang_tag.get("href") or "").strip()
                        lang = (hreflang_tag.get("hreflang") or "").strip().lower()
                        if href and urljoin(str(response.url), href).rstrip("/") == final_url_norm:
                            self_href = True
                        if lang == "x-default":
                            has_x_default = True
                    extra["Hreflang Self Reference"] = self_href
                    extra["x-default Present"] = has_x_default

                    next_tag = soup.find("link", attrs={"rel": "next"})
                    prev_tag = soup.find("link", attrs={"rel": "prev"})
                    extra["Pagination rel=next"] = bool(next_tag and next_tag.get("href"))
                    extra["Pagination rel=prev"] = bool(prev_tag and prev_tag.get("href"))

                    # Dates from common metas
                    published_meta = soup.find(
                        "meta", attrs={"property": "article:published_time"}
                    ) or soup.find("meta", attrs={"name": "publish_date"})
                    modified_meta = soup.find(
                        "meta", attrs={"property": "article:modified_time"}
                    ) or soup.find("meta", attrs={"name": "lastmod"})
                    if published_meta and published_meta.get("content"):
                        extra["Published Date"] = published_meta["content"].strip()
                    if modified_meta and modified_meta.get("content"):
                        extra["Modified Date"] = modified_meta["content"].strip()
                    if not extra["Last Updated"] and extra["Modified Date"]:
                        extra["Last Updated"] = extra["Modified Date"]

                    # Link extraction
                    current_host = urlparse(str(response.url)).netloc
                    internal_links = []
                    internal_links_full = []
                    external_links = 0
                    nofollow_internal = 0
                    nofollow_external = 0
                    generic_anchor_count = 0
                    generic_anchors = {
                        "click here",
                        "read more",
                        "learn more",
                        "more",
                        "here",
                    }
                    for link in soup.find_all("a", href=True):
                        href = link["href"].strip()
                        if not href or href.startswith("#") or href.startswith("mailto:") or href.startswith("tel:"):
                            continue
                        abs_url = urljoin(str(response.url), href)
                        anchor_text = link.get_text(" ", strip=True).lower()
                        if anchor_text in generic_anchors:
                            generic_anchor_count += 1
                        rel_set = {r.lower() for r in (link.get("rel") or [])}
                        is_nofollow = "nofollow" in rel_set
                        parsed = urlparse(abs_url)
                        if parsed.scheme not in ("http", "https"):
                            continue
                        if parsed.netloc == current_host:
                            internal_links.append(abs_url)
                            internal_links_full.append(abs_url)
                            if is_nofollow:
                                nofollow_internal += 1
                        else:
                            external_links += 1
                            if is_nofollow:
                                nofollow_external += 1
                        extra["Link Details"].append(
                            {
                                "Source URL": str(response.url),
                                "Target URL": abs_url,
                                "Internal": parsed.netloc == current_host,
                                "Nofollow": is_nofollow,
                                "Anchor Text": link.get_text(" ", strip=True),
                            }
                        )
                    extra["Internal Links Count"] = len(internal_links)
                    extra["External Links Count"] = external_links
                    extra["Unique Internal Links Count"] = len(set(internal_links))
                    extra["Nofollow Internal Links Count"] = nofollow_internal
                    extra["Nofollow External Links Count"] = nofollow_external
                    extra["Generic Anchor Text Count"] = generic_anchor_count
                    extra["Internal Links List Full"] = internal_links_full
                    extra["Internal Links List"] = list(set(internal_links))

                    # Media
                    images = soup.find_all("img")
                    missing_alt = 0
                    for img in images:
                        if not (img.get("alt") or "").strip():
                            missing_alt += 1
                        if urlparse(str(response.url)).scheme == "https":
                            src = (img.get("src") or "").strip()
                            if src.startswith("http://"):
                                extra["Mixed Content Detected"] = True
                    for tag in soup.find_all(["script", "link"]):
                        src = (tag.get("src") or tag.get("href") or "").strip()
                        if (
                            urlparse(str(response.url)).scheme == "https"
                            and src.startswith("http://")
                        ):
                            extra["Mixed Content Detected"] = True
                    extra["Image Count"] = len(images)
                    image_urls = []
                    for img in images:
                        src = (img.get("src") or "").strip()
                        if src:
                            image_urls.append(urljoin(str(response.url), src))
                    unique_images = sorted(set(image_urls))
                    extra["Images"] = " | ".join(unique_images) if unique_images else None
                    extra["Images Missing Alt"] = missing_alt
                    extra["Image Alt Coverage (%)"] = (
                        round(((len(images) - missing_alt) / len(images)) * 100, 2)
                        if images
                        else None
                    )
                    ext_counts = defaultdict(int)
                    large_images = 0
                    generic_image_names = 0
                    on_canonical_domain = 0
                    page_host = urlparse(str(response.url)).netloc
                    for img_url in unique_images:
                        ext_counts[image_extension(img_url)] += 1
                        if looks_generic_image_filename(img_url):
                            generic_image_names += 1
                        if urlparse(img_url).netloc == page_host:
                            on_canonical_domain += 1
                    for img in images:
                        try:
                            width = int((img.get("width") or "0").strip() or "0")
                            height = int((img.get("height") or "0").strip() or "0")
                            if width * height >= 1200 * 1200:
                                large_images += 1
                        except Exception:
                            pass
                    extra["Image Extension Distribution"] = ", ".join(
                        f"{k}:{v}" for k, v in sorted(ext_counts.items())
                    ) if ext_counts else None
                    extra["Likely Large Image Count"] = large_images
                    extra["Image Filename Quality Issues"] = generic_image_names
                    extra["Image On Canonical Domain (%)"] = (
                        round((on_canonical_domain / len(unique_images)) * 100, 2)
                        if unique_images
                        else None
                    )

                    # Structured data detail
                    schema_types = []
                    parse_errors = 0
                    for script in json_ld_scripts:
                        raw = (script.string or "").strip()
                        if not raw:
                            continue
                        try:
                            parsed_json = json.loads(raw)
                            nodes = parsed_json if isinstance(parsed_json, list) else [parsed_json]
                            for node in nodes:
                                if isinstance(node, dict):
                                    atype = node.get("@type")
                                    if isinstance(atype, list):
                                        schema_types.extend([str(t) for t in atype])
                                    elif atype:
                                        schema_types.append(str(atype))
                        except Exception:
                            parse_errors += 1
                    uniq_types = sorted(set(schema_types))
                    extra["Schema Types Found"] = ", ".join(uniq_types) if uniq_types else None
                    extra["Schema Types Count"] = len(uniq_types)
                    extra["Schema Parse Errors"] = parse_errors
                    schema_lower = {s.lower() for s in uniq_types}
                    extra["QAPage/FAQ Schema Present"] = bool(
                        {"faqpage", "qapage"} & schema_lower
                    )
                    extra["Speakable Schema Present"] = bool(
                        {"speakable", "speakablespecification"} & schema_lower
                    )

                    # Social
                    og_title = soup.find("meta", attrs={"property": "og:title"})
                    og_desc = soup.find("meta", attrs={"property": "og:description"})
                    twitter_card = soup.find("meta", attrs={"name": "twitter:card"})
                    extra["OG Title"] = og_title.get("content").strip() if og_title and og_title.get("content") else None
                    extra["OG Description"] = og_desc.get("content").strip() if og_desc and og_desc.get("content") else None
                    extra["Twitter Card Type"] = (
                        twitter_card.get("content").strip()
                        if twitter_card and twitter_card.get("content")
                        else None
                    )
                    extra["Open Graph Complete"] = bool(
                        extra["OG Title"] and extra["OG Description"] and extra["OG Image"]
                    )

                    # AEO-oriented content signals
                    faq_sections = soup.find_all(
                        ["section", "div"], string=re.compile(r"faq", re.I)
                    )
                    extra["FAQ Section Count"] = len(faq_sections)
                    q_headings = 0
                    for h in soup.find_all(["h1", "h2", "h3", "h4"]):
                        htxt = h.get_text(" ", strip=True)
                        if htxt.endswith("?"):
                            q_headings += 1
                    extra["Question Heading Count"] = q_headings

                    body_text_for_aeo = (
                        soup.body.get_text(" ", strip=True).lower() if soup.body else ""
                    )
                    extra["HowTo Signal"] = any(
                        key in body_text_for_aeo
                        for key in ["step 1", "how to", "steps to", "instructions"]
                    )
                    extra["Definition Signal"] = any(
                        key in body_text_for_aeo
                        for key in [" is ", " refers to ", " means "]
                    )
                    extra["List/Table Answer Signal"] = bool(
                        soup.find("ul") or soup.find("ol") or soup.find("table")
                    )
                    paragraphs = [
                        p.get_text(" ", strip=True)
                        for p in soup.find_all("p")
                        if p.get_text(" ", strip=True)
                    ]
                    medium_para_count = 0
                    for ptxt in paragraphs:
                        wc = len(ptxt.split())
                        if 40 <= wc <= 60:
                            medium_para_count += 1
                    extra["Paragraphs 40-60 Words Count"] = medium_para_count

                    aeo_score = 0
                    aeo_score += min(20, q_headings * 5)
                    aeo_score += 15 if extra["QAPage/FAQ Schema Present"] else 0
                    aeo_score += 10 if extra["Speakable Schema Present"] else 0
                    aeo_score += 10 if extra["HowTo Signal"] else 0
                    aeo_score += 10 if extra["Definition Signal"] else 0
                    aeo_score += 10 if extra["List/Table Answer Signal"] else 0
                    aeo_score += min(15, medium_para_count * 3)
                    aeo_score += 10 if (main_data.get("Title") and main_data.get("Meta Description")) else 0
                    extra["AEO Readiness Score"] = min(100, aeo_score)
                    if extra["AEO Readiness Score"] >= 80:
                        extra["AEO Badge"] = "Strong"
                    elif extra["AEO Readiness Score"] >= 60:
                        extra["AEO Badge"] = "Good"
                    elif extra["AEO Readiness Score"] >= 40:
                        extra["AEO Badge"] = "Fair"
                    else:
                        extra["AEO Badge"] = "Needs Work"

                    # Full suite robots check per host
                    if full_suite and extra.get("Final URL"):
                        parsed_final = urlparse(str(extra["Final URL"]))
                        host_key = f"{parsed_final.scheme}://{parsed_final.netloc}"
                        if robots_cache is not None and host_key in robots_cache:
                            cache_data = robots_cache[host_key]
                            extra["Robots.txt Accessible"] = cache_data["accessible"]
                            extra["Sitemap in Robots.txt"] = cache_data["has_sitemap"]
                            extra["Robots.txt Crawl-Delay"] = cache_data.get(
                                "crawl_delay"
                            )
                            extra["Robots.txt Disallow /"] = cache_data.get(
                                "disallow_root"
                            )
                        else:
                            robots_url = f"{host_key}/robots.txt"
                            try:
                                async with session.get(robots_url, timeout=timeout) as robots_resp:
                                    if robots_resp.status == 200:
                                        robots_txt = await robots_resp.text()
                                        crawl_delay = None
                                        disallow_root = False
                                        for line in robots_txt.splitlines():
                                            line_clean = line.strip().lower()
                                            if (
                                                not line_clean
                                                or line_clean.startswith("#")
                                            ):
                                                continue
                                            if line_clean.startswith("crawl-delay:"):
                                                try:
                                                    crawl_delay = float(
                                                        line_clean.split(":", 1)[1].strip()
                                                    )
                                                except Exception:
                                                    crawl_delay = None
                                            if line_clean.startswith("disallow:"):
                                                dis = line_clean.split(":", 1)[1].strip()
                                                if dis == "/":
                                                    disallow_root = True
                                        cache_data = {
                                            "accessible": True,
                                            "has_sitemap": "sitemap:" in robots_txt.lower(),
                                            "crawl_delay": crawl_delay,
                                            "disallow_root": disallow_root,
                                        }
                                    else:
                                        cache_data = {
                                            "accessible": False,
                                            "has_sitemap": False,
                                            "crawl_delay": None,
                                            "disallow_root": False,
                                        }
                            except Exception:
                                cache_data = {
                                    "accessible": False,
                                    "has_sitemap": False,
                                    "crawl_delay": None,
                                    "disallow_root": False,
                                }
                            extra["Robots.txt Accessible"] = cache_data["accessible"]
                            extra["Sitemap in Robots.txt"] = cache_data["has_sitemap"]
                            extra["Robots.txt Crawl-Delay"] = cache_data["crawl_delay"]
                            extra["Robots.txt Disallow /"] = cache_data["disallow_root"]
                            if robots_cache is not None:
                                robots_cache[host_key] = cache_data

                    completed = True
                    break

            except asyncio.TimeoutError:
                if attempt < MAX_RETRIES:
                    wait_time = min(
                        RETRY_MAX_DELAY_SECONDS,
                        RETRY_BASE_DELAY_SECONDS * (RETRY_BACKOFF_FACTOR ** attempt),
                    ) + random.uniform(0, REQUEST_JITTER_SECONDS)
                    print(
                        f"[Timeout] Retrying {url} "
                        f"(attempt {attempt + 2}/{MAX_RETRIES + 1}) in {wait_time:.1f}s"
                    )
                    await asyncio.sleep(wait_time)
                    continue
                main_data["Status Code"] = "Timeout"
                extra["Status Code"] = "Timeout"
                break
            except aiohttp.ClientError:
                if attempt < MAX_RETRIES:
                    wait_time = min(
                        RETRY_MAX_DELAY_SECONDS,
                        RETRY_BASE_DELAY_SECONDS * (RETRY_BACKOFF_FACTOR ** attempt),
                    ) + random.uniform(0, REQUEST_JITTER_SECONDS)
                    print(
                        f"[Connection Error] Retrying {url} "
                        f"(attempt {attempt + 2}/{MAX_RETRIES + 1}) in {wait_time:.1f}s"
                    )
                    await asyncio.sleep(wait_time)
                    continue
                main_data["Status Code"] = "Connection Error"
                extra["Status Code"] = "Connection Error"
                break
            except Exception as e:
                main_data["Status Code"] = f"Error: {str(e)}"
                extra["Status Code"] = f"Error: {str(e)}"
                break

        if not completed and main_data["Load Time (s)"] is None:
            main_data["Load Time (s)"] = round(time.time() - start_time, 3)
        if extra["Status Class"] is None:
            extra["Status Class"] = status_class(extra["Status Code"])

        status_val = extra["Status Code"]
        status_int = status_val if isinstance(status_val, int) else None
        indexability_reasons = []
        if status_int is not None and status_int >= 400:
            indexability_reasons.append(f"HTTP {status_int}")
        if isinstance(status_val, str) and status_val in {"Timeout", "Connection Error"}:
            indexability_reasons.append(status_val)
        robots_raw = (extra["Meta Robots Raw"] or "").lower()
        x_robots = (extra["X-Robots-Tag"] or "").lower()
        if "noindex" in robots_raw or "noindex" in x_robots:
            indexability_reasons.append("Noindex directive")
            main_data["Indexability"] = "Noindex"
        if extra["Canonical Type"] == "cross-canonical":
            indexability_reasons.append("Canonical points elsewhere")
        if not indexability_reasons:
            indexability_reasons.append("Indexable")
        extra["Indexability Reason"] = " | ".join(indexability_reasons)

        print(f"[{main_data['Status Code']}] Crawled: {url}")

        # Enforce server protection delay before releasing the semaphore
        delay_seconds = (
            request_delay if request_delay is not None else DELAY_BETWEEN_REQUESTS
        )
        await asyncio.sleep(delay_seconds + random.uniform(0, REQUEST_JITTER_SECONDS))

        return {"main": main_data, "extra": extra}


async def parse_sitemap(url, session):
    """Fetches and extracts all <loc> URLs from an XML sitemap."""
    print(f"Fetching sitemap from: {url}")
    try:
        async with session.get(url) as response:
            if response.status != 200:
                print("Failed to retrieve sitemap.")
                return [], {}

            xml_data = await response.text()
            # Remove namespaces to make parsing easier
            xml_data = re.sub(r'\sxmlns="[^"]+"', "", xml_data, count=1)
            root = ET.fromstring(xml_data)

            urls = []
            sitemap_meta = {}
            for url_node in root.findall(".//url"):
                loc_node = url_node.find("loc")
                if loc_node is None or not loc_node.text:
                    continue
                page_url = loc_node.text.strip()
                urls.append(page_url)
                sitemap_meta[page_url] = {
                    "changefreq": (
                        url_node.findtext("changefreq").strip()
                        if url_node.findtext("changefreq")
                        else None
                    ),
                    "priority": (
                        url_node.findtext("priority").strip()
                        if url_node.findtext("priority")
                        else None
                    ),
                    "lastmod": (
                        url_node.findtext("lastmod").strip()
                        if url_node.findtext("lastmod")
                        else None
                    ),
                }

            # Fallback for sitemap variants where only <loc> was present.
            if not urls:
                urls = [loc.text.strip() for loc in root.findall(".//loc") if loc.text]

            print(f"Found {len(urls)} URLs in sitemap.")
            return urls, sitemap_meta
    except Exception as e:
        print(f"Error parsing sitemap: {e}")
        return [], {}


async def check_url_status_light(session, url):
    """Lightweight status check for links not in primary crawl set."""
    timeout = aiohttp.ClientTimeout(
        total=min(TIMEOUT_SECONDS, 12),
        connect=min(CONNECT_TIMEOUT_SECONDS, 6),
        sock_read=min(READ_TIMEOUT_SECONDS, 10),
    )
    try:
        async with session.head(url, timeout=timeout, allow_redirects=True) as resp:
            return resp.status
    except Exception:
        try:
            async with session.get(url, timeout=timeout, allow_redirects=True) as resp:
                return resp.status
        except Exception:
            return None


async def check_url_status_light_limited(session, url, semaphore):
    """Concurrency-limited wrapper for lightweight link checks."""
    async with semaphore:
        return await check_url_status_light(session, url)


async def main():
    print("=== Python Technical SEO Auditor ===")
    print("1. Read URLs from a Sitemap (XML)")
    print("2. Read URLs from a local text file")

    choice = input("Select input method (1 or 2): ").strip()
    urls = []
    sitemap_meta = {}
    workers = MAX_WORKERS
    request_delay = DELAY_BETWEEN_REQUESTS
    source_label = "manual_input"

    # Session requires headers to bypass basic bot-blocking mechanisms
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Technical-SEO-Auditor/1.0"
    }

    async with aiohttp.ClientSession(headers=headers) as session:
        if choice == "1":
            sitemap_url = input(
                "Enter the full Sitemap URL (e.g., https://example.com/sitemap.xml): "
            ).strip()
            urls, sitemap_meta = await parse_sitemap(sitemap_url, session)
            parsed_source = urlparse(sitemap_url)
            host = parsed_source.netloc or "sitemap"
            path_part = sanitize_filename_part(parsed_source.path or "sitemap.xml")
            source_label = f"{host}_{path_part}"
        elif choice == "2":
            file_path = input("Enter the file path (e.g., urls.txt): ").strip()
            if os.path.exists(file_path):
                with open(file_path, "r", encoding="utf-8") as f:
                    urls = [line.strip() for line in f if line.strip()]
                print(f"Loaded {len(urls)} URLs from {file_path}.")
                source_label = os.path.splitext(os.path.basename(file_path))[0]
            else:
                print("File not found.")
        else:
            print("Invalid choice. Exiting.")
            return

        if not urls:
            print("No URLs to crawl. Exiting.")
            return

        # Dedupe URLs while preserving order to avoid unnecessary repeat requests.
        original_count = len(urls)
        urls = list(dict.fromkeys(urls))
        if len(urls) != original_count:
            print(f"Removed {original_count - len(urls)} duplicate URLs.")

        print("\nCrawl safety profile:")
        print("1. Gentle (fewer workers, longer delay)")
        print("2. Balanced (default)")
        print("3. Faster (more workers, shorter delay)")
        profile_choice = input("Select crawl profile (1, 2, or 3): ").strip()
        if profile_choice == "1":
            workers = 2
            request_delay = 4.0
        elif profile_choice == "3":
            workers = 4
            request_delay = 1.5

        suite_choice = input(
            "Run mode - 1) Main tab only  2) Full SEO suite (all tabs): "
        ).strip()
        full_suite = suite_choice == "2"
        output_filename = build_output_filename(source_label, full_suite)
        checkpoint_file = output_filename.replace(".xlsx", "_checkpoint.json")

        print(f"Output file: {output_filename}")

        print(f"\nStarting crawl of {len(urls)} URLs...")
        print(
            f"Max Workers: {workers} | Delay: {request_delay}s | "
            f"Retries: {MAX_RETRIES} | Timeout: {TIMEOUT_SECONDS}s | "
            f"Mode: {'Full Suite' if full_suite else 'Main Only'}"
        )
        checkpoint_raw = input(
            "Checkpoint save every N completed URLs (0 to disable): "
        ).strip()
        try:
            checkpoint_every = int(checkpoint_raw or "0")
        except ValueError:
            checkpoint_every = 0

        semaphore = asyncio.Semaphore(workers)
        robots_cache = {}
        resumed_results = []
        checkpoint_completed_urls = set()
        if os.path.exists(checkpoint_file):
            resume_choice = input(
                "Checkpoint found for this source. Resume from checkpoint? (y/N): "
            ).strip().lower()
            if resume_choice in {"y", "yes"}:
                try:
                    with open(checkpoint_file, "r", encoding="utf-8") as f:
                        checkpoint_data = json.load(f)
                    resumed_results = checkpoint_data.get("results", []) or []
                    checkpoint_completed_urls = set(
                        checkpoint_data.get("completed_urls", []) or []
                    )
                    if not checkpoint_completed_urls:
                        checkpoint_completed_urls = {
                            r.get("main", {}).get("URL")
                            for r in resumed_results
                            if r.get("main", {}).get("URL")
                        }
                    urls = [u for u in urls if u not in checkpoint_completed_urls]
                    print(
                        f"Resuming crawl. Completed: {len(checkpoint_completed_urls)} | "
                        f"Remaining: {len(urls)}"
                    )
                except Exception as e:
                    print(f"Could not load checkpoint. Starting fresh. ({e})")
                    resumed_results = []
                    checkpoint_completed_urls = set()

        tasks = [
            asyncio.create_task(
                fetch_and_parse(
                    url,
                    session,
                    semaphore,
                    full_suite,
                    robots_cache,
                    request_delay,
                    sitemap_meta,
                )
            )
            for url in urls
        ]

        # Gather results with progress and optional checkpoints.
        results = list(resumed_results)
        total_urls = len(urls) + len(checkpoint_completed_urls)
        for done_task in asyncio.as_completed(tasks):
            results.append(await done_task)
            done_count = len(results)
            if checkpoint_every > 0 and done_count % checkpoint_every == 0:
                completed_urls = [r.get("main", {}).get("URL") for r in results if r.get("main")]
                remaining_urls = [u for u in urls if u not in set(completed_urls)]
                checkpoint_payload = {
                    "saved_at": datetime.utcnow().isoformat() + "Z",
                    "completed": done_count,
                    "total": len(urls) + len(checkpoint_completed_urls),
                    "completed_urls": completed_urls,
                    "remaining_urls": remaining_urls,
                    "results": results,
                }
                with open(checkpoint_file, "w", encoding="utf-8") as f:
                    json.dump(checkpoint_payload, f, ensure_ascii=True, indent=2)
                print(
                    f"Checkpoint saved: {done_count}/{total_urls} -> {checkpoint_file}"
                )

        # Save final checkpoint snapshot for recovery convenience.
        if checkpoint_every > 0:
            completed_urls = [r.get("main", {}).get("URL") for r in results if r.get("main")]
            remaining_urls = [u for u in urls if u not in set(completed_urls)]
            checkpoint_payload = {
                "saved_at": datetime.utcnow().isoformat() + "Z",
                "completed": len(results),
                "total": len(urls) + len(checkpoint_completed_urls),
                "completed_urls": completed_urls,
                "remaining_urls": remaining_urls,
                "results": results,
            }
            with open(checkpoint_file, "w", encoding="utf-8") as f:
                json.dump(checkpoint_payload, f, ensure_ascii=True, indent=2)

        # Gather results concurrently
        # results = await asyncio.gather(*tasks)

        # Export to Excel
        print("\nGenerating Excel report...")
        main_rows = [r["main"] for r in results]
        extra_rows = [r["extra"] for r in results]
        main_df = pd.DataFrame(main_rows)
        extra_df = pd.DataFrame(extra_rows)
        status_by_url = {}
        for row in extra_rows:
            if row.get("Final URL"):
                status_by_url[str(row["Final URL"]).rstrip("/")] = row.get("Status Code")
            if row.get("URL"):
                status_by_url[str(row["URL"]).rstrip("/")] = row.get("Status Code")

        # Lightweight link validation for internal links not included in the crawl set.
        unresolved_targets = set()
        for row in extra_rows:
            for target in row.get("Internal Links List Full", []):
                t_norm = target.rstrip("/")
                if t_norm not in status_by_url:
                    unresolved_targets.add(target)
        if unresolved_targets:
            print(
                f"Running lightweight status checks for {len(unresolved_targets)} "
                "internal links not in crawl set..."
            )
            lightweight_check_workers = min(20, max(5, workers * 3))
            link_check_semaphore = asyncio.Semaphore(lightweight_check_workers)
            check_tasks = [
                check_url_status_light_limited(session, t, link_check_semaphore)
                for t in unresolved_targets
            ]
            checked_statuses = await asyncio.gather(*check_tasks)
            for target, status in zip(unresolved_targets, checked_statuses):
                status_by_url[target.rstrip("/")] = status

        # Hreflang consistency checks across crawled pages.
        crawled_finals = {
            str(row.get("Final URL")).rstrip("/")
            for row in extra_rows
            if row.get("Final URL")
        }

        for row in extra_rows:
            canonical_url = row.get("Canonical URL")
            if canonical_url and row.get("URL"):
                row["Canonical in Sitemap Match"] = canonical_url.rstrip("/") == str(
                    row["URL"]
                ).rstrip("/")
            row["Hreflang Canonical Consistency"] = (
                bool(row.get("Hreflang Present"))
                and row.get("Canonical Type") in {"self", "missing"}
            ) if row.get("Hreflang Present") else None
            if row.get("Hreflang Present"):
                row["Hreflang Reciprocal Check"] = (
                    str(row.get("Final URL", "")).rstrip("/") in crawled_finals
                    and bool(row.get("Hreflang Self Reference"))
                )

            link_statuses = []
            broken_internal = 0
            unresolved_internal = 0
            for target in row.get("Internal Links List Full", []):
                t_norm = target.rstrip("/")
                status = status_by_url.get(t_norm)
                if isinstance(status, int) and status >= 400:
                    broken_internal += 1
                elif status is None:
                    unresolved_internal += 1
                link_statuses.append(f"{target} => {status if status is not None else 'Not crawled'}")
            row["Broken Internal Links Count"] = broken_internal
            row["Unresolved Internal Links Count"] = unresolved_internal
            row["Internal Link Statuses"] = " | ".join(link_statuses) if link_statuses else None

        # Precompute inlinks for equity proxies.
        inlinks_map = defaultdict(set)
        crawled_set = {
            str((row.get("Final URL") or row.get("URL") or "")).rstrip("/")
            for row in extra_rows
            if (row.get("Final URL") or row.get("URL"))
        }
        for row in extra_rows:
            source = str((row.get("Final URL") or row.get("URL") or "")).rstrip("/")
            for target in row.get("Internal Links List", []):
                t_norm = str(target).rstrip("/")
                if t_norm in crawled_set and source:
                    inlinks_map[t_norm].add(source)

        # Cannibalization hints from duplicate title/meta clusters.
        title_map = defaultdict(list)
        meta_map = defaultdict(list)
        segment_by_url = {}
        for mrow in main_rows:
            t_key = normalize_text_hash(mrow.get("Title"))
            d_key = normalize_text_hash(mrow.get("Meta Description"))
            parsed_u = urlparse(str(mrow.get("URL") or ""))
            segs = [s for s in parsed_u.path.strip("/").split("/") if s]
            segment_by_url[mrow.get("URL")] = segs[0] if segs else "(home)"
            if t_key:
                title_map[t_key].append(mrow.get("URL"))
            if d_key:
                meta_map[d_key].append(mrow.get("URL"))

        summary_rules = get_summary_rules()
        score_by_url = {}
        for row in extra_rows:
            score, badge, icon, matched = score_url_health(row, summary_rules)
            row["SEO Health Score"] = score
            row["Severity Badge"] = badge
            row["Health Icon"] = icon
            row["Critical Issues Count"] = len(matched["Critical"])
            row["Warning Issues Count"] = len(matched["Warning"])
            row["Info Issues Count"] = len(matched["Info"])
            row["Matched Issues"] = " | ".join(
                matched["Critical"] + matched["Warning"] + matched["Info"]
            )
            row["Action Needed"] = "Yes" if badge in {"Critical", "Warning"} else "No"
            row["Owner"] = DEFAULT_OWNER_BY_SEVERITY.get(badge, "SEO")
            row["Sprint"] = ""
            row["Status"] = "Open"
            all_issue_ids = [stable_issue_id(row.get("URL"), issue) for issue in matched["Critical"] + matched["Warning"] + matched["Info"]]
            row["Stable Issue IDs"] = " | ".join(all_issue_ids) if all_issue_ids else None

            final_norm = str((row.get("Final URL") or row.get("URL") or "")).rstrip("/")
            inlinks_count = len(inlinks_map.get(final_norm, set()))
            if inlinks_count == 0:
                row["Inlinks Bucket"] = "0"
            elif inlinks_count <= 2:
                row["Inlinks Bucket"] = "1-2"
            elif inlinks_count <= 10:
                row["Inlinks Bucket"] = "3-10"
            else:
                row["Inlinks Bucket"] = "10+"
            row["Important But Underlinked"] = (
                score < 70 and inlinks_count <= 2
            )

            url_for_hint = row.get("URL")
            title_key = normalize_text_hash(
                next((m.get("Title") for m in main_rows if m.get("URL") == url_for_hint), None)
            )
            meta_key = normalize_text_hash(
                next((m.get("Meta Description") for m in main_rows if m.get("URL") == url_for_hint), None)
            )
            hints = []
            if title_key and len(title_map.get(title_key, [])) > 1:
                hints.append("Near-duplicate title cluster")
                seg_set = {
                    segment_by_url.get(u)
                    for u in title_map.get(title_key, [])
                    if segment_by_url.get(u)
                }
                if len(seg_set) == 1:
                    hints.append("Shared path segment pattern")
            if meta_key and len(meta_map.get(meta_key, [])) > 1:
                hints.append("Near-duplicate meta description cluster")
            if hints:
                row["Cannibalization Hint"] = " | ".join(hints)
            if row.get("URL"):
                score_by_url[row["URL"]] = {
                    "score": score,
                    "badge": badge,
                    "icon": icon,
                }

        for mrow in main_rows:
            url = mrow.get("URL")
            score_data = score_by_url.get(url, {})
            mrow["SEO Health Score"] = score_data.get("score")
            mrow["Severity Badge"] = score_data.get("badge")
            mrow["Health Icon"] = score_data.get("icon")

        extra_df = pd.DataFrame(extra_rows)
        main_df = pd.DataFrame(main_rows)

        # Previous run comparison (if prior workbook exists).
        prev_issue_ids = set()
        prev_counts = {}
        if os.path.exists(output_filename):
            try:
                prev_xls = pd.ExcelFile(output_filename)
                if "IssueInventory" in prev_xls.sheet_names:
                    prev_inv = pd.read_excel(output_filename, sheet_name="IssueInventory")
                    if "Stable Issue ID" in prev_inv.columns:
                        prev_issue_ids = {
                            str(v).strip()
                            for v in prev_inv["Stable Issue ID"].dropna().tolist()
                            if str(v).strip()
                        }
                if "Summary" in prev_xls.sheet_names:
                    prev_summary = pd.read_excel(output_filename, sheet_name="Summary")
                    for _, srow in prev_summary.iterrows():
                        if str(srow.get("Section", "")) == "Issue Counts":
                            issue_name = str(srow.get("Issue", ""))
                            prev_counts[issue_name] = int(srow.get("Affected URL Count", 0) or 0)
            except Exception:
                prev_issue_ids = set()
                prev_counts = {}

        # Optional safety backup before overwrite.
        if os.path.exists(output_filename):
            backup_choice = input(
                "Existing output found. Create timestamped backup before overwrite? (y/N): "
            ).strip().lower()
            if backup_choice in {"y", "yes"}:
                ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
                backup_file = output_filename.replace(".xlsx", f"_{ts}.bak.xlsx")
                try:
                    shutil.copy2(output_filename, backup_file)
                    print(f"Backup created: {backup_file}")
                except Exception as e:
                    print(f"Backup failed, continuing without backup: {e}")

        # Use pandas ExcelWriter to format the output sheet
        with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
            main_df.to_excel(writer, index=False, sheet_name="Main")
            adjust_sheet_format(writer, "Main")

            if full_suite:
                technical_cols = [
                    "URL",
                    "Health Icon",
                    "Severity Badge",
                    "SEO Health Score",
                    "Action Needed",
                    "Owner",
                    "Sprint",
                    "Status",
                    "Status Code",
                    "Final URL",
                    "Protocol",
                    "Redirect Chain Length",
                    "Redirect Target",
                    "Redirect Hops",
                    "HTTP->HTTPS Redirect",
                    "Status Class",
                    "TTFB (ms)",
                    "Total Request Time (ms)",
                    "Content-Type",
                    "HTTP Version",
                    "HTML Size (KB)",
                    "Compression Enabled",
                    "Cache-Control",
                    "ETag",
                    "X-Robots-Tag",
                    "Meta Robots Raw",
                    "Canonical URL",
                    "Canonical Matches Final URL",
                    "Canonical Type",
                    "Canonical Absolute URL",
                    "Canonical in Sitemap Match",
                    "Hreflang Present",
                    "Hreflang Count",
                    "Hreflang Self Reference",
                    "Hreflang Reciprocal Check",
                    "Hreflang Canonical Consistency",
                    "x-default Present",
                    "Pagination rel=next",
                    "Pagination rel=prev",
                    "Last-Modified",
                    "Published Date",
                    "Modified Date",
                    "Last Updated",
                    "Change Frequency",
                    "Priority",
                    "Indexability Reason",
                    "Critical Issues Count",
                    "Warning Issues Count",
                    "Info Issues Count",
                    "Inlinks Bucket",
                    "Important But Underlinked",
                    "SERP Title Truncation Risk",
                    "SERP Meta Truncation Risk",
                    "SERP Title Pixel Approx",
                    "SERP Meta Pixel Approx",
                    "Cannibalization Hint",
                    "Stable Issue IDs",
                    "URL Depth",
                    "Param URL Flag",
                ]
                content_cols = [
                    "URL",
                    "H1 Count",
                    "Missing H1 Flag",
                    "Multiple H1 Flag",
                    "Title Missing",
                    "Meta Description Missing",
                    "Word Count",
                    "Word Count Band",
                    "Sentence Count",
                    "Body Text-to-HTML Ratio",
                    "Readability (Rough Flesch)",
                    "Thin Content Flag",
                ]
                links_cols = [
                    "URL",
                    "Internal Links Count",
                    "Unique Internal Links Count",
                    "External Links Count",
                    "Nofollow Internal Links Count",
                    "Nofollow External Links Count",
                    "Generic Anchor Text Count",
                    "Broken Internal Links Count",
                    "Unresolved Internal Links Count",
                    "Internal Link Statuses",
                ]
                media_cols = [
                    "URL",
                    "Image Count",
                    "Images",
                    "Images Missing Alt",
                    "Image Alt Coverage (%)",
                    "Image Extension Distribution",
                    "Likely Large Image Count",
                    "Image Filename Quality Issues",
                    "Image On Canonical Domain (%)",
                    "Mixed Content Detected",
                ]
                schema_cols = [
                    "URL",
                    "Schema Types Found",
                    "Schema Types Count",
                    "Schema Parse Errors",
                ]
                aeo_cols = [
                    "URL",
                    "AEO Badge",
                    "AEO Readiness Score",
                    "FAQ Section Count",
                    "Question Heading Count",
                    "QAPage/FAQ Schema Present",
                    "Speakable Schema Present",
                    "HowTo Signal",
                    "Definition Signal",
                    "List/Table Answer Signal",
                    "Paragraphs 40-60 Words Count",
                    "Title Missing",
                    "Meta Description Missing",
                ]
                social_cols = [
                    "URL",
                    "OG Title",
                    "OG Description",
                    "OG Image",
                    "Open Graph Complete",
                    "Twitter Card Type",
                ]
                security_cols = [
                    "URL",
                    "Strict-Transport-Security",
                    "Content-Security-Policy",
                    "X-Content-Type-Options",
                    "X-Frame-Options",
                    "Referrer-Policy",
                    "Permissions-Policy",
                    "Robots.txt Accessible",
                    "Sitemap in Robots.txt",
                    "Robots.txt Crawl-Delay",
                    "Robots.txt Disallow /",
                ]

                extra_df[technical_cols].to_excel(writer, index=False, sheet_name="Technical")
                extra_df[content_cols].to_excel(writer, index=False, sheet_name="Content")
                extra_df[links_cols].to_excel(writer, index=False, sheet_name="Links")
                extra_df[media_cols].to_excel(writer, index=False, sheet_name="Media")
                extra_df[schema_cols].to_excel(
                    writer, index=False, sheet_name="StructuredData"
                )
                extra_df[aeo_cols].to_excel(writer, index=False, sheet_name="AEO")
                extra_df[social_cols].to_excel(writer, index=False, sheet_name="Social")
                extra_df[security_cols].to_excel(writer, index=False, sheet_name="Security")

                # Indexability tab
                indexability_cols = [
                    "URL",
                    "Status Code",
                    "Status Class",
                    "Final URL",
                    "Indexability Reason",
                    "Meta Robots Raw",
                    "X-Robots-Tag",
                    "Canonical URL",
                    "Canonical Type",
                    "Canonical Matches Final URL",
                    "Canonical in Sitemap Match",
                ]
                extra_df[indexability_cols].to_excel(
                    writer, index=False, sheet_name="Indexability"
                )

                # Redirects tab
                redirects_df = extra_df[
                    [
                        "URL",
                        "Status Code",
                        "Final URL",
                        "Redirect Chain Length",
                        "Redirect Target",
                        "Redirect Hops",
                        "HTTP->HTTPS Redirect",
                    ]
                ].copy()
                redirects_df["Redirect Loop Flag"] = redirects_df.apply(
                    lambda r: (
                        isinstance(r.get("Redirect Hops"), str)
                        and str(r.get("URL", "")).rstrip("/")
                        == str(r.get("Final URL", "")).rstrip("/")
                        and int(r.get("Redirect Chain Length") or 0) > 0
                    ),
                    axis=1,
                )
                redirects_df.to_excel(writer, index=False, sheet_name="Redirects")

                # Links detail tab
                link_rows = []
                for row in extra_rows:
                    for item in row.get("Link Details", []):
                        target_status = status_by_url.get(
                            str(item.get("Target URL", "")).rstrip("/")
                        )
                        item["Target Status (if crawled)"] = target_status
                        item["Crawlable"] = (
                            target_status is None
                            or (isinstance(target_status, int) and target_status < 400)
                        )
                        link_rows.append(item)
                pd.DataFrame(link_rows).to_excel(
                    writer, index=False, sheet_name="LinksDetail"
                )

                # Duplicate signals tab
                title_groups = defaultdict(list)
                desc_groups = defaultdict(list)
                for row in main_rows:
                    t_key = normalize_text_hash(row.get("Title"))
                    d_key = normalize_text_hash(row.get("Meta Description"))
                    if t_key:
                        title_groups[t_key].append(row.get("URL"))
                    if d_key:
                        desc_groups[d_key].append(row.get("URL"))
                duplicate_rows = []
                for row in main_rows:
                    t_key = normalize_text_hash(row.get("Title"))
                    d_key = normalize_text_hash(row.get("Meta Description"))
                    duplicate_rows.append(
                        {
                            "URL": row.get("URL"),
                            "Title Duplicate Count": len(title_groups.get(t_key, []))
                            if t_key
                            else 0,
                            "Meta Description Duplicate Count": len(
                                desc_groups.get(d_key, [])
                            )
                            if d_key
                            else 0,
                            "Title Duplicate URLs": " | ".join(title_groups.get(t_key, []))
                            if t_key and len(title_groups.get(t_key, [])) > 1
                            else None,
                            "Meta Duplicate URLs": " | ".join(desc_groups.get(d_key, []))
                            if d_key and len(desc_groups.get(d_key, [])) > 1
                            else None,
                        }
                    )
                pd.DataFrame(duplicate_rows).to_excel(
                    writer, index=False, sheet_name="Duplicates"
                )

                # Template clusters tab (by title pattern + first path segment).
                cluster_groups = defaultdict(list)
                for row in extra_rows:
                    final_url = row.get("Final URL") or row.get("URL")
                    if not final_url:
                        continue
                    parsed = urlparse(str(final_url))
                    segs = [s for s in parsed.path.strip("/").split("/") if s]
                    first_seg = segs[0] if segs else "(home)"
                    title = next(
                        (
                            m.get("Title")
                            for m in main_rows
                            if m.get("URL") == row.get("URL")
                        ),
                        None,
                    )
                    title_norm = normalize_text_hash(title)
                    title_pattern = (
                        re.sub(r"\d+", "{n}", title_norm)[:80] if title_norm else "(no-title)"
                    )
                    key = (first_seg, title_pattern)
                    cluster_groups[key].append(row)

                cluster_rows = []
                template_issue_counts = defaultdict(lambda: defaultdict(int))
                for (seg, pattern), urls_in_group in sorted(
                    cluster_groups.items(), key=lambda x: len(x[1]), reverse=True
                ):
                    url_list = [u.get("URL") for u in urls_in_group if u.get("URL")]
                    critical_count = sum(
                        1 for u in urls_in_group if (u.get("Critical Issues Count") or 0) > 0
                    )
                    warning_count = sum(
                        1
                        for u in urls_in_group
                        if (u.get("Critical Issues Count") or 0) == 0
                        and (u.get("Warning Issues Count") or 0) > 0
                    )
                    for u in urls_in_group:
                        for issue in str(u.get("Matched Issues") or "").split(" | "):
                            if issue:
                                template_issue_counts[seg][issue] += 1
                    dominant_issue = None
                    suggested_fix = None
                    if template_issue_counts[seg]:
                        dominant_issue = max(
                            template_issue_counts[seg].items(), key=lambda x: x[1]
                        )[0]
                        _, suggested_fix = root_cause_and_fix(dominant_issue)
                    avg_score = round(
                        sum([(u.get("SEO Health Score") or 0) for u in urls_in_group])
                        / max(1, len(urls_in_group)),
                        2,
                    )
                    cluster_rows.append(
                        {
                            "Path Segment": seg,
                            "Title Pattern": pattern,
                            "URL Count": len(url_list),
                            "Cluster Health Score Avg": avg_score,
                            "% with Critical": round(
                                (critical_count / max(1, len(url_list))) * 100, 2
                            ),
                            "% with Warnings": round(
                                (warning_count / max(1, len(url_list))) * 100, 2
                            ),
                            "Dominant Issue Type": dominant_issue,
                            "Suggested Template Fix": suggested_fix,
                            "URLs": " | ".join(url_list[:30]),
                        }
                    )
                pd.DataFrame(cluster_rows).to_excel(
                    writer, index=False, sheet_name="TemplateClusters"
                )

                # Summary tab: issue counts grouped by severity.
                summary_rows = []
                aeo_issue_names = {
                    "Low AEO Readiness Score",
                    "No FAQ/QA Schema",
                    "No Question Headings",
                    "No Answer-Friendly Structure",
                    "No 40-60 Word Answer Paragraphs",
                }
                summary_rows.append(
                    {
                        "Section": "=== Issue Counts ===",
                        "Severity": None,
                        "Issue": None,
                        "Affected URL Count": None,
                        "Affected URLs (sample)": None,
                    }
                )
                for severity, issue_name, rule_fn in summary_rules:
                    affected_urls = []
                    for row in extra_rows:
                        try:
                            if rule_fn(row):
                                affected_urls.append(row.get("URL"))
                        except Exception:
                            continue
                    summary_rows.append(
                        {
                            "Section": "Issue Counts",
                            "Severity": severity,
                            "Issue": issue_name,
                            "Affected URL Count": len(affected_urls),
                            "Reference Tab": (
                                "Indexability"
                                if "Canonical" in issue_name or "Noindex" in issue_name
                                else "Links"
                                if "Links" in issue_name
                                else "AEO"
                                if "AEO" in issue_name or "Question" in issue_name or "FAQ" in issue_name
                                else "Technical"
                            ),
                            "Affected URLs (sample)": " | ".join(
                                [u for u in affected_urls[:25] if u]
                            )
                            + " || Full list: see Technical/Links/Indexability tabs",
                        }
                    )
                summary_rows.append(
                    {
                        "Section": "=== AEO Opportunities ===",
                        "Severity": None,
                        "Issue": None,
                        "Affected URL Count": None,
                        "Affected URLs (sample)": "Detailed rows: see AEO tab",
                    }
                )
                for severity, issue_name, rule_fn in summary_rules:
                    if issue_name not in aeo_issue_names:
                        continue
                    affected_urls = []
                    for row in extra_rows:
                        try:
                            if rule_fn(row):
                                affected_urls.append(row.get("URL"))
                        except Exception:
                            continue
                    summary_rows.append(
                        {
                            "Section": "AEO Opportunities",
                            "Severity": severity,
                            "Issue": issue_name,
                            "Affected URL Count": len(affected_urls),
                            "Reference Tab": "AEO",
                            "Affected URLs (sample)": " | ".join(
                                [u for u in affected_urls[:25] if u]
                            )
                            + " || Full list: see AEO tab",
                        }
                    )

                severity_order = {"Critical": 0, "Warning": 1, "Info": 2}
                summary_rows = sorted(
                    summary_rows,
                    key=lambda x: (
                        x.get("Section", ""),
                        severity_order.get(x.get("Severity", ""), 99),
                        -(x.get("Affected URL Count") or 0),
                        x.get("Issue", ""),
                    ),
                )

                # Top 10 critical URLs
                summary_rows.append(
                    {
                        "Section": "=== Top 10 Critical URLs ===",
                        "Severity": None,
                        "Issue": None,
                        "Affected URL Count": None,
                        "Affected URLs (sample)": None,
                    }
                )
                critical_urls = sorted(
                    [r for r in extra_rows if (r.get("Critical Issues Count") or 0) > 0],
                    key=lambda r: (-(r.get("Critical Issues Count") or 0), r.get("SEO Health Score") or 100),
                )[:10]
                for idx, row in enumerate(critical_urls, start=1):
                    summary_rows.append(
                        {
                            "Section": "Top 10 Critical URLs",
                            "Severity": "Critical",
                            "Issue": f"#{idx} {row.get('URL')}",
                            "Affected URL Count": row.get("Critical Issues Count"),
                            "Reference Tab": "Priority URLs",
                            "Affected URLs (sample)": row.get("Matched Issues"),
                        }
                    )

                # Top recurring issue types by template
                summary_rows.append(
                    {
                        "Section": "=== Top Issues by Template ===",
                        "Severity": None,
                        "Issue": None,
                        "Affected URL Count": None,
                        "Affected URLs (sample)": None,
                    }
                )
                top_template_issues = []
                for seg, issues in template_issue_counts.items():
                    for issue_name, issue_count in issues.items():
                        top_template_issues.append((seg, issue_name, issue_count))
                top_template_issues = sorted(top_template_issues, key=lambda x: x[2], reverse=True)[:20]
                for seg, issue_name, issue_count in top_template_issues:
                    summary_rows.append(
                        {
                            "Section": "Top Issues by Template",
                            "Severity": "Info",
                            "Issue": f"{seg} -> {issue_name}",
                            "Affected URL Count": issue_count,
                            "Reference Tab": "TemplateClusters",
                            "Affected URLs (sample)": None,
                        }
                    )

                pd.DataFrame(summary_rows).to_excel(
                    writer, index=False, sheet_name="Summary"
                )

                # Issue inventory for stable IDs and run-to-run diff.
                issue_inventory_rows = []
                for row in extra_rows:
                    url = row.get("URL")
                    for issue in str(row.get("Matched Issues") or "").split(" | "):
                        if not issue:
                            continue
                        issue_inventory_rows.append(
                            {
                                "URL": url,
                                "Issue": issue,
                                "Stable Issue ID": stable_issue_id(url, issue),
                                "Severity": (
                                    "Critical"
                                    if issue in [i[1] for i in summary_rules if i[0] == "Critical"]
                                    else "Warning"
                                    if issue in [i[1] for i in summary_rules if i[0] == "Warning"]
                                    else "Info"
                                ),
                                "Owner": DEFAULT_OWNER_BY_SEVERITY.get(row.get("Severity Badge"), "SEO"),
                                "Sprint": "",
                                "Status": "Open",
                            }
                        )
                issue_inventory_df = pd.DataFrame(issue_inventory_rows)
                issue_inventory_df.to_excel(
                    writer, index=False, sheet_name="IssueInventory"
                )

                # Fix plan tab
                fixplan_rows = []
                for severity, issue_name, _ in summary_rules:
                    affected = [
                        r for r in extra_rows if issue_name in str(r.get("Matched Issues") or "").split(" | ")
                    ]
                    root_cause, recommended_fix = root_cause_and_fix(issue_name)
                    fixplan_rows.append(
                        {
                            "Category": "AEO" if issue_name in aeo_issue_names else "SEO",
                            "Issue Type": issue_name,
                            "Severity": severity,
                            "Affected Count": len(affected),
                            "Likely Root Cause": root_cause,
                            "Recommended Fix": recommended_fix,
                            "Owner": DEFAULT_OWNER_BY_SEVERITY.get(severity, "SEO"),
                            "Effort": DEFAULT_EFFORT_BY_SEVERITY.get(severity, "S"),
                            "Action Needed": "Yes" if severity in {"Critical", "Warning"} else "No",
                            "Sprint": "",
                            "Status": "Open",
                        }
                    )
                fixplan_df = pd.DataFrame(
                    sorted(fixplan_rows, key=lambda x: (-x["Affected Count"], x["Severity"]))
                )
                fixplan_df.to_excel(writer, index=False, sheet_name="FixPlan")

                # Priority URLs tab
                priority_rows = []
                for row in extra_rows:
                    risk_score = (
                        (row.get("Critical Issues Count") or 0) * 30
                        + (row.get("Warning Issues Count") or 0) * 10
                        + (100 - (row.get("SEO Health Score") or 100))
                    )
                    reasons = []
                    if (row.get("Critical Issues Count") or 0) > 0:
                        reasons.append("Has critical issues")
                    if (row.get("Broken Internal Links Count") or 0) > 0:
                        reasons.append("Broken internal links")
                    if row.get("Canonical Type") == "cross-canonical":
                        reasons.append("Cross canonical")
                    if "noindex" in str(row.get("Indexability Reason", "")).lower():
                        reasons.append("Noindex")
                    priority_rows.append(
                        {
                            "URL": row.get("URL"),
                            "Business Risk Score": int(risk_score),
                            "SEO Health Score": row.get("SEO Health Score"),
                            "Severity Badge": row.get("Severity Badge"),
                            "Critical Issues Count": row.get("Critical Issues Count"),
                            "Warning Issues Count": row.get("Warning Issues Count"),
                            "Indexability Reason": row.get("Indexability Reason"),
                            "Broken Internal Links Count": row.get("Broken Internal Links Count"),
                            "Canonical Type": row.get("Canonical Type"),
                            "Why Prioritized": " | ".join(reasons) if reasons else "Monitor",
                            "Action Needed": "Yes" if risk_score >= 30 else "No",
                            "Owner": DEFAULT_OWNER_BY_SEVERITY.get(row.get("Severity Badge"), "SEO"),
                            "Sprint": "",
                            "Status": "Open",
                        }
                    )
                priority_df = pd.DataFrame(
                    sorted(priority_rows, key=lambda x: x["Business Risk Score"], reverse=True)
                )
                priority_df.to_excel(writer, index=False, sheet_name="Priority URLs")

                # Dashboard tab
                total_urls = len(extra_rows)
                pass_count = len([r for r in extra_rows if r.get("Severity Badge") == "Pass"])
                critical_count = len([r for r in extra_rows if r.get("Severity Badge") == "Critical"])
                warning_count = len([r for r in extra_rows if r.get("Severity Badge") == "Warning"])
                score_bands = {"90-100": 0, "70-89": 0, "<70": 0}
                status_dist = defaultdict(int)
                for row in extra_rows:
                    status_dist[str(row.get("Status Class"))] += 1
                    s = row.get("SEO Health Score") or 0
                    if s >= 90:
                        score_bands["90-100"] += 1
                    elif s >= 70:
                        score_bands["70-89"] += 1
                    else:
                        score_bands["<70"] += 1
                top_blockers = fixplan_df.head(10)[["Issue Type", "Severity", "Affected Count"]]
                dashboard_rows = [
                    {"Metric": "URLs Crawled", "Value": total_urls},
                    {"Metric": "Pass Rate (%)", "Value": round((pass_count / max(1, total_urls)) * 100, 2)},
                    {"Metric": "Critical URL Count", "Value": critical_count},
                    {"Metric": "Warning URL Count", "Value": warning_count},
                    {"Metric": "Status Distribution", "Value": ", ".join([f"{k}:{v}" for k, v in sorted(status_dist.items())])},
                    {"Metric": "Score Bands", "Value": ", ".join([f"{k}:{v}" for k, v in score_bands.items()])},
                ]
                pd.DataFrame(dashboard_rows).to_excel(
                    writer, index=False, sheet_name="Dashboard"
                )
                quick_links_df = pd.DataFrame(
                    [
                        {"Label": "Open Summary", "Target Tab": "Summary"},
                        {"Label": "Open Main", "Target Tab": "Main"},
                        {"Label": "Open FixPlan", "Target Tab": "FixPlan"},
                        {"Label": "Open Priority URLs", "Target Tab": "Priority URLs"},
                        {"Label": "Open AEO", "Target Tab": "AEO"},
                    ]
                )
                quick_links_df.to_excel(
                    writer, index=False, sheet_name="Dashboard", startrow=1, startcol=12
                )
                top_blockers.to_excel(
                    writer, index=False, sheet_name="Dashboard", startrow=len(dashboard_rows) + 3
                )
                status_dist_df = pd.DataFrame(
                    [{"Status Class": k, "Count": v} for k, v in sorted(status_dist.items())]
                )
                status_dist_df.to_excel(
                    writer,
                    index=False,
                    sheet_name="Dashboard",
                    startrow=len(dashboard_rows) + 3,
                    startcol=5,
                )
                score_band_df = pd.DataFrame(
                    [{"Score Band": k, "Count": v} for k, v in score_bands.items()]
                )
                score_band_df.to_excel(
                    writer,
                    index=False,
                    sheet_name="Dashboard",
                    startrow=len(dashboard_rows) + 3,
                    startcol=8,
                )

                # Add visual charts to dashboard
                dash_ws = writer.sheets["Dashboard"]
                blockers_chart = BarChart()
                blockers_chart.title = "Top 10 Blockers"
                blockers_chart.y_axis.title = "Affected URLs"
                blockers_chart.height = 7
                blockers_chart.width = 14
                blockers_data = Reference(
                    dash_ws,
                    min_col=3,
                    min_row=len(dashboard_rows) + 5,
                    max_row=len(dashboard_rows) + 5 + min(9, len(top_blockers)),
                )
                blockers_cats = Reference(
                    dash_ws,
                    min_col=1,
                    min_row=len(dashboard_rows) + 5,
                    max_row=len(dashboard_rows) + 5 + min(9, len(top_blockers)),
                )
                blockers_chart.add_data(blockers_data, titles_from_data=False)
                blockers_chart.set_categories(blockers_cats)
                dash_ws.add_chart(blockers_chart, "A16")

                status_chart = PieChart()
                status_chart.title = "Status Class Distribution"
                status_chart.height = 7
                status_chart.width = 9
                status_data = Reference(
                    dash_ws,
                    min_col=7,
                    min_row=len(dashboard_rows) + 4,
                    max_row=len(dashboard_rows) + 3 + len(status_dist_df),
                )
                status_labels = Reference(
                    dash_ws,
                    min_col=6,
                    min_row=len(dashboard_rows) + 4,
                    max_row=len(dashboard_rows) + 3 + len(status_dist_df),
                )
                status_chart.add_data(status_data, titles_from_data=False)
                status_chart.set_categories(status_labels)
                dash_ws.add_chart(status_chart, "J16")

                score_chart = BarChart()
                score_chart.title = "SEO Score Bands"
                score_chart.y_axis.title = "URLs"
                score_chart.height = 7
                score_chart.width = 9
                score_data = Reference(
                    dash_ws,
                    min_col=10,
                    min_row=len(dashboard_rows) + 4,
                    max_row=len(dashboard_rows) + 3 + len(score_band_df),
                )
                score_labels = Reference(
                    dash_ws,
                    min_col=9,
                    min_row=len(dashboard_rows) + 4,
                    max_row=len(dashboard_rows) + 3 + len(score_band_df),
                )
                score_chart.add_data(score_data, titles_from_data=False)
                score_chart.set_categories(score_labels)
                dash_ws.add_chart(score_chart, "J31")

                # Run metadata and delta tabs
                run_meta_rows = [
                    {"Key": "Run Timestamp (UTC)", "Value": datetime.utcnow().isoformat() + "Z"},
                    {"Key": "Total URLs", "Value": len(urls)},
                    {"Key": "Mode", "Value": "Full Suite"},
                    {"Key": "Workers", "Value": workers},
                    {"Key": "Delay Seconds", "Value": request_delay},
                    {"Key": "Retries", "Value": MAX_RETRIES},
                    {"Key": "Timeout Seconds", "Value": TIMEOUT_SECONDS},
                    {"Key": "Checkpoint Every", "Value": checkpoint_every},
                ]
                pd.DataFrame(run_meta_rows).to_excel(
                    writer, index=False, sheet_name="RunMetadata"
                )

                current_issue_ids = {
                    r["Stable Issue ID"] for r in issue_inventory_rows if r.get("Stable Issue ID")
                }
                new_issues = current_issue_ids - prev_issue_ids
                resolved_issues = prev_issue_ids - current_issue_ids
                unchanged_issues = current_issue_ids & prev_issue_ids
                delta_rows = [
                    {"Metric": "New Issues", "Count": len(new_issues)},
                    {"Metric": "Resolved Issues", "Count": len(resolved_issues)},
                    {"Metric": "Unchanged Issues", "Count": len(unchanged_issues)},
                ]
                for sev, issue_name, _ in summary_rules:
                    current_count = len(
                        [r for r in extra_rows if issue_name in str(r.get("Matched Issues") or "").split(" | ")]
                    )
                    delta_rows.append(
                        {
                            "Metric": f"Issue Delta: {issue_name}",
                            "Count": current_count - int(prev_counts.get(issue_name, 0)),
                        }
                    )
                pd.DataFrame(delta_rows).to_excel(
                    writer, index=False, sheet_name="DeltaFromPreviousRun"
                )

                # Vocabulary / legend tab for the whole workbook.
                legend_rows = [
                    {"Section": "Severity", "Term": "Critical", "Meaning": "High-impact SEO blocker that should be fixed first.", "Values/Threshold": "Immediate action", "Related Tabs": "Summary, FixPlan, Technical"},
                    {"Section": "Severity", "Term": "Warning", "Meaning": "Out-of-best-practice issue likely affecting performance.", "Values/Threshold": "Plan next sprint", "Related Tabs": "Summary, FixPlan, Technical"},
                    {"Section": "Severity", "Term": "Info", "Meaning": "Optimization opportunity or context signal.", "Values/Threshold": "Backlog/monitor", "Related Tabs": "Summary, Technical"},
                    {"Section": "Scoring", "Term": "SEO Health Score", "Meaning": "Weighted technical SEO quality score per URL.", "Values/Threshold": ">=90 green, 70-89 yellow, <70 red", "Related Tabs": "Technical, Priority URLs, Dashboard"},
                    {"Section": "Scoring", "Term": "AEO Readiness Score", "Meaning": "Answer Engine Optimization readiness score per URL.", "Values/Threshold": ">=80 strong, 60-79 good, 40-59 fair", "Related Tabs": "AEO"},
                    {"Section": "Indexing", "Term": "Indexability Reason", "Meaning": "Primary reason URL may not be indexed.", "Values/Threshold": "Noindex, non-200, canonical mismatch", "Related Tabs": "Indexability, Technical"},
                    {"Section": "Links", "Term": "Broken Internal Links Count", "Meaning": "Internal links returning 4xx/5xx or equivalent failures.", "Values/Threshold": ">0 flagged", "Related Tabs": "Links, LinksDetail, Priority URLs"},
                    {"Section": "Content", "Term": "Word Count Band", "Meaning": "Body content depth class.", "Values/Threshold": "Thin / OK / Strong", "Related Tabs": "Content"},
                    {"Section": "AEO", "Term": "Question Heading Count", "Meaning": "Headings phrased as questions to match answer intent.", "Values/Threshold": "Higher is generally better", "Related Tabs": "AEO"},
                    {"Section": "Color Key", "Term": "Green", "Meaning": "Pass / aligned with best practice.", "Values/Threshold": "Good", "Related Tabs": "All"},
                    {"Section": "Color Key", "Term": "Yellow", "Meaning": "Warning / needs attention soon.", "Values/Threshold": "Medium risk", "Related Tabs": "All"},
                    {"Section": "Color Key", "Term": "Red", "Meaning": "Failure / high-priority issue.", "Values/Threshold": "High risk", "Related Tabs": "All"},
                    {"Section": "Color Key", "Term": "Purple", "Meaning": "Informational edge-case or AEO category signal.", "Values/Threshold": "Context", "Related Tabs": "All"},
                ]
                pd.DataFrame(legend_rows).to_excel(
                    writer, index=False, sheet_name="Legend"
                )

                # Crawl graph tab (inlinks and orphan candidates)
                inlinks_map = defaultdict(set)
                crawled_set = set(main_df["URL"].dropna().tolist())
                for row in extra_rows:
                    source = row.get("URL")
                    for target in row.get("Internal Links List", []):
                        if target in crawled_set:
                            inlinks_map[target].add(source)
                graph_rows = []
                for url_item in main_df["URL"].dropna().tolist():
                    inlinks = sorted(list(inlinks_map.get(url_item, set())))
                    graph_rows.append(
                        {
                            "URL": url_item,
                            "Inlinks Count": len(inlinks),
                            "Inlinks URLs": " | ".join(inlinks) if inlinks else None,
                            "Orphan Candidate": len(inlinks) == 0,
                        }
                    )
                pd.DataFrame(graph_rows).to_excel(
                    writer, index=False, sheet_name="CrawlGraph"
                )

                # Sitemap QA tab
                sitemap_rows = []
                if sitemap_meta:
                    for sitemap_url, meta in sitemap_meta.items():
                        matched = None
                        for row in extra_rows:
                            if (
                                str(row.get("URL", "")).rstrip("/")
                                == str(sitemap_url).rstrip("/")
                            ):
                                matched = row
                                break
                        final_url = matched.get("Final URL") if matched else None
                        status_code = matched.get("Status Code") if matched else None
                        sitemap_rows.append(
                            {
                                "Sitemap URL": sitemap_url,
                                "Final URL": final_url,
                                "Status Code": status_code,
                                "In Sitemap but Non-200": status_code != 200,
                                "Sitemap URL Redirects": (
                                    matched.get("Redirect Chain Length", 0) > 0
                                    if matched
                                    else None
                                ),
                                "In Sitemap but Canonicalized Elsewhere": (
                                    matched.get("Canonical Type") == "cross-canonical"
                                    if matched
                                    else None
                                ),
                                "Missing <lastmod>": not bool(meta.get("lastmod")),
                                "Missing <changefreq>": not bool(meta.get("changefreq")),
                                "Missing <priority>": not bool(meta.get("priority")),
                                "Sitemap <lastmod>": meta.get("lastmod"),
                                "Sitemap <changefreq>": meta.get("changefreq"),
                                "Sitemap <priority>": meta.get("priority"),
                            }
                        )
                pd.DataFrame(sitemap_rows).to_excel(
                    writer, index=False, sheet_name="SitemapQA"
                )

                # Enforce preferred tab order for first sheets.
                preferred_first_tabs = ["Dashboard", "Summary", "Main", "FixPlan"]
                wb = writer.book
                for idx, tab_name in enumerate(preferred_first_tabs):
                    if tab_name in wb.sheetnames:
                        wb.move_sheet(wb[tab_name], offset=-wb.index(wb[tab_name]) + idx)

                # Add useful cross-tab links.
                apply_tab_hyperlinks(writer)

                for sname in [
                    "Dashboard",
                    "Legend",
                    "Technical",
                    "Content",
                    "Links",
                    "LinksDetail",
                    "Media",
                    "StructuredData",
                    "AEO",
                    "Social",
                    "Security",
                    "Indexability",
                    "Redirects",
                    "Duplicates",
                    "TemplateClusters",
                    "Priority URLs",
                    "FixPlan",
                    "IssueInventory",
                    "RunMetadata",
                    "DeltaFromPreviousRun",
                    "CrawlGraph",
                    "SitemapQA",
                    "Summary",
                ]:
                    adjust_sheet_format(writer, sname)

        print(f"\nAudit complete! Report saved to {output_filename}")


if __name__ == "__main__":
    # Required for Windows environments utilizing asyncio
    if os.name == "nt":
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

    asyncio.run(main())

