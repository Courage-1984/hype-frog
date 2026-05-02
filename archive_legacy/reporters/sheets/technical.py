from __future__ import annotations

from openpyxl.utils import get_column_letter


def collapse_technical_deep_dive_columns(
    worksheet, sheet_name: str, *, header_index_fn
) -> None:
    if sheet_name != "Technical" or worksheet.max_column <= 1:
        return
    headers = header_index_fn(worksheet)
    deep_dive_headers = [
        "Hreflang Present",
        "Hreflang Count",
        "Hreflang Self Reference",
        "Hreflang Reciprocal Check",
        "Hreflang Canonical Consistency",
        "x-default Present",
        "Pagination rel=next",
        "Pagination rel=prev",
    ]
    deep_dive_cols = sorted([headers[h] for h in deep_dive_headers if h in headers])
    if not deep_dive_cols:
        return

    range_start = deep_dive_cols[0]
    prev_col = deep_dive_cols[0]
    for col_idx in deep_dive_cols[1:]:
        if col_idx == prev_col + 1:
            prev_col = col_idx
            continue
        worksheet.column_dimensions.group(
            get_column_letter(range_start),
            get_column_letter(prev_col),
            hidden=True,
            outline_level=1,
        )
        range_start = col_idx
        prev_col = col_idx
    worksheet.column_dimensions.group(
        get_column_letter(range_start),
        get_column_letter(prev_col),
        hidden=True,
        outline_level=1,
    )
