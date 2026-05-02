from __future__ import annotations

from collections import Counter
from collections import defaultdict
import math
import os
from typing import Any

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.views import Selection
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from reporters.formatting import (
    apply_fixplan_workflow_formatting,
    apply_global_conditional_formatting,
    ensure_auto_filter,
    ensure_freeze_header,
)
from reporters.sheets.links import (
    add_url_navigation_links as _add_url_navigation_links_impl,
    apply_cross_sheet_links as _apply_cross_sheet_links_impl,
    is_safe_hyperlink_target as _is_safe_hyperlink_target_impl,
    normalize_url_for_match as _normalize_url_for_match_impl,
    sanitize_excel_url as _sanitize_excel_url_impl,
)
from reporters.sheets.schema import add_schema_header_tooltips as _add_schema_header_tooltips_impl
from reporters.sheets.technical import (
    collapse_technical_deep_dive_columns as _collapse_technical_deep_dive_columns_impl,
)
from reporters.sheets import apply_workbook_toc_and_links
from utils import normalize_url_key

STD_NAVY = "2F3A4A"
STD_WHITE = "FFFFFF"
STD_BLUE = "2F6FA3"
DATA_HEAVY_TABS = {
    "Main",
    "Technical",
    "AEO",
    "AIOSEO",
    "FixPlan",
    "Summary",
    "Schema & Metadata",
    "Content Optimization Hub",
}


def _env_bool(name: str, default: bool) -> bool:
    raw = os.getenv(name)
    if raw is None:
        return default
    return str(raw).strip().lower() in {"1", "true", "yes", "y", "on"}


DEBUG_EXCEL_ISOLATION_MODE = _env_bool("HF_DEBUG_EXCEL_ISOLATION_MODE", False)
DISABLE_DATA_VALIDATION = _env_bool("HF_DISABLE_DATA_VALIDATION", False)
DISABLE_CONDITIONAL_FORMATTING = _env_bool("HF_DISABLE_CONDITIONAL_FORMATTING", False)
DISABLE_EXTERNAL_LINKS_AND_IMAGES = _env_bool(
    "HF_DISABLE_EXTERNAL_LINKS_AND_IMAGES", False
)
DISABLE_NON_CORE_FREEZE_PANES = _env_bool("HF_DISABLE_NON_CORE_FREEZE_PANES", False)


def _header_index(worksheet) -> dict[str, int]:
    return {
        str(cell.value): idx
        for idx, cell in enumerate(worksheet[1], start=1)
        if cell.value
    }


def _add_back_to_dashboard_link(worksheet, sheet_name: str) -> None:
    if DEBUG_EXCEL_ISOLATION_MODE:
        return
    if sheet_name in {"Dashboard", "Content Optimization Hub"}:
        return
    target_col = worksheet.max_column + 1
    target_ref = f"{get_column_letter(target_col)}1"
    worksheet[target_ref] = "BACK TO DASHBOARD"
    worksheet[target_ref].hyperlink = "#Dashboard!A1"
    worksheet[target_ref].style = "Hyperlink"
    worksheet[target_ref].font = Font(color=STD_BLUE, underline="single", bold=True)
    worksheet[target_ref].alignment = Alignment(horizontal="left")


def _is_safe_hyperlink_target(target: str) -> bool:
    return _is_safe_hyperlink_target_impl(
        target, disable_external_links_and_images=DISABLE_EXTERNAL_LINKS_AND_IMAGES
    )


def _to_int(value: Any, default: int = 0) -> int:
    try:
        return int(float(value))
    except Exception:
        return default


def _sort_worksheet_rows(worksheet, key_fn) -> None:
    if worksheet.max_row <= 2:
        return
    rows = [
        list(row)
        for row in worksheet.iter_rows(
            min_row=2, max_row=worksheet.max_row, values_only=True
        )
    ]
    rows.sort(key=key_fn)
    for row_idx, row_values in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_values, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)


def _apply_intelligent_sorting(worksheet, sheet_name: str) -> None:
    headers = _header_index(worksheet)
    if sheet_name == "FixPlan":
        severity_rank = {
            "Critical": 0,
            "High": 1,
            "Warning": 2,
            "Medium": 3,
            "Low": 4,
            "Observation": 5,
        }
        pcol = headers.get("Priority Score")
        scol = headers.get("Severity")
        ucol = headers.get("URL")
        if pcol:
            _sort_worksheet_rows(
                worksheet,
                key_fn=lambda r: (
                    -_to_int(r[pcol - 1], 0),
                    severity_rank.get(str(r[scol - 1]) if scol else "", 99),
                    str(r[ucol - 1] or "") if ucol else "",
                ),
            )
    elif sheet_name in {"Main", "Technical"}:
        sccol = headers.get("Status Code")
        ucol = headers.get("URL")
        if sccol:
            _sort_worksheet_rows(
                worksheet,
                key_fn=lambda r: (
                    -_to_int(r[sccol - 1], 0),
                    str(r[ucol - 1] or "") if ucol else "",
                ),
            )
    elif sheet_name == "AIOSEO":
        severity_rank = {"Critical": 0, "Warning": 1, "Observation": 2}
        sev_col = headers.get("Severity")
        pri_col = headers.get("Priority Score")
        panel_col = headers.get("AIOSEO Panel")
        ucol = headers.get("URL")
        if sev_col:
            _sort_worksheet_rows(
                worksheet,
                key_fn=lambda r: (
                    severity_rank.get(str(r[sev_col - 1]) if sev_col else "", 99),
                    -_to_int(r[pri_col - 1], 0) if pri_col else 0,
                    str(r[panel_col - 1] or "") if panel_col else "",
                    str(r[ucol - 1] or "") if ucol else "",
                ),
            )


def _hide_noisy_columns(worksheet, sheet_name: str) -> None:
    noisy_tokens_by_sheet = {
        "Main": ["json-ld", "schema", "raw", "headers", "paragraph", "text extract"],
        "Technical": [
            "json-ld",
            "schema",
            "raw",
            "headers",
            "html",
            "paragraph",
            "text extract",
        ],
        "Content": ["paragraph", "full text", "raw", "json", "headers"],
    }
    tokens = noisy_tokens_by_sheet.get(sheet_name)
    if not tokens:
        return
    for idx, cell in enumerate(worksheet[1], start=1):
        header = str(cell.value or "").lower()
        if any(tok in header for tok in tokens):
            worksheet.column_dimensions[get_column_letter(idx)].hidden = True


def _friendly_metric_label(header: str) -> str:
    return (
        header.replace("_", " ")
        .replace("URL", "URL")
        .replace("SEO", "SEO")
        .replace("AEO", "AEO")
        .strip()
    )


def _tooltip_for_header(header: str) -> str:
    h = (header or "").strip()
    lower = h.lower()
    if not h:
        return "Column descriptor for this worksheet section."
    if "url" in lower:
        return f"{_friendly_metric_label(h)}. Use this URL to inspect the page directly or jump to related tabs."
    if "status code" in lower:
        return "HTTP response code returned for this page/resource. Fix: Resolve 4xx/5xx errors and reduce unnecessary redirects."
    if "status class" in lower:
        return "Grouped HTTP class (2xx/3xx/4xx/5xx). Use this to quickly triage crawl health and error concentration."
    if "health score" in lower:
        return "Composite SEO quality score for this URL. Higher is better. Guide: 90+ strong, 70-89 needs tuning, below 70 high priority."
    if "pass rate" in lower:
        return "Share of URLs marked as pass across the crawl. Use together with Error Rate and critical issue counts for decisions."
    if "severity" in lower:
        return "Issue impact level. Fix: Resolve Critical first, then Warning, then improvement opportunities."
    if "priority score" in lower:
        return "Execution priority score combining impact and effort. Fix: Start from highest scores."
    if "affected count" in lower:
        return "How many URLs are impacted by this issue. Larger counts usually indicate template-level or systemic problems."
    if "ttfb" in lower:
        return "Time to First Byte: server responsiveness signal. Fix: optimise backend performance, caching, and CDN usage."
    if "load time" in lower or "request time" in lower:
        return "Page/request speed metric. Fix: optimise assets, server response time, and blocking resources."
    if "indexability" in lower or "canonical" in lower:
        return "Indexing/canonicalization signal. Fix: ensure indexable pages use correct canonicals and non-conflicting directives."
    if "meta robots" in lower or "x-robots" in lower:
        return "Robots directives that influence indexation. Fix: remove unintended noindex/nofollow values on important pages."
    if "word count" in lower or "readability" in lower:
        return "Content quality depth metric. Fix: expand thin content and improve clarity for search intent."
    if "link" in lower or "anchor" in lower:
        return "Link quality and crawl-path metric. Fix: repair broken links and improve internal linking relevance."
    if "open in main" in lower or "technical view" in lower or "view details" in lower:
        return "Navigation helper. Click to jump directly to the related record in another tab."
    if "schema" in lower or "json-ld" in lower:
        return "Structured data signal. Fix: add valid schema types and correct parsing/validation errors."
    if "owner" in lower or "sprint" in lower or "status" == lower:
        return (
            "Workflow management field for planning and tracking remediation progress."
        )
    if "section" in lower or "reference tab" in lower:
        return "Summary grouping/navigation field. Use with hyperlinks to jump into detailed tabs."
    return f"{_friendly_metric_label(h)}. Review this metric to identify risk, then use linked tabs for details and remediation."


def _add_all_header_tooltips(worksheet) -> None:
    if DISABLE_DATA_VALIDATION:
        return
    headers = _header_index(worksheet)
    for header, col_idx in headers.items():
        ref = f"{get_column_letter(col_idx)}1"
        dv = DataValidation(
            type="custom", formula1="TRUE", showInputMessage=True, allow_blank=True
        )
        dv.promptTitle = _friendly_metric_label(header)[:32] or "Column"
        dv.prompt = _tooltip_for_header(header)
        worksheet.add_data_validation(dv)
        dv.add(ref)


def _add_url_navigation_links(writer, worksheet, sheet_name: str) -> None:
    _add_url_navigation_links_impl(
        writer,
        worksheet,
        sheet_name,
        debug_excel_isolation_mode=DEBUG_EXCEL_ISOLATION_MODE,
        disable_external_links_and_images=DISABLE_EXTERNAL_LINKS_AND_IMAGES,
        header_index_fn=_header_index,
    )


def _sanitize_excel_url(url_value: Any) -> str:
    return _sanitize_excel_url_impl(url_value)


def _normalize_url_for_match(url_value: Any) -> str:
    return _normalize_url_for_match_impl(url_value)


def _ranges_overlap(range_a, range_b) -> bool:
    return not (
        range_a.max_row < range_b.min_row
        or range_b.max_row < range_a.min_row
        or range_a.max_col < range_b.min_col
        or range_b.max_col < range_a.min_col
    )


def _audit_non_overlapping_merges(worksheet) -> None:
    merge_ranges = list(worksheet.merged_cells.ranges)
    if len(merge_ranges) < 2:
        return
    kept = []
    for merge_range in merge_ranges:
        if any(_ranges_overlap(merge_range, existing) for existing in kept):
            worksheet.unmerge_cells(str(merge_range))
            continue
        kept.append(merge_range)


def _audit_freeze_merge_conflicts(worksheet) -> None:
    freeze = worksheet.freeze_panes
    if not freeze:
        return
    freeze_ref = freeze if isinstance(freeze, str) else freeze.coordinate
    freeze_row, freeze_col = coordinate_to_tuple(freeze_ref)
    for merge_range in list(worksheet.merged_cells.ranges):
        if (
            merge_range.min_row <= freeze_row <= merge_range.max_row
            and merge_range.min_col <= freeze_col <= merge_range.max_col
        ):
            worksheet.unmerge_cells(str(merge_range))


def _set_freeze_panes_safe(worksheet, value: str | None) -> None:
    view = worksheet.views.sheetView[0]
    if not view.selection:
        view.selection = [Selection(activeCell="A1", sqref="A1")]
    worksheet.freeze_panes = value
    _sanitize_sheet_view_selection(worksheet)


def _sanitize_sheet_view_selection(worksheet) -> None:
    view = worksheet.views.sheetView[0]
    pane = worksheet.sheet_view.pane
    if not view.selection:
        return
    x_split = float(getattr(pane, "xSplit", 0) or 0) if pane is not None else 0.0
    y_split = float(getattr(pane, "ySplit", 0) or 0) if pane is not None else 0.0

    valid_panes: set[str | None] = {None, ""}
    if x_split > 0 and y_split > 0:
        valid_panes.update({"topRight", "bottomLeft", "bottomRight"})
    elif x_split > 0 and y_split == 0:
        valid_panes.add("topRight")
    elif x_split == 0 and y_split > 0:
        valid_panes.add("bottomLeft")

    sanitized = [sel for sel in view.selection if getattr(sel, "pane", None) in valid_panes]
    if not sanitized:
        sanitized = [Selection(activeCell="A1", sqref="A1")]
    view.selection = sanitized


def _normalize_table_headers(worksheet, header_row: int = 1) -> None:
    seen: dict[str, int] = {}
    for col_idx in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=header_row, column=col_idx)
        raw = cell.value
        if (
            raw is None
            or (isinstance(raw, float) and math.isnan(raw))
            or not str(raw).strip()
        ):
            header = f"Column_{col_idx}"
        else:
            header = str(raw).replace("\n", " ").replace("\r", " ").strip()
            if len(header) > 255:
                header = header[:255].strip() or f"Column_{col_idx}"
        if header in seen:
            seen[header] += 1
            header = f"{header}_{seen[header]}"
        else:
            seen[header] = 0
        cell.value = header


def _compute_exact_table_ref(worksheet, header_row: int) -> str | None:
    if worksheet.max_row < header_row + 1:
        return None
    max_col = 1
    for row_idx in range(header_row, worksheet.max_row + 1):
        for col_idx in range(worksheet.max_column, 0, -1):
            value = worksheet.cell(row=row_idx, column=col_idx).value
            if value is not None and str(value) != "":
                if col_idx > max_col:
                    max_col = col_idx
                break
    if max_col < 1:
        return None
    return f"A{header_row}:{get_column_letter(max_col)}{worksheet.max_row}"


def _apply_mock_table_styling(
    worksheet, min_col: int, max_col: int, min_row: int, max_row: int
) -> None:
    if min_col > max_col or min_row > max_row:
        return
    if max_row >= min_row + 1:
        ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        worksheet.auto_filter.ref = ref

    header_fill = PatternFill(
        start_color=STD_NAVY, end_color=STD_NAVY, fill_type="solid"
    )
    header_font = Font(color=STD_WHITE, bold=True)
    for col_idx in range(min_col, max_col + 1):
        cell = worksheet.cell(row=min_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    band_fill = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
    for row_idx in range(min_row + 1, max_row + 1):
        if row_idx % 2 == 0:
            for col_idx in range(min_col, max_col + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.fill.fill_type is None:
                    cell.fill = band_fill

    if max_row >= min_row:
        if worksheet.title == "Content Optimization Hub":
            _set_freeze_panes_safe(worksheet, "D3")
        else:
            _set_freeze_panes_safe(worksheet, "A2")


def _apply_fixplan_interactivity(writer, worksheet) -> None:
    if DEBUG_EXCEL_ISOLATION_MODE:
        return
    headers = _header_index(worksheet)
    if worksheet.max_row <= 1:
        return
    if "Status" not in headers:
        new_col = worksheet.max_column + 1
        worksheet.cell(row=1, column=new_col, value="Status")
        for r in range(2, worksheet.max_row + 1):
            worksheet.cell(row=r, column=new_col, value="To Do")
        headers = _header_index(worksheet)
    status_col = headers.get("Status")
    if status_col:
        _apply_status_dropdown(worksheet, status_col)

    owner_col = headers.get("Agency Owner")
    if owner_col:
        owners = sorted(
            {
                str(worksheet.cell(r, owner_col).value).strip()
                for r in range(2, worksheet.max_row + 1)
                if worksheet.cell(r, owner_col).value
            }
        )
        if not owners:
            owners = ["Copy Writer", "Dev", "Server/Host"]
        owner_list = ",".join(owners[:8])
        owner_dv = DataValidation(
            type="list", formula1=f'"{owner_list}"', allow_blank=True
        )
        worksheet.add_data_validation(owner_dv)
        owner_dv.add(
            f"{get_column_letter(owner_col)}2:{get_column_letter(owner_col)}{worksheet.max_row}"
        )

    url_col = headers.get("URL")
    jump_col = headers.get("Jump to Details")
    details_col = headers.get("View Details")
    if not details_col:
        details_col = worksheet.max_column + 1
        worksheet.cell(row=1, column=details_col, value="View Details")
        headers = _header_index(worksheet)
    for row_idx in range(2, worksheet.max_row + 1):
        if url_col:
            url_cell = worksheet.cell(row=row_idx, column=url_col)
            url = str(url_cell.value or "").strip()
            if url.startswith(("http://", "https://")) and _is_safe_hyperlink_target(
                url
            ):
                url_cell.hyperlink = url
                url_cell.style = "Hyperlink"
                url_cell.alignment = Alignment(wrap_text=True, vertical="top")
        if jump_col and url_col:
            url = str(worksheet.cell(row=row_idx, column=url_col).value or "").strip()
            jump_cell = worksheet.cell(row=row_idx, column=jump_col)
            if url:
                if len(url) <= 255:
                    jump_cell.value = f'=IFERROR(HYPERLINK("#Main!A"&MATCH("{url}",Main!A:A,0),"Jump to Main"),HYPERLINK("#Main!A1","Open Main"))'
                else:
                    jump_cell.value = "Jump to Main"
            else:
                jump_cell.value = '=HYPERLINK("#Main!A1","Open Main")'
        if details_col and url_col:
            url_col_letter = get_column_letter(url_col)
            worksheet.cell(
                row=row_idx,
                column=details_col,
                value=f'=IFERROR(HYPERLINK("#Main!A"&MATCH({url_col_letter}{row_idx},Main!A:A,0),"View URL Details"),HYPERLINK("#Main!A1","View URL Details"))',
            )
        affected_col = headers.get("Affected URLs")
        detail_tab_col = headers.get("Detail Reference Tab")
        if affected_col and detail_tab_col:
            affected_cell = worksheet.cell(row=row_idx, column=affected_col)
            detail_tab = str(
                worksheet.cell(row=row_idx, column=detail_tab_col).value or ""
            ).strip()
            affected_text = str(affected_cell.value or "").strip()
            if affected_text.startswith("SEE DETAILS IN ") and detail_tab:
                affected_cell.value = (
                    f'=HYPERLINK("#{detail_tab}!A1","SEE DETAILS IN {detail_tab}")'
                )


def _apply_status_dropdown(worksheet, status_col: int) -> None:
    if DISABLE_DATA_VALIDATION:
        return
    if status_col <= 0 or worksheet.max_row <= 1:
        return
    status_dv = DataValidation(
        type="list", formula1='"To Do,In Progress,Fixed"', allow_blank=True
    )
    worksheet.add_data_validation(status_dv)
    status_dv.add(
        f"{get_column_letter(status_col)}2:{get_column_letter(status_col)}{worksheet.max_row}"
    )


def _style_dashboard(worksheet, writer) -> None:
    worksheet.sheet_view.showGridLines = False
    _set_freeze_panes_safe(worksheet, "A2")
    worksheet._charts = []
    light_header_fill = PatternFill("solid", fgColor="E5E7EB")
    headers = _header_index(worksheet)
    # Clear legacy helper blocks outside the B:C dashboard grid.
    for row_idx in range(1, 80):
        for col_idx in range(4, max(worksheet.max_column + 1, 26)):
            worksheet.cell(row=row_idx, column=col_idx, value=None)
    metric_col = headers.get("Metric")
    value_col = headers.get("Value")
    total_urls = 0
    pass_rate_pct = 0.0
    if metric_col and value_col:
        metric_to_value = {}
        for row_idx in range(2, worksheet.max_row + 1):
            metric_name = str(
                worksheet.cell(row=row_idx, column=metric_col).value or ""
            ).strip()
            if metric_name:
                metric_to_value[metric_name] = worksheet.cell(
                    row=row_idx, column=value_col
                ).value
        # Clear dataframe-rendered dashboard grid and rebuild as clean report blocks.
        if worksheet.max_row > 0:
            worksheet.delete_rows(1, worksheet.max_row)
        for col_letter, width in {
            "A": 35,
            "B": 15,
            "C": 5,
            "D": 30,
            "E": 15,
            "F": 5,
            "G": 15,
            "H": 15,
            "I": 15,
            "J": 15,
            "K": 15,
        }.items():
            worksheet.column_dimensions[col_letter].width = width
        total_urls = _to_int(metric_to_value.get("URLs Crawled"), 0)
        try:
            pass_rate_pct = float(metric_to_value.get("Pass Rate (%)", 0) or 0.0)
        except Exception:
            pass_rate_pct = 0.0
        critical_urls = _to_int(metric_to_value.get("Critical URL Count"), 0)
        warning_urls = _to_int(metric_to_value.get("Warning URL Count"), 0)

        title = worksheet["A1"]
        title.value = "Executive SEO & AEO Performance Report"
        worksheet["A2"] = "Executive SEO & AEO Dashboard"
        worksheet["A2"].font = Font(color=STD_NAVY, bold=True, size=12)
        title.font = Font(color=STD_NAVY, bold=True, size=16)
        title.alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[1].height = 30

        header_fill = PatternFill("solid", fgColor="ADD8E6")
        header_font = Font(color="000000", bold=True, size=12)
        value_fill = PatternFill("solid", fgColor="DCE3EA")

        worksheet["A4"] = "EXECUTIVE METRICS"
        worksheet["B4"] = "Value"
        for ref in ("A4", "B4"):
            worksheet[ref].fill = header_fill
            worksheet[ref].font = header_font
            worksheet[ref].alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        worksheet["A5"] = '=HYPERLINK("#\'Main\'!A1","Total URLs")'
        worksheet["B5"] = total_urls
        worksheet["A6"] = '=HYPERLINK("#\'Summary\'!A1","Overall Health Score")'
        worksheet["B6"] = "=IFERROR(0,0)"
        worksheet["B6"].number_format = "0.00%"
        worksheet["A7"] = '=HYPERLINK("#\'Summary\'!A1","SEO Pass Rate %")'
        worksheet["B7"] = "=IFERROR(0,0)"
        worksheet["B7"].number_format = "0.00%"
        worksheet["A8"] = '=HYPERLINK("#\'Technical\'!A1","Pass URLs")'
        worksheet["B8"] = 0
        worksheet["A9"] = '=HYPERLINK("#\'FixPlan\'!A1","Critical URLs")'
        worksheet["B9"] = critical_urls
        worksheet["A10"] = '=HYPERLINK("#\'Technical\'!A1","Warning URLs")'
        worksheet["B10"] = warning_urls
        worksheet["A11"] = '=HYPERLINK("#\'Technical\'!A1","Error Rate % (4xx/5xx)")'
        worksheet["B11"] = "=IFERROR(0,0)"
        worksheet["B11"].number_format = "0.00%"
        worksheet["A12"] = (
            '=HYPERLINK("#\'Technical\'!A1","Crawl Success Rate % (2xx)")'
        )
        worksheet["B12"] = "=IFERROR(0,0)"
        worksheet["B12"].number_format = "0.00%"
        worksheet["A13"] = '=HYPERLINK("#\'Technical\'!A1","Critical URL Rate %")'
        worksheet["B13"] = "=IFERROR(0,0)"
        worksheet["B13"].number_format = "0.00%"
        worksheet["A14"] = '=HYPERLINK("#\'Technical\'!A1","Warning URL Rate %")'
        worksheet["B14"] = "=IFERROR(0,0)"
        worksheet["B14"].number_format = "0.00%"
        worksheet["A15"] = "Projected Health Score % (if all To Do done)"
        worksheet["B15"] = "=IFERROR(0,0)"
        worksheet["B15"].number_format = "0.00%"
        worksheet["A16"] = "Projected Pass Rate % (if all To Do done)"
        worksheet["B16"] = "=IFERROR(0,0)"
        worksheet["B16"].number_format = "0.00%"
        worksheet["A17"] = (
            '=HYPERLINK("#\'Content Optimization Hub\'!A1","Content Hub Readiness (%)")'
        )
        worksheet["B17"] = (
            "=IFERROR(COUNTIF('Content Optimization Hub'!A:A,\"Completed\")/COUNTA('Content Optimization Hub'!E:E),0)"
        )
        worksheet["B17"].number_format = "0.00%"
        worksheet["A18"] = '=HYPERLINK("#\'Schema & Metadata\'!A1","URLs with Schema")'
        worksheet["B18"] = 0
        worksheet["A19"] = '=HYPERLINK("#\'Links\'!A1","Broken Internal Links")'
        worksheet["B19"] = 0
        for ref in (
            "A5",
            "B5",
            "A6",
            "B6",
            "A7",
            "B7",
            "A8",
            "B8",
            "A9",
            "B9",
            "A10",
            "B10",
            "A11",
            "B11",
            "A12",
            "B12",
            "A13",
            "B13",
            "A14",
            "B14",
            "A15",
            "B15",
            "A16",
            "B16",
            "A17",
            "B17",
            "A18",
            "B18",
            "A19",
            "B19",
        ):
            worksheet[ref].fill = value_fill
            worksheet[ref].font = Font(
                color="1F2937", bold=True, size=12 if ref.startswith("A") else 14
            )
            worksheet[ref].alignment = Alignment(horizontal="center", vertical="center")
        for ref in (
            "A5",
            "A6",
            "A7",
            "A8",
            "A9",
            "A10",
            "A11",
            "A12",
            "A13",
            "A14",
            "A17",
            "A18",
            "A19",
        ):
            worksheet[ref].font = Font(
                color=STD_BLUE, underline="single", bold=True, size=12
            )
        for row in range(5, 20):
            worksheet[f"A{row}"].alignment = Alignment(
                horizontal="left", vertical="center"
            )
            worksheet.row_dimensions[row].height = 24

    # In-cell dashboard tables: status and severity summaries.
    status_buckets = {
        "200 OK": 0,
        "3xx Redirects": 0,
        "4xx Errors": 0,
        "5xx Errors": 0,
        "Other": 0,
    }
    pass_urls = 0
    critical_urls = 0
    warning_urls = 0
    avg_ttfb_ms = 0.0
    if "Technical" in writer.book.sheetnames:
        tech_ws = writer.book["Technical"]
        tech_headers = _header_index(tech_ws)
        sc_col = tech_headers.get("Status Code")
        sev_col = tech_headers.get("Severity Badge")
        ttfb_col = tech_headers.get("TTFB (ms)")
        schema_count_col = tech_headers.get("Schema Types Count")
        broken_links_col = tech_headers.get("Broken Internal Links Count")
        ttfb_values: list[float] = []
        schema_urls = 0
        broken_links_total = 0
        if sc_col:
            for r in range(2, tech_ws.max_row + 1):
                code = _to_int(tech_ws.cell(row=r, column=sc_col).value, 0)
                if 200 <= code < 300:
                    status_buckets["200 OK"] += 1
                elif 300 <= code < 400:
                    status_buckets["3xx Redirects"] += 1
                elif 400 <= code < 500:
                    status_buckets["4xx Errors"] += 1
                elif 500 <= code < 600:
                    status_buckets["5xx Errors"] += 1
                elif code:
                    status_buckets["Other"] += 1
                if sev_col:
                    sev_val = (
                        str(tech_ws.cell(row=r, column=sev_col).value or "")
                        .strip()
                        .lower()
                    )
                    if sev_val == "critical":
                        critical_urls += 1
                    elif sev_val == "warning":
                        warning_urls += 1
                if ttfb_col:
                    raw_ttfb = tech_ws.cell(row=r, column=ttfb_col).value
                    try:
                        if raw_ttfb is not None and str(raw_ttfb).strip() != "":
                            ttfb_values.append(float(raw_ttfb))
                    except Exception:
                        pass
                if schema_count_col:
                    schema_urls += (
                        1
                        if _to_int(tech_ws.cell(row=r, column=schema_count_col).value, 0) > 0
                        else 0
                    )
                if broken_links_col:
                    broken_links_total += _to_int(
                        tech_ws.cell(row=r, column=broken_links_col).value, 0
                    )
        if ttfb_values:
            avg_ttfb_ms = round(sum(ttfb_values) / len(ttfb_values), 2)
        crit_col = tech_headers.get("Critical Issues Count")
        warn_col = tech_headers.get("Warning Issues Count")
        if crit_col and warn_col:
            for r in range(2, tech_ws.max_row + 1):
                crit_count = _to_int(tech_ws.cell(row=r, column=crit_col).value, 0)
                warn_count = _to_int(tech_ws.cell(row=r, column=warn_col).value, 0)
                # Treat Observation-only URLs as pass; failing means critical/warning exists.
                if crit_count == 0 and warn_count == 0:
                    pass_urls += 1
        else:
            sev_badge_col = tech_headers.get("Severity Badge")
            if sev_badge_col:
                for r in range(2, tech_ws.max_row + 1):
                    sev = (
                        str(tech_ws.cell(row=r, column=sev_badge_col).value or "")
                        .strip()
                        .lower()
                    )
                    if sev in {"pass", "info"}:
                        pass_urls += 1

    severity_counts = Counter()
    if "FixPlan" in writer.book.sheetnames:
        fix_ws = writer.book["FixPlan"]
        fix_headers = _header_index(fix_ws)
        sev_col = fix_headers.get("Severity")
        if sev_col:
            for r in range(2, fix_ws.max_row + 1):
                sev = str(fix_ws.cell(r, sev_col).value or "").strip().lower()
                if sev == "critical":
                    severity_counts["Critical"] += 1
                elif sev in {"high", "warning"}:
                    severity_counts["High"] += 1
                elif sev == "medium":
                    severity_counts["Medium"] += 1
                else:
                    severity_counts["Low"] += 1

    # Real KPI formulas based on parsed status buckets and available score columns.
    status_total = sum(status_buckets.values())
    error_count = status_buckets["4xx Errors"] + status_buckets["5xx Errors"]
    success_count = status_buckets["200 OK"]
    crawl_denominator = max(1, status_total or total_urls)
    pass_rate_pct = round((pass_urls / crawl_denominator) * 100, 2)
    critical_rate_pct = round((critical_urls / crawl_denominator) * 100, 2)
    warning_rate_pct = round((warning_urls / crawl_denominator) * 100, 2)
    worksheet["B8"] = pass_urls
    worksheet["B9"] = critical_urls
    worksheet["B10"] = warning_urls
    worksheet["B18"] = schema_urls if "schema_urls" in locals() else 0
    worksheet["B19"] = broken_links_total if "broken_links_total" in locals() else 0
    worksheet["B11"] = f"=IFERROR({error_count}/{crawl_denominator},0)"
    worksheet["B11"].number_format = "0.00%"
    worksheet["B12"] = f"=IFERROR({success_count}/{crawl_denominator},0)"
    worksheet["B12"].number_format = "0.00%"
    worksheet["B13"] = f"=IFERROR({critical_urls}/{crawl_denominator},0)"
    worksheet["B13"].number_format = "0.00%"
    worksheet["B14"] = f"=IFERROR({warning_urls}/{crawl_denominator},0)"
    worksheet["B14"].number_format = "0.00%"
    worksheet["B7"] = f"=IFERROR({pass_urls}/{crawl_denominator},0)"
    worksheet["B7"].number_format = "0.00%"
    # Keep labels readable with left alignment and breathing room.
    for row in range(5, 15):
        worksheet[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[row].height = 24

    avg_health_score = None
    if "Technical" in writer.book.sheetnames:
        technical_ws = writer.book["Technical"]
        technical_headers = _header_index(technical_ws)
        score_col = technical_headers.get("SEO Health Score")
        if score_col:
            score_values: list[float] = []
            for r in range(2, technical_ws.max_row + 1):
                raw = technical_ws.cell(row=r, column=score_col).value
                try:
                    if raw is not None and str(raw).strip() != "":
                        score_values.append(float(raw))
                except Exception:
                    pass
            if score_values:
                avg_health_score = sum(score_values) / len(score_values)
    if avg_health_score is not None:
        worksheet["B6"] = f"=IFERROR({round(avg_health_score, 2)}/100,0)"
    else:
        worksheet["B6"] = f"=IFERROR({pass_rate_pct}/100,0)"
    worksheet["B6"].number_format = "0.00%"

    table_header_fill = PatternFill("solid", fgColor="ADD8E6")
    table_header_font = Font(color="000000", bold=True, size=11)
    for ref, val in {
        "D4": "Status",
        "E4": "Count",
        "D12": "Severity",
        "E12": "Count",
    }.items():
        cell = worksheet[ref]
        cell.value = val
        cell.fill = table_header_fill
        cell.font = table_header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    status_rows = [
        ("200 OK", status_buckets["200 OK"], "C6EFCE"),
        ("3xx Redirects", status_buckets["3xx Redirects"], "FFCC99"),
        ("4xx Errors", status_buckets["4xx Errors"], "FFC1C1"),
        ("5xx Errors", status_buckets["5xx Errors"], "FFC1C1"),
        ("Other", status_buckets["Other"], "FFCC99"),
    ]
    for idx, (label, count, color) in enumerate(status_rows, start=5):
        worksheet[f"D{idx}"] = label
        worksheet[f"E{idx}"] = count
        worksheet[f"D{idx}"].fill = PatternFill("solid", fgColor=color)
        worksheet[f"E{idx}"].fill = PatternFill("solid", fgColor=color)

    sev_rows = [
        ("Critical", severity_counts.get("Critical", 0), "FFC1C1"),
        ("Warning", severity_counts.get("High", 0), "FFCC99"),
        ("Medium", severity_counts.get("Medium", 0), "FFCC99"),
        ("Low", severity_counts.get("Low", 0), "C6EFCE"),
    ]
    for idx, (label, count, color) in enumerate(sev_rows, start=13):
        worksheet[f"D{idx}"] = label
        worksheet[f"E{idx}"] = count
        worksheet[f"D{idx}"].fill = PatternFill("solid", fgColor=color)
        worksheet[f"E{idx}"].fill = PatternFill("solid", fgColor=color)

    for row in range(5, 16):
        for col in ("D", "E"):
            worksheet[f"{col}{row}"].alignment = Alignment(
                horizontal="center", vertical="center"
            )

    # High-priority operational snapshot.
    worksheet["J4"] = "PRIORITY SNAPSHOT"
    worksheet["K4"] = "Value"
    for ref in ("J4", "K4"):
        worksheet[ref].fill = table_header_fill
        worksheet[ref].font = table_header_font
        worksheet[ref].alignment = Alignment(horizontal="center", vertical="center")
    top_issue_name = "N/A"
    top_issue_affected = 0
    owner_rollup: dict[str, dict[str, int]] = defaultdict(
        lambda: {
            "issue_rows": 0,
            "affected_urls": 0,
            "critical": 0,
            "warning": 0,
            "info": 0,
        }
    )
    if "FixPlan" in writer.book.sheetnames:
        fix_ws = writer.book["FixPlan"]
        fix_headers = _header_index(fix_ws)
        issue_col = fix_headers.get("Issue Type")
        affected_col = fix_headers.get("Affected Count")
        owner_col = fix_headers.get("Owner")
        sev_col = fix_headers.get("Severity")
        if issue_col and affected_col:
            for r in range(2, fix_ws.max_row + 1):
                affected = _to_int(fix_ws.cell(row=r, column=affected_col).value, 0)
                if affected > top_issue_affected:
                    top_issue_affected = affected
                    top_issue_name = str(
                        fix_ws.cell(row=r, column=issue_col).value or "N/A"
                    )
                owner_name = (
                    str(
                        fix_ws.cell(row=r, column=owner_col).value or "Unassigned"
                    ).strip()
                    if owner_col
                    else "Unassigned"
                )
                sev_val = (
                    str(fix_ws.cell(row=r, column=sev_col).value or "").strip().lower()
                    if sev_col
                    else ""
                )
                owner_rollup[owner_name]["issue_rows"] += 1
                owner_rollup[owner_name]["affected_urls"] += affected
                if sev_val == "critical":
                    owner_rollup[owner_name]["critical"] += 1
                elif sev_val in {"warning", "high", "medium"}:
                    owner_rollup[owner_name]["warning"] += 1
                else:
                    owner_rollup[owner_name]["info"] += 1
    worksheet["J5"] = "Top Blocking Issue"
    worksheet["K5"] = top_issue_name
    worksheet["J6"] = "Top Issue Affected URLs"
    worksheet["K6"] = top_issue_affected
    worksheet["J7"] = "4xx/5xx URLs"
    worksheet["K7"] = status_buckets["4xx Errors"] + status_buckets["5xx Errors"]
    worksheet["J8"] = "Avg TTFB (ms)"
    worksheet["K8"] = avg_ttfb_ms
    for row in range(5, 9):
        worksheet[f"J{row}"].fill = PatternFill("solid", fgColor="F5F7FA")
        worksheet[f"K{row}"].fill = PatternFill("solid", fgColor="F5F7FA")

    # Pivot-style owner summary block.
    worksheet["G4"] = "OWNER ISSUE SUMMARY"
    worksheet["G5"] = "Owner"
    worksheet["H5"] = "Issue Rows"
    worksheet["I5"] = "Affected URLs"
    worksheet["J5"] = "Critical"
    worksheet["K5"] = "Warning"
    for ref in ("G4",):
        worksheet[ref].fill = light_header_fill
        worksheet[ref].font = table_header_font
        worksheet[ref].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    owner_header_fill = PatternFill("solid", fgColor="ADD8E6")
    for ref in ("G5", "H5", "I5", "J5", "K5"):
        worksheet[ref].fill = owner_header_fill
        worksheet[ref].font = Font(color="000000", bold=True, size=11)
        worksheet[ref].alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    worksheet.merge_cells("G4:K4")

    owner_rows_sorted = sorted(
        owner_rollup.items(),
        key=lambda x: (
            -x[1]["affected_urls"],
            -x[1]["critical"],
            -x[1]["warning"],
            x[0],
        ),
    )
    owner_start_row = 6
    for idx, (owner_name, metrics) in enumerate(
        owner_rows_sorted[:8], start=owner_start_row
    ):
        worksheet[f"G{idx}"] = f'=HYPERLINK("#\'FixPlan\'!A1","{owner_name}")'
        worksheet[f"H{idx}"] = metrics["issue_rows"]
        worksheet[f"I{idx}"] = metrics["affected_urls"]
        worksheet[f"J{idx}"] = metrics["critical"]
        worksheet[f"K{idx}"] = metrics["warning"]
        for col in ("G", "H", "I", "J", "K"):
            worksheet[f"{col}{idx}"].fill = PatternFill("solid", fgColor="F5F7FA")
            worksheet[f"{col}{idx}"].alignment = Alignment(
                horizontal="center", vertical="center"
            )
    if not owner_rows_sorted:
        worksheet["G6"] = "No owner data"
        worksheet["G6"].fill = PatternFill("solid", fgColor="F5F7FA")
        worksheet["G6"].alignment = Alignment(horizontal="center", vertical="center")

    for col, width in {"G": 15, "H": 15, "I": 15, "J": 15, "K": 15}.items():
        worksheet.column_dimensions[col].width = width

    # KPI emphasis colors.
    overall_health = float(avg_health_score or pass_rate_pct)
    health_fill = (
        "C6EFCE"
        if overall_health >= 80
        else "FFEB9C" if overall_health >= 60 else "FFC7CE"
    )
    worksheet["B6"].fill = PatternFill("solid", fgColor=health_fill)
    worksheet["B7"].fill = PatternFill(
        "solid",
        fgColor=(
            "C6EFCE"
            if pass_rate_pct >= 70
            else "FFEB9C" if pass_rate_pct >= 40 else "FFC7CE"
        ),
    )
    worksheet["B11"].fill = PatternFill(
        "solid", fgColor="FFC7CE" if error_count > 0 else "C6EFCE"
    )
    worksheet["B12"].fill = PatternFill(
        "solid",
        fgColor=(
            "C6EFCE" if success_count >= max(1, crawl_denominator * 0.9) else "FFEB9C"
        ),
    )
    worksheet["B13"].fill = PatternFill(
        "solid",
        fgColor=(
            "FFC7CE"
            if critical_rate_pct >= 10
            else "FFEB9C" if critical_rate_pct > 0 else "C6EFCE"
        ),
    )
    worksheet["B14"].fill = PatternFill(
        "solid",
        fgColor=(
            "FFC7CE"
            if warning_rate_pct >= 50
            else "FFEB9C" if warning_rate_pct > 20 else "C6EFCE"
        ),
    )
    projected_pass_rate_pct = min(
        100.0,
        pass_rate_pct
        + ((critical_urls * 1.0 + warning_urls * 0.75) / max(1, crawl_denominator))
        * 100.0,
    )
    projected_health_pct = min(
        100.0,
        (overall_health if overall_health <= 100 else 100.0)
        + ((100.0 - overall_health) * 0.6),
    )
    worksheet["B15"] = projected_health_pct / 100.0
    worksheet["B15"].number_format = "0.00%"
    worksheet["B16"] = projected_pass_rate_pct / 100.0
    worksheet["B16"].number_format = "0.00%"
    worksheet["B15"].fill = PatternFill(
        "solid",
        fgColor=(
            "C6EFCE"
            if projected_health_pct >= 80
            else "FFEB9C" if projected_health_pct >= 60 else "FFC7CE"
        ),
    )
    worksheet["B16"].fill = PatternFill(
        "solid",
        fgColor=(
            "C6EFCE"
            if projected_pass_rate_pct >= 70
            else "FFEB9C" if projected_pass_rate_pct >= 40 else "FFC7CE"
        ),
    )
    worksheet["B17"].fill = PatternFill("solid", fgColor="C6EFCE")

    # Top issue list with direct links.
    worksheet["B44"] = "TOP ISSUES TO FIX FIRST"
    worksheet["C44"] = "Affected URLs"
    for ref in ("B44", "C44"):
        worksheet[ref].fill = table_header_fill
        worksheet[ref].font = table_header_font
        worksheet[ref].alignment = Alignment(horizontal="center", vertical="center")
    if "FixPlan" in writer.book.sheetnames:
        fix_ws = writer.book["FixPlan"]
        fix_headers = _header_index(fix_ws)
        issue_col = fix_headers.get("Issue Type")
        affected_col = fix_headers.get("Affected Count")
        priority_col = fix_headers.get("Priority Score")
        top_rows: list[tuple[str, int, int, int]] = []
        if issue_col and affected_col:
            for r in range(2, fix_ws.max_row + 1):
                issue_name = str(
                    fix_ws.cell(row=r, column=issue_col).value or ""
                ).strip()
                affected = _to_int(fix_ws.cell(row=r, column=affected_col).value, 0)
                priority = (
                    _to_int(fix_ws.cell(row=r, column=priority_col).value, 0)
                    if priority_col
                    else 0
                )
                if issue_name:
                    top_rows.append((issue_name, affected, priority, r))
        top_rows.sort(key=lambda x: (-x[1], -x[2], x[0]))
        for idx, (issue_name, affected, _priority, source_row) in enumerate(
            top_rows[:5], start=45
        ):
            worksheet[f"B{idx}"] = (
                f'=HYPERLINK("#FixPlan!A{source_row}","{issue_name}")'
            )
            worksheet[f"C{idx}"] = affected
            worksheet[f"B{idx}"].fill = PatternFill("solid", fgColor="F5F7FA")
            worksheet[f"C{idx}"].fill = PatternFill("solid", fgColor="F5F7FA")
            worksheet[f"B{idx}"].font = Font(
                color=STD_BLUE, underline="single", bold=True
            )

    # Quick navigation block.
    quick_nav_fill = PatternFill("solid", fgColor=STD_NAVY)
    worksheet["J12"] = "Quick Navigation"
    worksheet["K12"] = "Open"
    for ref in ("J12", "K12"):
        worksheet[ref].fill = quick_nav_fill
        worksheet[ref].font = Font(color="000000", bold=True, size=11)
        worksheet[ref].alignment = Alignment(horizontal="center", vertical="center")
    quick_links = [
        ("Fix Plan", "#FixPlan!A1"),
        ("Main URL Data", "#Main!A1"),
        ("Technical Diagnostics", "#Technical!A1"),
        ("Indexability", "#Indexability!A1"),
        ("AEO Opportunities", "#AEO!A1"),
    ]
    quick_links.append(("AIOSEO Action Queue", "#AIOSEO!A1"))
    for idx, (label, target) in enumerate(quick_links, start=13):
        worksheet[f"J{idx}"] = label
        worksheet[f"K{idx}"] = f'=HYPERLINK("{target}","Open")'
        worksheet[f"K{idx}"].font = Font(color=STD_BLUE, underline="single", bold=True)
        worksheet[f"J{idx}"].fill = PatternFill("solid", fgColor="F5F7FA")
        worksheet[f"K{idx}"].fill = PatternFill("solid", fgColor="F5F7FA")

    worksheet["J20"] = "BUSINESS IMPACT SUMMARY"
    worksheet["J21"] = (
        f"{status_buckets['4xx Errors'] + status_buckets['5xx Errors']} error URLs detected "
        f"({status_buckets['4xx Errors']} 4xx / {status_buckets['5xx Errors']} 5xx). "
        f"Critical issue volume is {critical_urls} URLs and warning volume is {warning_urls}. "
        f"Top blocker: {top_issue_name} affecting {top_issue_affected} URLs."
    )
    worksheet.merge_cells("J20:K20")
    worksheet.merge_cells("J21:K24")
    worksheet["J20"].fill = table_header_fill
    worksheet["J20"].font = table_header_font
    worksheet["J20"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["J21"].fill = PatternFill("solid", fgColor="F5F7FA")
    worksheet["J21"].alignment = Alignment(
        horizontal="left", vertical="top", wrap_text=True
    )

    # Traditional SEO vs AEO Readiness block.
    worksheet["E20"] = "Traditional SEO vs. 2026 AEO Readiness"
    worksheet["E21"] = "Dimension"
    worksheet["F21"] = "Score"
    for ref in ("E20", "E21", "F21"):
        worksheet[ref].fill = table_header_fill
        worksheet[ref].font = table_header_font
        worksheet[ref].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.merge_cells("E20:F20")

    traditional_score = max(
        0.0, min(100.0, (success_count / max(1, crawl_denominator)) * 100.0)
    )
    if avg_health_score is not None:
        traditional_score = round(
            (traditional_score * 0.4) + (avg_health_score * 0.6), 2
        )
    aeo_score_values: list[float] = []
    if "AEO" in writer.book.sheetnames:
        aeo_ws = writer.book["AEO"]
        aeo_headers = _header_index(aeo_ws)
        aeo_col = aeo_headers.get("AEO Readiness Score")
        if aeo_col:
            for r in range(2, aeo_ws.max_row + 1):
                try:
                    raw = aeo_ws.cell(row=r, column=aeo_col).value
                    if raw is not None and str(raw).strip() != "":
                        aeo_score_values.append(float(raw))
                except Exception:
                    pass
    aeo_readiness = (
        round(sum(aeo_score_values) / len(aeo_score_values), 2)
        if aeo_score_values
        else 0.0
    )
    worksheet["E22"] = "Traditional SEO"
    worksheet["F22"] = traditional_score / 100.0
    worksheet["F22"].number_format = "0.00%"
    worksheet["E23"] = "2026 AEO Readiness"
    worksheet["F23"] = aeo_readiness / 100.0
    worksheet["F23"].number_format = "0.00%"
    for ref in ("E22", "F22", "E23", "F23"):
        worksheet[ref].fill = PatternFill("solid", fgColor="F5F7FA")
        worksheet[ref].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["E24"] = "Strategic Narrative"
    worksheet["E25"] = (
        "High SEO / Low AEO suggests the site is visible to humans but invisible to AI answer engines."
        if traditional_score >= 70 and aeo_readiness < 60
        else "SEO and AEO signals are moving together; continue balancing crawl health with answer-first content."
    )
    worksheet.merge_cells("E25:F27")
    worksheet["E24"].fill = table_header_fill
    worksheet["E24"].font = table_header_font
    worksheet["E24"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["E25"].fill = PatternFill("solid", fgColor="F5F7FA")
    worksheet["E25"].alignment = Alignment(
        horizontal="left", vertical="top", wrap_text=True
    )

    # Dashboard KPI tooltips on result cells.
    dashboard_tooltips = {
        "C5": "Total URLs crawled in this run. Calculated as the number of audited URL rows.",
        "C6": "Overall Health Score. Calculated as average SEO Health Score across Technical URLs; fallback to SEO Pass Rate if score data is unavailable.",
        "C7": "SEO Pass Rate %. Calculated as Pass URLs divided by Total URLs.",
        "C8": "Pass URL count. URL is pass when it has no Critical and no Warning issues.",
        "C9": "Critical URL count from Technical severity badge.",
        "C10": "Warning URL count from Technical severity badge.",
        "C11": "HTTP Error Rate %. Calculated as (4xx URLs + 5xx URLs) / Total URLs.",
        "C12": "Crawl Success Rate %. Calculated as 2xx URLs / Total URLs.",
        "C13": "Critical URL Rate %. Calculated as Critical URLs / Total URLs.",
        "C14": "Warning URL Rate %. Calculated as Warning URLs / Total URLs.",
        "C15": "Projected Health Score if all current To Do items are completed in this cycle.",
        "C16": "Projected Pass Rate if all current To Do items are completed in this cycle.",
        "C17": "Content Hub Readiness %. Calculated as completed content items divided by total tracked content URLs.",
        "C32": "Most widespread issue from FixPlan by affected URL count.",
        "C33": "Number of URLs impacted by the top blocking issue.",
        "C34": "Total URLs returning client/server errors (4xx + 5xx).",
        "C35": "Average Time to First Byte across Technical URLs (ms).",
        "C45": "Affected URL count for the highest-priority issue (linked to FixPlan).",
        "C46": "Affected URL count for the next issue in the priority list.",
        "C47": "Affected URL count for the next issue in the priority list.",
        "C48": "Affected URL count for the next issue in the priority list.",
        "C49": "Affected URL count for the next issue in the priority list.",
        "E5": "Owner responsible for remediation. Click to open FixPlan.",
        "F5": "Number of issue rows assigned to this owner.",
        "G5": "Total affected URLs across this owner's assigned issues.",
        "H5": "Count of critical issue types assigned to this owner.",
        "I5": "Count of warning issue types assigned to this owner.",
        "J5": "Count of info issue types assigned to this owner.",
    }
    for ref, message in dashboard_tooltips.items():
        dv = DataValidation(
            type="custom", formula1="TRUE", showInputMessage=True, allow_blank=True
        )
        dv.promptTitle = f"KPI {ref}"
        dv.prompt = message
        worksheet.add_data_validation(dv)
        dv.add(ref)
    for row_idx in range(4, 60):
        worksheet.row_dimensions[row_idx].height = 20
    # Ensure all dashboard header cells wrap for 1080p readability.
    for row_idx in range(1, 70):
        for col_idx in range(1, 12):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value and cell.font and cell.font.bold:
                cell.alignment = Alignment(
                    horizontal=(
                        cell.alignment.horizontal if cell.alignment else "center"
                    ),
                    vertical=cell.alignment.vertical if cell.alignment else "center",
                    wrap_text=True,
                )


def _reorder_columns(worksheet, sheet_name: str) -> None:
    preferred_orders = {
        "Main": [
            "Health Icon",
            "URL",
            "Status Code",
            "Indexability",
            "Load Time (s)",
            "Title",
            "Meta Description",
            "Word Count (Body)",
            "SEO Health Score",
            "Severity Badge",
            "Action Needed",
            "Owner",
            "Status",
            "Sprint",
        ],
        "Technical": [
            "URL",
            "Status Code",
            "Status Class",
            "SEO Health Score",
            "Severity Badge",
            "Action Needed",
            "Indexability Reason",
            "TTFB (ms)",
            "Total Request Time (ms)",
            "Final URL",
            "Canonical URL",
            "Canonical Type",
            "Owner",
            "Status",
            "Sprint",
        ],
        "FixPlan": [
            "Issue Type",
            "Severity",
            "Priority Score",
            "Affected Count",
            "Affected URLs",
            "Detail Reference Tab",
            "Resolution Type",
            "URL",
            "Recommended Fix",
            "Likely Root Cause",
            "Owner",
            "Agency Owner",
            "Effort",
            "Est. Hours",
            "Est. Sprint Points",
            "Aging/Priority",
            "Status",
            "Verified By",
            "Date Resolved",
            "Revenue Risk",
            "Action Needed",
            "Jump to Details",
            "View Details",
            "Sprint",
        ],
        "Summary": [
            "Section",
            "Severity",
            "Issue",
            "Affected URL Count",
            "Reference Tab",
            "Affected URLs (sample)",
        ],
        "Priority URLs": [
            "URL",
            "Business Risk Score",
            "Severity Badge",
            "SEO Health Score",
            "GSC Impressions",
            "GSC CTR",
            "Revenue Intent",
            "Critical Issues Count",
            "Warning Issues Count",
            "Action Needed",
            "Why Prioritized",
            "Owner",
            "Status",
            "Sprint",
        ],
        "Content": [
            "URL",
            "Word Count",
            "Word Count Band",
            "Readability (Rough Flesch)",
            "H1 Count",
            "Missing H1 Flag",
            "Multiple H1 Flag",
            "Title Missing",
            "Meta Description Missing",
            "Thin Content Flag",
        ],
        "Links": [
            "URL",
            "Internal Links Count",
            "Unique Internal Links Count",
            "Broken Internal Links Count",
            "Unresolved Internal Links Count",
            "Generic Anchor Text Count",
            "External Links Count",
            "Nofollow Internal Links Count",
            "Nofollow External Links Count",
            "Internal Link Statuses",
        ],
        "AIOSEO": [
            "URL",
            "WordPress Post ID",
            "Direct Edit Link",
            "AIOSEO Panel",
            "Severity",
            "Issue",
            "Priority Score",
            "Current Value",
            "Recommended Target",
            "How to Fix in AIOSEO",
            "Reference Tab",
            "Reference Field",
            "Action Needed",
            "Owner",
            "Status",
            "Est. Hours",
            "Stable Issue ID",
        ],
        "Content Optimization Hub": [
            "Action Required",
            "Status",
            "Assigned Owner",
            "Content Cluster ID",
            "URL",
            "Current SEO Score",
            "Projected SEO Score",
            "Elementor Builder Link",
            "Target Keywords",
            "Current Page Copy Snippet",
            "Current Title",
            "Proposed Title (50-60 Chars)",
            "Title Count",
            "Current Meta Desc",
            "Proposed Meta Desc (120-160 Chars)",
            "Desc Count",
            "Current H-Tag Structure",
            "Proposed H-Tag Fixes",
            "AEO Answer Block Draft",
            "FAQ/QA Draft",
            "Current OG-Image URL",
            "OG Image Preview",
            "Social Share Note",
        ],
        "AEO": [
            "URL",
            "AEO Readiness Score",
            "AEO Badge",
            "Why It Matters",
            "FAQ Section Count",
            "Question Heading Count",
            "Paragraphs 40-60 Words Count",
            "QAPage/FAQ Schema Present",
            "Speakable Schema Present",
            "HowTo Signal",
            "Definition Signal",
            "List/Table Answer Signal",
        ],
        "Indexability": [
            "URL",
            "Status Code",
            "Status Class",
            "Indexability Reason",
            "Canonical URL",
            "Canonical Type",
            "Canonical Matches Final URL",
            "Canonical in Sitemap Match",
            "Meta Robots Raw",
            "X-Robots-Tag",
            "Final URL",
        ],
        "IssueInventory": [
            "URL",
            "Issue",
            "Severity",
            "Reference Tab",
            "Stable Issue ID",
            "Owner",
            "Status",
            "Sprint",
            "Open in Main",
            "Open in Reference",
        ],
    }
    preferred = preferred_orders.get(sheet_name)
    if not preferred or worksheet.max_row < 1:
        return
    current_headers = [
        worksheet.cell(row=1, column=i).value
        for i in range(1, worksheet.max_column + 1)
    ]
    if not any(h in current_headers for h in preferred):
        return
    ordered_headers = [h for h in preferred if h in current_headers]
    ordered_headers.extend([h for h in current_headers if h not in ordered_headers])
    if ordered_headers == current_headers:
        return
    idx_map = [current_headers.index(h) + 1 for h in ordered_headers]
    rows = []
    for row_idx in range(1, worksheet.max_row + 1):
        rows.append(
            [worksheet.cell(row=row_idx, column=src_col).value for src_col in idx_map]
        )
    for row_idx, row_vals in enumerate(rows, start=1):
        for col_idx, val in enumerate(row_vals, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=val)


def _collapse_technical_deep_dive_columns(worksheet, sheet_name: str) -> None:
    _collapse_technical_deep_dive_columns_impl(
        worksheet, sheet_name, header_index_fn=_header_index
    )


def _apply_cross_sheet_links(writer, worksheet, sheet_name: str) -> None:
    _apply_cross_sheet_links_impl(
        writer,
        worksheet,
        sheet_name,
        debug_excel_isolation_mode=DEBUG_EXCEL_ISOLATION_MODE,
        header_index_fn=_header_index,
    )


def _add_header_tooltips(worksheet) -> None:
    _add_schema_header_tooltips_impl(
        worksheet,
        disable_data_validation=DISABLE_DATA_VALIDATION,
        header_index_fn=_header_index,
    )


def _apply_wrapped_row_heights(worksheet) -> None:
    headers = _header_index(worksheet)
    wrapped_cols = []
    for header_name in (
        "URL",
        "Final URL",
        "Canonical URL",
        "Affected URLs",
        "Internal Link Statuses",
        "How to Fix in AIOSEO",
    ):
        col_idx = headers.get(header_name)
        if col_idx:
            wrapped_cols.append(col_idx)
    if not wrapped_cols:
        return

    for row_idx in range(2, worksheet.max_row + 1):
        max_lines = 1
        for col_idx in wrapped_cols:
            value = worksheet.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            text = str(value)
            explicit_lines = text.count("\n") + 1
            estimated_wrap_lines = max(1, int(len(text) / 50) + 1)
            max_lines = max(max_lines, explicit_lines, estimated_wrap_lines)
        if max_lines > 1:
            worksheet.row_dimensions[row_idx].height = min(120, 15 * max_lines)


def _apply_south_african_formats(worksheet) -> None:
    headers = _header_index(worksheet)
    percent_headers = {
        "Pass Rate (%)",
        "SEO Pass Rate %",
        "Error Rate % (4xx/5xx)",
        "Crawl Success Rate % (2xx)",
        "Critical URL Rate %",
        "Warning URL Rate %",
        "Pass Rate",
    }
    integer_headers = {
        "URLs Crawled",
        "Value",
        "Critical URL Count",
        "Warning URL Count",
        "Pass URLs",
        "Critical URLs",
        "Warning URLs",
        "Top Issue Affected URLs",
        "Affected Count",
        "Priority Score",
        "Est. Sprint Points",
    }
    decimal_headers = {
        "SEO Health Score",
        "AEO Readiness Score",
        "TTFB (ms)",
        "Total Request Time (ms)",
        "Avg TTFB (ms)",
    }
    date_like_tokens = ("date", "timestamp", "lastmod", "updated")

    for col_idx, cell in enumerate(worksheet[1], start=1):
        header = str(cell.value or "").strip()
        col_letter = get_column_letter(col_idx)
        rng_start = 2
        rng_end = worksheet.max_row
        if rng_end < rng_start:
            continue
        if header in percent_headers or "%" in header:
            fmt = "[$-en-ZA]0.00%"
        elif header in decimal_headers:
            fmt = "[$-en-ZA]#,##0.00"
        elif header in integer_headers:
            fmt = "[$-en-ZA]#,##0"
        elif any(token in header.lower() for token in date_like_tokens):
            fmt = "[$-en-ZA]dd/mm/yyyy hh:mm:ss"
        else:
            continue
        for row_idx in range(rng_start, rng_end + 1):
            worksheet.cell(row=row_idx, column=col_idx).number_format = fmt


def adjust_sheet_format(writer, sheet_name):
    worksheet = writer.sheets[sheet_name]
    _reorder_columns(worksheet, sheet_name)
    header_fill = PatternFill(
        start_color=STD_NAVY, end_color=STD_NAVY, fill_type="solid"
    )
    header_font = Font(color=STD_WHITE, bold=True)
    bad_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    warn_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    good_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    traffic_warn_fill = PatternFill(
        start_color="F4B183", end_color="F4B183", fill_type="solid"
    )
    edge_fill = PatternFill(start_color="D9D2E9", end_color="D9D2E9", fill_type="solid")
    zebra_fill = PatternFill(
        start_color="F7F7F7", end_color="F7F7F7", fill_type="solid"
    )
    headers = [cell.value for cell in worksheet[1]]
    if sheet_name in {"FixPlan", "Main", "Technical", "AIOSEO"}:
        _apply_intelligent_sorting(worksheet, sheet_name)
    # Apply table view settings near the end after all sheet mutations.
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    is_wide_sheet = worksheet.max_column >= 30

    def parse_bool(value):
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.strip().lower() in {"true", "1", "yes", "y"}
        return bool(value)

    def is_bad_header(h):
        h = (h or "").lower()
        return any(
            t in h
            for t in [
                "error",
                "broken",
                "missing",
                "noindex",
                "disallow",
                "thin",
                "mixed content",
                "cross-canonical",
                "issue",
                "non-200",
                "loop",
                "out of",
            ]
        )

    def is_edge_header(h):
        h = (h or "").lower()
        return any(
            t in h for t in ["redirect chain", "param url", "edge", "unresolved"]
        )

    def is_good_header(h):
        h = (h or "").lower()
        return any(
            t in h
            for t in [
                "accessible",
                "match",
                "enabled",
                "complete",
                "present",
                "indexable",
                "coverage (%)",
            ]
        )

    for row_idx in range(2, worksheet.max_row + 1):
        row_has_issue = False
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            val = cell.value
            h = str(header) if header is not None else ""
            is_url_like = (
                "url" in h.lower()
                or h.lower().endswith("urls")
                or h.lower() in {"final url", "canonical url"}
            )
            if is_wide_sheet and is_url_like:
                cell.alignment = Alignment(
                    wrap_text=False, shrink_to_fit=True, vertical="top"
                )
            elif "url" in h.lower() or "hops" in h.lower() or "images" == h.lower():
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")
            if "%" in h:
                try:
                    pct = float(val)
                    if pct == 100.0:
                        cell.fill = good_fill
                    elif pct < 80:
                        cell.fill = bad_fill
                        row_has_issue = True
                    else:
                        cell.fill = warn_fill
                except Exception:
                    pass
            if h in {"Status Code", "Target Status (if crawled)"} and isinstance(
                val, int
            ):
                if val >= 400:
                    cell.fill = bad_fill
                    row_has_issue = True
                elif val >= 300:
                    cell.fill = warn_fill
                elif 200 <= val < 300:
                    cell.fill = good_fill
            if isinstance(val, bool) or (
                isinstance(val, str) and val.strip().lower() in {"true", "false"}
            ):
                flag = parse_bool(val)
                if is_bad_header(h):
                    cell.fill = bad_fill if flag else good_fill
                    row_has_issue = row_has_issue or flag
                elif is_good_header(h):
                    cell.fill = good_fill if flag else warn_fill
                    row_has_issue = row_has_issue or (not flag)
                elif is_edge_header(h) and flag:
                    cell.fill = edge_fill
            if h in {
                "Broken Internal Links Count",
                "Image Filename Quality Issues",
                "Generic Anchor Text Count",
            }:
                try:
                    if int(val or 0) > 0:
                        cell.fill = bad_fill
                        row_has_issue = True
                    else:
                        cell.fill = good_fill
                except Exception:
                    pass
            if h in {"Word Count Band"} and isinstance(val, str):
                band = val.lower()
                if band == "thin":
                    cell.fill = bad_fill
                    row_has_issue = True
                elif band == "ok":
                    cell.fill = warn_fill
                elif band == "strong":
                    cell.fill = good_fill
            if h in {"Indexability Reason"} and isinstance(val, str):
                if "indexable" in val.lower() and "noindex" not in val.lower():
                    cell.fill = good_fill
                else:
                    cell.fill = bad_fill
                    row_has_issue = True
            if h in {"Severity", "Severity Badge"} and isinstance(val, str):
                sev = val.strip().lower()
                if sev == "critical":
                    cell.fill = bad_fill
                    row_has_issue = True
                elif sev == "warning":
                    cell.fill = traffic_warn_fill
                    row_has_issue = True
                elif sev in {"info", "observation"}:
                    cell.fill = edge_fill
                elif sev == "pass":
                    cell.fill = good_fill
            if h in {"SEO Health Score", "SEO Score", "Technical Health", "Copy Score"}:
                try:
                    score = float(val)
                    if score < 70:
                        cell.fill = bad_fill
                        row_has_issue = True
                    elif score < 90:
                        cell.fill = traffic_warn_fill
                        row_has_issue = True
                    else:
                        cell.fill = good_fill
                except Exception:
                    pass
            if h == "Status" and isinstance(val, str):
                st = val.strip().lower()
                if st in {"to do", "todo", "open"}:
                    cell.fill = bad_fill
                    row_has_issue = True
                elif st == "in progress":
                    cell.fill = traffic_warn_fill
                    row_has_issue = True
                elif st in {"fixed", "done", "closed"}:
                    cell.fill = good_fill
            if (
                h == "Direct Edit Link"
                and isinstance(val, str)
                and val.startswith(("http://", "https://"))
            ):
                if _is_safe_hyperlink_target(val):
                    cell.hyperlink = val
                    cell.style = "Hyperlink"
                    cell.fill = PatternFill(
                        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
                    )
                    cell.font = Font(color="FFFFFF", bold=True, underline="single")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            if h in {"Owner", "Agency Owner"} and isinstance(val, str):
                owner = val.strip().lower()
                if "dev" in owner:
                    cell.fill = PatternFill(
                        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
                    )
                elif "copy" in owner:
                    cell.fill = PatternFill(
                        start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"
                    )
                elif "server" in owner or "host" in owner:
                    cell.fill = PatternFill(
                        start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
                    )
        if not row_has_issue:
            worksheet.cell(row=row_idx, column=1).fill = good_fill
        if sheet_name != "Dashboard" and row_idx % 2 == 0:
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.fill.fill_type is None:
                    cell.fill = zebra_fill
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)
    # Keep URL-like columns readable and ensure wrapped text is visible.
    headers = _header_index(worksheet)
    for header_name, width in {
        "URL": 45,
        "Final URL": 45,
        "Canonical URL": 45,
        "Affected URLs": 55,
        "How to Fix in AIOSEO": 55,
    }.items():
        col_idx = headers.get(header_name)
        if col_idx:
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(
                worksheet.column_dimensions[get_column_letter(col_idx)].width or width,
                width,
            )
    if is_wide_sheet:
        for header_name in ("URL", "Final URL", "Canonical URL", "Affected URLs"):
            col_idx = headers.get(header_name)
            if col_idx:
                worksheet.column_dimensions[get_column_letter(col_idx)].width = 36
    _apply_wrapped_row_heights(worksheet)
    if sheet_name == "FixPlan":
        apply_fixplan_workflow_formatting(worksheet)
        _apply_fixplan_interactivity(writer, worksheet)
    _hide_noisy_columns(worksheet, sheet_name)
    _apply_south_african_formats(worksheet)
    if not DISABLE_CONDITIONAL_FORMATTING:
        apply_global_conditional_formatting(worksheet)
    _collapse_technical_deep_dive_columns(worksheet, sheet_name)
    _add_url_navigation_links(writer, worksheet, sheet_name)
    _apply_cross_sheet_links(writer, worksheet, sheet_name)
    if sheet_name == "AIOSEO":
        status_col = _header_index(worksheet).get("Status")
        if status_col:
            _apply_status_dropdown(worksheet, status_col)
    _add_back_to_dashboard_link(worksheet, sheet_name)
    if sheet_name == "Content Optimization Hub":
        # Insert instruction row and keep editorial controls pinned.
        worksheet.insert_rows(1)
        max_col_letter = get_column_letter(worksheet.max_column)
        worksheet.merge_cells(f"A1:{max_col_letter}1")
        worksheet["A1"] = (
            "CONTENT HUB INSTRUCTIONS: 1. Draft in 'Proposed' columns. | "
            "2. Watch 'Count' for Green. | "
            "3. Click 'Elementor' link. | "
            "4. Mark 'Status' as 'Completed'. | "
            "NOTE: If images show '#BLOCKED!', click the 'Security Warning' bar at the top of Excel and select 'Enable Content'."
        )
        worksheet["A1"].hyperlink = None
        worksheet["A1"].fill = PatternFill(
            start_color="BFE9E4", end_color="BFE9E4", fill_type="solid"
        )
        worksheet["A1"].font = Font(color=STD_NAVY, bold=True)
        worksheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
        worksheet.row_dimensions[1].height = 28
        _set_freeze_panes_safe(worksheet, "F3")
        hub_headers = {
            str(cell.value): idx
            for idx, cell in enumerate(worksheet[2], start=1)
            if cell.value
        }
        legacy_seo_col = hub_headers.get("SEO Score")
        if legacy_seo_col and "Current SEO Score" not in hub_headers:
            worksheet.cell(row=2, column=legacy_seo_col, value="Current SEO Score")
            projected_col = legacy_seo_col + 1
            worksheet.insert_cols(projected_col)
            worksheet.cell(row=2, column=projected_col, value="Projected SEO Score")
            for r in range(3, worksheet.max_row + 1):
                worksheet.cell(row=r, column=projected_col, value="")
            hub_headers = {
                str(cell.value): idx
                for idx, cell in enumerate(worksheet[2], start=1)
                if cell.value
            }
        action_required_col = hub_headers.get("Action Required")
        status_col = hub_headers.get("Status")
        has_data_rows = worksheet.max_row >= 3
        if status_col:
            if not DISABLE_DATA_VALIDATION:
                dv = DataValidation(
                    type="list",
                    formula1='"To Do,In Progress,Review,Completed"',
                    allow_blank=True,
                )
                worksheet.add_data_validation(dv)
                if has_data_rows:
                    dv.add(
                        f"{get_column_letter(status_col)}3:{get_column_letter(status_col)}{worksheet.max_row}"
                    )
            if has_data_rows and not DISABLE_CONDITIONAL_FORMATTING:
                status_letter = get_column_letter(status_col)
                worksheet.conditional_formatting.add(
                    f"{status_letter}3:{status_letter}{worksheet.max_row}",
                    FormulaRule(
                        formula=[f'LOWER({status_letter}3)="completed"'],
                        stopIfTrue=True,
                        fill=PatternFill("solid", fgColor="00B050"),
                        font=Font(color="000000"),
                    ),
                )
                worksheet.conditional_formatting.add(
                    f"{status_letter}3:{status_letter}{worksheet.max_row}",
                    FormulaRule(
                        formula=[f'LOWER({status_letter}3)="in progress"'],
                        stopIfTrue=True,
                        fill=PatternFill("solid", fgColor="FFFF00"),
                        font=Font(color="000000"),
                    ),
                )
                worksheet.conditional_formatting.add(
                    f"{status_letter}3:{status_letter}{worksheet.max_row}",
                    FormulaRule(
                        formula=[f'LOWER({status_letter}3)="review"'],
                        stopIfTrue=True,
                        fill=PatternFill("solid", fgColor="FFC000"),
                        font=Font(color="000000"),
                    ),
                )
                worksheet.conditional_formatting.add(
                    f"{status_letter}3:{status_letter}{worksheet.max_row}",
                    FormulaRule(
                        formula=[f'LOWER({status_letter}3)="to do"'],
                        stopIfTrue=True,
                        fill=PatternFill("solid", fgColor="FF0000"),
                        font=Font(color="FFFFFF"),
                    ),
                )
        headers = hub_headers
        assigned_owner_col = headers.get("Assigned Owner")
        if assigned_owner_col and not DISABLE_DATA_VALIDATION:
            owner_dv = DataValidation(
                type="list",
                formula1='"Unassigned,Copywriter,SEO Specialist,Developer"',
                allow_blank=True,
            )
            worksheet.add_data_validation(owner_dv)
            if has_data_rows:
                owner_dv.add(
                    f"{get_column_letter(assigned_owner_col)}3:{get_column_letter(assigned_owner_col)}{worksheet.max_row}"
                )
                for r in range(3, worksheet.max_row + 1):
                    cell = worksheet.cell(row=r, column=assigned_owner_col)
                    if not str(cell.value or "").strip():
                        cell.value = "Unassigned"
        if assigned_owner_col and has_data_rows and not DISABLE_CONDITIONAL_FORMATTING:
            owner_letter = get_column_letter(assigned_owner_col)
            owner_range = f"{owner_letter}3:{owner_letter}{worksheet.max_row}"
            worksheet.conditional_formatting.add(
                owner_range,
                FormulaRule(
                    formula=[f'LOWER({owner_letter}3)="copywriter"'],
                    stopIfTrue=True,
                    fill=PatternFill("solid", fgColor="BDD7EE"),
                    font=Font(color="000000"),
                ),
            )
            worksheet.conditional_formatting.add(
                owner_range,
                FormulaRule(
                    formula=[f'LOWER({owner_letter}3)="seo specialist"'],
                    stopIfTrue=True,
                    fill=PatternFill("solid", fgColor="C6E0B4"),
                    font=Font(color="000000"),
                ),
            )
            worksheet.conditional_formatting.add(
                owner_range,
                FormulaRule(
                    formula=[f'LOWER({owner_letter}3)="developer"'],
                    stopIfTrue=True,
                    fill=PatternFill("solid", fgColor="D9E1F2"),
                    font=Font(color="000000"),
                ),
            )
            worksheet.conditional_formatting.add(
                owner_range,
                FormulaRule(
                    formula=[f'LOWER({owner_letter}3)="unassigned"'],
                    stopIfTrue=True,
                    fill=PatternFill("solid", fgColor="D9D9D9"),
                    font=Font(color="000000"),
                ),
            )
        proposed_cols = [
            headers.get("Current Title"),
            headers.get("Proposed Title (50-60 Chars)"),
            headers.get("Current Meta Desc"),
            headers.get("Proposed Meta Desc (120-160 Chars)"),
            headers.get("Current H-Tag Structure"),
            headers.get("Current Page Copy Snippet"),
            headers.get("Social Share Note"),
            headers.get("Proposed H-Tag Fixes"),
            headers.get("AEO Answer Block Draft"),
            headers.get("FAQ/QA Draft"),
        ]
        black_font = Font(color="000000")
        for r in range(3, worksheet.max_row + 1):
            for c in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=r, column=c)
                cell.font = black_font
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col_idx in [c for c in proposed_cols if c]:
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = max(40, 48)
            for r in range(3, worksheet.max_row + 1):
                cell = worksheet.cell(row=r, column=col_idx)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.fill = PatternFill(
                    start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
                )
                cell.font = black_font
        current_cols = [
            headers.get("Current Title"),
            headers.get("Current Meta Desc"),
            headers.get("Current H-Tag Structure"),
            headers.get("Current Page Copy Snippet"),
            headers.get("Current OG-Image URL"),
            headers.get("Current SEO Score"),
        ]
        readonly_fill = PatternFill(
            start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"
        )
        for col_idx in [c for c in current_cols if c]:
            for r in range(3, worksheet.max_row + 1):
                cell = worksheet.cell(row=r, column=col_idx)
                cell.fill = readonly_fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.font = black_font
        title_prop_col = headers.get("Proposed Title (50-60 Chars)")
        title_count_col = headers.get("Title Count")
        desc_prop_col = headers.get("Proposed Meta Desc (120-160 Chars)")
        desc_count_col = headers.get("Desc Count")
        if action_required_col and title_prop_col and desc_prop_col:
            action_letter = get_column_letter(action_required_col)
            for r in range(3, worksheet.max_row + 1):
                formula_string = f'=IF(AND(LEN(J{r})>0, LEN(M{r})>0), "Ready to Publish", "Needs Copy")'
                worksheet.cell(row=r, column=1).value = formula_string
            if not DISABLE_CONDITIONAL_FORMATTING:
                red_fill = PatternFill(
                    start_color="FF0000", end_color="FF0000", fill_type="solid"
                )
                white_font = Font(color="FFFFFF", bold=True)
                green_fill = PatternFill(
                    start_color="00FF00", end_color="00FF00", fill_type="solid"
                )
                black_font = Font(color="000000", bold=True)
                worksheet.conditional_formatting.add(
                    f"A3:A{worksheet.max_row}",
                    CellIsRule(
                        operator="equal",
                        formula=['"Needs Copy"'],
                        fill=red_fill,
                        font=white_font,
                    ),
                )
                worksheet.conditional_formatting.add(
                    f"A3:A{worksheet.max_row}",
                    CellIsRule(
                        operator="equal",
                        formula=['"Ready to Publish"'],
                        fill=green_fill,
                        font=black_font,
                    ),
                )
        if title_prop_col and title_count_col:
            tp_letter = get_column_letter(title_prop_col)
            for r in range(3, worksheet.max_row + 1):
                worksheet.cell(
                    row=r, column=title_count_col, value=f"=LEN({tp_letter}{r})"
                )
            if not DISABLE_CONDITIONAL_FORMATTING:
                tc_letter = get_column_letter(title_count_col)
                worksheet.conditional_formatting.add(
                    f"{tc_letter}3:{tc_letter}{worksheet.max_row}",
                    FormulaRule(
                        formula=[f"AND({tc_letter}3>=50,{tc_letter}3<=60)"],
                        stopIfTrue=True,
                        fill=PatternFill("solid", fgColor="C6EFCE"),
                    ),
                )
                worksheet.conditional_formatting.add(
                    f"{tc_letter}3:{tc_letter}{worksheet.max_row}",
                    FormulaRule(
                        formula=[f"{tc_letter}3>60"],
                        stopIfTrue=True,
                        fill=PatternFill("solid", fgColor="FFC7CE"),
                    ),
                )
        if desc_prop_col and desc_count_col:
            dp_letter = get_column_letter(desc_prop_col)
            for r in range(3, worksheet.max_row + 1):
                worksheet.cell(
                    row=r, column=desc_count_col, value=f"=LEN({dp_letter}{r})"
                )
            if not DISABLE_CONDITIONAL_FORMATTING:
                dc_letter = get_column_letter(desc_count_col)
                worksheet.conditional_formatting.add(
                    f"{dc_letter}3:{dc_letter}{worksheet.max_row}",
                    FormulaRule(
                        formula=[f"AND({dc_letter}3>=120,{dc_letter}3<=160)"],
                        stopIfTrue=True,
                        fill=PatternFill("solid", fgColor="C6EFCE"),
                    ),
                )
                worksheet.conditional_formatting.add(
                    f"{dc_letter}3:{dc_letter}{worksheet.max_row}",
                    FormulaRule(
                        formula=[f"{dc_letter}3>160"],
                        stopIfTrue=True,
                        fill=PatternFill("solid", fgColor="FFC7CE"),
                    ),
                )
        for key in ("Status", "Assigned Owner", "URL", "Elementor Builder Link"):
            cidx = headers.get(key)
            if cidx:
                col_letter = get_column_letter(cidx)
                worksheet.column_dimensions[col_letter].width = (
                    18 if key in {"Status", "Assigned Owner"} else 42
                )
        if action_required_col:
            worksheet.column_dimensions[
                get_column_letter(action_required_col)
            ].width = 20
        current_score_col = headers.get("Current SEO Score")
        projected_score_col = headers.get("Projected SEO Score")
        if current_score_col:
            worksheet.column_dimensions[get_column_letter(current_score_col)].width = 18
            for r in range(3, worksheet.max_row + 1):
                cell = worksheet.cell(row=r, column=current_score_col)
                cell.fill = readonly_fill
                cell.font = black_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        if projected_score_col:
            worksheet.column_dimensions[
                get_column_letter(projected_score_col)
            ].width = 18
            for r in range(3, worksheet.max_row + 1):
                cell = worksheet.cell(row=r, column=projected_score_col)
                cell.value = f'=IF(A{r}="Ready to Publish", MIN(V{r}+25, 100), V{r})'
                cell.fill = PatternFill(
                    start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
                )
                cell.font = black_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for key in ("Current Title", "Current Meta Desc", "Current Page Copy Snippet"):
            cidx = headers.get(key)
            if cidx:
                worksheet.column_dimensions[get_column_letter(cidx)].width = max(
                    40, worksheet.column_dimensions[get_column_letter(cidx)].width or 40
                )
        elm_col = headers.get("Elementor Builder Link")
        if elm_col:
            worksheet.column_dimensions[get_column_letter(elm_col)].width = 42
        ct_col = headers.get("Current Title")
        cmd_col = headers.get("Current Meta Desc")
        ch1_col = headers.get("Current H-Tag Structure")
        if ct_col and cmd_col and ch1_col:
            for col_idx in (ct_col, cmd_col, ch1_col):
                worksheet.column_dimensions[get_column_letter(col_idx)].outlineLevel = 1
                worksheet.column_dimensions[get_column_letter(col_idx)].hidden = False
        # Editorial palette
        soft_header = PatternFill(
            start_color="6A5ACD", end_color="6A5ACD", fill_type="solid"
        )
        for cell in worksheet[2]:
            cell.fill = soft_header
            cell.font = Font(color="FFFFFF", bold=True)
        header_comments = {
            "URL": "The live page being audited.",
            "Elementor Builder Link": "Opens the Elementor editor directly, bypassing the standard WordPress dashboard.",
            "Current Title": "Live crawled <title> value from the page.",
            "Target Keywords": "Your AIOSEO Focus Keyphrase. Keep this topic in mind while drafting.",
            "Current Page Copy Snippet": "The first 250 characters of the actual page body text. Use this to understand the page context.",
            "Proposed Title (50-60 Chars)": "SEO RULE: 50-60 characters. Put the most important keyword at the beginning.",
            "Title Count": "Green = Good. Red = Too long (will be cut off by Google) or too short.",
            "Current Meta Desc": "Live crawled meta description from the page.",
            "Proposed Meta Desc (120-160 Chars)": "SEO RULE: 120-160 characters. Must include a Call-To-Action (e.g., 'Read more', 'Register here').",
            "Desc Count": "Green = Good. Red = Too long (will be cut off by Google) or too short.",
            "Current H-Tag Structure": "Current H1/H2/H3 structure scraped from the page.",
            "Current OG-Image URL": "The image shared on Social Media (Facebook/LinkedIn).",
            "OG Image Preview": "Live preview of the social share image. Requires Microsoft 365 and an active internet connection to render.",
            "Proposed H-Tag Fixes": "Ensure exactly ONE H1. Use H2s for main sections, and H3s for sub-sections. Format headings as questions where possible.",
            "AEO Answer Block Draft": "A 40-60 word, highly factual paragraph designed to win AI overviews/Featured Snippets. Do not use marketing fluff here.",
            "FAQ/QA Draft": "Draft Question & Answer pairs. These will be loaded into the AIOSEO FAQ schema block in Elementor.",
        }
        for h, msg in header_comments.items():
            cidx = headers.get(h)
            if cidx:
                worksheet.cell(row=2, column=cidx).comment = Comment(
                    msg, "SEO Audit Bot"
                )
        if action_required_col:
            worksheet.cell(row=2, column=action_required_col).comment = Comment(
                "Auto-calculated from proposed title and meta description completion.",
                "SEO Audit Bot",
            )
        url_col = headers.get("URL")
        elm_col = headers.get("Elementor Builder Link")
        og_url_col = headers.get("Current OG-Image URL")
        og_preview_col = headers.get("OG Image Preview")
        if url_col:
            for r in range(3, worksheet.max_row + 1):
                cell = worksheet.cell(row=r, column=url_col)
                value = str(cell.value or "").strip()
                if value.startswith(
                    ("http://", "https://")
                ) and _is_safe_hyperlink_target(value):
                    cell.hyperlink = value
                    cell.style = "Hyperlink"
        if elm_col:
            for r in range(3, worksheet.max_row + 1):
                cell = worksheet.cell(row=r, column=elm_col)
                value = str(cell.value or "").strip()
                if value.startswith(
                    ("http://", "https://")
                ) and _is_safe_hyperlink_target(value):
                    cell.hyperlink = value
                    cell.style = "Hyperlink"
                    cell.fill = PatternFill(
                        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
                    )
                    cell.font = Font(color="FFFFFF", bold=True, underline="single")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        open_in_main_col = headers.get("Open in Main")
        if not open_in_main_col:
            open_in_main_col = worksheet.max_column + 1
            worksheet.cell(row=2, column=open_in_main_col, value="Open in Main")
        main_ws = writer.book["Main"] if "Main" in writer.book.sheetnames else None
        if main_ws:
            main_headers = _header_index(main_ws)
            main_url_col = main_headers.get("URL", 1)
            for main_row in range(2, main_ws.max_row + 1):
                main_cell = main_ws.cell(row=main_row, column=main_url_col)
                main_cell.value = _normalize_url_for_match(main_cell.value)
        url_col = headers.get("URL")
        if url_col:
            for r in range(3, worksheet.max_row + 1):
                worksheet.cell(row=r, column=url_col).value = _normalize_url_for_match(
                    worksheet.cell(row=r, column=url_col).value
                )
        for r in range(3, worksheet.max_row + 1):
            worksheet.cell(
                row=r,
                column=open_in_main_col,
                value=f'=IFERROR(HYPERLINK("#\'Main\'!A" & MATCH(E{r}, \'Main\'!$A:$A, 0), "Open URL"), "Missing from Main")',
            )
        if og_preview_col and og_url_col:
            og_url_letter = get_column_letter(og_url_col)
            preview_letter = get_column_letter(og_preview_col)
            og_url_col_letter = get_column_letter(og_url_col)
            worksheet.column_dimensions[og_url_col_letter].width = 42
            worksheet.column_dimensions[preview_letter].width = 25
            for r in range(3, worksheet.max_row + 1):
                og_url_cell = worksheet.cell(row=r, column=og_url_col)
                og_url_cell.value = _sanitize_excel_url(og_url_cell.value)
                worksheet.cell(
                    row=r,
                    column=og_preview_col,
                    value=(
                        str(og_url_cell.value or "")
                        if (
                            DEBUG_EXCEL_ISOLATION_MODE
                            or DISABLE_EXTERNAL_LINKS_AND_IMAGES
                        )
                        else f'=IF(LEN({og_url_letter}{r})>0, _xlfn.IMAGE({og_url_letter}{r}), "")'
                    ),
                )
                pcell = worksheet.cell(row=r, column=og_preview_col)
                pcell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                worksheet.row_dimensions[r].height = 100
        target_keywords_col = headers.get("Target Keywords")
        if target_keywords_col:
            north_star_fill = PatternFill(
                start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
            )
            for r in range(3, worksheet.max_row + 1):
                worksheet.cell(row=r, column=target_keywords_col).fill = north_star_fill
        # Enable filters on the Hub headers.
        if worksheet.max_row >= 3 and worksheet.max_column >= 1:
            max_col = worksheet.max_column
            max_row = worksheet.max_row
            worksheet.auto_filter.ref = f"A2:{get_column_letter(max_col)}{max_row}"
        # Ensure instructional row remains plain merged text with no navigation links.
        # Only the top-left merged anchor cell is writable in openpyxl.
        worksheet["A1"].hyperlink = None
    if sheet_name == "Quick Reference Guide":
        _set_freeze_panes_safe(worksheet, "A2")
        for col_letter, width in {"A": 42, "B": 36, "C": 96, "D": 58}.items():
            worksheet.column_dimensions[col_letter].width = width
        for cell in worksheet[1]:
            cell.font = Font(color="1F2937", bold=True)
            cell.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=False
            )
        section_fill = PatternFill(
            start_color=STD_NAVY, end_color=STD_NAVY, fill_type="solid"
        )
        for row_idx in range(2, worksheet.max_row + 1):
            section_text = str(
                worksheet.cell(row=row_idx, column=1).value or ""
            ).strip()
            item_text = str(worksheet.cell(row=row_idx, column=2).value or "").strip()
            if section_text.startswith("[") and section_text.endswith("]"):
                for col_idx in range(1, 5):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = section_fill
                    cell.font = Font(color=STD_WHITE, bold=True)
                    cell.alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )
                worksheet.row_dimensions[row_idx].height = 24
            else:
                for col_idx in range(1, 5):
                    worksheet.cell(row=row_idx, column=col_idx).alignment = Alignment(
                        horizontal="left", vertical="top", wrap_text=True
                    )
                if item_text:
                    worksheet.cell(row=row_idx, column=2).font = Font(
                        color="1F2937", bold=True
                    )
                    worksheet.row_dimensions[row_idx].height = 44
        worksheet.protection.sheet = True
        worksheet.protection.enable()
        worksheet.protection.formatCells = False
        worksheet.protection.formatColumns = False
        worksheet.protection.formatRows = False
        worksheet.protection.insertColumns = False
        worksheet.protection.insertRows = False
        worksheet.protection.deleteColumns = False
        worksheet.protection.deleteRows = False
        worksheet.protection.sort = False
        worksheet.protection.autoFilter = True
    if sheet_name == "PSI Performance":
        headers = _header_index(worksheet)
        mobile_lcp_col = headers.get("Mobile LCP")
        mobile_cls_col = headers.get("Mobile CLS")
        for header_name, tooltip in {
            "Mobile LCP": "Largest Contentful Paint. Target: < 2.5 seconds.",
            "Mobile CLS": "Cumulative Layout Shift. Target: < 0.1.",
        }.items():
            cidx = headers.get(header_name)
            if cidx:
                worksheet.cell(row=1, column=cidx).comment = Comment(
                    tooltip, "SEO Audit Bot"
                )

        score_cols = [
            cidx
            for h, cidx in headers.items()
            if ("score" in str(h).lower())
            and ("desktop" in str(h).lower() or "mobile" in str(h).lower())
        ]
        for cidx in score_cols:
            col = get_column_letter(cidx)
            rng = f"{col}2:{col}{worksheet.max_row}"
            if not DISABLE_CONDITIONAL_FORMATTING:
                worksheet.conditional_formatting.add(
                    rng,
                    FormulaRule(
                        formula=[f"{col}2>=90"],
                        stopIfTrue=True,
                        fill=PatternFill("solid", fgColor="C6EFCE"),
                        font=Font(color="006100"),
                    ),
                )
                worksheet.conditional_formatting.add(
                    rng,
                    FormulaRule(
                        formula=[f"AND({col}2>=50,{col}2<90)"],
                        stopIfTrue=True,
                        fill=PatternFill("solid", fgColor="FFEB9C"),
                        font=Font(color="9C6500"),
                    ),
                )
                worksheet.conditional_formatting.add(
                    rng,
                    FormulaRule(
                        formula=[f"{col}2<50"],
                        stopIfTrue=True,
                        fill=PatternFill("solid", fgColor="FFC7CE"),
                        font=Font(color="9C0006"),
                    ),
                )
        if mobile_lcp_col and not DISABLE_CONDITIONAL_FORMATTING:
            col = get_column_letter(mobile_lcp_col)
            rng = f"{col}2:{col}{worksheet.max_row}"
            worksheet.conditional_formatting.add(
                rng,
                FormulaRule(
                    formula=[f"{col}2<=2.5"],
                    stopIfTrue=True,
                    fill=PatternFill("solid", fgColor="C6EFCE"),
                ),
            )
            worksheet.conditional_formatting.add(
                rng,
                FormulaRule(
                    formula=[f"{col}2>4.0"],
                    stopIfTrue=True,
                    fill=PatternFill("solid", fgColor="FFC7CE"),
                ),
            )
    if sheet_name == "AEO":
        _audit_non_overlapping_merges(worksheet)
    _add_all_header_tooltips(worksheet)
    if sheet_name in DATA_HEAVY_TABS:
        _add_header_tooltips(worksheet)
    if sheet_name == "Dashboard":
        if not DEBUG_EXCEL_ISOLATION_MODE:
            _style_dashboard(worksheet, writer)
    if sheet_name != "Dashboard":
        header_row = 2 if sheet_name == "Content Optimization Hub" else 1
        _normalize_table_headers(worksheet, header_row=header_row)
        header_values = [
            worksheet.cell(row=header_row, column=c).value
            for c in range(1, worksheet.max_column + 1)
        ]
        valid_table_headers = all(
            isinstance(v, str) and v.strip() for v in header_values
        )
        if (
            worksheet.max_row > header_row
            and worksheet.max_column > 0
            and valid_table_headers
        ):
            ref_string = _compute_exact_table_ref(worksheet, header_row)
            if ref_string:
                start_ref, end_ref = ref_string.split(":")
                min_row, min_col = coordinate_to_tuple(start_ref)
                max_row, max_col = coordinate_to_tuple(end_ref)
                _apply_mock_table_styling(
                    worksheet,
                    min_col=min_col,
                    max_col=max_col,
                    min_row=min_row,
                    max_row=max_row,
                )
    # Ensure final, exact filter/freeze settings are computed from final shape.
    ensure_auto_filter(worksheet)
    if sheet_name != "Dashboard":
        ensure_freeze_header(worksheet)
    _sanitize_sheet_view_selection(worksheet)
    _audit_non_overlapping_merges(worksheet)
    _audit_freeze_merge_conflicts(worksheet)


def apply_tab_hyperlinks(writer):
    apply_workbook_toc_and_links(
        writer,
        debug_excel_isolation_mode=DEBUG_EXCEL_ISOLATION_MODE,
        disable_non_core_freeze_panes=DISABLE_NON_CORE_FREEZE_PANES,
        std_navy=STD_NAVY,
        std_white=STD_WHITE,
        std_blue=STD_BLUE,
    )


__all__ = ["adjust_sheet_format", "apply_tab_hyperlinks"]
